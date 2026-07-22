#!/usr/bin/env python3
'\n工信部数据、语音跟踪统计工具 V2.5.35_CC\nGUI: 双标签(速率统计 / 语音VQI) → 选择文件(MMF+可选基准) → 处理 → 内嵌查看结果\nCLI: python3 this.py --cmd\n'
_BK='已请求取消，将在当前阶段结束后停止...'
_BJ='请先选择MMF文件'
_BI='QTextEdit{font-size:9pt;}'
_BH='QPushButton{background-color:#e8f0fe;border:1px solid #c0c0c0;border-radius:3px;padding:2px 8px;}'
_BG='QTextEdit{font-size:10pt;background:#fff;border:2px solid #666;border-radius:4px;padding:4px;}'
_BF='可以将文件或是文件夹拖入，支持各类文件类型'
_BE='Microsoft YaHei'
_BD='PingFang SC'
_BC='biz_order_table'
_BB='last_time'
_BA='用户跟踪ID=(\\d+)'
_B9='(\\d{4}-\\d{2}-\\d{2}\\s+\\d{2}:\\d{2}:\\d{2})[\\.\\(](\\d+)'
_B8='tool_type'
_B7='machine_code'
_B6='application/json; charset=utf-8'
_B5='Content-Type'
_B4='hubei_monitor_host.txt'
_B3='[\\t,，;；/\\s\\\\]+'
_B2='user_common_monitoring'
_B1='height'
_B0='width'
_A_='#CCE5FF'
_Az='#D9E1F2'
_Ay='#E2EFDA'
_Ax='dur_pct'
_Aw='dur_max'
_Av='dur_min'
_Au='flow_pct'
_At='flow_max'
_As='flow_min'
_Ar='stable'
_Aq='/.../'
_Ap='message_type'
_Ao='Message Type'
_An='openpyxl'
_Am='calamine'
_Al='open "'
_Ak='5GDataVoice'
_Aj='HubeiTool'
_Ai='当前未配置任何手机号-TraceID映射，结果中将无法显示手机号。\n是否继续处理？'
_Ah='手机号映射为空'
_Ag='平均呼叫建立时长(s)'
_Af='通话轮次'
_Ae='削峰平均速率(Mbps)'
_Ad='toPython'
_Ac='%Y%m%d-%H%M'
_Ab='V2.5.35_CC'
_Aa='.zip'
_AZ='.xls'
_AY='语音业务-原始文件'
_AX='MOS值'
_AW='.mmf'
_AV='text_wrap'
_AU='Uplink'
_AT='Downlink'
_AS='HH:mm:ss'
_AR='%Y-%m-%d %H:%M:%S'
_AQ='MOS3.5计算明细'
_AP='语音指标统计'
_AO='_in_call_period'
_AN='数据业务'
_AM='bold'
_AL='平均速率(Mbps)'
_AK='row_count'
_AJ='.csv'
_AI='.xlsx'
_AH='服务小区平均RSRP(dBm)(通话时)'
_AG='time_range'
_AF='汇总-数据和语音'
_AE='上行均值速率'
_AD='上行前10%峰值速率'
_AC='下行均值速率'
_AB='下行前10%峰值速率'
_AA='优质通话占比(MOS3.5)'
_A9='num_format'
_A8='速率(Mbps)'
_A7='None'
_A6='ul_avg'
_A5='ul_peak'
_A4='utf-8'
_A3='未知'
_A2='dl_avg'
_A1='dl_peak'
_A0='RLC'
_z='vqi'
_y='servRSRP值'
_x='时间'
_w='是否优质(MOS>3.5)'
_v='用户取消'
_u='_time_only'
_t='trace_id'
_i='是'
_s='电信'
_r='联通'
_q='et'
_d='data'
_p='bg_color'
_o='full_path'
_n='dur'
_m='微信小包发送'
_l='/'
_k='QCI'
_j='s'
_h='operator'
_g='应用商店小文件下载'
_e='Time'
_c='border'
_b='MessageType'
_a='\n'
_Z='vcenter'
_Y='center'
_X='valign'
_W='align'
_V='phone'
_U='微信大包发送'
_T='业务类型'
_S='ul'
_R='应用商店大文件下载'
_Q='coerce'
_P='dl'
_O='呼叫建立时长(s)'
_N='st'
_M='手机号'
_L='FTP上传'
_K='FTP下载'
_J='业务'
_I='结束时间'
_H='轮次'
_G='开始时间'
_F='_source_file'
_E='运营商'
_D=False
_C='_t'
_B=True
_A=None
import sys,os,re,json,hashlib,socket,uuid,subprocess,urllib.request,threading,pandas as pd,numpy as np
from datetime import datetime,timedelta,time
from collections import Counter,defaultdict
_FMT_DT=_AR
_FMT_HMS=_AS
_EXE_DIR=os.path.dirname(os.path.abspath(__file__))
_MONITOR_PORT=5002
_TOOL_TYPE='xianjia'
_HOST_FILE=os.path.join(_EXE_DIR,_B4)
MONITOR_HOST=_A
if os.path.exists(_HOST_FILE):
	with open(_HOST_FILE,'r')as _f:MONITOR_HOST=_f.read().strip()
if not MONITOR_HOST:
	_alt=os.path.join(os.path.expanduser('~'),_B4)
	if os.path.exists(_alt):
		with open(_alt,'r')as _f:MONITOR_HOST=_f.read().strip()
if not MONITOR_HOST:MONITOR_HOST='47.109.101.79'
_LICENSE_HOME=os.path.join(os.path.expanduser('~'),'.hubei_xianjia')
os.makedirs(_LICENSE_HOME,exist_ok=_B)
_LICENSE_FILE=os.path.join(_LICENSE_HOME,'.hb_license.json')
_OFFLINE_FILE=os.path.join(_LICENSE_HOME,'.offline_count')
_LICENSE_DATA={}
def _load_license():
	global _LICENSE_DATA
	if os.path.exists(_LICENSE_FILE):
		try:
			with open(_LICENSE_FILE,'r',encoding=_A4)as f:_LICENSE_DATA=json.load(f)
		except Exception:pass
	return _LICENSE_DATA
def _save_license():
	try:
		with open(_LICENSE_FILE,'w',encoding=_A4)as f:json.dump(_LICENSE_DATA,f,ensure_ascii=_D)
	except Exception:pass
def _get_machine_code():
	'生成本机唯一机器码';stable_id=_A
	try:
		out=subprocess.check_output(['ioreg','-d2','-c','IOPlatformExpertDevice'],stderr=subprocess.DEVNULL,timeout=5).decode(_A4,errors='ignore')
		for line in out.split(_a):
			if'IOPlatformUUID'in line:stable_id=line.strip().split('"')[-2];break
	except Exception:pass
	if not stable_id:
		try:
			out=subprocess.check_output('wmic csproduct get uuid',shell=_B,timeout=5).decode();lines=out.strip().split(_a)
			if len(lines)>1:stable_id=lines[1].strip()
		except Exception:pass
	if not stable_id:
		id_file=os.path.join(_LICENSE_HOME,'.machine_id')
		if os.path.exists(id_file):
			with open(id_file,'r')as f:stable_id=f.read().strip()
		if not stable_id:
			stable_id=uuid.uuid4().hex
			with open(id_file,'w')as f:f.write(stable_id)
	raw=f"{socket.gethostname()}:{stable_id}";return hashlib.sha256(raw.encode()).hexdigest()[:16]
def _http_get(path,timeout=5):
	'HTTP GET 到鉴权服务器'
	try:req=urllib.request.Request(f"http://{MONITOR_HOST}:{_MONITOR_PORT}{path}",headers={_B5:_B6});resp=urllib.request.urlopen(req,timeout=timeout);return json.loads(resp.read().decode())
	except Exception:return
def _http_post(path,payload,timeout=5):
	'HTTP POST 到鉴权服务器'
	try:req=urllib.request.Request(f"http://{MONITOR_HOST}:{_MONITOR_PORT}{path}",data=json.dumps(payload,ensure_ascii=_D,default=str).encode(_A4),headers={_B5:_B6},method='POST');resp=urllib.request.urlopen(req,timeout=timeout);return json.loads(resp.read().decode())
	except Exception:return
def _read_offline_count():
	try:
		if os.path.exists(_OFFLINE_FILE):return int(open(_OFFLINE_FILE).read().strip())
	except Exception:pass
	return 0
def _inc_offline_count():
	c=_read_offline_count()+1
	try:
		with open(_OFFLINE_FILE,'w')as f:f.write(str(c))
	except Exception:pass
	return c
_load_license()
_machine_code=_get_machine_code()
def check_license():
	'鉴权主函数：联网验证→注册→100次配额→离线5次兜底。返回(ok, msg)';E='online';D='limit';C='usage_count';B='status';A='ok';result=_http_get(f"/api/check_license?code={_machine_code}&tool_type={_TOOL_TYPE}")
	if result and result.get(A):
		if result.get(B)=='banned':return _D,'使用权限已被禁止'
		if result.get(B)=='unknown':
			reg=register_machine()
			if reg and reg.get(A):
				result=_http_get(f"/api/check_license?code={_machine_code}&tool_type={_TOOL_TYPE}")
				if not result or not result.get(A):return _D,'注册后鉴权失败，请重试'
			else:_reg_err=reg.get('error','未知错误')if reg else'无网络';return _D,f"自动注册失败: {_reg_err}"
		if result.get('quota_exceeded'):_used=result.get(C,'?');_lim=result.get(D,'?');return _D,f"配额已用完({_used}/{_lim})"
		remain=max(0,result.get(D,100)-result.get(C,0));global _LICENSE_DATA;_LICENSE_DATA.update({'last_check':str(datetime.now()),E:_B});_save_license()
		try:os.remove(_OFFLINE_FILE)
		except Exception:pass
		return _B,f"鉴权通过(云端余{remain}次)"
	offline_count=_read_offline_count();data=_load_license()
	if data.get(E):
		if offline_count<5:_inc_offline_count();return _B,f"离线模式(剩余{5-offline_count-1}次)"
		else:return _D,'离线次数已用完(5次)，请联网鉴权'
	else:return _D,'首次使用需联网鉴权'
def register_machine():
	'首次联网时自动注册机器（静默）'
	try:hostname=socket.gethostname();return _http_post('/api/register',{_B7:_machine_code,_B8:_TOOL_TYPE,'name':hostname,'employee_id':'xianjia','hostname':hostname})
	except Exception:return
def consume_quota():
	'处理完成后上报消费配额（静默）'
	def _post():_http_post('/api/report_done',{_B7:_machine_code,_B8:_TOOL_TYPE})
	threading.Thread(target=_post,daemon=_B).start()
def _read_file(f,nrows=_A):
	'读取xlsx/xls/csv/mmf/tmf文件，返回DataFrame\n    nrows: 可选，只读取前n行（用于快速判断文件类型）\n    ';A='latin-1';low=f.lower()
	if low.endswith(_AJ):
		for enc in['gbk',_A4,A]:
			try:return pd.read_csv(f,encoding=enc,nrows=nrows)
			except Exception:continue
		return pd.read_csv(f,encoding=A,on_bad_lines='skip',nrows=nrows)
	try:return pd.read_excel(f,engine=_Am,nrows=nrows)
	except Exception:
		try:return pd.read_excel(f,nrows=nrows)
		except Exception:
			try:return pd.read_excel(f,engine=_An,nrows=nrows)
			except Exception:return pd.read_excel(f,engine='xlrd')
def _has_message_type_col(df):
	'判断DataFrame是否有MessageType列（支持多种列名）'
	for col in[_b,_Ao,_Ap,'消息类型']:
		if col in df.columns:return _B
	if len(df.columns)>=2 and _e in df.columns and'Source'in df.columns:return _D
	return _D
_PARSE_RE=re.compile(_B9)
_PARSE_MS_RE=re.compile('(\\d{4}-\\d{2}-\\d{2}\\s+\\d{2}:\\d{2}:\\d{2})\\.(\\d+)')
def parse_time(t):
	m=_PARSE_RE.match(str(t))
	if m:return datetime.strptime(m.group(1),_AR)+timedelta(milliseconds=int(m.group(2)))
def _parse_time_vec(s):"向量化解析时间列(替代 apply(parse_time),数万行从数十秒→毫秒级)。\n    格式: '2026-07-19 18:27:28(123)' 或 '...18:27:28.123' → datetime+毫秒";ext=s.astype(str).str.extract(_B9);base=pd.to_datetime(ext[0],format=_AR,errors=_Q);ms=pd.to_numeric(ext[1],errors=_Q).fillna(0);return base+pd.to_timedelta(ms,unit='ms')
def _add_calc_comment(cell,formula_desc,calc_process,source_col,numerator_vals=_A,denominator_vals=_A,numerator_desc='分子',denominator_desc='分母'):
	'为单元格添加计算批注，显示公式、分子分母值'
	if numerator_vals is _A or len(numerator_vals)==0:vs_str='[]'
	else:
		vs=[str(round(v,2))for v in numerator_vals]
		if len(vs)<=20:vs_str=_l.join(vs)
		else:vs_str=_l.join(vs[:5])+_Aq+_l.join(vs[-5:])
	if denominator_vals is _A or len(denominator_vals)==0:dv_str='[]'
	else:
		dv=[str(round(v,2))for v in denominator_vals]
		if len(dv)<=20:dv_str=_l.join(dv)
		else:dv_str=_l.join(dv[:5])+_Aq+_l.join(dv[-5:])
	txt=f"""公式: {formula_desc}
计算过程: {calc_process}

{numerator_desc}: {vs_str}
{denominator_desc}: {dv_str}

数据来源: {source_col} 列""";from openpyxl.comments import Comment;c=Comment(txt,'Tool');c.width=500;c.height=250;cell.comment=c
DL_ACTIVE_BIG,UL_ACTIVE_BIG=5e1,1e1
DL_ACTIVE_SMALL,UL_ACTIVE_SMALL=5.,5.
GAP_MERGE=3.
FTP_PAIR_GAP=9e1
STORE_L_DL_MIN,STORE_L_DL_MAX=500,99999
WX_L_DUR_MIN,WX_L_DUR_MAX,WX_L_UL_MIN=12,90,40
DL_BIZ={_K,_g,_R}
UL_BIZ={_L,_m,_U}
REQ_BIZ={_K,_L,_g,_R,_m,_U}
def detect_segments(ts,dl,ul,dl_th,ul_th,gap,r1=_A,r2=_A):
	active=(dl>dl_th)|(ul>ul_th);segs=[];i=0;n=len(ts)
	while i<n:
		if r1 is not _A and ts[i]<r1:i+=1;continue
		if r2 is not _A and ts[i]>r2:break
		if active[i]:
			si=i;j=i+1
			while j<n:
				if r2 is not _A and ts[j]>r2:break
				if not active[j]:
					k=j
					while k<n and not active[k]:
						if(ts[k]-ts[j-1])/np.timedelta64(1,_j)>gap:break
						k+=1
					if k<n and active[k]and(r2 is _A or ts[k]<=r2):j=k;continue
					else:break
				j+=1
			st=ts[si];et=ts[j-1];dur=(et-st)/np.timedelta64(1,_j)+1;segs.append({_N:st,_q:et,_n:dur,_P:float(dl[si:j].sum()),_S:float(ul[si:j].sum())});i=j
		else:i+=1
	return segs
def _extract_trace_id(path):"从文件路径中提取用户跟踪ID, 如 '用户跟踪ID=553' → '553'";import re;m=re.search(_BA,str(path));return m.group(1)if m else _A
def _lookup_phone(trace_id,phone_trace_map):
	'在phone_trace_map中精确查找trace_id对应的手机号'
	if not trace_id or not phone_trace_map:return''
	for entry in phone_trace_map:
		if entry[_t]==trace_id:return entry[_V]
	return''
def process(files,qci_list=_A,dl_clip=1000,ul_clip=200,callback=_A,cancel_check=_A,progress_cb=_A,base_file=_A,time_filter=_A,ftp_duration=25,merge_raw=_D,add_annotations=_B,phone_trace_map=_A,biz_order=_A,biz_params=_A):
	A4='上传速率(Mbps)';A3='下载速率（Mbps）';A2='#E5FFCC';A1='font_size';A0='df_detail_complete';z='has_base';y='_write_order';x='df_summary_complete';t='df_summary_all';u='业务时长中位值(s)';q='业务时长平均值(s)';p='Uplink RLC Throughput(bps)';o='Downlink RLC Throughput(bps)';n='微信小包';m='微信大包';l='应用宝小包';k='应用宝大包';c='整理后文件';b='max';V='ALL_FTP_DETAIL_';U='#DAEEF3';a='df_detail';Z='数值.1';Y='时长类指标';X='数值';W='速率类指标';T='永川参考数据.xlsx';S='有效';R='-';Q='ul_rate';P='dl_rate';O='mark_dur';N='mark_et';M='mark_st';L='持续时长';K='上传速率';J='下载速率';I='FTP候选';H='业务识别';G='_s';F='nan';E='_ul';D='_dl';C='businessType';B='biz';A='rnd'
	if qci_list is _A:qci_list=[5,6,7]
	ftp_dur_min=ftp_duration-3.;ftp_dur_max=ftp_duration+3.;'统一的处理函数：从MMF文件生成5Sheet输出'
	def log(msg,pct=_A):
		if cancel_check and cancel_check():raise KeyboardInterrupt(_v)
		import time as _time;_t=_time.time()
		if not hasattr(log,_BB):log.last_time=_t
		elapsed=_t-log.last_time if pct is not _A else 0;prefix=f"({elapsed:.1f}秒) "if pct is not _A and elapsed>0 else'';print(msg)
		if callback:callback(f"{prefix}{msg}")
		if pct is not _A and progress_cb:progress_cb(pct)
		log.last_time=_t
	import zipfile,tempfile,shutil,glob;_tmp_dirs=[]
	def _expand_zip(zip_path):
		'递归解压 zip(含嵌套 zip)，返回所有 xlsx/xls 路径';_tmp=tempfile.mkdtemp(prefix='mmf_unzip_');_tmp_dirs.append(_tmp);_xs=[]
		try:
			with zipfile.ZipFile(zip_path)as _zf:_zf.extractall(_tmp)
			_xs=glob.glob(os.path.join(_tmp,'**','*.xlsx'),recursive=_B)+glob.glob(os.path.join(_tmp,'**','*.xls'),recursive=_B);_inner_zips=glob.glob(os.path.join(_tmp,'**','*.zip'),recursive=_B)
			for _iz in _inner_zips:_xs.extend(_expand_zip(_iz))
		except Exception as _e:log(f"  解压失败 {os.path.basename(zip_path)}: {_e}")
		return _xs
	_expanded=[]
	for _f in files or[]:
		if str(_f).lower().endswith(_Aa):_xs=_expand_zip(_f);log(f"  解压 {os.path.basename(_f)}: 找到 {len(_xs)} 个Excel(含嵌套)",5);_expanded.extend(_xs)
		else:_expanded.append(_f)
	_filtered=[]
	for _f in _expanded:
		_bn=os.path.basename(str(_f));_bnl=_bn.lower()
		if _bn.startswith('.~')or'savemsg'in _bnl or _AJ in _bnl:continue
		if'mmf'in _bnl:_filtered.append(_f)
	skipped=len(_expanded)-len(_filtered)
	if skipped>0:log(f"  自动跳过 {skipped} 个文件(语音CSV/临时文件)",5)
	files=_filtered;nf=len(files)if files else 1;log(f"加载MMF文件（共 {nf} 个）...",5);dfs=[]
	for(fi,f)in enumerate(files):
		if cancel_check and cancel_check():raise KeyboardInterrupt(_v)
		basename=os.path.basename(f)
		if basename.startswith(V):log(f"  跳过非MMF文件 {fi+1}/{nf}: {basename}",5+int(15*(fi+1)/nf));continue
		log(f"  读取文件 {fi+1}/{nf}: {basename}",5+int(15*(fi+1)/nf));df_i=_read_file(f);rlc_dl=[c for c in df_i.columns if _A0 in c and _AT in c];rlc_ul=[c for c in df_i.columns if _A0 in c and _AU in c];_dl_val=pd.to_numeric(df_i[rlc_dl[0]],errors=_Q).fillna(0)/1e6 if rlc_dl else 0;_ul_val=pd.to_numeric(df_i[rlc_ul[0]],errors=_Q).fillna(0)/1e6 if rlc_ul else 0;df_i=df_i.assign(_t=_parse_time_vec(df_i[_e]),_source_file=os.path.basename(f),_dl=_dl_val,_ul=_ul_val);dfs.append(df_i)
	if not dfs:
		log('警告: 所有文件均为非MMF文件名(ALL_FTP_DETAIL_)，无可处理数据',100);version=_Ab;timestamp=datetime.now().strftime(_Ac);out=f"联通/电信最终输出_{version}_{timestamp}.xlsx"
		with pd.ExcelWriter(out,engine=_An)as w:pd.DataFrame({'提示':['所有文件均为非MMF文件名(ALL_FTP_DETAIL_)，无可处理数据。请选择MMF文件后重试。']}).to_excel(w,sheet_name='提示',index=_D)
		log(f"完成(空): {out}",100);return out
	df=pd.concat(dfs,ignore_index=_B);log('  合并完成，按时间排序...',22);data_file_info={};dfs_idx=0
	for(fi,f)in enumerate(files):
		fname=os.path.basename(f)
		if fname.startswith(V):continue
		df_i=dfs[dfs_idx];dfs_idx+=1;dl_vals=df_i[D].dropna().values;ul_vals=df_i[E].dropna().values;dl_nonzero=dl_vals[dl_vals>0];ul_nonzero=ul_vals[ul_vals>0];dl_peak=0
		if len(dl_nonzero)>0:tn=max(1,int(len(dl_nonzero)*.1));dl_sorted=np.sort(dl_nonzero)[::-1];dl_peak=round(min(float(np.mean(dl_sorted[:tn])),1000),2)
		dl_avg=0
		if len(dl_nonzero)>0:dl_avg=round(min(float(np.mean(dl_nonzero)),1000),2)
		ul_peak=0
		if len(ul_nonzero)>0:tn=max(1,int(len(ul_nonzero)*.1));ul_sorted=np.sort(ul_nonzero)[::-1];ul_peak=round(min(float(np.mean(ul_sorted[:tn])),200),2)
		ul_avg=0
		if len(ul_nonzero)>0:ul_avg=round(min(float(np.mean(ul_nonzero)),200),2)
		times=df_i[_C].dropna();time_range_str=''
		if len(times)>0:time_range_str=f"{times.min().strftime(_FMT_DT)} ~ {times.max().strftime(_FMT_DT)}"
		data_file_info[fname]={_h:_detect_operator(f),_AK:len(df_i),_AG:time_range_str,_A1:dl_peak,_A2:dl_avg,_A5:ul_peak,_A6:ul_avg,_o:f}
	op_map={fname:info[_h]for(fname,info)in data_file_info.items()};df[_E]=df[_F].map(op_map);df=df.sort_values(_C,kind=_Ar).reset_index(drop=_B)
	if time_filter:
		use_time,t_start,t_end=time_filter
		if use_time and t_start is not _A and t_end is not _A:t_start_t=t_start.toPython()if hasattr(t_start,_Ad)else t_start;t_end_t=t_end.toPython()if hasattr(t_end,_Ad)else t_end;df[_u]=df[_C].dt.time;mask=(df[_u]>=t_start_t)&(df[_u]<=t_end_t);df=df[mask].copy();df=df.drop(_u,axis=1);log(f"  时间段过滤: {t_start_t} ~ {t_end_t}, 剩余 {len(df)} 行",22)
	log('QCI过滤+秒聚合...',28);df[_k]=pd.to_numeric(df[_k],errors=_Q);df7=df[df[_k].isin(qci_list)].dropna(subset=[_C]).copy()
	if len(df7)==0:
		available_qci=df[_k].dropna().unique();avail_str=', '.join([str(int(float(q)))for q in available_qci if pd.notna(q)and str(q).strip()not in('',F,_A7)])
		if avail_str:
			log(f"警告: 数据中无QCI={qci_list}的记录，可用QCI: {avail_str}",28);log(f"建议: 重新选择QCI后再次运行",28);version=_Ab;timestamp=datetime.now().strftime(_Ac);out=f"联通/电信最终输出_{version}_{timestamp}.xlsx"
			with pd.ExcelWriter(out,engine=_An)as w:pd.DataFrame({'提示':[f"数据中无QCI={qci_list}的记录，可用QCI: {avail_str}"]}).to_excel(w,sheet_name='提示',index=_D)
			log(f"完成(空): {out}",100);return out
		else:raise ValueError('数据中无有效的QCI记录')
	df7[G]=df7[_C].dt.floor(_j);sec=df7.groupby(G).agg({D:b,E:b}).sort_index().reset_index()
	if len(sec)>0:_full_ts=pd.date_range(sec[G].min(),sec[G].max(),freq=_j);sec=sec.set_index(G).reindex(_full_ts).reset_index(names=G)
	dl_v,ul_v,ts_v=sec[D].fillna(0).values,sec[E].fillna(0).values,sec[G].values;log('业务识别...',35);big=detect_segments(ts_v,dl_v,ul_v,DL_ACTIVE_BIG,UL_ACTIVE_BIG,GAP_MERGE);_DEFAULT_BIZ_ORDER=[_L,_K,_m,_U,_g,_R];BIZ_ORDER=biz_order if biz_order else _DEFAULT_BIZ_ORDER
	def _match_strict(s,biz):
		'严格匹配：业务类型的标准特征';d,dl,ul=s[_n],s[_P],s[_S]
		if biz==_L:return ul>dl and ftp_dur_min<=d<=ftp_dur_max
		if biz==_K:return dl>ul and ftp_dur_min<=d<=ftp_dur_max
		if biz==_R:return dl>ul and STORE_L_DL_MIN<=dl<=STORE_L_DL_MAX and d>6
		if biz==_g:return dl>ul and dl>=200 and 2<=d<=4
		if biz==_U:return ul>dl and WX_L_DUR_MIN<=d<=WX_L_DUR_MAX and ul>=WX_L_UL_MIN
		if biz==_m:return ul>dl and ul>=40 and 1<=d<=3
		return _D
	def _match_relaxed(s,biz):
		'宽松匹配：降低阈值兜底';d,dl,ul=s[_n],s[_P],s[_S]
		if biz==_L:return ul>dl and ftp_dur_min-5<=d<=ftp_dur_max+5
		if biz==_K:return dl>ul and ftp_dur_min-5<=d<=ftp_dur_max+5
		if biz==_R:return dl>ul and dl>=300 and d>3
		if biz==_g:return dl>ul and dl>=50 and d>=2
		if biz==_U:return ul>dl and d>=5 and ul>=25
		if biz==_m:return ul>dl and ul>=20 and d>=1
		return _D
	_big_sorted=sorted(big,key=lambda x:x[_N]);_pos=0
	for s in _big_sorted:
		s[A]=_A;expected=BIZ_ORDER[_pos];first_biz=BIZ_ORDER[0]
		if _match_strict(s,expected):s[B]=I if expected in(_K,_L)else expected;_pos=(_pos+1)%len(BIZ_ORDER);continue
		if expected!=first_biz and _match_strict(s,first_biz):s[B]=I if first_biz in(_K,_L)else first_biz;_pos=1;continue
		_found=_D;_big_bizs={_K,_L,_R,_U}
		for(_i,_biz)in enumerate(BIZ_ORDER):
			if _biz in _big_bizs and _match_strict(s,_biz):s[B]=I if _biz in(_K,_L)else _biz;_pos=(_i+1)%len(BIZ_ORDER);_found=_B;break
		if _found:continue
		if _match_relaxed(s,expected):s[B]=I if expected in(_K,_L)else expected;_pos=(_pos+1)%len(BIZ_ORDER);continue
		if expected!=first_biz and _match_relaxed(s,first_biz):s[B]=I if first_biz in(_K,_L)else first_biz;_pos=1;continue
		if biz_params:
			d,dl,ul=s[_n],s[_P],s[_S]
			for(_biz,_bp)in biz_params.items():
				_fmin,_fmax=_bp[_As],_bp[_At];_fpct,_dmin,_dmax=_bp[_Au],_bp[_Av],_bp[_Aw];_dpct=_bp[_Ax];_flow=ul if _biz in UL_BIZ else dl;_flow_ok=_flow>=_fmin*(1-_fpct)and _flow<=_fmax*(1+_fpct);_dur_ok=d>=_dmin*(1-_dpct)and d<=_dmax*(1+_dpct)
				if _flow_ok and _dur_ok:
					s[B]=I if _biz in(_K,_L)else _biz
					try:_pos=(BIZ_ORDER.index(_biz)+1)%len(BIZ_ORDER)
					except ValueError:pass
					break
			else:s[B]=_A3
			continue
		s[B]=_A3
	log('FTP配对+小业务识别...',40);fc=[s for s in big if s[B]==I];paired=set();fp=[]
	for i in range(len(fc)):
		if id(fc[i])in paired:continue
		for j in range(i+1,len(fc)):
			if id(fc[j])in paired:continue
			g=(fc[j][_N]-fc[i][_q])/np.timedelta64(1,_j)
			if 0<=g<FTP_PAIR_GAP:fp.append((fc[i],fc[j]));paired.add(id(fc[i]));paired.add(id(fc[j]));break
			elif g>=FTP_PAIR_GAP:break
	def fd(s):return _K if s[_P]>s[_S]else _L
	for s in fc:
		if id(s)not in paired:
			if STORE_L_DL_MIN<=s[_P]<=STORE_L_DL_MAX and s[_P]>s[_S]:s[B]=_R
			else:s[B]=fd(s)
	al=sorted(big,key=lambda x:x[_N])
	for s in al:s[_J]=s[B]
	rn=0;pi=[]
	for(f1,f2)in fp:rn+=1;f1[A]=rn;f1[_J]=fd(f1);f2[A]=rn;f2[_J]=fd(f2);pi.append((rn,f2[_q],f1[_N]))
	if not fp:
		log('无FTP配对，使用大业务作为轮次起点...',42);large_bizs=[s for s in al if s[_J]in(_R,_U,_K,_L)]
		for(i,s)in enumerate(large_bizs):rn+=1;s[A]=rn;pi.append((rn,s[_q],s[_N]))
		log(f"  识别到 {rn} 个大业务作为轮次",43)
	for(idx,(rnd,pe,_))in enumerate(pi):
		ns=pi[idx+1][2]if idx+1<len(pi)else al[-1][_q]+np.timedelta64(1,_j)
		for s in al:
			if s[_q]>pe and s[_N]<ns and s[A]is _A and s[_J]in(_R,_U):s[A]=rnd
	tb=sorted([s for s in al if s[_J]in(_K,_L,_R,_U)],key=lambda x:x[_N]);P2={_R:_g,_U:_m};small=[]
	for(idx,(rnd,pe,ps))in enumerate(pi):
		we=pi[idx+1][2]if idx+1<len(pi)else tb[-1][_q]+np.timedelta64(1,_j);bw=sorted([s for s in tb if s[_N]>=pe and s[_N]<we and s[_J]in(_R,_U)],key=lambda x:x[_N]);prev=pe
		for bs in bw:
			st2=P2.get(bs[_J])
			if st2 and prev<bs[_N]and(bs[_N]-prev)/np.timedelta64(1,_j)>1:
				gs=detect_segments(ts_v,dl_v,ul_v,DL_ACTIVE_SMALL,UL_ACTIVE_SMALL,GAP_MERGE,prev,bs[_N]);vd=[s for s in gs if s[_P]+s[_S]>=10 and s[_n]>=.5]
				if st2==_g:
					dls=[s for s in vd if s[_P]>s[_S]and s[_n]>=2]
					if dls:bm=max(dls,key=lambda x:x[_P]);bm[_J]=st2;bm[A]=rnd;small.append(bm)
				else:
					uls=[s for s in vd if s[_S]>s[_P]and s[_n]>=2]
					if uls:bm=max(uls,key=lambda x:x[_S]);bm[_J]=st2;bm[A]=rnd;small.append(bm)
			prev=bs[_q]
	if pi:
		first_ftp_start=pi[0][2];data_start=ts_v[0]
		if first_ftp_start>data_start+np.timedelta64(30,_j):
			pre_big=sorted([s for s in al if s[_N]<first_ftp_start and s[_J]in(_R,_U)],key=lambda x:x[_N]);rng_start=data_start
			if pre_big:
				for bs in pre_big:
					st2=P2.get(bs[_J])
					if st2 and(bs[_N]-rng_start)/np.timedelta64(1,_j)>1:
						gs=detect_segments(ts_v,dl_v,ul_v,DL_ACTIVE_SMALL,UL_ACTIVE_SMALL,GAP_MERGE,rng_start,bs[_N]);vd=[s for s in gs if s[_P]+s[_S]>=10 and s[_n]>=.5]
						if st2==_g:
							dls=[s for s in vd if s[_P]>s[_S]and s[_n]>=2]
							if dls:bm=max(dls,key=lambda x:x[_P]);bm[_J]=st2;bm[A]=0;small.append(bm)
						else:
							uls=[s for s in vd if s[_S]>s[_P]and s[_n]>=2]
							if uls:bm=max(uls,key=lambda x:x[_S]);bm[_J]=st2;bm[A]=0;small.append(bm)
					rng_start=bs[_q]
			if first_ftp_start>rng_start+np.timedelta64(1,_j):
				gs=detect_segments(ts_v,dl_v,ul_v,DL_ACTIVE_SMALL,UL_ACTIVE_SMALL,GAP_MERGE,rng_start,first_ftp_start);vd=[s for s in gs if s[_P]+s[_S]>=10 and s[_n]>=.5];dls=[s for s in vd if s[_P]>s[_S]and s[_n]>=2]
				if dls:bm=max(dls,key=lambda x:x[_P]);bm[_J]=_g;bm[A]=0;small.append(bm)
				uls=[s for s in vd if s[_S]>s[_P]and s[_n]>=2]
				if uls:bm=max(uls,key=lambda x:x[_S]);bm[_J]=_m;bm[A]=0;small.append(bm)
	for s in al:
		if s[_J]in(_g,_m,_A3)and s.get(A)is _A:
			et=s[_q];assigned=_A
			for(idx,(rnd,pe,ps))in enumerate(pi):
				ns=pi[idx+1][2]if idx+1<len(pi)else s[_N]+np.timedelta64(999,_j)
				if et>pe and s[_N]<ns:assigned=rnd;break
			if assigned is _A and pi:s[A]=pi[-1][0]
			elif assigned is not _A:s[A]=assigned
			else:rn+=1;s[A]=rn
	af=sorted(al+small,key=lambda x:x[_N])
	for s in af:s[S]=s[_J]is not _A and s.get(A)is not _A
	rm=defaultdict(list)
	for s in af:
		if s[S]and s[_J]and s.get(A)is not _A:rm[s[A]].append(s)
	for(rnd,segs)in rm.items():
		seen=set();cnt=0
		for s in sorted(segs,key=lambda x:x[_N]):
			if s[_J]==_A3:continue
			if s[_J]in seen or cnt>=6:s[S]=_D
			else:seen.add(s[_J]);cnt+=1
	valid=[s for s in af if s[S]and s[_J]];log(f"识别完成: {len(valid)}个业务段, {rn}轮");df=pd.concat([df,pd.DataFrame({J:df[D],K:df[E],H:_A,L:_A,_A8:_A},index=df.index)],axis=1)
	for(si,seg)in enumerate(valid):
		if cancel_check and si%20==0 and cancel_check():raise KeyboardInterrupt(_v)
		st,et,biz=seg[_N],seg[_q],seg[_J];sm=(df[_C]>=st)&(df[_C]<=et+timedelta(seconds=1))&df[_k].isin(qci_list);col=D if biz in DL_BIZ else E;cv_all=df.loc[sm,col].fillna(0);_ftp_th=1e1 if biz in(_K,_L)else .0;nz=cv_all>_ftp_th;mm=sm&nz.reindex(df.index,fill_value=_D)
		if not mm.any():continue
		if biz in(_K,_L):
			flow_vals=df.loc[mm,col].fillna(0).tolist();flow_times=df.loc[mm,_C].tolist()
			if len(flow_vals)==0:continue
			first_idx=0;last_idx=len(flow_vals)-1;first_t=flow_times[first_idx];last_t=flow_times[last_idx];md=(last_t-first_t).total_seconds()
			if md<9.5 or md>10.5:
				ft0=st-timedelta(seconds=3);pm=(df[_C]>=ft0)&(df[_C]<first_t)&df[_k].isin(qci_list)
				if pm.any():
					pv=df.loc[pm,col].fillna(0);pv_nz=pv>_ftp_th
					if pv_nz.any():first_t=df.loc[pv[pv_nz].index[0],_C];md=(last_t-first_t).total_seconds()
				et0=et+timedelta(seconds=3);pm2=(df[_C]>last_t)&(df[_C]<=et0)&df[_k].isin(qci_list)
				if pm2.any():
					pv2=df.loc[pm2,col].fillna(0);pv2_nz=pv2>_ftp_th
					if pv2_nz.any():last_t=df.loc[pv2[pv2_nz].index[-1],_C];md=(last_t-first_t).total_seconds()
			_ftp_end_th=1e1 if biz in(_K,_L)else .0;mm_final=(df[_C]>=first_t)&(df[_C]<=last_t)&df[_k].isin(qci_list)&(df[col].fillna(0)>_ftp_end_th)
			if biz in(_K,_L)and mm_final.any():
				_seg_vals=df.loc[mm_final,col].fillna(0);_peak=_seg_vals.max()
				if _peak>0:
					_rel_th=max(_peak*.25,1e1);_rel_mask=_seg_vals>=_rel_th
					if _rel_mask.any():
						_rel_idx=_seg_vals.index[_rel_mask];mm_final=mm_final&(df.index>=_rel_idx[0])&(df.index<=_rel_idx[-1]);_new_times=df.loc[mm_final,_C]
						if len(_new_times)>0:first_t,last_t=_new_times.min(),_new_times.max()
			md=(last_t-first_t).total_seconds();df.loc[mm_final,H]=biz;df.loc[mm_final,L]=round(md,3);vs=pd.to_numeric(df.loc[mm_final,J]if biz in DL_BIZ else df.loc[mm_final,K],errors=_Q).fillna(0).tolist();ar=sum(vs)/md if md>0 else 0;df.loc[mm_final,_A8]=round(ar,2);seg[M],seg[N],seg[O]=first_t,last_t,md;continue
		_nonftp_vals=df.loc[mm,col].fillna(0)
		if len(_nonftp_vals)>1:
			_pk=_nonftp_vals.max()
			if _pk>0:
				_th=max(_pk*.1,5.);_active_mask=_nonftp_vals>=_th
				if _active_mask.any():_ai=_nonftp_vals.index[_active_mask];mm=mm&(df.index>=_ai[0])&(df.index<=_ai[-1])
		if not mm.any():continue
		first_t=df.loc[mm,_C].min();last_t=df.loc[mm,_C].max();md=(last_t-first_t).total_seconds();df.loc[mm,H]=biz;df.loc[mm,L]=round(md,3)
		if biz in DL_BIZ:vs=pd.to_numeric(df.loc[mm,J],errors=_Q).fillna(0).tolist()
		else:vs=pd.to_numeric(df.loc[mm,K],errors=_Q).fillna(0).tolist()
		ar=sum(vs)/md if md>0 else 0;df.loc[mm,_A8]=round(ar,2);seg[M],seg[N],seg[O]=first_t,last_t,md
	for(f1,f2)in fp:
		dur1=f1.get(O,0);dur2=f2.get(O,0)
		if dur1<=0 or dur2<=0:continue
		diff=abs(dur1-dur2)
		if diff<=.6:continue
		_long,_short=(f1,f2)if dur1>dur2 else(f2,f1);_long_dur,_short_dur=max(dur1,dur2),min(dur1,dur2);_lcol=D if _long[_J]in DL_BIZ else E;_lmask=(df[_C]>=_long[M])&(df[_C]<=_long[N])&df[_k].isin(qci_list);_lvals=df.loc[_lmask,_lcol].fillna(0)
		if len(_lvals)<=_short_dur:continue
		_trim_n=int(len(_lvals)-_short_dur)
		if _trim_n<=0:continue
		_candidates=pd.concat([_lvals.head(_trim_n),_lvals.tail(_trim_n)]);_trim_idx=set(_candidates.nsmallest(_trim_n).index)
		if _trim_idx:
			_new_mask=_lmask&~df.index.isin(_trim_idx)
			if _new_mask.any():
				df.loc[_lmask,H]=_A;df.loc[_new_mask,H]=_long[_J];_new_t=df.loc[_new_mask,_C];_new_dur=(_new_t.max()-_new_t.min()).total_seconds();df.loc[_new_mask,L]=round(_new_dur,3)
				if _new_dur>0:_long_col=J if _long[_J]in DL_BIZ else K;_vs=pd.to_numeric(df.loc[_new_mask,_long_col],errors=_Q).fillna(0).tolist();df.loc[_new_mask,_A8]=round(sum(_vs)/_new_dur,2)
				_long[M],_long[N],_long[O]=_new_t.min(),_new_t.max(),_new_dur
	rd=defaultdict(set)
	for s in valid:
		r=s.get(A)
		if r is not _A:rd[r].add(s[_J])
	rl={};ii=0
	for r in sorted(rd.keys()):
		if rd[r]>=REQ_BIZ:rl[r]=str(r)
		else:ii+=1;rl[r]=f"不完整{ii}"
	df[_H]=_A;qci_mask=df[_k].isin(qci_list);biz_notna_mask=df[H].notna()
	for s in valid:
		r=s.get(A);lb=rl.get(r)if r is not _A else''
		if lb:ms2,me2=s.get(M,s[_N]),s.get(N,s[_q]);mm2=(df[_C]>=ms2)&(df[_C]<=me2)&qci_mask&biz_notna_mask;df.loc[mm2,_H]=lb
	operators_in_data=df[_E].dropna().unique();operators_list=[op for op in operators_in_data if op in(_r,_s)]
	if not operators_list:log('警告: 无法识别运营商，所有数据将合并处理',47);operators_list=[_A3]
	_sep=', ';log(f"识别到运营商: {_sep.join(operators_list)}",47);log('加载基准数据...',53);per_operator_bases={};_base_src=base_file if base_file and os.path.exists(base_file)else _A
	if _base_src is _A:
		for _try in[T,'重大-永川参考数据.xlsx',os.path.join(os.path.dirname(os.path.dirname(__file__)),'输入文件',T),os.path.join(os.path.expanduser('~'),T)]:
			if os.path.exists(_try):_base_src=_try;break
	if _base_src is _A:
		log('未找到基准文件，跳过基准对齐',53)
		for op in operators_list:per_operator_bases[op]=pd.DataFrame()
	else:
		try:
			_xls=pd.ExcelFile(_base_src,engine=_Am);_sn=_xls.sheet_names;_sheet=c if c in _sn else'呼叫详情'if'呼叫详情'in _sn else _sn[0];raw=pd.read_excel(_base_src,sheet_name=_sheet,engine=_Am);log(f"加载基准文件: {os.path.basename(_base_src)} [{_sheet}], {len(raw)}行",53);_biz_map={k:_R,l:_g,m:_U,n:_m,_K:_K,_L:_L}
			def _build_base_segments(raw_df,op_prefix):
				'从永川基准数据构建某运营商的业务段（含RLC吞吐率平均值，用于速率辅助对齐）';D='业务时长(秒)';C='上行RLC吞吐率(Mbps)';B='下行RLC吞吐率(Mbps)';A='采集时间';biz_col=f"{op_prefix}业务"
				if biz_col not in raw_df.columns:log(f"  {op_prefix}基准: 未找到{biz_col}列，跳过",53);return pd.DataFrame()
				dl_rate_col=f"{op_prefix}下行RLC吞吐率(Mbps)";ul_rate_col=f"{op_prefix}上行RLC吞吐率(Mbps)";has_dl_rate=dl_rate_col in raw_df.columns;has_ul_rate=ul_rate_col in raw_df.columns
				if has_dl_rate or has_ul_rate:log(f"  {op_prefix}基准: 检测到RLC吞吐率列，启用速率辅助对齐",53)
				_sub_cols=[A,biz_col]
				if has_dl_rate:_sub_cols.append(dl_rate_col)
				if has_ul_rate:_sub_cols.append(ul_rate_col)
				sub=raw_df[_sub_cols].copy();sub=sub.rename(columns={A:_G,biz_col:_T});sub[_T]=sub[_T].map(_biz_map).fillna(sub[_T]);segments=[];cur_biz,seg_start=_A,0
				for i in range(len(sub)):
					biz=sub.loc[i,_T]if pd.notna(sub.loc[i,_T])else _A
					if biz!=cur_biz:
						if cur_biz and seg_start<i:
							seg_slice=sub.iloc[seg_start:i];seg_dl=float(seg_slice[dl_rate_col].mean())if has_dl_rate else _A;seg_ul=float(seg_slice[ul_rate_col].mean())if has_ul_rate else _A;_t_start=pd.to_datetime(sub.iloc[seg_start][_G]);_t_end=pd.to_datetime(sub.iloc[i-1][_G]);_dur=round((_t_end-_t_start).total_seconds(),2)if pd.notna(_t_start)and pd.notna(_t_end)else 0
							if cur_biz in DL_BIZ:_rate_indicator=B;_rate_val=round(seg_dl,2)if seg_dl is not _A else 0
							elif cur_biz in UL_BIZ:_rate_indicator=C;_rate_val=round(seg_ul,2)if seg_ul is not _A else 0
							else:_rate_indicator='';_rate_val=0
							segments.append({_G:sub.iloc[seg_start][_G],_I:sub.iloc[i-1][_G],_T:cur_biz,W:_rate_indicator,X:_rate_val,Y:D,Z:_dur,P:seg_dl,Q:seg_ul})
						cur_biz,seg_start=biz,i
				if cur_biz and seg_start<len(sub):
					seg_slice=sub.iloc[seg_start:];seg_dl=float(seg_slice[dl_rate_col].mean())if has_dl_rate else _A;seg_ul=float(seg_slice[ul_rate_col].mean())if has_ul_rate else _A;_t_start=pd.to_datetime(sub.iloc[seg_start][_G]);_t_end=pd.to_datetime(sub.iloc[len(sub)-1][_G]);_dur=round((_t_end-_t_start).total_seconds(),2)if pd.notna(_t_start)and pd.notna(_t_end)else 0
					if cur_biz in DL_BIZ:_rate_indicator=B;_rate_val=round(seg_dl,2)if seg_dl is not _A else 0
					elif cur_biz in UL_BIZ:_rate_indicator=C;_rate_val=round(seg_ul,2)if seg_ul is not _A else 0
					else:_rate_indicator='';_rate_val=0
					segments.append({_G:sub.iloc[seg_start][_G],_I:sub.iloc[len(sub)-1][_G],_T:cur_biz,W:_rate_indicator,X:_rate_val,Y:D,Z:_dur,P:seg_dl,Q:seg_ul})
				base_df=pd.DataFrame(segments)
				if len(base_df)>0:log(f"  {op_prefix}基准段: {len(base_df)}个",53)
				return base_df
			for op in operators_list:
				if op==_r:base_op=_build_base_segments(raw,_r)
				elif op==_s:base_op=_build_base_segments(raw,_s)
				else:base_op=pd.DataFrame()
				per_operator_bases[op]=base_op
		except Exception as e:
			log(f"永川参考加载失败: {e}",53)
			for op in operators_list:per_operator_bases[op]=pd.DataFrame()
	def _process_one_operator(df_op,op_name,base_op,has_base_op):
		'处理单个运营商的数据：构建详细过程 + 基准对齐 + 汇总';V='end_idx';U='start_idx';T='基准_数值.1';S='基准_时长类指标';O='基准_数值';N='基准_速率类指标';M='基准_结束时间';I='基准_开始时间';A='基准_业务类型';op_file_to_path={}
		for(fname,info)in data_file_info.items():
			if info[_h]==op_name or op_name==_A3:op_file_to_path[fname]=info[_o]
		trace_to_phone={}
		if phone_trace_map:
			for entry in phone_trace_map:trace_to_phone[entry[_t]]=entry[_V]
		_source_file_to_phone={}
		for sf in df_op[_F].unique():_source_file_to_phone[sf]=trace_to_phone.get(_extract_trace_id(op_file_to_path.get(sf,'')),'')
		OC=[J,K,_A8,L,_H];base_cols=[A,I,M,N,O,S,T];special_cols=[_k,o,p];exclude_cols={'id',_C,D,E,G,C,H};base_col_dict={bc:_A for bc in base_cols};df_op=df_op.assign(筛选=0,id=range(1,len(df_op)+1),手机号=df_op[_F].map(_source_file_to_phone).fillna(''),businessType=df_op[H],**base_col_dict);_write_order=['筛选','id',_e]
		for c in OC:
			if c in df_op.columns:_write_order.append(c)
		_write_order.append(C)
		for bc in base_cols:
			if bc in df_op.columns:_write_order.append(bc)
		if _C in df_op.columns:_write_order.append(_C)
		_write_order.append(_M)
		for sc in special_cols:
			if sc in df_op.columns:_write_order.append(sc)
		for c in df_op.columns:
			if c not in _write_order and c not in exclude_cols:_write_order.append(c)
		df_detail=df_op
		if len(df_detail)>0 and _C in df_detail.columns:
			_t_vals=df_detail[_C].dropna()
			if len(_t_vals)>0:
				_existing_secs=set(_t_vals.dt.floor(_j));_all_secs=pd.date_range(_t_vals.min(),_t_vals.max(),freq=_j);_missing=[ts for ts in _all_secs if ts not in _existing_secs]
				if _missing:_empty_rows=pd.DataFrame({_C:_missing});df_detail=pd.concat([df_detail,_empty_rows],ignore_index=_B);df_detail=df_detail.sort_values(_C).reset_index(drop=_B)
		if has_base_op and len(base_op)>0 and _T in base_op.columns and C in df_detail.columns:
			mmf_segments=[];cur_biz=_A;seg_start=0
			for row_idx in range(len(df_detail)):
				biz_val=df_detail.loc[row_idx,C];biz=str(biz_val).strip()if pd.notna(biz_val)and str(biz_val).strip()not in('',F,_A7)else''
				if biz=='':continue
				if biz!=cur_biz:
					if cur_biz and seg_start<row_idx:mmf_segments.append({B:cur_biz,U:seg_start,V:row_idx-1})
					cur_biz=biz if biz else _A;seg_start=row_idx
			if cur_biz and seg_start<len(df_detail):mmf_segments.append({B:cur_biz,U:seg_start,V:len(df_detail)-1})
			if len(mmf_segments)==0:log(f"  {op_name}: MMF无可识别业务段，跳过基准对齐",53)
			else:
				base_biz_seq=base_op[_T].astype(str).tolist();base_starts=pd.to_datetime(base_op[_G]).tolist()if _G in base_op.columns else['']*len(base_op);base_ends=pd.to_datetime(base_op[_I]).tolist()if _I in base_op.columns else['']*len(base_op);base_rate_indicators=base_op[W].astype(str).tolist();base_rate_values=base_op[X].tolist();base_dur_indicators=base_op[Y].astype(str).tolist();base_dur_values=base_op[Z].tolist();has_rate_check=P in base_op.columns or Q in base_op.columns
				if has_rate_check:base_dl_rates=base_op[P].tolist()if P in base_op.columns else[_A]*len(base_op);base_ul_rates=base_op[Q].tolist()if Q in base_op.columns else[_A]*len(base_op)
				else:base_dl_rates=[_A]*len(base_op);base_ul_rates=[_A]*len(base_op)
				mmf_ptr=0;matching_pairs=[];matched_count=0;skipped_no_match=0
				for base_i in range(len(base_op)):
					if cancel_check and base_i%50==0 and cancel_check():raise KeyboardInterrupt(_v)
					b_biz=base_biz_seq[base_i];b_dl_rate=base_dl_rates[base_i];b_ul_rate=base_ul_rates[base_i];found_j=_A
					for j in range(mmf_ptr,len(mmf_segments)):
						if mmf_segments[j][B]!=b_biz:continue
						if has_rate_check and(b_dl_rate is not _A or b_ul_rate is not _A):
							m_start=mmf_segments[j][U];m_end=mmf_segments[j][V];mmf_dl_vals=pd.to_numeric(df_detail.loc[m_start:m_end,D],errors=_Q).dropna();mmf_ul_vals=pd.to_numeric(df_detail.loc[m_start:m_end,E],errors=_Q).dropna();mmf_dl_avg=float(mmf_dl_vals.mean())if len(mmf_dl_vals)>0 else 0;mmf_ul_avg=float(mmf_ul_vals.mean())if len(mmf_ul_vals)>0 else 0;rate_ok=_B
							if b_dl_rate is not _A and b_dl_rate>0 and mmf_dl_avg>0:
								ratio_dl=mmf_dl_avg/b_dl_rate
								if not .05<ratio_dl<2e1:rate_ok=_D
							if b_ul_rate is not _A and b_ul_rate>0 and mmf_ul_avg>0:
								ratio_ul=mmf_ul_avg/b_ul_rate
								if not .05<ratio_ul<2e1:rate_ok=_D
							if rate_ok:
								b_t=pd.to_datetime(base_starts[base_i]);m_t_start=df_detail.loc[m_start,_C];m_t_end=df_detail.loc[m_end,_C]
								if abs((m_t_start-b_t).total_seconds())<=60 or abs((m_t_end-b_t).total_seconds())<=60:found_j=j;break
						else:found_j=j;break
					if found_j is not _A:matching_pairs.append((base_i,found_j));mmf_ptr=found_j+1;matched_count+=1
					else:skipped_no_match+=1
				if matched_count>0 or skipped_no_match>0:log(f"  {op_name}基准对齐: 业务序列匹配{matched_count}段, 跳过{skipped_no_match}段",53)
				for(base_i,mmf_j)in matching_pairs:
					lo=mmf_segments[mmf_j][U];hi=mmf_segments[mmf_j][V]
					if lo<=hi:
						b_start=pd.to_datetime(base_starts[base_i]);b_end=pd.to_datetime(base_ends[base_i])
						if pd.notna(b_start)and pd.notna(b_end)and _C in df_detail.columns:_time_mask=(df_detail[_C]>=b_start)&(df_detail[_C]<=b_end);df_detail.loc[_time_mask,A]=base_biz_seq[base_i];df_detail.loc[_time_mask,I]=str(base_starts[base_i]);df_detail.loc[_time_mask,M]=str(base_ends[base_i]);df_detail.loc[_time_mask,N]=base_rate_indicators[base_i];df_detail.loc[_time_mask,O]=base_rate_values[base_i];df_detail.loc[_time_mask,S]=base_dur_indicators[base_i];df_detail.loc[_time_mask,T]=base_dur_values[base_i]
						else:df_detail.loc[lo:hi,A]=base_biz_seq[base_i];df_detail.loc[lo:hi,I]=str(base_starts[base_i]);df_detail.loc[lo:hi,M]=str(base_ends[base_i]);df_detail.loc[lo:hi,N]=base_rate_indicators[base_i];df_detail.loc[lo:hi,O]=base_rate_values[base_i];df_detail.loc[lo:hi,S]=base_dur_indicators[base_i];df_detail.loc[lo:hi,T]=base_dur_values[base_i]
		check_cols=[_A8,L,_H,C,A,I,M,N,O,S,T];has=pd.Series(_D,index=df_detail.index)
		for col in check_cols:s=df_detail[col];ss=s.astype(str).str.strip();has=has|s.notna()&(ss!='')&(ss.str.lower()!=F)
		df_detail['筛选']=has.astype(int).values
		def _calc_summary(sub_detail,dl_clip,ul_clip):
			'从 df_detail 子集计算汇总行（含子段修剪：剔除段首尾低活跃行）';D='削峰TOP10%峰值速率';B='5M/100M以下占比(%)';A='阈值以上占比(%)';sr=[]
			for biz in sorted(DL_BIZ|UL_BIZ):
				sub=sub_detail[sub_detail[C]==biz].dropna(subset=[_C]).sort_values(_C)
				if len(sub)==0:continue
				dl_cn=[c for c in sub.columns if _A0 in c and _AT in c];ul_cn=[c for c in sub.columns if _A0 in c and _AU in c]
				if biz in DL_BIZ and dl_cn:
					cn=dl_cn[0];raw_rates=pd.to_numeric(sub[cn],errors=_Q).fillna(0)/1e6
					if len(raw_rates)>1:
						_peak=raw_rates.max()
						if _peak>0:
							_th=max(_peak*.1,5.);_mask=raw_rates>=_th;_active_idx=raw_rates.index[_mask]
							if len(_active_idx)>0:raw_rates=raw_rates.loc[_active_idx[0]:_active_idx[-1]]
					rates=raw_rates[raw_rates>0]
				elif biz in UL_BIZ and ul_cn:
					cn=ul_cn[0];raw_rates=pd.to_numeric(sub[cn],errors=_Q).fillna(0)/1e6
					if len(raw_rates)>1:
						_peak=raw_rates.max()
						if _peak>0:
							_th=max(_peak*.1,5.);_mask=raw_rates>=_th;_active_idx=raw_rates.index[_mask]
							if len(_active_idx)>0:raw_rates=raw_rates.loc[_active_idx[0]:_active_idx[-1]]
					rates=raw_rates[raw_rates>0]
				else:rates=pd.Series([],dtype=float);cn=''
				ts2=sub[_C];durs=[]
				if len(ts2)>0:
					s1=ts2.iloc[0];p=s1
					for t in ts2.iloc[1:]:
						if(t-p).total_seconds()>5:durs.append((p-s1).total_seconds());s1=t
						p=t
					durs.append((p-s1).total_seconds())
				row={_T:biz};is_dl=biz in DL_BIZ;lim=dl_clip if is_dl else ul_clip;lo=100 if is_dl else 5
				if len(rates)>0:row[A]=round((rates>lim).sum()/len(rates)*100,2);row[B]=round((rates<lo).sum()/len(rates)*100,2);row[_AL]=round(rates.mean(),2);row[_Ae]=round(rates.clip(upper=lim).mean(),2);sv=rates.clip(upper=lim).sort_values(ascending=_D);tn=max(1,int(len(sv)*.1));row[D]=round(sv.head(tn).mean(),2)
				else:
					for k in[A,B,_AL,_Ae,D]:row[k]=R
				if durs:row[q]=round(np.mean(durs),2);row[u]=round(np.median(durs),0)
				phones_in_sub=sub[_M].dropna().unique();row[_M]='、'.join(sorted([str(p)for p in phones_in_sub]))if len(phones_in_sub)>0 else'';sr.append(row)
			return pd.DataFrame(sr)
		df_summary_all=_calc_summary(df_detail,dl_clip,ul_clip);complete_mask=df_detail[_H].astype(str).str.isdigit();df_detail_complete=df_detail[complete_mask];df_summary_complete=_calc_summary(df_detail_complete,dl_clip,ul_clip);return{a:df_detail,t:df_summary_all,x:df_summary_complete,y:_write_order,z:has_base_op,A0:df_detail_complete}
	log('按运营商处理...',50);all_results={}
	for op in operators_list:
		df_op=df[df[_E]==op].copy().reset_index(drop=_B)
		if len(df_op)==0:log(f"  {op}: 无数据，跳过",50);continue
		base_op=per_operator_bases.get(op,pd.DataFrame());has_base_op=len(base_op)>0;log(f"  处理 {op}: {len(df_op)} 行",55);all_results[op]=_process_one_operator(df_op,op,base_op,has_base_op)
	version=_Ab;timestamp=datetime.now().strftime(_Ac);ops_in_output=[op for op in operators_list if op in all_results]
	if len(ops_in_output)==1:out=f"{ops_in_output[0]}最终输出_{version}_{timestamp}.xlsx"
	else:out=f"联通电信最终输出_{version}_{timestamp}.xlsx"
	if os.path.dirname(out):os.makedirs(os.path.dirname(out),exist_ok=_B)
	else:os.makedirs('.',exist_ok=_B)
	log('写入Excel...',78);import xlsxwriter;from openpyxl.utils import get_column_letter as _gcl;from datetime import datetime as _dt_cls,date as _date_cls
	def _is_eo(v):
		'None 或 0（NaN 不算 — 与原 openpyxl 行为一致：isinstance int/float and v==0）'
		if v is _A:return _B
		if isinstance(v,bool):return _D
		if isinstance(v,(int,float))and not pd.isna(v)and v==0:return _B
		return _D
	def _ne(v):
		'非空（None/NaN/空串/nan字符串 都算空）'
		if v is _A:return _D
		if isinstance(v,float)and pd.isna(v):return _D
		if isinstance(v,str)and v.strip()in('',F,_A7):return _D
		return _B
	def _cval(v):
		'转 xlsxwriter 可写的值'
		if v is _A:return
		if isinstance(v,float)and pd.isna(v):return
		if isinstance(v,pd.Timestamp):return v.to_pydatetime()
		if isinstance(v,np.datetime64):return pd.Timestamp(v).to_pydatetime()
		if isinstance(v,(_dt_cls,_date_cls)):return v
		if isinstance(v,np.integer):return int(v)
		if isinstance(v,np.floating):return float(v)
		if isinstance(v,str)and v=='':return
		return v
	def _ws_write(ws,row,col,v,fmt=_A):
		'智能写入，处理 NaN/Timestamp/None';cv=_cval(v)
		if cv is _A:ws.write_blank(row,col,_A,fmt);return
		if isinstance(cv,bool):ws.write_boolean(row,col,cv,fmt);return
		if isinstance(cv,(int,float)):ws.write_number(row,col,float(cv),fmt);return
		if isinstance(cv,(_dt_cls,_date_cls)):ws.write_datetime(row,col,cv,fmt);return
		if isinstance(cv,str):ws.write_string(row,col,cv,fmt);return
		ws.write_string(row,col,str(cv),fmt)
	def _add_val(lst,v):
		'快速将 numpy 值转为 Python 原生类型并加到列表(内联避免函数调用)'
		if v is _A:lst.append('');return
		if isinstance(v,float):
			if v!=v:lst.append('');return
			lst.append(v);return
		if isinstance(v,(int,bool)):lst.append(v);return
		if isinstance(v,np.integer):lst.append(int(v));return
		if isinstance(v,np.floating):
			fv=float(v)
			if fv!=fv:lst.append('');return
			lst.append(fv);return
		if isinstance(v,(pd.Timestamp,np.datetime64)):lst.append(str(v));return
		s=str(v);lst.append(''if s in('',F,_A7)else s)
	_out_tmp=out+'.tmp'
	with xlsxwriter.Workbook(_out_tmp,{'strings_to_numbers':_D,'nan_inf_to_errors':_B,'default_date_format':'yyyy-mm-dd hh:mm:ss'})as wb:
		fmt_hdr_d=wb.add_format({_AV:_B});fmt_num=wb.add_format({_A9:'0'});fmt_d=wb.add_format({_A9:'0',_p:U});fmt_e=wb.add_format({_A9:'0',_p:_Ay});fmt_hdr_s=wb.add_format({_AM:_B,_W:_Y,_X:_Z,_p:_Az,_c:1,_AV:_B});fmt_ta=wb.add_format({_AM:_B,_W:_Y,_X:_Z,A1:11,_p:_Ay,_c:1});fmt_tc=wb.add_format({_AM:_B,_W:_Y,_X:_Z,A1:11,_p:U,_c:1});fmt_d_s=wb.add_format({_W:_Y,_X:_Z,_c:1});fmt_d_url=wb.add_format({_W:_Y,_X:_Z,_c:1,'font_color':'#0000FF','underline':1});fmt_hdr_h=wb.add_format({_AM:_B,_W:_Y,_X:_Z,_p:_Az,_c:1,_AV:_B});fmt_biz={bn:wb.add_format({_p:_A_ if i%2==0 else A2,_c:1,_W:_Y,_X:_Z})for(i,bn)in enumerate(BIZ_ORDER)};_biz_color_num={bn:wb.add_format({_p:_A_ if i%2==0 else A2,_c:1,_W:_Y,_X:_Z,_A9:'0'})for(i,bn)in enumerate(BIZ_ORDER)};fmt_y=wb.add_format({_p:'#FFFF00',_c:1,_W:_Y,_X:_Z});fmt_center=wb.add_format({_W:_Y,_X:_Z})
		for op in operators_list:
			if op not in all_results:continue
			if cancel_check and cancel_check():raise KeyboardInterrupt(_v)
			res=all_results[op];df_detail=res[a];df_summary_all=res[t];df_summary_complete=res[x];_write_order=res[y];has_base=res[z];df_detail_complete=res[A0];detail_sheet_name=f"{op}-详细过程";summary_sheet_name=f"{op}-汇总";log(f"  写入 {detail_sheet_name}...",82);ws=wb.add_worksheet(detail_sheet_name);cols=_write_order;n_detail=len(df_detail);n_col_detail=len(cols);_write_arrays=[df_detail[c].values for c in cols];col_3_idx=cols.index(J)if J in cols else 3;col_4_idx=cols.index(K)if K in cols else 4;col_biz_idx=cols.index(C)if C in cols else _A
			for(ci,col_name)in enumerate(cols):
				if ci==col_3_idx:ws.write(0,ci,A3,fmt_hdr_d)
				elif ci==col_4_idx:ws.write(0,ci,A4,fmt_hdr_d)
				else:ws.write(0,ci,col_name,fmt_hdr_d)
			_fmt_d_no_color=wb.add_format({_A9:'0'});_fmt_e_no_color=wb.add_format({_A9:'0'})
			if n_col_detail>col_3_idx:ws.set_column(col_3_idx,col_3_idx,_A,_fmt_d_no_color)
			if n_col_detail>col_4_idx:ws.set_column(col_4_idx,col_4_idx,_A,_fmt_e_no_color)
			_d_arr=_write_arrays[col_3_idx]if n_col_detail>col_3_idx else _A;_e_arr=_write_arrays[col_4_idx]if n_col_detail>col_4_idx else _A
			for i in range(n_detail):
				if cancel_check and i>0 and i%5000==0 and cancel_check():raise KeyboardInterrupt(_v)
				row0=i+1;d_orig=_d_arr[i]if _d_arr is not _A else _A;e_orig=_e_arr[i]if _e_arr is not _A else _A;d_dash=_is_eo(d_orig);e_dash=_is_eo(e_orig);row_vals=[]
				for ci in range(n_col_detail):
					v=_write_arrays[ci][i]
					if ci==0:row_vals.append(0 if d_dash else int(v)if not pd.isna(v)else'')
					elif ci==col_3_idx:row_vals.append(_A)
					elif ci==col_4_idx:row_vals.append(_A)
					else:_add_val(row_vals,v)
				ws.write_row(row0,0,row_vals);_biz_v_for_color=_write_arrays[col_biz_idx][i]if col_biz_idx is not _A and n_col_detail>col_biz_idx else _A;_has_biz=_biz_v_for_color is not _A and not(isinstance(_biz_v_for_color,float)and pd.isna(_biz_v_for_color))and str(_biz_v_for_color).strip()not in('',F,_A7);_biz_str=str(_biz_v_for_color).strip()if _has_biz else'';_is_dl_biz=_biz_str in DL_BIZ if _has_biz else _D;_is_ul_biz=_biz_str in UL_BIZ if _has_biz else _D;_biz_color=fmt_biz.get(_biz_str,fmt_d_s)if _has_biz else _A
				if n_col_detail>col_3_idx:
					if d_dash:ws.write_string(row0,col_3_idx,R)
					else:
						_dv=float(_d_arr[i])if _d_arr is not _A else 0
						if not pd.isna(_dv):ws.write_number(row0,col_3_idx,_dv,_biz_color_num.get(_biz_str,_fmt_d_no_color)if _is_dl_biz else _fmt_d_no_color)
				if n_col_detail>col_4_idx:
					if e_dash:ws.write_string(row0,col_4_idx,R)
					else:
						_ev=float(_e_arr[i])if _e_arr is not _A else 0
						if not pd.isna(_ev):ws.write_number(row0,col_4_idx,_ev,_biz_color_num.get(_biz_str,_fmt_e_no_color)if _is_ul_biz else _fmt_e_no_color)
				if col_biz_idx is not _A and n_col_detail>col_biz_idx:
					_biz_v=str(_write_arrays[col_biz_idx][i])if _write_arrays[col_biz_idx][i]is not _A and not(isinstance(_write_arrays[col_biz_idx][i],float)and pd.isna(_write_arrays[col_biz_idx][i]))else''
					if _biz_v:_biz_fmt=fmt_biz.get(_biz_v,fmt_d_s);ws.write(row0,col_biz_idx,_biz_v,_biz_fmt)
			for ci in range(n_col_detail):
				name=cols[ci]
				if ci==col_3_idx:name=A3
				elif ci==col_4_idx:name=A4
				width=max(min(len(str(name))+2,30),10);hidden=not has_base and 9<=ci<=16
				if hidden:ws.set_column(ci,ci,width,_A,{'hidden':_B})
				else:ws.set_column(ci,ci,width)
			ws.freeze_panes(1,9);ws.autofilter(0,0,n_detail,n_col_detail-1);ws.filter_column(0,'x == 1');log(f"  写入 {summary_sheet_name}...",88)
			if cancel_check and cancel_check():raise KeyboardInterrupt(_v)
			ws_sum=wb.add_worksheet(summary_sheet_name);hdrs=[_T,'应用层FTP上传/下载速率\n阈值以上占比(%)','应用层FTP上传/下载速率\n5M/100M以下占比(%)','应用层平均上传/下载\n速率(Mbps)','削峰应用层平均上传/下载\n速率(Mbps)','上行/下行削峰TOP10%\n峰值速率',q,u,_M]
			for(ci,h)in enumerate(hdrs):ws_sum.write(0,ci,h,fmt_hdr_s)
			for ci in range(9):ws_sum.set_column(ci,ci,20)
			ws_sum.merge_range(1,0,1,8,'▼ 所有识别业务',fmt_ta)
			for ri in range(len(df_summary_all)):
				for ci in range(9):
					cv=df_summary_all.iloc[ri,ci]if ci<len(df_summary_all.columns)else _A
					if isinstance(cv,float)and pd.isna(cv):cv=_A
					_ws_write(ws_sum,ri+2,ci,cv,fmt_d_s)
			all_data_rows=len(df_summary_all);comp_start_excel=all_data_rows+4;comp_start0=comp_start_excel-1;ws_sum.merge_range(comp_start0,0,comp_start0,8,'▼ 仅完整轮次（代码侧6业务齐全）',fmt_tc)
			for ri in range(len(df_summary_complete)):
				for ci in range(9):
					cv=df_summary_complete.iloc[ri,ci]if ci<len(df_summary_complete.columns)else _A
					if isinstance(cv,float)and pd.isna(cv):cv=_A
					_ws_write(ws_sum,ri+comp_start0+1,ci,cv,fmt_d_s)
			last_row_sum=comp_start0+len(df_summary_complete)
			for r in range(0,last_row_sum+2):ws_sum.set_row(r,22)
			ws_sum.freeze_panes(2,0)
			def _fmt_vals(vals,prec=2):
				if vals is _A or len(vals)==0:return'[]'
				vs=[round(v,prec)if prec>0 else int(round(v))for v in vals]
				if len(vs)<=20:return _l.join(str(v)for v in vs)
				return _l.join(str(v)for v in vs[:5])+_Aq+_l.join(str(v)for v in vs[-5:])
			def _add_comments(sub_detail,row_offset):
				'row_offset: 1-based Excel 行偏移（与原代码一致）'
				for(bi,biz)in enumerate(sorted(DL_BIZ|UL_BIZ)):
					if cancel_check and cancel_check():raise KeyboardInterrupt(_v)
					sub=sub_detail[sub_detail[C]==biz].dropna(subset=[_C]).sort_values(_C)
					if len(sub)==0:continue
					dl_cn=[c for c in sub.columns if _A0 in c and _AT in c];ul_cn=[c for c in sub.columns if _A0 in c and _AU in c]
					if biz in DL_BIZ and dl_cn:cn=dl_cn[0]
					elif biz in UL_BIZ and ul_cn:cn=ul_cn[0]
					else:continue
					rates=pd.to_numeric(sub[cn],errors=_Q).fillna(0)/1e6
					if len(rates)>1:
						_peak=rates.max()
						if _peak>0:
							_th=max(_peak*.1,5.);_mask=rates>=_th;_active_idx=rates.index[_mask]
							if len(_active_idx)>0:rates=rates.loc[_active_idx[0]:_active_idx[-1]]
					rates=rates[rates>0];ts2=sub[_C];durs=[]
					if len(ts2)>0:
						s1=ts2.iloc[0];p=s1
						for t in ts2.iloc[1:]:
							if(t-p).total_seconds()>5:durs.append((p-s1).total_seconds());s1=t
							p=t
						durs.append((p-s1).total_seconds())
					is_dl=biz in DL_BIZ;lim=dl_clip if is_dl else ul_clip;lo=100 if is_dl else 5;n=len(rates);sr=sub.index;r_start=sr[0]+2;r_end=sr[-1]+2;row_idx_excel=bi+row_offset;row0=row_idx_excel-1;df_sum_ref=df_summary_all if row_offset==3 else df_summary_complete
					if bi>=len(df_sum_ref):continue
					for cell_col in range(2,8):
						col0=cell_col-1;cv=df_sum_ref.iloc[bi,col0]if col0<len(df_sum_ref.columns)else _A
						if cv is _A or str(cv)in(F,_A7,R,''):continue
						cp=cell_col-2
						if cp in(0,1):
							if cp==0:mol=[r for r in rates if r>lim];cnt_label=f"计数(rate>{lim})={len(mol)}";op_cn=f"COUNT(rate > {lim}) / COUNT(all) x 100\n= 统计速率大于{lim}的采样数 / 总采样数 x 100"
							else:mol=[r for r in rates if r<lo];cnt_label=f"计数(rate<{lo})={len(mol)}";op_cn=f"COUNT(rate < {lo}) / COUNT(all) x 100\n= 统计速率小于{lo}的采样数 / 总采样数 x 100"
							calc=f"= {len(mol)} / {n} x 100"
						elif cp in(2,3,4):
							if cp==2:mol=list(rates);sum_v=sum(rates);sum_label=f"求和={sum_v:.1f}";cnt_label=f"计数={n}";op_cn=f"SUM(rate) / COUNT(all)\n= 速率求和 / 总采样数";calc=f"= {sum_v:.1f} / {n}"
							elif cp==3:clipped=rates.clip(upper=lim);mol=list(clipped);sum_v=sum(clipped);sum_label=f"削峰后求和={sum_v:.1f}";cnt_label=f"计数={n}";op_cn=f"SUM(CLIP(rate, 0, {lim})) / COUNT(all)\n= 削峰后速率求和 / 总采样数";calc=f"= {sum_v:.1f} / {n}"
							else:clipped_sv=rates.clip(upper=lim).sort_values(ascending=_D);tn=max(1,int(len(clipped_sv)*.1));mol=list(clipped_sv.head(tn));sum_v=sum(mol);sum_label=f"TOP10%求和(已削峰)={sum_v:.1f}";cnt_label=f"TOP10%计数={len(mol)}";op_cn=f"SUM(CLIP(rate,0,{lim}) TOP10%) / COUNT(TOP10%)\n= 削峰后前10%速率求和 / 前10%个数";calc=f"= {sum_v:.1f} / {len(mol)}"
						else:continue
						label_str=f"{sum_label}; {cnt_label}"if'sum_label'in locals()else cnt_label;txt=f"""公式: {op_cn}
计算过程: {calc}
结果: {cv}

{label_str}
值: {_fmt_vals(mol)}

分子来源: {cn} 列, 第 {r_start}~{r_end} 行 (详见详细过程 {_gcl(cell_col)}{r_start})
原始数据行数: {len(sub)}""";ws_sum.write_comment(row0,col0,txt,{_B0:500,_B1:220})
				for(bi,biz)in enumerate(sorted(DL_BIZ|UL_BIZ)):
					sub=sub_detail[sub_detail[C]==biz].dropna(subset=[_C]).sort_values(_C)
					if len(sub)==0:continue
					ts2=sub[_C];dur_vals=[]
					if len(ts2)>0:
						s1=ts2.iloc[0];p=s1
						for t in ts2.iloc[1:]:
							if(t-p).total_seconds()>5:dur_vals.append((p-s1).total_seconds());s1=t
							p=t
						dur_vals.append((p-s1).total_seconds())
					sr=sub.index;r_start=sr[0]+2;r_end=sr[-1]+2;row_idx_excel=bi+row_offset;row0=row_idx_excel-1;df_sum_ref=df_summary_all if row_offset==3 else df_summary_complete
					if bi>=len(df_sum_ref):continue
					for cell_col in[7,8]:
						col0=cell_col-1;cv=df_sum_ref.iloc[bi,col0]if col0<len(df_sum_ref.columns)else _A
						if cv is _A or str(cv)in(F,_A7,R,''):continue
						if cell_col==7:txt=f"""AVERAGE(duration)
= 业务段时间平均值
计数={len(dur_vals)}
= {cv}

值: {_fmt_vals(dur_vals)}

来源: 业务时长(秒) 列, 第 {r_start}~{r_end} 行 (详见详细过程 col_H{r_start})"""
						else:txt=f"""MEDIAN(duration)
= 业务段时间中位数
计数={len(dur_vals)}
= {cv}

值: {_fmt_vals(dur_vals)}

来源: 业务时长(秒) 列, 第 {r_start}~{r_end} 行"""
						ws_sum.write_comment(row0,col0,txt,{_B0:420,_B1:180})
			_add_comments(df_detail,3);_add_comments(df_detail_complete,comp_start_excel+1)
		log('  写入分业务RLC速率汇总...',92);ws_rlc=wb.add_worksheet('分业务RLC速率汇总');_rlc_biz_short={_L:_L,_K:_K,_U:m,_m:n,_R:k,_g:l};_rlc_headers=['联通业务','联通下行RLC吞吐率(Mbps)','联通上行RLC吞吐率(Mbps)','电信业务','电信下行RLC吞吐率(Mbps)','电信上行RLC吞吐率(Mbps)']
		for(ci,h)in enumerate(_rlc_headers):ws_rlc.write(0,ci,h,fmt_hdr_s)
		_fmt_rlc=wb.add_format({_A9:'0.00',_W:_Y,_X:_Z,_c:1});_fmt_rlc_hdr=wb.add_format({_AM:_B,_W:_Y,_X:_Z,_p:U,_c:1,_AV:_B})
		def _calc_rlc_avg(detail,biz):
			'从detail DataFrame计算某业务的RLC均值(Mbps), 返回(dl_avg, ul_avg, 有值)';_biz_rows=detail[detail[C]==biz]if C in detail.columns else pd.DataFrame();dl_avg,ul_avg=_A,_A
			if len(_biz_rows)==0:return dl_avg,ul_avg
			dl_cn=o;ul_cn=p
			if dl_cn in _biz_rows.columns and biz in DL_BIZ:
				_rates=pd.to_numeric(_biz_rows[dl_cn],errors=_Q).dropna()
				if len(_rates)>0:dl_avg=round(float(_rates.mean())/1e6,2)
			if ul_cn in _biz_rows.columns and biz in UL_BIZ:
				_rates=pd.to_numeric(_biz_rows[ul_cn],errors=_Q).dropna()
				if len(_rates)>0:ul_avg=round(float(_rates.mean())/1e6,2)
			return dl_avg,ul_avg
		_rlc_data={}
		for op in operators_list:
			if op not in all_results:continue
			_rlc_data[op]={};_detail=all_results[op][a]
			for biz in BIZ_ORDER:dl,ul=_calc_rlc_avg(_detail,biz);_rlc_data[op][biz]=dl,ul
		ri=1
		for biz in BIZ_ORDER:
			short_name=_rlc_biz_short.get(biz,biz);_lt_data=_rlc_data.get(_r,{}).get(biz,(_A,_A));ws_rlc.write(ri,0,short_name,fmt_d_s)
			if _lt_data[0]is not _A:ws_rlc.write_number(ri,1,_lt_data[0],_fmt_rlc)
			if _lt_data[1]is not _A:ws_rlc.write_number(ri,2,_lt_data[1],_fmt_rlc)
			_dx_data=_rlc_data.get(_s,{}).get(biz,(_A,_A));ws_rlc.write(ri,3,short_name,fmt_d_s)
			if _dx_data[0]is not _A:ws_rlc.write_number(ri,4,_dx_data[0],_fmt_rlc)
			if _dx_data[1]is not _A:ws_rlc.write_number(ri,5,_dx_data[1],_fmt_rlc)
			ri+=1
		ws_rlc.set_column(0,0,14);ws_rlc.set_column(1,2,26);ws_rlc.set_column(3,3,14);ws_rlc.set_column(4,5,26);ws_rlc.freeze_panes(1,0);log('  分业务RLC速率汇总完成',93)
	log(f"完成! {out}",100)
	if os.path.exists(_out_tmp):os.replace(_out_tmp,out)
	for _d in _tmp_dirs:
		try:shutil.rmtree(_d)
		except Exception:pass
	return out
def _detect_operator(filepath,df=_A):
	'从文件路径和数据内容判断运营商：联通/电信\n\n    规则：\n    1. 优先从路径名判断（电信/联通），避免"电信区域-联通用户"误判\n    2. 如果路径名无法判断且有数据df，通过QCI辅助（7→电信，5/6→联通）\n    3. 如果仍无法判断，返回\'未知\'\n    ';low=filepath.lower();has_dx=_s in low or'dx'in low;has_lt=_r in low or'lt'in low or'cu'in low
	if has_lt and not has_dx:return _r
	if has_dx and not has_lt:return _s
	if has_lt and has_dx:
		if'联通用户'in low or'联通-'in low:return _r
		if'电信用户'in low or'电信-'in low:return _s
		return _r if low.rfind(_r)>low.rfind(_s)else _s
	if df is not _A and _k in df.columns:
		qci_vals=df[_k].dropna().unique()
		if 7 in qci_vals:return _s
		if 5 in qci_vals or 6 in qci_vals:return _r
	return _A3
def _detect_operator_by_path(filepath):
	'从文件夹/文件名判断运营商（增强版，支持数字编号目录名）\n\n    规则：\n    1. 文件夹名含"联通"或"电信"→直接判断\n    2. 文件夹名含数字编号如"07191845"搭配"联通-0719"→联通\n    3. 也可通过mmf里的QCI判断（有7→电信，有5/6→联通）\n    ';low=filepath.lower()
	if _s in low or'dx'in low:return _s
	if _r in low or'lt'in low or'cu'in low:return _r
	return _A3
def _extract_call_type_from_path(filepath):
	'从文件夹名提取业务类型/主被叫信息\n\n    规则：\n    - VoNR主叫 → voNR-主叫\n    - VoNR被叫 → voNR-被叫\n    - ViNR微信视频-主叫 → viNR-主叫\n    - ViNR微信视频-被叫 → viNR-被叫\n    - 数据-553 → 数据业务（或含"user_common_monitoring"/".mmf"）\n    注意：不因路径中有"最新数据"/"数据文件"之类的词而误判\n    '
	if not filepath:return''
	path_lower=str(filepath).lower()
	if'vonr'in path_lower or'vnr'in path_lower:
		if'主叫'in path_lower:return'voNR-主叫'
		elif'被叫'in path_lower:return'voNR-被叫'
	if'vinr'in path_lower or'微信视频'in path_lower:
		if'主叫'in path_lower:return'viNR-主叫'
		elif'被叫'in path_lower:return'viNR-被叫'
	if path_lower.endswith(_AW)or _B2 in path_lower:return _AN
	if _AN in path_lower or'/数据/'in path_lower or'数据-'in path_lower:return _AN
	return''
def _extract_caller_type_from_path(filepath):
	'从文件夹名提取主被叫（用于按文件名表）\n\n    规则：\n    - 含"主叫" → 主叫（优先）\n    - 含"被叫" → 被叫\n    - 数据业务（精确匹配: ends with .mmf, 或含"数据-"等） → 空\n    注意：不因路径中有"最新数据"/"数据文件"之类的词而误判\n    '
	if not filepath:return''
	path_lower=str(filepath).lower()
	if path_lower.endswith(_AW)or _B2 in path_lower:return''
	if _AN in path_lower or'/数据/'in path_lower or'数据-'in path_lower:return''
	if'主叫'in path_lower:return'主叫'
	if'被叫'in path_lower:return'被叫'
	return''
def _detect_file_type(filename):
	'判断文件类型：voice/voice_vqi 或 data/monitoring\n\n    通过文件名和列名综合判断：\n    - 如果文件名包含nr_virtual_ue_trace或.tmf，或有MessageType列 → 返回"语音"\n    - 否则 → 返回"数据"\n    ';name=filename.lower()
	if'nr_virtual_ue_trace'in name or name.endswith('.tmf'):return'语音'
	if _B2 in name or name.endswith(_AW):return'数据'
	return'数据'
_VQI_RE=re.compile('DlE2eVqi=(\\d+)')
_RSRP_RE=re.compile('servRSRP=(-?\\d+)(?:,(-?\\d+))?')
def _build_anno_text(col_name,row_op,ri,df_voice_out,df_vqi,voice_records,vqi_records,rsrp_records,df_file_stats,op_voice_cache=_A,op_vqi_cache=_A,op_rsrp_cache=_A):
	'构建汇总-数据和语音表的详细批注（使用缓存避免重复过滤）'
	if col_name==_O:
		op_vrs=(op_voice_cache or{}).get(row_op,[vr for vr in voice_records if vr.get(_E)==row_op]);op_vrs_with_val=[vr for vr in op_vrs if vr.get(_O)!='']
		if op_vrs_with_val:
			details=[]
			for vr in op_vrs_with_val:rnd=vr.get(_H,'');start=vr.get(_G,'');end=vr.get(_I,'');dur=vr.get(_O,'');details.append(f"轮次{rnd}: {start}~{end}, 时长={dur}s")
			detail_txt=_a.join(details);vals=[float(vr[_O])for vr in op_vrs_with_val];avg=round(sum(vals)/len(vals),2);return f"""公式: 平均呼叫建立时长(s)
数据来源: 语音指标统计 -> 呼叫建立时长(s) 列
计算过程: 对 {row_op} 所有轮次的呼叫建立时长取平均值
分子(各轮次呼叫建立时长):
{detail_txt}
分母(轮次数): {len(vals)}
计算结果: {avg}"""
	elif col_name==_AA:
		op_vqi=(op_vqi_cache or{}).get(row_op,[r for r in vqi_records if r.get(_E)==row_op]);total=len(op_vqi);quality=sum(1 for r in op_vqi if r.get(_w)==_i)
		if total>0:
			details=[]
			for r in op_vqi:t=r.get(_x,'');mos=r.get(_AX,'');details.append(f"{t}, mos={mos}")
			detail_txt=_a.join(details);return f"""公式: 优质通话占比(MOS3.5)
数据来源: MOS3.5计算明细 -> 是否优质(MOS>3.5) 列
计算过程: 优质通话数 / 总通话数
分子(所有通话):
{detail_txt}
分母(总通话数): {total}
计算结果: {round(quality/total*100,2)}%"""
	elif col_name==_AH:
		op_rsrp=(op_rsrp_cache or{}).get(row_op,[r for r in rsrp_records if r.get(_E)==row_op]);op_rsrp_in_call=[r for r in op_rsrp if r.get(_AO)];vals=[r[_y]for r in op_rsrp_in_call]
		if vals:nums=_l.join([str(v)for v in vals]);avg=round(sum(vals)/len(vals),2);return f"""公式: 服务小区平均RSRP(dBm)(通话时)
数据来源: Detail Info -> servRSRP 列(RRC_MEAS_RPRT行，仅通话时段)
计算过程: 对 {row_op} 通话时段内的RSRP值取平均
分子(各RSRP值): {nums}
分母(RSRP数量): {len(vals)}
计算结果: {avg}"""
	return f"公式: {col_name}\n计算过程: 从原始数据统计\n\n数据来源: {col_name} 列"
def _build_file_anno_text(col_name,fname,voice_records,vqi_records,rsrp_records,fname_voice_cache=_A,fname_vqi_cache=_A,fname_rsrp_cache=_A):
	'构建按文件名表的详细批注（使用缓存避免重复过滤）'
	if col_name==_Af:
		f_vrs=(fname_voice_cache or{}).get(fname,[vr for vr in voice_records if vr.get(_F)==fname])
		if f_vrs:
			rounds_detail=[]
			for vr in f_vrs:rnd=vr.get(_H,'');start=vr.get(_G,'');end=vr.get(_I,'');rounds_detail.append(f"轮次{rnd}: {start} ~ {end}")
			detail_txt=_a.join(rounds_detail);return f"公式: 通话轮次\n数据来源: 语音指标统计 -> 轮次/开始时间/结束时间 列\n计算过程: 从 {fname} 的语音通话记录中统计\n详情:\n{detail_txt}"
	elif col_name==_Ag:
		f_vrs=(fname_voice_cache or{}).get(fname,[vr for vr in voice_records if vr.get(_F)==fname]);f_vrs_with_val=[vr for vr in f_vrs if vr.get(_O)!='']
		if f_vrs_with_val:
			details=[]
			for vr in f_vrs_with_val:rnd=vr.get(_H,'');start=vr.get(_G,'');end=vr.get(_I,'');dur=vr.get(_O,'');details.append(f"轮次{rnd}: {start}~{end}, 时长={dur}s")
			detail_txt=_a.join(details);vals=[float(vr[_O])for vr in f_vrs_with_val];avg=round(sum(vals)/len(vals),2);return f"""公式: 平均呼叫建立时长(s)
数据来源: 语音指标统计 -> 呼叫建立时长(s) 列
计算过程: 对 {fname} 各轮次的呼叫建立时长取平均
分子(各轮次呼叫建立时长):
{detail_txt}
分母(轮次数): {len(vals)}
计算结果: {avg}"""
	elif col_name==_AA:
		f_vqi=(fname_vqi_cache or{}).get(fname,[r for r in vqi_records if r.get(_F)==fname]);total=len(f_vqi);quality=sum(1 for r in f_vqi if r.get(_w)==_i)
		if total>0:
			details=[]
			for r in f_vqi:t=r.get(_x,'');mos=r.get(_AX,'');details.append(f"{t}, mos={mos}")
			detail_txt=_a.join(details);return f"""公式: 优质通话占比(MOS3.5)
数据来源: MOS3.5计算明细 -> 是否优质(MOS>3.5) 列
计算过程: 优质通话数 / 总通话数
分子(所有通话):
{detail_txt}
分母(总通话数): {total}
计算结果: {round(quality/total*100,2)}%"""
	elif col_name==_AH:
		f_rsrp=(fname_rsrp_cache or{}).get(fname,[r for r in rsrp_records if r.get(_F)==fname]);f_rsrp_in_call=[r for r in f_rsrp if r.get(_AO)];vals=[r[_y]for r in f_rsrp_in_call]
		if vals:nums=_l.join([str(v)for v in vals]);avg=round(sum(vals)/len(vals),2);return f"""公式: 服务小区平均RSRP(dBm)(通话时)
数据来源: Detail Info -> servRSRP 列(RRC_MEAS_RPRT行，仅通话时段)
计算过程: 对 {fname} 通话时段内的RSRP值取平均
分子(各RSRP值): {nums}
分母(RSRP数量): {len(vals)}
计算结果: {avg}"""
	return f"公式: {col_name}\n数据来源: {col_name} 列"
def _build_data_biz_anno(col_name,fname,data_file_info):
	'为数据业务指标列生成批注（下行/上行峰值、均值等）';info=data_file_info.get(fname,{})
	if not info:return
	row_cnt=info.get(_AK,0);_dl_peak=info.get(_A1,'');_dl_avg=info.get(_A2,'');_ul_peak=info.get(_A5,'');_ul_avg=info.get(_A6,'');formulas={_AB:f"公式: 下行RLC速率降序排列后取前10%采样点求平均\nTop10%均值={_dl_peak}\n分母: 总采样数={row_cnt}\n来源: 数据文件 RLC Downlink 列",_AC:f"公式: AVERAGE(下行RLC速率采样, 不含0值)\n均值={_dl_avg}\n分母: 总采样数={row_cnt}\n来源: 数据文件 RLC Downlink 列",_AD:f"公式: 上行RLC速率降序排列后取前10%采样点求平均\nTop10%均值={_ul_peak}\n分母: 总采样数={row_cnt}\n来源: 数据文件 RLC Uplink 列",_AE:f"公式: AVERAGE(上行RLC速率采样, 不含0值)\n均值={_ul_avg}\n分母: 总采样数={row_cnt}\n来源: 数据文件 RLC Uplink 列"};return formulas.get(col_name)
def _build_op_anno_text(col_name,op,voice_records,vqi_records,rsrp_records,df_file_stats,op_voice_cache=_A,op_vqi_cache=_A,op_rsrp_cache=_A):
	'构建按运营商表的详细批注（使用缓存避免重复过滤）'
	if col_name==_Af:
		op_vrs=(op_voice_cache or{}).get(op,[vr for vr in voice_records if vr.get(_E)==op])
		if op_vrs:
			rounds_detail=[]
			for vr in op_vrs:rnd=vr.get(_H,'');start=vr.get(_G,'');end=vr.get(_I,'');rounds_detail.append(f"轮次{rnd}: {start} ~ {end}")
			detail_txt=_a.join(rounds_detail);return f"公式: 通话轮次\n数据来源: 语音指标统计 -> 轮次/开始时间/结束时间 列\n计算过程: 从 {op} 的语音通话记录中统计\n详情:\n{detail_txt}"
	elif col_name==_Ag:
		op_vrs=(op_voice_cache or{}).get(op,[vr for vr in voice_records if vr.get(_E)==op]);op_vrs=[vr for vr in op_vrs if vr.get(_O)!=''];vals=[float(vr[_O])for vr in op_vrs]
		if vals:nums=_l.join([str(v)for v in vals]);avg=round(sum(vals)/len(vals),2);return f"""公式: 平均呼叫建立时长(s)
数据来源: 语音指标统计 -> 呼叫建立时长(s) 列
计算过程: 对 {op} 各轮次的呼叫建立时长取平均
分子(各轮次呼叫建立时长): {nums}
分母(轮次数): {len(vals)}
计算结果: {avg}"""
	elif col_name==_AA:
		op_vqi=(op_vqi_cache or{}).get(op,[r for r in vqi_records if r.get(_E)==op]);total=len(op_vqi);quality=sum(1 for r in op_vqi if r.get(_w)==_i)
		if total>0:return f"""公式: 优质通话占比(MOS3.5)
数据来源: MOS3.5计算明细 -> 是否优质(MOS>3.5) 列
计算过程: 优质通话数 / 总通话数
分子(优质通话数): {quality}
分母(总通话数): {total}
计算结果: {round(quality/total*100,2)}%"""
	elif col_name=='服务小区平均RSRP(dBm)':
		op_rsrp=(op_rsrp_cache or{}).get(op,[r for r in rsrp_records if r.get(_E)==op]);vals=[r[_y]for r in op_rsrp]
		if vals:nums=_l.join([str(v)for v in vals]);avg=round(sum(vals)/len(vals),2);return f"""公式: 服务小区平均RSRP(dBm)
数据来源: Detail Info -> servRSRP 列(RRC_MEAS_RPRT行)
计算过程: 对 {op} 的所有RSRP值取平均
分子(各RSRP值): {nums}
分母(RSRP数量): {len(vals)}
计算结果: {avg}"""
	return f"公式: {col_name}\n数据来源: {col_name} 列"
def _extract_vqi_vale(detail_info):
	'从Detail Info提取DlE2eVqi值，排除65535'
	if pd.isna(detail_info):return
	m=_VQI_RE.search(str(detail_info))
	if m:val=int(m.group(1));return val if val!=65535 else _A
def _extract_serv_rsrp(detail_info):
	'从Detail Info提取servRSRP值并返回最大值\n\n    格式示例: servRSRP=-96,-91 或 servRSRP=-96\n    如果有两个值，取最大的那个（例如-96,-91中取-91，因为-91 > -96）\n    如果只有一个值，直接返回该值\n    '
	if pd.isna(detail_info):return
	m=_RSRP_RE.search(str(detail_info))
	if m:
		val1=int(m.group(1));val2=int(m.group(2))if m.group(2)else _A
		if val2 is not _A:return max(val1,val2)
		return val1
def process_vqi(files,callback=_A,cancel_check=_A,progress_cb=_A,time_filter=_A,time_range=_A,merge_raw=_D,add_annotations=_B,phone_trace_map=_A):
	'VQI处理函数：自动分类语音/数据文件，生成语音统计输出\n\n    参数:\n        files: 输入文件列表\n        callback: 日志回调函数\n        cancel_check: 取消检查函数\n        progress_cb: 进度回调函数\n        time_filter: 时间段过滤 (use_time, start_time, end_time)\n        time_range: 时间起止范围，用于汇总表。若为None则从数据自动计算\n    ';A1='数据量(MB)';A0='方向(上下行)';z='业务类型(微信小/微信大/商店小/商店大)';y='采样数';x='最小速率';u='最大速率';q='前10%峰值';p='业务时长(s)';o='前10%峰值速率';n='author';l='文件类型';k='结束时间_vr';j='_dur_reason';h='SIP_RSP_200';g='SIP_REQ_BYE';d='数据_微信商店业务原始数据';c='数据_FTP上传原始数据';b='数据_FTP下载原始数据';a='数据_业务时长原始数据';Z='数据_速率原始数据';Y='类型(主叫/viNR-主叫/viNR-被叫/voNR-主叫/voNR-被叫/数据业务)';X='业务起止时间';W='时间范围';V='行数';U='_ul_rlc';T='_dl_rlc';S='quality';R='按文件名';Q='total';P='商店小app平均速率(Mbps)';O='商店小app平均时延(s)';N='商店大app平均速率(Mbps)';M='商店大app平均时延(s)';L='微信小文件发送平均速率(Mbps)';K='微信小文件发送平均时延(s)';J='微信大文件发送平均速率(Mbps)';I='微信大文件上传中位数时长(s)';H='服务小区平均RSRP(dBm)(全部)';G='DlE2eVqi原始值';F='BYE时间';E='INVITE时间';D='否';C='是否完整通话';B='主被叫';A='文件名'
	def log(msg,pct=_A):
		if cancel_check and cancel_check():raise KeyboardInterrupt(_v)
		import time as _time;_t=_time.time()
		if not hasattr(log,_BB):log.last_time=_t
		elapsed=_t-log.last_time if pct is not _A else 0;prefix=f"({elapsed:.1f}秒) "if pct is not _A and elapsed>0 else'';print(msg)
		if callback:callback(f"{prefix}{msg}")
		if pct is not _A and progress_cb:progress_cb(pct)
		log.last_time=_t
	import zipfile,tempfile,shutil,glob;_tmp_dirs=[]
	def _expand_zip(zip_path):
		_tmp=tempfile.mkdtemp(prefix='vqi_unzip_');_tmp_dirs.append(_tmp);_xs=[]
		try:
			with zipfile.ZipFile(zip_path)as _zf:_zf.extractall(_tmp)
			_xs=glob.glob(os.path.join(_tmp,'**','*.xlsx'),recursive=_B)+glob.glob(os.path.join(_tmp,'**','*.xls'),recursive=_B)+glob.glob(os.path.join(_tmp,'**','*.mmf'),recursive=_B)+glob.glob(os.path.join(_tmp,'**','*.tmf'),recursive=_B);_inner_zips=glob.glob(os.path.join(_tmp,'**','*.zip'),recursive=_B)
			for _iz in _inner_zips:_xs.extend(_expand_zip(_iz))
		except Exception as _e:log(f"  解压失败 {os.path.basename(zip_path)}: {_e}")
		return _xs
	all_files=[]
	for _f in files or[]:
		if str(_f).lower().endswith(_Aa):_xs=_expand_zip(_f);log(f"  解压 {os.path.basename(_f)}: 找到 {len(_xs)} 个文件(含嵌套)",5);all_files.extend(_xs)
		else:all_files.append(_f)
	if not all_files:log('无有效文件',100);return
	log(f"共 {len(all_files)} 个文件",5);voice_files=[];data_files=[];data_files_skipped=0
	for f in all_files:
		try:
			df_test=_read_file(f,nrows=3);has_msg=_has_message_type_col(df_test);msg_type_status='[OK] MessageType列'if has_msg else'[NO] 无MessageType列';log(f"  文件 {os.path.basename(f)}: {msg_type_status}",6)
			if has_msg:voice_files.append(f)
			else:data_files.append(f)
		except Exception as e:log(f"  文件 {os.path.basename(f)}: 读取失败 - {e}",6);data_files_skipped+=1
	log(f"  语音文件: {len(voice_files)} 个, 数据文件: {len(data_files)} 个, 跳过的数据文件: {data_files_skipped} 个",8);voice_dfs=[];voice_file_info={}
	for(fi,f)in enumerate(voice_files):
		log(f"  读取语音文件 {fi+1}/{len(voice_files)}: {os.path.basename(f)}",10+int(10*(fi+1)/max(len(voice_files),1)));operator=_detect_operator(f)
		try:df_i=_read_file(f)
		except Exception:log(f"    读取失败: {f}",10);continue
		if _Ao in df_i.columns and _b not in df_i.columns:df_i=df_i.rename(columns={_Ao:_b})
		elif _Ap in df_i.columns and _b not in df_i.columns:df_i=df_i.rename(columns={_Ap:_b})
		df_i[_C]=_parse_time_vec(df_i[_e]);df_i[_E]=operator;df_i[_T]='语音业务';filename_base=os.path.basename(f);rel_path=f;work_dir=os.getcwd()
		if rel_path.startswith(work_dir):rel_path=rel_path[len(work_dir)+1:]
		df_i[_F]=rel_path;voice_dfs.append(df_i);times=df_i[_C].dropna();time_range_str=''
		if len(times)>0:time_range_str=f"{times.min().strftime(_FMT_DT)} ~ {times.max().strftime(_FMT_DT)}"
		voice_file_info[rel_path]={_h:operator,_AK:len(df_i),_AG:time_range_str,'time_start':times.min()if len(times)>0 else _A,'time_end':times.max()if len(times)>0 else _A,_o:f}
	if voice_dfs:df_voice_raw=pd.concat(voice_dfs,ignore_index=_B);df_voice_raw=df_voice_raw.sort_values(_C,kind=_Ar).reset_index(drop=_B)
	else:df_voice_raw=pd.DataFrame()
	log(f"语音数据合并完成: {len(df_voice_raw)} 行",22);log('处理数据文件...',23);data_dfs=[];data_file_info={}
	for(fi,f)in enumerate(data_files):
		if cancel_check and cancel_check():raise KeyboardInterrupt(_v)
		log(f"  读取数据文件 {fi+1}/{len(data_files)}: {os.path.basename(f)}",24+int(5*(fi+1)/max(len(data_files),1)));operator=_detect_operator(f)
		try:df_i=_read_file(f)
		except Exception:log(f"    读取失败: {f}",24);continue
		has_time=_e in df_i.columns;has_qci=_k in df_i.columns;has_rlc=any(_A0 in c and(_AT in c or _AU in c)for c in df_i.columns)
		if not(has_time and(has_qci or has_rlc)):log(f"    跳过非数据文件(无Time/QCI/RLC列)",24);data_files_skipped+=1;continue
		df_i[_C]=_parse_time_vec(df_i[_e]);df_i[_E]=operator;df_i[_T]=_AN;rlc_dl=[c for c in df_i.columns if _A0 in c and _AT in c];rlc_ul=[c for c in df_i.columns if _A0 in c and _AU in c]
		if rlc_dl:df_i[T]=pd.to_numeric(df_i[rlc_dl[0]],errors=_Q).fillna(0)/1e6
		else:df_i[T]=0
		if rlc_ul:df_i[U]=pd.to_numeric(df_i[rlc_ul[0]],errors=_Q).fillna(0)/1e6
		else:df_i[U]=0
		data_dfs.append(df_i);dl_vals=df_i[T].dropna().values;ul_vals=df_i[U].dropna().values;dl_nonzero=dl_vals[dl_vals>0];ul_nonzero=ul_vals[ul_vals>0];dl_peak=0
		if len(dl_nonzero)>0:tn=max(1,int(len(dl_nonzero)*.1));dl_sorted=np.sort(dl_nonzero)[::-1];dl_peak=round(min(float(np.mean(dl_sorted[:tn])),1000),2)
		dl_avg=0
		if len(dl_nonzero)>0:dl_avg=round(min(float(np.mean(dl_nonzero)),1000),2)
		ul_peak=0
		if len(ul_nonzero)>0:tn=max(1,int(len(ul_nonzero)*.1));ul_sorted=np.sort(ul_nonzero)[::-1];ul_peak=round(min(float(np.mean(ul_sorted[:tn])),200),2)
		ul_avg=0
		if len(ul_nonzero)>0:ul_avg=round(min(float(np.mean(ul_nonzero)),200),2)
		times=df_i[_C].dropna();time_range_str=''
		if len(times)>0:time_range_str=f"{times.min().strftime(_FMT_DT)} ~ {times.max().strftime(_FMT_DT)}"
		data_file_info[os.path.basename(f)]={_h:operator,_AK:len(df_i),_AG:time_range_str,_A1:dl_peak,_A2:dl_avg,_A5:ul_peak,_A6:ul_avg,_o:f}
	for(fname,info)in data_file_info.items():trace_id=_extract_trace_id(info.get(_o,''));info[_V]=_lookup_phone(trace_id,phone_trace_map)
	if data_dfs:df_data_raw=pd.concat(data_dfs,ignore_index=_B);df_data_raw=df_data_raw.sort_values(_C,kind=_Ar).reset_index(drop=_B)
	else:df_data_raw=pd.DataFrame()
	log(f"数据文件处理完成: {len(data_dfs)} 个文件, {len(df_data_raw)} 行",28);data_operator_indicators={};actual_ops=set(info[_h]for info in data_file_info.values())
	for op in sorted(actual_ops):
		op_data=[info for(fname,info)in data_file_info.items()if info[_h]==op]
		if op_data:dl_peaks=[d[_A1]for d in op_data if d[_A1]>0];dl_avgs=[d[_A2]for d in op_data if d[_A2]>0];ul_peaks=[d[_A5]for d in op_data if d[_A5]>0];ul_avgs=[d[_A6]for d in op_data if d[_A6]>0];data_operator_indicators[op]={_AB:round(np.mean(dl_peaks),2)if dl_peaks else'',_AC:round(np.mean(dl_avgs),2)if dl_avgs else'',_AD:round(np.mean(ul_peaks),2)if ul_peaks else'',_AE:round(np.mean(ul_avgs),2)if ul_avgs else''}
		else:data_operator_indicators[op]={_AB:'',_AC:'',_AD:'',_AE:''}
	if time_filter and len(df_voice_raw)>0:
		use_time,t_start,t_end=time_filter
		if use_time and t_start is not _A and t_end is not _A:t_start_t=t_start.toPython()if hasattr(t_start,_Ad)else t_start;t_end_t=t_end.toPython()if hasattr(t_end,_Ad)else t_end;df_voice_raw[_u]=df_voice_raw[_C].dt.time;mask=(df_voice_raw[_u]>=t_start_t)&(df_voice_raw[_u]<=t_end_t);df_voice_raw=df_voice_raw[mask].copy();df_voice_raw=df_voice_raw.drop(_u,axis=1);log(f"  时间段过滤: {t_start_t} ~ {t_end_t}, 剩余 {len(df_voice_raw)} 行",22)
	if _b in df_voice_raw.columns and len(df_voice_raw)>0:df_msg=df_voice_raw[df_voice_raw[_b].notna()&(df_voice_raw[_b].astype(str).str.strip()!='')].copy()
	else:df_msg=pd.DataFrame()
	log(f"  MessageType非空行数: {len(df_msg)}",28);log('识别语音通话...',35);voice_records=[];rnd=0
	if len(df_msg)>0 and _b in df_msg.columns and _E in df_msg.columns:
		operators_in_data=df_msg[_E].dropna().unique()
		for op in operators_in_data:
			if str(op)in(_A3,'',_A7):continue
			df_op=df_msg[df_msg[_E]==op].copy()
			if len(df_op)==0:continue
			log(f"  处理{op}信令({len(df_op)}行)...",36);invite_mask=df_op[_b].astype(str)=='SIP_REQ_INVITE';invite_positions=df_op.index[invite_mask].tolist();n_op=len(df_op);valid_invite_positions=[]
			for idx_val in invite_positions:
				if valid_invite_positions:
					prev_val=valid_invite_positions[-1]
					if abs((df_op.loc[idx_val,_C]-df_op.loc[prev_val,_C]).total_seconds())<1.:continue
				valid_invite_positions.append(idx_val)
			invite_positions=valid_invite_positions;log(f"  找到 {len(invite_positions)} 个有效{op}INVITE",36)
			for(ii,pos)in enumerate(invite_positions):
				t1=df_op.loc[pos,_C];t1_time=df_op.loc[pos,_e];operator=str(op);next_pos=invite_positions[ii+1]if ii+1<len(invite_positions)else df_op.index[-1]+1;end_j=next_pos-1
				if pos+1>end_j:full_range2=df_op.loc[pos+1:];bye2=full_range2[full_range2[_b].astype(str).isin([g,h])]if len(full_range2)>0 else full_range2.iloc[0:0];rnd+=1;voice_records.append({_H:rnd,_E:operator,_G:t1_time,_I:t1_time,_O:'',C:_i if len(bye2)>0 else D,B:_extract_caller_type_from_path(df_op.loc[pos].get(_F,'')),_F:df_op.loc[pos].get(_F,'')});continue
				search_range=df_op.loc[pos+1:end_j];r180=search_range[search_range[_b].astype(str)=='SIP_RSP_180']
				if len(r180)>0 and _C in r180.columns:t2=r180.iloc[0][_C]
				else:r183=search_range[search_range[_b].astype(str)=='SIP_RSP_183'];t2=r183.iloc[0][_C]if len(r183)>0 and _C in r183.columns else _A
				full_range=df_op.loc[pos+1:];bye=full_range[full_range[_b].astype(str).isin([g,h])];bye_time_raw=_A
				if len(bye)>0:bye_time_raw=bye.iloc[0][_e]if _e in bye.columns else _A;_bye_t=bye.iloc[0][_C]if _C in bye.columns else _A;last_time_raw=bye_time_raw
				else:last_time_raw=search_range.iloc[-1][_e]if len(search_range)>0 and _e in search_range.columns else t1_time
				call_setup_duration='';_dur_reason=''
				if t2 is not _A:call_setup_duration=round((t2-t1).total_seconds(),2)
				else:_dur_reason='无SIP_RSP_180/183(未经过振铃/会话进行阶段,可能被叫立即应答或拒绝)'
				is_complete=_i if len(bye)>0 else D;rnd+=1;voice_records.append({_H:rnd,_E:operator,_G:t1_time,_I:last_time_raw,E:t1_time,F:bye_time_raw if len(bye)>0 else'',B:_extract_caller_type_from_path(df_op.loc[pos].get(_F,'')),_O:call_setup_duration,j:_dur_reason,C:is_complete,_F:df_op.loc[pos].get(_F,'')})
	log(f"识别完成: {len(voice_records)} 轮语音通话",50);_file_rnd={}
	for vr in sorted(voice_records,key=lambda x:(x.get(_F,''),str(x.get(_G,'')))):sf=vr.get(_F,'');_file_rnd[sf]=_file_rnd.get(sf,0)+1;vr[_H]=_file_rnd[sf]
	voice_records.sort(key=lambda x:(x.get(_F,''),int(x.get(_H,0))));log('提取DlE2eVqi值...',55);vqi_records=[];detail_col=_A
	if len(df_voice_raw)>0:
		for col_name in['Detail Info','详细信息','DetailInfo','detail_info']:
			if col_name in df_voice_raw.columns:detail_col=col_name;break
	if detail_col and len(df_voice_raw)>0:
		detail_vals=df_voice_raw[detail_col].values;time_vals=df_voice_raw[_e].values if _e in df_voice_raw.columns else _A;op_vals=df_voice_raw[_E].values if _E in df_voice_raw.columns else _A;sf_vals=df_voice_raw[_F].values if _F in df_voice_raw.columns else _A
		for i in range(len(detail_vals)):
			v=detail_vals[i]
			if v is _A or pd.isna(v):continue
			s=str(v)
			if'DlE2eVqi='not in s:continue
			m=_VQI_RE.search(s)
			if m:
				vqi=int(m.group(1))
				if vqi==65535:continue
				mos=vqi/1e2;is_quality=_i if mos>3.5 else D;vqi_records.append({_x:time_vals[i]if time_vals is not _A else'',_E:op_vals[i]if op_vals is not _A else'',G:vqi,_AX:round(mos,2),_w:is_quality,_F:sf_vals[i]if sf_vals is not _A else''})
	_vr_lookup={}
	for vr in voice_records:
		sf=vr.get(_F,'');st=vr.get(_G,'');et=vr.get(_I,'')
		if sf and st:_vr_lookup.setdefault(sf,[]).append((st,et,vr.get(_H,'')))
	for r in vqi_records:
		sf=r.get(_F,'');t=r.get(_x,'')
		if sf and t:
			for(st,et,rnd)in _vr_lookup.get(sf,[]):
				if st<=t<=(et if et else st):r[_H]=rnd;r['开始时间_vr']=st;r[k]=et;break
	df_vqi=pd.DataFrame(vqi_records)
	if len(df_vqi)>0:total=len(df_vqi);quality=(df_vqi[_w]==_i).sum();quality_pct=round(quality/total*100,2)
	else:total,quality,quality_pct=0,0,0
	log('提取服务小区RSRP值...',58);rsrp_records=[];call_periods_by_file={}
	for vr in voice_records:
		fname=vr.get(_F,'')
		if not fname:continue
		if vr.get(C)!=_i:continue
		start_t=vr.get(_G,'');end_t=vr.get(_I,'')
		if start_t and end_t:
			if fname not in call_periods_by_file:call_periods_by_file[fname]=[]
			call_periods_by_file[fname].append((start_t,end_t))
	if detail_col and len(df_voice_raw)>0 and _b in df_voice_raw.columns:
		msg_types=df_voice_raw[_b].values;rrc_indices=[]
		for i in range(len(msg_types)):
			val=msg_types[i]
			if val is not _A and not pd.isna(val)and str(val)=='RRC_MEAS_RPRT':rrc_indices.append(i)
		log(f"  筛选到 {len(rrc_indices)} 个RRC_MEAS_RPRT行",59)
		if len(rrc_indices)>0:
			detail_vals=df_voice_raw[detail_col].values;time_vals=df_voice_raw[_e].values if _e in df_voice_raw.columns else _A;op_vals=df_voice_raw[_E].values if _E in df_voice_raw.columns else _A;sf_vals=df_voice_raw[_F].values if _F in df_voice_raw.columns else _A
			for idx in rrc_indices:
				dv=detail_vals[idx]
				if dv is _A or pd.isna(dv):continue
				m=_RSRP_RE.search(str(dv))
				if m:
					v1=int(m.group(1));v2=int(m.group(2))if m.group(2)else _A;rsrp_val=max(v1,v2)if v2 is not _A else v1;in_call=_D;sf=sf_vals[idx]if sf_vals is not _A else'';t_str=time_vals[idx]if time_vals is not _A else''
					if sf and t_str:
						periods=call_periods_by_file.get(sf,[])
						for(s_t,e_t)in periods:
							if s_t and e_t and s_t<=t_str<=e_t:in_call=_B;break
					rsrp_records.append({_x:t_str,_E:op_vals[idx]if op_vals is not _A else'',_y:rsrp_val,_F:sf,_AO:in_call})
	avg_rsrp=_A
	if rsrp_records:rsrp_vals=[r[_y]for r in rsrp_records];avg_rsrp=round(sum(rsrp_vals)/len(rsrp_vals),2);log(f"  提取到 {len(rsrp_records)} 个RSRP值，平均值: {avg_rsrp}",58)
	df_raw_out=df_voice_raw.copy()
	if _C in df_raw_out.columns:df_raw_out=df_raw_out.drop(_C,axis=1)
	if _u in df_raw_out.columns:df_raw_out=df_raw_out.drop(_u,axis=1)
	df_voice_out=pd.DataFrame(voice_records)
	if len(df_voice_out)==0:df_voice_out=pd.DataFrame(columns=[_H,_E,_G,_I,_O,C])
	if len(df_vqi)>0:summary_row=pd.DataFrame([{_x:'汇总',_E:'',G:'',_AX:'',_w:f"分子={quality} / 分母={total} = {quality_pct}%"}]);df_vqi_out=pd.concat([df_vqi,summary_row],ignore_index=_B)
	else:df_vqi_out=pd.DataFrame(columns=[_x,_E,G,_AX,_w])
	file_stats=[]
	for(fname_base,f_info)in voice_file_info.items():
		operator=f_info[_h];file_row_count=f_info[_AK];file_time_range_str=f_info[_AG];caller_type=_extract_caller_type_from_path(f_info.get(_o,''));file_voice_records=[vr for vr in voice_records if vr.get(_F)==fname_base];file_rounds=len(file_voice_records);setup_vals=[]
		for vr in file_voice_records:
			if vr.get(_O)!=''and pd.notna(vr.get(_O)):
				try:setup_vals.append(float(vr[_O]))
				except:pass
		file_avg_setup=round(np.mean(setup_vals),2)if setup_vals else'';file_vqi_records=[r for r in vqi_records if r.get(_F)==fname_base];file_quality_pct=''
		if file_vqi_records:
			total_file=len(file_vqi_records);quality_file=sum(1 for r in file_vqi_records if r.get(_w)==_i)
			if total_file>0:file_quality_pct=f"{round(quality_file/total_file*100,2)}%"
		file_rsrp_records=[r for r in rsrp_records if r.get(_F)==fname_base];file_avg_rsrp=''
		if file_rsrp_records:rsrp_vals=[r[_y]for r in file_rsrp_records];file_avg_rsrp=round(sum(rsrp_vals)/len(rsrp_vals),2)
		file_rsrp_call=[r for r in rsrp_records if r.get(_F)==fname_base and r.get(_AO)];file_avg_rsrp_call=''
		if file_rsrp_call:rsrp_call_vals=[r[_y]for r in file_rsrp_call];file_avg_rsrp_call=round(sum(rsrp_call_vals)/len(rsrp_call_vals),2)
		voice_trace_id=_extract_trace_id(f_info.get(_o,''));voice_phone=_lookup_phone(voice_trace_id,phone_trace_map);file_stats.append({A:fname_base,_E:operator,l:'语音',B:caller_type,_M:voice_phone,V:file_row_count,W:file_time_range_str,_Af:file_rounds,_Ag:file_avg_setup,_AA:file_quality_pct,H:file_avg_rsrp,_AH:file_avg_rsrp_call,_AB:'',_AC:'',_AD:'',_AE:'',I:'',J:'',K:'',L:'',M:'',N:'',O:'',P:''})
	for(fname,info)in data_file_info.items():data_trace_id=_extract_trace_id(info.get(_o,''));data_phone=_lookup_phone(data_trace_id,phone_trace_map);caller_type='';file_stats.append({A:fname,_E:info[_h],l:'数据',B:caller_type,_M:data_phone,V:info[_AK],W:info[_AG],_Af:0,_Ag:'',_AA:'',H:'',_AH:'',_AB:info[_A1],_AC:info[_A2],_AD:info[_A5],_AE:info[_A6],I:'',J:'',K:'',L:'',M:'',N:'',O:'',P:''})
	df_file_stats=pd.DataFrame(file_stats)
	if time_range:time_range_str=time_range
	elif time_filter and time_filter[0]and time_filter[1]is not _A and time_filter[2]is not _A:time_range_str=f"{time_filter[1].toString(_FMT_HMS)} ~ {time_filter[2].toString(_FMT_HMS)}"
	elif len(df_voice_raw)>0 and _C in df_voice_raw.columns and df_voice_raw[_C].notna().any():t_min=df_voice_raw[_C].min();t_max=df_voice_raw[_C].max();time_range_str=f"{t_min.strftime(_FMT_DT)} ~ {t_max.strftime(_FMT_DT)}"
	else:time_range_str=''
	summary_data=[];summary_data.append({_E:'标准',A:'',X:'',Y:'',_M:'',_O:1,_AA:'95%',H:'',_AH:'',_AB:1000,_AC:800,_AD:200,_AE:160,I:30,J:50,K:2,L:20,M:20,N:400,O:4,P:200})
	for(fname_base,f_info)in voice_file_info.items():
		operator=f_info[_h];call_type=_extract_call_type_from_path(f_info.get(_o,''))
		if not call_type:call_type='voNR-主叫'
		f_vrs=[vr for vr in voice_records if vr.get(_F)==fname_base];biz_start='';biz_end=''
		if f_vrs:
			if f_vrs[0].get(_G):biz_start=str(f_vrs[0][_G])
			if f_vrs[-1].get(_I):biz_end=str(f_vrs[-1][_I])
		_tr=f_info.get(_AG,'');biz_time_range=f"{biz_start} ~ {biz_end}"if biz_start and biz_end else _tr;trace_id=_extract_trace_id(f_info.get(_o,''));phone=_lookup_phone(trace_id,phone_trace_map);setup_vals=[]
		for vr in f_vrs:
			if vr.get(_O)!=''and pd.notna(vr.get(_O)):
				try:setup_vals.append(float(vr[_O]))
				except:pass
		avg_setup=round(np.mean(setup_vals),2)if setup_vals else'';f_vqi=[r for r in vqi_records if r.get(_F)==fname_base];quality_pct=''
		if f_vqi:
			total_f=len(f_vqi);quality_f=sum(1 for r in f_vqi if r.get(_w)==_i)
			if total_f>0:quality_pct=f"{round(quality_f/total_f*100,2)}%"
		f_rsrp=[r for r in rsrp_records if r.get(_F)==fname_base];avg_rsrp=''
		if f_rsrp:rsrp_vals=[r[_y]for r in f_rsrp];avg_rsrp=round(sum(rsrp_vals)/len(rsrp_vals),2)
		f_rsrp_call=[r for r in rsrp_records if r.get(_F)==fname_base and r.get(_AO)];avg_rsrp_call=''
		if f_rsrp_call:rsrp_call_vals=[r[_y]for r in f_rsrp_call];avg_rsrp_call=round(sum(rsrp_call_vals)/len(rsrp_call_vals),2)
		summary_data.append({_E:operator,A:fname_base,X:biz_time_range,Y:call_type,_M:phone,_O:avg_setup,_AA:quality_pct,H:avg_rsrp,_AH:avg_rsrp_call,_AB:'',_AC:'',_AD:'',_AE:'',I:'',J:'',K:'',L:'',M:'',N:'',O:'',P:''})
	for(fname,info)in data_file_info.items():operator=info[_h];call_type=_AN;biz_time_range=info.get(_AG,'');trace_id=_extract_trace_id(info.get(_o,''));phone=_lookup_phone(trace_id,phone_trace_map);summary_data.append({_E:operator,A:fname,X:biz_time_range,Y:call_type,_M:phone,_O:'',_AA:'',H:'',_AH:'',_AB:info[_A1],_AC:info[_A2],_AD:info[_A5],_AE:info[_A6],I:'',J:'',K:'',L:'',M:'',N:'',O:'',P:''})
	df_summary_data_voice=pd.DataFrame(summary_data);version=_Ab;timestamp=datetime.now().strftime(_Ac);out=f"语音VQI输出_{version}_{timestamp}.xlsx";log('写入输出文件...',70);log('  引擎初始化...',71);_out_tmp_vqi=out.replace(_AI,'_tmp.xlsx')
	with pd.ExcelWriter(_out_tmp_vqi,engine='xlsxwriter')as w:
		workbook=w.book;hdr_fmt=workbook.add_format({_AM:_B,_W:_Y,_X:_Z,_p:_Az,_c:1,_AV:_B});center_fmt=workbook.add_format({_W:_Y,_X:_Z,_c:1});green_fmt=workbook.add_format({_W:_Y,_X:_Z,_c:1,_p:_Ay});blue_fmt=workbook.add_format({_W:_Y,_X:_Z,_c:1,_p:_A_});anno_fmt=workbook.add_format({_W:_Y,_X:_Z,_c:1,_A9:'0'})
		for(sheet_name,df_data,colwidths)in[(_AP,df_voice_out,20),(_AQ,df_vqi_out,22),(_AF,df_summary_data_voice,18),(R,df_file_stats,20)]:
			if len(df_data)==0:continue
			log(f"  写入{sheet_name}({len(df_data)}行x{len(df_data.columns)}列)...",72);ws=workbook.add_worksheet(sheet_name)
			for(ci,col_name)in enumerate(df_data.columns):ws.write(0,ci,str(col_name),hdr_fmt)
			data_array=df_data.values;ncols=len(df_data.columns)
			for ri in range(len(df_data)):
				row_vals=data_array[ri]
				for ci in range(ncols):
					val=row_vals[ci]
					if val is _A or isinstance(val,float)and pd.isna(val):ws.write_blank(ri+1,ci,_A,center_fmt)
					elif isinstance(val,(int,float)):
						if sheet_name==_AQ and ri+1==len(df_data):ws.write(ri+1,ci,val,green_fmt)
						elif sheet_name==_AF and ri==1:ws.write(ri+1,ci,val,blue_fmt)
						else:ws.write(ri+1,ci,val,center_fmt)
					else:ws.write(ri+1,ci,str(val),center_fmt)
			ws.set_column(0,ncols-1,colwidths);ws.freeze_panes(1,0);log(f"  {sheet_name} 完成",74)
			if sheet_name==_AP and len(df_data.columns)>=5:
				for ri in range(len(df_data)):
					val=data_array[ri,4]
					if val is _A or isinstance(val,float)and pd.isna(val):reason=voice_records[ri].get(j,'')if ri<len(voice_records)else'';ws.write_comment(ri+1,4,reason or'无RSP_180/183',{n:'Tool'})
		if merge_raw and len(df_raw_out)>0:
			raw_n=len(df_raw_out);log(f"  语音原始数据共{raw_n}行",75)
			for _c_drop in[_C,_u]:
				if _c_drop in df_raw_out.columns:df_raw_out=df_raw_out.drop(_c_drop,axis=1)
			if raw_n>500000:csv_out=out.replace(_AI,'_原始数据.csv');log(f"  原始数据{raw_n}行超过50万, 导出为CSV...",76);df_raw_out.to_csv(csv_out,index=_D,encoding='utf-8-sig');log(f"  原始数据CSV写入完成: {csv_out}",79)
			else:log(f"  写入{raw_n}行x{len(df_raw_out.columns)}列...",77);df_raw_out.to_excel(w,sheet_name=_AY,index=_D);ws_raw=w.sheets[_AY];ws_raw.freeze_panes(1,0);log(f"  原始数据写入完成",79)
		if add_annotations:
			log('  添加批注...',85);_op_voice_cache={};_op_vqi_cache={};_op_rsrp_cache={}
			for op in set(r.get(_E,'')for r in voice_records):_op_voice_cache[op]=[vr for vr in voice_records if vr.get(_E)==op]
			for op in set(r.get(_E,'')for r in vqi_records):_op_vqi_cache[op]=[r for r in vqi_records if r.get(_E)==op]
			for op in set(r.get(_E,'')for r in rsrp_records):_op_rsrp_cache[op]=[r for r in rsrp_records if r.get(_E)==op]
			_fname_voice_cache={};_fname_vqi_cache={};_fname_rsrp_cache={}
			for fname in set(r.get(_F,'')for r in voice_records):_fname_voice_cache[fname]=[vr for vr in voice_records if vr.get(_F)==fname]
			for fname in set(r.get(_F,'')for r in vqi_records):_fname_vqi_cache[fname]=[r for r in vqi_records if r.get(_F)==fname]
			for fname in set(r.get(_F,'')for r in rsrp_records):_fname_rsrp_cache[fname]=[r for r in rsrp_records if r.get(_F)==fname]
			for sheet_name in[_AF,R]:
				try:ws=workbook.get_worksheet_by_name(sheet_name)
				except:continue
				log(f"  批注: {sheet_name}",86);df_map={_AF:df_summary_data_voice,R:df_file_stats};df_anno=df_map.get(sheet_name)
				if df_anno is _A or len(df_anno)==0:continue
				anno_data=df_anno.values;anno_cols=list(df_anno.columns)
				for ri in range(len(df_anno)):
					for ci in range(1,len(anno_cols)):
						val=anno_data[ri,ci]
						if isinstance(val,(int,float))and not(isinstance(val,float)and pd.isna(val)):
							col_name=anno_cols[ci]
							if sheet_name==_AF:row_op=df_anno.iloc[ri,0];txt=_build_anno_text(col_name,row_op,ri,df_voice_out,df_vqi,voice_records,vqi_records,rsrp_records,df_file_stats,_op_voice_cache,_op_vqi_cache,_op_rsrp_cache)
							elif sheet_name==R:
								fname=df_anno.iloc[ri,0];txt=_build_file_anno_text(col_name,fname,voice_records,vqi_records,rsrp_records,_fname_voice_cache,_fname_vqi_cache,_fname_rsrp_cache)
								if not txt:txt=_build_data_biz_anno(col_name,fname,data_file_info)
							else:txt=f"公式: {col_name}\n计算过程: 从原始数据统计\n\n数据来源: {col_name} 列"
							ws.write_comment(ri+1,ci,txt,{n:'Tool',_B0:550,_B1:250})
		log('  写入数据明细sheet...',80);voice_file_to_path={fname:info[_o]for(fname,info)in voice_file_info.items()};data_file_to_path={fname:info[_o]for(fname,info)in data_file_info.items()};all_file_to_path={**voice_file_to_path,**data_file_to_path};trace_to_phone={}
		if phone_trace_map:
			for entry in phone_trace_map:trace_to_phone[entry[_t]]=entry[_V]
		ws_call_setup=workbook.add_worksheet('语音_呼叫建立时长原始数据');hdr_call_setup=[A,_E,_M,_H,B,E,F,_G,_I,_O]
		for(ci,col_name)in enumerate(hdr_call_setup):ws_call_setup.write(0,ci,col_name,hdr_fmt)
		for(ri,vr)in enumerate(voice_records):
			sf=vr.get(_F,'');phone=trace_to_phone.get(_extract_trace_id(all_file_to_path.get(sf,'')),'');row_vals=[sf,vr.get(_E,''),phone,vr.get(_H,''),vr.get(B,''),vr.get(E,''),vr.get(F,''),vr.get(_G,''),vr.get(_I,''),vr.get(_O,'')]
			for(ci,val)in enumerate(row_vals):
				if val is _A or isinstance(val,float)and pd.isna(val):ws_call_setup.write_blank(ri+1,ci,_A,center_fmt)
				else:ws_call_setup.write(ri+1,ci,str(val),center_fmt)
		ws_call_setup.set_column(0,0,40);ws_call_setup.set_column(1,2,12);ws_call_setup.set_column(3,9,18);ws_call_setup.freeze_panes(1,0);ws_mos=workbook.add_worksheet('语音_MOS原始数据');hdr_mos=[A,_E,_M,_H,_G,_I,'DlE2eVqi','MOS值(=Vqi/100)',_w]
		for(ci,col_name)in enumerate(hdr_mos):ws_mos.write(0,ci,col_name,hdr_fmt)
		for(ri,r)in enumerate(vqi_records):
			e2e=r.get(G,'');mos_val=round(e2e/100,2)if isinstance(e2e,(int,float))and e2e!=65535 else'';quality=_i if isinstance(e2e,(int,float))and e2e>350 else D;sf=r.get(_F,'');phone=trace_to_phone.get(_extract_trace_id(all_file_to_path.get(sf,'')),'');row_vals=[sf,r.get(_E,''),phone,r.get(_H,''),r.get(_x,''),r.get(k,''),e2e,mos_val,quality]
			for(ci,val)in enumerate(row_vals):
				if val is _A or isinstance(val,float)and pd.isna(val):ws_mos.write_blank(ri+1,ci,_A,center_fmt)
				else:ws_mos.write(ri+1,ci,str(val),center_fmt)
		ws_mos.set_column(0,0,40);ws_mos.set_column(1,2,12);ws_mos.set_column(3,8,18);ws_mos.freeze_panes(1,0);ws_rsrp=workbook.add_worksheet('语音_RSRP原始数据');hdr_rsrp=[A,_E,_M,_x,'服务小区RSRP(dBm)','是否在通话中(是/否)']
		for(ci,col_name)in enumerate(hdr_rsrp):ws_rsrp.write(0,ci,col_name,hdr_fmt)
		for(ri,r)in enumerate(rsrp_records):
			sf=r.get(_F,'');phone=trace_to_phone.get(_extract_trace_id(all_file_to_path.get(sf,'')),'');row_vals=[sf,r.get(_E,''),phone,r.get(_x,''),r[_y],_i if r.get(_AO)else D]
			for(ci,val)in enumerate(row_vals):
				if val is _A or isinstance(val,float)and pd.isna(val):ws_rsrp.write_blank(ri+1,ci,_A,center_fmt)
				else:ws_rsrp.write(ri+1,ci,str(val),center_fmt)
		ws_rsrp.set_column(0,0,40);ws_rsrp.set_column(1,2,12);ws_rsrp.set_column(3,5,18);ws_rsrp.freeze_panes(1,0);ws_round=workbook.add_worksheet('语音_通话轮次原始数据');hdr_round=[A,_E,_M,_H,B,E,F,_G,_I,C]
		for(ci,col_name)in enumerate(hdr_round):ws_round.write(0,ci,col_name,hdr_fmt)
		for(ri,vr)in enumerate(voice_records):
			sf=vr.get(_F,'');phone=trace_to_phone.get(_extract_trace_id(all_file_to_path.get(sf,'')),'');row_vals=[sf,vr.get(_E,''),phone,vr.get(_H,''),vr.get(B,''),vr.get(E,''),vr.get(F,''),vr.get(_G,''),vr.get(_I,''),vr.get(C,D)]
			for(ci,val)in enumerate(row_vals):
				if val is _A or isinstance(val,float)and pd.isna(val):ws_round.write_blank(ri+1,ci,_A,center_fmt)
				else:ws_round.write(ri+1,ci,str(val),center_fmt)
		ws_round.set_column(0,0,40);ws_round.set_column(1,2,12);ws_round.set_column(3,9,18);ws_round.freeze_panes(1,0);ws_quality=workbook.add_worksheet('语音_优质通话占比原始数据');hdr_quality=[A,_E,_M,'总通话数','优质通话数','优质通话占比(%)']
		for(ci,col_name)in enumerate(hdr_quality):ws_quality.write(0,ci,col_name,hdr_fmt)
		fname_quality={}
		for r in vqi_records:
			fn=r.get(_F,'');e2e=r.get(G,'');phone=trace_to_phone.get(_extract_trace_id(all_file_to_path.get(fn,'')),'')
			if fn not in fname_quality:fname_quality[fn]={Q:0,S:0,'op':r.get(_E,''),_V:phone}
			fname_quality[fn][Q]+=1
			if isinstance(e2e,(int,float))and e2e>350:fname_quality[fn][S]+=1
		ri=0
		for(fn,qd)in fname_quality.items():
			pct=round(qd[S]/qd[Q]*100,2)if qd[Q]>0 else 0;row_vals=[fn,qd['op'],qd[_V],qd[Q],qd[S],pct]
			for(ci,val)in enumerate(row_vals):
				if val is _A or isinstance(val,float)and pd.isna(val):ws_quality.write_blank(ri+1,ci,_A,center_fmt)
				else:ws_quality.write(ri+1,ci,str(val),center_fmt)
			ri+=1
		ws_quality.set_column(0,0,40);ws_quality.set_column(1,2,12);ws_quality.set_column(3,5,18);ws_quality.freeze_panes(1,0);ws_voice_dur=workbook.add_worksheet('语音_语音业务时长原始数据');hdr_voice_dur=[A,_E,_M,_H,B,_G,_I,'通话时长(s)']
		for(ci,col_name)in enumerate(hdr_voice_dur):ws_voice_dur.write(0,ci,col_name,hdr_fmt)
		for(ri,vr)in enumerate(voice_records):
			start=vr.get(_G,'');end=vr.get(_I,'');dur=''
			if start and end:
				try:
					m1=_PARSE_RE.match(str(start));m2=_PARSE_RE.match(str(end))
					if m1 and m2:t1=datetime.strptime(m1.group(1),_AR)+timedelta(milliseconds=int(m1.group(2)));t2=datetime.strptime(m2.group(1),_AR)+timedelta(milliseconds=int(m2.group(2)));dur=round((t2-t1).total_seconds(),2)
				except:pass
			sf=vr.get(_F,'');phone=trace_to_phone.get(_extract_trace_id(all_file_to_path.get(sf,'')),'');row_vals=[sf,vr.get(_E,''),phone,vr.get(_H,''),vr.get(B,''),start,end,dur]
			for(ci,val)in enumerate(row_vals):
				if val is _A or isinstance(val,float)and pd.isna(val):ws_voice_dur.write_blank(ri+1,ci,_A,center_fmt)
				else:ws_voice_dur.write(ri+1,ci,str(val),center_fmt)
		ws_voice_dur.set_column(0,0,40);ws_voice_dur.set_column(1,2,12);ws_voice_dur.set_column(3,7,18);ws_voice_dur.freeze_panes(1,0);log('  写入数据业务明细sheet...',81)
		if data_dfs:
			ws_rate=workbook.add_worksheet(Z);hdr_rate=[A,_E,_M,_H,_T,_G,_I,_AL,_Ae,o]
			for(ci,col_name)in enumerate(hdr_rate):ws_rate.write(0,ci,col_name,hdr_fmt)
			ri=0
			for(fname,info)in data_file_info.items():
				op=info.get(_h,'');phone=info.get(_V,'');row_vals=[fname,op,phone,'','','','',info.get(_A2,''),info.get('dl_clip_avg',''),info.get(_A1,'')]
				for(ci,val)in enumerate(row_vals):
					if val is _A or isinstance(val,float)and pd.isna(val):ws_rate.write_blank(ri+1,ci,_A,center_fmt)
					else:ws_rate.write(ri+1,ci,str(val),center_fmt)
				ri+=1
			ws_rate.set_column(0,0,35);ws_rate.set_column(1,2,10);ws_rate.set_column(3,9,15);ws_rate.freeze_panes(1,0);ws_biz_dur=workbook.add_worksheet(a);hdr_biz_dur=[A,_E,_M,_H,_T,_G,_I,p]
			for(ci,col_name)in enumerate(hdr_biz_dur):ws_biz_dur.write(0,ci,col_name,hdr_fmt)
			ri=0
			for(fname,info)in data_file_info.items():
				op=info.get(_h,'');phone=info.get(_V,'');row_vals=[fname,op,phone,'','','','','']
				for(ci,val)in enumerate(row_vals):
					if val is _A or isinstance(val,float)and pd.isna(val):ws_biz_dur.write_blank(ri+1,ci,_A,center_fmt)
					else:ws_biz_dur.write(ri+1,ci,str(val),center_fmt)
				ri+=1
			ws_biz_dur.set_column(0,0,35);ws_biz_dur.set_column(1,2,10);ws_biz_dur.set_column(3,7,15);ws_biz_dur.freeze_panes(1,0);ws_ftp_dl=workbook.add_worksheet(b);hdr_ftp=[A,_E,_M,_H,_G,_I,_AL,q,u,x,y]
			for(ci,col_name)in enumerate(hdr_ftp):ws_ftp_dl.write(0,ci,col_name,hdr_fmt)
			ri=0
			for(fname,info)in data_file_info.items():
				op=info.get(_h,'');phone=info.get(_V,'');row_vals=[fname,op,phone,'','','',info.get(_A2,''),info.get(_A1,''),'','','']
				for(ci,val)in enumerate(row_vals):
					if val is _A or isinstance(val,float)and pd.isna(val):ws_ftp_dl.write_blank(ri+1,ci,_A,center_fmt)
					else:ws_ftp_dl.write(ri+1,ci,str(val),center_fmt)
				ri+=1
			ws_ftp_dl.set_column(0,0,35);ws_ftp_dl.set_column(1,2,10);ws_ftp_dl.set_column(3,10,15);ws_ftp_dl.freeze_panes(1,0);ws_ftp_ul=workbook.add_worksheet(c)
			for(ci,col_name)in enumerate(hdr_ftp):ws_ftp_ul.write(0,ci,col_name,hdr_fmt)
			ri=0
			for(fname,info)in data_file_info.items():
				op=info.get(_h,'');phone=info.get(_V,'');row_vals=[fname,op,phone,'','','',info.get(_A6,''),info.get(_A5,''),'','','']
				for(ci,val)in enumerate(row_vals):
					if val is _A or isinstance(val,float)and pd.isna(val):ws_ftp_ul.write_blank(ri+1,ci,_A,center_fmt)
					else:ws_ftp_ul.write(ri+1,ci,str(val),center_fmt)
				ri+=1
			ws_ftp_ul.set_column(0,0,35);ws_ftp_ul.set_column(1,2,10);ws_ftp_ul.set_column(3,10,15);ws_ftp_ul.freeze_panes(1,0);ws_wechat=workbook.add_worksheet(d);hdr_wechat=[A,_E,_M,_H,z,A0,_G,_I,A1,_A8,'时延(s)']
			for(ci,col_name)in enumerate(hdr_wechat):ws_wechat.write(0,ci,col_name,hdr_fmt)
			ws_wechat.set_column(0,0,35);ws_wechat.set_column(1,2,10);ws_wechat.set_column(3,10,25);ws_wechat.freeze_panes(1,0)
		else:
			for sheet_name in[Z,a,b,c,d]:
				ws_empty=workbook.add_worksheet(sheet_name)
				if sheet_name==Z:hdr=[A,_E,_M,_H,_T,_G,_I,_AL,_Ae,o]
				elif sheet_name==a:hdr=[A,_E,_M,_H,_T,_G,_I,p]
				elif sheet_name in[b,c]:hdr=[A,_E,_M,_H,_G,_I,_AL,q,u,x,y]
				elif sheet_name==d:hdr=[A,_E,_M,_H,z,A0,_G,_I,A1,_A8,'时延(s)']
				else:hdr=[A,_E,_M,'QCI值',V,W]
				for(ci,col_name)in enumerate(hdr):ws_empty.write(0,ci,col_name,hdr_fmt)
				ws_empty.set_column(0,0,35);ws_empty.set_column(1,2,10);ws_empty.set_column(3,len(hdr)-1,18);ws_empty.freeze_panes(1,0)
		log('  数据明细sheet写入完成',82)
	log(f"完成! {out}",100)
	if os.path.exists(_out_tmp_vqi):os.replace(_out_tmp_vqi,out)
	for _d in _tmp_dirs:
		try:shutil.rmtree(_d)
		except Exception:pass
	return out
try:from PySide6.QtWidgets import QApplication,QMainWindow,QWidget,QVBoxLayout,QHBoxLayout,QPushButton,QLabel,QFileDialog,QMessageBox,QProgressBar,QTextEdit,QGroupBox,QCheckBox,QSpinBox,QDialog,QSplitter,QGridLayout,QTabWidget,QTableWidget,QTableWidgetItem,QHeaderView,QTimeEdit,QComboBox,QLineEdit,QFrame,QListWidget,QListWidgetItem;from PySide6.QtCore import Qt,QThread,Signal,Slot,QTime,QEvent,QTimer,QSettings;from PySide6.QtGui import QFont;GUI_OK=_B
except:GUI_OK=_D
if GUI_OK:
	class WorkerThread(QThread):
		progress=Signal(str);progress_pct=Signal(int);done=Signal(str)
		def __init__(self,files,qci_list,dl_clip,ul_clip,base_file=_A,time_filter=_A,ftp_duration=25,merge_raw=_D,add_annotations=_B,phone_trace_map=_A,biz_order=_A,biz_params=_A):super().__init__();self.files=files;self.qci_list=qci_list;self.dl_clip=dl_clip;self.ul_clip=ul_clip;self.base_file=base_file;self.time_filter=time_filter;self.ftp_duration=ftp_duration;self.merge_raw=merge_raw;self.add_annotations=add_annotations;self.phone_trace_map=phone_trace_map or[];self.biz_order=biz_order;self.biz_params=biz_params or{};self._cancel=_D;self._pct=0
		def cancel(self):self._cancel=_B
		def _ask_process_type(self):
			'弹出处理类型选择对话框';msg=QMessageBox(self);msg.setWindowTitle('选择处理类型');msg.setText('请选择要处理的业务类型：');msg.setIcon(QMessageBox.Question);btn_data=msg.addButton('仅数据业务',QMessageBox.ActionRole);btn_voice=msg.addButton('仅语音VQI',QMessageBox.ActionRole);btn_both=msg.addButton('两者都处理',QMessageBox.ActionRole);btn_cancel=msg.addButton('取消',QMessageBox.RejectRole);msg.setDefaultButton(btn_both);msg.exec_();clicked=msg.clickedButton()
			if clicked==btn_data:return 1
			if clicked==btn_voice:return 2
			if clicked==btn_both:return 3
			return 0
		def _do_data_run(self):
			'执行数据业务处理(5G速率统计)';ok,msg=check_license()
			if not ok:QMessageBox.warning(self,'鉴权失败',f"无法使用：{msg}\n请联网后重试。");return
			self._sync_phone_mapping(_d)
			if not self.phone_trace_map_data:
				ret=QMessageBox.question(self,_Ah,_Ai,QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No,QMessageBox.StandardButton.No)
				if ret!=QMessageBox.StandardButton.Yes:return
			self.result_tabs.clear();self.result_tabs.clear();self.lt.clear();qci_list=[]
			if self.cb6.isChecked():qci_list.append(6)
			if self.cb7.isChecked():qci_list.append(7)
			if not qci_list:qci_list=[7]
			time_filter=_A
			if self.cb_time.isChecked():t_start=self.time_start.time();t_end=self.time_end.time();time_filter=_B,t_start,t_end
			ftp_duration=self.ftp_duration_spin.value();biz_order=[];biz_params={}
			if hasattr(self,_BC):
				for r in range(self.biz_order_table.rowCount()):
					_biz_item=self.biz_order_table.item(r,0)
					if not _biz_item or not _biz_item.text():continue
					_biz=_biz_item.text();biz_order.append(_biz);_vals=[self.biz_order_table.item(r,c).text().strip()if self.biz_order_table.item(r,c)else''for c in range(1,7)]
					if all(_vals):biz_params[_biz]={_As:float(_vals[0]),_At:float(_vals[1]),_Au:float(_vals[2])/1e2,_Av:float(_vals[3]),_Aw:float(_vals[4]),_Ax:float(_vals[5])/1e2}
			self.br.setEnabled(_D);self.bcancel.setEnabled(_B);self.pb.setValue(0);self.lt.append(f"开始处理数据业务... QCI={qci_list}, FTP时长={ftp_duration}秒")
			if time_filter:self.lt.append(f"时间段过滤: {t_start.toString(_FMT_HMS)} ~ {t_end.toString(_FMT_HMS)}")
			self.thread=WorkerThread(self.files,qci_list,self.sd.value(),self.su.value(),base_file=self.base_file,time_filter=time_filter,ftp_duration=ftp_duration,merge_raw=self.cb_merge_raw.isChecked(),add_annotations=self.cb_annotations.isChecked(),phone_trace_map=self.phone_trace_map_data,biz_order=biz_order,biz_params=biz_params);self.thread.progress.connect(self.upd);self.thread.progress_pct.connect(lambda p:self.pb.setValue(p));self.thread.done.connect(self.dn);self.thread.start()
		def _do_vqi_run(self):
			'执行语音VQI处理';ok,msg=check_license()
			if not ok:QMessageBox.warning(self,'鉴权失败',f"无法使用：{msg}\n请联网后重试。");return
			self._sync_phone_mapping(_z)
			if not self.phone_trace_map_vqi:
				ret=QMessageBox.question(self,_Ah,_Ai,QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No,QMessageBox.StandardButton.No)
				if ret!=QMessageBox.StandardButton.Yes:return
			vqi_files=list(self.vqi_files)if hasattr(self,'vqi_files')and self.vqi_files else list(self.files);time_filter=_A
			if self.vqi_cb_time.isChecked()if hasattr(self,'vqi_cb_time')else _D:time_filter=_B,self.vqi_time_start.time(),self.vqi_time_end.time()
			self.vqi_br.setEnabled(_D)if hasattr(self,'vqi_br')else _A;self.bcancel.setEnabled(_B);self.pb.setValue(0);self.lt.append('开始处理语音VQI...');self.vqi_thread=VqiWorkerThread(vqi_files,time_filter=time_filter,merge_raw=self.vqi_cb_merge_raw.isChecked()if hasattr(self,'vqi_cb_merge_raw')else _B,add_annotations=self.vqi_cb_annotations.isChecked()if hasattr(self,'vqi_cb_annotations')else _B,phone_trace_map=self.phone_trace_map_vqi);self.vqi_thread.progress.connect(self.upd);self.vqi_thread.progress_pct.connect(lambda p:self.pb.setValue(p));self.vqi_thread.done.connect(self.vqi_dn);self.vqi_thread.start()
		def run(self):
			try:out=process(self.files,self.qci_list,self.dl_clip,self.ul_clip,callback=self.l,cancel_check=lambda:self._cancel,progress_cb=lambda p:self.progress_pct.emit(int(p)),base_file=self.base_file,time_filter=self.time_filter,ftp_duration=self.ftp_duration,merge_raw=self.merge_raw,add_annotations=self.add_annotations,phone_trace_map=self.phone_trace_map,biz_order=self.biz_order,biz_params=self.biz_params);self.done.emit(''if self._cancel else out)
			except KeyboardInterrupt:self.done.emit('')
			except Exception as e:import traceback;self.progress.emit(f"处理出错: {e}");self.progress.emit(traceback.format_exc());self.done.emit('')
		def l(self,msg):self.progress.emit(msg)
	class VqiWorkerThread(QThread):
		progress=Signal(str);progress_pct=Signal(int);done=Signal(str)
		def __init__(self,files,time_filter=_A,merge_raw=_D,add_annotations=_B,phone_trace_map=_A):super().__init__();self.files=files;self.time_filter=time_filter;self._cancel=_D;self.merge_raw=merge_raw;self.add_annotations=add_annotations;self.phone_trace_map=phone_trace_map or[]
		def cancel(self):self._cancel=_B
		def run(self):
			try:out=process_vqi(self.files,callback=self.l,cancel_check=lambda:self._cancel,progress_cb=lambda p:self.progress_pct.emit(int(p)),time_filter=self.time_filter,merge_raw=self.merge_raw,add_annotations=self.add_annotations,phone_trace_map=self.phone_trace_map);self.done.emit(''if self._cancel else out)
			except KeyboardInterrupt:self.done.emit('')
			except Exception as e:import traceback;self.progress.emit(f"VQI处理出错: {e}");self.progress.emit(traceback.format_exc());self.done.emit('')
		def l(self,msg):self.progress.emit(msg)
	class MainWindow(QMainWindow):
		def __init__(self):super().__init__();self.setWindowTitle('工信部数据、语音跟踪统计工具 V2.5.35_CC');self.files=[];self.last_out=_A;self.base_file=_A;self.vqi_files=[];self.vqi_base_file=_A;self.vqi_last_out=_A;self.phone_trace_map_data=[];self.phone_trace_map_vqi=[];scr=QApplication.primaryScreen();self.scr_w=scr.geometry().width()if scr else 1440;sh=scr.geometry().height()if scr else 1080;self.fs=max(11,int(round(11*sh/1080)));self.setMinimumSize(1200,750);self._build();self._load_phone_mapping(_d);self._load_phone_mapping(_z);self.showMaximized()
		def _font(self,big=0):
			if sys.platform=='darwin':_fn=_BD
			else:_fn=_BE
			return QFont(_fn,self.fs+big,QFont.Bold if big>=2 else QFont.Normal)
		@staticmethod
		def _extract_trace_id(path):"从文件路径中提取用户跟踪ID, 如 '用户跟踪ID=553' → '553'";import re;m=re.search(_BA,str(path));return m.group(1)if m else _A
		def _lookup_phone(self,trace_id,phone_map=_A):
			'查找trace_id对应的手机号(实例级备用,默认用data map)';pm=phone_map if phone_map is not _A else self.phone_trace_map_data
			if not trace_id or not pm:return''
			for entry in pm:
				if entry[_t]==trace_id:return entry.get(_V,'')
			return''
		def _phone_map_file(self,which):'返回映射json文件路径(config目录,与工具一起走)';name='手机号映射_数据业务.json'if which==_d else'手机号映射_语音VQI.json';return os.path.join(os.path.dirname(os.path.abspath(__file__)),'..','config',name)
		def _save_phone_mapping(self,which):
			'保存对应tab的映射到json(失败不报错,不影响主功能)'
			if which not in(_d,_z):return
			table=self.phone_table if which==_d else self.vqi_phone_table;rows=[]
			for row in range(table.rowCount()):
				phone_item=table.item(row,0);tid_item=table.item(row,1);phone=phone_item.text().strip()if phone_item else'';tid=tid_item.text().strip()if tid_item else''
				if phone and tid:rows.append({_V:phone,_t:tid})
			try:
				path=self._phone_map_file(which);os.makedirs(os.path.dirname(path),exist_ok=_B)
				with open(path,'w',encoding=_A4)as f:json.dump(rows,f,ensure_ascii=_D,indent=2)
			except Exception:pass
		def _load_phone_mapping(self,which):
			'启动时从json加载映射并填充到对应表格(文件损坏/缺失则空启动不报错)'
			if which not in(_d,_z):return
			try:
				path=self._phone_map_file(which)
				if not os.path.exists(path):return
				with open(path,'r',encoding=_A4)as f:rows=json.load(f)
				table=self.phone_table if which==_d else self.vqi_phone_table;pairs=[(r.get(_V,''),r.get(_t,''))for r in rows if r.get(_V)and r.get(_t)]
				if pairs:self._add_rows_to_table(pairs,table)
			except Exception:pass
		def _remove_phone_mapping_row(self,table_widget,btn,which=_A):
			'删除按钮所在行并持久化(动态按按钮位置查行,避免删中间行后行号失效)';row=table_widget.indexAt(btn.pos()).row()
			if row>=0:table_widget.removeRow(row)
			if which:self._save_phone_mapping(which)
		def _build_phone_mapping_group(self,table_widget,text_edit,add_btn,clear_btn,which=_A):
			'构建手机号映射UI组（左右并排，外框紧凑不超内容）';gp=QGroupBox('手机号-TraceID映射');_main=QHBoxLayout(gp);_main.setContentsMargins(2,2,2,2);_main.setSpacing(4);_lcol=QVBoxLayout();_lcol.setSpacing(2);_lcol.addWidget(QLabel('手动添加：'));table_widget.setColumnCount(3);table_widget.setHorizontalHeaderLabels([_M,'Trace ID','操作']);table_widget.horizontalHeader().setStretchLastSection(_D);_settings=QSettings(_Aj,_Ak)
			for(_ci,_cw)in enumerate([130,70,55]):
				_saved=_settings.value(f"phone_col_{which}_{_ci}",_cw)
				try:_cw=int(_saved)
				except:pass
				table_widget.horizontalHeader().setSectionResizeMode(_ci,QHeaderView.Interactive);table_widget.setColumnWidth(_ci,_cw)
			def _save_phone_cols(_ci,_os,_ns):QSettings(_Aj,_Ak).setValue(f"phone_col_{which}_{_ci}",_ns)
			table_widget.horizontalHeader().sectionResized.connect(_save_phone_cols);table_widget.setMinimumHeight(120);_lcol.addWidget(table_widget);add_btn.setText('添加映射');add_btn.clicked.connect(lambda:self._add_phone_mapping_row(table_widget,which));_lcol.addWidget(add_btn);_main.addLayout(_lcol,2);_rcol=QVBoxLayout();_rcol.setSpacing(2);_rcol.addWidget(QLabel('拖入/粘贴映射：'));text_edit.setMinimumHeight(120);text_edit.setPlaceholderText('拖入Excel/Txt文件，或粘贴文本\n(支持空格/逗号/斜杠分隔),自动填表...');text_edit.setStyleSheet('QTextEdit{font-size:9pt;background:#fff;border:1px solid #999;border-radius:3px;padding:2px;}');_rcol.addWidget(text_edit);clear_btn.setText('清空映射');clear_btn.clicked.connect(lambda:self._clear_phone_mapping(table_widget,text_edit,which));_rcol.addWidget(clear_btn);_main.addLayout(_rcol,1);table_widget.cellChanged.connect(lambda:self._save_phone_mapping(which)if which else _A);timer=QTimer();timer.setSingleShot(_B);timer.setInterval(300);timer.timeout.connect(lambda:self._parse_textedit_into_table(text_edit,table_widget,which));text_edit.textChanged.connect(timer.start);text_edit._parse_timer=timer;return gp
		def _add_phone_mapping_row(self,table_widget,which=_A):'在映射表格中添加一行空行';row=table_widget.rowCount();table_widget.insertRow(row);table_widget.setItem(row,0,QTableWidgetItem(''));table_widget.setItem(row,1,QTableWidgetItem(''));del_btn=QPushButton('删除');del_btn.clicked.connect(lambda checked,b=del_btn:self._remove_phone_mapping_row(table_widget,b,which));table_widget.setCellWidget(row,2,del_btn)
		def _clear_phone_mapping(self,table_widget,text_edit,which=_A):
			'清空手机号映射（表格+文本框+内存+持久化文件）';table_widget.setRowCount(0);text_edit.clear()
			if which==_d:self.phone_trace_map_data=[]
			elif which==_z:self.phone_trace_map_vqi=[]
			if which:self._save_phone_mapping(which)
		def _sync_phone_mapping(self,which=_A):
			"从UI同步映射到内存(which='data'/'vqi'指定单tab;None=两个都同步)"
			def _sync_one(table,te):
				result=[]
				for row in range(table.rowCount()):
					phone_item=table.item(row,0);tid_item=table.item(row,1);phone=phone_item.text().strip()if phone_item else'';tid=tid_item.text().strip()if tid_item else''
					if phone and tid and not any(e[_t]==tid for e in result):result.append({_V:phone,_t:tid})
				txt=te.toPlainText().strip()
				if txt:
					for line in txt.split(_a):
						line=line.strip()
						if line and os.path.isfile(line):self._parse_mapping_file(line,table)
					pairs=self._parse_mapping_text_to_pairs(txt)
					if pairs:self._add_rows_to_table(pairs,table)
					for row in range(table.rowCount()):
						phone_item=table.item(row,0);tid_item=table.item(row,1);phone=phone_item.text().strip()if phone_item else'';tid=tid_item.text().strip()if tid_item else''
						if phone and tid and not any(e[_t]==tid for e in result):result.append({_V:phone,_t:tid})
				return result
			if which==_d or which is _A:self.phone_trace_map_data=_sync_one(self.phone_table,self.phone_textedit);self._save_phone_mapping(_d)
			if which==_z or which is _A:self.phone_trace_map_vqi=_sync_one(self.vqi_phone_table,self.vqi_phone_textedit);self._save_phone_mapping(_z)
		def _parse_mapping_file(self,file_path,table_widget):
			'解析映射文件（Excel两列 / Txt两列Tab分隔），列顺序自适应';low=file_path.lower();rows=[]
			if low.endswith((_AI,_AZ)):
				try:
					import openpyxl;wb=openpyxl.load_workbook(file_path,read_only=_B,data_only=_B);ws=wb.active;data_rows=list(ws.iter_rows(values_only=_B))
					if not data_rows:return
					rows=self._detect_columns_and_parse(data_rows)
				except Exception as e:return
			else:
				try:
					with open(file_path,'r',encoding=_A4)as f:raw_lines=f.readlines()
					data_rows=[]
					for line in raw_lines:
						line=line.strip()
						if line:parts=re.split(_B3,line);data_rows.append([p for p in parts if p])
					if data_rows:rows=self._detect_columns_and_parse(data_rows)
				except Exception as e:return
			self._add_rows_to_table(rows,table_widget)
		@staticmethod
		def _order_pair(a,b):
			'两值配对,11位数字当手机号放前(用户:13271860243是手机号,562是trace id)'
			if len(a)==11 and a.isdigit():return a,b
			if len(b)==11 and b.isdigit():return b,a
			return a,b
		def _parse_mapping_text_to_pairs(self,txt):
			'解析多行映射文本为(phone,trace_id)对。\n            支持两种格式:①同行多列(空格/逗号/斜杠/Tab分隔) ②换行交替(手机号一行、ID一行)。可混合。';pairs=[];pending=_A
			for raw_line in txt.split(_a):
				line=raw_line.strip()
				if not line or os.path.isfile(line):continue
				parts=re.split(_B3,line);parts=[p for p in parts if p]
				if len(parts)>=2:
					if pending:pairs.append(self._order_pair(pending,parts[0]));pending=_A;rest=parts[1:]
					else:rest=parts
					for i in range(0,len(rest)-1,2):pairs.append(self._order_pair(rest[i],rest[i+1]))
				elif len(parts)==1:
					if pending is _A:pending=parts[0]
					else:pairs.append(self._order_pair(pending,parts[0]));pending=_A
			return[(p,t)for(p,t)in pairs if p and t]
		def _parse_mapping_line(self,line,table_widget):
			'解析单行文本(Tab/空格/逗号/分号/斜杠等分隔符,两列:手机号+TraceID)';parts=re.split(_B3,line.strip());parts=[p for p in parts if p]
			if len(parts)>=2:
				phone,tid=self._order_pair(parts[0],parts[1])
				if phone and tid:self._add_rows_to_table([(phone,tid)],table_widget)
		def _parse_textedit_into_table(self,text_edit,table_widget,which=_A):
			'文本框内容变化时即时解析并增量填表(同行多列+换行交替都支持,去重,填完自动保存)';txt=text_edit.toPlainText().strip()
			if not txt:return
			for line in txt.split(_a):
				line=line.strip()
				if line and os.path.isfile(line):self._parse_mapping_file(line,table_widget)
			pairs=self._parse_mapping_text_to_pairs(txt)
			if pairs:self._add_rows_to_table(pairs,table_widget)
			if which:self._save_phone_mapping(which)
		@staticmethod
		def _detect_columns_and_parse(data_rows):
			'自动检测列顺序：手机号（11位数字）vs TraceID（纯数字非11位）'
			if not data_rows:return[]
			header=data_rows[0];col_idx=[0,1]
			if len(header)>=2:
				h0=str(header[0]).strip().lower();h1=str(header[1]).strip().lower();phone_keywords=['手机',_V,'mobile','号码','tel'];trace_keywords=['trace','跟踪','id','用户跟踪'];h0_is_phone=any(k in h0 for k in phone_keywords);h1_is_phone=any(k in h1 for k in phone_keywords);h0_is_trace=any(k in h0 for k in trace_keywords);h1_is_trace=any(k in h1 for k in trace_keywords)
				if h0_is_phone and h1_is_trace:col_idx=[0,1]
				elif h0_is_trace and h1_is_phone:col_idx=[1,0]
				else:
					col_idx=[0,1]
					if len(data_rows)>1:
						v0=str(data_rows[1][0]).strip()if len(data_rows[1])>0 else'';v1=str(data_rows[1][1]).strip()if len(data_rows[1])>1 else''
						if v0.isdigit()and len(v0)==11 and v1.isdigit()and len(v1)!=11:col_idx=[0,1]
						elif v0.isdigit()and len(v0)!=11 and v1.isdigit()and len(v1)==11:col_idx=[1,0]
			results=[];start_row=1 if len(data_rows)>1 else 0
			for row in data_rows[start_row:]:
				if len(row)>=2:
					phone=str(row[col_idx[0]]).strip()if len(row)>col_idx[0]else'';tid=str(row[col_idx[1]]).strip()if len(row)>col_idx[1]else''
					if phone and tid:results.append((phone,tid))
			return results
		def _add_rows_to_table(self,rows,table_widget):
			'将解析结果添加到表格（去重）；批量写入时屏蔽cellChanged避免频繁保存';which=_d if table_widget is self.phone_table else _z;table_widget.blockSignals(_B)
			try:
				existing=set()
				for row in range(table_widget.rowCount()):
					phone_item=table_widget.item(row,0);tid_item=table_widget.item(row,1)
					if phone_item and tid_item:existing.add((phone_item.text().strip(),tid_item.text().strip()))
				for(phone,tid)in rows:
					if(phone,tid)not in existing:row=table_widget.rowCount();table_widget.insertRow(row);table_widget.setItem(row,0,QTableWidgetItem(phone));table_widget.setItem(row,1,QTableWidgetItem(tid));del_btn=QPushButton('删除');del_btn.clicked.connect(lambda checked,b=del_btn:self._remove_phone_mapping_row(table_widget,b,which));table_widget.setCellWidget(row,2,del_btn);existing.add((phone,tid))
			finally:table_widget.blockSignals(_D)
		def eventFilter(self,obj,event):
			if event.type()==QEvent.DragEnter:self.dragEnterEvent(event);return _B
			elif event.type()==QEvent.Drop:self.dropEvent(event);return _B
			return super().eventFilter(obj,event)
		def dragEnterEvent(self,event):
			if event.mimeData().hasUrls():event.acceptProposedAction()
			else:event.ignore()
		def dropEvent(self,event):
			A='vqi_lt';new=[]
			for u in event.mimeData().urls():
				f=u.toLocalFile()
				if os.path.isdir(f):
					for(root,dirs,files)in os.walk(f):
						for fn in files:
							low=fn.lower()
							if low.endswith((_AI,_AZ,_Aa,_AJ,_AW,'.tmf')):new.append(os.path.join(root,fn))
				else:
					low=f.lower()
					if low.endswith((_AI,_AZ,_Aa,_AJ,_AW,'.tmf')):new.append(f)
			if new:
				_time_found=_A;_time_patterns=['(\\d{1,2})[点点:：](\\d{1,2})[到至\\-~]+(\\d{1,2})[点点:：](\\d{1,2})','(\\d{1,2})时(\\d{1,2})分[到至\\-~]+(\\d{1,2})时(\\d{1,2})分','(\\d{1,2})点(\\d{1,2})?[到至\\-~]+(\\d{1,2})点(\\d{1,2})?']
				for f in new:
					for pat in _time_patterns:
						m=re.search(pat,f)
						if m:
							g=m.groups()
							if len(g)==4 and(g[1]is _A or g[3]is _A):h1,h2=int(g[0]),int(g[2]);m1=int(g[1])if g[1]else 0;m2=int(g[3])if g[3]else 0
							else:h1,m1,h2,m2=int(g[0]),int(g[1]),int(g[2]),int(g[3])
							if 0<=h1<24 and 0<=m1<60 and 0<=h2<24 and 0<=m2<60:
								if _time_found and _time_found!=(h1,m1,h2,m2):self.lt.append('检测到多个不同时间范围，请手动设置');_time_found=_A;break
								_time_found=h1,m1,h2,m2
					if _time_found is _A and pat!=_time_patterns[-1]:continue
				if _time_found:
					h1,m1,h2,m2=_time_found;current_tab=self.centralWidget().currentIndex()
					if current_tab==1:
						self.vqi_cb_time.setChecked(_B);self.vqi_time_start.setTime(QTime(h1,m1));self.vqi_time_end.setTime(QTime(h2,m2))
						if hasattr(self,A):self.vqi_lt.append(f"自动识别时间: {h1:02d}:{m1:02d} ~ {h2:02d}:{m2:02d}")
					else:self.cb_time.setChecked(_B);self.time_start.setTime(QTime(h1,m1));self.time_end.setTime(QTime(h2,m2));self.lt.append(f"自动识别时间: {h1:02d}:{m1:02d} ~ {h2:02d}:{m2:02d}")
				current_tab=self.centralWidget().currentIndex()
				if current_tab==1:
					self.vqi_files=list(dict.fromkeys((self.vqi_files or[])+new));display_lines=[os.path.basename(f)for f in self.vqi_files];self.vqi_te.setText(_a.join(display_lines));self.vqi_br.setEnabled(_B)
					if hasattr(self,A):self.vqi_lt.append(f"拖入 {len(new)} 个文件到VQI，当前共 {len(self.vqi_files)} 个")
				else:self.files=list(dict.fromkeys((self.files or[])+new));display_lines=[os.path.basename(f)for f in self.files];self.te.setText(_a.join(display_lines));self.br.setEnabled(_B);self.lt.append(f"拖入 {len(new)} 个文件，当前共 {len(self.files)} 个")
			event.acceptProposedAction()
		def _setup_phone_textedit_drop(self,text_edit):
			'为手机号映射QTextEdit设置独立的拖放处理';A='.txt';text_edit.setAcceptDrops(_B)
			def drag_enter(e):
				if e.mimeData().hasUrls():
					for u in e.mimeData().urls():
						f=u.toLocalFile();low=f.lower()
						if low.endswith((_AI,_AZ,A,_AJ)):e.acceptProposedAction();return
				e.ignore()
			def drop(e):
				paths=[]
				for u in e.mimeData().urls():
					f=u.toLocalFile();low=f.lower()
					if low.endswith((_AI,_AZ,A,_AJ)):paths.append(f)
				if paths:text_edit.setText(_a.join(paths));e.acceptProposedAction()
				else:e.ignore()
			text_edit.dragEnterEvent=drag_enter;text_edit.dropEvent=drop
		def disable_child_drops(self,widget):
			for child in widget.findChildren(QWidget):child.setAcceptDrops(_D);child.installEventFilter(self);self.disable_child_drops(child)
		def _build(self):self.setAcceptDrops(_B);tabs=QTabWidget();tabs.addTab(self._build_rate_tab(),'工信部数据统计工具');tabs.addTab(self._build_vqi_tab(),'工信部语音VQI工具');self.setCentralWidget(tabs);self.disable_child_drops(self)
		def _build_rate_tab(self):
			C='冗余%';B='方案二';A='方案一';sp=QSplitter(Qt.Horizontal);left=QWidget();ll=QVBoxLayout(left);ll.setContentsMargins(4,4,4,4);ll.setSpacing(4);title=QLabel('工信部数据统计工具 V2.5.35_CC');title.setFont(self._font(2));title.setAlignment(Qt.AlignCenter);ll.addWidget(title);vs=QSplitter(Qt.Vertical);vs.setChildrenCollapsible(_D);gf=QGroupBox('输入文件');gfl=QVBoxLayout(gf);gfl.setContentsMargins(4,4,4,4);gfl.setSpacing(2);bf=QHBoxLayout();bf.setSpacing(4);self.bs=QPushButton('选择MMF文件');self.bs.clicked.connect(self.sf);self.bc=QPushButton('清空');self.bc.clicked.connect(self.cf);bf.addWidget(self.bs);bf.addWidget(self.bc);gfl.addLayout(bf);self.te=QTextEdit();self.te.setReadOnly(_B);self.te.setMaximumHeight(300);self.te.setPlaceholderText(_BF);self.te.setStyleSheet(_BG);gfl.addWidget(self.te);vs.addWidget(gf);gp=QGroupBox('参数设置');_outer=QVBoxLayout(gp);_outer.setContentsMargins(4,4,4,4);_outer.setSpacing(3);_params_top=QHBoxLayout();_params_top.setSpacing(6);_params_top.addWidget(QLabel('QCI:'));self.cb6=QCheckBox('6');self.cb7=QCheckBox('7');self.cb6.setChecked(_B);self.cb7.setChecked(_B);_params_top.addWidget(self.cb6);_params_top.addWidget(self.cb7);_params_top.addWidget(QLabel('下行削峰:'));self.sd=QSpinBox();self.sd.setRange(100,5000);self.sd.setValue(1000);self.sd.setMaximumWidth(65);_params_top.addWidget(self.sd);_params_top.addWidget(QLabel('上行削峰:'));self.su=QSpinBox();self.su.setRange(50,1000);self.su.setValue(200);self.su.setMaximumWidth(65);_params_top.addWidget(self.su);_params_top.addWidget(QLabel('FTP:'));self.ftp_duration_spin=QSpinBox();self.ftp_duration_spin.setRange(5,120);self.ftp_duration_spin.setValue(25);self.ftp_duration_spin.setSuffix('秒');self.ftp_duration_spin.setMaximumWidth(65);_params_top.addWidget(self.ftp_duration_spin);_params_top.addStretch();_outer.addLayout(_params_top);_params_mid=QHBoxLayout();_params_mid.setSpacing(6);self.cb_time=QCheckBox();_params_mid.addWidget(self.cb_time);_params_mid.addWidget(QLabel('开始:'));self.time_start=QTimeEdit();self.time_start.setDisplayFormat(_AS);self.time_start.setTime(QTime(0,0,0));self.time_start.editingFinished.connect(lambda:self.cb_time.setChecked(_B));_params_mid.addWidget(self.time_start);_params_mid.addWidget(QLabel('结束:'));self.time_end=QTimeEdit();self.time_end.setDisplayFormat(_AS);self.time_end.setTime(QTime(23,59,59));self.time_end.editingFinished.connect(lambda:self.cb_time.setChecked(_B));_params_mid.addWidget(self.time_end);self.cb_merge_raw=QCheckBox('合并原始数据');self.cb_merge_raw.setChecked(_D);_params_mid.addWidget(self.cb_merge_raw);self.cb_annotations=QCheckBox('添加批注');self.cb_annotations.setChecked(_B);_params_mid.addWidget(self.cb_annotations);_params_mid.addStretch();_outer.addLayout(_params_mid);_BIZ_SCHEMES={A:[_L,_K,_m,_U,_g,_R],B:[_K,_L,_g,_R,_m,_U]};self._current_biz_scheme=A;_scheme_label_row=QHBoxLayout();_scheme_label_row.addWidget(QLabel('快速方案:'));_btn_s1=QPushButton(A);_btn_s2=QPushButton(B);_scheme_label_row.addWidget(_btn_s1);_scheme_label_row.addWidget(_btn_s2);_scheme_label_row.addStretch();_outer.addLayout(_scheme_label_row);_outer.addWidget(QLabel('业务执行顺序:'));self.biz_order_table=QTableWidget();self.biz_order_table.setColumnCount(7);self.biz_order_table.setHorizontalHeaderLabels([_J,'流量从(Mbps)','到',C,'时长从(秒)','到',C]);hh=self.biz_order_table.horizontalHeader();_biz_settings=QSettings(_Aj,_Ak);_biz_defaults=[90,55,40,0,55,40,0]
			for(_ci,_cw)in enumerate(_biz_defaults):
				if _ci in(3,6):hh.setSectionResizeMode(_ci,QHeaderView.Stretch)
				else:
					_saved=_biz_settings.value(f"biz_col_{_ci}",_cw)
					try:_cw=int(_saved)
					except:pass
					hh.setSectionResizeMode(_ci,QHeaderView.Interactive);self.biz_order_table.setColumnWidth(_ci,_cw)
			def _save_biz_cols(_ci,_os,_ns):QSettings(_Aj,_Ak).setValue(f"biz_col_{_ci}",_ns)
			hh.sectionResized.connect(_save_biz_cols);self.biz_order_table.verticalHeader().setDefaultSectionSize(22);self.biz_order_table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed);self.biz_order_table.horizontalHeader().setSectionsClickable(_D);self.biz_order_table.setMaximumHeight(155);self.biz_order_table.setSelectionMode(QTableWidget.MultiSelection);self.biz_order_table.setSelectionBehavior(QTableWidget.SelectRows);self.biz_order_table.setEditTriggers(QTableWidget.DoubleClicked);self.biz_order_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff);_ALL_BIZ=[_L,_K,_m,_U,_g,_R]
			for r in range(6):
				self.biz_order_table.insertRow(r);_item=QTableWidgetItem(_ALL_BIZ[r]);self.biz_order_table.setItem(r,0,_item)
				for c in range(1,7):_item=QTableWidgetItem('');_item.setTextAlignment(Qt.AlignCenter);self.biz_order_table.setItem(r,c,_item)
			_biz_btn_row=QHBoxLayout();_biz_btn_row.addWidget(self.biz_order_table);_btn_up=QPushButton('↑上移');_btn_up.setMaximumWidth(55);_btn_down=QPushButton('↓下移');_btn_down.setMaximumWidth(55);_btn_clear=QPushButton('清空');_btn_clear.setMaximumWidth(55);_biz_btn_style='QPushButton{background-color:#e8f0fe;border:1px solid #c0c0c0;border-radius:3px;padding:2px 4px;}';_btn_up.setStyleSheet(_biz_btn_style);_btn_down.setStyleSheet(_biz_btn_style);_btn_clear.setStyleSheet(_biz_btn_style);_btn_up.setFont(self._font());_btn_down.setFont(self._font());_btn_clear.setFont(self._font());_btn_up.clicked.connect(lambda:self._move_biz_item(-1));_btn_down.clicked.connect(lambda:self._move_biz_item(1));_btn_clear.clicked.connect(self._clear_biz_params);_ud_btns=QVBoxLayout();_ud_btns.addWidget(_btn_up);_ud_btns.addWidget(_btn_down);_ud_btns.addWidget(_btn_clear);_biz_btn_row.addLayout(_ud_btns);_outer.addLayout(_biz_btn_row)
			def _apply_scheme(name):
				C='background-color: #c8e6c9; font-weight: bold;';self._current_biz_scheme=name;order=_BIZ_SCHEMES[name]
				for r in range(6):
					_item=self.biz_order_table.item(r,0)
					if _item:_item.setText(order[r])
				_btn_s1.setStyleSheet(C if name==A else'');_btn_s2.setStyleSheet(C if name==B else'')
			_btn_s1.clicked.connect(lambda:_apply_scheme(A));_btn_s2.clicked.connect(lambda:_apply_scheme(B));_apply_scheme(A);vs.addWidget(gp);self.phone_table=QTableWidget();self.phone_textedit=QTextEdit();self.phone_add_btn=QPushButton();self.phone_clear_btn=QPushButton();gp_phone=self._build_phone_mapping_group(self.phone_table,self.phone_textedit,self.phone_add_btn,self.phone_clear_btn,_d);vs.addWidget(gp_phone);self._setup_phone_textedit_drop(self.phone_textedit);gr=QFrame();gr.setFrameStyle(QFrame.Box);grl=QVBoxLayout(gr);grl.setContentsMargins(2,2,2,2);grl.setSpacing(1);_btn_row=QHBoxLayout();_btn_row.setSpacing(4);self.br=QPushButton('开始处理');self.br.clicked.connect(self.run);self.br.setEnabled(_D);self.bcancel=QPushButton('取消');self.bcancel.clicked.connect(self.cancel);self.bcancel.setEnabled(_D);self.babout=QPushButton('关于');self.babout.clicked.connect(self.about);_btn_style=_BH;self.br.setStyleSheet(_btn_style);self.bcancel.setStyleSheet(_btn_style);self.babout.setStyleSheet(_btn_style);self.br.setFont(self._font());self.bcancel.setFont(self._font());self.babout.setFont(self._font());_btn_row.addWidget(self.br,1);_btn_row.addWidget(self.bcancel,1);_btn_row.addWidget(self.babout,1);grl.addLayout(_btn_row);pb_open=QHBoxLayout();self.pb=QProgressBar();pb_open.addWidget(self.pb);self.bopen=QPushButton('打开输出文件');self.bopen.clicked.connect(self.open_out);self.bopen.setEnabled(_D);pb_open.addWidget(self.bopen);grl.addLayout(pb_open);self.lt=QTextEdit();self.lt.setReadOnly(_B);self.lt.setMaximumHeight(100);self.lt.setStyleSheet(_BI);grl.addWidget(self.lt);vs.addWidget(gr);vs.setSizes([200,350,200,100]);ll.addWidget(vs,1);left.setFont(self._font());right=QWidget();rl=QVBoxLayout(right);rh=QLabel('结果数据（内嵌查看，默认筛选A=1）');rh.setFont(self._font(1));rl.addWidget(rh);self.result_tabs=QTabWidget();rl.addWidget(self.result_tabs);right.setFont(self._font());sp.addWidget(left);sp.addWidget(right);sp.setStretchFactor(0,1);sp.setStretchFactor(1,5);sp.setSizes([int(self.scr_w*.167),int(self.scr_w*.833)]);return sp
		def _build_vqi_tab(self):sp=QSplitter(Qt.Horizontal);left=QWidget();ll=QVBoxLayout(left);ll.setContentsMargins(4,4,4,4);ll.setSpacing(4);title=QLabel('工信部语音VQI工具 V2.5.35_CC');title.setFont(self._font(2));title.setAlignment(Qt.AlignCenter);ll.addWidget(title);vs=QSplitter(Qt.Vertical);vs.setChildrenCollapsible(_D);gf=QGroupBox('输入文件');gfl=QVBoxLayout(gf);gfl.setContentsMargins(4,4,4,4);gfl.setSpacing(2);bf=QHBoxLayout();bf.setSpacing(4);self.vqi_bs=QPushButton('选择MMF文件');self.vqi_bs.clicked.connect(self.vqi_sf);self.vqi_bc=QPushButton('清空');self.vqi_bc.clicked.connect(self.vqi_cf);bf.addWidget(self.vqi_bs);bf.addWidget(self.vqi_bc);gfl.addLayout(bf);self.vqi_te=QTextEdit();self.vqi_te.setReadOnly(_B);self.vqi_te.setMaximumHeight(300);self.vqi_te.setPlaceholderText(_BF);self.vqi_te.setStyleSheet(_BG);gfl.addWidget(self.vqi_te);vs.addWidget(gf);gp=QGroupBox('参数设置');gpl=QGridLayout(gp);gpl.setContentsMargins(4,4,4,4);gpl.setSpacing(2);self.vqi_cb_time=QCheckBox();gpl.addWidget(self.vqi_cb_time,0,0);gpl.addWidget(QLabel('开始时间:'),0,1);self.vqi_time_start=QTimeEdit();self.vqi_time_start.setDisplayFormat(_AS);self.vqi_time_start.setTime(QTime(0,0,0));self.vqi_time_start.editingFinished.connect(lambda:self.vqi_cb_time.setChecked(_B));gpl.addWidget(self.vqi_time_start,0,2);gpl.addWidget(QLabel('结束时间:'),1,1);self.vqi_time_end=QTimeEdit();self.vqi_time_end.setDisplayFormat(_AS);self.vqi_time_end.setTime(QTime(23,59,59));self.vqi_time_end.editingFinished.connect(lambda:self.vqi_cb_time.setChecked(_B));gpl.addWidget(self.vqi_time_end,1,2);self.vqi_cb_merge_raw=QCheckBox('合并原始数据到结果文件');self.vqi_cb_merge_raw.setChecked(_D);gpl.addWidget(self.vqi_cb_merge_raw,2,0,1,3);self.vqi_cb_annotations=QCheckBox('添加批注到结果文件');self.vqi_cb_annotations.setChecked(_B);gpl.addWidget(self.vqi_cb_annotations,3,0,1,3);vs.addWidget(gp);self.vqi_phone_table=QTableWidget();self.vqi_phone_textedit=QTextEdit();self.vqi_phone_add_btn=QPushButton();self.vqi_phone_clear_btn=QPushButton();gp_vqi_phone=self._build_phone_mapping_group(self.vqi_phone_table,self.vqi_phone_textedit,self.vqi_phone_add_btn,self.vqi_phone_clear_btn,_z);vs.addWidget(gp_vqi_phone);self._setup_phone_textedit_drop(self.vqi_phone_textedit);gr=QFrame();gr.setFrameStyle(QFrame.Box);grl=QVBoxLayout(gr);grl.setContentsMargins(2,2,2,2);grl.setSpacing(1);_btn_row=QHBoxLayout();_btn_row.setSpacing(4);self.vqi_br=QPushButton('开始处理');self.vqi_br.clicked.connect(self.vqi_run);self.vqi_br.setEnabled(_D);self.vqi_bcancel=QPushButton('取消');self.vqi_bcancel.clicked.connect(self.vqi_cancel);self.vqi_bcancel.setEnabled(_D);self.vqi_babout=QPushButton('关于');self.vqi_babout.clicked.connect(self.about);_vqi_btn_style=_BH;self.vqi_br.setStyleSheet(_vqi_btn_style);self.vqi_bcancel.setStyleSheet(_vqi_btn_style);self.vqi_babout.setStyleSheet(_vqi_btn_style);self.vqi_br.setFont(self._font());self.vqi_bcancel.setFont(self._font());self.vqi_babout.setFont(self._font());_btn_row.addWidget(self.vqi_br,1);_btn_row.addWidget(self.vqi_bcancel,1);_btn_row.addWidget(self.vqi_babout,1);grl.addLayout(_btn_row);pb_open=QHBoxLayout();self.vqi_pb=QProgressBar();pb_open.addWidget(self.vqi_pb);self.vqi_bopen=QPushButton('打开输出文件');self.vqi_bopen.clicked.connect(self.vqi_open_out);self.vqi_bopen.setEnabled(_D);pb_open.addWidget(self.vqi_bopen);grl.addLayout(pb_open);self.vqi_lt=QTextEdit();self.vqi_lt.setReadOnly(_B);self.vqi_lt.setMaximumHeight(100);self.vqi_lt.setStyleSheet(_BI);grl.addWidget(self.vqi_lt);vs.addWidget(gr);vs.setSizes([150,150,200,100]);ll.addWidget(vs,1);left.setFont(self._font());right=QWidget();rl=QVBoxLayout(right);rh=QLabel('结果数据（内嵌查看）');rh.setFont(self._font(1));rl.addWidget(rh);self.vqi_result_tabs=QTabWidget();self.vqi_tbl_raw=QTableWidget();self.vqi_tbl_raw.setSelectionBehavior(QTableWidget.SelectItems);self.vqi_tbl_voice=QTableWidget();self.vqi_tbl_voice.setSelectionBehavior(QTableWidget.SelectItems);self.vqi_tbl_mos=QTableWidget();self.vqi_tbl_mos.setSelectionBehavior(QTableWidget.SelectItems);self.vqi_result_tabs.addTab(self.vqi_tbl_raw,'原始数据');self.vqi_result_tabs.addTab(self.vqi_tbl_voice,_AP);self.vqi_result_tabs.addTab(self.vqi_tbl_mos,_AQ);self.vqi_tbl_data_raw=QTableWidget();self.vqi_tbl_data_raw.setSelectionBehavior(QTableWidget.SelectItems);self.vqi_result_tabs.addTab(self.vqi_tbl_data_raw,'数据业务-原始文件');rl.addWidget(self.vqi_result_tabs);right.setFont(self._font());sp.addWidget(left);sp.addWidget(right);sp.setStretchFactor(0,1);sp.setStretchFactor(1,5);sp.setSizes([int(self.scr_w*.167),int(self.scr_w*.833)]);return sp
		def vqi_sf(self):
			fs,_=QFileDialog.getOpenFileNames(self,'选择文件','','Excel/CSV/压缩 (*.xlsx *.xls *.csv *.zip);;All (*)')
			if fs:self.vqi_files=list(dict.fromkeys((self.vqi_files or[])+fs));self.vqi_te.setText(_a.join(os.path.basename(f)for f in self.vqi_files));self.vqi_br.setEnabled(_B)
		def vqi_cf(self):self.vqi_files=[];self.vqi_te.clear();self.vqi_br.setEnabled(_D)
		def vqi_run(self):
			ok,msg=check_license()
			if not ok:QMessageBox.warning(self,'鉴权失败',f"无法使用：{msg}\n请联网后重试。");return
			if not self.vqi_files:QMessageBox.warning(self,'提示',_BJ);return
			self._sync_phone_mapping(_z)
			if not self.phone_trace_map_vqi:
				ret=QMessageBox.question(self,_Ah,_Ai,QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No,QMessageBox.StandardButton.No)
				if ret!=QMessageBox.StandardButton.Yes:return
			self.vqi_tbl_raw.setRowCount(0);self.vqi_tbl_raw.setColumnCount(0);self.vqi_tbl_voice.setRowCount(0);self.vqi_tbl_voice.setColumnCount(0);self.vqi_tbl_mos.setRowCount(0);self.vqi_tbl_mos.setColumnCount(0);self.vqi_tbl_data_raw.setRowCount(0);self.vqi_tbl_data_raw.setColumnCount(0);self.vqi_lt.clear();time_filter=_A
			if self.vqi_cb_time.isChecked():t_start=self.vqi_time_start.time();t_end=self.vqi_time_end.time();time_filter=_B,t_start,t_end
			self.vqi_br.setEnabled(_D);self.vqi_bcancel.setEnabled(_B);self.vqi_pb.setValue(0);self.vqi_lt.append(f"开始处理...")
			if time_filter:self.vqi_lt.append(f"时间段过滤: {t_start.toString(_FMT_HMS)} ~ {t_end.toString(_FMT_HMS)}")
			self.vqi_thread=VqiWorkerThread(self.vqi_files,time_filter=time_filter,merge_raw=self.vqi_cb_merge_raw.isChecked(),add_annotations=self.vqi_cb_annotations.isChecked(),phone_trace_map=self.phone_trace_map_vqi);self.vqi_thread.progress.connect(self.vqi_upd);self.vqi_thread.progress_pct.connect(lambda p:self.vqi_pb.setValue(p));self.vqi_thread.done.connect(self.vqi_dn);self.vqi_thread.start()
		def vqi_cancel(self):
			if hasattr(self,'vqi_thread')and self.vqi_thread.isRunning():self.vqi_thread.cancel();self.vqi_lt.append(_BK)
		@Slot(str)
		def vqi_upd(self,msg):self.vqi_lt.append(msg)
		@Slot(str)
		def vqi_dn(self,path):
			self.vqi_bcancel.setEnabled(_D)
			if path:self.vqi_pb.setValue(100);self.vqi_last_out=path;self.vqi_bopen.setEnabled(_B);self.vqi_lt.append(f"完成! {path}");self.show_vqi_result(path);consume_quota();os.system(_Al+path+'"')
			else:self.vqi_pb.setValue(0);self.vqi_lt.append('已取消。')
			self.vqi_br.setEnabled(_B)
		def vqi_open_out(self):
			if self.vqi_last_out:os.system(_Al+self.vqi_last_out+'"')
		def show_vqi_result(self,path):
			try:
				import openpyxl;wb=openpyxl.load_workbook(path,data_only=_B);sheet_list=wb.sheetnames
				if _AY in sheet_list:self._fill_vqi_table(self.vqi_tbl_raw,wb[_AY],path=path,sheet=_AY)
				else:self.vqi_tbl_raw.setRowCount(0);self.vqi_tbl_raw.setColumnCount(0)
				if _AP in sheet_list:self._fill_vqi_table(self.vqi_tbl_voice,wb[_AP],path=path,sheet=_AP)
				else:self.vqi_tbl_voice.setRowCount(0);self.vqi_tbl_voice.setColumnCount(0)
				if _AQ in sheet_list:self._fill_vqi_table(self.vqi_tbl_mos,wb[_AQ],path=path,sheet=_AQ)
				elif _AF in sheet_list:self._fill_vqi_table(self.vqi_tbl_mos,wb[_AF],path=path,sheet=_AF)
				else:self.vqi_tbl_mos.setRowCount(0);self.vqi_tbl_mos.setColumnCount(0)
				cnt=len(sheet_list);self.vqi_lt.append(f"结果已内嵌显示（{cnt}个Sheet）")
			except Exception as e:self.vqi_lt.append(f"内嵌显示失败: {e}")
		def _fill_vqi_table(self,tbl,ws,path,sheet):
			from PySide6.QtGui import QColor;tbl.clearContents();tbl.setRowCount(0);rows=list(ws.iter_rows())
			if not rows:return
			headers=[str(c.value)if c.value is not _A else''for c in rows[0]];tbl.setColumnCount(len(headers));tbl.setHorizontalHeaderLabels(headers);r_idx=0
			for row in rows[1:]:
				tbl.insertRow(r_idx)
				for(ci,cell)in enumerate(row):
					v=cell.value;item=QTableWidgetItem(''if v is _A else str(v))
					try:
						fill=cell.fill
						if fill and fill.start_color and fill.start_color.rgb:
							rgb=str(fill.start_color.rgb)
							if rgb not in('00000000','0'):item.setBackground(QColor('#'+rgb[-6:]))
					except Exception:pass
					tbl.setItem(r_idx,ci,item)
				r_idx+=1
			tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive);tbl.setSortingEnabled(_D)
		def sf(self):
			fs,_=QFileDialog.getOpenFileNames(self,'选择文件','','所有支持文件 (*.xlsx *.xls *.csv *.zip *.mmf *.tmf);;Excel/CSV (*.xlsx *.xls *.csv);;All (*)')
			if fs:
				for f in fs:
					fn=os.path.basename(f).lower()
					if'基准'in fn or'呼叫详情'in fn or'整理'in fn:self.base_file=f;self.lt.append(f"已选择基准文件: {os.path.basename(f)}")
					else:self.files=list(dict.fromkeys((self.files or[])+[f]))
				if self.files:self.te.setText(_a.join(os.path.basename(f)for f in self.files));self.br.setEnabled(_B)
				if self.base_file:self.te.append(f"\n[基准] {os.path.basename(self.base_file)}")
		def cf(self):self.files=[];self.base_file=_A;self.te.clear();self.br.setEnabled(_D)
		@staticmethod
		def _swap_rows(table,r1,r2):
			'交换QTableWidget中两行的文本（不移动widget，只交换显示内容避免崩溃）'
			for c in range(table.columnCount()):
				i1=table.item(r1,c);i2=table.item(r2,c);w1=table.cellWidget(r1,c);w2=table.cellWidget(r2,c);t1=i1.text()if i1 else w1.text()if w1 else'';t2=i2.text()if i2 else w2.text()if w2 else''
				if i1:i1.setText(t2)
				elif w1:w1.setText(t2)
				if i2:i2.setText(t1)
				elif w2:w2.setText(t1)
		def _move_biz_item(self,direction):
			'移动业务顺序表格中的选中行（支持多行一起移动）: direction=-1上移, 1下移';rows=sorted(set(r.row()for r in self.biz_order_table.selectedIndexes()))
			if not rows:return
			if direction==-1:
				if rows[0]<=0:return
				for r in rows:MainWindow._swap_rows(self.biz_order_table,r,r-1)
				self.biz_order_table.clearSelection()
				for r in rows:self.biz_order_table.selectRow(r-1)
			else:
				if rows[-1]>=self.biz_order_table.rowCount()-1:return
				for r in reversed(rows):MainWindow._swap_rows(self.biz_order_table,r,r+1)
				self.biz_order_table.clearSelection()
				for r in rows:self.biz_order_table.selectRow(r+1)
		def _clear_biz_params(self):
			'清空所有业务辅助参数（流量/时长输入框）'
			for r in range(self.biz_order_table.rowCount()):
				for c in range(1,7):
					_item=self.biz_order_table.item(r,c)
					if _item:_item.setText('')
		def run(self):
			if not self.files:QMessageBox.warning(self,'提示',_BJ);return
			self._sync_phone_mapping(_d)
			if not self.phone_trace_map_data:
				ret=QMessageBox.question(self,_Ah,_Ai,QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No,QMessageBox.StandardButton.No)
				if ret!=QMessageBox.StandardButton.Yes:return
			self.result_tabs.clear();self.result_tabs.clear();self.lt.clear();qci_list=[]
			if self.cb6.isChecked():qci_list.append(6)
			if self.cb7.isChecked():qci_list.append(7)
			if not qci_list:qci_list=[7]
			time_filter=_A
			if self.cb_time.isChecked():t_start=self.time_start.time();t_end=self.time_end.time();time_filter=_B,t_start,t_end
			ftp_duration=self.ftp_duration_spin.value();biz_order=[];biz_params={}
			if hasattr(self,_BC):
				for r in range(self.biz_order_table.rowCount()):
					_biz_item=self.biz_order_table.item(r,0)
					if not _biz_item or not _biz_item.text():continue
					_biz=_biz_item.text();biz_order.append(_biz);_vals=[self.biz_order_table.item(r,c).text().strip()if self.biz_order_table.item(r,c)else''for c in range(1,7)]
					if all(_vals):biz_params[_biz]={_As:float(_vals[0]),_At:float(_vals[1]),_Au:float(_vals[2])/1e2,_Av:float(_vals[3]),_Aw:float(_vals[4]),_Ax:float(_vals[5])/1e2}
			self.br.setEnabled(_D);self.bcancel.setEnabled(_B);self.pb.setValue(0);self.lt.append(f"开始处理数据业务... QCI={qci_list}, FTP时长={ftp_duration}秒")
			if time_filter:self.lt.append(f"时间段过滤: {t_start.toString(_FMT_HMS)} ~ {t_end.toString(_FMT_HMS)}")
			self.thread=WorkerThread(self.files,qci_list,self.sd.value(),self.su.value(),base_file=self.base_file,time_filter=time_filter,ftp_duration=ftp_duration,merge_raw=self.cb_merge_raw.isChecked(),add_annotations=self.cb_annotations.isChecked(),phone_trace_map=self.phone_trace_map_data,biz_order=biz_order,biz_params=biz_params);self.thread.progress.connect(self.upd);self.thread.progress_pct.connect(lambda p:self.pb.setValue(p));self.thread.done.connect(self.dn);self.thread.start()
		def cancel(self):
			if hasattr(self,'thread')and self.thread.isRunning():self.thread.cancel();self.lt.append(_BK)
		@Slot(str)
		def upd(self,msg):self.lt.append(msg)
		@Slot(str)
		def dn(self,path):
			self.bcancel.setEnabled(_D)
			if path:self.pb.setValue(100);self.last_out=path;self.bopen.setEnabled(_B);self.lt.append(f"完成! {path}");self.show_result(path);consume_quota();os.system(_Al+path+'"')
			else:self.pb.setValue(0);self.lt.append('已取消。')
			self.br.setEnabled(_B)
		def open_out(self):
			if self.last_out:os.system(_Al+self.last_out+'"')
		def show_result(self,path):
			try:
				import openpyxl;from PySide6.QtGui import QColor;wb=openpyxl.load_workbook(path,data_only=_B);sheet_names=wb.sheetnames;self.result_tabs.clear();_biz_sheets=[s for s in sheet_names if'语音'not in s and s not in('提示',)]
				for sn in _biz_sheets:tbl=QTableWidget();tbl.setSelectionBehavior(QTableWidget.SelectItems);self.result_tabs.addTab(tbl,sn);self._fill_table(tbl,wb[sn],filter_a1='-详细过程'in sn,path=path,sheet=sn)
				self.lt.append('结果已内嵌显示（详细过程默认筛选A=1）')
			except Exception as e:self.lt.append(f"内嵌显示失败: {e}")
		def _fill_table(self,tbl,ws,filter_a1,path,sheet):
			from PySide6.QtGui import QColor;tbl.clearContents();tbl.setRowCount(0);rows=list(ws.iter_rows())
			if not rows:return
			headers=[str(c.value)if c.value is not _A else''for c in rows[0]];tbl.setColumnCount(len(headers));tbl.setHorizontalHeaderLabels(headers);r_idx=0
			for(ri,row)in enumerate(rows[1:],start=2):
				if filter_a1:
					a=row[0].value if row else _A
					if a is _A or str(a)!='1':continue
				tbl.insertRow(r_idx)
				for(ci,cell)in enumerate(row):
					v=cell.value;item=QTableWidgetItem(''if v is _A else str(v))
					try:
						fill=cell.fill
						if fill and fill.start_color and fill.start_color.rgb:
							rgb=str(fill.start_color.rgb)
							if rgb not in('00000000','0'):item.setBackground(QColor('#'+rgb[-6:]))
					except Exception:pass
					try:
						if cell.comment:item.setToolTip(cell.comment.text)
					except Exception:pass
					tbl.setItem(r_idx,ci,item)
				r_idx+=1
			tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive);tbl.setSortingEnabled(_D)
		def about(self):d=QDialog(self);d.setWindowTitle('关于');d.resize(560,600);v=QVBoxLayout(d);te=QTextEdit();te.setReadOnly(_B);te.setHtml('<h3 style=\'text-align:center\'>工信部数据、语音跟踪统计工具</h3><p><b>版本:</b> V2.5.35_CC &nbsp; <b>开发者:</b> 孙晓军 &nbsp; <b>联系方式:</b> 317827@qq.com</p><p><b>鸣谢：</b>丁先甲，谭润林，赵永川，廖洋（排名不分先后）</p><hr/><h4>更新记录</h4><p><b>2026-07-21</b></p><ul><li><b>V2.5.35</b> 防误判优化（短脉冲不再识别为业务）；界面微调（冻结列移位、结果区tab数量跟随输出、速率列整数显示）</li>\n<li><b>V2.5.34</b> 业务识别纠错（FTP配对时长短边自动对齐、首尾弱信号自动剔除）；防误判加固（应用商店类最短识别时长收紧）</li>\n<li><b>V2.5.33</b> 业务段自动修剪（去掉头尾无效数据点）；状态机增加兜底扫描（解决遗漏大业务的问题）</li>\n<li><b>V2.5.32</b> 业务识别规则细化（应用商店类和微信类增加最短时长限制，防止1秒短脉冲误判）</li>\n<li><b>V2.5.31</b> 新增子段修剪功能（计算均值时自动跳过首尾低活跃秒）；无法判断的数据标注为未知不再强行分类；支持用户自定义业务流量和时长辅助判断参数</li>\n<li><b>V2.5.30</b> 业务执行顺序升级为表格配置（支持每项业务单独设置流量范围、时长范围、上下浮动百分比）；界面大改版（参数区横排紧凑、手机号映射左右并排、速率列整数显示）</li>\n<li><b>V2.5.29</b> 界面布局重构（参数设置从竖排改为左右双栏）；手机号映射改为左右并排更省空间；去掉冗长的操作说明文字</li>\n<li><b>V2.5.28</b> 修正削峰Top10%峰值速率的计算方式（先按设定上限裁切再取前10%平均，之前是取原始值没有裁切）</li>\n<li><b>V2.5.27</b> 业务识别升级为顺序状态机（按设定的业务执行顺序依次识别，不再使用固定规则，大幅减少误判和遗漏）</li>\n<li><b>V2.5.26</b> 修复应用商店大文件下载被误判为FTP下载的问题（调整识别优先级）</li>\n<li><b>V2.5.25</b> 新增业务顺序快速切换（方案一和方案二两套预设，一键切换）</li>\n<li><b>V2.5.24</b> 增强短段识别（4秒高下载突发等不再遗漏）；新增未分类段自动归类</li>\n<li><b>V2.5.23</b> 基准列数据填充完善（不仅填业务名和时间，速率指标和时长也一并填入）；基准文件加载兼容性提升；速率校验放宽减少误过滤</li>\n<li><b>V2.5.22</b> RLC速率汇总表改为横表（联通和电信各独立列，数值精确到2位小数）；详细过程基准列补全实际数据</li>\n<li><b>V2.5.20</b> 新增分业务RLC速率汇总Sheet（按业务类型统计RLC吞吐率，联通/电信分开，自动转换单位）</li>\n<li><b>V2.5.19</b> 基准对齐算法升级（从时间匹配改为业务序列匹配，解决永川基准与实测数据时间不同步的问题）</li>\n<li><b>V2.5.18</b> 永川基准对齐增强（引入速率辅助校验提高匹配准确度）；业务执行顺序改为可拖拽排序；放宽对齐时间窗口</li>\n</ul><p><b>2026-07-19</b></p><ul><li><b>V2.5.17_CC</b> 性能优化: DataFrame碎片化修复(assign替代逐列赋值)、详细过程Sheet改用write_row批量写(避免6亿次Python调用→560万次)、语音VQI原始数据Sheet改用xlsxwriter流式写入+constant_memory(522万行不再卡死)、列宽固定宽度、NaN用nan_inf_to_errors</li>\n<li><b>V2.5.16_原</b> 完整通话规则改为只看BYE(不再找200);RSRP拆分为(全部)和(通话时)两列;批注增加每轮次明细;汇总表/按文件名表列顺序调整(新增主被叫/手机号/类型);删除运营商表;新增12个数据明细sheet(语音6+数据6);运营商检测增加QCI兜底(路径无法判断时按QCI:7→电信/5或6→联通);默认值调整(QCI6+7都勾选,合并原始数据默认开,添加批注默认开);数据业务写入从openpyxl改为xlsxwriter逐行写(提速31%,可中断);手机号列构建性能优化(解决大量文件时卡顿);临时文件安全写入(取消不留坏文件)</li>\n<li><b>V2.5.13</b> 手机号-TraceID映射解析增强：粘贴/拖入文本支持多种分隔符(Tab/空格/逗号/分号/斜杠等)；文本框内容变化后停止输入约300ms自动解析并增量填入表格(防抖避免逐字中间态,去重不重复,填完自动持久化)</li>\n<li><b>V2.5.12</b> 取消响应优化：生成汇总/写入详细过程/原始数据写入/批注/对比横表/着色等长循环全部加密取消检查点，点取消后通常1~3秒内停止且开始按钮立即恢复可用；手机号-TraceID映射持久化：5G速率tab与语音VQI tab各存一份独立JSON（关闭后重开自动恢复，增删改自动保存，映射文件损坏或缺失时空启动不报错）；开始处理前若未配置任何映射弹窗提醒是否继续；基准文件缺失改为中性提示且不影响主流程</li>\n<li><b>V2.5.11</b> 修复手机号映射合并问题：手动添加和拖入/粘贴的手机号映射现在都能参与计算；优化语音VQI批注性能：预构建过滤缓存避免重复遍历，大幅减少运营商/文件名批注卡顿；QCI 6和7改为独立可选不再互斥；处理开始时清空结果数据显示区；数据业务处理跳过非MMF文件名(如ALL_FTP_DETAIL_*.xlsx)；修复跳过非MMF文件时data_file_info索引错位bug并增加空数据保护</li>\n<li><b>V2.5.10</b> 新增手机号-TraceID映射功能：参数设置支持手动添加/拖入ExcelTXT/粘贴文本三种方式填写手机号与TraceID对应关系；自动从文件路径提取"用户跟踪ID=xxx"匹配手机号；按文件名/按运营商/汇总-数据和语音Sheet新增手机号列</li>\n</li>\n</ul><ul><li>批注优化：汇总-数据和语音、按文件名、按运营商表的批注显示详细计算过程(分子/分母/具体数值用/分隔)</li>\n<li>汇总-数据和语音表：文件名列改为逐行显示具体文件名(换行)，起止时间列改为每文件第一轮INVITE开始~最后一轮结束</li>\n<li>按运营商表：新增文件名和时间范围列(换行显示)</li>\n<li>按文件名表：新增微信/商店等业务指标列(当前留空)</li>\n<li>默认勾选改为默认不勾选(合并原始数据、添加批注)</li>\n<li>原始数据sheet取消20000行限制，全量写入</li>\n<li>修复跨运营商信令干扰：按运营商(电信/联通)分组独立搜索INVITE→180→BYE→200，避免联通的BYE匹配到电信的200</li>\n<li>修复搜索边界公式bug：非连续索引时最后一段信令不遗漏</li>\n<li>修复列不存在时KeyError崩溃：r180/bye列存在性检查</li>\n<li>修复get_loc对不连续索引抛出KeyError：try-except保护</li>\n<li>文件名改为显示相对路径(含文件夹信息)，如"语音类输入文件/电信主叫SIP/SaveMsg_xxx.csv"</li>\n</ul><p><b>2026-07-18</b></p><ul><li><b>V2.5.8</b> 性能优化：基准对齐使用二分查找(searchsorted)替代向量化abs计算，大幅提升大文件处理速度；新增服务小区平均RSRP统计(从RRC_MEAS_RPRT提取servRSRP，双值取最大)；修复RSRP正则支持单值无逗号格式；VQI提取使用.values批量优化；xlsxwriter引擎替代openpyxl(write_row批量写入)；批注添加功能(复选框控制)；自动打开结果文件</li>\n<li><b>V2.5.7</b> 修复按文件名统计表显示0的问题：1)为每条voice_record/vqi_record添加_source_file字段；2)使用voice_file_info字典记录每个文件的基本信息；3)按_source_file字段精确匹配统计每个文件的轮次/MOS指标</li>\n<li><b>V2.5.6</b> 新增"合并原始数据到结果文件"复选框(速率统计/VQI标签各一),默认勾选</li>\n<li>新增批量计算批注:汇总-数据和语音、按文件名、按运营商Sheet的数值单元格自动添加计算批注</li>\n<li>数据文件处理:无MessageType列的数据文件(mmf)不再跳过,读取并计算下行/上行速率指标</li>\n<li>VQI输出Sheet优化:新增"数据业务-原始文件"Sheet,按文件名/按运营商表含完整数据业务指标</li>\n<li>性能优化:文件类型检测只读前3行(nrows=3);缓存文件列表避免重复读取</li>\n<li><b>V2.5.5</b> 修复文件类型显示为unknown问题：_detect_file_type返回语音/数据而非unknown</li><li>按文件名表新增指标列：通话轮次、平均呼叫建立时长(s)、优质通话占比(MOS3.5)</li><li>按运营商表新增完整指标列：下行前10%峰值速率等12列（当前留空）</li><li>语音指标统计sheet中对空呼叫建立时长单元格加批注无RSP_180</li><li>平均呼叫建立时长只统计有值的轮次</li><li><b>V2.5.3</b> 呼叫建立时长单位从毫秒改为秒，精确到小数点后2位</li><li>新增汇总-数据和语音Sheet：标准行+电信+联通，包含语音指标和数据业务指标列</li><li>新增按文件名Sheet：每个文件一行，包含文件名、运营商、文件类型、行数、时间范围</li><li>新增按运营商Sheet：电信/联通各一行，包含语音文件数、通话轮次、平均呼叫建立时长、优质通话占比</li><li>process_vqi函数新增time_range参数，支持外部传入时间起止</li><li><b>V2.5.2</b> 全面重构语音VQI处理逻辑：自动分类语音/数据文件、识别INVITE(取最后一个连续)/180/BYE/200完整通话</li><li>新增MOS3.5优质通话占比计算：从Detail Info提取DlE2eVqi值，排除65535，大于3.5为优质</li><li>VQI输出3个Sheet：原始数据(加运营商/业务类型列)、语音指标统计、MOS3.5计算明细</li><li>支持.tmf文件(语音)和.mmf文件(数据)的自动分类</li><li><b>V2.5.1</b> 修复拖入文件不起作用的问题；支持csv文件类型</li><li><b>V2.5.0</b> 新增语音VQI标签页：实现完整的语音VQI处理功能</li><li>VQI功能：读取MMF文件，筛选MessageType非空行，识别SIP信令（INVITE/180/BYE/200），计算呼叫建立时长，判定完整通话</li><li>VQI输出：原始数据sheet（MessageType非空的所有行）+ 语音统计sheet（轮次/开始时间/结束时间/呼叫建立时长/是否完整通话）</li><li>FTP时长选项：参数设置新增FTP时长下拉选择（10秒/20秒/30秒），默认20秒，动态调整FTP候选检测阈值</li><li>Sheet重命名：详细过程→数据业务-详细过程，对比横表→数据业务-对比</li></ul><p><b>2026-07-17</b></p><ul><li><b>V2.4.5</b> 参数设置增加时间段限制：复选框启用后，可设置开始/结束时间(HH:mm:ss)</li><li>启用时只处理该时间段内的数据，未启用则全量处理</li><li><b>V2.4.4</b> 支持用户自定义基准文件</li></ul><p><b>V2.4.2</b></p><ul><li>工具改名：工信部数据统计工具 → 工信部数据、语音跟踪统计工具</li><li>主界面改为双标签(QTabWidget)：标签1=工信部数据统计工具，标签2=5G虚拟用户跟踪用户语音VQI工具(空壳待接入算法)</li><li>速率统计标签布局调整：左侧上设置+下运行(进度/日志)，右侧腾出内嵌显示结果3个sheet(详细过程/汇总/对比横表)</li><li>右侧内嵌QTableWidget：默认筛选A=1、复制xlsx颜色、批注(hover显示)</li><li>输入文件区：选择/清空按钮紧挨着，拖入框高度翻倍(300px)、边框加粗明显</li><li>关于对话框：更新记录按天汇总，同一天多次更新合并列出</li></ul><p><b>V2.4.1</b> <i>(2026-07-17)</i></p><ul><li>商店小文件识别修复：GAP_MERGE 2.0→3.0，防止商店大段被切碎误判FTP；8783段等真小文件恢复识别</li><li>去掉商店小文件上行占比约束（RLC/MAC占比均不可靠），回到max(dl)选取</li><li>对比横表改就近匹配：找离基准最近的同业务段，距离≤300秒，完整轮0缺失</li><li>支持嵌套zip：递归解压(zip里套zip)</li><li>取消按钮修复：写Excel各阶段加检查点，长循环每5000行检查</li><li>清理重复sort_values；fmt_time改预编译正则</li></ul><p><b>V2.4.0</b> <i>(2026-07-16)</i></p><ul><li>性能优化：预编译正则 + usecols 只读必要列，MMF加载提速</li><li>汇总批注：每格显示公式+实际数值（hover查看）</li><li>D/E列颜色标记：详细过程 D列浅蓝/E列浅绿</li></ul><p><b>V2.3.0</b> <i>(2026-07-16)</i></p><ul><li>性能优化：基准对齐/筛选列向量化，处理耗时 107秒→14秒(约7.7倍)</li><li>新增：取消按钮(阶段点中断)、拖拽文件入窗口、真实进度百分比</li><li>对比竖表美化：业务/轮次配色、合并同轮次、冻结首行、行高</li></ul><p><b>V2.2.0</b> <i>(2026-07-16)</i></p><ul><li>对比增加「偏差」行/列：(代码−基准)/基准 百分比，速率与时长各算</li><li>对比横表每轮由2列扩为3列(代码+基准+偏差)</li><li>新增「对比竖表」sheet(横表转置)</li><li>主界面改左右分栏(左35%设置/右65%结果)，全屏+自适应分辨率+高分屏字号放大+可拖动+打开输出按钮</li></ul><p><b>V2.1.0</b> <i>(2026-07-16)</i></p><ul><li>对比横表匹配改为「业务连续段」判定 + 按 FTP下载 切轮</li><li>不再用固定时间窗口(-5~+60s)和硬切6条分组</li><li>整轮6业务完整且全对上才编号，否则标「不计轮次」</li></ul><p><b>V2.0.9</b></p><ul><li>修复对比横表代码业务结束时间缺失</li><li>新增第一个 FTP对 之前不完整轮次的识别</li><li>版本号统一更新(文件名/关于/需求文档)</li></ul><p><b>V2.0.1</b></p><ul><li>QCI 支持 5/6/7 多选；主界面增加「关于」</li></ul><p><b>V2.0.0</b></p><ul><li>从零重写，三页向导改单页</li><li>自动识别 FTP/商店/微信 6 类业务</li><li>从原始 MMF 文件直接生成详细过程/对比/汇总 3 个 Sheet</li><li>削峰阈值可调，轮次完整性检查</li></ul>');v.addWidget(te);bb=QPushButton('关闭');bb.clicked.connect(d.accept);v.addWidget(bb);d.exec_()
if __name__=='__main__':
	if GUI_OK and len(sys.argv)==1:
		if hasattr(Qt,'AA_EnableHighDpiScaling'):QApplication.setAttribute(Qt.AA_EnableHighDpiScaling,_B)
		if hasattr(Qt,'AA_UseHighDpiPixmaps'):QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps,_B)
		app=QApplication(sys.argv);app.setStyle('Fusion');_sys_font=QFont()
		if sys.platform=='darwin':_sys_font=QFont(_BD,12)
		elif sys.platform=='win32':_sys_font=QFont(_BE,10)
		app.setFont(_sys_font);w=MainWindow();w.show();sys.exit(app.exec())
	else:out=process(['联通/mmf20260703115328-电信.xlsx','联通/mmf20260703115334-电信.xlsx'],qci_list=[5,6,7],dl_clip=1000,ul_clip=200);print(f"输出: {out}")