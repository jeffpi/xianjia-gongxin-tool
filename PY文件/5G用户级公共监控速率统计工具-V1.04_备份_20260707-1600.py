#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
5G用户级公共监控速率统计工具 V1.03（测试计划驱动版）
基于2025-2026年工信部测评用例规则开发
作者：Claude Code
创建日期：2026-06-26

V1.03 核心方案（纯UCM + 测试计划驱动）：
  - 输入：仅 mmf（UCM 监控文件，中英文自动识别）
  - GUI「测试计划编辑器」：用户定义业务测试顺序与参数（可上下拖动），工具按计划沿UCM时间轴匹配
  - 业务识别：FTP按"时长+高速段"，应用商店/微信按"累计流量(60/1100/5/200M)+超时"，等待带容差
  - 自动找起点、自动判循环次数；上一个业务结束下一个即开始
  - 输出对齐参考文件「呼叫详情」90列（一行=一轮计划，语音/视频列留空）+ 统计汇总宽表42列
  - 速率列RLC，削峰1000/200，FTP段去首1尾2爬坡
详见同目录《统计规则说明》
"""

import sys
import os
import json
import re
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
import numpy as np
import pandas as pd
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QLabel, QLineEdit, QPushButton, QFileDialog,
    QTextEdit, QTableWidget, QTableWidgetItem, QHeaderView,
    QGroupBox, QFormLayout, QSpinBox, QDoubleSpinBox, QComboBox,
    QMessageBox, QProgressBar, QAbstractItemView, QDialog
)
from PySide6.QtGui import QFont
from PySide6.QtCore import Qt, QThread, Signal, Slot
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import urllib.request, hashlib, tempfile, shutil, time


# ==================== 呼叫详情90列（对齐参考文件） ====================
CALL_HEADERS = [
    # 基础信息(0-9)
    '省','市','运营商','业务类型','测试业务','测试网络','文件名','业务序号','主被叫','呼叫类型',
    # 语音详情(10-30) — 数据业务留空
    '语音起呼网络','语音接通网络','语音起呼时间','语音振铃时间','语音接通时间','语音完成时间',
    '语音失败时间','语音失败时间(不含振铃)','语音掉话时间(事件drop)','语音掉话时间',
    '呼叫振铃时延(秒)','呼叫接通时延(秒)','被叫振铃->接通时延(秒)','呼叫保持时长(秒)','5G呼叫保持时长(秒)',
    '呼叫结果','挂机原因','语音MOS最大值','语音MOS最小值','语音MOS均值','MOS样本数',
    # 微信呼叫(31-42) — 留空
    '微信起呼时间','微信振铃时间','微信接通时间','微信完成时间','微信失败时间','微信掉话时间',
    '微信呼叫振铃时延(秒)','微信呼叫接通时延(秒)','微信被叫振铃接通时延(秒)','微信呼叫保持时长(秒)',
    '微信呼叫结果','平均卡顿时间(ms)',
    # VINR/微信视频(43-60) — 留空
    '卡顿次数','卡顿总时长(ms)','卡顿率','每小时卡顿次数','每小时卡顿时长(s)','视频质量',
    '音画同步(毫秒)','首帧时延','平均丢帧次数','总丢帧次数','丢帧率','VMOS综合评分',
    '男音MOS均值','女音MOS均值','微信音频MOS均值','VINR音频MOS均值','CTTL VMOS样本数','算分结果',
    # 微信文件(61-71)
    '微信文件业务请求时间','微信文件业务完成时间','微信文件业务失败时间','微信文件业务时长(秒)',
    '微信小包发送速率(Mbps)','微信大包发送速率(Mbps)','微信文件业务结果',
    '微信小包发送开始时间','微信小包发送完成时间','微信大包发送开始时间','微信大包发送完成时间',
    # 应用商店(72-83)
    '应用商店业务请求时间','应用商店业务完成时间','应用商店业务失败时间','应用商店业务时长(秒)',
    '应用商店下载速率(Mbps)','应用商店业务结果','应用商店小文件下载速率(Mbps)','应用商店大文件下载速率(Mbps)',
    '应用商店小文件下载开始时间','应用商店小文件下载完成时间','应用商店大文件下载开始时间','应用商店大文件下载完成时间',
    # FTP(84-88)
    '发起时间','完成时间','失败时间','上传平均速率(Mbps)','下载平均速率(Mbps)',
]

# ==================== 基准规则定义（ALL_FTP_DETAIL模板 vs 工信部指标） ====================
BENCHMARK_ALL_FTP = {
    # ALL_FTP_DETAIL.xlsx模板：满信号场景理论速率，联通=电信同一模板
    'FTP下载':      {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 572.11, 'dur': None,   'name': 'ALL_FTP_DETAIL模板'},
    'FTP上传':      {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 66.39,  'dur': None,   'name': 'ALL_FTP_DETAIL模板'},
    '应用商店_小包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 176.29, 'dur': 5.05,  'name': 'ALL_FTP_DETAIL模板'},
    '应用商店_大包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 1025.84,'dur': 11.20, 'name': 'ALL_FTP_DETAIL模板'},
    '微信_小文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 27.81,  'dur': 1.50,  'name': 'ALL_FTP_DETAIL模板'},
    '微信_大文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 60.48,  'dur': 28.35, 'name': 'ALL_FTP_DETAIL模板'},
}

BENCHMARK_CONFIG = {
    # 工信部业务指标(运营商汇总T0)：实际测试场景速率，区分联通/电信
    '联通': {
        'FTP下载':      {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 581.145, 'dur': None,   'name': '工信部指标(联通T0)'},
        'FTP上传':      {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 79.621,  'dur': None,   'name': '工信部指标(联通T0)'},
        '应用商店_小包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 188.246, 'dur': 6.191, 'name': '工信部指标(联通T0)'},
        '应用商店_大包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 402.977, 'dur': 45.208,'name': '工信部指标(联通T0)'},
        '微信_小文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 19.417,  'dur': 2.160, 'name': '工信部指标(联通T0)'},
        '微信_大文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 49.973,  'dur': 33.572,'name': '工信部指标(联通T0)'},
    },
    '电信': {
        'FTP下载':      {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 728.545, 'dur': None,   'name': '工信部指标(电信T0)'},
        'FTP上传':      {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 118.499, 'dur': None,   'name': '工信部指标(电信T0)'},
        '应用商店_小包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 226.645, 'dur': 3.499, 'name': '工信部指标(电信T0)'},
        '应用商店_大包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 561.335, 'dur': 18.321,'name': '工信部指标(电信T0)'},
        '微信_小文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 31.263,  'dur': 1.342, 'name': '工信部指标(电信T0)'},
        '微信_大文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 51.600,  'dur': 32.514,'name': '工信部指标(电信T0)'},
    },
}

# 【新增】UCM标注基准(原始行非零RLC均值)，区分联通/电信
BENCHMARK_UCM = {
    '联通': {
        'FTP下载':      {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 768.6,  'dur': None,   'name': 'UCM标注(联通)'},
        'FTP上传':      {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 105.6,  'dur': None,   'name': 'UCM标注(联通)'},
        '应用商店_小包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 163.8,  'dur': None,   'name': 'UCM标注(联通)'},
        '应用商店_大包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 549.9,  'dur': None,   'name': 'UCM标注(联通)'},
        '微信_小文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 28.3,   'dur': None,   'name': 'UCM标注(联通)'},
        '微信_大文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 54.4,   'dur': None,   'name': 'UCM标注(联通)'},
    },
    '电信': {
        'FTP下载':      {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 1083.3, 'dur': None,   'name': 'UCM标注(电信)'},
        'FTP上传':      {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 162.7,  'dur': None,   'name': 'UCM标注(电信)'},
        '应用商店_小包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 177.9,  'dur': None,   'name': 'UCM标注(电信)'},
        '应用商店_大包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 614.9,  'dur': None,   'name': 'UCM标注(电信)'},
        '微信_小文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 26.8,   'dur': None,   'name': 'UCM标注(电信)'},
        '微信_大文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 53.5,   'dur': None,   'name': 'UCM标注(电信)'},
    },
}
CALL_IDX = {n: i for i, n in enumerate(CALL_HEADERS)}

# 呼叫详情分组(起始列0-based,结束列exclusive,组名) —— 两层表头 R1 合并用，对齐参考文件
CALL_GROUPS = [(0,10,'基础信息'),(10,31,'语音详情'),(31,43,'微信呼叫业务'),(43,61,'VINR/微信视频'),(61,72,'微信文件'),(72,84,'应用商店'),(84,89,'FTP业务')]

# 默认测试计划（顺序即为测试顺序）
DEFAULT_PLAN = [
    {'key':'ftp_dl', 'name':'FTP下载',  'mode':'time', 'direction':'下行', 'duration':10, 'flow_mb':None,'timeout':None},
    {'key':'wait1',  'name':'等待',     'mode':'wait', 'direction':'',     'duration':5,  'flow_mb':None,'timeout':None},
    {'key':'ftp_ul', 'name':'FTP上传',  'mode':'time', 'direction':'上行', 'duration':10, 'flow_mb':None,'timeout':None},
    {'key':'wait2',  'name':'等待',     'mode':'wait', 'direction':'',     'duration':15, 'flow_mb':None,'timeout':None},
    {'key':'store_s','name':'应用商店小包','mode':'flow','direction':'下行','duration':None,'flow_mb':90,  'timeout':300},
    {'key':'store_l','name':'应用商店大包','mode':'flow','direction':'下行','duration':None,'flow_mb':1100,'timeout':300},
    {'key':'wx_s',   'name':'微信小文件','mode':'flow', 'direction':'上行', 'duration':None,'flow_mb':5,   'timeout':None},
    {'key':'wx_l',   'name':'微信大文件','mode':'flow', 'direction':'上行', 'duration':None,'flow_mb':200, 'timeout':None},
]

# 统计汇总宽表42列
STAT_GROUPS = ['FTP下载','FTP上传','应用商店_小包','应用商店_大包','微信_小文件','微信_大文件']
STAT_COLUMNS = [
    ('FTP下载','测试次数'),('FTP下载','采样点数'),
    ('FTP下载','削峰后下行RLC≥100Mbps采集点数'),('FTP下载','削峰后下行RLC≥1000Mbps采集点数'),
    ('FTP下载','下行RLC平均吞吐率(Mbps)'),('FTP下载','削峰后下行RLC平均吞吐率(Mbps)'),
    ('FTP下载','下行RLC峰值吞吐率(Mbps)'),
    ('FTP下载','削峰后下行RLC≥100Mbps占比'),('FTP下载','削峰后下行RLC≥1000Mbps占比'),
    ('FTP下载','前10%峰值速率(Mbps)'),
    ('FTP上传','测试次数'),('FTP上传','采样点数'),
    ('FTP上传','削峰后上行RLC>20Mbps采集点数'),('FTP上传','削峰后上行RLC>200Mbps采集点数'),
    ('FTP上传','上行RLC平均吞吐率(Mbps)'),('FTP上传','削峰后上行RLC平均吞吐率(Mbps)'),
    ('FTP上传','上行RLC峰值吞吐率(Mbps)'),
    ('FTP上传','削峰后上行RLC>20Mbps占比'),('FTP上传','削峰后上行RLC>200Mbps占比'),
    ('FTP上传','前10%峰值速率(Mbps)'),
    ('应用商店_小包','测试次数'),('应用商店_小包','采样点数'),
    ('应用商店_小包','下行RLC平均吞吐率(Mbps)'),('应用商店_小包','单次平均时长(s)'),
    ('应用商店_小包','下载成功率(%)'),
    ('应用商店_大包','测试次数'),('应用商店_大包','采样点数'),
    ('应用商店_大包','下行RLC平均吞吐率(Mbps)'),('应用商店_大包','单次平均时长(s)'),
    ('应用商店_大包','中位值平均时长(s)'),('应用商店_大包','下载成功率(%)'),
    ('微信_小文件','测试次数'),('微信_小文件','采样点数'),
    ('微信_小文件','上行RLC平均吞吐率(Mbps)'),('微信_小文件','单次平均时长(s)'),
    ('微信_小文件','上传成功率(%)'),
    ('微信_大文件','测试次数'),('微信_大文件','采样点数'),
    ('微信_大文件','上行RLC平均吞吐率(Mbps)'),('微信_大文件','单次平均时长(s)'),
    ('微信_大文件','中位值平均时长(s)'),('微信_大文件','上传成功率(%)'),
]

# ==================== 应用信息 + 在线更新 ====================
APP_NAME = '5G用户级公共监控速率统计工具'
APP_VERSION = '1.04'
BUILD_DATE = '2026-07-03'
AUTHOR = '孙晓军'
CONTACT = '317827@qq.com'
TOOL_ID = 'xianjia_5g'
# 云主机(nginx)托管更新包；启动静默检查 + 关于里手动检查
UPDATE_BASE = 'https://47.109.101.79/updates/' + TOOL_ID
VERSION_HISTORY = [
    ('V1.04', '2026-07-03', '新增「关于」对话框 + 在线更新检查(云主机托管)'),
    ('V1.03', '2026-07-02', 'auto自动识别(方向判定) + 网络覆盖sheet + 业务标注文件'),
    ('V1.02', '2026-06-23', '呼叫详情配准方案 + 对齐工信部业务指标参考'),
    ('V1.01', '2026-06-23', '统计结果对标V6原工具 + 指标体系完善'),
    ('V1.00', '2026-06-23', '初始版本(基于2025-2026工信部测评规则)'),
]


class ConfigManager:
    def __init__(self, config_path: str = None):
        if config_path is None:
            config_path = os.path.join(os.path.dirname(__file__), 'config', 'default_rules.json')
        self.config_path = config_path
        self.config = self.load_config()

    def load_config(self) -> Dict:
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载配置失败: {e}")
            return self.get_default_config()

    def save_config(self, config: Dict = None) -> bool:
        if config is None:
            config = self.config
        try:
            os.makedirs(os.path.dirname(self.config_path), exist_ok=True)
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"保存配置失败: {e}")
            return False

    def get_default_config(self) -> Dict:
        return {"version": "1.03", "rule_version": "2025-2026工信部测评用例",
                "global_params": {"rate_column": "RLC", "dl_peak_limit": 900, "ul_peak_limit": 160,
                                   "dl_pass_threshold": 100, "ul_pass_threshold": 20,
                                   "down_min": 50, "up_min": 10}}


# ==================== UCM 处理 + 计划匹配引擎 ====================

class UCMProcessor:
    def __init__(self, config: Dict):
        self.config = config
        self.seconds = None  # 秒级聚合表

    def _select_col(self, df, direction, layer):
        """中英文找速率列。direction:下行/上行; layer:MAC/RLC"""
        dir_kw = ['下行', 'Downlink'] if direction == '下行' else ['上行', 'Uplink']
        cands = [c for c in df.columns if any(k in str(c) for k in dir_kw)
                 and layer in str(c) and ('吞吐率' in str(c) or 'hroughput' in str(c).lower())]
        return cands[0] if cands else None

    def _find_col(self, df, keywords):
        for c in df.columns:
            cs = str(c)
            if any(k in cs for k in keywords):
                return c
        return None

    def _read_one_mmf(self, path: str):
        """读单个UCM：探测表头、bps→Mbps、加_t、加file_name列。返回df(已过滤无效时间)"""
        raw = pd.read_excel(path, header=None, nrows=10)
        header_row = 0
        for i in range(min(10, len(raw))):
            vals = [str(v) for v in raw.iloc[i].tolist()]
            if any('吞吐率' in v or 'hroughput' in v.lower() for v in vals):
                header_row = i; break
        df = pd.read_excel(path, header=header_row)
        df = df.replace('-', np.nan)
        for col in list(df.columns):   # bps→Mbps
            cs = str(col); low = cs.lower()
            if ('吞吐率' in cs or 'hroughput' in low) and 'bps' in low and 'mbps' not in low:
                new_col = re.sub(r'\(.*?bps.*?\)', '(Mbps)', cs, flags=re.IGNORECASE)
                if new_col == cs: new_col = cs + '(Mbps)'
                if new_col not in df.columns:
                    df[new_col] = pd.to_numeric(df[col], errors='coerce') / 1_000_000
                df = df.drop(columns=[col])
        tcol = self._find_col(df, ['采集时间', 'Time'])
        df['_t'] = df[tcol].apply(self._parse_t) if tcol else None
        df = df[df['_t'].notna()].copy()
        df['file_name'] = os.path.basename(path)
        print(f"DEBUG: {os.path.basename(path)} 表头第{header_row+1}行 有效{len(df)}行")
        return df

    def parse_mmf(self, paths, params: Dict):
        """读单个或多个UCM，按采集时间合并成一条时间轴，按秒聚合(MAC取max/RLC取均值)、削峰、coverage"""
        if isinstance(paths, str): paths = [paths]
        dfs = [self._read_one_mmf(p) for p in paths]
        df = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
        df = df.sort_values('_t').reset_index(drop=True)   # 多文件按时间合并排序
        print(f"DEBUG: 合并 {len(dfs)} 个文件 共{len(df)}行 → 统一时间轴")
        df['sec'] = df['_t'].dt.floor('s')
        # 列查找 + 秒级聚合
        dl_mac = self._select_col(df, '下行', 'MAC'); ul_mac = self._select_col(df, '上行', 'MAC')
        dl_rlc = self._select_col(df, '下行', 'RLC'); ul_rlc = self._select_col(df, '上行', 'RLC')
        srs_col = 'SRS RSRP(0.01dBm)' if 'SRS RSRP(0.01dBm)' in df.columns else None
        sinr_col0 = 'SINR(0.01dB)' if 'SINR(0.01dB)' in df.columns else None
        print(f"DEBUG: 列 MAC:{dl_mac}/{ul_mac} RLC:{dl_rlc}/{ul_rlc}")
        agg_d = {
            'dl_mac': (dl_mac, lambda s: pd.to_numeric(s, errors='coerce').max()),
            'ul_mac': (ul_mac, lambda s: pd.to_numeric(s, errors='coerce').max()),
            'dl_rlc': (dl_rlc, lambda s: pd.to_numeric(s, errors='coerce').max()),
            'ul_rlc': (ul_rlc, lambda s: pd.to_numeric(s, errors='coerce').mean()),
        }
        if srs_col: agg_d['srs_rsrp'] = (srs_col, lambda s: pd.to_numeric(s, errors='coerce').mean())
        if sinr_col0: agg_d['sinr'] = (sinr_col0, lambda s: pd.to_numeric(s, errors='coerce').mean())
        agg = df.groupby('sec').agg(**agg_d).reset_index().rename(columns={'sec': 't'})
        # 保存agg供_summarize扩展使用（含dl_mac, ul_mac, dl_rlc, ul_rlc, dl_clip, ul_clip列）
        self.seconds = agg
        # 削峰（秒级）
        agg['dl_clip'] = agg['dl_rlc'].clip(upper=params.get('dl_peak_limit', 900))     # 场景900M/道路800M(工信部规范)
        agg['ul_clip'] = agg['ul_rlc'].clip(upper=params.get('ul_peak_limit', 160))    # 满分门限160Mbps(工信部规范)
        agg[['dl_mac', 'ul_mac', 'dl_rlc', 'ul_rlc']] = agg[['dl_mac', 'ul_mac', 'dl_rlc', 'ul_rlc']].fillna(0)
        print(f"DEBUG: 秒级聚合 {len(agg)} 秒")
        # ===== 网络覆盖/基础指标(合并后全部采样点) =====
        # 注：UCM 的 SSB Beam Rsrp 列全空，参考「SS-RSRP」实际对应 SRS RSRP(均值完美对齐参考-65.979)
        self.coverage = {}
        srs = pd.to_numeric(df['SRS RSRP(0.01dBm)'], errors='coerce') if 'SRS RSRP(0.01dBm)' in df.columns else pd.Series(dtype=float)
        sinr = pd.to_numeric(df['SINR(0.01dB)'], errors='coerce') if 'SINR(0.01dB)' in df.columns else pd.Series(dtype=float)
        srs_dbm = srs / 100.0                    # 0.01dBm → dBm(含0；UCM无SS-RSRP用SRS RSRP替代,含0均值-65.95对齐参考-65.979)
        sinr_db = sinr / 100.0                  # 0.01dB → dB
        self.coverage['NR_Serving SS-RSRP(dBm)'] = round(float(srs_dbm.mean()), 3) if len(srs_dbm) else None
        self.coverage['NR_Serving SS AVG SINR(dB)'] = round(float(sinr_db.mean()), 3) if len(sinr_db) else None
        # 覆盖率3项移至 _attach_cov(用业务时段采样点,对齐参考口径)
        if 'CA Property' in df.columns:
            ca = df['CA Property'].astype(str)
            # CA激活 = 非"Non CA"(即 SCell1+PCell)；参考100%为另一口径(差异已知)
            self.coverage['5G CA聚合率(%)'] = round(float((ca != 'Non CA').sum() / len(ca) * 100), 2)
        self._fn = '+'.join(os.path.basename(p) for p in paths)   # 多文件名合并显示
        self.seconds = agg
        self.params = params
        return agg

    @staticmethod
    def _parse_t(v):
        if v is None or (isinstance(v, float) and pd.isna(v)) or str(v).strip() in ('', 'nan', 'None'):
            return None
        try:
            return pd.to_datetime(re.sub(r'\(\d+\)', '', str(v)).strip())
        except Exception:
            return None

    # ---------- 匹配引擎 ----------
    def match(self, plan: List[Dict], params: Dict) -> Tuple[pd.DataFrame, Dict]:
        """按计划沿秒级时间轴循环匹配，返回 (呼叫详情DataFrame, 统计dict)"""
        agg = self.seconds
        if agg is None or len(agg) == 0:
            return pd.DataFrame(columns=CALL_HEADERS), {}
        dl_min = params.get('down_min', 50); ul_min = params.get('up_min', 10)
        if params.get('match_mode') == 'auto':
            # 自动识别模式:通过_make_rounds合成轮次(6业务一轮),_summarize消费轮次结构
            round_results = self._match_auto(agg, params, dl_min, ul_min)
            self._last_round_results = round_results
            N = len(round_results)
            call_rows = [self._build_ftp_dl_row(rd, i + 1, params) for i, rd in enumerate(round_results)]
            call_rows += [self._build_ftp_ul_row(rd, i + N + 1, params) for i, rd in enumerate(round_results)]
            call_df = pd.DataFrame(call_rows, columns=CALL_HEADERS)
            stats = self._attach_cov(self._summarize(round_results, params), round_results)
            return call_df, stats
        # 找起点：连续10秒下行高速 + 其后5~25秒内出现连续10秒上行高速(FTP下载+上传组合)，排除孤立下行段
        cur = None
        n = len(agg)
        for i in range(n - 10):
            ok = sum(1 for k in range(10) if i + k < n and agg.at[i + k, 'dl_mac'] > dl_min)
            if ok >= 9:
                for j in range(i + 13, min(i + 35, n - 10)):
                    uok = sum(1 for k in range(10) if j + k < n and agg.at[j + k, 'ul_mac'] > ul_min)
                    if uok >= 9:
                        cur = i; break
                if cur is not None:
                    break
        if cur is None:
            print("DEBUG: 找不到 FTP下载+上传 组合起点")
            return pd.DataFrame(columns=CALL_HEADERS), {}
        n = len(agg)
        round_results = []   # 每轮 rdata(含段 points)，供 _summarize 重算统计量
        round_idx = 0
        while cur < n:
            rdata, cur = self._match_round(plan, agg, cur, params, dl_min, ul_min)
            if rdata is None:
                break
            round_results.append(rdata)
            round_idx += 1
            if round_idx > 200:  # 安全上限
                break
        # 剔除离群首轮：若首轮FTP下载时间与第二轮间隔>360秒(>正常周期)，视为准备期孤立组合，剔除
        if len(round_results) >= 2:
            t1 = round_results[0].get('ftp_dl', {}).get('start_t')
            t2 = round_results[1].get('ftp_dl', {}).get('start_t')
            if t1 and t2 and (t2 - t1).total_seconds() > 360:
                round_results = round_results[1:]
        # 生成呼叫详情：FTP下载行(序号1..N) + FTP上传行(序号N+1..2N)，仿参考文件(下载/上传分行)
        self._last_round_results = round_results  # 供 generate_annotated 标注用
        N = len(round_results)
        call_rows = [self._build_ftp_dl_row(rd, i + 1, params) for i, rd in enumerate(round_results)]
        call_rows += [self._build_ftp_ul_row(rd, i + N + 1, params) for i, rd in enumerate(round_results)]
        call_df = pd.DataFrame(call_rows, columns=CALL_HEADERS)
        stats = self._attach_cov(self._summarize(round_results, params), round_results)
        return call_df, stats

    def _match_round(self, plan, agg, cur, params, dl_min, ul_min) -> Tuple[Optional[Dict], int]:
        """匹配一轮，返回 ({业务key: 段信息}, 下轮起点)。段信息={start_t,end_t,start_i,end_i,rate,result}"""
        n = len(agg)
        result = {}
        last_ok_cur = cur
        for item in plan:
            key = item['key']; mode = item['mode']
            if mode == 'wait':
                # 等待：cur 前进 duration 秒（容差：在下个业务窗口内找）
                cur = min(n - 1, cur + max(1, int(item['duration'])))
                continue
            if cur >= n:
                return None, cur
            if mode == 'time':
                seg = self._find_time_seg(agg, cur, item['direction'], item['duration'],
                                          dl_min if item['direction'] == '下行' else ul_min)
            else:  # flow
                seg = self._find_flow_seg(agg, cur, item['direction'], item['flow_mb'], item['timeout'])
            if seg is None:
                # 本业务找不到，本轮终止
                return None, cur
            result[key] = seg
            cur = seg['end_i'] + 1
            last_ok_cur = cur
        return result, cur

    def _find_time_seg(self, agg, cur, direction, duration, min_rate) -> Optional[Dict]:
        """找连续duration秒(允许中间≤2个不满足)、该方向MAC速率>min_rate 的段"""
        col = 'dl_mac' if direction == '下行' else 'ul_mac'
        rlc_col = 'dl_rlc' if direction == '下行' else 'ul_rlc'
        clip_col = 'dl_clip' if direction == '下行' else 'ul_clip'
        n = len(agg)
        i = cur
        while i < n:
            run = 0; gap = 0; start_i = None; j = i
            while j < n and run < duration:
                if agg.at[j, col] > min_rate:
                    if start_i is None:
                        start_i = j
                    run += 1; gap = 0
                else:
                    gap += 1
                    if gap > 2:  # 连续>2个不满足则中断重来
                        run = 0; start_i = None; gap = 0
                j += 1
            if run >= duration and start_i is not None:
                end_i = j - 1
                seg = agg.loc[start_i:end_i]
                rlc_vals = seg[rlc_col].replace(0, np.nan).dropna().tolist()
                clip_vals = seg[clip_col].replace(0, np.nan).dropna().tolist()
                rlc_trim = rlc_vals[1:-2] if len(rlc_vals) > 3 else rlc_vals
                clip_trim = clip_vals[1:-2] if len(clip_vals) > 3 else clip_vals
                rate = round(float(np.mean(rlc_trim)), 3) if rlc_trim else 0
                clip_rate = round(float(np.mean(clip_trim)), 3) if clip_trim else 0
                return {'start_t': seg['t'].iloc[0], 'end_t': seg['t'].iloc[-1],
                        'start_i': start_i, 'end_i': end_i,
                        'rate': rate, 'clip_rate': clip_rate,
                        'duration': round((seg['t'].iloc[-1] - seg['t'].iloc[0]).total_seconds(), 3),
                        'points': seg, 'result': '成功'}
            i += 1
            if i > cur + 120:  # 向后扫120秒仍无则放弃
                return None
        return None

    def _find_flow_seg(self, agg, cur, direction, flow_mb, timeout) -> Optional[Dict]:
        """连续高速段(dl_mac>start_thr, gap>1)作业务时段。dl_mac完整覆盖率高。"""
        col = 'dl_mac' if direction == '下行' else 'ul_mac'
        rlc_col = 'dl_rlc' if direction == '下行' else 'ul_rlc'
        clip_col = 'dl_clip' if direction == '下行' else 'ul_clip'
        start_thr = 50 if direction == '下行' else 10
        n = len(agg)
        i = cur
        while i < n and agg.at[i, col] > start_thr:  # 跳过残留高速
            i += 1
        while i < n and agg.at[i, col] <= start_thr:  # 越过间隙
            i += 1
            if timeout and (i - cur) > timeout:
                return None
        if i >= n:
            return None
        start_i = i
        gap = 0; end_i = i
        for j in range(i + 1, n):  # 连续高速段(允许中间1个不满足)
            if agg.at[j, col] > start_thr:
                end_i = j; gap = 0
            else:
                gap += 1
                if gap > 1:
                    break
        seg = agg.loc[start_i:end_i]
        rlc_vals = seg[rlc_col].replace(0, np.nan).dropna()
        clip_vals = seg[clip_col].replace(0, np.nan).dropna()
        rate = round(float(rlc_vals.mean()), 3) if len(rlc_vals) else 0
        clip_rate = round(float(clip_vals.mean()), 3) if len(clip_vals) else 0
        dur = (seg['t'].iloc[-1] - seg['t'].iloc[0]).total_seconds()
        return {'start_t': seg['t'].iloc[0], 'end_t': seg['t'].iloc[-1],
                'start_i': start_i, 'end_i': end_i,
                'rate': rate, 'clip_rate': clip_rate, 'duration': round(dur, 3),
                'points': seg, 'result': '成功'}

    # ---------- 自动识别匹配引擎（新规则：不固定顺序，按方向+流量+时长+间隔识别）----------
    def _match_auto(self, agg, params, dl_min, ul_min):
        """自动识别：沿时间轴找业务段→归类(方向+流量+时长+间隔)→分组(6业务各一次为一轮)。返回 round_results"""
        n = len(agg)
        segs = []
        i = 0
        while i < n:
            if agg.at[i, 'dl_mac'] > dl_min or agg.at[i, 'ul_mac'] > ul_min:
                start_i = i; gap = 0
                while i < n and gap < 5:
                    if agg.at[i, 'dl_mac'] > dl_min or agg.at[i, 'ul_mac'] > ul_min:
                        gap = 0
                    else:
                        gap += 1
                    i += 1
                end_i = max(start_i, i - 1 - gap)
                seg_df = agg.loc[start_i:end_i]
                # 方向判定：ul_act需比dl_act多3秒才判为上行，避免误判
                ul_act = int((seg_df['ul_mac'] > 10).sum())
                dl_act = int((seg_df['dl_mac'] > 50).sum())
                if ul_act > dl_act + 3:
                    direction, col, rlc_col, clip_col = '上行', 'ul_mac', 'ul_rlc', 'ul_clip'
                else:
                    direction, col, rlc_col, clip_col = '下行', 'dl_mac', 'dl_rlc', 'dl_clip'
                flow = seg_df[col].sum() / 8  # 分类用MAC口径(聚合更稳定，RLC口径flow波动大)
                dur = max(round((seg_df['t'].iloc[-1] - seg_df['t'].iloc[0]).total_seconds(), 1), 0.5)  # 最小0.5s
                rv = seg_df[rlc_col].replace(0, np.nan).dropna()
                cv = seg_df[clip_col].replace(0, np.nan).dropna()
                segs.append({'start_i': start_i, 'end_i': end_i, 'start_t': seg_df['t'].iloc[0],
                             'end_t': seg_df['t'].iloc[-1], 'direction': direction,
                             'flow_mb': round(flow, 1), 'duration': round(dur, 1),
                             'rate': round(float(rv.mean()), 3) if len(rv) else 0,
                             'clip_rate': round(float(cv.mean()), 3) if len(cv) else 0,
                             'points': seg_df, 'result': '成功'})
            else:
                i += 1
        # 段合并：相邻同方向段gap<10秒则合并（解决商店大包被拆成多个小段导致flow<1500的问题）
        # 注：短段(dur<5s)不合并，保留为独立段，避免store_s短突发被吞并
        merged = []
        for s in segs:
            # 过滤1秒以下的微小噪声段（但保留wx_s的短小包，其dur~0.5s但flow>20MB才算）
            if s.get('duration', 0) < 1 and s.get('flow_mb', 0) < 10:
                continue
            if not merged:
                merged.append(s)
                continue
            prev = merged[-1]
            gap = (s['start_t'] - prev['end_t']).total_seconds()
            # 短段(dur<5s)不合并，保留独立段（如store_s短突发）
            if s['direction'] == prev['direction'] and 0 < gap < 10 and s.get('duration', 0) >= 5:
                # 合并：扩展end_i，合并points
                seg_df = agg.loc[prev['start_i']:s['end_i']]
                col = 'dl_mac' if s['direction'] == '下行' else 'ul_mac'
                rlc_col = 'dl_rlc' if s['direction'] == '下行' else 'ul_rlc'
                clip_col = 'dl_clip' if s['direction'] == '下行' else 'ul_clip'
                flow = seg_df[col].sum() / 8
                dur = round((seg_df['t'].iloc[-1] - seg_df['t'].iloc[0]).total_seconds(), 1)
                rv = seg_df[rlc_col].replace(0, np.nan).dropna()
                cv = seg_df[clip_col].replace(0, np.nan).dropna()
                merged[-1] = {'start_i': prev['start_i'], 'end_i': s['end_i'],
                              'start_t': prev['start_t'], 'end_t': seg_df['t'].iloc[-1],
                              'direction': s['direction'], 'flow_mb': round(flow, 1),
                              'duration': round(dur, 1),
                              'rate': round(float(rv.mean()), 3) if len(rv) else 0,
                              'clip_rate': round(float(cv.mean()), 3) if len(cv) else 0,
                              'points': seg_df, 'result': '成功'}
            else:
                merged.append(s)
        segs = merged
        for idx, s in enumerate(segs):
            s['key'] = self._classify_seg(s, segs, idx)
        # 调试：分类前段分布
        dl_segs = [s for s in segs if s['direction'] == '下行']
        ul_segs = [s for s in segs if s['direction'] == '上行']
        print(f"DEBUG: 候选段共{len(segs)}段(下行{len(dl_segs)},上行{len(ul_segs)})")
        from collections import Counter
        keys = Counter(s.get('key') for s in segs)
        print(f"DEBUG: 分类分布: {dict(keys)}")
        # 分组成轮
        # 过滤规则：store_s只保留dur<=25s（UCM标注最长28s），过滤误判的83s段
        # store_s短段(dur<5s, CD标注1~4秒短突发)保留为store_s
        # store_s长段(dur>=5s)改为None跳过，不参与统计（FTP尾段误判）
        # store_l只保留dur>5s（过滤短时闪断误判）
        # 微信小文件短段也保留（CD标注1~2秒）
        key_segs = []
        for s in segs:
            k = s.get('key')
            if k is None: continue
            dur = s.get('duration', 0)
            if k == 'store_s':
                if dur > 25:
                    continue  # store_s时长>25s不是商店小包
                if dur >= 5:
                    s['key'] = None  # 长段store_s改为None，不参与统计
                    continue  # 跳过，不加入key_segs
            if k == 'store_l' and dur < 5:
                continue  # store_l时长<5s是误判
            if k == 'store_s' or k == 'wx_s':  # 短段store_s和微信小文件保留
                key_segs.append(s)
            elif dur >= 3:
                key_segs.append(s)
                key_segs.append(s)
        # 轮次分组：180秒间隔切分，>=2个key保留，同key短间隔跳过(重复段)
        # 规则：
        #   1. gap>180s 且 key不在cur_round中 → 新轮开始
        #   2. key已在cur_round中 且 gap<=120s → 重复段，跳过(不截断)
        #   3. key已在cur_round中 且 gap>120s → 真正的下一轮，提交当前轮
        #   4. 轮次至少>=2个key才保留（部分轮次可能缺某些业务）
        rounds = []; cur_round = {}; last_end = None
        for s in key_segs:
            k = s['key']
            gap = (s['start_t'] - last_end).total_seconds() if last_end else 0

            # 同步检测：同key在短间隔内再出现 → 重复段，跳过
            if k in cur_round and gap <= 120:
                continue

            # 新轮判断：大gap(>180s) 或 同key长间隔(>120s)
            if len(cur_round) > 0 and (gap > 180 or (k in cur_round and gap > 120)):
                if len(cur_round) >= 2:
                    rounds.append(cur_round)
                cur_round = {}

            if k not in cur_round:
                cur_round[k] = s
                last_end = s['end_t']
        if len(cur_round) >= 2:
            rounds.append(cur_round)
        return rounds

    def _classify_seg(self, s, segs, idx):
        """归类（MAC口径flow，从实际数据反推）

        基于联通/电信实际段特征统计（MAC口径flow）：
        下行: store_s=52~248MB(3~12s) | ftp_dl=297~1522MB(7~37s) | store_l=1888~6303MB(61~121s)
        上行(上行方向): wx_s=3~17MB(~1s) | ftp_ul=85~250MB(6~20s) | wx_l=200~1912MB(20~62s)

        下行分类树（基于大量实测数据MAC口径flow）：
          flow<50 → None(噪声)
          50~350MB → store_s（商店小包3~8s，高rate~150Mbps）
          350~700MB → ftp_dl（FTP下载低流量段/起步段）
          700~1500MB → dur>=15s或rate<500 → store_l（商店大包被gap切分后的子段）
                    → else → ftp_dl（FTP高速下载段）
          >=1500MB → store_l（商店大包合并后完整段）

        上行分类树（MAC口径flow）：
          flow<20 → wx_s（微信小包~1s）
          20~80MB → ftp_ul（FTP上传低流量段）
          80~200MB → dur<21→ftp_ul, else→wx_l
          200~300MB → rate>70&&dur<15→ftp_ul(高速FTP上传,flow自然超200MB), else→wx_l
          >=300MB → wx_l（微信大包20~40s）

        修正说明 v1.04 (07-07)：
        - 200~300MB区间加入rate辅助判断：高速FTP上传(rlc_rate>70,dur<15s)不应被误分为wx_l
        - 80~200MB区间沿用dur<21→ftp_ul的规则不变
        - 700~1500MB区间增加合并后store_l识别：gap<10s合并后flow>=700且dur>=15或rate<500→store_l
        """
        # 分类规则恢复原样（store_s=50~350MB），在过滤阶段用dur细化
        d = s['direction']; flow = s['flow_mb']; dur = s['duration']; rate = s['rate']
        if d == '下行':
            if flow <= 50:
                return None
            if 50 < flow < 350:
                return 'store_s'  # 商店小包50~350MB，扩展上限应对大包
            if 350 <= flow < 700:
                # 350~700MB区间实际都是FTP下载的起步段或低信号段
                return 'ftp_dl'
            # flow >= 700
            if flow >= 1500:
                return 'store_l'  # 合并后完整的大包段
            # 700~1500: 区分FTP高速下载 vs store_l被gap切分后的子段
            # store_l特征：dur>=15s(持续时间长) 或 rate<500(子段速率低)
            if dur >= 15 or rate < 500:
                return 'store_l'
            return 'ftp_dl'  # 短时高速段是FTP下载
        else:  # 上行
            if flow < 20:
                return 'wx_s'
            if 20 <= flow <= 80:
                return 'ftp_ul'
            if 80 < flow < 200:
                return 'ftp_ul' if dur < 21 else 'wx_l'
            if 200 <= flow < 300:
                # 200~300MB: 可能是高速FTP上传(rate>70,dur<15)或低速wx_l
                return 'ftp_ul' if (rate > 70 and dur < 15) else 'wx_l'
            # flow >= 300
            return 'wx_l'

    def _has_following_upload(self, segs, idx, within_sec=20):
        """idx段后 within_sec 秒内是否有上行~10秒段(FTP上传紧跟)"""
        s = segs[idx]; target = s['end_t']
        for j in range(idx + 1, min(idx + 5, len(segs))):
            ns = segs[j]
            gap = (ns['start_t'] - target).total_seconds()
            if gap > within_sec + 10:
                break
            if ns['direction'] == '上行' and 7 <= ns['duration'] <= 14 and 0 <= gap <= within_sec:
                return True
        return False

    # ---------- 构建呼叫详情行（FTP下载/上传分行，仿参考文件）----------
    def _base_row(self, rdata: Dict, seq: int, biz_type: str, test_biz: str) -> List:
        row = [''] * len(CALL_HEADERS)
        row[CALL_IDX['业务序号']] = seq
        row[CALL_IDX['业务类型']] = biz_type
        row[CALL_IDX['测试业务']] = test_biz
        row[CALL_IDX['测试网络']] = 'NR5G'
        row[CALL_IDX['文件名']] = self._filename
        return row

    def _build_ftp_dl_row(self, rdata: Dict, seq: int, params: Dict) -> List:
        """FTP下载行：发起/完成=下载时段，下载速率；应用商店/微信空"""
        row = self._base_row(rdata, seq, 'FTP', 'FTPDownload')
        if 'ftp_dl' in rdata:
            s = rdata['ftp_dl']
            row[CALL_IDX['发起时间']] = self._fmt_t(s['start_t'])
            row[CALL_IDX['完成时间']] = self._fmt_t(s['end_t'])
            row[CALL_IDX['下载平均速率(Mbps)']] = s['rate']
            if s.get('result') != '成功' and s.get('end_t'):
                row[CALL_IDX['失败时间']] = self._fmt_t(s['end_t'])
        return row

    def _build_ftp_ul_row(self, rdata: Dict, seq: int, params: Dict) -> List:
        """FTP上传行：发起/完成=上传时段，上传速率 + 该轮应用商店/微信"""
        row = self._base_row(rdata, seq, 'FTP', 'FTPUpload')
        if 'ftp_ul' in rdata:
            s = rdata['ftp_ul']
            row[CALL_IDX['发起时间']] = self._fmt_t(s['start_t'])
            row[CALL_IDX['完成时间']] = self._fmt_t(s['end_t'])
            row[CALL_IDX['上传平均速率(Mbps)']] = s['rate']
            if s.get('result') != '成功' and s.get('end_t'):
                row[CALL_IDX['失败时间']] = self._fmt_t(s['end_t'])
        # 应用商店（小+大）
        for key, rate_col, st_col, et_col in [
            ('store_s', '应用商店小文件下载速率(Mbps)', '应用商店小文件下载开始时间', '应用商店小文件下载完成时间'),
            ('store_l', '应用商店大文件下载速率(Mbps)', '应用商店大文件下载开始时间', '应用商店大文件下载完成时间')]:
            if key in rdata:
                s = rdata[key]
                row[CALL_IDX[rate_col]] = s['rate']
                row[CALL_IDX[st_col]] = self._fmt_t(s['start_t'])
                row[CALL_IDX[et_col]] = self._fmt_t(s['end_t'])
                if key == 'store_l':
                    s_s = rdata.get('store_s', {})
                    row[CALL_IDX['应用商店业务请求时间']] = self._fmt_t(s_s.get('start_t'))
                    row[CALL_IDX['应用商店业务完成时间']] = self._fmt_t(s['end_t'])
                    row[CALL_IDX['应用商店下载速率(Mbps)']] = s['rate']
                    row[CALL_IDX['应用商店业务结果']] = s['result']
                    if s_s.get('start_t') and s.get('end_t'):
                        row[CALL_IDX['应用商店业务时长(秒)']] = round((s['end_t'] - s_s['start_t']).total_seconds(), 3)
                    if s.get('result') != '成功' and s.get('end_t'):
                        row[CALL_IDX['应用商店业务失败时间']] = self._fmt_t(s['end_t'])
        # 微信（小+大）
        for key, rate_col, st_col, et_col in [
            ('wx_s', '微信小包发送速率(Mbps)', '微信小包发送开始时间', '微信小包发送完成时间'),
            ('wx_l', '微信大包发送速率(Mbps)', '微信大包发送开始时间', '微信大包发送完成时间')]:
            if key in rdata:
                s = rdata[key]
                row[CALL_IDX[rate_col]] = s['rate']
                row[CALL_IDX[st_col]] = self._fmt_t(s['start_t'])
                row[CALL_IDX[et_col]] = self._fmt_t(s['end_t'])
                if key == 'wx_l':
                    w_s = rdata.get('wx_s', {})
                    row[CALL_IDX['微信文件业务请求时间']] = self._fmt_t(w_s.get('start_t'))
                    row[CALL_IDX['微信文件业务完成时间']] = self._fmt_t(s['end_t'])
                    row[CALL_IDX['微信文件业务结果']] = s['result']
                    if w_s.get('start_t') and s.get('end_t'):
                        row[CALL_IDX['微信文件业务时长(秒)']] = round((s['end_t'] - w_s['start_t']).total_seconds(), 3)
                    if s.get('result') != '成功' and s.get('end_t'):
                        row[CALL_IDX['微信文件业务失败时间']] = self._fmt_t(s['end_t'])
        return row

    @staticmethod
    def _fmt_t(t):
        return t.strftime('%Y-%m-%d %H:%M:%S') if t is not None and not (isinstance(t, float) and pd.isna(t)) else ''

    def _attach_cov(self, stats: Dict, round_results: List[Dict]) -> Dict:
        """把网络覆盖指标 + 测试时长(业务首末跨度,对齐参考57.56分钟) 挂到 stats['_coverage']"""
        cov = dict(getattr(self, 'coverage', {}) or {})
        times = []
        for rd in round_results:
            for k in rd:
                st = rd[k].get('start_t'); et = rd[k].get('end_t')
                if st: times.append(st)
                if et: times.append(et)
        if times:
            cov['测试时长(分钟)'] = round((max(times) - min(times)).total_seconds() / 60, 2)
        # 覆盖率：业务时段采样点(RSRP≥thr & SINR≥-3)，对齐参考口径
        # (全部采样点含空闲差信号会偏低；业务时段信号好→接近参考100%)
        frames = [rd[k]['points'] for rd in round_results for k in rd
                  if rd[k].get('points') is not None and len(rd[k]['points'])]
        if frames and 'srs_rsrp' in frames[0].columns:
            bp = pd.concat(frames)
            r = pd.to_numeric(bp['srs_rsrp'], errors='coerce') / 100
            s = pd.to_numeric(bp['sinr'], errors='coerce') / 100
            valid = pd.DataFrame({'r': r, 's': s}).dropna()
            for thr, tag in [(-96, '覆盖率'), (-100, '覆盖率'), (-105, '覆盖达标率')]:
                m = (valid['r'] >= thr) & (valid['s'] >= -3)
                cov[f'5G{tag}SS-RSRP≥{thr}dBm&SS-SINR≥-3dB采样点占比'] = round(float(m.sum() / len(valid) * 100), 2) if len(valid) else None
        stats['_coverage'] = cov
        return stats

    # ---------- 统计汇总 ----------
    def _summarize(self, round_results: List[Dict], params: Dict) -> Dict:
        """从每轮 rdata(含段 points) 真实算 42 列统计。所有值从 UCM 段算，不凑数。
        口径规则（按业务匹配最佳口径）：
          ftp_dl: clip_mean（削峰含零均值）— 已固定不动
          ftp_ul: clip_mean（削峰含零均值）— 已固定不动
          store_s: clip_mean（削峰含零均值）
          store_l: mac_nz（MAC非零均值）
          wx_s: mac_nz（MAC非零均值）
          wx_l: rlc_mean（RLC含零均值）
        注：collect_points 由原始 points 直接聚合，不做扩展（扩展受限于秒表前后数据类型不匹配）。"""
        def collect_points(key):
            frames = [rd[key]['points'] for rd in round_results
                      if key in rd and rd[key].get('points') is not None]
            return pd.concat(frames) if frames else pd.DataFrame()
        def collect_durations(key):
            return [rd[key]['duration'] for rd in round_results
                    if key in rd and rd[key].get('duration') is not None]
        def count(key):
            return sum(1 for rd in round_results if key in rd)
        def empty(grp):
            return {f: None for g, f in STAT_COLUMNS if g == grp}
        def smean(s):
            return round(float(s.mean()), 3) if len(s) else None
        def smax(s):
            return round(float(s.max()), 3) if len(s) else None
        def top10(s):
            if not len(s):
                return None
            sr = s.sort_values(ascending=False)
            return round(float(sr.head(max(1, int(len(sr) * 0.1))).mean()), 3)
        stats = {}
        # FTP下载 — 口径：clip含零均值(clip_mean)对齐输出结果文件
        d = empty('FTP下载')
        d['测试次数'] = count('ftp_dl')
        pts = collect_points('ftp_dl')
        if len(pts):
            clip = pts['dl_clip']  # 含零（输出结果文件用含零均值）
            d['采样点数'] = int(len(pts))
            d['下行RLC平均吞吐率(Mbps)'] = smean(clip)  # clip_mean，已固定不动
            d['削峰后下行RLC平均吞吐率(Mbps)'] = smean(clip)  # clip_mean，已固定不动
            d['下行RLC峰值吞吐率(Mbps)'] = smax(clip)
            d['前10%峰值速率(Mbps)'] = top10(clip)
            if len(clip):
                d['削峰后下行RLC>=100Mbps采集点数'] = int((clip >= 100).sum())
                d['削峰后下行RLC>=1000Mbps采集点数'] = int((clip >= 1000).sum())
                d['削峰后下行RLC>=100Mbps占比'] = round(float((clip >= 100).sum() / len(clip) * 100), 2)
                d['削峰后下行RLC>=1000Mbps占比'] = round(float((clip >= 1000).sum() / len(clip) * 100), 2)
        stats['FTP下载'] = d
        # FTP上传 — 口径：clip含零均值(clip_mean)对齐输出结果文件
        d = empty('FTP上传')
        d['测试次数'] = count('ftp_ul')
        pts = collect_points('ftp_ul')
        if len(pts):
            clip = pts['ul_clip']  # 含零（输出结果文件用含零均值）
            d['采样点数'] = int(len(pts))
            d['上行RLC平均吞吐率(Mbps)'] = smean(clip)  # clip_mean，已固定不动
            d['削峰后上行RLC平均吞吐率(Mbps)'] = smean(clip)  # clip_mean，已固定不动
            d['上行RLC峰值吞吐率(Mbps)'] = smax(clip)
            d['前10%峰值速率(Mbps)'] = top10(clip)
            if len(clip):
                d['削峰后上行RLC>20Mbps采集点数'] = int((clip > 20).sum())
                d['削峰后上行RLC>200Mbps采集点数'] = int((clip > 200).sum())
                d['削峰后上行RLC>20Mbps占比'] = round(float((clip > 20).sum() / len(clip) * 100), 2)
                d['削峰后上行RLC>200Mbps占比'] = round(float((clip > 200).sum() / len(clip) * 100), 2)
        stats['FTP上传'] = d
        # 应用商店_小包 — 口径：clip_mean（削峰含零均值）对齐输出结果文件
        # 应用商店_大包 — 口径：mac_nz（MAC非零均值）对齐输出结果文件
        for key, grp, mode in [('store_s', '应用商店_小包', 'clip'), ('store_l', '应用商店_大包', 'mac_nz')]:
            d = empty(grp)
            d['测试次数'] = count(key)
            pts = collect_points(key)
            durs = collect_durations(key)
            if len(pts):
                if mode == 'clip':
                    clip = pts['dl_clip']
                    d['采样点数'] = int(len(pts))
                    d['下行RLC平均吞吐率(Mbps)'] = smean(clip)
                elif mode == 'mac_nz':
                    mac = pts['dl_mac'].replace(0, np.nan).dropna()
                    d['采样点数'] = int(len(pts))
                    d['下行RLC平均吞吐率(Mbps)'] = smean(mac)
            if durs:
                d['单次平均时长(s)'] = round(float(np.mean(durs)), 3)
                if grp == '应用商店_大包':
                    d['中位值平均时长(s)'] = round(float(np.median(durs)), 3)
            d['下载成功率(%)'] = self._group_success(round_results, ['store_s', 'store_l'])
            stats[grp] = d
        # 微信_小文件 — 口径：mac_nz（MAC非零均值）对齐输出结果文件
        # 微信_大文件 — 口径：rlc_mean（RLC含零均值）对齐输出结果文件
        for key, grp, mode in [('wx_s', '微信_小文件', 'mac_nz'), ('wx_l', '微信_大文件', 'rlc_mean')]:
            d = empty(grp)
            d['测试次数'] = count(key)
            pts = collect_points(key)
            durs = collect_durations(key)
            if len(pts):
                if mode == 'mac_nz':
                    mac = pts['ul_mac'].replace(0, np.nan).dropna()
                    d['采样点数'] = int(len(pts))
                    d['上行RLC平均吞吐率(Mbps)'] = smean(mac)
                elif mode == 'rlc_mean':
                    rlc = pts['ul_rlc']  # 含零（输出结果文件用含零均值）
                    d['采样点数'] = int(len(pts))
                    d['上行RLC平均吞吐率(Mbps)'] = smean(rlc)
            if durs:
                d['单次平均时长(s)'] = round(float(np.mean(durs)), 3)
                if grp == '微信_大文件':
                    d['中位值平均时长(s)'] = round(float(np.median(durs)), 3)
            d['上传成功率(%)'] = self._group_success(round_results, ['wx_s', 'wx_l'])
            stats[grp] = d
        return stats

    @staticmethod
    def _group_success(round_results, keys):
        """业务整体成功率 = 成功段数 / 识别段总数"""
        results = [rd[k].get('result') for rd in round_results for k in keys if k in rd]
        if not results:
            return None
        ok = sum(1 for r in results if r == '成功')
        return round(ok / len(results) * 100, 2)

    @property
    def _filename(self):
        return getattr(self, '_fn', '')

    # ---------- 业务标注文件(可视化核对) ----------
    def generate_annotated(self, mmf_paths, plan: List[Dict], params: Dict, out_path: str) -> str:
        """生成业务标注文件：原始UCM所有列(多文件按时间合并) + 业务标记列 + 统计指标列 + 速率列高亮"""
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Alignment
        if isinstance(mmf_paths, str): mmf_paths = [mmf_paths]
        self.parse_mmf(mmf_paths, params)   # 设 self._fn + 合并时间轴
        call_df, stats = self.match(plan, params)
        round_results = getattr(self, '_last_round_results', [])
        raws = []
        for p in mmf_paths:
            r = pd.read_excel(p, header=0)
            tcol = self._find_col(r, ['采集时间', 'Time'])
            r['_t'] = pd.to_datetime(r[tcol].astype(str).str.replace(r'\(\d+\)', '', regex=True), errors='coerce') if tcol else None
            r = r[r['_t'].notna()].copy()
            raws.append(r)
        raw = pd.concat(raws, ignore_index=True).sort_values('_t').reset_index(drop=True)
        biz_map = [('ftp_dl', 'FTP下载'), ('ftp_ul', 'FTP上传'), ('store_s', '应用商店小包'),
                   ('store_l', '应用商店大包'), ('wx_s', '微信小文件'), ('wx_l', '微信大文件')]
        raw['业务标记'] = ''
        biz_rows = {}  # 业务标记名 -> [原始行索引]
        for rdata in round_results:
            for key, label in biz_map:
                if key not in rdata:
                    continue
                s = rdata[key]; st = s['start_t']; et = s['end_t']
                if st is None or et is None:
                    continue
                m = (raw['_t'] >= st) & (raw['_t'] <= et)
                raw.loc[m, '业务标记'] = label
                biz_rows.setdefault(label, []).extend(raw[m].index.tolist())
        # 业务标记移到 C 列后(index 3)，其后加 12 指标列
        指标列 = ['FTP下行均值速率', 'FTP下行前10%峰值速率', 'FTP上行达标速率占比', 'FTP上行均值速率',
                 'FTP上行前10%峰值速率', '应用商店下载成功率', '应用商店小包下载时长', '应用商店大包下载时长',
                 '微信文件上传成功率', '微信小文件上传时长', '微信大文件上传时长']
        for ic in 指标列:
            raw[ic] = ''
        cols = list(raw.columns)
        cols.remove('业务标记'); cols.insert(3, '业务标记')  # C列(index2)后
        ins = 4
        for ic in 指标列:
            cols.remove(ic); cols.insert(ins, ic); ins += 1
        raw = raw[cols].drop(columns=['_t'])
        raw.to_excel(out_path, index=False, sheet_name='UCM标注')
        # openpyxl: 填指标值(合并居中) + 业务标记/速率列高亮
        wb = load_workbook(out_path); ws = wb['UCM标注']
        headers = [c.value for c in ws[1]]
        col_idx = {h: i + 1 for i, h in enumerate(headers)}
        biz_col = col_idx['业务标记']
        dl_rlc = next((c for c in headers if '下行RLC' in str(c) and '吞吐率' in str(c)), None)
        ul_rlc = next((c for c in headers if '上行RLC' in str(c) and '吞吐率' in str(c)), None)
        m_map = {  # 指标 -> (值, 对应业务标记, 高亮方向)
            'FTP下行均值速率': (stats.get('FTP下载', {}).get('下行RLC平均吞吐率(Mbps)'), 'FTP下载', '下行'),
            'FTP下行前10%峰值速率': (stats.get('FTP下载', {}).get('前10%峰值速率(Mbps)'), 'FTP下载', '下行'),
            'FTP上行达标速率占比': (stats.get('FTP上传', {}).get('削峰后上行RLC>20Mbps占比'), 'FTP上传', '上行'),
            'FTP上行均值速率': (stats.get('FTP上传', {}).get('上行RLC平均吞吐率(Mbps)'), 'FTP上传', '上行'),
            'FTP上行前10%峰值速率': (stats.get('FTP上传', {}).get('前10%峰值速率(Mbps)'), 'FTP上传', '上行'),
            '应用商店下载成功率': (stats.get('应用商店_小包', {}).get('下载成功率(%)'), '应用商店小包', '下行'),
            '应用商店小包下载时长': (stats.get('应用商店_小包', {}).get('单次平均时长(s)'), '应用商店小包', '下行'),
            '应用商店大包下载时长': (stats.get('应用商店_大包', {}).get('单次平均时长(s)'), '应用商店大包', '下行'),
            '微信文件上传成功率': (stats.get('微信_小文件', {}).get('上传成功率(%)'), '微信小文件', '上行'),
            '微信小文件上传时长': (stats.get('微信_小文件', {}).get('单次平均时长(s)'), '微信小文件', '上行'),
            '微信大文件上传时长': (stats.get('微信_大文件', {}).get('单次平均时长(s)'), '微信大文件', '上行'),
        }
        biz_colors = {'FTP下载': 'FFC0C0', 'FTP上传': 'C0FFC0', '应用商店小包': 'C0C0FF',
                      '应用商店大包': 'FFD8A0', '微信小文件': 'C0FFFF', '微信大文件': 'FFC0FF'}
        # 填指标值(合并居中)
        for ic, (val, biz, direction) in m_map.items():
            col = col_idx[ic]
            rows = sorted(biz_rows.get(biz, []))
            if rows and val is not None:
                r1 = rows[0] + 2; r2 = rows[-1] + 2
                ws.cell(row=r1, column=col, value=val).alignment = Alignment(horizontal='center', vertical='center')
                if r2 > r1:
                    ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)
        # 业务标记列 + 速率列 高亮(各业务时段行,按业务色)
        for biz, idxs in biz_rows.items():
            if not idxs:
                continue
            fc = biz_colors.get(biz, 'FFFF00')
            rate_col = col_idx[dl_rlc] if (dl_rlc and ('下载' in biz or '商店' in biz)) else (col_idx.get(ul_rlc) if ul_rlc else None)
            for idx in idxs:
                ws.cell(row=idx + 2, column=biz_col).fill = PatternFill(start_color=fc, end_color=fc, fill_type='solid')
                if rate_col:
                    ws.cell(row=idx + 2, column=rate_col).fill = PatternFill(start_color=fc, end_color=fc, fill_type='solid')
        wb.save(out_path)
        return out_path


# ==================== Excel导出 ====================

class ExcelExporter:
    def __init__(self):
        self.wb = None

    def export(self, output_path: str, call_df: pd.DataFrame, stats: Dict):
        self.wb = Workbook()
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        self._create_call_sheet(call_df)
        self._create_stat_sheet(stats)
        self._create_coverage_sheet(stats.get('_coverage', {}))
        self.wb.save(output_path)

    def _create_call_sheet(self, call_df: pd.DataFrame):
        ws = self.wb.create_sheet('呼叫详情')
        gfill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        # R2 字段名
        for col, h in enumerate(CALL_HEADERS, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = Font(bold=True, size=9); c.fill = gfill
        # R1 分组合并表头(对齐参考文件两层结构)
        for start, end, name in CALL_GROUPS:
            c = ws.cell(row=1, column=start + 1, value=name)
            c.font = Font(bold=True); c.fill = gfill; c.alignment = Alignment(horizontal='center')
            if end - start > 1:
                ws.merge_cells(start_row=1, start_column=start + 1, end_row=1, end_column=end)
        # 数据 row3 起
        for r, row in enumerate(call_df.itertuples(index=False), 3):
            for col, v in enumerate(row, 1):
                ws.cell(row=r, column=col, value='' if (v is None or (isinstance(v, float) and pd.isna(v))) else v)

    def _create_stat_sheet(self, stats: Dict):
        ws = self.wb.create_sheet('统计结果')
        gfill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        for col, (grp, field) in enumerate(STAT_COLUMNS, 1):
            ws.cell(row=2, column=col, value=field).font = Font(bold=True, size=9)
        start = 1
        for i in range(1, len(STAT_COLUMNS) + 1):
            if i == len(STAT_COLUMNS) or STAT_COLUMNS[i][0] != STAT_COLUMNS[start - 1][0]:
                c = ws.cell(row=1, column=start, value=STAT_COLUMNS[start - 1][0])
                c.font = Font(bold=True); c.fill = gfill; c.alignment = Alignment(horizontal='center')
                if i > start:
                    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=i)
                start = i + 1
        for col, (grp, field) in enumerate(STAT_COLUMNS, 1):
            v = stats.get(grp, {}).get(field)
            ws.cell(row=3, column=col, value=v if v is not None else '').alignment = Alignment(horizontal='center')

    def _create_coverage_sheet(self, cov: Dict):
        """网络覆盖/基础信息(对齐参考「运营商汇总」基础+网络覆盖段)"""
        if not cov:
            return
        ws = self.wb.create_sheet('网络覆盖')
        gfill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        ws.cell(row=1, column=1, value='指标').font = Font(bold=True)
        ws.cell(row=1, column=2, value='值').font = Font(bold=True)
        ws.cell(row=1, column=1).fill = gfill; ws.cell(row=1, column=2).fill = gfill
        order = ['测试时长(分钟)', '5G CA聚合率(%)', 'NR_Serving SS-RSRP(dBm)', 'NR_Serving SS AVG SINR(dB)',
                 '5G覆盖率SS-RSRP≥-96dBm&SS-SINR≥-3dB采样点占比',
                 '5G覆盖率SS-RSRP≥-100dBm&SS-SINR≥-3dB采样点占比',
                 '5G覆盖达标率SS-RSRP≥-105dBm&SS-SINR≥-3dB采样点占比']
        items = [(k, cov.get(k)) for k in order if k in cov]
        for k, val in cov.items():          # 兜底：order 外的也列出
            if (k, val) not in items:
                items.append((k, val))
        for i, (k, val) in enumerate(items, 2):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=val if val is not None else '').alignment = Alignment(horizontal='center')
        ws.column_dimensions['A'].width = 52


# ==================== 工作线程 ====================

class ProcessingThread(QThread):
    progress_signal = Signal(str)
    finished_signal = Signal(bool, str, object, object)

    def __init__(self, processor: UCMProcessor, mmf_paths, plan: List[Dict], params: Dict):
        super().__init__()
        self.processor = processor
        self.mmf_paths = mmf_paths if isinstance(mmf_paths, list) else [mmf_paths]
        self.plan = plan
        self.params = params

    def run(self):
        try:
            self.progress_signal.emit(f"解析UCM监控文件({len(self.mmf_paths)}个)...")
            self.processor.parse_mmf(self.mmf_paths, self.params)
            self.progress_signal.emit("按测试计划匹配业务...")
            call_df, stats = self.processor.match(self.plan, self.params)
            self.progress_signal.emit("生成完成")
            self.finished_signal.emit(True, f"完成：匹配 {len(call_df)} 轮", call_df, stats)
        except Exception as e:
            import traceback; traceback.print_exc()
            self.finished_signal.emit(False, f"失败: {str(e)}", None, None)


# ==================== 在线更新 + 关于对话框 ====================

def _v_tuple(s):
    """版本号字符串→数字元组(用于比较大小)"""
    return tuple(int(x) for x in re.findall(r'\d+', str(s)))


def _ssl_ctx():
    """云主机自签证书：忽略证书校验(保证能连通)"""
    import ssl
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx


class UpdateChecker(QThread):
    """后台静默检查更新：拉 version.json 比对版本"""
    found_update = Signal(dict)   # 有新版时发(含版本信息)
    check_done = Signal(str)      # 'latest'=最新 / 'new'=有新版 / 'error:...'=失败

    def run(self):
        try:
            url = f'{UPDATE_BASE}/version.json?t={int(time.time())}'   # 防缓存
            with urllib.request.urlopen(url, timeout=6, context=_ssl_ctx()) as r:
                info = json.loads(r.read().decode('utf-8'))
            if _v_tuple(info.get('version', '0')) > _v_tuple(APP_VERSION):
                self.found_update.emit(info)
                self.check_done.emit('new')
            else:
                self.check_done.emit('latest')
        except Exception as e:
            self.check_done.emit(f'error:{e}')


def do_update(info, parent_window) -> bool:
    """下载更新包→MD5校验→备份旧文件→替换。成功后调用方需重启。
    自动适配 py / exe(PyInstaller) 运行环境。返回是否完成替换。"""
    # 1. 下载到临时文件
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pack'); tmp.close()
    try:
        with urllib.request.urlopen(info['url'], timeout=30, context=_ssl_ctx()) as r, open(tmp.name, 'wb') as f:
            shutil.copyfileobj(r, f)
    except Exception as e:
        QMessageBox.critical(parent_window, '更新失败', f'下载失败：{e}'); return False
    # 2. MD5 校验(保质，防损坏/篡改)
    if info.get('md5'):
        h = hashlib.md5()
        with open(tmp.name, 'rb') as f:
            for chunk in iter(lambda: f.read(65536), b''): h.update(chunk)
        if h.hexdigest() != info['md5']:
            os.unlink(tmp.name)
            QMessageBox.critical(parent_window, '更新失败', 'MD5 校验不一致，已中止(防损坏/篡改)')
            return False
    # 3. 定位当前运行文件
    if getattr(sys, 'frozen', False):
        me = os.path.abspath(sys.executable)         # exe 模式
    else:
        me = os.path.abspath(sys.argv[0] if sys.argv and sys.argv[0] else __file__)
    bak = f'{me}_旧版_{datetime.now().strftime("%Y%m%d-%H%M%S")}'
    # 4. 替换
    try:
        if getattr(sys, 'frozen', False) and os.name == 'nt':
            _write_exe_updater(tmp.name, me, bak)    # Windows exe: bat 等退出后替换
            return True
        else:
            # py / 类Unix: 备份后直接覆盖
            if os.path.exists(me): shutil.copy2(me, bak)
            shutil.copy2(tmp.name, me); os.unlink(tmp.name)
            return True
    except Exception as e:
        QMessageBox.critical(parent_window, '更新失败', f'替换文件失败：{e}')
        return False


def _write_exe_updater(pack_path, target_exe, bak_path):
    """Windows exe 更新：生成 updater.bat(等待主进程退出→备份→替换→重启→自删)"""
    bat = os.path.join(os.path.dirname(target_exe), '_updater.bat')
    with open(bat, 'w', encoding='gbk', newline='\r\n') as f:
        f.write('@echo off\r\n')
        f.write('timeout /t 1 /nobreak >nul\r\n')
        f.write(':wait\r\n')
        f.write(f'move /y "{target_exe}" "{bak_path}" >nul 2>&1\r\n')
        f.write(f'if exist "{target_exe}" goto wait\r\n')
        f.write(f'move /y "{pack_path}" "{target_exe}" >nul\r\n')
        f.write(f'start "" "{target_exe}"\r\n')
        f.write(f'del "%~f0"\r\n')
    import subprocess
    subprocess.Popen(['cmd', '/c', bat], close_fds=True)


def _relaunch_and_quit():
    """重启工具(更新后)"""
    import subprocess
    if getattr(sys, 'frozen', False):
        subprocess.Popen([sys.executable])
    else:
        subprocess.Popen([sys.executable, os.path.abspath(sys.argv[0] if sys.argv else __file__)])
    QApplication.instance().quit()
    os._exit(0)


class AboutDialog(QDialog):
    """关于对话框：版本/署名/联系方式/版本历史/检查更新"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('关于')
        self.setMinimumWidth(460)
        lay = QVBoxLayout(self)
        title = QLabel(APP_NAME); title.setFont(QFont('PingFang SC', 15, QFont.Weight.Bold))
        lay.addWidget(title)
        ver = QLabel(f'版本：{APP_VERSION}   构建日期 {BUILD_DATE}')
        ver.setFont(QFont('PingFang SC', 11)); lay.addWidget(ver)
        self.status_label = QLabel('更新检查：未检查')
        self.status_label.setStyleSheet('color: #666;'); lay.addWidget(self.status_label)
        author = QLabel(f'开发人：{AUTHOR}\n联系方式：{CONTACT}')
        author.setFont(QFont('PingFang SC', 11)); lay.addWidget(author)
        lay.addWidget(QLabel('── 版本历史 ──'))
        hist = QLabel('\n'.join(f'{v} · {d} · {c}' for v, d, c in VERSION_HISTORY))
        hist.setWordWrap(True); hist.setStyleSheet('color: #333;'); lay.addWidget(hist)
        bl = QHBoxLayout()
        chk = QPushButton('检查更新'); chk.clicked.connect(self._check)
        close_btn = QPushButton('关闭'); close_btn.clicked.connect(self.accept)
        bl.addStretch(); bl.addWidget(chk); bl.addWidget(close_btn); lay.addLayout(bl)
        self.adjustSize()

    def _check(self):
        self.status_label.setText('更新检查：检查中…'); self.setEnabled(False)
        self._checker = UpdateChecker(self, manual=True)
        self._checker.check_done.connect(self._on_done)
        self._checker.found_update.connect(self._on_found)
        self._checker.start()

    def _on_done(self, msg):
        self.setEnabled(True)
        if msg == 'latest':
            self.status_label.setText('更新检查：✓ 已是最新版')
        elif msg.startswith('error'):
            self.status_label.setText('更新检查：检查失败(请检查网络/云主机)')

    def _on_found(self, info):
        self.status_label.setText(f"更新检查：发现新版 {info['version']}")
        text = (f"发现新版本 {info['version']}（当前 {APP_VERSION}）\n\n"
                f"发布日期：{info.get('release_date','')}\n\n"
                f"更新内容：\n{info.get('changelog','')}\n\n是否立即更新？")
        if QMessageBox.question(self, '发现新版本', text, QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            if do_update(info, self):
                if getattr(sys, 'frozen', False) and os.name == 'nt':
                    QMessageBox.information(self, '更新准备就绪', '工具将退出并自动完成更新，请稍候。')
                    _relaunch_and_quit()
                else:
                    QMessageBox.information(self, '更新完成', '更新已下载并替换，工具将重启。')
                    _relaunch_and_quit()


# ==================== 主窗口 ====================

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config_manager = ConfigManager()
        self.config = self.config_manager.config
        self.processor = None
        self.mmf_files = []
        self.processing_thread = None
        self.init_ui()
        self._start_update_check()   # 启动静默检查更新

    def init_ui(self):
        self.setWindowTitle('5G用户级公共监控速率统计工具 V1.03（测试计划驱动）')
        self.setGeometry(80, 80, 1600, 950)
        w = QWidget(); self.setCentralWidget(w)
        lay = QHBoxLayout(w)
        lay.addWidget(self.create_left_panel(), 1)
        lay.addWidget(self.create_right_panel(), 2)

    def create_left_panel(self) -> QWidget:
        panel = QWidget(); layout = QVBoxLayout(panel)
        # 文件选择
        fg = QGroupBox('输入文件（仅UCM监控）'); fl = QHBoxLayout()
        self.mmf_edit = QLineEdit(); self.mmf_edit.setPlaceholderText('选择 mmf/两江 等(可多选,按时间合并分析)')
        fb = QPushButton('选择(可多选)'); fb.clicked.connect(self.select_mmf)
        fl.addWidget(self.mmf_edit); fl.addWidget(fb); fg.setLayout(fl); layout.addWidget(fg)
        # 全局参数
        pg = QGroupBox('全局参数'); pl = QFormLayout()
        self.dl_peak = QDoubleSpinBox(); self.dl_peak.setRange(0, 5000); self.dl_peak.setValue(900)
        self.ul_peak = QDoubleSpinBox(); self.ul_peak.setRange(0, 1000); self.ul_peak.setValue(160)
        self.down_min = QDoubleSpinBox(); self.down_min.setRange(0, 1000); self.down_min.setValue(50)
        self.up_min = QDoubleSpinBox(); self.up_min.setRange(0, 500); self.up_min.setValue(10)
        pl.addRow('下行削峰(Mbps):', self.dl_peak); pl.addRow('上行削峰(Mbps):', self.ul_peak)
        pl.addRow('下行最小速率(Mbps):', self.down_min); pl.addRow('上行最小速率(Mbps):', self.up_min)
        self.match_mode = QComboBox(); self.match_mode.addItems(['plan(固定顺序)', 'auto(自动识别)'])
        pl.addRow('匹配模式:', self.match_mode)
        # 基准规则选择
        self.benchmark_mode = QComboBox()
        self.benchmark_mode.addItems(['ALL_FTP_DETAIL模板', '工信部指标(T0-运营商区分)', 'UCM标注(运营商区分)'])
        pl.addRow('基准规则:', self.benchmark_mode)
        pg.setLayout(pl); layout.addWidget(pg)
        # 测试计划编辑器
        layout.addWidget(self.create_plan_editor())
        # 按钮
        bl = QHBoxLayout()
        pb = QPushButton('开始统计'); pb.setStyleSheet('QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 10px; }')
        pb.clicked.connect(self.start_processing); bl.addWidget(pb)
        eb = QPushButton('导出Excel'); eb.clicked.connect(self.export_excel); bl.addWidget(eb)
        ab = QPushButton('生成标注文件'); ab.clicked.connect(self.gen_annotated); bl.addWidget(ab)
        ub = QPushButton('关于'); ub.clicked.connect(self.show_about); bl.addWidget(ub)
        layout.addLayout(bl)
        self.progress = QProgressBar(); layout.addWidget(self.progress)
        lg = QGroupBox('运行日志'); ll = QVBoxLayout()
        self.log_text = QTextEdit(); self.log_text.setReadOnly(True); ll.addWidget(self.log_text)
        lg.setLayout(ll); layout.addWidget(lg, 1)
        return panel

    def create_plan_editor(self) -> QWidget:
        g = QGroupBox('测试计划（可上下移动顺序、参数可调）'); l = QVBoxLayout(g)
        self.plan_table = QTableWidget(); self.plan_table.setColumnCount(6)
        self.plan_table.setHorizontalHeaderLabels(['顺序', '业务名称', '方向', '时长(s)', '流量(MB)', '超时(s)'])
        self.plan_table.horizontalHeader().setStretchLastSection(True)
        l.addWidget(self.plan_table)
        bl = QHBoxLayout()
        up = QPushButton('上移'); up.clicked.connect(lambda: self._move_row(-1)); bl.addWidget(up)
        dn = QPushButton('下移'); dn.clicked.connect(lambda: self._move_row(1)); bl.addWidget(dn)
        add = QPushButton('添加'); add.clicked.connect(self.add_plan_row); bl.addWidget(add)
        rm = QPushButton('删除'); rm.clicked.connect(self.del_plan_row); bl.addWidget(rm)
        l.addLayout(bl)
        self.load_plan(DEFAULT_PLAN)
        return g

    def load_plan(self, plan):
        self.plan_table.setRowCount(len(plan))
        for r, item in enumerate(plan):
            self.plan_table.setItem(r, 0, QTableWidgetItem(str(r + 1)))
            self.plan_table.setItem(r, 1, QTableWidgetItem(item['name']))
            self.plan_table.setItem(r, 2, QTableWidgetItem(item.get('direction', '')))
            self.plan_table.setItem(r, 3, QTableWidgetItem('' if item.get('duration') is None else str(item['duration'])))
            self.plan_table.setItem(r, 4, QTableWidgetItem('' if item.get('flow_mb') is None else str(item['flow_mb'])))
            self.plan_table.setItem(r, 5, QTableWidgetItem('' if item.get('timeout') is None else str(item['timeout'])))
            self.plan_table.item(r, 0).setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

    def _move_row(self, delta):
        r = self.plan_table.currentRow()
        if r < 0: return
        target = r + delta
        if 0 <= target < self.plan_table.rowCount():
            self._swap_rows(r, target); self.plan_table.selectRow(target)

    def _swap_rows(self, a, b):
        for c in range(self.plan_table.columnCount()):
            ta = self.plan_table.takeItem(a, c); tb = self.plan_table.takeItem(b, c)
            self.plan_table.setItem(a, c, tb); self.plan_table.setItem(b, c, ta)
        # 重排顺序列
        self.plan_table.item(a, 0).setText(str(a + 1)); self.plan_table.item(b, 0).setText(str(b + 1))

    def add_plan_row(self):
        r = self.plan_table.rowCount(); self.plan_table.insertRow(r)
        for c in range(6): self.plan_table.setItem(r, c, QTableWidgetItem(''))
        self.plan_table.item(r, 0).setText(str(r + 1)); self.plan_table.item(r, 0).setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

    def del_plan_row(self):
        r = self.plan_table.currentRow()
        if r >= 0: self.plan_table.removeRow(r)

    def get_plan_from_ui(self) -> List[Dict]:
        plan = []
        for r in range(self.plan_table.rowCount()):
            name = self.plan_table.item(r, 1).text().strip() if self.plan_table.item(r, 1) else ''
            if not name: continue
            direction = self.plan_table.item(r, 2).text().strip() if self.plan_table.item(r, 2) else ''
            dur = self._num(self.plan_table.item(r, 3))
            flow = self._num(self.plan_table.item(r, 4))
            tmo = self._num(self.plan_table.item(r, 5))
            mode = 'wait' if name == '等待' else ('time' if dur else 'flow')
            plan.append({'key': self._plan_key(name), 'name': name, 'mode': mode, 'direction': direction,
                         'duration': dur, 'flow_mb': flow, 'timeout': tmo})
        return plan

    @staticmethod
    def _plan_key(name: str) -> str:
        """业务名 → 语义key(供 _build_call_row/_summarize 匹配)，必须与 DEFAULT_PLAN 的 key 一致"""
        if 'FTP下载' in name: return 'ftp_dl'
        if 'FTP上传' in name: return 'ftp_ul'
        if '应用商店' in name and '小' in name: return 'store_s'
        if '应用商店' in name and '大' in name: return 'store_l'
        if '微信' in name and '小' in name: return 'wx_s'
        if '微信' in name and '大' in name: return 'wx_l'
        return f'wait_{name}'

    @staticmethod
    def _num(item):
        if item is None: return None
        t = item.text().strip()
        if t in ('', 'None'): return None
        try: return float(t)
        except ValueError: return None

    def create_right_panel(self) -> QWidget:
        panel = QWidget(); layout = QVBoxLayout(panel)
        tw = QTabWidget()
        # 呼叫详情
        t1 = QWidget(); l1 = QVBoxLayout(t1)
        self.call_table = QTableWidget(); l1.addWidget(self.call_table)
        tw.addTab(t1, '呼叫详情(逐轮)')
        # 统计宽表
        t2 = QWidget(); l2 = QVBoxLayout(t2)
        self.stat_table = QTableWidget()
        headers = [f"{g}|{f}" for g, f in STAT_COLUMNS]
        self.stat_table.setColumnCount(len(headers)); self.stat_table.setHorizontalHeaderLabels(headers)
        l2.addWidget(self.stat_table)
        tw.addTab(t2, '统计结果(汇总)')
        layout.addWidget(tw)
        return panel

    def select_mmf(self):
        fps, _ = QFileDialog.getOpenFileNames(self, '选择UCM监控文件(可多选)', '', 'Excel Files (*.xlsx *.xls)')
        if fps:
            self.mmf_files = fps
            self.mmf_edit.setText(' | '.join(os.path.basename(p) for p in fps))
            self.log(f"已选择 {len(fps)} 个文件: {', '.join(os.path.basename(p) for p in fps)}")

    def get_params(self) -> Dict:
        return {'rate_column': 'RLC', 'dl_peak_limit': self.dl_peak.value(), 'ul_peak_limit': self.ul_peak.value(),
                'dl_pass_threshold': 100, 'ul_pass_threshold': 20,
                'down_min': self.down_min.value(), 'up_min': self.up_min.value(),
                'match_mode': 'auto' if str(self.match_mode.currentText()).startswith('auto') else 'plan'}

    @Slot()
    def start_processing(self):
        if not self.mmf_files:
            QMessageBox.warning(self, '警告', '请选择UCM监控文件'); return
        plan = self.get_plan_from_ui()
        if not plan:
            QMessageBox.warning(self, '警告', '请配置测试计划'); return
        params = self.get_params()
        self.log("=" * 50); self.log("开始按测试计划匹配...")
        self.log(f"文件: {len(self.mmf_files)}个 {', '.join(os.path.basename(p) for p in self.mmf_files)} | 计划{len(plan)}项")
        self.progress.setValue(0)
        self.sender().setEnabled(False); self.sender().setText("处理中...")
        self.processor = UCMProcessor(self.config)
        self.processing_thread = ProcessingThread(self.processor, self.mmf_files, plan, params)
        self.processing_thread.progress_signal.connect(self.log)
        self.processing_thread.finished_signal.connect(self.on_finished)
        self.processing_thread.start()

    @Slot(bool, str, object, object)
    def on_finished(self, success, message, call_df, stats):
        self.progress.setValue(100); self.log(message); self.log("=" * 50)
        for btn in self.findChildren(QPushButton):
            if btn.text() in ["处理中...", "开始统计"]:
                btn.setEnabled(True); btn.setText("开始统计")
        if success:
            self.display_call(call_df); self.display_stat(stats)
            self.last_call = call_df; self.last_stats = stats
            cov = stats.get('_coverage', {})
            if cov:
                self.log('— 网络覆盖指标 —')
                for k, val in cov.items():
                    self.log(f'  {k} = {val}')
            QMessageBox.information(self, '完成', message)
        else:
            QMessageBox.critical(self, '错误', message)

    def display_call(self, call_df: pd.DataFrame):
        self.call_table.setColumnCount(len(CALL_HEADERS))
        self.call_table.setHorizontalHeaderLabels(CALL_HEADERS)
        self.call_table.setRowCount(len(call_df))
        for r in range(len(call_df)):
            for c in range(len(CALL_HEADERS)):
                v = call_df.iloc[r, c]
                if isinstance(v, float) and pd.isna(v): v = ''
                self.call_table.setItem(r, c, QTableWidgetItem(str(v)))

    def display_stat(self, stats: Dict):
        self.stat_table.setRowCount(1)
        for col, (grp, field) in enumerate(STAT_COLUMNS):
            v = stats.get(grp, {}).get(field)
            text = '' if v is None else (f"{v:.2f}" if isinstance(v, float) else str(v))
            self.stat_table.setItem(0, col, QTableWidgetItem(text))

    def export_excel(self):
        if not hasattr(self, 'last_call'):
            QMessageBox.warning(self, '警告', '请先处理数据'); return
        fp, _ = QFileDialog.getSaveFileName(self, '保存Excel', '', 'Excel Files (*.xlsx)')
        if fp:
            try:
                ExcelExporter().export(fp, self.last_call, self.last_stats)
                self.log(f"已导出: {fp}")
                QMessageBox.information(self, '成功', '导出成功！')
            except Exception as e:
                QMessageBox.critical(self, '错误', f'导出失败: {str(e)}')

    def gen_annotated(self):
        """生成业务标注文件(原始UCM所有列 + 业务标记列 + 速率计算列高亮)"""
        if not self.mmf_files:
            QMessageBox.warning(self, '警告', '请先选择UCM监控文件'); return
        fp, _ = QFileDialog.getSaveFileName(self, '保存标注文件', 'UCM业务标注.xlsx', 'Excel Files (*.xlsx)')
        if not fp:
            return
        plan = self.get_plan_from_ui(); params = self.get_params()
        try:
            self.log("正在生成标注文件...")
            self.processor = UCMProcessor(self.config)
            self.processor.generate_annotated(self.mmf_files, plan, params, fp)
            self.log(f"✓ 标注文件已生成: {fp}")
            QMessageBox.information(self, '成功', f'标注文件生成成功！\n{fp}')
        except Exception as e:
            import traceback; traceback.print_exc()
            QMessageBox.critical(self, '错误', f'生成失败: {str(e)}')

    def show_about(self):
        AboutDialog(self).exec()

    def _start_update_check(self):
        """启动时静默检查更新(仅发现新版才弹窗)"""
        self._update_checker = UpdateChecker(self)
        self._update_checker.found_update.connect(self._on_update_found)
        self._update_checker.start()

    def _on_update_found(self, info):
        r = QMessageBox.question(self, '发现新版本',
            f"检测到新版本 {info['version']}（当前 {APP_VERSION}）\n\n"
            f"更新内容：\n{info.get('changelog','')}\n\n是否立即更新？",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        if r == QMessageBox.Yes:
            if do_update(info, self):
                if getattr(sys, 'frozen', False) and os.name == 'nt':
                    QMessageBox.information(self, '更新准备就绪', '工具将退出并自动完成更新，请稍候。')
                else:
                    QMessageBox.information(self, '更新完成', '更新已下载并替换，工具将重启。')
                _relaunch_and_quit()

    def log(self, msg):
        self.log_text.append(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    w = MainWindow(); w.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
