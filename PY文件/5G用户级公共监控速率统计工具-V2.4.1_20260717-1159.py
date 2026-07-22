#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
5G用户级公共监控速率统计工具 V2.4.1
GUI: 选择MMF文件 → 设置QCI/削峰阈值 → 生成5Sheet输出
CLI: python3 this.py --cmd
"""
import sys, os, re
import pandas as pd, numpy as np
from datetime import datetime, timedelta
from collections import Counter, defaultdict

# ===== 预编译正则（性能优化）=====
_PARSE_RE = re.compile(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})[\.\(](\d+)')
_PARSE_MS_RE = re.compile(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\.(\d+)')

def parse_time(t):
    m = _PARSE_RE.match(str(t))
    if m:
        return datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S') + timedelta(milliseconds=int(m.group(2)))
    return None

# ===== 阈值 =====
DL_ACTIVE_BIG, UL_ACTIVE_BIG = 50.0, 10.0
DL_ACTIVE_SMALL, UL_ACTIVE_SMALL = 5.0, 5.0
GAP_MERGE = 3.0
FTP_DUR_MIN, FTP_DUR_MAX, FTP_PAIR_GAP = 8.5, 13.0, 15.0
STORE_L_DL_MIN, STORE_L_DL_MAX = 8800, 10100
WX_L_DUR_MIN, WX_L_DUR_MAX, WX_L_UL_MIN = 15, 60, 800
DL_BIZ = {'FTP下载', '应用商店小文件下载', '应用商店大文件下载'}
UL_BIZ = {'FTP上传', '微信小包发送', '微信大包发送'}
REQ_BIZ = {'FTP下载', 'FTP上传', '应用商店小文件下载', '应用商店大文件下载', '微信小包发送', '微信大包发送'}


def detect_segments(ts, dl, ul, dl_th, ul_th, gap, r1=None, r2=None):
    active = (dl > dl_th) | (ul > ul_th); segs = []; i = 0; n = len(ts)
    while i < n:
        if r1 is not None and ts[i] < r1: i += 1; continue
        if r2 is not None and ts[i] > r2: break
        if active[i]:
            si = i; j = i + 1
            while j < n:
                if r2 is not None and ts[j] > r2: break
                if not active[j]:
                    k = j
                    while k < n and not active[k]:
                        if (ts[k] - ts[j - 1]) / np.timedelta64(1, 's') > gap: break
                        k += 1
                    if k < n and active[k] and (r2 is None or ts[k] <= r2): j = k; continue
                    else: break
                j += 1
            st = ts[si]; et = ts[j - 1]; dur = (et - st) / np.timedelta64(1, 's') + 1
            segs.append({'st': st, 'et': et, 'dur': dur, 'dl': float(dl[si:j].sum()), 'ul': float(ul[si:j].sum())})
            i = j
        else: i += 1
    return segs


def process(files, qci_list=None, dl_clip=1000, ul_clip=200, callback=None, cancel_check=None, progress_cb=None):
    if qci_list is None: qci_list = [5, 6, 7]
    """统一的处理函数：从MMF文件生成5Sheet输出"""
    def log(msg, pct=None):
        if cancel_check and cancel_check():
            raise KeyboardInterrupt('用户取消')
        print(msg)
        if callback: callback(msg)
        if pct is not None and progress_cb: progress_cb(pct)

    # 展开压缩文件(zip 自动解压，支持嵌套 zip-in-zip 递归解压)
    import zipfile, tempfile, shutil, glob
    _tmp_dirs = []
    def _expand_zip(zip_path):
        """递归解压 zip(含嵌套 zip)，返回所有 xlsx/xls 路径"""
        _tmp = tempfile.mkdtemp(prefix='mmf_unzip_'); _tmp_dirs.append(_tmp)
        _xs = []
        try:
            with zipfile.ZipFile(zip_path) as _zf: _zf.extractall(_tmp)
            _xs = glob.glob(os.path.join(_tmp, '**', '*.xlsx'), recursive=True) + \
                  glob.glob(os.path.join(_tmp, '**', '*.xls'), recursive=True)
            # 内嵌 zip 递归解压
            _inner_zips = glob.glob(os.path.join(_tmp, '**', '*.zip'), recursive=True)
            for _iz in _inner_zips:
                _xs.extend(_expand_zip(_iz))
        except Exception as _e:
            log(f"  解压失败 {os.path.basename(zip_path)}: {_e}")
        return _xs

    _expanded = []
    for _f in (files or []):
        if str(_f).lower().endswith('.zip'):
            _xs = _expand_zip(_f)
            log(f"  解压 {os.path.basename(_f)}: 找到 {len(_xs)} 个Excel(含嵌套)", 5)
            _expanded.extend(_xs)
        else:
            _expanded.append(_f)
    files = _expanded
    nf = len(files) if files else 1
    log(f"加载MMF文件（共 {nf} 个）...", 5)
    dfs = []
    for fi, f in enumerate(files):
        log(f"  读取文件 {fi+1}/{nf}: {os.path.basename(f)}", 5 + int(15 * (fi+1) / nf))
        # 读全部列（详细过程需展示mmf所有列）；优先用 calamine 引擎(比openpyxl快5-8倍)
        try:
            df_i = pd.read_excel(f, engine='calamine')
        except Exception:
            df_i = pd.read_excel(f)
        rlc_dl = [c for c in df_i.columns if 'RLC' in c and 'Downlink' in c]
        rlc_ul = [c for c in df_i.columns if 'RLC' in c and 'Uplink' in c]
        df_i['_t'] = df_i['Time'].apply(parse_time)
        if rlc_dl:
            df_i['_dl'] = pd.to_numeric(df_i[rlc_dl[0]], errors='coerce').fillna(0) / 1e6
        else:
            df_i['_dl'] = 0
        if rlc_ul:
            df_i['_ul'] = pd.to_numeric(df_i[rlc_ul[0]], errors='coerce').fillna(0) / 1e6
        else:
            df_i['_ul'] = 0
        dfs.append(df_i)
    df = pd.concat(dfs, ignore_index=True)
    log("  合并完成，按时间排序...", 22)
    # 按 C列(Time/_t) 升序排列，保证详细过程表按时间顺序
    df = df.sort_values('_t', kind='stable').reset_index(drop=True)

    log("QCI过滤+秒聚合...", 28)
    df7 = df[df['QCI'].isin(qci_list)].dropna(subset=['_t']).copy()
    df7['_s'] = df7['_t'].dt.floor('s')
    sec = df7.groupby('_s').agg({'_dl': 'max', '_ul': 'max'}).sort_index().reset_index()
    dl_v, ul_v, ts_v = sec['_dl'].fillna(0).values, sec['_ul'].fillna(0).values, sec['_s'].values

    log("业务识别...", 35)
    big = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_BIG, UL_ACTIVE_BIG, GAP_MERGE)

    def cb(s):
        d, dl, ul = s['dur'], s['dl'], s['ul']
        if FTP_DUR_MIN <= d <= FTP_DUR_MAX: return 'FTP候选'
        if dl > ul and STORE_L_DL_MIN <= dl <= STORE_L_DL_MAX and d > 6: return '应用商店大文件下载'
        if WX_L_DUR_MIN <= d <= WX_L_DUR_MAX and ul > dl and ul >= WX_L_UL_MIN: return '微信大包发送'
    for s in big: s['biz'] = cb(s); s['rnd'] = None

    log("FTP配对+小业务识别...", 40)
    fc = [s for s in big if s['biz'] == 'FTP候选']; paired = set(); fp = []
    for i in range(len(fc)):
        if id(fc[i]) in paired: continue
        for j in range(i + 1, len(fc)):
            if id(fc[j]) in paired: continue
            g = (fc[j]['st'] - fc[i]['et']) / np.timedelta64(1, 's')
            if 0 <= g < FTP_PAIR_GAP:
                fp.append((fc[i], fc[j])); paired.add(id(fc[i])); paired.add(id(fc[j])); break
            elif g >= FTP_PAIR_GAP: break

    def fd(s): return 'FTP下载' if s['dl'] > s['ul'] else 'FTP上传'
    for s in fc:
        if id(s) not in paired:
            if STORE_L_DL_MIN <= s['dl'] <= STORE_L_DL_MAX and s['dl'] > s['ul']: s['biz'] = '应用商店大文件下载'
            else: s['biz'] = fd(s)

    al = sorted(big, key=lambda x: x['st'])
    for s in al: s['业务'] = s['biz']
    rn = 0; pi = []
    for f1, f2 in fp:
        rn += 1; f1['rnd'] = rn; f1['业务'] = fd(f1); f2['rnd'] = rn; f2['业务'] = fd(f2)
        pi.append((rn, f2['et'], f1['st']))
    for idx, (rnd, pe, _) in enumerate(pi):
        ns = pi[idx + 1][2] if idx + 1 < len(pi) else (al[-1]['et'] + np.timedelta64(1, 's'))
        for s in al:
            if s['et'] > pe and s['st'] < ns and s['rnd'] is None and s['业务'] in ('应用商店大文件下载', '微信大包发送'):
                s['rnd'] = rnd

    tb = sorted([s for s in al if s['业务'] in ('FTP下载', 'FTP上传', '应用商店大文件下载', '微信大包发送')],
                key=lambda x: x['st'])
    P2 = {'应用商店大文件下载': '应用商店小文件下载', '微信大包发送': '微信小包发送'}
    small = []
    for idx, (rnd, pe, ps) in enumerate(pi):
        we = pi[idx + 1][2] if idx + 1 < len(pi) else (tb[-1]['et'] + np.timedelta64(1, 's'))
        bw = sorted([s for s in tb if s['st'] >= pe and s['st'] < we and s['业务'] in (
        '应用商店大文件下载', '微信大包发送')], key=lambda x: x['st'])
        prev = pe
        for bs in bw:
            st2 = P2.get(bs['业务'])
            if st2 and prev < bs['st'] and (bs['st'] - prev) / np.timedelta64(1, 's') > 1:
                gs = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_SMALL, UL_ACTIVE_SMALL, GAP_MERGE, prev, bs['st'])
                vd = [s for s in gs if s['dl'] + s['ul'] >= 10 and s['dur'] >= 0.5]
                if st2 == '应用商店小文件下载':
                    dls = [s for s in vd if s['dl'] > s['ul']]
                    if dls: bm = max(dls, key=lambda x: x['dl']); bm['业务'] = st2; bm['rnd'] = rnd; small.append(bm)
                else:
                    uls = [s for s in vd if s['ul'] > s['dl']]
                    if uls: bm = max(uls, key=lambda x: x['ul']); bm['业务'] = st2; bm['rnd'] = rnd; small.append(bm)
            prev = bs['et']

    # 检测第一个FTP对之前的不完整轮次
    if pi:
        first_ftp_start = pi[0][2]
        data_start = ts_v[0]
        if first_ftp_start > data_start + np.timedelta64(30, 's'):
            # 找第一个FTP对之前的大业务(应用商店大/微信大)
            pre_big = sorted([s for s in al if s['st'] < first_ftp_start and s['业务'] in (
                '应用商店大文件下载', '微信大包发送')], key=lambda x: x['st'])
            rng_start = data_start
            if pre_big:
                for bs in pre_big:
                    st2 = P2.get(bs['业务'])
                    if st2 and (bs['st'] - rng_start) / np.timedelta64(1, 's') > 1:
                        gs = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_SMALL, UL_ACTIVE_SMALL,
                                             GAP_MERGE, rng_start, bs['st'])
                        vd = [s for s in gs if s['dl'] + s['ul'] >= 10 and s['dur'] >= 0.5]
                        if st2 == '应用商店小文件下载':
                            dls = [s for s in vd if s['dl'] > s['ul']]
                            if dls: bm = max(dls, key=lambda x: x['dl']); bm['业务'] = st2; bm['rnd'] = 0; small.append(bm)
                        else:
                            uls = [s for s in vd if s['ul'] > s['dl']]
                            if uls: bm = max(uls, key=lambda x: x['ul']); bm['业务'] = st2; bm['rnd'] = 0; small.append(bm)
                    rng_start = bs['et']
            # 最后一个大业务(或数据开头) 到 first_ftp_start 之间
            if first_ftp_start > rng_start + np.timedelta64(1, 's'):
                gs = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_SMALL, UL_ACTIVE_SMALL,
                                     GAP_MERGE, rng_start, first_ftp_start)
                vd = [s for s in gs if s['dl'] + s['ul'] >= 10 and s['dur'] >= 0.5]
                dls = [s for s in vd if s['dl'] > s['ul']]
                if dls: bm = max(dls, key=lambda x: x['dl']); bm['业务'] = '应用商店小文件下载'; bm['rnd'] = 0; small.append(bm)
                uls = [s for s in vd if s['ul'] > s['dl']]
                if uls: bm = max(uls, key=lambda x: x['ul']); bm['业务'] = '微信小包发送'; bm['rnd'] = 0; small.append(bm)

    af = sorted(al + small, key=lambda x: x['st'])
    for s in af: s['有效'] = s['业务'] is not None and s.get('rnd') is not None
    rm = defaultdict(list)
    for s in af:
        if s['有效'] and s['业务'] and s.get('rnd') is not None: rm[s['rnd']].append(s)
    for rnd, segs in rm.items():
        seen = set(); cnt = 0
        for s in sorted(segs, key=lambda x: x['st']):
            if s['业务'] in seen or cnt >= 6: s['有效'] = False
            else: seen.add(s['业务']); cnt += 1

    valid = [s for s in af if s['有效'] and s['业务']]
    log(f"识别完成: {len(valid)}个业务段, {rn}轮")

    # ===== 标记 =====
    df = pd.concat([df, pd.DataFrame({
        '下载速率': df['_dl'], '上传速率': df['_ul'],
        '业务识别': None, '持续时长': None, '速率(Mbps)': None,
    }, index=df.index)], axis=1)

    for seg in valid:
        st, et, biz = seg['st'], seg['et'], seg['业务']
        sm = (df['_t'] >= st) & (df['_t'] <= et + timedelta(seconds=1)) & (df['QCI'].isin(qci_list))
        col = '_dl' if biz in DL_BIZ else '_ul'

        # 所有业务：只标记有流量的行(>0)
        cv_all = df.loc[sm, col].fillna(0)
        nz = cv_all > 0
        mm = sm & nz.reindex(df.index, fill_value=False)

        if not mm.any():
            continue  # 无流量则跳过

        # FTP：从第一个有流量的行开始，检查前后小流量补齐到10秒
        if biz in ('FTP下载', 'FTP上传'):
            # 找所有有流量的行
            flow_vals = df.loc[mm, col].fillna(0).tolist()
            flow_times = df.loc[mm, '_t'].tolist()

            if len(flow_vals) == 0:
                continue

            # 起点：从第一个有流量的行开始
            first_idx = 0

            # 终点：最后一个有流量的行
            last_idx = len(flow_vals) - 1

            # 计算时长
            first_t = flow_times[first_idx]
            last_t = flow_times[last_idx]
            md = (last_t - first_t).total_seconds()

            # 如果时长不是10秒左右，检查前后是否有小流量可以补齐
            # 向前找：在segment开始时间之前3秒内，找有流量的行
            if md < 9.5 or md > 10.5:
                # 向前找小流量
                ft0 = st - timedelta(seconds=3)
                pm = (df['_t'] >= ft0) & (df['_t'] < first_t) & (df['QCI'].isin(qci_list))
                if pm.any():
                    pv = df.loc[pm, col].fillna(0)
                    pv_nz = pv > 0
                    if pv_nz.any():
                        # 有前置流量，扩展起点
                        first_t = df.loc[pv[pv_nz].index[0], '_t']
                        md = (last_t - first_t).total_seconds()

                # 向后找小流量
                et0 = et + timedelta(seconds=3)
                pm2 = (df['_t'] > last_t) & (df['_t'] <= et0) & (df['QCI'].isin(qci_list))
                if pm2.any():
                    pv2 = df.loc[pm2, col].fillna(0)
                    pv2_nz = pv2 > 0
                    if pv2_nz.any():
                        # 有后置流量，扩展终点
                        last_t = df.loc[pv2[pv2_nz].index[-1], '_t']
                        md = (last_t - first_t).total_seconds()

            # 重新计算标记范围
            mm_final = (df['_t'] >= first_t) & (df['_t'] <= last_t) & (df['QCI'].isin(qci_list)) & (df[col].fillna(0) > 0)
            md = (last_t - first_t).total_seconds()

            df.loc[mm_final, '业务识别'] = biz
            df.loc[mm_final, '持续时长'] = round(md, 3)
            vs = pd.to_numeric(df.loc[mm_final, '下载速率'] if biz in DL_BIZ else df.loc[mm_final, '上传速率'], errors='coerce').fillna(0).tolist()
            ar = sum(vs) / md if md > 0 else 0
            df.loc[mm_final, '速率(Mbps)'] = round(ar, 2)
            seg['mark_st'], seg['mark_et'], seg['mark_dur'] = first_t, last_t, md
            continue

        # 其他业务：直接标记有流量的行
        first_t = df.loc[mm, '_t'].min()
        last_t = df.loc[mm, '_t'].max()
        md = (last_t - first_t).total_seconds()

        df.loc[mm, '业务识别'] = biz
        df.loc[mm, '持续时长'] = round(md, 3)
        if biz in DL_BIZ: vs = pd.to_numeric(df.loc[mm, '下载速率'], errors='coerce').fillna(0).tolist()
        else: vs = pd.to_numeric(df.loc[mm, '上传速率'], errors='coerce').fillna(0).tolist()
        ar = sum(vs) / md if md > 0 else 0
        df.loc[mm, '速率(Mbps)'] = round(ar, 2)
        seg['mark_st'], seg['mark_et'], seg['mark_dur'] = first_t, last_t, md

    rd = defaultdict(set)
    for s in valid:
        r = s.get('rnd')
        if r is not None: rd[r].add(s['业务'])
    rl = {}; ii = 0
    for r in sorted(rd.keys()):
        if rd[r] >= REQ_BIZ: rl[r] = str(r)
        else: ii += 1; rl[r] = f"不完整{ii}"
    df['轮次'] = None
    for s in valid:
        r = s.get('rnd'); lb = rl.get(r) if r is not None else ''
        if lb:
            ms2, me2 = s.get('mark_st', s['st']), s.get('mark_et', s['et'])
            # 只在有业务标记的行填轮次
            mm2 = (df['_t'] >= ms2) & (df['_t'] <= me2) & (df['QCI'].isin(qci_list)) & df['业务识别'].notna()
            df.loc[mm2, '轮次'] = lb

    # ===== 详细过程 =====
    # 先创建筛选列（A列）
    # E-N列：速率(Mbps)、持续时长、轮次、businessType、基准_业务类型、基准_开始时间、基准_结束时间、基准_速率类指标、基准_数值、基准_时长类指标、基准_数值.1
    OC = ['下载速率', '上传速率', '速率(Mbps)', '持续时长', '轮次']
    log("生成详细过程...", 48)
    df_detail = pd.DataFrame({'筛选': None, 'id': range(1, len(df) + 1), 'Time': df['Time']})
    for c in OC: df_detail[c] = df[c].values if c in df.columns else None
    df_detail['businessType'] = df['业务识别'].values

    # 加载基准数据并标记到详细过程
    log("加载基准数据...", 53)
    has_base = True
    try:
        base = pd.read_excel('工信部业务指标_呼叫详情_整理.xlsx', sheet_name='电信').sort_values('开始时间')
        if base is None or len(base) == 0:
            has_base = False
    except Exception:
        has_base = False
        base = pd.DataFrame()
    log("基准对齐...", 55)
    # 初始化基准列
    base_cols = ['基准_业务类型', '基准_开始时间', '基准_结束时间', '基准_速率类指标', '基准_数值', '基准_时长类指标', '基准_数值.1']
    for bc in base_cols:
        df_detail[bc] = None

    # 对齐基准数据到详细过程（向量化：预取整列，避免逐行 df.loc，提速约50倍）
    ts = df['_t']
    ts_valid = ts[ts.notna()]
    for _, brow in base.iterrows():
        b_start = pd.to_datetime(brow['开始时间'])
        b_end = pd.to_datetime(brow['结束时间'])
        b_biz = str(brow['业务类型'])
        if len(ts_valid) == 0:
            break
        # 一次向量化求 b_start 最近行
        diff_s = (ts_valid - b_start).abs()
        best_start_idx = diff_s.idxmin()
        if diff_s.loc[best_start_idx].total_seconds() >= 5:
            continue
        # 从 best_start_idx 起向量化求 b_end 最近行
        sub = ts_valid[ts_valid.index >= best_start_idx]
        best_end_idx = sub.index[(sub - b_end).abs().argmin()] if len(sub) > 0 else best_start_idx
        # 批量切片赋值（替代逐行 loc）
        lo, hi = best_start_idx, min(best_end_idx, len(df_detail) - 1)
        if lo <= hi:
            df_detail.loc[lo:hi, '基准_业务类型'] = b_biz
            df_detail.loc[lo:hi, '基准_开始时间'] = str(brow['开始时间'])
            df_detail.loc[lo:hi, '基准_结束时间'] = str(brow['结束时间'])
            df_detail.loc[lo:hi, '基准_速率类指标'] = str(brow['速率类指标'])
            df_detail.loc[lo:hi, '基准_数值'] = brow['数值']
            df_detail.loc[lo:hi, '基准_时长类指标'] = str(brow['时长类指标'])
            df_detail.loc[lo:hi, '基准_数值.1'] = brow['数值.1']

    # 添加_t列
    df_detail['_t'] = df['_t'].values

    # 将QCI、Downlink RLC、Uplink RLC三列移到_t之后
    special_cols = ['QCI', 'Downlink RLC Throughput(bps)', 'Uplink RLC Throughput(bps)']
    for sc in special_cols:
        if sc in df.columns:
            df_detail[sc] = df[sc].values

    # 添加其他列
    for c in df.columns:
        if c not in df_detail.columns and c not in ('id', '_t', '_dl', '_ul', '_s', 'businessType', '业务识别'):
            df_detail[c] = df[c].values

    # 计算筛选列（向量化）：check_cols 任一非空则为1
    check_cols = ['速率(Mbps)', '持续时长', '轮次', 'businessType',
                  '基准_业务类型', '基准_开始时间', '基准_结束时间', '基准_速率类指标', '基准_数值', '基准_时长类指标', '基准_数值.1']
    has = pd.Series(False, index=df_detail.index)
    for col in check_cols:
        s = df_detail[col]
        ss = s.astype(str).str.strip()
        has = has | (s.notna() & (ss != '') & (ss.str.lower() != 'nan'))
    df_detail['筛选'] = has.astype(int).values

    # ===== 汇总 =====
    sr = []
    sum_raw_data = {}  # 用于批注：{biz: {rates, durs, col_name, r_start, r_end, sub}}
    for biz in sorted(DL_BIZ | UL_BIZ):
        sub = df_detail[df_detail['businessType'] == biz].dropna(subset=['_t']).sort_values('_t')
        if len(sub) == 0: continue
        dl_cn = [c for c in sub.columns if 'RLC' in c and 'Downlink' in c]
        ul_cn = [c for c in sub.columns if 'RLC' in c and 'Uplink' in c]
        if biz in DL_BIZ and dl_cn:
            col_name = dl_cn[0]
            rates_raw = pd.to_numeric(sub[col_name], errors='coerce').fillna(0) / 1e6
            rates = rates_raw[rates_raw > 0]
        elif biz in UL_BIZ and ul_cn:
            col_name = ul_cn[0]
            rates_raw = pd.to_numeric(sub[col_name], errors='coerce').fillna(0) / 1e6
            rates = rates_raw[rates_raw > 0]
        else:
            rates = pd.Series([], dtype=float); col_name = ''
        # 时长
        ts2 = sub['_t']; durs = []
        if len(ts2) > 0:
            s1 = ts2.iloc[0]; p = s1
            for t in ts2.iloc[1:]:
                if (t - p).total_seconds() > 5: durs.append((p - s1).total_seconds() + 1); s1 = t
                p = t
            durs.append((p - s1).total_seconds() + 1)
        # 记录原始数据位置（在详细过程df中的行号）
        r_start = sub.index[0] + 2  # +2 因为Excel第1行表头, 第2行索引0
        r_end = sub.index[-1] + 2
        sum_raw_data[biz] = {'rates': rates, 'durs': durs, 'col_name': col_name,
                             'r_start': r_start, 'r_end': r_end, 'sub': sub}
        row = {'业务类型': biz}
        if biz in DL_BIZ and len(rates) > 0:
            row['应用层FTP下载速率1000M以上占比(%)'] = round((rates > dl_clip).sum() / len(rates) * 100, 2)
            row['应用层FTP下载速率100M以下占比(%)'] = round((rates < 100).sum() / len(rates) * 100, 2)
            row['应用层平均下载速率(Mbps)'] = round(rates.mean(), 2)
            row['削峰应用层平均下载速率(Mbps)'] = round(rates.clip(upper=dl_clip).mean(), 2)
            sv = rates.sort_values(ascending=False); tn = max(1, int(len(sv) * 0.1))
            row['下行削峰TOP10%峰值速率'] = round(sv.head(tn).mean(), 2)
        elif biz in UL_BIZ and len(rates) > 0:
            row['应用层FTP上传速率200M以上占比(%)'] = round((rates > ul_clip).sum() / len(rates) * 100, 2)
            row['应用层FTP上传速率5M以下占比(%)'] = round((rates < 5).sum() / len(rates) * 100, 2)
            row['应用层平均上传速率'] = round(rates.mean(), 2)
            row['削峰应用层平均上传速率(Mbps)'] = round(rates.clip(upper=ul_clip).mean(), 2)
            sv = rates.sort_values(ascending=False); tn = max(1, int(len(sv) * 0.1))
            row['上行削峰TOP10%峰值速率'] = round(sv.head(tn).mean(), 2)
        if durs:
            row['业务时长平均值(s)'] = round(np.mean(durs), 2)
            row['业务时长中位值(s)'] = round(np.median(durs), 0)
        sr.append(row)
    df_summary = pd.DataFrame(sr)

    # ===== 汇总批注 + D/E列颜色 =====
    # 收集每业务的原始数据，用于批注
    raw_data = {}
    for biz in sorted(DL_BIZ | UL_BIZ):
        sub = df_detail[df_detail['businessType'] == biz].dropna(subset=['_t']).sort_values('_t')
        if len(sub) == 0: continue
        dl_cn = [c for c in sub.columns if 'RLC' in c and 'Downlink' in c]
        ul_cn = [c for c in sub.columns if 'RLC' in c and 'Uplink' in c]
        if biz in DL_BIZ and dl_cn:
            rates = pd.to_numeric(sub[dl_cn[0]], errors='coerce').fillna(0) / 1e6
            raw_data[biz] = {'rates': rates, 'durs': [], 'col': dl_cn[0], 'rows': (sub.index.min(), sub.index.max())}
        elif biz in UL_BIZ and ul_cn:
            rates = pd.to_numeric(sub[ul_cn[0]], errors='coerce').fillna(0) / 1e6
            raw_data[biz] = {'rates': rates, 'durs': [], 'col': ul_cn[0], 'rows': (sub.index.min(), sub.index.max())}
        ts2 = sub['_t']; durs = []
        if len(ts2) > 0:
            s1 = ts2.iloc[0]; p = s1
            for t in ts2.iloc[1:]:
                if (t - p).total_seconds() > 5: durs.append((p - s1).total_seconds() + 1); s1 = t
                p = t
            durs.append((p - s1).total_seconds() + 1)
        if biz in raw_data:
            raw_data[biz]['durs'] = durs

    # 等 ExcelWriter 写入完毕后，对汇总 sheet 加批注
    # 标注在 ExcelWriter 的 with 块内完成

    # ===== 对比横表（基准驱动，逐行匹配）=====
    BIZ_ORDER = ['FTP下载', 'FTP上传', '应用商店小文件下载', '应用商店大文件下载', '微信小包发送', '微信大包发送']
    rate_nm = {'FTP下载': '下载平均速率(Mbps)', 'FTP上传': '上传平均速率(Mbps)',
               '应用商店小文件下载': '小文件下载速率(Mbps)', '应用商店大文件下载': '大文件下载速率(Mbps)',
               '微信小包发送': '微信小包发送速率(Mbps)', '微信大包发送': '微信大包发送速率(Mbps)'}

    def fmt_time(t_str):
        m = _PARSE_RE.match(str(t_str))
        if m:
            dt = datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S')
            return f"{m.group(1)}({m.group(2)})", dt
        # 尝试毫秒格式 .xxx
        m2 = _PARSE_MS_RE.match(str(t_str))
        if m2:
            dt = datetime.strptime(m2.group(1), '%Y-%m-%d %H:%M:%S')
            return f"{m2.group(1)}({m2.group(2)[:3]})", dt
        return str(t_str), None

    def fmt_base_time(t):
        """基准时间格式化为 2026-07-02 22:27:36.557 样式(点+3位毫秒)，方便CTRL+F对照基准文件"""
        if t is None: return ''
        try:
            ts = pd.Timestamp(t)
            if pd.isna(ts): return str(t)
            return ts.strftime('%Y-%m-%d %H:%M:%S.') + f"{ts.microsecond // 1000:03d}"
        except Exception:
            return str(t)

    # 给df_detail添加_t列便于时间搜索
    df_detail['_t'] = df['_t'].values

    # 基准已在上方(详细过程模块)读取并按开始时间排序，此处直接复用，避免重复读盘
    base_sorted = base.reset_index(drop=True)

    # 按业务序列切轮：每轮以 FTP下载 开头(标准顺序
    # FTP下载→FTP上传→应用商店小文件→应用商店大文件→微信小包→微信大包)。
    # 第一个 FTP下载 之前的业务归为开头不完整轮次；末尾不足6个为末尾不完整轮次。
    # (替代 V2.0.9 的"每6条硬切一组"——硬切遇到开头/末尾不完整轮次会整体错位)
    base_rounds = []
    cur = []
    for i in range(len(base_sorted)):
        biz = str(base_sorted.iloc[i]['业务类型'])
        if biz == 'FTP下载' and cur:
            base_rounds.append(cur); cur = []
        cur.append(base_sorted.iloc[i])
    if cur:
        base_rounds.append(cur)

    # 逐轮逐业务匹配(基准驱动)。判定"对得上"不用固定时间窗口，而看业务连续段：
    #   对基准轮内业务 Bk，匹配区间由相邻业务边界决定——
    #     下界 = 上一个基准业务结束时间(轮首无下界)
    #     上界 = 下一个基准业务开始时间(轮末用自身结束时间)
    #   代码侧同业务段只要开始时间落在该区间且未被占用即算对得上；
    #   整轮6业务完整且全部对上才编号，否则该轮标"不计轮次"。
    all_items = []
    rnd_counter = 0
    used = set()  # 已占用的 df_detail 行索引，避免一代码段被多轮重复匹配
    for rnd_bizs in base_rounds:
        n = len(rnd_bizs)
        all_matched = True
        grp_items = []

        for k in range(n):
            brow = rnd_bizs[k]
            b_biz = str(brow['业务类型'])
            b_start_str = str(brow['开始时间'])
            b_end_str = str(brow['结束时间'])
            b_start = pd.to_datetime(brow['开始时间'])
            b_end = pd.to_datetime(brow['结束时间'])

            # 就近匹配(B方案)：在所有未占用的同业务行里，找离 b_start 最近的连续段
            # 距离<=300秒(约一轮间隔)才算同轮，避免跨轮串；used 防一段被多轮重复占用
            cand = df_detail[(df_detail['businessType'] == b_biz) &
                             (~df_detail.index.isin(used))]
            code_data = None
            if len(cand) > 0:
                dist = (cand['_t'] - b_start).abs()
                best_idx = dist.idxmin()
                if dist.loc[best_idx] <= pd.Timedelta(seconds=300):
                    # 取 best_idx 所在的连续行块(行号连续的同业务未占用行)
                    cand_idx_sorted = cand.index.sort_values()
                    pos = cand_idx_sorted.get_loc(best_idx)
                    lo = pos
                    while lo - 1 >= 0 and cand_idx_sorted[lo - 1] == cand_idx_sorted[lo] - 1:
                        lo -= 1
                    hi = pos
                    while hi + 1 < len(cand_idx_sorted) and cand_idx_sorted[hi + 1] == cand_idx_sorted[hi] + 1:
                        hi += 1
                    seg_idxs = list(cand_idx_sorted[lo:hi + 1])
                    used.update(seg_idxs)
                    cand_s = df_detail.loc[seg_idxs].sort_values('_t')
                    first = cand_s.iloc[0]
                    last = cand_s.iloc[-1]

                    c_start = str(first['Time'])
                    c_end = str(last['Time'])
                    c_start_fmt, _ = fmt_time(c_start)
                    c_end_fmt, _ = fmt_time(c_end)

                    code_data = {
                        '开始时间': c_start_fmt if c_start_fmt else c_start,
                        '结束时间': c_end_fmt if c_end_fmt else c_end,
                        '速率类指标': rate_nm.get(b_biz, ''),
                        '数值': first.get('速率(Mbps)'),
                        '时长类指标': '业务时长(秒)',
                        '数值.1': first.get('持续时长'),
                    }
                else:
                    all_matched = False
            else:
                all_matched = False

            # 基准数据(开始/结束时间用 .557 点样式，与基准文件一致便于CTRL+F)
            base_data = {
                '开始时间': fmt_base_time(brow['开始时间']),
                '结束时间': fmt_base_time(brow['结束时间']),
                '速率类指标': str(brow.get('速率类指标', '')),
                '数值': brow.get('数值'),
                '时长类指标': str(brow.get('时长类指标', '')),
                '数值.1': brow.get('数值.1'),
            }

            grp_items.append({
                'biz': b_biz,
                '基准': base_data,
                '代码': code_data,
            })

        complete = (n == 6)
        rnd_label = str(rnd_counter + 1) if (complete and all_matched) else '不计轮次'
        if complete and all_matched:
            rnd_counter += 1
        for item in grp_items: item['rnd_label'] = rnd_label
        all_items.extend(grp_items)

    # 偏差计算辅助函数
    def _pct(cv, bv):
        try:
            if pd.isna(cv) or pd.isna(bv): return '-'
            cf = float(cv); bf = float(bv)
            return '-' if bf == 0 else round((cf - bf) / bf * 100, 2)
        except Exception:
            return '-'

    def _time_diff(c_time, b_time):
        try:
            m = _PARSE_RE.match(str(c_time)); n = _PARSE_RE.match(str(b_time))
            if m and n:
                cdt = datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S') + timedelta(milliseconds=int(m.group(2)))
                bdt = datetime.strptime(n.group(1), '%Y-%m-%d %H:%M:%S') + timedelta(milliseconds=int(n.group(2)))
                return round((cdt - bdt).total_seconds(), 2)
            return '-'
        except Exception:
            return '-'

    # 构建df_compare：每业务2行(基准+代码)，偏差在右侧3列横铺
    cr = []
    for item in all_items:
        rnd = item['rnd_label']
        biz = item['biz']
        bd = item['基准']
        cd = item['代码']
        rpct = _pct(cd['数值'], bd['数值']) if cd else '-'
        dpct = _pct(cd['数值.1'], bd['数值.1']) if cd else '-'
        tdiff = _time_diff(cd['开始时间'], bd['开始时间']) if cd else '-'
        # 基准行（偏差列有值）
        cr.append({'轮次': rnd, '业务类型': biz, '来源': '基准',
                   '开始时间': bd['开始时间'], '结束时间': bd['结束时间'],
                   '速率类指标': bd['速率类指标'], '数值': bd['数值'],
                   '时长类指标': bd['时长类指标'], '数值.1': bd['数值.1'],
                   '速率偏差(%)': rpct, '时长偏差(%)': dpct, '时间差(秒)': tdiff})
        # 代码行（偏差列留空，由合并单元格覆盖）
        if cd:
            cr.append({'轮次': rnd, '业务类型': biz, '来源': '代码',
                       '开始时间': cd['开始时间'], '结束时间': cd['结束时间'],
                       '速率类指标': cd['速率类指标'], '数值': cd['数值'],
                       '时长类指标': cd['时长类指标'], '数值.1': cd['数值.1'],
                       '速率偏差(%)': '', '时长偏差(%)': '', '时间差(秒)': ''})
        else:
            cr.append({'轮次': rnd, '业务类型': biz, '来源': '代码',
                       '开始时间': '缺失', '结束时间': '缺失',
                       '速率类指标': '缺失', '数值': '缺失',
                       '时长类指标': '缺失', '数值.1': '缺失',
                       '速率偏差(%)': '', '时长偏差(%)': '', '时间差(秒)': ''})

    df_compare = pd.DataFrame(cr).reset_index(drop=True)

    # ===== 汇总：所有业务 =====
    def _calc_summary(sub_detail, dl_clip, ul_clip):
        """从 df_detail 子集计算汇总行"""
        sr = []
        for biz in sorted(DL_BIZ | UL_BIZ):
            sub = sub_detail[sub_detail['businessType'] == biz].dropna(subset=['_t']).sort_values('_t')
            if len(sub) == 0: continue
            dl_cn = [c for c in sub.columns if 'RLC' in c and 'Downlink' in c]
            ul_cn = [c for c in sub.columns if 'RLC' in c and 'Uplink' in c]
            if biz in DL_BIZ and dl_cn:
                cn = dl_cn[0]; rates = pd.to_numeric(sub[cn], errors='coerce').fillna(0) / 1e6; rates = rates[rates > 0]
            elif biz in UL_BIZ and ul_cn:
                cn = ul_cn[0]; rates = pd.to_numeric(sub[cn], errors='coerce').fillna(0) / 1e6; rates = rates[rates > 0]
            else: rates = pd.Series([], dtype=float); cn = ''
            ts2 = sub['_t']; durs = []
            if len(ts2) > 0:
                s1 = ts2.iloc[0]; p = s1
                for t in ts2.iloc[1:]:
                    if (t - p).total_seconds() > 5: durs.append((p - s1).total_seconds() + 1); s1 = t
                    p = t
                durs.append((p - s1).total_seconds() + 1)
            row = {'业务类型': biz}
            is_dl = biz in DL_BIZ; lim = dl_clip if is_dl else ul_clip; lo = 100 if is_dl else 5
            if len(rates) > 0:
                row['阈值以上占比(%)'] = round((rates > lim).sum() / len(rates) * 100, 2)
                row['5M/100M以下占比(%)'] = round((rates < lo).sum() / len(rates) * 100, 2)
                row['平均速率(Mbps)'] = round(rates.mean(), 2)
                row['削峰平均速率(Mbps)'] = round(rates.clip(upper=lim).mean(), 2)
                sv = rates.sort_values(ascending=False); tn = max(1, int(len(sv) * 0.1))
                row['削峰TOP10%峰值速率'] = round(sv.head(tn).mean(), 2)
            else:
                for k in ['阈值以上占比(%)','5M/100M以下占比(%)','平均速率(Mbps)','削峰平均速率(Mbps)','削峰TOP10%峰值速率']:
                    row[k] = '-'
            if durs:
                row['业务时长平均值(s)'] = round(np.mean(durs), 2)
                row['业务时长中位值(s)'] = round(np.median(durs), 0)
            sr.append(row)
        return pd.DataFrame(sr)

    log("生成汇总...", 70)
    df_summary_all = _calc_summary(df_detail, dl_clip, ul_clip)
    complete_mask = df_detail['轮次'].astype(str).str.isdigit()
    df_detail_complete = df_detail[complete_mask]
    df_summary_complete = _calc_summary(df_detail_complete, dl_clip, ul_clip)

    # 生成输出文件名
    version = 'V2.4.1'
    timestamp = datetime.now().strftime('%Y%m%d-%H%M')
    out = f'联通/电信最终输出_{version}_{timestamp}.xlsx'

    log("写入详细过程...", 78)
    with pd.ExcelWriter(out, engine='openpyxl') as w:
        df_detail.to_excel(w, sheet_name='详细过程', index=False)

        from openpyxl.styles import Alignment, PatternFill, Border, Side, Font as OpFont
        from openpyxl.utils import get_column_letter as _gcl
        from openpyxl.comments import Comment
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # ===== 详细过程格式 =====
        ws = w.sheets['详细过程']
        # D/E列：无流量填"-"，有值设0.00格式；D="-"的行 A列设0(随A=1筛选隐藏)
        for row_idx in range(2, ws.max_row + 1):
            if cancel_check and row_idx % 5000 == 0 and cancel_check():
                raise KeyboardInterrupt('用户取消')
            d_is_dash = False
            for col_letter in ['D', 'E']:
                cell = ws[f'{col_letter}{row_idx}']
                v = cell.value
                if v is None or (isinstance(v, (int, float)) and v == 0):
                    orig = df_detail.iloc[row_idx - 2, (4 if col_letter == 'D' else 5) - 1] if row_idx - 2 < len(df_detail) else None
                    if orig is None or (isinstance(orig, (int, float)) and orig == 0):
                        cell.value = '-'
                        if col_letter == 'D': d_is_dash = True
                    else:
                        cell.number_format = '0.00'
                else:
                    cell.number_format = '0.00'
            if d_is_dash:
                ws.cell(row=row_idx, column=1).value = 0
        for col_idx in range(1, ws.max_column + 1):
            mx = len(str(ws.cell(row=1, column=col_idx).value or ''))
            for r in range(2, min(102, ws.max_row + 1)):
                cv = ws.cell(row=r, column=col_idx).value
                if cv is not None: mx = max(mx, min(len(str(cv)), 30))
            ws.column_dimensions[_gcl(col_idx)].width = max(mx + 2, 10)
        for cell in ws[1]: cell.alignment = Alignment(wrap_text=True)
        ws.cell(row=1, column=4).value = '下载速率（Mbps）'
        ws.cell(row=1, column=5).value = '上传速率(Mbps)'
        # 冻结到第1行 I列(A~H列+表头冻结)
        ws.freeze_panes = 'I2'
        from openpyxl.worksheet.filters import FilterColumn
        ws.auto_filter.ref = ws.dimensions
        ws.auto_filter.add_filter_column(0, ['1'])
        # 无基准时隐藏 J~Q 列(基准7列 J~P + _t Q)
        if not has_base:
            for col_idx in range(10, 18):  # J=10 ~ Q=17
                ws.column_dimensions[_gcl(col_idx)].hidden = True
        # D/E 列颜色 (F~I条件)
        fill_d = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
        fill_e = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        for col_letter, fill in [('D', fill_d), ('E', fill_e)]:
            for cell in ws[col_letter]:
                if cell.row <= 1: continue
                if cell.value is None or str(cell.value) in ('nan','None','-',''): continue
                r = cell.row
                if any(ws.cell(row=r,column=c).value is not None and str(ws.cell(row=r,column=c).value) not in ('nan','None','')
                       for c in range(6,10)):
                    cell.fill = fill

        # ===== 汇总格式 =====
        log("写入汇总...", 85)
        if cancel_check and cancel_check(): raise KeyboardInterrupt('用户取消')
        wb = w.book
        ws_sum = wb.create_sheet('汇总')
        hdr_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        # 表头(合并格式)
        hdrs = ['业务类型',
                '应用层FTP上传/下载速率\n阈值以上占比(%)',
                '应用层FTP上传/下载速率\n5M/100M以下占比(%)',
                '应用层平均上传/下载\n速率(Mbps)',
                '削峰应用层平均上传/下载\n速率(Mbps)',
                '上行/下行削峰TOP10%\n峰值速率',
                '业务时长平均值(s)',
                '业务时长中位值(s)']
        for ci, h in enumerate(hdrs, 1):
            c = ws_sum.cell(row=1, column=ci, value=h); c.alignment = center; c.fill = hdr_fill
            c.font = OpFont(bold=True); c.border = border
        for ci in range(1, 9):
            ws_sum.column_dimensions[_gcl(ci)].width = 20

        # "所有业务"标题行
        ws_sum.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        tc = ws_sum.cell(row=2, column=1, value='▼ 所有识别业务')
        tc.alignment = center; tc.font = OpFont(bold=True, size=11)
        tc.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        for ci in range(1, 9): ws_sum.cell(row=2, column=ci).border = border

        # 汇总数据行(从row3开始)
        for ri in range(len(df_summary_all)):
            for ci in range(1, 9):
                cv = df_summary_all.iloc[ri, ci - 1] if ci - 1 < len(df_summary_all.columns) else None
                cell = ws_sum.cell(row=ri + 3, column=ci, value=cv if not (isinstance(cv, float) and pd.isna(cv)) else '')
                cell.alignment = center; cell.border = border
        all_data_rows = len(df_summary_all)

        # "仅完整轮次"标题行
        comp_start = all_data_rows + 4
        ws_sum.merge_cells(start_row=comp_start, start_column=1, end_row=comp_start, end_column=8)
        tc2 = ws_sum.cell(row=comp_start, column=1, value='▼ 仅完整轮次（代码侧6业务齐全）')
        tc2.alignment = center; tc2.font = OpFont(bold=True, size=11)
        tc2.fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
        for ci in range(1, 9): ws_sum.cell(row=comp_start, column=ci).border = border

        for ri in range(len(df_summary_complete)):
            for ci in range(1, 9):
                cv = df_summary_complete.iloc[ri, ci - 1] if ci - 1 < len(df_summary_complete.columns) else None
                cell = ws_sum.cell(row=ri + comp_start + 1, column=ci, value=cv if not (isinstance(cv, float) and pd.isna(cv)) else '')
                cell.alignment = center; cell.border = border

        # 自适应行高 + 冻结
        for r in range(1, comp_start + all_data_rows + 2):
            ws_sum.row_dimensions[r].height = 22
        ws_sum.freeze_panes = 'A3'

        # ===== 汇总批注 + 超链接 =====
        def _fmt_vals(vals, prec=2):
            if vals is None or len(vals) == 0: return '[]'
            vs = [round(v, prec) if prec > 0 else int(round(v)) for v in vals]
            if len(vs) <= 20:
                return '/'.join(str(v) for v in vs)
            return '/'.join(str(v) for v in vs[:5]) + '/.../' + '/'.join(str(v) for v in vs[-5:])

        def _add_comments(sub_detail, row_offset):
            for bi, biz in enumerate(sorted(DL_BIZ | UL_BIZ)):
                sub = sub_detail[sub_detail['businessType'] == biz].dropna(subset=['_t']).sort_values('_t')
                if len(sub) == 0: continue
                dl_cn = [c for c in sub.columns if 'RLC' in c and 'Downlink' in c]
                ul_cn = [c for c in sub.columns if 'RLC' in c and 'Uplink' in c]
                if biz in DL_BIZ and dl_cn: cn = dl_cn[0]
                elif biz in UL_BIZ and ul_cn: cn = ul_cn[0]
                else: continue
                rates = pd.to_numeric(sub[cn], errors='coerce').fillna(0) / 1e6; rates = rates[rates > 0]
                ts2 = sub['_t']; durs = []
                if len(ts2) > 0:
                    s1 = ts2.iloc[0]; p = s1
                    for t in ts2.iloc[1:]:
                        if (t - p).total_seconds() > 5: durs.append((p - s1).total_seconds() + 1); s1 = t
                        p = t
                    durs.append((p - s1).total_seconds() + 1)
                is_dl = biz in DL_BIZ; lim = dl_clip if is_dl else ul_clip; lo = 100 if is_dl else 5
                n = len(rates)
                sr = sub.index; r_start = sr[0] + 2; r_end = sr[-1] + 2
                row_idx = bi + row_offset

                for cell_col in range(2, 8):
                    cell = ws_sum.cell(row=row_idx, column=cell_col)
                    cv = cell.value
                    if cv is None or str(cv) in ('nan','None','-',''): continue
                    cp = cell_col - 2
                    if cp in (0, 1):  # above/below pct → 计数
                        if cp == 0:
                            mol = [r for r in rates if r > lim]
                            cnt_label = f"计数(rate>{lim})={len(mol)}"
                            op_cn = f"COUNT(rate > {lim}) / COUNT(all) x 100\n= 统计速率大于{lim}的采样数 / 总采样数 x 100"
                        else:
                            mol = [r for r in rates if r < lo]
                            cnt_label = f"计数(rate<{lo})={len(mol)}"
                            op_cn = f"COUNT(rate < {lo}) / COUNT(all) x 100\n= 统计速率小于{lo}的采样数 / 总采样数 x 100"
                        calc = f"= {len(mol)} / {n} x 100"
                    elif cp in (2, 3, 4):  # avg/clip/top10 → 求和
                        if cp == 2:
                            mol = list(rates); sum_v = sum(rates)
                            sum_label = f"求和={sum_v:.1f}"; cnt_label = f"计数={n}"
                            op_cn = "SUM(rate) / COUNT(rate)\n= 各速率求和 / 总采样数"
                            calc = f"= {sum_v:.1f} / {n}"
                        elif cp == 3:
                            clipped = [min(r, lim) for r in rates]; sum_v = sum(clipped); mol = clipped
                            sum_label = f"求和(MIN≤{lim})={sum_v:.1f}"; cnt_label = f"计数={n}"
                            op_cn = f"SUM(MIN(rate,{lim})) / COUNT(rate)\n= 各速率取{lim}上限后求和 / 总采样数"
                            calc = f"= {sum_v:.1f} / {n}"
                        elif cp == 4:
                            sr_rates = sorted(rates, reverse=True)
                            nt = max(1, int(len(sr_rates) * 0.1))
                            top_vals = sr_rates[:nt]; sum_v = sum(top_vals); mol = top_vals
                            sum_label = f"求和(TOP10%)={sum_v:.1f}"; cnt_label = f"计数(TOP10%)={nt}"
                            op_cn = "AVERAGE(top 10%)\n= 前10%最大速率的平均值"
                            calc = f"= {sum_v:.1f} / {nt}"
                    elif cp == 5:
                        continue
                    else:
                        continue
                    col_letter = _gcl(sub.columns.get_loc(cn) + 1)
                    txt = (f"{op_cn}\n{calc}\n= {cv}\n\n"
                           f"分子: {sum_label if cp in (2,3,4) else cnt_label}\n"
                           f"分子值: {_fmt_vals(mol)}\n"
                           f"分母: {cnt_label if cp in (2,3,4) else f'计数={n}'}\n"
                           f"分母值: {_fmt_vals(list(rates))}\n\n"
                           f"分子来源: {cn} 列, 第 {r_start}~{r_end} 行\n"
                           f"分母来源: {cn} 列, 第 {r_start}~{r_end} 行")
                    c = Comment(txt, "Tool"); cell.comment = c; c.width = 500; c.height = 220
                    cell.hyperlink = f"#'详细过程'!{col_letter}{r_start}"

                # G/H 列：业务时长批注
                dur_vals = [d for d in durs if d > 0]
                if not dur_vals: continue
                for dur_col, dur_label in [(7, 'AVERAGE'), (8, 'MEDIAN')]:
                    cell = ws_sum.cell(row=row_idx, column=dur_col)
                    cv = cell.value
                    if cv is None or str(cv) in ('nan','None','-'): continue
                    if dur_col == 7:
                        dur_v = round(np.mean(dur_vals), 2)
                        txt = (f"AVERAGE(duration)\n= 各业务段时间求和 / 总段数\n"
                               f"求和={sum(dur_vals):.1f} / 计数={len(dur_vals)}\n= {cv}\n\n"
                               f"值: {_fmt_vals(dur_vals)}\n\n"
                               f"来源: 业务时长(秒) 列, 第 {r_start}~{r_end} 行")
                    else:
                        dur_v = round(np.median(dur_vals), 0)
                        txt = (f"MEDIAN(duration)\n= 业务段时间中位数\n"
                               f"计数={len(dur_vals)}\n= {cv}\n\n"
                               f"值: {_fmt_vals(dur_vals)}\n\n"
                               f"来源: 业务时长(秒) 列, 第 {r_start}~{r_end} 行")
                    c = Comment(txt, "Tool"); cell.comment = c; c.width = 420; c.height = 180

        # 上表批注
        _add_comments(df_detail, 3)
        # 下表批注
        _add_comments(df_detail_complete, comp_start + 1)

        # ===== 对比横表 =====
        log("写入对比横表...", 93)
        if cancel_check and cancel_check(): raise KeyboardInterrupt('用户取消')
        wb = w.book
        ws_h = wb.create_sheet('对比横表')
        hdr = ['轮次', '业务类型', '来源', '开始时间', '结束时间', '速率类指标', '数值', '时长类指标', '数值.1',
               '速率偏差(%)', '时长偏差(%)', '时间差(秒)']
        for ci, h in enumerate(hdr, 1):
            c = ws_h.cell(row=1, column=ci, value=h)
            c.alignment = center; c.fill = hdr_fill; c.font = OpFont(bold=True); c.border = border

        color1 = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        color2 = PatternFill(start_color='E5FFCC', end_color='E5FFCC', fill_type='solid')
        color_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        biz_color_map = {bn: (color1 if i % 2 == 0 else color2) for i, bn in enumerate(BIZ_ORDER)}

        hr = 2; rnd_start_row = 2; prev_rnd = None
        for ci in range(0, len(df_compare)):
            row = df_compare.iloc[ci]
            rnd = str(row['轮次']); biz = row['业务类型']; src = str(row['来源'])

            # 轮次切换：合并上一轮的轮次列
            if rnd != prev_rnd and prev_rnd is not None:
                ws_h.merge_cells(start_row=rnd_start_row, start_column=1, end_row=hr - 1, end_column=1)
                ws_h.cell(row=rnd_start_row, column=1).alignment = center
                rnd_start_row = hr
            prev_rnd = rnd

            if '缺失' in str(row.get('开始时间', '')): fill = color_yellow
            else: fill = biz_color_map.get(biz, color1)

            vals = [rnd if hr == rnd_start_row else '', biz, src,
                    str(row['开始时间']) if pd.notna(row.get('开始时间')) else '',
                    str(row['结束时间']) if pd.notna(row.get('结束时间')) else '',
                    str(row['速率类指标']) if pd.notna(row.get('速率类指标')) else '',
                    row['数值'] if pd.notna(row.get('数值')) else '',
                    str(row['时长类指标']) if pd.notna(row.get('时长类指标')) else '',
                    row['数值.1'] if pd.notna(row.get('数值.1')) else '',
                    row['速率偏差(%)'] if pd.notna(row.get('速率偏差(%)')) else '',
                    row['时长偏差(%)'] if pd.notna(row.get('时长偏差(%)')) else '',
                    row['时间差(秒)'] if pd.notna(row.get('时间差(秒)')) else '']
            for ri, v in enumerate(vals):
                c = ws_h.cell(row=hr, column=ri + 1, value=v if v != '' else '')
                c.alignment = center; c.fill = fill; c.border = border
            hr += 1
        if prev_rnd is not None and hr > rnd_start_row:
            ws_h.merge_cells(start_row=rnd_start_row, start_column=1, end_row=hr - 1, end_column=1)
            ws_h.cell(row=rnd_start_row, column=1).alignment = center

        # 每2行合并：业务类型(2) + 偏差列(10,11,12)
        for r in range(2, hr, 2):
            if r + 1 > hr - 1: break
            for mc in [2, 10, 11, 12]:
                ws_h.merge_cells(start_row=r, start_column=mc, end_row=r + 1, end_column=mc)
                ws_h.cell(row=r, column=mc).alignment = center

        for cidx in range(1, len(hdr) + 1):
            mx = 0
            for ridx in range(1, min(hr, 120)):
                cv = ws_h.cell(row=ridx, column=cidx).value
                if cv is not None: mx = max(mx, min(len(str(cv)), 30))
            ws_h.column_dimensions[_gcl(cidx)].width = max(mx + 2, 10)
        ws_h.freeze_panes = 'C2'

    log(f"完成! {out}", 100)
    # 清理解压临时目录
    for _d in _tmp_dirs:
        try: shutil.rmtree(_d)
        except Exception: pass
    return out


# ===== GUI 包装 =====
try:
    from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QPushButton, QLabel, QFileDialog, QMessageBox, QProgressBar, QTextEdit,
        QGroupBox, QCheckBox, QSpinBox, QDialog, QSplitter, QGridLayout)
    from PySide6.QtCore import Qt, QThread, Signal, Slot
    from PySide6.QtGui import QFont
    GUI_OK = True
except:
    GUI_OK = False


if GUI_OK:
    class WorkerThread(QThread):
        progress = Signal(str)
        progress_pct = Signal(int)
        done = Signal(str)

        def __init__(self, files, qci_list, dl_clip, ul_clip):
            super().__init__()
            self.files = files; self.qci_list = qci_list; self.dl_clip = dl_clip; self.ul_clip = ul_clip
            self._cancel = False; self._pct = 0

        def cancel(self): self._cancel = True

        def run(self):
            try:
                out = process(self.files, self.qci_list, self.dl_clip, self.ul_clip,
                              callback=self.l, cancel_check=lambda: self._cancel,
                              progress_cb=lambda p: self.progress_pct.emit(int(p)))
                self.done.emit('' if self._cancel else out)
            except KeyboardInterrupt:
                self.done.emit('')

        def l(self, msg):
            self.progress.emit(msg)

    class MainWindow(QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("5G用户级公共监控速率统计工具 V2.4.1")
            self.files = []; self.last_out = None
            # 自适应：按屏幕高度缩放字号，高分屏放大避免看不清
            scr = QApplication.primaryScreen()
            self.scr_w = scr.geometry().width() if scr else 1440
            sh = scr.geometry().height() if scr else 1080
            self.fs = max(11, int(round(11 * sh / 1080)))
            self.setMinimumSize(1000, 680)
            self._build()
            self.showMaximized()   # 全屏主界面

        def _font(self, big=0):
            return QFont("PingFang SC", self.fs + big, QFont.Bold if big >= 2 else QFont.Normal)

        def dragEnterEvent(self, event):
            event.accept() if event.mimeData().hasUrls() else event.ignore()

        def dropEvent(self, event):
            new = [u.toLocalFile() for u in event.mimeData().urls()
                   if u.toLocalFile().lower().endswith(('.xlsx', '.xls', '.zip'))]
            if new:
                self.files = list(dict.fromkeys((self.files or []) + new))
                self.te.setText("\n".join(self.files)); self.br.setEnabled(True)

        def _build(self):
            self.setAcceptDrops(True)
            sp = QSplitter(Qt.Horizontal)
            # ===== 左侧：设置区（35%）=====
            left = QWidget(); ll = QVBoxLayout(left)
            title = QLabel("5G用户级公共监控速率统计工具  V2.4.1")
            title.setFont(self._font(3)); title.setAlignment(Qt.AlignCenter); ll.addWidget(title)

            gf = QGroupBox("输入文件"); gfl = QVBoxLayout(gf)
            bf = QHBoxLayout()
            self.bs = QPushButton("选择MMF文件"); self.bs.clicked.connect(self.sf)
            self.bc = QPushButton("清空"); self.bc.clicked.connect(self.cf)
            bf.addWidget(self.bs); bf.addWidget(self.bc); gfl.addLayout(bf)
            self.te = QTextEdit(); self.te.setReadOnly(True); self.te.setMaximumHeight(90)
            gfl.addWidget(self.te); ll.addWidget(gf)

            gp = QGroupBox("参数设置"); gpl = QGridLayout(gp)
            gpl.addWidget(QLabel("QCI:"), 0, 0)
            self.cb5 = QCheckBox("5"); self.cb6 = QCheckBox("6"); self.cb7 = QCheckBox("7")
            self.cb7.setChecked(True); self.cb6.setChecked(True); self.cb5.setChecked(True)
            gpl.addWidget(self.cb5, 0, 1); gpl.addWidget(self.cb6, 0, 2); gpl.addWidget(self.cb7, 0, 3)
            gpl.addWidget(QLabel("下行削峰(Mbps):"), 1, 0)
            self.sd = QSpinBox(); self.sd.setRange(100, 5000); self.sd.setValue(1000); gpl.addWidget(self.sd, 1, 1, 1, 3)
            gpl.addWidget(QLabel("上行削峰(Mbps):"), 2, 0)
            self.su = QSpinBox(); self.su.setRange(50, 1000); self.su.setValue(200); gpl.addWidget(self.su, 2, 1, 1, 3)
            ll.addWidget(gp)

            gh = QGroupBox("运行"); ghl = QHBoxLayout(gh)
            self.br = QPushButton("开始处理"); self.br.clicked.connect(self.run); self.br.setEnabled(False)
            self.br.setStyleSheet("QPushButton{background-color:#4CAF50;color:white;font-weight:bold;padding:8px 16px;}")
            self.bcancel = QPushButton("取消"); self.bcancel.clicked.connect(self.cancel); self.bcancel.setEnabled(False)
            self.bcancel.setStyleSheet("QPushButton{background-color:#f44336;color:white;padding:8px 12px;}")
            self.babout = QPushButton("关于"); self.babout.clicked.connect(self.about)
            ghl.addWidget(self.br); ghl.addWidget(self.bcancel); ghl.addWidget(self.babout); ll.addWidget(gh)
            ll.addStretch(1)
            left.setFont(self._font())

            # ===== 右侧：结果显示区（65%）=====
            right = QWidget(); rl = QVBoxLayout(right)
            rh = QLabel("运行结果"); rh.setFont(self._font(2)); rl.addWidget(rh)
            topbar = QHBoxLayout()
            self.bopen = QPushButton("打开输出文件"); self.bopen.clicked.connect(self.open_out); self.bopen.setEnabled(False)
            topbar.addStretch(1); topbar.addWidget(self.bopen); rl.addLayout(topbar)
            self.pb = QProgressBar(); rl.addWidget(self.pb)
            self.lt = QTextEdit(); self.lt.setReadOnly(True); rl.addWidget(self.lt, 1)
            right.setFont(self._font())

            sp.addWidget(left); sp.addWidget(right)
            sp.setStretchFactor(0, 35); sp.setStretchFactor(1, 65)
            sp.setSizes([int(self.scr_w * 0.35), int(self.scr_w * 0.65)])
            self.setCentralWidget(sp)

        def sf(self):
            fs, _ = QFileDialog.getOpenFileNames(self, "选择MMF文件", "", "Excel/压缩 (*.xlsx *.xls *.zip);;All (*)")
            if fs:
                self.files = list(dict.fromkeys((self.files or []) + fs))
                self.te.setText("\n".join(self.files)); self.br.setEnabled(True)

        def cf(self): self.files = []; self.te.clear(); self.br.setEnabled(False)

        def run(self):
            if not self.files: QMessageBox.warning(self, "提示", "请先选择MMF文件"); return
            qci_list = []
            if self.cb5.isChecked(): qci_list.append(5)
            if self.cb6.isChecked(): qci_list.append(6)
            if self.cb7.isChecked(): qci_list.append(7)
            if not qci_list: qci_list = [7]
            self.br.setEnabled(False); self.bcancel.setEnabled(True)
            self.pb.setValue(0); self.lt.append(f"开始处理... QCI={qci_list}")
            self.thread = WorkerThread(self.files, qci_list, self.sd.value(), self.su.value())
            self.thread.progress.connect(self.upd)
            self.thread.progress_pct.connect(lambda p: self.pb.setValue(p))
            self.thread.done.connect(self.dn); self.thread.start()

        def cancel(self):
            if hasattr(self, 'thread') and self.thread.isRunning():
                self.thread.cancel(); self.lt.append("已请求取消，将在当前阶段结束后停止...")

        @Slot(str)
        def upd(self, msg): self.lt.append(msg)

        @Slot(str)
        def dn(self, path):
            self.bcancel.setEnabled(False)
            if path:
                self.pb.setValue(100)
                self.last_out = path; self.bopen.setEnabled(True)
                self.lt.append(f"完成! {path}"); os.system(f'open "{path}"')
            else:
                self.pb.setValue(0); self.lt.append("已取消。")
            self.br.setEnabled(True)

        def open_out(self):
            if self.last_out: os.system(f'open "{self.last_out}"')

        def about(self):
            d = QDialog(self); d.setWindowTitle("关于")
            d.resize(560, 520)
            v = QVBoxLayout(d)
            te = QTextEdit(); te.setReadOnly(True)
            te.setHtml(
                "<h3 style='text-align:center'>5G用户级公共监控速率统计工具</h3>"
                "<p><b>版本:</b> V2.4.1 &nbsp; <b>开发者:</b> 孙晓军 &nbsp; "
                "<b>联系方式:</b> 317827@qq.com</p><hr/>"
                "<h4>更新记录</h4>"
                "<p><b>V2.4.1</b> <i>(2026-07-17)</i></p>"
                "<ul>"
                "<li>商店小文件识别修复：GAP_MERGE 2.0→3.0，防止商店大段被切碎误判FTP；8783段等真小文件恢复识别</li>"
                "<li>去掉商店小文件上行占比约束（RLC/MAC占比均不可靠），回到max(dl)选取</li>"
                "<li>对比横表改就近匹配：找离基准最近的同业务段，距离≤300秒，完整轮0缺失</li>"
                "<li>支持嵌套zip：递归解压(zip里套zip)</li>"
                "<li>取消按钮修复：写Excel各阶段加检查点，长循环每5000行检查</li>"
                "<li>清理重复sort_values；fmt_time改预编译正则</li>"
                "</ul>"
                "<p><b>V2.4.0</b> <i>(2026-07-16)</i></p>"
                "<ul>"
                "<li>性能优化：预编译正则 + usecols 只读必要列，MMF加载提速</li>"
                "<li>汇总批注：每格显示公式+实际数值（hover查看）</li>"
                "<li>D/E列颜色标记：详细过程 D列浅蓝/E列浅绿</li>"
                "</ul>"
                "<p><b>V2.3.0</b> <i>(2026-07-16)</i></p>"
                "<ul>"
                "<li>性能优化：基准对齐/筛选列向量化，处理耗时 107秒→14秒(约7.7倍)</li>"
                "<li>新增：取消按钮(阶段点中断)、拖拽文件入窗口、真实进度百分比</li>"
                "<li>对比竖表美化：业务/轮次配色、合并同轮次、冻结首行、行高</li>"
                "</ul>"
                "<p><b>V2.2.0</b> <i>(2026-07-16)</i></p>"
                "<ul>"
                "<li>对比增加「偏差」行/列：(代码−基准)/基准 百分比，速率与时长各算</li>"
                "<li>对比横表每轮由2列扩为3列(代码+基准+偏差)</li>"
                "<li>新增「对比竖表」sheet(横表转置)</li>"
                "<li>主界面改左右分栏(左35%设置/右65%结果)，全屏+自适应分辨率+高分屏字号放大+可拖动+打开输出按钮</li>"
                "</ul>"
                "<p><b>V2.1.0</b> <i>(2026-07-16)</i></p>"
                "<ul>"
                "<li>对比横表匹配改为「业务连续段」判定 + 按 FTP下载 切轮</li>"
                "<li>不再用固定时间窗口(-5~+60s)和硬切6条分组</li>"
                "<li>整轮6业务完整且全对上才编号，否则标「不计轮次」</li>"
                "</ul>"
                "<p><b>V2.0.9</b></p>"
                "<ul>"
                "<li>修复对比横表代码业务结束时间缺失</li>"
                "<li>新增第一个 FTP对 之前不完整轮次的识别</li>"
                "<li>版本号统一更新(文件名/关于/需求文档)</li>"
                "</ul>"
                "<p><b>V2.0.1</b></p>"
                "<ul><li>QCI 支持 5/6/7 多选；主界面增加「关于」</li></ul>"
                "<p><b>V2.0.0</b></p>"
                "<ul>"
                "<li>从零重写，三页向导改单页</li>"
                "<li>自动识别 FTP/商店/微信 6 类业务</li>"
                "<li>从原始 MMF 文件直接生成详细过程/对比/汇总 3 个 Sheet</li>"
                "<li>削峰阈值可调，轮次完整性检查</li>"
                "</ul>"
            )
            v.addWidget(te)
            bb = QPushButton("关闭"); bb.clicked.connect(d.accept)
            v.addWidget(bb)
            d.exec_()


if __name__ == '__main__':
    if GUI_OK and len(sys.argv) == 1:
        app = QApplication(sys.argv); app.setStyle('Fusion')
        w = MainWindow(); w.show(); sys.exit(app.exec())
    else:
        out = process(
            ['联通/mmf20260703115328-电信.xlsx', '联通/mmf20260703115334-电信.xlsx'],
            qci_list=[5,6,7], dl_clip=1000, ul_clip=200
        )
        print(f"输出: {out}")