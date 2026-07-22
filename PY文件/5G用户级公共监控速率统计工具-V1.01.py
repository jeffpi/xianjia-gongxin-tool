#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
5G用户级公共监控速率统计工具 V1.01
基于2025-2026年工信部测评用例规则开发
作者：Claude Code
创建日期：2026-06-23

V1.01 变更（对标 V6 原工具 + 完善指标体系）：
  - 统计结果改为 13 列扩充窄表，覆盖 FTP下载/上传、应用商店_小包/大包、微信_小文件/大文件 6 个细分业务
  - 速率指标统一用「削峰后 RLC 吞吐率」：均值速率、前10%峰值速率、达标速率占比、满分门限占比
  - 新增应用商店下载成功率 / 微信上传成功率（流量完整性代理判定，业务整体）
  - 新增小包/小文件平均时长、大包/大文件中位数时长（修复 duration 原为采样点数的 bug，改为首末采样时间差秒数）
  - 速率列可选 MAC/RLC（默认 RLC，对标 V6）
  - 业务规则表新增「连续点数」列
详见同目录《统计规则说明_V1.01.md》
"""

import sys
import os
import json
import re
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
import numpy as np
import pandas as pd
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QLabel, QLineEdit, QPushButton, QFileDialog,
    QTextEdit, QTableWidget, QTableWidgetItem, QHeaderView,
    QGroupBox, QFormLayout, QSpinBox, QDoubleSpinBox, QComboBox,
    QCheckBox, QMessageBox, QProgressBar, QSplitter
)
from PySide6.QtCore import Qt, QThread, Signal, Slot
from PySide6.QtGui import QFont, QColor
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ==================== 固定业务行定义（统计结果窄表按此顺序输出） ====================
# 每行：业务分组 / 业务类型(=display_name) / 方向 / 包类型
BIZ_ROWS = [
    {'分组': 'FTP', '类型': 'FTP下载', '方向': '下行', '包': '—'},
    {'分组': 'FTP', '类型': 'FTP上传', '方向': '上行', '包': '—'},
    {'分组': '应用商店', '类型': '应用商店_小包', '方向': '下行', '包': '小包'},
    {'分组': '应用商店', '类型': '应用商店_大包', '方向': '下行', '包': '大包'},
    {'分组': '微信', '类型': '微信_小文件', '方向': '上行', '包': '小文件'},
    {'分组': '微信', '类型': '微信_大文件', '方向': '上行', '包': '大文件'},
]

# 统计结果 13 列表头
STAT_HEADERS = [
    '业务分组', '业务类型', '方向', '包类型', '测试次数', '采样点数',
    '均值速率(Mbps)', '前10%峰值速率(Mbps)', '达标速率占比(%)',
    '满分门限占比(%)', '成功率(%)', '平均时长(s)', '中位数时长(s)'
]


# ==================== 配置管理 ====================

class ConfigManager:
    """配置文件管理器"""

    def __init__(self, config_path: str = None):
        if config_path is None:
            config_path = os.path.join(os.path.dirname(__file__), 'config', 'default_rules.json')
        self.config_path = config_path
        self.config = self.load_config()

    def load_config(self) -> Dict:
        """加载配置文件"""
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载配置文件失败: {e}")
            return self.get_default_config()

    def save_config(self, config: Dict = None) -> bool:
        """保存配置文件"""
        if config is None:
            config = self.config
        try:
            os.makedirs(os.path.dirname(self.config_path), exist_ok=True)
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"保存配置文件失败: {e}")
            return False

    def get_default_config(self) -> Dict:
        """获取默认配置"""
        return {
            "version": "1.01",
            "rule_version": "2025-2026工信部测评用例",
            "global_params": {
                "qci": 0,
                "rate_column": "RLC",
                "start_time": "auto",
                "end_time": "auto",
                "dl_peak_limit": 1000,
                "ul_peak_limit": 200,
                "dl_pass_threshold": 100,
                "ul_pass_threshold": 20,
                "scenario_type": "场景",
                "road_mean_threshold": 800
            },
            "business_rules": [],
            "voice_rules": [],
            "ott_rules": []
        }


# ==================== 数据处理引擎 ====================

class DataProcessor:
    """数据处理核心引擎"""

    def __init__(self, config: Dict):
        self.config = config
        self.raw_data = None
        self.cleaned_data = None
        self.filtered_data = None
        self.rate_details = None
        self.statistics = None

    # ---------- 辅助：速率列选择 ----------
    def _select_rate_col(self, df: pd.DataFrame, direction: str, pref: str) -> Optional[str]:
        """
        选择某方向的速率列。
        direction: '下行' / '上行'
        pref: 'RLC' / 'MAC'（优先选含该关键字的列；找不到则取任一吞吐率列）
        """
        candidates = [c for c in df.columns if direction in str(c) and '吞吐率' in str(c)]
        for c in candidates:
            if pref in str(c):
                return c
        return candidates[0] if candidates else None

    def parse_input_file(self, file_path: str) -> pd.DataFrame:
        """
        解析输入文件
        跳过前7行元数据，从第8行读取表头（header=7）
        """
        try:
            print("DEBUG: Reading Excel file...")
            df = pd.read_excel(file_path, header=7)

            print(f"DEBUG: Loaded {len(df)} rows, {len(df.columns)} columns")

            if '采集时间' in df.columns:
                df['采集时间_解析'] = df['采集时间'].apply(self._parse_time)
                print(f"DEBUG: Parsed collection time for {df['采集时间_解析'].notna().sum()} rows")
            else:
                print("DEBUG: '采集时间' column not found!")

            return df
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise Exception(f"解析输入文件失败: {e}")

    def _parse_time(self, time_str: str) -> Optional[datetime]:
        """解析采集时间字符串，格式：YYYY-MM-DD HH:MM:SS (xxx)"""
        if pd.isna(time_str) or time_str == "N/A":
            return None
        try:
            match = re.match(r'^(.+?)\s*\(\d+\)$', str(time_str))
            if match:
                time_str = match.group(1)
            return pd.to_datetime(time_str)
        except:
            return None

    def clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        数据清洗
        1. "-" 前向填充
        2. bps→Mbps转换（RLC吞吐率）
        """
        df = df.copy()
        df = df.replace('-', np.nan).ffill()

        for col in df.columns:
            if 'RLC吞吐率' in col and 'Mbps' not in col:
                new_col = col.replace('bps', 'Mbps')
                if new_col != col and new_col not in df.columns:
                    df[new_col] = df[col] / 1_000_000

        return df

    def filter_data(self, df: pd.DataFrame, params: Dict) -> pd.DataFrame:
        """
        过滤筛选：时间窗 / QCI / 载波聚合属性
        """
        df = df.copy()
        print(f"DEBUG: filter_data input: {len(df)} rows")

        if '采集时间_解析' in df.columns:
            if params.get('start_time') and params['start_time'] != 'auto':
                start_time = pd.to_datetime(params['start_time'])
                df = df[df['采集时间_解析'] >= start_time]
            if params.get('end_time') and params['end_time'] != 'auto':
                end_time = pd.to_datetime(params['end_time'])
                df = df[df['采集时间_解析'] <= end_time]

        qci_value = params.get('qci', 0)
        if 'QCI' in df.columns and qci_value and qci_value > 0:
            before_qci = len(df)
            df = df[df['QCI'] == qci_value]
            print(f"DEBUG: QCI={qci_value} filter: {before_qci} -> {len(df)} rows")

        if '载波聚合属性' in df.columns:
            before_ca = len(df)
            df = df[df['载波聚合属性'].isin(['非CA用户', 'CA用户主小区'])]
            print(f"DEBUG: CA filter: {before_ca} -> {len(df)} rows")

        print(f"DEBUG: filter_data output: {len(df)} rows")
        return df

    def identify_direction(self, df: pd.DataFrame, rules: List[Dict], params: Dict) -> pd.DataFrame:
        """
        方向识别（按 params['rate_column'] 选择速率列，默认 RLC）
        1. 计算上下行比率判定方向
        2. 同一秒同向保留最大速率点
        """
        print(f"DEBUG: identify_direction input rows: {len(df)}")
        if df.empty:
            return df

        df = df.copy()
        pref = params.get('rate_column', 'RLC')
        dl_col = self._select_rate_col(df, '下行', pref)
        ul_col = self._select_rate_col(df, '上行', pref)

        if not dl_col or not ul_col:
            print(f"DEBUG: Missing columns - dl: {dl_col}, ul: {ul_col}")
            return df

        print(f"DEBUG: Using columns (pref={pref}) - DL: {dl_col}, UL: {ul_col}")

        df[dl_col] = pd.to_numeric(df[dl_col], errors='coerce').fillna(0)
        df[ul_col] = pd.to_numeric(df[ul_col], errors='coerce').fillna(0)

        up_min = min([r.get('min_rate', 3) for r in rules if r.get('direction') == '上传'], default=3)
        down_min = min([r.get('min_rate', 10) for r in rules if r.get('direction') == '下载'], default=10)
        print(f"DEBUG: up_min={up_min}, down_min={down_min}")

        df['方向'] = ''
        df['速率_有效'] = 0.0

        down_count = 0
        up_count = 0

        for idx, row in df.iterrows():
            dl_rate = row[dl_col]
            ul_rate = row[ul_col]

            if dl_rate > 0 and ul_rate > 0:
                u_d_ratio = ul_rate / dl_rate if dl_rate > 0 else 0
                d_u_ratio = dl_rate / ul_rate if ul_rate > 0 else 0

                if u_d_ratio > 10 and ul_rate > up_min:
                    df.loc[idx, '方向'] = '上行'
                    df.loc[idx, '速率_有效'] = ul_rate
                    up_count += 1
                elif d_u_ratio > 10 and dl_rate > down_min:
                    df.loc[idx, '方向'] = '下行'
                    df.loc[idx, '速率_有效'] = dl_rate
                    down_count += 1
            elif dl_rate > down_min:
                df.loc[idx, '方向'] = '下行'
                df.loc[idx, '速率_有效'] = dl_rate
                down_count += 1
            elif ul_rate > up_min:
                df.loc[idx, '方向'] = '上行'
                df.loc[idx, '速率_有效'] = ul_rate
                up_count += 1

        print(f"DEBUG: Direction identified - down: {down_count}, up: {up_count}")

        # 同一秒同向保留最大速率点
        if '采集时间_解析' in df.columns:
            df['时间秒'] = df['采集时间_解析'].dt.floor('s')
            before_dedup = len(df)
            valid_points = df[df['方向'] != '']
            if len(valid_points) > 0:
                df = df.loc[valid_points.groupby(['时间秒', '方向'])['速率_有效'].idxmax()]
                df = df.drop(columns=['时间秒'])
                print(f"DEBUG: Dedup: {before_dedup} -> {len(df)} rows")

        print(f"DEBUG: identify_direction output rows: {len(df)}")
        return df

    def segment_business(self, df: pd.DataFrame, rules: List[Dict]) -> pd.DataFrame:
        """
        业务分段
        1. 同向连续点为一段
        2. 相邻时间差 > 8 秒截断
        3. 时长 = 段内首末采样时间差（秒）
        4. 按流量区间/秒数匹配细分业务类型（含大/小包）
        """
        if df.empty or '方向' not in df.columns:
            return df

        df = df.copy()
        df = df[df['方向'] != ''].sort_values('采集时间_解析')

        df['段ID'] = 0
        current_segment = 0
        last_time = None
        last_direction = None

        for idx, row in df.iterrows():
            current_time = row['采集时间_解析']
            current_direction = row['方向']

            if last_time is None:
                current_segment += 1
            elif current_direction != last_direction:
                current_segment += 1
            elif (current_time - last_time).total_seconds() > 8:
                current_segment += 1

            df.loc[idx, '段ID'] = current_segment
            last_time = current_time
            last_direction = current_direction

        # 计算每段的细分业务类型
        seg_map = {}
        for segment_id, segment_df in df.groupby('段ID'):
            if segment_df.empty:
                continue
            direction = segment_df['方向'].iloc[0]
            # 时长 = 首末采样时间差（秒）
            duration = (segment_df['采集时间_解析'].max() - segment_df['采集时间_解析'].min()).total_seconds()
            total_flow = segment_df['速率_有效'].sum()  # Mbps·s
            business_type = self._match_business(direction, duration, total_flow, rules)
            seg_map[segment_id] = business_type

        df['业务类型'] = df['段ID'].map(seg_map).fillna('未知业务')
        return df

    def _match_business(self, direction: str, duration: float, total_flow: float, rules: List[Dict]) -> str:
        """
        匹配细分业务类型（返回 display_name，6 类之一）
        - 优先级1：流量落在某业务完整区间 [flow_min, flow_max] → 该业务（完整）
        - 优先级1.5：流量落在「业务组宽区间」（小包flow_min ~ 大包flow_max）但不在完整区间
                     → 归该组大包（大包下载不完整，如 140~1100MB）
        - 优先级2：FTP（秒数 + 最小速率）
        - 其余：未知业务
        """
        direction_map = {'下行': '下载', '上行': '上传'}
        mapped_direction = direction_map.get(direction, direction)
        same_direction_rules = [r for r in rules if r.get('direction') == mapped_direction and r.get('enabled', True)]

        flow_mb = total_flow / 8  # Mbps·s → MB

        # 优先级1：完整区间命中（应用商店/微信等有 flow 区间的业务）
        for rule in same_direction_rules:
            flow_min = rule.get('flow_min')
            flow_max = rule.get('flow_max')
            if flow_min is not None and flow_max is not None:
                calc_seconds = rule.get('calc_seconds')
                if calc_seconds is None or duration >= calc_seconds:
                    if flow_min <= flow_mb <= flow_max:
                        return rule.get('display_name', rule['name'])

        # 优先级1.5：业务组宽区间（识别大包不完整段）
        # 收集同方向、有 flow 区间的规则，按业务分组取 [min(flow_min), max(flow_max)]
        flow_rules = [r for r in same_direction_rules if r.get('flow_min') is not None and r.get('flow_max') is not None]
        if flow_rules:
            group_min = min(r['flow_min'] for r in flow_rules)
            group_max = max(r['flow_max'] for r in flow_rules)
            if group_min <= flow_mb <= group_max:
                # 落在宽区间但未在完整区间 → 归「大包」(流量偏大者)
                big_rules = [r for r in flow_rules if r.get('package') in ('大包', '大文件')]
                if big_rules:
                    return big_rules[0].get('display_name', big_rules[0]['name'])

        # 优先级2：FTP（秒数 + 最小速率）
        for rule in same_direction_rules:
            if 'FTP' in rule['name']:
                calc_seconds = rule.get('calc_seconds')
                min_rate = rule.get('min_rate')
                if calc_seconds is not None and duration < calc_seconds:
                    continue
                avg_rate = total_flow / duration if duration > 0 else 0
                if min_rate is not None and avg_rate >= min_rate:
                    return rule.get('display_name', rule['name'])

        return '未知业务'

    def apply_peak_limiting(self, df: pd.DataFrame, params: Dict) -> pd.DataFrame:
        """削峰处理：每点 min(原始速率, 削峰阈值)，按 params['rate_column'] 选列"""
        df = df.copy()

        dl_limit = params.get('dl_peak_limit', 1000)
        ul_limit = params.get('ul_peak_limit', 200)
        pref = params.get('rate_column', 'RLC')

        dl_rate_col = self._select_rate_col(df, '下行', pref)
        ul_rate_col = self._select_rate_col(df, '上行', pref)

        print(f"DEBUG: Peak limiting (pref={pref}) - DL: {dl_rate_col}, UL: {ul_rate_col}, limits: {dl_limit}/{ul_limit}")

        # 确保数值类型
        if dl_rate_col:
            df[dl_rate_col] = pd.to_numeric(df[dl_rate_col], errors='coerce')
            df['削峰后下行RLC'] = df[dl_rate_col].apply(lambda x: min(x, dl_limit) if pd.notna(x) else x)
        if ul_rate_col:
            df[ul_rate_col] = pd.to_numeric(df[ul_rate_col], errors='coerce')
            df['削峰后上行RLC'] = df[ul_rate_col].apply(lambda x: min(x, ul_limit) if pd.notna(x) else x)

        return df

    def calculate_statistics(self, df: pd.DataFrame, params: Dict, rules: List[Dict]) -> List[Dict]:
        """
        统计计算 → 返回 6 行（固定 BIZ_ROWS 顺序），每行 13 列字段。
        速率类：采样点级「削峰后 RLC」列统计
        时长类 / 成功率：段级聚合（每段首末时间差 + 流量完整性）
        """
        dl_pass = params.get('dl_pass_threshold', 100)
        ul_pass = params.get('ul_pass_threshold', 20)
        dl_peak = params.get('dl_peak_limit', 1000)
        ul_peak = params.get('ul_peak_limit', 200)

        # ----- 段级聚合 -----
        seg = pd.DataFrame()
        if not df.empty and '段ID' in df.columns and '业务类型' in df.columns:
            valid = df[df['业务类型'] != '未知业务'].copy()
            if not valid.empty:
                seg = valid.groupby('段ID').agg(
                    方向=('方向', 'first'),
                    业务类型=('业务类型', 'first'),
                    时长=('采集时间_解析', lambda x: (x.max() - x.min()).total_seconds()),
                    总流量MB=('速率_有效', lambda s: s.sum() / 8),
                ).reset_index()

        # 每个细分业务对应的规则（用于完整性判定：flow_min/flow_max）
        rule_by_type = {}
        for r in rules:
            dn = r.get('display_name')
            if dn:
                rule_by_type[dn] = r

        # 业务整体成功率：应用商店 / 微信
        success_rate_store = self._group_success_rate(seg, '应用商店', rule_by_type)
        success_rate_wechat = self._group_success_rate(seg, '微信', rule_by_type)

        results = []
        for row_def in BIZ_ROWS:
            biz_type = row_def['类型']
            direction = row_def['方向']
            group = row_def['分组']

            # 该业务的采样点级数据
            sub = df[df['业务类型'] == biz_type] if '业务类型' in df.columns else pd.DataFrame()
            seg_sub = seg[seg['业务类型'] == biz_type] if not seg.empty else pd.DataFrame()

            # 选削峰后列
            clipped_col = '削峰后下行RLC' if direction == '下行' else '削峰后上行RLC'
            clip_series = pd.Series(dtype=float)
            if clipped_col in sub.columns:
                clip_series = pd.to_numeric(sub[clipped_col], errors='coerce').dropna()

            stats = {
                '业务分组': group,
                '业务类型': biz_type,
                '方向': direction,
                '包类型': row_def['包'],
                '测试次数': int(seg_sub['段ID'].nunique()) if not seg_sub.empty else 0,
                '采样点数': int(len(sub)),
                '均值速率': round(float(clip_series.mean()), 2) if len(clip_series) > 0 else None,
                '前10%峰值速率': None,
                '达标速率占比': None,
                '满分门限占比': None,
                '成功率': None,
                '平均时长': None,
                '中位数时长': None,
            }

            # FTP 行：速率四件套
            if group == 'FTP':
                if len(clip_series) > 0:
                    pass_thr = dl_pass if direction == '下行' else ul_pass
                    peak_thr = dl_peak if direction == '下行' else ul_peak
                    sorted_rates = clip_series.sort_values(ascending=False)
                    top10_count = max(1, int(len(sorted_rates) * 0.1))
                    stats['前10%峰值速率'] = round(float(sorted_rates.head(top10_count).mean()), 2)
                    stats['达标速率占比'] = round(float((clip_series >= pass_thr).sum() / len(clip_series) * 100), 2)
                    stats['满分门限占比'] = round(float((clip_series >= peak_thr).sum() / len(clip_series) * 100), 2)

            # 应用商店 / 微信 行：成功率（业务整体） + 时长
            if group == '应用商店':
                stats['成功率'] = success_rate_store
            elif group == '微信':
                stats['成功率'] = success_rate_wechat

            if group in ('应用商店', '微信'):
                # 时长：仅完整段
                complete_durations = self._complete_durations(seg_sub, rule_by_type.get(biz_type))
                if row_def['包'] in ('小包', '小文件') and len(complete_durations) > 0:
                    stats['平均时长'] = round(float(np.mean(complete_durations)), 2)
                elif row_def['包'] in ('大包', '大文件') and len(complete_durations) > 0:
                    stats['中位数时长'] = round(float(np.median(complete_durations)), 2)

            results.append(stats)

        return results

    def _complete_durations(self, seg_sub: pd.DataFrame, rule: Optional[Dict]) -> List[float]:
        """返回该细分业务「完整段」（流量落在 flow 区间）的时长列表"""
        if seg_sub is None or seg_sub.empty or not rule:
            return []
        flow_min = rule.get('flow_min')
        flow_max = rule.get('flow_max')
        if flow_min is None or flow_max is None:
            # 无区间（如 FTP）→ 全部视为完整
            return seg_sub['时长'].tolist()
        mask = (seg_sub['总流量MB'] >= flow_min) & (seg_sub['总流量MB'] <= flow_max)
        return seg_sub.loc[mask, '时长'].tolist()

    def _group_success_rate(self, seg: pd.DataFrame, group: str, rule_by_type: Dict) -> Optional[float]:
        """
        业务整体成功率（流量完整性代理判定）：
          成功率 = 完整段数 / 该组识别段总数（含大包不完整段）
        group: '应用商店' / '微信'
        """
        if seg is None or seg.empty:
            return None
        types = [r['类型'] for r in BIZ_ROWS if r['分组'] == group]
        group_seg = seg[seg['业务类型'].isin(types)]
        if group_seg.empty:
            return None
        total = len(group_seg)
        complete = 0
        for t in types:
            sub = group_seg[group_seg['业务类型'] == t]
            complete += len(self._complete_durations(sub, rule_by_type.get(t)))
        if total == 0:
            return None
        return round(complete / total * 100, 2)

    def process(self, file_path: str, params: Dict, rules: List[Dict]) -> Tuple[pd.DataFrame, pd.DataFrame, List[Dict]]:
        """
        完整处理流程
        返回：(原表整理, 速率详表, 统计结果[List[Dict]])
        """
        self.raw_data = self.parse_input_file(file_path)
        self.cleaned_data = self.clean_data(self.raw_data)
        self.filtered_data = self.filter_data(self.cleaned_data, params)
        rate_df = self.identify_direction(self.filtered_data, rules, params)
        rate_df = self.segment_business(rate_df, rules)
        rate_df = self.apply_peak_limiting(rate_df, params)
        self.statistics = self.calculate_statistics(rate_df, params, rules)
        self.rate_details = rate_df
        return self.cleaned_data, rate_df, self.statistics


# ==================== Excel导出 ====================

class ExcelExporter:
    """Excel导出器"""

    def __init__(self):
        self.wb = None

    def export(self, output_path: str, cleaned_data: pd.DataFrame, rate_details: pd.DataFrame, statistics: List[Dict]):
        """导出Excel文件，包含3个sheet"""
        self.wb = Workbook()
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        self._create_statistics_sheet(statistics)
        self._create_rate_details_sheet(rate_details)
        self._create_cleaned_data_sheet(cleaned_data)
        self.wb.save(output_path)

    def _create_statistics_sheet(self, statistics: List[Dict]):
        """创建统计结果sheet（13列窄表）"""
        ws = self.wb.create_sheet('统计结果')

        for col, header in enumerate(STAT_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')

        for row, stats in enumerate(statistics, 2):
            ws.cell(row=row, column=1, value=stats.get('业务分组', ''))
            ws.cell(row=row, column=2, value=stats.get('业务类型', ''))
            ws.cell(row=row, column=3, value=stats.get('方向', ''))
            ws.cell(row=row, column=4, value=stats.get('包类型', ''))
            ws.cell(row=row, column=5, value=stats.get('测试次数', 0))
            ws.cell(row=row, column=6, value=stats.get('采样点数', 0))
            self._write_num(ws, row, 7, stats.get('均值速率'))
            self._write_num(ws, row, 8, stats.get('前10%峰值速率'))
            self._write_num(ws, row, 9, stats.get('达标速率占比'))
            self._write_num(ws, row, 10, stats.get('满分门限占比'))
            self._write_num(ws, row, 11, stats.get('成功率'))
            self._write_num(ws, row, 12, stats.get('平均时长'))
            self._write_num(ws, row, 13, stats.get('中位数时长'))

    @staticmethod
    def _write_num(ws, row, col, value):
        """写入数值，None 留空"""
        if value is None:
            ws.cell(row=row, column=col, value='')
        else:
            ws.cell(row=row, column=col, value=value)

    def _create_rate_details_sheet(self, rate_details: pd.DataFrame):
        """创建速率详表sheet"""
        ws = self.wb.create_sheet('速率详表')
        if rate_details is None or rate_details.empty:
            ws.cell(row=1, column=1, value='无数据')
            return

        key_columns = ['序号', '采集时间', '来源', '基站标识', '小区标识', '业务类型',
                       '下行RLC吞吐率', '上行RLC吞吐率', '削峰后下行RLC', '削峰后上行RLC', '方向']
        available_columns = [col for col in key_columns if col in rate_details.columns]

        for col, header in enumerate(available_columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        for row_idx, row_data in rate_details.iterrows():
            for col, col_name in enumerate(available_columns, 1):
                value = row_data.get(col_name, '')
                if pd.isna(value):
                    value = ''
                ws.cell(row=row_idx + 2, column=col, value=value)

    def _create_cleaned_data_sheet(self, cleaned_data: pd.DataFrame):
        """创建原表整理sheet"""
        ws = self.wb.create_sheet('原表整理')
        if cleaned_data is None or cleaned_data.empty:
            ws.cell(row=1, column=1, value='无数据')
            return

        for col, header in enumerate(cleaned_data.columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        for row_idx, row_data in cleaned_data.iterrows():
            for col, value in enumerate(row_data, 1):
                if pd.isna(value):
                    value = ''
                ws.cell(row=row_idx + 2, column=col, value=value)


# ==================== 工作线程 ====================

class ProcessingThread(QThread):
    """数据处理工作线程"""
    progress_signal = Signal(str)
    finished_signal = Signal(bool, str, object, object, object)

    def __init__(self, processor: DataProcessor, file_path: str, params: Dict, rules: List[Dict]):
        super().__init__()
        self.processor = processor
        self.file_path = file_path
        self.params = params
        self.rules = rules

    def run(self):
        try:
            self.progress_signal.emit("正在解析输入文件...")
            cleaned, rate_details, statistics = self.processor.process(
                self.file_path, self.params, self.rules
            )
            self.progress_signal.emit("正在生成结果...")
            print(f"DEBUG: Processing complete, statistics rows: {len(statistics)}")
            self.finished_signal.emit(True, "处理完成", cleaned, rate_details, statistics)
        except Exception as e:
            print(f"DEBUG: ProcessingThread error: {e}")
            import traceback
            traceback.print_exc()
            self.finished_signal.emit(False, f"处理失败: {str(e)}", None, None, None)


# ==================== 主窗口 ====================

class MainWindow(QMainWindow):
    """主窗口"""

    def __init__(self):
        super().__init__()
        self.config_manager = ConfigManager()
        self.config = self.config_manager.config
        self.processor = None
        self.input_files = []
        self.processing_thread = None

        self.init_ui()
        self.load_config_to_ui()

    def init_ui(self):
        """初始化UI"""
        self.setWindowTitle('5G用户级公共监控速率统计工具 V1.01')
        self.setGeometry(100, 100, 1500, 900)

        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)

        main_layout.addWidget(self.create_left_panel(), 1)
        main_layout.addWidget(self.create_right_panel(), 2)

    def create_left_panel(self) -> QWidget:
        """创建左侧配置面板"""
        panel = QWidget()
        layout = QVBoxLayout(panel)

        tab_widget = QTabWidget()
        tab_widget.addTab(self.create_global_params_tab(), '全局参数')
        tab_widget.addTab(self.create_business_rules_tab(), '业务规则')
        layout.addWidget(tab_widget)

        file_group = QGroupBox('文件输入')
        file_layout = QVBoxLayout()
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText('选择输入文件...')
        file_layout.addWidget(self.file_path_edit)

        file_btn_layout = QHBoxLayout()
        select_file_btn = QPushButton('选择文件')
        select_file_btn.clicked.connect(self.select_input_file)
        file_btn_layout.addWidget(select_file_btn)
        file_layout.addLayout(file_btn_layout)
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)

        btn_layout = QHBoxLayout()
        process_btn = QPushButton('开始统计')
        process_btn.setStyleSheet('QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 10px; }')
        process_btn.clicked.connect(self.start_processing)
        btn_layout.addWidget(process_btn)

        export_btn = QPushButton('导出Excel')
        export_btn.clicked.connect(self.export_excel)
        btn_layout.addWidget(export_btn)
        layout.addLayout(btn_layout)

        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)

        log_group = QGroupBox('运行日志')
        log_layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout)
        layout.addWidget(log_group, 1)

        return panel

    def create_global_params_tab(self) -> QWidget:
        """创建全局参数选项卡"""
        tab = QWidget()
        layout = QFormLayout(tab)

        self.qci_spin = QSpinBox()
        self.qci_spin.setRange(0, 99)
        self.qci_spin.setValue(0)
        self.qci_spin.setSpecialValueText("全部")
        layout.addRow('QCI(0=全部):', self.qci_spin)

        # 速率列选择（V1.01 新增，默认 RLC 对标 V6）
        self.rate_column_combo = QComboBox()
        self.rate_column_combo.addItems(['RLC', 'MAC'])
        layout.addRow('速率取值列:', self.rate_column_combo)

        self.dl_peak_spin = QDoubleSpinBox()
        self.dl_peak_spin.setRange(0, 5000)
        self.dl_peak_spin.setValue(1000)
        layout.addRow('下行削峰速率(Mbps):', self.dl_peak_spin)

        self.ul_peak_spin = QDoubleSpinBox()
        self.ul_peak_spin.setRange(0, 1000)
        self.ul_peak_spin.setValue(200)
        layout.addRow('上行削峰速率(Mbps):', self.ul_peak_spin)

        self.dl_pass_spin = QDoubleSpinBox()
        self.dl_pass_spin.setRange(0, 1000)
        self.dl_pass_spin.setValue(100)
        layout.addRow('下行达标门限(Mbps):', self.dl_pass_spin)

        self.ul_pass_spin = QDoubleSpinBox()
        self.ul_pass_spin.setRange(0, 100)
        self.ul_pass_spin.setValue(20)
        layout.addRow('上行达标门限(Mbps):', self.ul_pass_spin)

        self.scenario_combo = QComboBox()
        self.scenario_combo.addItems(['场景', '道路'])
        layout.addRow('场景类型:', self.scenario_combo)

        self.road_threshold_spin = QDoubleSpinBox()
        self.road_threshold_spin.setRange(0, 2000)
        self.road_threshold_spin.setValue(800)
        layout.addRow('道路均值门限(Mbps):', self.road_threshold_spin)

        return tab

    def create_business_rules_tab(self) -> QWidget:
        """创建业务规则选项卡（9列，含「连续点数」）"""
        tab = QWidget()
        layout = QVBoxLayout(tab)

        self.rules_table = QTableWidget()
        self.rules_table.setColumnCount(9)
        self.rules_table.setHorizontalHeaderLabels([
            '业务名称', '方向', '计算秒数', '最小速率', '连续点数',
            '流量Min', '流量Max', '中位流量', '启用'
        ])
        self.rules_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.rules_table)

        btn_layout = QHBoxLayout()
        add_rule_btn = QPushButton('添加规则')
        add_rule_btn.clicked.connect(self.add_business_rule)
        btn_layout.addWidget(add_rule_btn)
        delete_rule_btn = QPushButton('删除规则')
        delete_rule_btn.clicked.connect(self.delete_business_rule)
        btn_layout.addWidget(delete_rule_btn)
        layout.addLayout(btn_layout)

        return tab

    def create_right_panel(self) -> QWidget:
        """创建右侧结果展示面板"""
        panel = QWidget()
        layout = QVBoxLayout(panel)

        tab_widget = QTabWidget()
        tab_widget.addTab(self.create_statistics_tab(), '统计结果')
        tab_widget.addTab(self.create_details_tab(), '速率详表')
        tab_widget.addTab(self.create_cleaned_tab(), '原表整理')
        layout.addWidget(tab_widget)

        return panel

    def create_statistics_tab(self) -> QWidget:
        """创建统计结果选项卡（13列）"""
        tab = QWidget()
        layout = QVBoxLayout(tab)

        self.stats_table = QTableWidget()
        self.stats_table.setColumnCount(len(STAT_HEADERS))
        self.stats_table.setHorizontalHeaderLabels(STAT_HEADERS)
        self.stats_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.stats_table)

        return tab

    def create_details_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        self.details_table = QTableWidget()
        layout.addWidget(self.details_table)
        return tab

    def create_cleaned_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        self.cleaned_table = QTableWidget()
        layout.addWidget(self.cleaned_table)
        return tab

    def load_config_to_ui(self):
        """加载配置到UI"""
        params = self.config.get('global_params', {})
        self.qci_spin.setValue(params.get('qci', 0))
        self.rate_column_combo.setCurrentText(params.get('rate_column', 'RLC'))
        self.dl_peak_spin.setValue(params.get('dl_peak_limit', 1000))
        self.ul_peak_spin.setValue(params.get('ul_peak_limit', 200))
        self.dl_pass_spin.setValue(params.get('dl_pass_threshold', 100))
        self.ul_pass_spin.setValue(params.get('ul_pass_threshold', 20))
        self.scenario_combo.setCurrentText(params.get('scenario_type', '场景'))
        self.road_threshold_spin.setValue(params.get('road_mean_threshold', 800))

        # 业务规则（9列：名称0 方向1 秒数2 最小速率3 连续点数4 流量Min5 流量Max6 中位流量7 启用8）
        rules = self.config.get('business_rules', [])
        self.rules_table.setRowCount(len(rules))
        for row, rule in enumerate(rules):
            self.rules_table.setItem(row, 0, QTableWidgetItem(rule.get('name', '')))
            self.rules_table.setItem(row, 1, QTableWidgetItem(rule.get('direction', '')))
            self.rules_table.setItem(row, 2, QTableWidgetItem(self._fmt(rule.get('calc_seconds'))))
            self.rules_table.setItem(row, 3, QTableWidgetItem(self._fmt(rule.get('min_rate'))))
            self.rules_table.setItem(row, 4, QTableWidgetItem(self._fmt(rule.get('continuous_count'))))
            self.rules_table.setItem(row, 5, QTableWidgetItem(self._fmt(rule.get('flow_min'))))
            self.rules_table.setItem(row, 6, QTableWidgetItem(self._fmt(rule.get('flow_max'))))
            self.rules_table.setItem(row, 7, QTableWidgetItem(self._fmt(rule.get('median_flow'))))
            enabled_item = QTableWidgetItem()
            enabled_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            enabled_item.setCheckState(Qt.Checked if rule.get('enabled', True) else Qt.Unchecked)
            self.rules_table.setItem(row, 8, enabled_item)

    @staticmethod
    def _fmt(v) -> str:
        """数值格式化用于表格显示（None → 空）"""
        if v is None:
            return ''
        return str(v)

    def select_input_file(self):
        """选择输入文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, '选择输入文件', '', 'Excel Files (*.xlsx *.xls)'
        )
        if file_path:
            self.input_files = [file_path]
            self.file_path_edit.setText(file_path)
            self.log_message(f"已选择文件: {file_path}")

    def get_params_from_ui(self) -> Dict:
        """从UI获取参数"""
        return {
            'qci': self.qci_spin.value(),
            'rate_column': self.rate_column_combo.currentText(),
            'dl_peak_limit': self.dl_peak_spin.value(),
            'ul_peak_limit': self.ul_peak_spin.value(),
            'dl_pass_threshold': self.dl_pass_spin.value(),
            'ul_pass_threshold': self.ul_pass_spin.value(),
            'scenario_type': self.scenario_combo.currentText(),
            'road_mean_threshold': self.road_threshold_spin.value()
        }

    def get_rules_from_ui(self) -> List[Dict]:
        """从UI获取业务规则（9列）"""
        def safe_num(text: str) -> Optional[float]:
            if not text or str(text).strip() in ('', 'None'):
                return None
            try:
                return float(text)
            except ValueError:
                return None

        # 建立 display_name/group/package 的默认映射（按业务名关键字）
        def defaults_for(name: str, direction: str) -> Dict:
            name_l = name or ''
            if 'FTP' in name_l and direction == '下载':
                return {'group': 'FTP', 'package': None, 'display_name': 'FTP下载'}
            if 'FTP' in name_l and direction == '上传':
                return {'group': 'FTP', 'package': None, 'display_name': 'FTP上传'}
            if '泰尔' in name_l:
                return {'group': '应用商店', 'package': '小包', 'display_name': '应用商店_小包'}
            if '我的世界' in name_l:
                return {'group': '应用商店', 'package': '大包', 'display_name': '应用商店_大包'}
            if '5M' in name_l or '5M' in name_l.upper():
                return {'group': '微信', 'package': '小文件', 'display_name': '微信_小文件'}
            if '200M' in name_l or '200M' in name_l.upper():
                return {'group': '微信', 'package': '大文件', 'display_name': '微信_大文件'}
            return {'group': '', 'package': None, 'display_name': name_l}

        rules = []
        for row in range(self.rules_table.rowCount()):
            name_item = self.rules_table.item(row, 0)
            if name_item is None:
                continue
            name = name_item.text().strip()
            if not name:
                continue
            direction = self.rules_table.item(row, 1).text().strip()
            enabled_item = self.rules_table.item(row, 8)
            enabled = enabled_item.checkState() == Qt.Checked if enabled_item else True
            d = defaults_for(name, direction)
            calc_s = safe_num(self.rules_table.item(row, 2).text())
            rules.append({
                'id': row + 1,
                'name': name,
                'display_name': d['display_name'],
                'group': d['group'],
                'package': d['package'],
                'direction': direction,
                'calc_seconds': calc_s,
                'min_rate': safe_num(self.rules_table.item(row, 3).text()) or 0,
                'continuous_count': safe_num(self.rules_table.item(row, 4).text()),
                'flow_min': safe_num(self.rules_table.item(row, 5).text()),
                'flow_max': safe_num(self.rules_table.item(row, 6).text()),
                'median_flow': safe_num(self.rules_table.item(row, 7).text()),
                'enabled': enabled,
            })
        return rules

    @Slot()
    def start_processing(self):
        """开始处理"""
        print("DEBUG: start_processing called")
        if not self.input_files:
            QMessageBox.warning(self, '警告', '请先选择输入文件')
            return

        file_path = self.input_files[0]
        params = self.get_params_from_ui()
        rules = self.get_rules_from_ui()

        print(f"DEBUG: file_path = {file_path}")
        print(f"DEBUG: params = {params}")
        print(f"DEBUG: rules count = {len(rules)}")

        self.log_message("=" * 50)
        self.log_message("开始处理...")
        self.log_message(f"文件: {os.path.basename(file_path)}")
        self.log_message(f"QCI: {params['qci']}，速率列: {params['rate_column']}")
        self.log_message(f"下行削峰: {params['dl_peak_limit']} Mbps，上行削峰: {params['ul_peak_limit']} Mbps")
        self.progress_bar.setValue(0)

        self.sender().setEnabled(False)
        self.sender().setText("处理中...")

        self.processor = DataProcessor(self.config)
        self.processing_thread = ProcessingThread(self.processor, file_path, params, rules)
        self.processing_thread.progress_signal.connect(self.update_progress)
        self.processing_thread.finished_signal.connect(self.on_processing_finished)
        self.processing_thread.start()

    @Slot(str)
    def update_progress(self, message: str):
        self.log_message(message)
        current = self.progress_bar.value()
        if current < 90:
            self.progress_bar.setValue(min(current + 10, 90))

    @Slot(bool, str, object, object, object)
    def on_processing_finished(self, success: bool, message: str, cleaned: pd.DataFrame,
                               rate_details: pd.DataFrame, statistics: List[Dict]):
        """处理完成回调"""
        self.progress_bar.setValue(100)
        self.log_message(message)
        self.log_message("=" * 50)

        for btn in self.findChildren(QPushButton):
            if btn.text() in ["处理中...", "开始统计"]:
                btn.setEnabled(True)
                btn.setText("开始统计")

        if success:
            self.display_statistics(statistics)
            self.display_rate_details(rate_details)
            self.display_cleaned_data(cleaned)

            self.last_cleaned = cleaned
            self.last_rate_details = rate_details
            self.last_statistics = statistics

            self.log_message(f"✓ 统计完成！共 {len(statistics)} 个业务行")
            QMessageBox.information(self, '完成', '数据处理完成！')
        else:
            self.log_message(f"✗ 处理失败: {message}")
            QMessageBox.critical(self, '错误', f'处理失败: {message}')

    def display_statistics(self, statistics: List[Dict]):
        """显示统计结果（13列）"""
        self.stats_table.setRowCount(len(statistics))
        keys = ['业务分组', '业务类型', '方向', '包类型', '测试次数', '采样点数',
                '均值速率', '前10%峰值速率', '达标速率占比', '满分门限占比',
                '成功率', '平均时长', '中位数时长']
        for row, stats in enumerate(statistics):
            for col, k in enumerate(keys):
                v = stats.get(k)
                if v is None:
                    text = ''
                elif isinstance(v, float):
                    text = f"{v:.2f}"
                else:
                    text = str(v)
                self.stats_table.setItem(row, col, QTableWidgetItem(text))

    def display_rate_details(self, rate_details: pd.DataFrame):
        """显示速率详表"""
        if rate_details is None or rate_details.empty:
            return
        key_columns = ['序号', '采集时间', '来源', '业务类型', '下行RLC吞吐率', '上行RLC吞吐率']
        available_columns = [col for col in key_columns if col in rate_details.columns]
        self.details_table.setColumnCount(len(available_columns))
        self.details_table.setHorizontalHeaderLabels(available_columns)
        self.details_table.setRowCount(min(len(rate_details), 1000))
        for row in range(min(len(rate_details), 1000)):
            for col, col_name in enumerate(available_columns):
                value = rate_details.iloc[row][col_name]
                if pd.isna(value):
                    value = ''
                self.details_table.setItem(row, col, QTableWidgetItem(str(value)))

    def display_cleaned_data(self, cleaned_data: pd.DataFrame):
        """显示原表整理"""
        if cleaned_data is None or cleaned_data.empty:
            return
        self.cleaned_table.setColumnCount(len(cleaned_data.columns))
        self.cleaned_table.setHorizontalHeaderLabels(cleaned_data.columns)
        self.cleaned_table.setRowCount(min(len(cleaned_data), 100))
        for row in range(min(len(cleaned_data), 100)):
            for col, col_name in enumerate(cleaned_data.columns):
                value = cleaned_data.iloc[row][col_name]
                if pd.isna(value):
                    value = ''
                self.cleaned_table.setItem(row, col, QTableWidgetItem(str(value)))

    def export_excel(self):
        """导出Excel"""
        if not hasattr(self, 'last_statistics') or self.last_statistics is None:
            QMessageBox.warning(self, '警告', '请先处理数据')
            return
        output_path, _ = QFileDialog.getSaveFileName(
            self, '保存Excel文件', '', 'Excel Files (*.xlsx)'
        )
        if output_path:
            try:
                exporter = ExcelExporter()
                exporter.export(output_path, self.last_cleaned, self.last_rate_details, self.last_statistics)
                self.log_message(f"Excel已导出: {output_path}")
                QMessageBox.information(self, '成功', 'Excel文件导出成功！')
            except Exception as e:
                QMessageBox.critical(self, '错误', f'导出失败: {str(e)}')

    def add_business_rule(self):
        """添加业务规则（9列）"""
        row = self.rules_table.rowCount()
        self.rules_table.insertRow(row)
        for col in range(8):
            self.rules_table.setItem(row, col, QTableWidgetItem(''))
        enabled_item = QTableWidgetItem()
        enabled_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
        enabled_item.setCheckState(Qt.Checked)
        self.rules_table.setItem(row, 8, enabled_item)

    def delete_business_rule(self):
        """删除业务规则"""
        current_row = self.rules_table.currentRow()
        if current_row >= 0:
            self.rules_table.removeRow(current_row)

    def log_message(self, message: str):
        """输出日志消息"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.log_text.append(f"[{timestamp}] {message}")


# ==================== 主程序入口 ====================

def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
