#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
V5 业务识别算法（完善版）
- 两层段检测：大业务(严格阈值) + 小业务(空隙低阈值补检测)
- F列持续时长根据H列业务段时间计算
- 新增"速率(Mbps)"列
- H列识别不参考基准列(J-P)
- 通过表头名称定位列
"""
import pandas as pd
import numpy as np
import re
from datetime import datetime, timedelta
from collections import Counter, defaultdict

INPUT_FILE = '联通/mmf20260703-电信-合并-带基准_V3.xlsx'
OUTPUT_FILE = '联通/mmf20260703-电信-合并-带基准_V5.xlsx'
QCI_SELECTED = 7

# ============ 阈值 ============
# 大业务活跃阈值
DL_ACTIVE_BIG = 50.0
UL_ACTIVE_BIG = 10.0
GAP_MERGE = 2.0  # 非活跃间隙<=2秒合并

# 小业务活跃阈值（空隙补检测）
DL_ACTIVE_SMALL = 5.0
UL_ACTIVE_SMALL = 5.0

# FTP（放宽到13秒，含爬坡的FTP可能达12秒）
FTP_DUR_MIN, FTP_DUR_MAX = 8.5, 13.0
FTP_PAIR_GAP = 15.0

# 商店大文件
STORE_L_DL_MIN, STORE_L_DL_MAX = 8800, 10100

# 微信大包
WX_L_DUR_MIN, WX_L_DUR_MAX = 15, 60
WX_L_UL_MIN = 800

# 间隔规律（辅助验证）
INTERVAL_RULES = {
    ('FTP下载', 'FTP上传'): 9.0, ('FTP上传', 'FTP下载'): 9.0,
    ('FTP下载', '应用商店小文件下载'): 100.0, ('FTP上传', '应用商店小文件下载'): 100.0,
    ('应用商店小文件下载', '应用商店大文件下载'): 23.0,
    ('应用商店大文件下载', '微信小包发送'): 90.0,
    ('微信小包发送', '微信大包发送'): 18.0,
    ('微信大包发送', 'FTP下载'): 22.0, ('微信大包发送', 'FTP上传'): 22.0,
}


def parse_time(t):
    try:
        t = str(t).strip()
        m = re.match(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\((\d+)\)', t)
        if m:
            return datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S') + timedelta(milliseconds=int(m.group(2)))
        return None
    except:
        return None


def find_col(df, *keywords):
    for col in df.columns:
        if all(kw in str(col) for kw in keywords):
            return col
    return None


def detect_segments(ts_sec, dl_sec, ul_sec, dl_thresh, ul_thresh, gap_merge, range_start=None, range_end=None):
    """通用段检测：DL>dl_thresh或UL>ul_thresh的连续秒"""
    active = (dl_sec > dl_thresh) | (ul_sec > ul_thresh)
    segments = []
    i = 0
    n = len(ts_sec)
    while i < n:
        # 限定范围
        if range_start is not None and ts_sec[i] < range_start:
            i += 1
            continue
        if range_end is not None and ts_sec[i] > range_end:
            break
        if active[i]:
            si = i; j = i + 1
            while j < n:
                if range_end is not None and ts_sec[j] > range_end:
                    break
                if not active[j]:
                    k = j
                    while k < n and not active[k]:
                        g = (ts_sec[k] - ts_sec[j-1]) / np.timedelta64(1, 's')
                        if g > gap_merge:
                            break
                        k += 1
                    if k < n and active[k] and (range_end is None or ts_sec[k] <= range_end):
                        j = k
                        continue
                    else:
                        break
                j += 1
            st = ts_sec[si]; et = ts_sec[j-1]
            dur = (et - st) / np.timedelta64(1, 's') + 1
            dl_sum = float(dl_sec[si:j].sum())
            ul_sum = float(ul_sec[si:j].sum())
            dl_peak = float(dl_sec[si:j].max())
            ul_peak = float(ul_sec[si:j].max())
            segments.append({
                'st': st, 'et': et, 'dur': dur,
                'dl': dl_sum, 'ul': ul_sum,
                'dl_peak': dl_peak, 'ul_peak': ul_peak,
            })
            i = j
        else:
            i += 1
    return segments


def main():
    print("=" * 80)
    print("V5 业务识别算法（完善版）")
    print("=" * 80)

    df = pd.read_excel(INPUT_FILE)
    print(f"\n[1] 读取: {len(df)} 行")

    time_col = find_col(df, 'Time')
    qci_col = find_col(df, 'QCI')
    dl_rlc_col = find_col(df, 'Downlink', 'RLC', 'Throughput')
    ul_rlc_col = find_col(df, 'Uplink', 'RLC', 'Throughput')
    assert all([time_col, qci_col, dl_rlc_col, ul_rlc_col]), "列定位失败"

    df['_c_time'] = df[time_col].apply(parse_time)
    df['_dl_mbps'] = pd.to_numeric(df[dl_rlc_col], errors='coerce') / 1e6
    df['_ul_mbps'] = pd.to_numeric(df[ul_rlc_col], errors='coerce') / 1e6

    # QCI过滤
    df7 = df[df[qci_col] == QCI_SELECTED].copy().sort_values('_c_time').reset_index(drop=True)
    print(f"[2] QCI={QCI_SELECTED}: {len(df7)} 行")

    # 按秒聚合
    df7['_sec'] = df7['_c_time'].dt.floor('s')
    sec_agg = df7.groupby('_sec').agg({'_dl_mbps': 'max', '_ul_mbps': 'max'}).reset_index().sort_values('_sec').reset_index(drop=True)
    print(f"[3] 按秒聚合: {len(sec_agg)} 秒")

    dl_sec = sec_agg['_dl_mbps'].fillna(0).values
    ul_sec = sec_agg['_ul_mbps'].fillna(0).values
    ts_sec = sec_agg['_sec'].values

    # ===== 大业务段检测 =====
    big_segs = detect_segments(ts_sec, dl_sec, ul_sec, DL_ACTIVE_BIG, UL_ACTIVE_BIG, GAP_MERGE)
    print(f"[4] 大业务段: {len(big_segs)} 个")

    # 大业务分类（FTP候选优先，配对后再判断商店大文件）
    def classify_big(seg):
        dur = seg['dur']; dl = seg['dl']; ul = seg['ul']
        is_dl = dl > ul
        # FTP候选（时长8.5-11.5秒，可能是FTP也可能是商店大文件，由配对决定）
        if FTP_DUR_MIN <= dur <= FTP_DUR_MAX:
            return 'FTP候选'
        # 商店大文件（非FTP时长，下载量8800-10100）
        if is_dl and STORE_L_DL_MIN <= dl <= STORE_L_DL_MAX and dur > 6:
            return '应用商店大文件下载'
        # 微信大包
        if WX_L_DUR_MIN <= dur <= WX_L_DUR_MAX and not is_dl and ul >= WX_L_UL_MIN:
            return '微信大包发送'
        return None

    for seg in big_segs:
        seg['biz初步'] = classify_big(seg)

    # ===== FTP候选配对 =====
    ftp_candidates = [s for s in big_segs if s['biz初步'] == 'FTP候选']
    paired = set()
    ftp_pairs = []
    for i in range(len(ftp_candidates)):
        if id(ftp_candidates[i]) in paired:
            continue
        for j in range(i+1, len(ftp_candidates)):
            if id(ftp_candidates[j]) in paired:
                continue
            gap = (ftp_candidates[j]['st'] - ftp_candidates[i]['et']) / np.timedelta64(1, 's')
            if 0 <= gap < FTP_PAIR_GAP:
                ftp_pairs.append((ftp_candidates[i], ftp_candidates[j]))
                paired.add(id(ftp_candidates[i]))
                paired.add(id(ftp_candidates[j]))
                break
            elif gap >= FTP_PAIR_GAP:
                break

    print(f"[5] FTP候选: {len(ftp_candidates)}, 配对: {len(ftp_pairs)} 对")

    def ftp_direction(seg):
        return 'FTP下载' if seg['dl'] > seg['ul'] else 'FTP上传'

    # 未配对FTP候选：下载量在商店大文件范围→商店大文件；否则→单独FTP
    for seg in ftp_candidates:
        if id(seg) not in paired:
            if STORE_L_DL_MIN <= seg['dl'] <= STORE_L_DL_MAX and seg['dl'] > seg['ul']:
                seg['biz初步'] = '应用商店大文件下载'
            else:
                seg['biz初步'] = ftp_direction(seg)

    # ===== 所有业务段（按时间排序）=====
    all_segs = sorted(big_segs, key=lambda x: x['st'])

    # 为每个段标记业务类型和轮次
    for seg in all_segs:
        seg['轮次'] = None
        seg['业务'] = seg['biz初步']

    # FTP对分配轮次
    round_num = 0
    pair_info = []
    for ftp1, ftp2 in ftp_pairs:
        round_num += 1
        ftp1['轮次'] = round_num; ftp1['业务'] = ftp_direction(ftp1)
        ftp2['轮次'] = round_num; ftp2['业务'] = ftp_direction(ftp2)
        pair_info.append((round_num, ftp2['et'], ftp1['st']))

    # 大业务（非FTP）分配到轮次
    for idx, (rnd, pair_end, _) in enumerate(pair_info):
        next_start = pair_info[idx+1][2] if idx+1 < len(pair_info) else (all_segs[-1]['et'] + np.timedelta64(1, 's') if all_segs else pair_end)
        for seg in all_segs:
            if seg['et'] > pair_end and seg['st'] < next_start:
                if seg['轮次'] is None and seg['业务'] in ('应用商店大文件下载', '微信大包发送'):
                    seg['轮次'] = rnd

    # 文件开头（FTP对之前）
    if pair_info:
        first_start = pair_info[0][2]
        for seg in all_segs:
            if seg['et'] <= first_start and seg['业务'] in ('应用商店大文件下载', '微信大包发送'):
                seg['轮次'] = 0  # 不算轮次

    # ===== 小业务补检测（按位置约束：在每个大业务前找对应小业务）=====
    typed_big = sorted(
        [s for s in all_segs if s['业务'] in ('FTP下载', 'FTP上传', '应用商店大文件下载', '微信大包发送')],
        key=lambda x: x['st']
    )

    # 商店大文件前面→商店小文件，微信大包前面→微信小包
    PRECEDING_SMALL = {
        '应用商店大文件下载': '应用商店小文件下载',
        '微信大包发送': '微信小包发送',
    }

    small_segs = []

    for idx, (rnd, pair_end, pair_start) in enumerate(pair_info):
        window_end = pair_info[idx+1][2] if idx+1 < len(pair_info) else (typed_big[-1]['et'] + np.timedelta64(1, 's') if typed_big else pair_end)

        big_in_window = sorted(
            [s for s in typed_big if s['st'] >= pair_end and s['st'] < window_end and s['业务'] in ('应用商店大文件下载', '微信大包发送')],
            key=lambda x: x['st']
        )

        prev_end = pair_end
        for big_seg in big_in_window:
            small_type = PRECEDING_SMALL.get(big_seg['业务'])
            if small_type and prev_end < big_seg['st']:
                gap_dur = (big_seg['st'] - prev_end) / np.timedelta64(1, 's')
                if gap_dur > 1:
                    segs = detect_segments(ts_sec, dl_sec, ul_sec, DL_ACTIVE_SMALL, UL_ACTIVE_SMALL, GAP_MERGE, prev_end, big_seg['st'])
                    valid = [s for s in segs if s['dl'] + s['ul'] >= 10 and s['dur'] >= 0.5]
                    if small_type == '应用商店小文件下载':
                        dl_s = [s for s in valid if s['dl'] > s['ul']]
                        if dl_s:
                            best = max(dl_s, key=lambda x: x['dl'])
                            best['业务'] = small_type; best['轮次'] = rnd
                            small_segs.append(best)
                    else:
                        ul_s = [s for s in valid if s['ul'] > s['dl']]
                        if ul_s:
                            best = max(ul_s, key=lambda x: x['ul'])
                            best['业务'] = small_type; best['轮次'] = rnd
                            small_segs.append(best)
            prev_end = big_seg['et']

    # 合并所有段（轮次=0的文件开头段不标记）
    all_segs_final = sorted(all_segs + small_segs, key=lambda x: x['st'])
    for seg in all_segs_final:
        seg['有效'] = seg['业务'] is not None
        # 轮次=0或None的段不标记（在第一个完整6业务开始之前）
        if seg.get('轮次') is None or seg.get('轮次') == 0:
            seg['有效'] = False

    # ===== 每轮最多6个业务（位置约束）=====
    from collections import defaultdict
    round_segs_map = defaultdict(list)
    for seg in all_segs_final:
        if seg['有效'] and seg['业务']:
            rnd = seg.get('轮次')
            if rnd is not None and rnd > 0:
                round_segs_map[rnd].append(seg)

    for rnd, segs in round_segs_map.items():
        segs_sorted = sorted(segs, key=lambda x: x['st'])
        # 每类业务最多1个，总共最多6个
        seen_types = set()
        count = 0
        for seg in segs_sorted:
            biz = seg['业务']
            if biz in seen_types or count >= 6:
                seg['有效'] = False  # 超过6个或重复类型，标记无效
            else:
                seen_types.add(biz)
                count += 1

    valid_biz = [s for s in all_segs_final if s['有效'] and s['业务']]
    print(f"[6] 有效业务: {len(valid_biz)} 个")
    cnt = Counter(s['业务'] for s in valid_biz)
    for k, v in sorted(cnt.items()):
        print(f"  {k}: {v} 个")
    print(f"  总轮数: {round_num}")

    # ===== 标记输出（标记范围=段内高速行，从第一高速行到最后高速行）=====
    biz_col = '业务识别'
    df[biz_col] = pd.Series([None] * len(df), dtype='object')

    dur_col = '持续时长'
    df[dur_col] = pd.Series([None] * len(df), dtype='object')  # 清空旧值

    rate_col = '速率(Mbps)'
    df[rate_col] = pd.Series([None] * len(df), dtype='object')

    dl_rate_col = '下载速率'
    ul_rate_col = '上传速率'
    if dl_rate_col not in df.columns:
        df[dl_rate_col] = df['_dl_mbps']
    if ul_rate_col not in df.columns:
        df[ul_rate_col] = df['_ul_mbps']

    # 段信息补充：高速部分的起止时间和时长
    biz_segments = sorted(valid_biz, key=lambda x: x['st'])
    for seg in biz_segments:
        st = seg['st']; et = seg['et']
        biz = seg['业务']
        seg_mask = (df['_c_time'] >= st) & (df['_c_time'] <= et + timedelta(seconds=1)) & (df[qci_col] == QCI_SELECTED)

        # 下载/上行速率列
        if biz in ('FTP下载', '应用商店小文件下载', '应用商店大文件下载'):
            vals_series = pd.to_numeric(df.loc[seg_mask, '_dl_mbps'], errors='coerce').fillna(0)
            high_thresh = 50
        else:
            vals_series = pd.to_numeric(df.loc[seg_mask, '_ul_mbps'], errors='coerce').fillna(0)
            high_thresh = 10

        seg_rows = df[seg_mask]
        non_zero_mask = vals_series > 0
        high_mask_vals = vals_series > high_thresh

        if non_zero_mask.any():
            first_idx = non_zero_mask.idxmax()  # 第一个非零行索引
            # 尾部：找最后一个大值（>阈值），排除尾部<10的小值
            high_indices = vals_series[high_mask_vals].index
            last_vals = vals_series.iloc[-5:]  # 尾部最后几个值
            last_small = any(v > 0 and v < 10 for v in last_vals)
            if last_small and len(high_indices) > 0:
                last_idx = high_indices[-1]
            elif high_mask_vals.any():
                last_idx = high_mask_vals[high_mask_vals].index[-1]
            else:
                last_idx = non_zero_mask[non_zero_mask].index[-1]

            first_time = seg_rows.loc[first_idx, '_c_time']
            last_time = seg_rows.loc[last_idx, '_c_time']
        elif high_mask_vals.any():
            high_rows = seg_rows[high_mask_vals.values]
            first_time = high_rows['_c_time'].min()
            last_time = high_rows['_c_time'].max()
        else:
            first_time = seg_rows['_c_time'].min()
            last_time = seg_rows['_c_time'].max()

        # 时长：只有微信小包+1
        mark_dur = (last_time - first_time).total_seconds()
        if biz == '微信小包发送':
            mark_dur += 1

        # 扩展标记开始：大文件往前5秒找非零行（含爬坡）
        look_secs = 5 if biz in ('FTP下载', 'FTP上传', '应用商店大文件下载') else 2
        look_start = first_time - timedelta(seconds=look_secs)
        prev_mask = (df['_c_time'] >= look_start) & (df['_c_time'] < first_time) & (df[qci_col] == QCI_SELECTED)
        if prev_mask.any():
            pv_col = '_dl_mbps' if biz in ('FTP下载','应用商店小文件下载','应用商店大文件下载') else '_ul_mbps'
            pv = df.loc[prev_mask, pv_col].fillna(0)
            if (pv > 0).any():
                first_time = df.loc[pv[pv > 0].index[0], '_c_time']
        # 标记范围：大业务含爬坡零值，小业务排除<0.5噪声
        is_big = biz in ('FTP下载', 'FTP上传', '应用商店大文件下载', '微信大包发送')
        full_mask = (df['_c_time'] >= first_time) & (df['_c_time'] <= last_time) & (df[qci_col] == QCI_SELECTED)
        if is_big:
            mark_mask = full_mask
        else:
            col_name = '_dl_mbps' if biz == '应用商店小文件下载' else '_ul_mbps'
            full_vals = pd.to_numeric(df.loc[full_mask, col_name], errors='coerce').fillna(0)
            nz = pd.Series(False, index=df.index)
            nz.loc[full_mask[full_mask].index] = (full_vals > 0.5).values
            mark_mask = full_mask & nz
            if mark_mask.any():
                mt = df.loc[mark_mask, '_c_time']
                first_time = mt.min()
                mark_dur = (last_time - first_time).total_seconds()
                if biz == '微信小包发送':
                    mark_dur += 1

        df.loc[mark_mask, biz_col] = biz
        df.loc[mark_mask, dur_col] = round(mark_dur, 3)

        # 速率：排除最后一个<10的小值后求和/时长
        mark_df = df[mark_mask]
        if biz in ('FTP下载', '应用商店小文件下载', '应用商店大文件下载'):
            vals = pd.to_numeric(mark_df[dl_rate_col], errors='coerce').fillna(0).tolist()
        else:
            vals = pd.to_numeric(mark_df[ul_rate_col], errors='coerce').fillna(0).tolist()
        if len(vals) >= 2 and vals[-1] < 10 and all(v > 100 for v in vals[:-1] if v > 0):
            vals = vals[:-1]
        total = sum(vals)
        avg_rate = total / mark_dur if mark_dur > 0 else 0
        df.loc[mark_mask, rate_col] = round(avg_rate, 2)

        seg['mark_st'] = first_time
        seg['mark_et'] = last_time
        seg['mark_dur'] = mark_dur

    # 第一个FTP对前的区域清除标记
    if pair_info:
        first_ftp = pair_info[0][2]
        first_ftp_ts = pd.Timestamp(first_ftp)
        for col_name in [biz_col, dur_col, rate_col]:
            if col_name in df.columns:
                mask_clear = df['_c_time'] < first_ftp_ts
                df.loc[mask_clear, col_name] = None

    # ===== 轮次完整性检查 + 轮次列标记 =====
    REQUIRED_BIZ = {'FTP下载', 'FTP上传', '应用商店小文件下载', '应用商店大文件下载', '微信小包发送', '微信大包发送'}

    # 收集每个轮次的业务类型
    round_biz_types = {}  # {轮次号: set(业务类型)}
    for seg in valid_biz:
        rnd = seg.get('轮次')
        if rnd is not None and rnd > 0:
            if rnd not in round_biz_types:
                round_biz_types[rnd] = set()
            round_biz_types[rnd].add(seg['业务'])

    # 检查完整性，生成轮次标签
    round_labels = {}
    incomplete_idx = 0
    for rnd in sorted(round_biz_types.keys()):
        if round_biz_types[rnd] >= REQUIRED_BIZ:
            round_labels[rnd] = str(rnd)
        else:
            incomplete_idx += 1
            round_labels[rnd] = f"不完整{incomplete_idx}"

    # 文件开头（轮次=0或None，FTP缺失）→ 不完整轮次（单独编号）
    head_incomplete_idx = 0
    head_seg_rounds = set()
    for seg in valid_biz:
        rnd = seg.get('轮次')
        if rnd is None or rnd == 0:
            head_seg_rounds.add(id(seg))
    if head_seg_rounds:
        head_incomplete_idx += 1
        head_label = f"不完整{incomplete_idx + head_incomplete_idx}"

    print(f"\n[轮次完整性] 完整轮次: {sum(1 for v in round_labels.values() if not v.startswith('不完整'))}, 不完整: {sum(1 for v in round_labels.values() if v.startswith('不完整'))}")

    # 标记轮次列（用高速范围mark_st/mark_et）
    round_col = '轮次'
    df[round_col] = pd.Series([None] * len(df), dtype='object')
    for seg in valid_biz:
        st = seg.get('mark_st', seg['st'])
        et = seg.get('mark_et', seg['et'])
        rnd = seg.get('轮次')
        if rnd is not None and rnd > 0:
            label = round_labels.get(rnd, str(rnd))
        elif rnd == 0 or rnd is None:
            label = head_label if head_seg_rounds else None
        else:
            label = None
        if label:
            mask = (df['_c_time'] >= st) & (df['_c_time'] <= et) & (df[qci_col] == QCI_SELECTED)
            df.loc[mask, round_col] = label

    # ===== H列(临时项)空值标记：I列(业务识别)和K列(基准_业务类型)都为空→填"空值"=====
    temp_col = '临时项'
    df[temp_col] = pd.Series([None] * len(df), dtype='object')  # 强制object，覆盖旧值
    # 找基准_业务类型列
    base_biz_col = find_col(df, '基准', '业务类型')
    if base_biz_col is None:
        base_biz_col = '基准_业务类型'
    for i in range(len(df)):
        biz_val = df.at[i, biz_col] if pd.notna(df.at[i, biz_col]) else None
        base_val = df.at[i, base_biz_col] if base_biz_col in df.columns and pd.notna(df.at[i, base_biz_col]) else None
        if not biz_val and not base_val:
            df.at[i, temp_col] = '空值'

    # ===== 重新排列列（轮次列在H列业务识别后）=====
    cols = df.columns.tolist()
    time_idx = cols.index(time_col)
    prefix = cols[:time_idx + 1]
    used = set(prefix) | {'_c_time', '_dl_mbps', '_ul_mbps', '_sec'}

    new_cols = list(prefix)
    for c in ['下载速率', '上传速率', rate_col, dur_col, temp_col, biz_col, round_col]:
        if c in df.columns and c not in used:
            new_cols.append(c)
            used.add(c)
    for c in cols:
        if c not in used:
            new_cols.append(c)
            used.add(c)

    df_out = df[new_cols]

    # ===== 构建"详细过程"sheet数据 =====
    # 参考FTP_DETAIL_1样式：index, Time, 下载速率, 上传速率, 速率(Mbps), 持续时长, 轮次, businessType, 然后其余列(不含临时项)
    detail_cols = ['下载速率', '上传速率', rate_col, dur_col, '轮次']
    biz_type_col_name = 'businessType'
    df_out[biz_type_col_name] = df_out[biz_col]  # 复制业务识别为businessType

    # 详细过程的列序: A:id(idx), B:Time, C-F:下载/上传/速率/时长, G:轮次, H:businessType, I+:其余
    detail_prefix = list(df_out.columns[:time_idx+1])  # idx, Source, Time
    # 去掉Source列，用index作为id(仿FTP_DETAIL_1的id列)
    id_series = pd.Series(range(1, len(df_out)+1), name='id')
    detail_prefix2 = ['id', 'Time']
    detail_extra = detail_cols + [biz_type_col_name]
    used_detail = set(detail_prefix2 + detail_extra + ['Source', '临时项'])
    for c in detail_cols:
        if c not in df_out.columns:
            used_detail.discard(c)
    remaining_detail = [c for c in df_out.columns if c not in used_detail and c != 'Source']
    df_detail = pd.DataFrame({'id': id_series, 'Time': df_out['Time']})
    for c in detail_cols:
        if c in df_out.columns:
            df_detail[c] = df_out[c].values
    df_detail[biz_type_col_name] = df_out[biz_col].values
    for c in remaining_detail:
        if c in df_out.columns and c not in df_detail.columns:
            df_detail[c] = df_out[c].values
    # 保留_c_time用于汇总时长计算
    df_detail['_c_time'] = df['_c_time'].values  # 用于时长计算

    # ===== 构建"汇总"sheet数据 =====
    # 从详细过程计算：6类业务，13个指标
    DL_CLIP = 1000  # 下行削峰阈值（可调）
    UL_CLIP = 200   # 上行削峰阈值（可调）
    dl_biz = {'FTP下载', '应用商店小文件下载', '应用商店大文件下载'}
    ul_biz = {'FTP上传', '微信小包发送', '微信大包发送'}

    summary_rows = []
    for biz in sorted(dl_biz | ul_biz):
        sub = df_detail[df_detail[biz_type_col_name]==biz].copy()
        if len(sub) == 0: continue
        # RLC列（用于计算指标）
        dl_rlc_name = 'Downlink RLC Throughput(bps)'
        ul_rlc_name = 'Uplink RLC Throughput(bps)'
        if dl_rlc_name not in sub.columns: dl_rlc_name = [c for c in sub.columns if 'Downlink RLC' in c][0] if any('Downlink RLC' in c for c in sub.columns) else None
        if ul_rlc_name not in sub.columns: ul_rlc_name = [c for c in sub.columns if 'Uplink RLC' in c][0] if any('Uplink RLC' in c for c in sub.columns) else None

        # 速率值(Mbps) = RLC bps/1e6，排除0值
        if biz in dl_biz and dl_rlc_name:
            rates = pd.to_numeric(sub[dl_rlc_name], errors='coerce').fillna(0) / 1e6
            rates = rates[rates > 0]
        elif biz in ul_biz and ul_rlc_name:
            rates = pd.to_numeric(sub[ul_rlc_name], errors='coerce').fillna(0) / 1e6
            rates = rates[rates > 0]
        else:
            rates = pd.Series([], dtype=float)

        # 时长：按连续段（gap>5s断开，首尾差+1s），用_c_time（已解析的时间列）
        time_col_name = '_c_time' if '_c_time' in sub.columns else time_col
        sub_sorted = sub.dropna(subset=[time_col_name]).sort_values(time_col_name)
        t_series = sub_sorted[time_col_name].dropna()
        durations = []
        if len(t_series) > 0:
            st = t_series.iloc[0]; pt = st
            for t in t_series.iloc[1:]:
                if (t - pt).total_seconds() > 5:
                    durations.append((pt - st).total_seconds() + 1)
                    st = t
                pt = t
            durations.append((pt - st).total_seconds() + 1)

        # 计算指标
        row = {'业务类型': biz}
        if biz in dl_biz and len(rates) > 0:
            row['应用层FTP下载速率1000M以上占比(%)'] = round((rates > 1000).sum() / len(rates) * 100, 2)
            row['应用层FTP下载速率100M以下占比(%)'] = round((rates < 100).sum() / len(rates) * 100, 2)
            row['应用层平均下载速率(Mbps)'] = round(rates.mean(), 2)
            row['削峰应用层平均下载速率(Mbps)'] = round(rates.clip(upper=DL_CLIP).mean(), 2)
            sorted_vals = rates.sort_values(ascending=False)
            top_n = max(1, int(len(sorted_vals) * 0.1))
            row['下行削峰TOP10%峰值速率'] = round(sorted_vals.head(top_n).mean(), 2)
        elif biz in ul_biz and len(rates) > 0:
            row['应用层FTP上传速率200M以上占比(%)'] = round((rates > 200).sum() / len(rates) * 100, 2)
            row['应用层FTP上传速率5M以下占比(%)'] = round((rates < 5).sum() / len(rates) * 100, 2)
            row['应用层平均上传速率'] = round(rates.mean(), 2)
            row['削峰应用层平均上传速率(Mbps)'] = round(rates.clip(upper=UL_CLIP).mean(), 2)
            sorted_vals = rates.sort_values(ascending=False)
            top_n = max(1, int(len(sorted_vals) * 0.1))
            row['上行削峰TOP10%峰值速率'] = round(sorted_vals.head(top_n).mean(), 2)
        if durations:
            row['业务时长平均值(s)'] = round(np.mean(durations), 2)
            row['业务时长中位值(s)'] = round(np.median(durations), 0)
        summary_rows.append(row)

    df_summary = pd.DataFrame(summary_rows, columns=[
        '业务类型', '应用层FTP下载速率1000M以上占比(%)', '应用层FTP下载速率100M以下占比(%)',
        '应用层平均下载速率(Mbps)', '削峰应用层平均下载速率(Mbps)', '下行削峰TOP10%峰值速率',
        '应用层FTP上传速率200M以上占比(%)', '应用层FTP上传速率5M以下占比(%)',
        '应用层平均上传速率', '削峰应用层平均上传速率(Mbps)', '上行削峰TOP10%峰值速率',
        '业务时长平均值(s)', '业务时长中位值(s)',
    ])

    # ===== 构建"对比"sheet（已有逻辑，加排序）=====
    baseline_all = pd.read_excel('工信部业务指标_呼叫详情_整理.xlsx', sheet_name='电信')
    baseline_sorted = baseline_all.sort_values('开始时间').reset_index(drop=True)
    compare_rows = []
    for i, row in baseline_sorted.iterrows():
        compare_rows.append({'轮次': i//6+1, '来源': '基准', '业务类型': row['业务类型'],
            '开始时间': str(row['开始时间']), '结束时间': str(row['结束时间']),
            '速率类指标': row['速率类指标'], '数值': row['数值'],
            '时长类指标': row['时长类指标'], '数值2': row['数值.1']})
    rate_name_map = {
        'FTP下载': '下载平均速率(Mbps)', 'FTP上传': '上传平均速率(Mbps)',
        '应用商店小文件下载': '小文件下载速率(Mbps)', '应用商店大文件下载': '大文件下载速率(Mbps)',
        '微信小包发送': '微信小包发送速率(Mbps)', '微信大包发送': '微信大包发送速率(Mbps)',
    }
    for seg in valid_biz:
        rnd = seg.get('轮次')
        if rnd is not None and rnd > 0:
            mark_st = seg.get('mark_st', seg['st']); mark_et = seg.get('mark_et', seg['et'])
            mark_dur = seg.get('mark_dur', seg['dur'])
            mask = (df['_c_time'] >= mark_st) & (df['_c_time'] <= mark_et) & (df[qci_col] == QCI_SELECTED)
            rate_vals = pd.to_numeric(df.loc[mask, rate_col], errors='coerce').dropna()
            rv = rate_vals.iloc[0] if len(rate_vals) > 0 else None
            compare_rows.append({'轮次': rnd, '来源': '代码', '业务类型': seg['业务'],
                '开始时间': mark_st.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3] if hasattr(mark_st, 'strftime') else str(mark_st),
                '结束时间': mark_et.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3] if hasattr(mark_et, 'strftime') else str(mark_et),
                '速率类指标': rate_name_map.get(seg['业务'], ''), '数值': rv,
                '时长类指标': '业务时长(秒)', '数值2': round(mark_dur, 3)})
    df_compare = pd.DataFrame(compare_rows).sort_values(['轮次', '业务类型']).reset_index(drop=True)
    # 重命名数值2
    if '数值2' in df_compare.columns:
        df_compare = df_compare.rename(columns={'数值2': '数值.1'})

    # ===== 写入Excel（3 sheet）=====
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df_detail.to_excel(writer, sheet_name='详细过程', index=False)
        df_compare.to_excel(writer, sheet_name='对比', index=False)
        df_summary.to_excel(writer, sheet_name='汇总', index=False)

        wb = writer.book
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill

        # ---- 详细过程格式 ----
        ws_d = writer.sheets['详细过程']
        ws_d.freeze_panes = 'C2'
        # D/E列整数
        for cl in ['D','E']:
            if cl in {get_column_letter(i+1):True for i in range(15)}:
                for cell in ws_d[cl]:
                    if cell.row > 1: cell.number_format = '0'

        # ---- 对比格式：排序+颜色 ----
        ws_c = writer.sheets['对比']
        biz_c = ['C6E0B4', 'BDD7EE']
        seen_c = {}; pb = None; ci = 0
        for rn in range(2, len(df_compare)+2):
            v = ws_c.cell(row=rn, column=1).value  # A列轮次
            if v and v != pb:
                if v not in seen_c: seen_c[v] = biz_c[len(seen_c)%2]
                pb = v
            if v and v in seen_c:
                fill = PatternFill(start_color=seen_c[v], end_color=seen_c[v], fill_type='solid')
                for cc in range(1, 7): ws_c.cell(row=rn, column=cc).fill = fill

    print(f"\n[7] 已保存: {OUTPUT_FILE}")
    print("完成！")


if __name__ == '__main__':
    main()
