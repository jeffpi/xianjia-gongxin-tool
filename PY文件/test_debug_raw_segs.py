#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Deep debug: raw segments in store_l time windows
看看前3个缺失的store_l参考时间段内，gap=5切割出的原始段是什么
"""

import numpy as np
import pandas as pd
import openpyxl
from importlib import util as import_util

spec = import_util.spec_from_file_location('tool', '5G用户级公共监控速率统计工具-V1.04.py')
tool = import_util.module_from_spec(spec)
spec.loader.exec_module(tool)

uproc = tool.UCMProcessor({'dl_peak_limit': 900, 'ul_peak_limit': 160})
paths = ['联通/mmf20260703115328-电信.xlsx','联通/mmf20260703115334-电信.xlsx','联通/mmf20260703115336-联通.xlsx']
params = {'dl_peak_limit': 900, 'ul_peak_limit': 160, 'down_min': 50, 'up_min': 10, 'match_mode': 'auto'}
agg = uproc.parse_mmf(paths, params)

# 参考store_l时间段
wb = openpyxl.load_workbook('联通/呼叫详情_按时间排序-20260707.xlsx', data_only=True)
ws = wb['联通']
ref_entries = []
for r in range(2, ws.max_row+1):
    biz = str(ws.cell(r,2).value or '')
    if '应用商店' not in biz: continue
    ref_entries.append({'st':ws.cell(r,3).value, 'et':ws.cell(r,4).value, 'big':ws.cell(r,11).value})
ref_sl = [ref_entries[i] for i in range(2, len(ref_entries), 3)]

n = len(agg)
dl_min, ul_min = 50, 10

# 全量段切割
pos = 0
raw_segs = []
while pos < n:
    if agg.at[pos, 'dl_mac'] > dl_min or agg.at[pos, 'ul_mac'] > ul_min:
        start_i = pos; gap = 0
        while pos < n and gap < 5:
            if agg.at[pos, 'dl_mac'] > dl_min or agg.at[pos, 'ul_mac'] > ul_min:
                gap = 0
            else:
                gap += 1
            pos += 1
        end_i = max(start_i, pos-1-gap)
        seg_df = agg.loc[start_i:end_i]
        ul_act = int((seg_df['ul_mac']>10).sum())
        dl_act = int((seg_df['dl_mac']>50).sum())
        direction = '上行' if ul_act > dl_act+3 else '下行'
        col = 'dl_mac' if direction=='下行' else 'ul_mac'
        flow = seg_df[col].sum()/8
        dur = max(round((seg_df['t'].iloc[-1]-seg_df['t'].iloc[0]).total_seconds(),1), 0.5)
        rv = seg_df['dl_rlc' if direction=='下行' else 'ul_rlc'].replace(0,np.nan).dropna()
        rate = round(float(rv.mean()),3) if len(rv) else 0
        raw_segs.append({'start_t':seg_df['t'].iloc[0], 'end_t':seg_df['t'].iloc[-1], 'direction':direction, 'flow_mb':flow, 'duration':dur, 'rate':rate, 'dl_act':dl_act, 'ul_act':ul_act})
    else:
        pos += 1

# gap<10合并
merged = []
for s in raw_segs:
    if s.get('duration', 0) < 1 and s.get('flow_mb', 0) < 10:
        continue
    if not merged:
        merged.append(s.copy())
        continue
    prev = merged[-1]
    gap = (s['start_t'] - prev['end_t']).total_seconds()
    if s['direction'] == prev['direction'] and 0 < gap < 10:
        seg_df = agg.loc[prev['start_i'] if 'start_i' in prev else 0:s['end_i'] if 'end_i' in s else 0]
        # 但prev没有start_i/end_i了，重新算
        # 简化：用start_t/end_t找agg范围
        mask = (agg['t'] >= prev['start_t']) & (agg['t'] <= s['end_t'])
        seg_df2 = agg[mask]
        col = 'dl_mac' if s['direction']=='下行' else 'ul_mac'
        flow = seg_df2[col].sum()/8
        dur = max(round((seg_df2['t'].iloc[-1]-seg_df2['t'].iloc[0]).total_seconds(),1), 0.5)
        rv = seg_df2['dl_rlc' if s['direction']=='下行' else 'ul_rlc'].replace(0,np.nan).dropna()
        rate = round(float(rv.mean()),3) if len(rv) else 0
        merged[-1] = {'start_t': prev['start_t'], 'end_t': s['end_t'], 'direction': s['direction'], 'flow_mb': flow, 'duration': dur, 'rate': rate}
    else:
        merged.append(s.copy())

# 对每个缺失的store_l参考时间段，看原始段和合并段
for idx in [0, 1, 2, 5]:
    if idx >= len(ref_sl): continue
    sl = ref_sl[idx]
    rst = pd.to_datetime(sl['st'])
    ret = pd.to_datetime(sl['et'])
    print(f'SL#{idx} ref时间: {rst.strftime("%H:%M:%S")}-{ret.strftime("%H:%M:%S")}')

    # 原始段
    raw_overlap = [s for s in raw_segs if s['start_t'] <= ret and s['end_t'] >= rst]
    print(f'  === 原始段 (gap=5切割, 无合并) ===')
    for s in raw_overlap:
        f = s['flow_mb']
        if s['direction'] == '下行':
            if f <= 50: k='None'
            elif f < 350: k='store_s'
            elif f < 700: k='ftp_dl'
            elif f >= 1500: k='store_l'
            elif s['duration'] >= 15: k='store_l'
            elif s['rate'] < 500: k='store_l'
            else: k='ftp_dl'
        else:
            k = '上行'
        print(f'    {s["direction"]:4s}: flow={f:6.0f}MB dur={s["duration"]:5.1f}s rate={s["rate"]:6.1f} -> {k}  ({s["start_t"].strftime("%H:%M:%S")}-{s["end_t"].strftime("%H:%M:%S")})')

    # 合并后段
    merged_overlap = [s for s in merged if s['start_t'] <= ret and s['end_t'] >= rst]
    print(f'  === 合并后 (gap<10合并) ===')
    for s in merged_overlap:
        f = s['flow_mb']
        if s['direction'] == '下行':
            if f <= 50: k='None'
            elif f < 350: k='store_s'
            elif f < 700: k='ftp_dl'
            elif f >= 1500: k='store_l'
            elif s['duration'] >= 15: k='store_l'
            elif s['rate'] < 500: k='store_l'
            else: k='ftp_dl'
        else:
            k = '上行'
        print(f'    {s["direction"]:4s}: flow={f:6.0f}MB dur={s["duration"]:5.1f}s rate={s["rate"]:6.1f} -> {k}  ({s["start_t"].strftime("%H:%M:%S")}-{s["end_t"].strftime("%H:%M:%S")})')
    print()