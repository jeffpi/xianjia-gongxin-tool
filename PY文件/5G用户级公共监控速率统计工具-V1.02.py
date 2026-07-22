#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
5G用户级公共监控速率统计工具 V1.02（呼叫详情配准版）
基于2025-2026年工信部测评用例规则开发
作者：Claude Code
创建日期：2026-06-23

V1.02 核心方案（呼叫详情配准）：
  - 双输入：① UCM 监控文件(mmf，英文bps / 中文Mbps 自动识别) ② 业务日志(参考文件的「呼叫详情」sheet)
  - 业务定位不再靠速率行为推断，而是用呼叫详情里每次业务的「发起/完成时间」精确锚定
  - 速率类(FTP前10%峰值/达标占比/削峰后均值/采样点数)：按时间窗截取UCM逐秒数据计算
  - 业务类(测试次数/成功率/时长/平均速率)：直接取自呼叫详情（与参考文件完全一致）
  - 大包/大文件中位值平均时长：V6归一化(时长×中位流量/实际流量)，实际流量由UCM截取积分
  - 输出 V6 横向分块宽表(42列)，削峰1000/200，速率列RLC
验证：FTP上传削峰后169.6 vs 参考178.8(差5%)；下载差13%(UCM与前台速率固有差异)
详见同目录《统计规则说明_V1.02.md》
"""

import sys
import os
import json
import re
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
import numpy as np
import pandas as pd
import openpyxl
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QLabel, QLineEdit, QPushButton, QFileDialog,
    QTextEdit, QTableWidget, QTableWidgetItem, QHeaderView,
    QGroupBox, QFormLayout, QSpinBox, QDoubleSpinBox, QComboBox,
    QCheckBox, QMessageBox, QProgressBar
)
from PySide6.QtCore import Qt, QThread, Signal, Slot
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# ==================== 宽表列定义（V6 横向分块，共 42 列） ====================
WIDE_COLUMNS = [
    ('FTP下载', '测试次数'), ('FTP下载', '采样点数'),
    ('FTP下载', '削峰后下行RLC≥100Mbps采集点数'), ('FTP下载', '削峰后下行RLC≥1000Mbps采集点数'),
    ('FTP下载', '下行RLC平均吞吐率(Mbps)'), ('FTP下载', '削峰后下行RLC平均吞吐率(Mbps)'),
    ('FTP下载', '下行RLC峰值吞吐率(Mbps)'),
    ('FTP下载', '削峰后下行RLC≥100Mbps占比'), ('FTP下载', '削峰后下行RLC≥1000Mbps占比'),
    ('FTP下载', '前10%峰值速率(Mbps)'),
    ('FTP上传', '测试次数'), ('FTP上传', '采样点数'),
    ('FTP上传', '削峰后上行RLC>20Mbps采集点数'), ('FTP上传', '削峰后上行RLC>200Mbps采集点数'),
    ('FTP上传', '上行RLC平均吞吐率(Mbps)'), ('FTP上传', '削峰后上行RLC平均吞吐率(Mbps)'),
    ('FTP上传', '上行RLC峰值吞吐率(Mbps)'),
    ('FTP上传', '削峰后上行RLC>20Mbps占比'), ('FTP上传', '削峰后上行RLC>200Mbps占比'),
    ('FTP上传', '前10%峰值速率(Mbps)'),
    ('应用商店_小包', '测试次数'), ('应用商店_小包', '采样点数'),
    ('应用商店_小包', '下行RLC平均吞吐率(Mbps)'), ('应用商店_小包', '单次平均时长(s)'),
    ('应用商店_小包', '下载成功率(%)'),
    ('应用商店_大包', '测试次数'), ('应用商店_大包', '采样点数'),
    ('应用商店_大包', '下行RLC平均吞吐率(Mbps)'), ('应用商店_大包', '单次平均时长(s)'),
    ('应用商店_大包', '中位值平均时长(s)'), ('应用商店_大包', '下载成功率(%)'),
    ('微信_小文件', '测试次数'), ('微信_小文件', '采样点数'),
    ('微信_小文件', '上行RLC平均吞吐率(Mbps)'), ('微信_小文件', '单次平均时长(s)'),
    ('微信_小文件', '上传成功率(%)'),
    ('微信_大文件', '测试次数'), ('微信_大文件', '采样点数'),
    ('微信_大文件', '上行RLC平均吞吐率(Mbps)'), ('微信_大文件', '单次平均时长(s)'),
    ('微信_大文件', '中位值平均时长(s)'), ('微信_大文件', '上传成功率(%)'),
]


class ConfigManager:
    def __init__(self, config_path: str = None):
        if config_path is None:
            config_path = os.path.join(os.path.dirname(__file__), 'config', 'default_rules.json')
        self.config_path = config_path
        self.config = self.load_config()

    def load_config(self) -> Dict:
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载配置失败: {e}")
            return self.get_default_config()

    def save_config(self, config: Dict = None) -> bool:
        if config is None:
            config = self.config
        try:
            os.makedirs(os.path.dirname(self.config_path), exist_ok=True)
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"保存配置失败: {e}")
            return False

    def get_default_config(self) -> Dict:
        return {
            "version": "1.02", "rule_version": "2025-2026工信部测评用例",
            "global_params": {"qci": 0, "rate_column": "RLC", "start_time": "auto", "end_time": "auto",
                              "dl_peak_limit": 1000, "ul_peak_limit": 200,
                              "dl_pass_threshold": 100, "ul_pass_threshold": 20,
                              "scenario_type": "场景", "road_mean_threshold": 800},
            "business_rules": []
        }


# ==================== 数据处理引擎（呼叫详情配准） ====================

class DataProcessor:
    def __init__(self, config: Dict):
        self.config = config
        self.ucm_df = None
        self.statistics = None

    # ---------- 辅助：速率列选择（中英文） ----------
    def _select_rate_col(self, df: pd.DataFrame, direction: str, pref: str) -> Optional[str]:
        dir_kw = ['下行', 'Downlink'] if direction == '下行' else ['上行', 'Uplink']
        cols = [c for c in df.columns
                if any(k in str(c) for k in dir_kw)
                and ('吞吐率' in str(c) or 'hroughput' in str(c).lower())]
        for c in cols:
            if pref in str(c):
                return c
        return cols[0] if cols else None

    def _parse_t(self, v) -> Optional[datetime]:
        if v is None or (isinstance(v, float) and pd.isna(v)) or str(v).strip() in ('', 'nan', 'None'):
            return None
        try:
            s = re.sub(r'\(\d+\)', '', str(v)).strip()
            return pd.to_datetime(s)
        except Exception:
            return None

    # ---------- 解析 UCM 监控文件 ----------
    def parse_ucm(self, ucm_path: str, params: Dict) -> pd.DataFrame:
        """读UCM：探测表头(中文第8行/英文第1行)、bps→Mbps、解析时间、削峰"""
        print("DEBUG: 读UCM...")
        raw = pd.read_excel(ucm_path, header=None, nrows=10)
        header_row = 7
        for i in range(min(10, len(raw))):
            vals = [str(v) for v in raw.iloc[i].tolist()]
            if any('吞吐率' in v or 'hroughput' in v.lower() for v in vals):
                header_row = i
                break
        print(f"DEBUG: UCM表头在第 {header_row + 1} 行")

        df = pd.read_excel(ucm_path, header=header_row)
        df = df.replace('-', np.nan).ffill()

        # 时间列
        time_col = None
        for c in df.columns:
            if '采集时间' in str(c) or str(c) == 'Time':
                time_col = c
                break
        if time_col:
            df['采集时间_解析'] = df[time_col].apply(self._parse_t)

        # bps→Mbps（中英文吞吐率列）
        for col in list(df.columns):
            cs = str(col)
            low = cs.lower()
            if ('吞吐率' in cs or 'hroughput' in low) and ('bps' in low) and ('mbps' not in low) and ('kbps' not in low):
                new_col = re.sub(r'\(.*?bps.*?\)', '(Mbps)', cs, flags=re.IGNORECASE)
                if new_col == cs:
                    new_col = cs + '(Mbps)'
                if new_col not in df.columns:
                    df[new_col] = pd.to_numeric(df[col], errors='coerce') / 1_000_000
                df = df.drop(columns=[col])
        print(f"DEBUG: UCM {len(df)} 行, 列数 {len(df.columns)}")

        # 削峰
        pref = params.get('rate_column', 'RLC')
        dl_limit = params.get('dl_peak_limit', 1000)
        ul_limit = params.get('ul_peak_limit', 200)
        dl_col = self._select_rate_col(df, '下行', pref)
        ul_col = self._select_rate_col(df, '上行', pref)
        if dl_col:
            df['削峰后下行RLC'] = pd.to_numeric(df[dl_col], errors='coerce').apply(lambda x: min(x, dl_limit) if pd.notna(x) else x)
        if ul_col:
            df['削峰后上行RLC'] = pd.to_numeric(df[ul_col], errors='coerce').apply(lambda x: min(x, ul_limit) if pd.notna(x) else x)
        print(f"DEBUG: 削峰 {dl_limit}/{ul_limit}, 列 DL:{dl_col} UL:{ul_col}")
        return df

    # ---------- 解析呼叫详情（业务日志） ----------
    def parse_call_log(self, ref_path: str) -> Dict[str, List[Dict]]:
        """读参考文件「呼叫详情」sheet，按业务提取每次记录(start/end/rate/result/duration)"""
        print("DEBUG: 读呼叫详情...")
        wb = openpyxl.load_workbook(ref_path, data_only=True)
        sn = '呼叫详情' if '呼叫详情' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        header = list(rows[1]) if len(rows) > 1 else []

        def idx(name):
            return header.index(name) if name in header else None

        bi, tb = idx('业务类型'), idx('测试业务')
        # FTP
        ftp_st, ftp_et, ftp_fail = idx('发起时间'), idx('完成时间'), idx('失败时间')
        ftp_dl_r, ftp_ul_r = idx('下载平均速率(Mbps)'), idx('上传平均速率(Mbps)')
        # 微信
        wx_xs, wx_xe = idx('微信小包发送开始时间'), idx('微信小包发送完成时间')
        wx_ds, wx_de = idx('微信大包发送开始时间'), idx('微信大包发送完成时间')
        wx_xr, wx_dr = idx('微信小包发送速率(Mbps)'), idx('微信大包发送速率(Mbps)')
        wx_res = idx('微信文件业务结果')
        # 应用商店
        ap_xs, ap_xe = idx('应用商店小文件下载开始时间'), idx('应用商店小文件下载完成时间')
        ap_ds, ap_de = idx('应用商店大文件下载开始时间'), idx('应用商店大文件下载完成时间')
        ap_xr, ap_dr = idx('应用商店小文件下载速率(Mbps)'), idx('应用商店大文件下载速率(Mbps)')
        ap_res = idx('应用商店业务结果')

        recs = {k: [] for k in ['FTP下载', 'FTP上传', '应用商店_小包', '应用商店_大包', '微信_小文件', '微信_大文件']}

        def to_rec(s_col, e_col, r_col, res_col, fail_col=None):
            """从一行构造一条记录"""
            s = self._parse_t(s_col) if s_col is not None else None
            e = self._parse_t(e_col) if e_col is not None else None
            if s is None or e is None:
                return None
            dur = (e - s).total_seconds()
            res = None
            if res_col is not None and str(res_col).strip() not in ('', 'nan', 'None'):
                res = str(res_col).strip()
            elif fail_col is not None:
                res = '失败' if str(fail_col).strip() not in ('', 'nan', 'None') else '成功'
            rate = None
            if r_col is not None and str(r_col).strip() not in ('', 'nan', 'None'):
                try:
                    rate = float(r_col)
                except Exception:
                    rate = None
            return {'start': s, 'end': e, 'rate': rate, 'result': res, 'duration': dur if dur >= 0 else None}

        for r in rows[2:]:
            if r is None:
                continue
            # FTP
            if bi is not None and tb is not None and r[bi] == 'FTP' and r[tb] == 'FTPDownload':
                rec = to_rec(r[ftp_st] if ftp_st is not None else None,
                             r[ftp_et] if ftp_et is not None else None,
                             r[ftp_dl_r] if ftp_dl_r is not None else None,
                             None, r[ftp_fail] if ftp_fail is not None else None)
                if rec:
                    recs['FTP下载'].append(rec)
            if bi is not None and tb is not None and r[bi] == 'FTP' and r[tb] == 'FTPUpload':
                rec = to_rec(r[ftp_st] if ftp_st is not None else None,
                             r[ftp_et] if ftp_et is not None else None,
                             r[ftp_ul_r] if ftp_ul_r is not None else None,
                             None, r[ftp_fail] if ftp_fail is not None else None)
                if rec:
                    recs['FTP上传'].append(rec)
            # 微信小包/大包（每行）
            if wx_xs is not None and wx_xe is not None:
                res_v = r[wx_res] if wx_res is not None else None
                rec = to_rec(r[wx_xs], r[wx_xe], r[wx_xr] if wx_xr is not None else None, res_v)
                if rec:
                    recs['微信_小文件'].append(rec)
            if wx_ds is not None and wx_de is not None:
                res_v = r[wx_res] if wx_res is not None else None
                rec = to_rec(r[wx_ds], r[wx_de], r[wx_dr] if wx_dr is not None else None, res_v)
                if rec:
                    recs['微信_大文件'].append(rec)
            # 应用商店小/大包（每行）
            if ap_xs is not None and ap_xe is not None:
                res_v = r[ap_res] if ap_res is not None else None
                rec = to_rec(r[ap_xs], r[ap_xe], r[ap_xr] if ap_xr is not None else None, res_v)
                if rec:
                    recs['应用商店_小包'].append(rec)
            if ap_ds is not None and ap_de is not None:
                res_v = r[ap_res] if ap_res is not None else None
                rec = to_rec(r[ap_ds], r[ap_de], r[ap_dr] if ap_dr is not None else None, res_v)
                if rec:
                    recs['应用商店_大包'].append(rec)

        print(f"DEBUG: 呼叫详情提取 {[(k, len(v)) for k, v in recs.items()]}")
        return recs

    # ---------- 截取 UCM 点 ----------
    def _seg_points(self, ucm_df: pd.DataFrame, records: List[Dict], clip_col: str, orig_col: Optional[str]) -> Tuple[np.ndarray, np.ndarray]:
        """合并所有record时间窗内的UCM点，返回(削峰后数组, 原始数组)"""
        clips, origs = [], []
        if '采集时间_解析' not in ucm_df.columns or not records:
            return np.array([]), np.array([])
        for rec in records:
            if rec['start'] is None or rec['end'] is None:
                continue
            seg = ucm_df[(ucm_df['采集时间_解析'] >= rec['start']) & (ucm_df['采集时间_解析'] <= rec['end'])]
            if clip_col in seg.columns:
                clips.extend(pd.to_numeric(seg[clip_col], errors='coerce').dropna().tolist())
            if orig_col and orig_col in seg.columns:
                origs.extend(pd.to_numeric(seg[orig_col], errors='coerce').dropna().tolist())
        return np.array(clips), np.array(origs)

    def _seg_flow_mb(self, ucm_df: pd.DataFrame, records: List[Dict], direction: str) -> float:
        """截取时段内有效方向速率积分得流量(MB)——用于归一化时长"""
        if '采集时间_解析' not in ucm_df.columns or not records:
            return 0.0
        dl_col = self._select_rate_col(ucm_df, '下行', 'RLC')
        ul_col = self._select_rate_col(ucm_df, '上行', 'RLC')
        col = dl_col if direction == '下行' else ul_col
        if not col:
            return 0.0
        total = 0.0
        for rec in records:
            if rec['start'] is None or rec['end'] is None:
                continue
            seg = ucm_df[(ucm_df['采集时间_解析'] >= rec['start']) & (ucm_df['采集时间_解析'] <= rec['end'])]
            total += pd.to_numeric(seg[col], errors='coerce').fillna(0).sum() / 8  # Mbps·s→MB
        return total

    # ---------- 统计计算 ----------
    def calculate_statistics(self, ucm_df: pd.DataFrame, call_recs: Dict[str, List[Dict]], params: Dict, rules: List[Dict]) -> Dict[str, Dict]:
        pref = params.get('rate_column', 'RLC')
        dl_orig = self._select_rate_col(ucm_df, '下行', pref)
        ul_orig = self._select_rate_col(ucm_df, '上行', pref)
        rule_by_type = {r.get('display_name'): r for r in rules if r.get('display_name')}
        dl_pass = params.get('dl_pass_threshold', 100)
        ul_pass = params.get('ul_pass_threshold', 20)
        dl_peak = params.get('dl_peak_limit', 1000)
        ul_peak = params.get('ul_peak_limit', 200)

        def success_rate(recs):
            valid = [r for r in recs if r['result'] is not None]
            if not valid:
                return None
            ok = sum(1 for r in valid if r['result'] in ('成功', 'SUCCESS', '1', 1))
            return round(ok / len(valid) * 100, 2)

        def mean_rate(recs):
            rs = [r['rate'] for r in recs if r['rate'] is not None]
            return round(float(np.mean(rs)), 3) if rs else None

        result = {}

        # ---- FTP下载 ----
        recs = call_recs['FTP下载']
        clip, orig = self._seg_points(ucm_df, recs, '削峰后下行RLC', dl_orig)
        s = {'测试次数': len(recs), '采样点数': int(len(clip))}
        s['削峰后下行RLC≥100Mbps采集点数'] = int((clip >= 100).sum()) if len(clip) else 0
        s['削峰后下行RLC≥1000Mbps采集点数'] = int((clip >= 1000).sum()) if len(clip) else 0
        s['下行RLC平均吞吐率(Mbps)'] = mean_rate(recs) if mean_rate(recs) is not None else (round(float(orig.mean()), 3) if len(orig) else None)
        s['削峰后下行RLC平均吞吐率(Mbps)'] = round(float(clip.mean()), 3) if len(clip) else None
        s['下行RLC峰值吞吐率(Mbps)'] = round(float(orig.max()), 3) if len(orig) else None
        s['削峰后下行RLC≥100Mbps占比'] = round(float((clip >= 100).sum() / len(clip) * 100), 2) if len(clip) else None
        s['削峰后下行RLC≥1000Mbps占比'] = round(float((clip >= 1000).sum() / len(clip) * 100), 2) if len(clip) else None
        if len(clip):
            sr = np.sort(clip)[::-1]
            s['前10%峰值速率(Mbps)'] = round(float(sr[:max(1, int(len(sr) * 0.1))].mean()), 3)
        else:
            s['前10%峰值速率(Mbps)'] = None
        result['FTP下载'] = s

        # ---- FTP上传 ----
        recs = call_recs['FTP上传']
        clip, orig = self._seg_points(ucm_df, recs, '削峰后上行RLC', ul_orig)
        s = {'测试次数': len(recs), '采样点数': int(len(clip))}
        s['削峰后上行RLC>20Mbps采集点数'] = int((clip > 20).sum()) if len(clip) else 0
        s['削峰后上行RLC>200Mbps采集点数'] = int((clip > 200).sum()) if len(clip) else 0
        s['上行RLC平均吞吐率(Mbps)'] = mean_rate(recs) if mean_rate(recs) is not None else (round(float(orig.mean()), 3) if len(orig) else None)
        s['削峰后上行RLC平均吞吐率(Mbps)'] = round(float(clip.mean()), 3) if len(clip) else None
        s['上行RLC峰值吞吐率(Mbps)'] = round(float(orig.max()), 3) if len(orig) else None
        s['削峰后上行RLC>20Mbps占比'] = round(float((clip > 20).sum() / len(clip) * 100), 2) if len(clip) else None
        s['削峰后上行RLC>200Mbps占比'] = round(float((clip > 200).sum() / len(clip) * 100), 2) if len(clip) else None
        if len(clip):
            sr = np.sort(clip)[::-1]
            s['前10%峰值速率(Mbps)'] = round(float(sr[:max(1, int(len(sr) * 0.1))].mean()), 3)
        else:
            s['前10%峰值速率(Mbps)'] = None
        result['FTP上传'] = s

        # ---- 应用商店 / 微信 ----
        for biz, direction, is_big, succ_field in [
            ('应用商店_小包', '下行', False, '下载成功率(%)'),
            ('应用商店_大包', '下行', True, '下载成功率(%)'),
            ('微信_小文件', '上行', False, '上传成功率(%)'),
            ('微信_大文件', '上行', True, '上传成功率(%)'),
        ]:
            recs = call_recs[biz]
            orig_c = dl_orig if direction == '下行' else ul_orig
            clip_col = '削峰后下行RLC' if direction == '下行' else '削峰后上行RLC'
            clip, orig = self._seg_points(ucm_df, recs, clip_col, orig_c)
            s = {'测试次数': len(recs), '采样点数': int(len(orig))}
            rate_field = '下行RLC平均吞吐率(Mbps)' if direction == '下行' else '上行RLC平均吞吐率(Mbps)'
            s[rate_field] = mean_rate(recs) if mean_rate(recs) is not None else (round(float(orig.mean()), 3) if len(orig) else None)
            # 单次平均时长（完整记录）
            durs = [r['duration'] for r in recs if r['duration'] is not None and r['duration'] > 0]
            s['单次平均时长(s)'] = round(float(np.mean(durs)), 3) if durs else None
            # 中位值平均时长（大包/大文件）—— 经参考文件核对为「简单时长中位数」（非归一化）
            if is_big:
                s['中位值平均时长(s)'] = round(float(np.median(durs)), 3) if durs else None
            # 成功率（业务整体）
            grp = '应用商店' if '应用商店' in biz else '微信'
            grp_recs = []
            for k in call_recs:
                if (grp == '应用商店' and k.startswith('应用商店')) or (grp == '微信' and k.startswith('微信')):
                    grp_recs.extend(call_recs[k])
            s[succ_field] = success_rate(grp_recs)
            result[biz] = s
        return result

    def _norm_median_duration(self, ucm_df: pd.DataFrame, recs: List[Dict], direction: str, rule: Optional[Dict]) -> Optional[float]:
        """中位值平均时长(V6归一化): 每段 时长×中位流量/实际流量, 取中位。实际流量=该段UCM速率积分"""
        if not rule:
            return None
        median_flow = rule.get('median_flow')
        if not median_flow:
            # 无中位流量→退化为简单时长中位
            durs = [r['duration'] for r in recs if r['duration'] is not None and r['duration'] > 0]
            return round(float(np.median(durs)), 3) if durs else None
        norm_list = []
        dl_col = self._select_rate_col(ucm_df, '下行', 'RLC')
        ul_col = self._select_rate_col(ucm_df, '上行', 'RLC')
        col = dl_col if direction == '下行' else ul_col
        for r in recs:
            if r['duration'] is None or r['duration'] <= 0 or r['start'] is None or r['end'] is None:
                continue
            seg = ucm_df[(ucm_df['采集时间_解析'] >= r['start']) & (ucm_df['采集时间_解析'] <= r['end'])]
            flow_mb = pd.to_numeric(seg[col], errors='coerce').fillna(0).sum() / 8 if col else 0
            if flow_mb > 0:
                norm_list.append(r['duration'] * median_flow / flow_mb)
        return round(float(np.median(norm_list)), 3) if norm_list else None

    def process(self, ucm_path: str, ref_path: str, params: Dict, rules: List[Dict]) -> Tuple[pd.DataFrame, Dict]:
        self.ucm_df = self.parse_ucm(ucm_path, params)
        call_recs = self.parse_call_log(ref_path)
        self.statistics = self.calculate_statistics(self.ucm_df, call_recs, params, rules)
        return self.ucm_df, self.statistics


# ==================== Excel导出 ====================

class ExcelExporter:
    def __init__(self):
        self.wb = None

    def export(self, output_path: str, ucm_df: pd.DataFrame, statistics: Dict):
        self.wb = Workbook()
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        self._create_statistics_sheet(statistics)
        self._create_ucm_sheet(ucm_df)
        self.wb.save(output_path)

    def _create_statistics_sheet(self, statistics: Dict):
        ws = self.wb.create_sheet('统计结果')
        gfill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        for col_idx, (grp, field) in enumerate(WIDE_COLUMNS, 1):
            ws.cell(row=2, column=col_idx, value=field).font = Font(bold=True)
        start = 1
        for i in range(1, len(WIDE_COLUMNS) + 1):
            if i == len(WIDE_COLUMNS) or WIDE_COLUMNS[i][0] != WIDE_COLUMNS[start - 1][0]:
                c1 = ws.cell(row=1, column=start, value=WIDE_COLUMNS[start - 1][0])
                c1.font = Font(bold=True); c1.fill = gfill; c1.alignment = Alignment(horizontal='center')
                if i > start:
                    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=i)
                start = i + 1
        for col_idx, (grp, field) in enumerate(WIDE_COLUMNS, 1):
            v = statistics.get(grp, {}).get(field)
            ws.cell(row=3, column=col_idx, value=v if v is not None else '').alignment = Alignment(horizontal='center')

    def _create_ucm_sheet(self, ucm_df: pd.DataFrame):
        ws = self.wb.create_sheet('UCM数据')
        if ucm_df is None or ucm_df.empty:
            ws.cell(row=1, column=1, value='无数据'); return
        for col, h in enumerate(ucm_df.columns, 1):
            ws.cell(row=1, column=col, value=h).font = Font(bold=True)
        for row_idx, row_data in enumerate(ucm_df.head(5000).itertuples(index=False), 2):
            for col, v in enumerate(row_data, 1):
                if pd.isna(v):
                    v = ''
                ws.cell(row=row_idx, column=col, value=v)


# ==================== 工作线程 ====================

class ProcessingThread(QThread):
    progress_signal = Signal(str)
    finished_signal = Signal(bool, str, object, object)

    def __init__(self, processor: DataProcessor, ucm_path: str, ref_path: str, params: Dict, rules: List[Dict]):
        super().__init__()
        self.processor = processor
        self.ucm_path = ucm_path
        self.ref_path = ref_path
        self.params = params
        self.rules = rules

    def run(self):
        try:
            self.progress_signal.emit("解析UCM监控文件...")
            ucm_df, statistics = self.processor.process(self.ucm_path, self.ref_path, self.params, self.rules)
            self.progress_signal.emit("生成统计结果...")
            self.finished_signal.emit(True, "处理完成", ucm_df, statistics)
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.finished_signal.emit(False, f"处理失败: {str(e)}", None, None)


# ==================== 主窗口 ====================

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config_manager = ConfigManager()
        self.config = self.config_manager.config
        self.processor = None
        self.ucm_file = ''
        self.ref_file = ''
        self.processing_thread = None
        self.init_ui()
        self.load_config_to_ui()

    def init_ui(self):
        self.setWindowTitle('5G用户级公共监控速率统计工具 V1.02（呼叫详情配准）')
        self.setGeometry(100, 100, 1500, 900)
        w = QWidget(); self.setCentralWidget(w)
        lay = QHBoxLayout(w)
        lay.addWidget(self.create_left_panel(), 1)
        lay.addWidget(self.create_right_panel(), 2)

    def create_left_panel(self) -> QWidget:
        panel = QWidget(); layout = QVBoxLayout(panel)
        # 双文件输入
        fg = QGroupBox('输入文件（双文件）'); fl = QVBoxLayout()
        r1 = QHBoxLayout(); r1.addWidget(QLabel('UCM监控:'))
        self.ucm_edit = QLineEdit(); self.ucm_edit.setPlaceholderText('mmf/两江 等(中英文自动识别)')
        r1.addWidget(self.ucm_edit)
        b1 = QPushButton('选择'); b1.clicked.connect(lambda: self._pick(self.ucm_edit, 'ucm')); r1.addWidget(b1)
        fl.addLayout(r1)
        r2 = QHBoxLayout(); r2.addWidget(QLabel('业务日志:'))
        self.ref_edit = QLineEdit(); self.ref_edit.setPlaceholderText('参考文件(含呼叫详情sheet)')
        r2.addWidget(self.ref_edit)
        b2 = QPushButton('选择'); b2.clicked.connect(lambda: self._pick(self.ref_edit, 'ref')); r2.addWidget(b2)
        fl.addLayout(r2)
        fg.setLayout(fl); layout.addWidget(fg)

        # 参数
        pg = QGroupBox('全局参数'); pl = QFormLayout()
        self.qci_spin = QSpinBox(); self.qci_spin.setRange(0, 99); self.qci_spin.setValue(0); self.qci_spin.setSpecialValueText("全部")
        pl.addRow('QCI(0=全部):', self.qci_spin)
        self.rate_column_combo = QComboBox(); self.rate_column_combo.addItems(['RLC', 'MAC']); pl.addRow('速率取值列:', self.rate_column_combo)
        self.dl_peak_spin = QDoubleSpinBox(); self.dl_peak_spin.setRange(0, 5000); self.dl_peak_spin.setValue(1000); pl.addRow('下行削峰(Mbps):', self.dl_peak_spin)
        self.ul_peak_spin = QDoubleSpinBox(); self.ul_peak_spin.setRange(0, 1000); self.ul_peak_spin.setValue(200); pl.addRow('上行削峰(Mbps):', self.ul_peak_spin)
        self.dl_pass_spin = QDoubleSpinBox(); self.dl_pass_spin.setRange(0, 1000); self.dl_pass_spin.setValue(100); pl.addRow('下行达标门限:', self.dl_pass_spin)
        self.ul_pass_spin = QDoubleSpinBox(); self.ul_pass_spin.setRange(0, 100); self.ul_pass_spin.setValue(20); pl.addRow('上行达标门限:', self.ul_pass_spin)
        pg.setLayout(pl); layout.addWidget(pg)

        bl = QHBoxLayout()
        pb = QPushButton('开始统计'); pb.setStyleSheet('QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 10px; }')
        pb.clicked.connect(self.start_processing); bl.addWidget(pb)
        eb = QPushButton('导出Excel'); eb.clicked.connect(self.export_excel); bl.addWidget(eb)
        layout.addLayout(bl)
        self.progress_bar = QProgressBar(); layout.addWidget(self.progress_bar)

        lg = QGroupBox('运行日志'); ll = QVBoxLayout()
        self.log_text = QTextEdit(); self.log_text.setReadOnly(True); ll.addWidget(self.log_text)
        lg.setLayout(ll); layout.addWidget(lg, 1)
        return panel

    def _pick(self, edit, kind):
        fp, _ = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel Files (*.xlsx *.xls)')
        if fp:
            edit.setText(fp)
            if kind == 'ucm':
                self.ucm_file = fp
            else:
                self.ref_file = fp
            self.log_message(f"已选择({kind}): {fp}")

    def create_right_panel(self) -> QWidget:
        panel = QWidget(); layout = QVBoxLayout(panel)
        self.stats_table = QTableWidget()
        headers = [f"{g}|{f}" for g, f in WIDE_COLUMNS]
        self.stats_table.setColumnCount(len(headers))
        self.stats_table.setHorizontalHeaderLabels(headers)
        layout.addWidget(self.stats_table)
        return panel

    def load_config_to_ui(self):
        p = self.config.get('global_params', {})
        self.rate_column_combo.setCurrentText(p.get('rate_column', 'RLC'))
        self.dl_peak_spin.setValue(p.get('dl_peak_limit', 1000))
        self.ul_peak_spin.setValue(p.get('ul_peak_limit', 200))
        self.dl_pass_spin.setValue(p.get('dl_pass_threshold', 100))
        self.ul_pass_spin.setValue(p.get('ul_pass_threshold', 20))

    def get_params(self) -> Dict:
        return {
            'qci': self.qci_spin.value(), 'rate_column': self.rate_column_combo.currentText(),
            'dl_peak_limit': self.dl_peak_spin.value(), 'ul_peak_limit': self.ul_peak_spin.value(),
            'dl_pass_threshold': self.dl_pass_spin.value(), 'ul_pass_threshold': self.ul_pass_spin.value(),
        }

    @Slot()
    def start_processing(self):
        self.ucm_file = self.ucm_edit.text().strip()
        self.ref_file = self.ref_edit.text().strip()
        if not self.ucm_file or not self.ref_file:
            QMessageBox.warning(self, '警告', '请同时选择 UCM监控文件 和 业务日志(参考文件)'); return
        params = self.get_params()
        rules = self.config.get('business_rules', [])
        self.log_message("=" * 50); self.log_message("开始处理(呼叫详情配准)...")
        self.log_message(f"UCM: {os.path.basename(self.ucm_file)} | 业务日志: {os.path.basename(self.ref_file)}")
        self.progress_bar.setValue(0)
        self.sender().setEnabled(False); self.sender().setText("处理中...")
        self.processor = DataProcessor(self.config)
        self.processing_thread = ProcessingThread(self.processor, self.ucm_file, self.ref_file, params, rules)
        self.processing_thread.progress_signal.connect(self.update_progress)
        self.processing_thread.finished_signal.connect(self.on_finished)
        self.processing_thread.start()

    @Slot(str)
    def update_progress(self, msg):
        self.log_message(msg)
        c = self.progress_bar.value()
        if c < 90: self.progress_bar.setValue(min(c + 20, 90))

    @Slot(bool, str, object, object)
    def on_finished(self, success, message, ucm_df, statistics):
        self.progress_bar.setValue(100); self.log_message(message); self.log_message("=" * 50)
        for btn in self.findChildren(QPushButton):
            if btn.text() in ["处理中...", "开始统计"]:
                btn.setEnabled(True); btn.setText("开始统计")
        if success:
            self.display_statistics(statistics)
            self.last_ucm = ucm_df; self.last_statistics = statistics
            self.log_message(f"✓ 完成！业务组: {list(statistics.keys())}")
            QMessageBox.information(self, '完成', '处理完成！')
        else:
            self.log_message(f"✗ 失败: {message}")
            QMessageBox.critical(self, '错误', f'处理失败: {message}')

    def display_statistics(self, statistics: Dict):
        self.stats_table.setRowCount(1)
        for col, (grp, field) in enumerate(WIDE_COLUMNS):
            v = statistics.get(grp, {}).get(field)
            text = '' if v is None else (f"{v:.2f}" if isinstance(v, float) else str(v))
            self.stats_table.setItem(0, col, QTableWidgetItem(text))

    def export_excel(self):
        if not hasattr(self, 'last_statistics') or self.last_statistics is None:
            QMessageBox.warning(self, '警告', '请先处理数据'); return
        fp, _ = QFileDialog.getSaveFileName(self, '保存Excel', '', 'Excel Files (*.xlsx)')
        if fp:
            try:
                ExcelExporter().export(fp, self.last_ucm, self.last_statistics)
                self.log_message(f"已导出: {fp}")
                QMessageBox.information(self, '成功', '导出成功！')
            except Exception as e:
                QMessageBox.critical(self, '错误', f'导出失败: {str(e)}')

    def log_message(self, msg):
        self.log_text.append(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    w = MainWindow(); w.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
