#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
5G用户级公共监控速率统计工具 V2.0.9
GUI: 选择MMF文件 → 设置QCI/削峰阈值 → 生成4Sheet输出
CLI: python3 this.py --cmd
"""
import sys, os, re
import pandas as pd, numpy as np
from datetime import datetime, timedelta
from collections import Counter, defaultdict

# ===== 阈值 =====
DL_ACTIVE_BIG, UL_ACTIVE_BIG = 50.0, 10.0
DL_ACTIVE_SMALL, UL_ACTIVE_SMALL = 5.0, 5.0
GAP_MERGE = 2.0
FTP_DUR_MIN, FTP_DUR_MAX, FTP_PAIR_GAP = 8.5, 13.0, 15.0
STORE_L_DL_MIN, STORE_L_DL_MAX = 8800, 10100
WX_L_DUR_MIN, WX_L_DUR_MAX, WX_L_UL_MIN = 15, 60, 800
DL_BIZ = {'FTP下载', '应用商店小文件下载', '应用商店大文件下载'}
UL_BIZ = {'FTP上传', '微信小包发送', '微信大包发送'}
REQ_BIZ = {'FTP下载', 'FTP上传', '应用商店小文件下载', '应用商店大文件下载', '微信小包发送', '微信大包发送'}


def parse_time(t):
    try:
        t = str(t).strip()
        m = re.match(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\((\d+)\)', t)
        if m: return datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S') + timedelta(milliseconds=int(m.group(2)))
    except: pass


def detect_segments(ts, dl, ul, dl_th, ul_th, gap, r1=None, r2=None):
    active = (dl > dl_th) | (ul > ul_th); segs = []; i = 0; n = len(ts)
    while i < n:
        if r1 is not None and ts[i] < r1: i += 1; continue
        if r2 is not None and ts[i] > r2: break
        if active[i]:
            si = i; j = i + 1
            while j < n:
                if r2 is not None and ts[j] > r2: break
                if not active[j]:
                    k = j
                    while k < n and not active[k]:
                        if (ts[k] - ts[j - 1]) / np.timedelta64(1, 's') > gap: break
                        k += 1
                    if k < n and active[k] and (r2 is None or ts[k] <= r2): j = k; continue
                    else: break
                j += 1
            st = ts[si]; et = ts[j - 1]; dur = (et - st) / np.timedelta64(1, 's') + 1
            segs.append({'st': st, 'et': et, 'dur': dur, 'dl': float(dl[si:j].sum()), 'ul': float(ul[si:j].sum())})
            i = j
        else: i += 1
    return segs


def process(files, qci_list=None, dl_clip=1000, ul_clip=200, callback=None):
    if qci_list is None: qci_list = [5, 6, 7]
    """统一的处理函数：从MMF文件生成3Sheet输出"""
    def log(msg):
        print(msg)
        if callback: callback(msg)

    log("加载MMF文件...")
    dfs = []
    for f in files:
        df_i = pd.read_excel(f)
        df_i = df_i.assign(_t=df_i['Time'].apply(parse_time))
        rlc_dl = [c for c in df_i.columns if 'RLC' in c and 'Downlink' in c]
        rlc_ul = [c for c in df_i.columns if 'RLC' in c and 'Uplink' in c]
        df_i = df_i.assign(_dl=pd.to_numeric(df_i[rlc_dl[0]], errors='coerce') / 1e6,
                           _ul=pd.to_numeric(df_i[rlc_ul[0]], errors='coerce') / 1e6)
        dfs.append(df_i)
    df = pd.concat(dfs, ignore_index=True)

    log("QCI过滤+秒聚合...")
    df7 = df[df['QCI'].isin(qci_list)].dropna(subset=['_t']).copy()
    df7['_s'] = df7['_t'].dt.floor('s')
    sec = df7.groupby('_s').agg({'_dl': 'max', '_ul': 'max'}).sort_index().reset_index()
    dl_v, ul_v, ts_v = sec['_dl'].fillna(0).values, sec['_ul'].fillna(0).values, sec['_s'].values

    log("业务识别...")
    big = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_BIG, UL_ACTIVE_BIG, GAP_MERGE)

    def cb(s):
        d, dl, ul = s['dur'], s['dl'], s['ul']
        if FTP_DUR_MIN <= d <= FTP_DUR_MAX: return 'FTP候选'
        if dl > ul and STORE_L_DL_MIN <= dl <= STORE_L_DL_MAX and d > 6: return '应用商店大文件下载'
        if WX_L_DUR_MIN <= d <= WX_L_DUR_MAX and ul > dl and ul >= WX_L_UL_MIN: return '微信大包发送'
    for s in big: s['biz'] = cb(s); s['rnd'] = None

    fc = [s for s in big if s['biz'] == 'FTP候选']; paired = set(); fp = []
    for i in range(len(fc)):
        if id(fc[i]) in paired: continue
        for j in range(i + 1, len(fc)):
            if id(fc[j]) in paired: continue
            g = (fc[j]['st'] - fc[i]['et']) / np.timedelta64(1, 's')
            if 0 <= g < FTP_PAIR_GAP:
                fp.append((fc[i], fc[j])); paired.add(id(fc[i])); paired.add(id(fc[j])); break
            elif g >= FTP_PAIR_GAP: break

    def fd(s): return 'FTP下载' if s['dl'] > s['ul'] else 'FTP上传'
    for s in fc:
        if id(s) not in paired:
            if STORE_L_DL_MIN <= s['dl'] <= STORE_L_DL_MAX and s['dl'] > s['ul']: s['biz'] = '应用商店大文件下载'
            else: s['biz'] = fd(s)

    al = sorted(big, key=lambda x: x['st'])
    for s in al: s['业务'] = s['biz']
    rn = 0; pi = []
    for f1, f2 in fp:
        rn += 1; f1['rnd'] = rn; f1['业务'] = fd(f1); f2['rnd'] = rn; f2['业务'] = fd(f2)
        pi.append((rn, f2['et'], f1['st']))
    for idx, (rnd, pe, _) in enumerate(pi):
        ns = pi[idx + 1][2] if idx + 1 < len(pi) else (al[-1]['et'] + np.timedelta64(1, 's'))
        for s in al:
            if s['et'] > pe and s['st'] < ns and s['rnd'] is None and s['业务'] in ('应用商店大文件下载', '微信大包发送'):
                s['rnd'] = rnd

    tb = sorted([s for s in al if s['业务'] in ('FTP下载', 'FTP上传', '应用商店大文件下载', '微信大包发送')],
                key=lambda x: x['st'])
    P2 = {'应用商店大文件下载': '应用商店小文件下载', '微信大包发送': '微信小包发送'}
    small = []
    for idx, (rnd, pe, ps) in enumerate(pi):
        we = pi[idx + 1][2] if idx + 1 < len(pi) else (tb[-1]['et'] + np.timedelta64(1, 's'))
        bw = sorted([s for s in tb if s['st'] >= pe and s['st'] < we and s['业务'] in (
        '应用商店大文件下载', '微信大包发送')], key=lambda x: x['st'])
        prev = pe
        for bs in bw:
            st2 = P2.get(bs['业务'])
            if st2 and prev < bs['st'] and (bs['st'] - prev) / np.timedelta64(1, 's') > 1:
                gs = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_SMALL, UL_ACTIVE_SMALL, GAP_MERGE, prev, bs['st'])
                vd = [s for s in gs if s['dl'] + s['ul'] >= 10 and s['dur'] >= 0.5]
                if st2 == '应用商店小文件下载':
                    dls = [s for s in vd if s['dl'] > s['ul']]
                    if dls: bm = max(dls, key=lambda x: x['dl']); bm['业务'] = st2; bm['rnd'] = rnd; small.append(bm)
                else:
                    uls = [s for s in vd if s['ul'] > s['dl']]
                    if uls: bm = max(uls, key=lambda x: x['ul']); bm['业务'] = st2; bm['rnd'] = rnd; small.append(bm)
            prev = bs['et']

    # 检测第一个FTP对之前的不完整轮次
    if pi:
        first_ftp_start = pi[0][2]
        data_start = ts_v[0]
        if first_ftp_start > data_start + np.timedelta64(30, 's'):
            # 找第一个FTP对之前的大业务(应用商店大/微信大)
            pre_big = sorted([s for s in al if s['st'] < first_ftp_start and s['业务'] in (
                '应用商店大文件下载', '微信大包发送')], key=lambda x: x['st'])
            rng_start = data_start
            if pre_big:
                for bs in pre_big:
                    st2 = P2.get(bs['业务'])
                    if st2 and (bs['st'] - rng_start) / np.timedelta64(1, 's') > 1:
                        gs = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_SMALL, UL_ACTIVE_SMALL,
                                             GAP_MERGE, rng_start, bs['st'])
                        vd = [s for s in gs if s['dl'] + s['ul'] >= 10 and s['dur'] >= 0.5]
                        if st2 == '应用商店小文件下载':
                            dls = [s for s in vd if s['dl'] > s['ul']]
                            if dls: bm = max(dls, key=lambda x: x['dl']); bm['业务'] = st2; bm['rnd'] = 0; small.append(bm)
                        else:
                            uls = [s for s in vd if s['ul'] > s['dl']]
                            if uls: bm = max(uls, key=lambda x: x['ul']); bm['业务'] = st2; bm['rnd'] = 0; small.append(bm)
                    rng_start = bs['et']
            # 最后一个大业务(或数据开头) 到 first_ftp_start 之间
            if first_ftp_start > rng_start + np.timedelta64(1, 's'):
                gs = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_SMALL, UL_ACTIVE_SMALL,
                                     GAP_MERGE, rng_start, first_ftp_start)
                vd = [s for s in gs if s['dl'] + s['ul'] >= 10 and s['dur'] >= 0.5]
                dls = [s for s in vd if s['dl'] > s['ul']]
                if dls: bm = max(dls, key=lambda x: x['dl']); bm['业务'] = '应用商店小文件下载'; bm['rnd'] = 0; small.append(bm)
                uls = [s for s in vd if s['ul'] > s['dl']]
                if uls: bm = max(uls, key=lambda x: x['ul']); bm['业务'] = '微信小包发送'; bm['rnd'] = 0; small.append(bm)

    af = sorted(al + small, key=lambda x: x['st'])
    for s in af: s['有效'] = s['业务'] is not None and s.get('rnd') is not None
    rm = defaultdict(list)
    for s in af:
        if s['有效'] and s['业务'] and s.get('rnd') is not None: rm[s['rnd']].append(s)
    for rnd, segs in rm.items():
        seen = set(); cnt = 0
        for s in sorted(segs, key=lambda x: x['st']):
            if s['业务'] in seen or cnt >= 6: s['有效'] = False
            else: seen.add(s['业务']); cnt += 1

    valid = [s for s in af if s['有效'] and s['业务']]
    log(f"识别完成: {len(valid)}个业务段, {rn}轮")

    # ===== 标记 =====
    df = pd.concat([df, pd.DataFrame({
        '下载速率': df['_dl'], '上传速率': df['_ul'],
        '业务识别': None, '持续时长': None, '速率(Mbps)': None,
    }, index=df.index)], axis=1)

    for seg in valid:
        st, et, biz = seg['st'], seg['et'], seg['业务']
        sm = (df['_t'] >= st) & (df['_t'] <= et + timedelta(seconds=1)) & (df['QCI'].isin(qci_list))
        col = '_dl' if biz in DL_BIZ else '_ul'

        # 所有业务：只标记有流量的行(>0)
        cv_all = df.loc[sm, col].fillna(0)
        nz = cv_all > 0
        mm = sm & nz.reindex(df.index, fill_value=False)

        if not mm.any():
            continue  # 无流量则跳过

        # FTP：从第一个有流量的行开始，检查前后小流量补齐到10秒
        if biz in ('FTP下载', 'FTP上传'):
            # 找所有有流量的行
            flow_vals = df.loc[mm, col].fillna(0).tolist()
            flow_times = df.loc[mm, '_t'].tolist()

            if len(flow_vals) == 0:
                continue

            # 起点：从第一个有流量的行开始
            first_idx = 0

            # 终点：最后一个有流量的行
            last_idx = len(flow_vals) - 1

            # 计算时长
            first_t = flow_times[first_idx]
            last_t = flow_times[last_idx]
            md = (last_t - first_t).total_seconds()

            # 如果时长不是10秒左右，检查前后是否有小流量可以补齐
            # 向前找：在segment开始时间之前3秒内，找有流量的行
            if md < 9.5 or md > 10.5:
                # 向前找小流量
                ft0 = st - timedelta(seconds=3)
                pm = (df['_t'] >= ft0) & (df['_t'] < first_t) & (df['QCI'].isin(qci_list))
                if pm.any():
                    pv = df.loc[pm, col].fillna(0)
                    pv_nz = pv > 0
                    if pv_nz.any():
                        # 有前置流量，扩展起点
                        first_t = df.loc[pv[pv_nz].index[0], '_t']
                        md = (last_t - first_t).total_seconds()

                # 向后找小流量
                et0 = et + timedelta(seconds=3)
                pm2 = (df['_t'] > last_t) & (df['_t'] <= et0) & (df['QCI'].isin(qci_list))
                if pm2.any():
                    pv2 = df.loc[pm2, col].fillna(0)
                    pv2_nz = pv2 > 0
                    if pv2_nz.any():
                        # 有后置流量，扩展终点
                        last_t = df.loc[pv2[pv2_nz].index[-1], '_t']
                        md = (last_t - first_t).total_seconds()

            # 重新计算标记范围
            mm_final = (df['_t'] >= first_t) & (df['_t'] <= last_t) & (df['QCI'].isin(qci_list)) & (df[col].fillna(0) > 0)
            md = (last_t - first_t).total_seconds()

            df.loc[mm_final, '业务识别'] = biz
            df.loc[mm_final, '持续时长'] = round(md, 3)
            vs = pd.to_numeric(df.loc[mm_final, '下载速率'] if biz in DL_BIZ else df.loc[mm_final, '上传速率'], errors='coerce').fillna(0).tolist()
            ar = sum(vs) / md if md > 0 else 0
            df.loc[mm_final, '速率(Mbps)'] = round(ar, 2)
            seg['mark_st'], seg['mark_et'], seg['mark_dur'] = first_t, last_t, md
            continue

        # 其他业务：直接标记有流量的行
        first_t = df.loc[mm, '_t'].min()
        last_t = df.loc[mm, '_t'].max()
        md = (last_t - first_t).total_seconds()

        df.loc[mm, '业务识别'] = biz
        df.loc[mm, '持续时长'] = round(md, 3)
        if biz in DL_BIZ: vs = pd.to_numeric(df.loc[mm, '下载速率'], errors='coerce').fillna(0).tolist()
        else: vs = pd.to_numeric(df.loc[mm, '上传速率'], errors='coerce').fillna(0).tolist()
        ar = sum(vs) / md if md > 0 else 0
        df.loc[mm, '速率(Mbps)'] = round(ar, 2)
        seg['mark_st'], seg['mark_et'], seg['mark_dur'] = first_t, last_t, md

    rd = defaultdict(set)
    for s in valid:
        r = s.get('rnd')
        if r is not None: rd[r].add(s['业务'])
    rl = {}; ii = 0
    for r in sorted(rd.keys()):
        if rd[r] >= REQ_BIZ: rl[r] = str(r)
        else: ii += 1; rl[r] = f"不完整{ii}"
    df['轮次'] = None
    for s in valid:
        r = s.get('rnd'); lb = rl.get(r) if r is not None else ''
        if lb:
            ms2, me2 = s.get('mark_st', s['st']), s.get('mark_et', s['et'])
            # 只在有业务标记的行填轮次
            mm2 = (df['_t'] >= ms2) & (df['_t'] <= me2) & (df['QCI'].isin(qci_list)) & df['业务识别'].notna()
            df.loc[mm2, '轮次'] = lb

    # ===== 详细过程 =====
    # 先创建筛选列（A列）
    # E-N列：速率(Mbps)、持续时长、轮次、businessType、基准_业务类型、基准_开始时间、基准_结束时间、基准_速率类指标、基准_数值、基准_时长类指标、基准_数值.1
    OC = ['下载速率', '上传速率', '速率(Mbps)', '持续时长', '轮次']
    df_detail = pd.DataFrame({'筛选': None, 'id': range(1, len(df) + 1), 'Time': df['Time']})
    for c in OC: df_detail[c] = df[c].values if c in df.columns else None
    df_detail['businessType'] = df['业务识别'].values

    # 加载基准数据并标记到详细过程
    base = pd.read_excel('工信部业务指标_呼叫详情_整理.xlsx', sheet_name='电信').sort_values('开始时间')
    # 初始化基准列
    base_cols = ['基准_业务类型', '基准_开始时间', '基准_结束时间', '基准_速率类指标', '基准_数值', '基准_时长类指标', '基准_数值.1']
    for bc in base_cols:
        df_detail[bc] = None

    # 对齐基准数据到详细过程（使用df的_t列）
    for _, brow in base.iterrows():
        b_start = pd.to_datetime(brow['开始时间'])
        b_end = pd.to_datetime(brow['结束时间'])
        b_biz = str(brow['业务类型'])

        # 在df中找最接近的Time行（使用df的_t列）
        min_diff = 9999.0
        best_start_idx = None
        for idx in range(len(df)):
            t = df.loc[idx, '_t'] if '_t' in df.columns else None
            if t is not None and pd.notna(t):
                diff = abs((t - b_start).total_seconds())
                if diff < min_diff:
                    min_diff = diff
                    best_start_idx = idx

        # 标记基准信息
        if best_start_idx is not None and min_diff < 5:
            # 找到结束时间对应的行
            best_end_idx = best_start_idx
            for idx in range(best_start_idx, len(df)):
                t = df.loc[idx, '_t'] if '_t' in df.columns else None
                if t is not None and pd.notna(t):
                    if abs((t - b_end).total_seconds()) < 1:
                        best_end_idx = idx
                        break

            # 标记基准列（df_detail的索引与df相同）
            for idx in range(best_start_idx, min(best_end_idx + 1, len(df_detail))):
                df_detail.loc[idx, '基准_业务类型'] = b_biz
                df_detail.loc[idx, '基准_开始时间'] = str(brow['开始时间'])
                df_detail.loc[idx, '基准_结束时间'] = str(brow['结束时间'])
                df_detail.loc[idx, '基准_速率类指标'] = str(brow['速率类指标'])
                df_detail.loc[idx, '基准_数值'] = brow['数值']
                df_detail.loc[idx, '基准_时长类指标'] = str(brow['时长类指标'])
                df_detail.loc[idx, '基准_数值.1'] = brow['数值.1']

    # 添加_t列
    df_detail['_t'] = df['_t'].values

    # 将QCI、Downlink RLC、Uplink RLC三列移到_t之后
    special_cols = ['QCI', 'Downlink RLC Throughput(bps)', 'Uplink RLC Throughput(bps)']
    for sc in special_cols:
        if sc in df.columns:
            df_detail[sc] = df[sc].values

    # 添加其他列
    for c in df.columns:
        if c not in df_detail.columns and c not in ('id', '_t', '_dl', '_ul', '_s', 'businessType', '业务识别'):
            df_detail[c] = df[c].values

    # 计算筛选列：E-N列都为空则为0，否则为1
    check_cols = ['速率(Mbps)', '持续时长', '轮次', 'businessType',
                  '基准_业务类型', '基准_开始时间', '基准_结束时间', '基准_速率类指标', '基准_数值', '基准_时长类指标', '基准_数值.1']
    for idx in range(len(df_detail)):
        has_data = False
        for col in check_cols:
            val = df_detail.loc[idx, col]
            if val is not None and str(val).strip() != '' and str(val) != 'nan':
                has_data = True
                break
        df_detail.loc[idx, '筛选'] = 1 if has_data else 0

    # ===== 汇总 =====
    sr = []
    for biz in sorted(DL_BIZ | UL_BIZ):
        sub = df_detail[df_detail['businessType'] == biz].dropna(subset=['_t']).sort_values('_t')
        if len(sub) == 0: continue
        dl_cn = [c for c in sub.columns if 'RLC' in c and 'Downlink' in c]
        ul_cn = [c for c in sub.columns if 'RLC' in c and 'Uplink' in c]
        if biz in DL_BIZ and dl_cn:
            rates = pd.to_numeric(sub[dl_cn[0]], errors='coerce').fillna(0) / 1e6; rates = rates[rates > 0]
        elif biz in UL_BIZ and ul_cn:
            rates = pd.to_numeric(sub[ul_cn[0]], errors='coerce').fillna(0) / 1e6; rates = rates[rates > 0]
        else: rates = pd.Series([], dtype=float)
        ts2 = sub['_t']; durs = []
        if len(ts2) > 0:
            s1 = ts2.iloc[0]; p = s1
            for t in ts2.iloc[1:]:
                if (t - p).total_seconds() > 5: durs.append((p - s1).total_seconds() + 1); s1 = t
                p = t
            durs.append((p - s1).total_seconds() + 1)
        row = {'业务类型': biz}
        if biz in DL_BIZ and len(rates) > 0:
            row['应用层FTP下载速率1000M以上占比(%)'] = round((rates > dl_clip).sum() / len(rates) * 100, 2)
            row['应用层FTP下载速率100M以下占比(%)'] = round((rates < 100).sum() / len(rates) * 100, 2)
            row['应用层平均下载速率(Mbps)'] = round(rates.mean(), 2)
            row['削峰应用层平均下载速率(Mbps)'] = round(rates.clip(upper=dl_clip).mean(), 2)
            sv = rates.sort_values(ascending=False); tn = max(1, int(len(sv) * 0.1))
            row['下行削峰TOP10%峰值速率'] = round(sv.head(tn).mean(), 2)
        elif biz in UL_BIZ and len(rates) > 0:
            row['应用层FTP上传速率200M以上占比(%)'] = round((rates > ul_clip).sum() / len(rates) * 100, 2)
            row['应用层FTP上传速率5M以下占比(%)'] = round((rates < 5).sum() / len(rates) * 100, 2)
            row['应用层平均上传速率'] = round(rates.mean(), 2)
            row['削峰应用层平均上传速率(Mbps)'] = round(rates.clip(upper=ul_clip).mean(), 2)
            sv = rates.sort_values(ascending=False); tn = max(1, int(len(sv) * 0.1))
            row['上行削峰TOP10%峰值速率'] = round(sv.head(tn).mean(), 2)
        if durs:
            row['业务时长平均值(s)'] = round(np.mean(durs), 2)
            row['业务时长中位值(s)'] = round(np.median(durs), 0)
        sr.append(row)
    df_summary = pd.DataFrame(sr)

    # ===== 对比横表（基准驱动，逐行匹配）=====
    BIZ_ORDER = ['FTP下载', 'FTP上传', '应用商店小文件下载', '应用商店大文件下载', '微信小包发送', '微信大包发送']
    rate_nm = {'FTP下载': '下载平均速率(Mbps)', 'FTP上传': '上传平均速率(Mbps)',
               '应用商店小文件下载': '小文件下载速率(Mbps)', '应用商店大文件下载': '大文件下载速率(Mbps)',
               '微信小包发送': '微信小包发送速率(Mbps)', '微信大包发送': '微信大包发送速率(Mbps)'}

    def fmt_time(t_str):
        m = re.match(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})[\.\(](\d+)', str(t_str))
        if m:
            dt = datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S')
            return f"{m.group(1)}({m.group(2)})", dt
        # 尝试毫秒格式 .xxx
        m2 = re.match(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\.(\d+)', str(t_str))
        if m2:
            dt = datetime.strptime(m2.group(1), '%Y-%m-%d %H:%M:%S')
            return f"{m2.group(1)}({m2.group(2)[:3]})", dt
        return str(t_str), None

    # 给df_detail添加_t列便于时间搜索
    df_detail['_t'] = df['_t'].values

    # 读取基准文件，按开始时间排序
    base = pd.read_excel('工信部业务指标_呼叫详情_整理.xlsx', sheet_name='电信').sort_values('开始时间')

    # 每6条分一组
    base_groups = []
    for i in range(0, len(base), 6):
        grp = base.iloc[i:i+6]
        if len(grp) == 0: continue
        base_groups.append(grp)

    # 逐组逐业务匹配
    all_items = []
    rnd_counter = 0
    for grp_idx, grp in enumerate(base_groups):
        all_matched = True
        grp_items = []

        for _, brow in grp.iterrows():
            b_biz = str(brow['业务类型'])
            b_start_str = str(brow['开始时间'])
            b_end_str = str(brow['结束时间'])

            # 解析基准开始时间
            b_start_parsed = pd.to_datetime(brow['开始时间'])

            # 在df_detail中找时间在基准窗口附近、businessType相同的代码行
            t_min = b_start_parsed - pd.Timedelta(seconds=5)
            t_max = b_start_parsed + pd.Timedelta(seconds=60)

            code_rows = df_detail[(df_detail['_t'] >= t_min) &
                                   (df_detail['_t'] <= t_max) &
                                   (df_detail['businessType'] == b_biz)]

            if len(code_rows) > 0:
                first = code_rows.iloc[0]
                last = code_rows.iloc[-1]

                c_start = str(first['Time'])
                c_end = str(last['Time'])
                c_start_fmt, _ = fmt_time(c_start)
                c_end_fmt, _ = fmt_time(c_end)

                code_data = {
                    '开始时间': c_start_fmt if c_start_fmt else c_start,
                    '结束时间': c_end_fmt if c_end_fmt else c_end,
                    '速率类指标': rate_nm.get(b_biz, ''),
                    '数值': first.get('速率(Mbps)'),
                    '时长类指标': '业务时长(秒)',
                    '数值.1': first.get('持续时长'),
                }
            else:
                all_matched = False
                code_data = None

            # 基准数据
            b_start_fmt, _ = fmt_time(b_start_str)
            b_end_fmt, _ = fmt_time(b_end_str)

            base_data = {
                '开始时间': b_start_fmt if b_start_fmt else b_start_str,
                '结束时间': b_end_fmt if b_end_fmt else b_end_str,
                '速率类指标': str(brow.get('速率类指标', '')),
                '数值': brow.get('数值'),
                '时长类指标': str(brow.get('时长类指标', '')),
                '数值.1': brow.get('数值.1'),
            }

            grp_items.append({
                'biz': b_biz,
                '基准': base_data,
                '代码': code_data,
            })

        rnd_label = str(rnd_counter + 1) if all_matched else '不计轮次'
        if all_matched: rnd_counter += 1
        for item in grp_items: item['rnd_label'] = rnd_label
        all_items.extend(grp_items)

    # 构建df_compare（每业务2行：基准+代码）
    cr = []
    for item in all_items:
        rnd = item['rnd_label']
        biz = item['biz']
        bd = item['基准']
        cd = item['代码']
        cr.append({'轮次': rnd, '来源': '基准', '业务类型': biz,
            '开始时间': bd['开始时间'], '结束时间': bd['结束时间'],
            '速率类指标': bd['速率类指标'], '数值': bd['数值'],
            '时长类指标': bd['时长类指标'], '数值.1': bd['数值.1']})
        if cd:
            cr.append({'轮次': rnd, '来源': '代码', '业务类型': biz,
                '开始时间': cd['开始时间'], '结束时间': cd['结束时间'],
                '速率类指标': cd['速率类指标'], '数值': cd['数值'],
                '时长类指标': cd['时长类指标'], '数值.1': cd['数值.1']})
        else:
            cr.append({'轮次': rnd, '来源': '代码', '业务类型': biz,
                '开始时间': '缺失', '结束时间': '缺失',
                '速率类指标': '缺失', '数值': '缺失',
                '时长类指标': '缺失', '数值.1': '缺失'})

    df_compare = pd.DataFrame(cr).reset_index(drop=True)

    # ===== 汇总 =====
    # 生成带版本号和时间戳的输出文件名
    version = 'V2.0.9'
    timestamp = datetime.now().strftime('%Y%m%d-%H%M')
    out = f'联通/电信最终输出_{version}_{timestamp}.xlsx'

    with pd.ExcelWriter(out, engine='openpyxl') as w:
        df_detail.to_excel(w, sheet_name='详细过程', index=False)
        df_compare.to_excel(w, sheet_name='对比', index=False)
        df_summary.to_excel(w, sheet_name='汇总', index=False)

        # 详细过程格式
        from openpyxl.styles import Alignment
        ws = w.sheets['详细过程']

        # C/D列：无流量填"-"，有流量但显示为0保持原值
        # 注意：由于新增筛选列，原来的C/D列现在是D/E列
        for row_idx in range(2, ws.max_row + 1):
            for col_letter in ['D', 'E']:
                cell = ws[f'{col_letter}{row_idx}']
                val = cell.value
                if val is None or (isinstance(val, (int, float)) and val == 0):
                    # 检查原始数据是否真的为0
                    col_idx = 4 if col_letter == 'D' else 5
                    # 查看df_detail中的原始值
                    orig_val = df_detail.iloc[row_idx - 2, col_idx - 1] if row_idx - 2 < len(df_detail) else None
                    if orig_val is None or (isinstance(orig_val, (int, float)) and orig_val == 0):
                        cell.value = '-'
                    # 否则保持原值（可能有微小流量）

        # 自适应列宽：根据表头和前100行数据内容计算
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            # 表头
            header_val = ws.cell(row=1, column=col_idx).value
            if header_val:
                max_len = max(max_len, len(str(header_val)))
            # 前100行数据
            for row_idx in range(2, min(102, ws.max_row + 1)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_len = max(max_len, min(len(str(cell_val)), 30))  # 最多30字符
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        # 第一行换行
        for cell in ws[1]:
            cell.alignment = Alignment(wrap_text=True)

        # 冻结至M2（新增筛选列，列号后移）
        ws.freeze_panes = 'M2'

        # D列E列数值格式整数（不显示小数）- 列号因新增筛选列而后移
        for col_letter in ['D', 'E']:
            for cell in ws[col_letter]:
                if cell.row > 1 and cell.value != '-':
                    cell.number_format = '0'

        # 筛选只显示标记为1的行
        from openpyxl.worksheet.filters import FilterColumn
        # 设置自动筛选
        ws.auto_filter.ref = ws.dimensions
        # 筛选A列=1
        ws.auto_filter.add_filter_column(0, ['1'])

        # ===== 对比横表 =====
        wb = w.book
        ws_h = wb.create_sheet('对比横表')

        from openpyxl.styles import PatternFill, Alignment

        color1 = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        color2 = PatternFill(start_color='E5FFCC', end_color='E5FFCC', fill_type='solid')
        color_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        center = Alignment(horizontal='center', vertical='center')

        # 写入横表表头（第1列）
        for ri, h in enumerate(['轮次','来源','业务类型','开始时间','结束时间',
                                '速率类指标','数值','时长类指标','数值.1'], 1):
            ws_h.cell(row=ri, column=1, value=h).alignment = center

        # 按顺序写入：每2列一组（基准+代码），跟踪轮次边界用于合并
        col = 2
        color_toggle = 0
        prev_rnd = None
        rnd_start_col = 2  # 当前轮次的起始列

        for ci in range(0, len(df_compare), 2):
            if ci + 1 >= len(df_compare):
                break
            r1 = df_compare.iloc[ci]
            r2 = df_compare.iloc[ci + 1]

            rnd = r1['轮次']
            biz = r1['业务类型']

            # 轮次切换：先合并上一轮的行1
            if rnd != prev_rnd:
                if prev_rnd is not None and col > rnd_start_col:
                    ws_h.merge_cells(start_row=1, start_column=rnd_start_col,
                                     end_row=1, end_column=col - 1)
                color_toggle = 1 - color_toggle
                prev_rnd = rnd
                rnd_start_col = col

            fill = color_yellow if '不计' in str(rnd) else (color1 if color_toggle else color2)

            # 基准列 + 代码列
            for ri, key in enumerate(['轮次','来源','业务类型','开始时间','结束时间',
                                       '速率类指标','数值','时长类指标','数值.1']):
                c = ws_h.cell(row=ri+1, column=col, value=r1[key] if pd.notna(r1.get(key)) else '')
                c.fill = fill; c.alignment = center

            for ri, key in enumerate(['轮次','来源','业务类型','开始时间','结束时间',
                                       '速率类指标','数值','时长类指标','数值.1']):
                c = ws_h.cell(row=ri+1, column=col+1, value=r2[key] if pd.notna(r2.get(key)) else '')
                c.fill = fill; c.alignment = center

            # 合并第3行（业务类型跨基准+代码）
            ws_h.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)

            col += 2

        # 最后一轮的合并
        if prev_rnd is not None and col > rnd_start_col:
            ws_h.merge_cells(start_row=1, start_column=rnd_start_col,
                             end_row=1, end_column=col - 1)

    log(f"完成! {out}")
    return out


# ===== GUI 包装 =====
try:
    from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QPushButton, QLabel, QFileDialog, QMessageBox, QProgressBar, QTextEdit,
        QGroupBox, QCheckBox, QSpinBox)
    from PySide6.QtCore import Qt, QThread, Signal, Slot
    from PySide6.QtGui import QFont
    GUI_OK = True
except:
    GUI_OK = False


if GUI_OK:
    class WorkerThread(QThread):
        progress = Signal(str)
        done = Signal(str)

        def __init__(self, files, qci_list, dl_clip, ul_clip):
            super().__init__()
            self.files = files; self.qci_list = qci_list; self.dl_clip = dl_clip; self.ul_clip = ul_clip

        def run(self):
            out = process(self.files, self.qci_list, self.dl_clip, self.ul_clip, callback=self.l)
            self.done.emit(out)

        def l(self, msg):
            self.progress.emit(msg)

    class MainWindow(QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("5G用户级公共监控速率统计工具 V2.0.9")
            self.setMinimumSize(800, 600); self.resize(900, 650)
            self.files = []
            cw = QWidget(); self.setCentralWidget(cw); lo = QVBoxLayout(cw)
            lb = QLabel("5G用户级公共监控速率统计工具 V2.0.9")
            lb.setFont(QFont("Arial", 14, QFont.Bold)); lb.setAlignment(Qt.AlignCenter); lo.addWidget(lb)

            gf = QGroupBox("输入文件"); fl = QVBoxLayout(gf)
            bf = QHBoxLayout()
            self.bs = QPushButton("选择MMF文件"); self.bs.clicked.connect(self.sf); bf.addWidget(self.bs)
            self.bc = QPushButton("清空"); self.bc.clicked.connect(self.cf); bf.addWidget(self.bc)
            fl.addLayout(bf)
            self.te = QTextEdit(); self.te.setReadOnly(True); self.te.setMaximumHeight(80); fl.addWidget(self.te)
            lo.addWidget(gf)

            gp = QGroupBox("参数设置"); pl = QHBoxLayout(gp)
            pl.addWidget(QLabel("QCI:"))
            self.cb5 = QCheckBox("5"); self.cb6 = QCheckBox("6"); self.cb7 = QCheckBox("7")
            self.cb7.setChecked(True); self.cb6.setChecked(True); self.cb5.setChecked(True)
            pl.addWidget(self.cb5); pl.addWidget(self.cb6); pl.addWidget(self.cb7)
            pl.addWidget(QLabel("下行削峰:")); self.sd = QSpinBox(); self.sd.setRange(100, 5000); self.sd.setValue(1000)
            pl.addWidget(self.sd)
            pl.addWidget(QLabel("上行削峰:")); self.su = QSpinBox(); self.su.setRange(50, 1000); self.su.setValue(200)
            pl.addWidget(self.su)
            lo.addWidget(gp)

            gh = QGroupBox("运行"); hl = QHBoxLayout(gh)
            self.br = QPushButton("开始处理"); self.br.clicked.connect(self.run); self.br.setEnabled(False)
            self.babout = QPushButton("关于"); self.babout.clicked.connect(self.about)
            hl.addWidget(self.br); hl.addWidget(self.babout); lo.addWidget(gh)

            self.pb = QProgressBar(); lo.addWidget(self.pb)
            self.lt = QTextEdit(); self.lt.setReadOnly(True); lo.addWidget(self.lt)

        def sf(self):
            fs, _ = QFileDialog.getOpenFileNames(self, "选择MMF文件", "", "Excel (*.xlsx *.xls);;All (*)")
            if fs: self.files = fs; self.te.setText("\n".join(fs)); self.br.setEnabled(True)

        def cf(self): self.files = []; self.te.clear(); self.br.setEnabled(False)

        def run(self):
            if not self.files: QMessageBox.warning(self, "提示", "请先选择MMF文件"); return
            qci_list = []
            if self.cb5.isChecked(): qci_list.append(5)
            if self.cb6.isChecked(): qci_list.append(6)
            if self.cb7.isChecked(): qci_list.append(7)
            if not qci_list: qci_list = [7]
            self.br.setEnabled(False); self.lt.append(f"开始处理... QCI={qci_list}")
            self.thread = WorkerThread(self.files, qci_list, self.sd.value(), self.su.value())
            self.thread.progress.connect(self.upd); self.thread.done.connect(self.dn); self.thread.start()

        @Slot(str)
        def upd(self, msg): self.lt.append(msg); self.pb.setValue(min(self.pb.value() + 30, 95))

        @Slot(str)
        def dn(self, path):
            self.pb.setValue(100)
            if path: self.lt.append(f"完成! {path}"); os.system(f'open "{path}"')
            self.br.setEnabled(True)

        def about(self):
            msg = ("5G用户级公共监控速率统计工具\n版本: V2.0.9\n"
                   "开发者: 孙晓军  联系方式: 317827@qq.com\n\n"
                   "更新内容:\n"
                   "V2.0.9: 修复对比横表sheet代码业务结束时间缺失bug\n"
                   "  新增第一个FTP对之前不完整轮次的识别\n"
                   "  版本号统一更新(文件名/关于/需求文档)\n"
                   "V2.0.1: QCI支持5/6/7多选, 主界面增加关于\n"
                   "V2.0.0: 从零重写, 三页向导改单页, 自动识别FTP/商店/微信6类业务\n"
                   "  从原始MMF文件直接生成详细过程/对比/汇总3个Sheet\n"
                   "  削峰阈值可调, 轮次完整性检查")
            QMessageBox.about(self, "关于", msg)


if __name__ == '__main__':
    if GUI_OK and len(sys.argv) == 1:
        app = QApplication(sys.argv); app.setStyle('Fusion')
        w = MainWindow(); w.show(); sys.exit(app.exec())
    else:
        out = process(
            ['联通/mmf20260703115328-电信.xlsx', '联通/mmf20260703115334-电信.xlsx'],
            qci_list=[5,6,7], dl_clip=1000, ul_clip=200
        )
        print(f"输出: {out}")