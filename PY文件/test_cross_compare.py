#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
交叉对比：参考文件中的26个store_l vs 工具识别出的22个store_l
找出漏掉的4个以及原因
"""

import sys, os, json, re
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
import numpy as np
import pandas as pd

sys.path.insert(0, '/Users/sun/ClaudeCode/先甲工信部工具')
from importlib import util as import_util
spec = import_util.spec_from_file_location("tool",
    "/Users/sun/ClaudeCode/先甲工信部工具/5G用户级公共监控速率统计工具-V1.04.py")
tool = import_util.module_from_spec(spec)
spec.loader.exec_module(tool)

uproc = tool.UCMProcessor({'dl_peak_limit': 900, 'ul_peak_limit': 160})
paths = [
    '/Users/sun/ClaudeCode/先甲工信部工具/联通/mmf20260703115328-电信.xlsx',
    '/Users/sun/ClaudeCode/先甲工信部工具/联通/mmf20260703115334-电信.xlsx',
    '/Users/sun/ClaudeCode/先甲工信部工具/联通/mmf20260703115336-联通.xlsx'
]
params = {'dl_peak_limit': 900, 'ul_peak_limit': 160, 'down_min': 50, 'up_min': 10, 'match_mode': 'auto'}
agg = uproc.parse_mmf(paths, params)
uproc._fn = 'mmf_test'

# 从参考文件中提取26个store_l的时间段
import openpyxl
ref = '/Users/sun/ClaudeCode/先甲工信部工具/联通/呼叫详情_按时间排序-20260707.xlsx'
wb = openpyxl.load_workbook(ref, data_only=True)
ws = wb['联通']
ref_store_l = []
for r in range(2, ws.max_row + 1):
    biz = str(ws.cell(r, 2).value or '')
    if '应用商店' not in biz:
        continue
    # 每3行一组，第3行是store_l
    # 简单判断：没有大文件下载速率且名字是"应用商店大文件下载"
    # 或者用行号分组
    st = ws.cell(r, 3).value
    et = ws.cell(r, 4).value
    big = ws.cell(r, 11).value
    ref_store_l.append({'st': st, 'et': et, 'big_rate': big})

# 每3行一组，取第3行
groups = [(ref_store_l[i], ref_store_l[i+1], ref_store_l[i+2]) for i in range(0, len(ref_store_l), 3)]

# 工具识别出的store_l时间段
tool_store_l_times = []
for rd in uproc._match_auto(agg, params, 50, 10):
    for k, s in rd.items():
        if k == 'store_l':
            tool_store_l_times.append({
                'st': s['start_t'], 'et': s['end_t'],
                'flow': s['flow_mb'], 'dur': s['duration']
            })

print(f"参考文件 store_l 组数: {len(groups)}")
print(f"工具识别 store_l 数: {len(tool_store_l_times)}")

# 对每个参考store_l，看工具是否识别到了
print("\n===== 交叉对比 =====")
for i, (total, store_s, store_l) in enumerate(groups):
    ref_st = pd.to_datetime(store_l['st'])
    ref_et = pd.to_datetime(store_l['et'])
    # 找工具中时间最接近的store_l
    found = False
    for t in tool_store_l_times:
        # 允许5秒容差
        st_gap = abs((t['st'] - ref_st).total_seconds())
        et_gap = abs((t['et'] - ref_et).total_seconds())
        if st_gap < 5 and et_gap < 5:
            found = True
            break
    if not found:
        # 看看这个时间段内有什么下行段
        mask = (agg['t'] >= ref_st) & (agg['t'] <= ref_et)
        sub = agg[mask]
        dl_flow = sub['dl_mac'].sum() / 8
        dur = len(sub)
        avg_rate = sub['dl_rlc'].replace(0, np.nan).mean()
        print(f"  MISS 组{i}: {ref_st.strftime('%H:%M:%S')}-{ref_et.strftime('%H:%M:%S')} "
              f"flow={dl_flow:.0f}MB dur={dur}s rate={avg_rate:.1f}")

# 检查漏掉的到底是哪些段
print("\n===== 漏检分析：检查这些时间段内的段检测情况 =====")
# 先获取所有段的分类
n = len(agg)
segs = []
i = 0
dl_min, ul_min = 50, 10
while i < n:
    if agg.at[i, 'dl_mac'] > dl_min or agg.at[i, 'ul_mac'] > ul_min:
        start_i = i; gap = 0
        while i < n and gap < 5:
            if agg.at[i, 'dl_mac'] > dl_min or agg.at[i, 'ul_mac'] > ul_min:
                gap = 0
            else:
                gap += 1
            i += 1
        end_i = max(start_i, i - 1 - gap)
        seg_df = agg.loc[start_i:end_i]
        ul_act = int((seg_df['ul_mac'] > 10).sum())
        dl_act = int((seg_df['dl_mac'] > 50).sum())
        if ul_act > dl_act + 3:
            direction, col = '上行', 'ul_mac'
        else:
            direction, col = '下行', 'dl_mac'
        flow = seg_df[col].sum() / 8
        dur = max(round((seg_df['t'].iloc[-1] - seg_df['t'].iloc[0]).total_seconds(), 1), 0.5)
        rv = seg_df['dl_rlc' if direction == '下行' else 'ul_rlc'].replace(0, np.nan).dropna()
        rate = round(float(rv.mean()), 3) if len(rv) else 0
        segs.append({'start_i': start_i, 'end_i': end_i, 'start_t': seg_df['t'].iloc[0],
                     'end_t': seg_df['t'].iloc[-1], 'direction': direction,
                     'flow_mb': round(flow, 1), 'duration': round(dur, 1), 'rate': rate})
    else:
        i += 1

# 合并（gap<10）
merged = []
for s in segs:
    if s.get('duration', 0) < 1 and s.get('flow_mb', 0) < 10:
        continue
    if not merged:
        merged.append(s)
        continue
    prev = merged[-1]
    gap = (s['start_t'] - prev['end_t']).total_seconds()
    if s['direction'] == prev['direction'] and 0 < gap < 10:
        seg_df = agg.loc[prev['start_i']:s['end_i']]
        col = 'dl_mac' if s['direction'] == '下行' else 'ul_mac'
        rlc_col = 'dl_rlc' if s['direction'] == '下行' else 'ul_rlc'
        clip_col = 'dl_clip' if s['direction'] == '下行' else 'ul_clip'
        flow = seg_df[col].sum() / 8
        dur = round((seg_df['t'].iloc[-1] - seg_df['t'].iloc[0]).total_seconds(), 1)
        rv = seg_df[rlc_col].replace(0, np.nan).dropna()
        cv = seg_df[clip_col].replace(0, np.nan).dropna()
        merged[-1] = {'start_i': prev['start_i'], 'end_i': s['end_i'],
                      'start_t': prev['start_t'], 'end_t': seg_df['t'].iloc[-1],
                      'direction': s['direction'], 'flow_mb': round(flow, 1),
                      'duration': round(dur, 1),
                      'rate': round(float(rv.mean()), 3) if len(rv) else 0,
                      'clip_rate': round(float(cv.mean()), 3) if len(cv) else 0}

# 对每个漏掉的参考store_l，看合并后有哪些段
for i, (total, store_s, store_l) in enumerate(groups):
    ref_st = pd.to_datetime(store_l['st'])
    ref_et = pd.to_datetime(store_l['et'])
    found = False
    for t in tool_store_l_times:
        st_gap = abs((t['st'] - ref_st).total_seconds())
        et_gap = abs((t['et'] - ref_et).total_seconds())
        if st_gap < 5 and et_gap < 5:
            found = True
            break
    if not found:
        # 看这个时间段内合并后的段
        # 先找所有下行段，即使不在这个时间段内，只要和这个时间段重叠
        dl_segs = [s for s in merged if s['direction'] == '下行'
                   and s['start_t'] <= ref_et and s['end_t'] >= ref_st]
        print(f"\n MISS 组{i}: {ref_st.strftime('%H:%M:%S')}-{ref_et.strftime('%H:%M:%S')}")
        if not dl_segs:
            # 没有合并段覆盖？那看原始段
            print(f"    合并后无覆盖！检查原始段:")

            raw_segs_in_range = [s for s in segs if s['direction'] == '下行'
                                 and s['start_t'] <= ref_et and s['end_t'] >= ref_st]
            for s in raw_segs_in_range:
                print(f"    raw: flow={s['flow_mb']:6.0f}MB dur={s['duration']:5.1f}s rate={s['rate']:6.1f} "
                      f"({s['start_t'].strftime('%H:%M:%S')}-{s['end_t'].strftime('%H:%M:%S')})")
        else:
            for s in dl_segs:
                # 分类
                k = '?'
                f = s['flow_mb']
                if f <= 50: k = 'None'
                elif f < 350: k = 'store_s'
                elif f < 700: k = 'ftp_dl'
                elif f >= 1500: k = 'store_l'
                elif s['duration'] >= 15 or s['rate'] < 500: k = 'store_l'
                else: k = 'ftp_dl'
                print(f"    段: flow={s['flow_mb']:6.0f}MB dur={s['duration']:5.1f}s rate={s['rate']:6.1f} -> {k} "
                      f"({s['start_t'].strftime('%H:%M:%S')}-{s['end_t'].strftime('%H:%M:%S')})")