#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
5G用户级公共监控速率统计工具 V1.00
基于2025-2026年工信部测评用例规则开发
作者：Claude Code
创建日期：2026-06-23
"""

import sys
import os
import json
import re
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
import numpy as np
import pandas as pd
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QLabel, QLineEdit, QPushButton, QFileDialog,
    QTextEdit, QTableWidget, QTableWidgetItem, QHeaderView,
    QGroupBox, QFormLayout, QSpinBox, QDoubleSpinBox, QComboBox,
    QCheckBox, QMessageBox, QProgressBar, QSplitter
)
from PySide6.QtCore import Qt, QThread, Signal, Slot
from PySide6.QtGui import QFont, QColor
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ==================== 配置管理 ====================

class ConfigManager:
    """配置文件管理器"""

    def __init__(self, config_path: str = None):
        if config_path is None:
            config_path = os.path.join(os.path.dirname(__file__), 'config', 'default_rules.json')
        self.config_path = config_path
        self.config = self.load_config()

    def load_config(self) -> Dict:
        """加载配置文件"""
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载配置文件失败: {e}")
            return self.get_default_config()

    def save_config(self, config: Dict = None) -> bool:
        """保存配置文件"""
        if config is None:
            config = self.config
        try:
            os.makedirs(os.path.dirname(self.config_path), exist_ok=True)
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"保存配置文件失败: {e}")
            return False

    def get_default_config(self) -> Dict:
        """获取默认配置"""
        return {
            "version": "1.00",
            "rule_version": "2025-2026工信部测评用例",
            "global_params": {
                "qci": 9,
                "start_time": "auto",
                "end_time": "auto",
                "dl_peak_limit": 900,
                "ul_peak_limit": 160,
                "dl_pass_threshold": 100,
                "ul_pass_threshold": 20,
                "scenario_type": "场景",
                "road_mean_threshold": 800
            },
            "business_rules": [],
            "voice_rules": [],
            "ott_rules": []
        }


# ==================== 数据处理引擎 ====================

class DataProcessor:
    """数据处理核心引擎"""

    def __init__(self, config: Dict):
        self.config = config
        self.raw_data = None
        self.cleaned_data = None
        self.filtered_data = None
        self.rate_details = None
        self.statistics = None

    def parse_input_file(self, file_path: str) -> pd.DataFrame:
        """
        解析输入文件
        跳过前7行元数据，从第8行读取表头（header=7）
        """
        try:
            print("DEBUG: Reading Excel file...")
            # 读取Excel文件，跳过前7行，第8行作为表头（header=7）
            df = pd.read_excel(file_path, header=7)

            print(f"DEBUG: Loaded {len(df)} rows, {len(df.columns)} columns")
            print(f"DEBUG: First 5 columns: {df.columns[:5].tolist()}")

            # 解析采集时间（去除毫秒后缀）
            if '采集时间' in df.columns:
                df['采集时间_解析'] = df['采集时间'].apply(self._parse_time)
                print(f"DEBUG: Parsed collection time for {df['采集时间_解析'].notna().sum()} rows")
            else:
                print("DEBUG: '采集时间' column not found!")

            return df
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise Exception(f"解析输入文件失败: {e}")

    def _parse_time(self, time_str: str) -> Optional[datetime]:
        """解析采集时间字符串，格式：YYYY-MM-DD HH:MM:SS (xxx)"""
        if pd.isna(time_str) or time_str == "N/A":
            return None
        try:
            # 使用正则表达式提取时间部分
            match = re.match(r'^(.+?)\s*\(\d+\)$', str(time_str))
            if match:
                time_str = match.group(1)
            return pd.to_datetime(time_str)
        except:
            return None

    def clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        数据清洗
        1. "-" 前向填充
        2. bps→Mbps转换（RLC吞吐率）
        """
        df = df.copy()

        # 处理 "-" 前向填充
        df = df.replace('-', np.nan).ffill()

        # bps→Mbps转换
        for col in df.columns:
            if 'RLC吞吐率' in col and 'Mbps' not in col:
                new_col = col.replace('bps', 'Mbps')
                if new_col != col and new_col not in df.columns:
                    df[new_col] = df[col] / 1_000_000

        return df

    def filter_data(self, df: pd.DataFrame, params: Dict) -> pd.DataFrame:
        """
        过滤筛选
        1. 时间窗过滤
        2. QCI过滤
        3. 载波聚合属性过滤
        """
        df = df.copy()
        print(f"DEBUG: filter_data input: {len(df)} rows")

        # 时间窗过滤
        if '采集时间_解析' in df.columns:
            if params.get('start_time') and params['start_time'] != 'auto':
                start_time = pd.to_datetime(params['start_time'])
                df = df[df['采集时间_解析'] >= start_time]
            if params.get('end_time') and params['end_time'] != 'auto':
                end_time = pd.to_datetime(params['end_time'])
                df = df[df['采集时间_解析'] <= end_time]

        # QCI过滤（0表示不过滤）
        qci_value = params.get('qci', 0)
        if 'QCI' in df.columns and qci_value and qci_value > 0:
            before_qci = len(df)
            df = df[df['QCI'] == qci_value]
            print(f"DEBUG: QCI={qci_value} filter: {before_qci} -> {len(df)} rows")
        else:
            print(f"DEBUG: No QCI filtering (qci={qci_value})")

        # 载波聚合属性过滤
        if '载波聚合属性' in df.columns:
            before_ca = len(df)
            df = df[df['载波聚合属性'].isin(['非CA用户', 'CA用户主小区'])]
            print(f"DEBUG: CA filter: {before_ca} -> {len(df)} rows")

        print(f"DEBUG: filter_data output: {len(df)} rows")
        return df

    def identify_direction(self, df: pd.DataFrame, rules: List[Dict]) -> pd.DataFrame:
        """
        方向识别
        1. 计算上下行比率
        2. 判定上行/下行有效点
        3. 同一秒同向保留最大速率点
        """
        print(f"DEBUG: identify_direction input rows: {len(df)}")

        if df.empty:
            print("DEBUG: identify_direction: df is empty")
            return df

        df = df.copy()

        # 获取速率列（优先使用MAC吞吐率，因为覆盖率100%）
        # 支持带和不带后缀的列名
        dl_col = None
        ul_col = None

        for col in df.columns:
            if '下行MAC吞吐率' in col or '下行RLC吞吐率' in col:
                # 优先选择MAC
                if dl_col is None or ('MAC' in col and 'MAC' not in dl_col):
                    dl_col = col

        for col in df.columns:
            if '上行MAC吞吐率' in col or '上行RLC吞吐率' in col:
                # 优先选择MAC
                if ul_col is None or ('MAC' in col and 'MAC' not in ul_col):
                    ul_col = col

        if not dl_col or not ul_col:
            print(f"DEBUG: Missing columns - dl: {dl_col}, ul: {ul_col}")
            return df

        print(f"DEBUG: Using columns - DL: {dl_col}, UL: {ul_col}")

        # 确保速率列是数值类型
        df[dl_col] = pd.to_numeric(df[dl_col], errors='coerce').fillna(0)
        df[ul_col] = pd.to_numeric(df[ul_col], errors='coerce').fillna(0)

        # 获取最小速率阈值
        up_min = min([r.get('min_rate', 3) for r in rules if r.get('direction') == '上传'], default=3)
        down_min = min([r.get('min_rate', 10) for r in rules if r.get('direction') == '下载'], default=10)

        print(f"DEBUG: up_min={up_min}, down_min={down_min}")

        # 计算方向标记
        df['方向'] = ''
        df['速率_有效'] = 0.0

        down_count = 0
        up_count = 0

        for idx, row in df.iterrows():
            dl_rate = row[dl_col]
            ul_rate = row[ul_col]

            # 判定方向
            if dl_rate > 0 and ul_rate > 0:
                u_d_ratio = ul_rate / dl_rate if dl_rate > 0 else 0
                d_u_ratio = dl_rate / ul_rate if ul_rate > 0 else 0

                if u_d_ratio > 10 and ul_rate > up_min:
                    df.loc[idx, '方向'] = '上行'
                    df.loc[idx, '速率_有效'] = ul_rate
                    up_count += 1
                elif d_u_ratio > 10 and dl_rate > down_min:
                    df.loc[idx, '方向'] = '下行'
                    df.loc[idx, '速率_有效'] = dl_rate
                    down_count += 1
            elif dl_rate > down_min:
                df.loc[idx, '方向'] = '下行'
                df.loc[idx, '速率_有效'] = dl_rate
                down_count += 1
            elif ul_rate > up_min:
                df.loc[idx, '方向'] = '上行'
                df.loc[idx, '速率_有效'] = ul_rate
                up_count += 1

        print(f"DEBUG: Direction identified - down: {down_count}, up: {up_count}")

        # 同一秒同向保留最大速率点
        if '采集时间_解析' in df.columns:
            df['时间秒'] = df['采集时间_解析'].dt.floor('s')
            before_dedup = len(df)
            # 只对有方向的点进行去重
            valid_points = df[df['方向'] != '']
            if len(valid_points) > 0:
                df = df.loc[valid_points.groupby(['时间秒', '方向'])['速率_有效'].idxmax()]
                df = df.drop(columns=['时间秒'])
                print(f"DEBUG: Dedup: {before_dedup} -> {len(df)} rows")

        print(f"DEBUG: identify_direction output rows: {len(df)}")
        return df

    def segment_business(self, df: pd.DataFrame, rules: List[Dict]) -> pd.DataFrame:
        """
        业务分段
        1. 同向连续点为一段
        2. 相邻时间差 > 8秒截断
        3. 按流量区间和时长匹配业务类型
        """
        if df.empty or '方向' not in df.columns:
            return df

        df = df.copy()
        df = df[df['方向'] != ''].sort_values('采集时间_解析')

        # 添加段ID
        df['段ID'] = 0
        current_segment = 0
        last_time = None
        last_direction = None

        for idx, row in df.iterrows():
            current_time = row['采集时间_解析']
            current_direction = row['方向']

            if last_time is None:
                current_segment += 1
            elif current_direction != last_direction:
                current_segment += 1
            elif (current_time - last_time).total_seconds() > 8:
                current_segment += 1

            df.loc[idx, '段ID'] = current_segment
            last_time = current_time
            last_direction = current_direction

        # 计算每段的统计信息
        segments = []
        for segment_id, segment_df in df.groupby('段ID'):
            if segment_df.empty:
                continue

            direction = segment_df['方向'].iloc[0]
            duration = len(segment_df)
            total_flow = segment_df['速率_有效'].sum()

            # 匹配业务类型
            business_type = self._match_business(direction, duration, total_flow, rules)

            segments.append({
                '段ID': segment_id,
                '方向': direction,
                '时长': duration,
                '总流量': total_flow,
                '业务类型': business_type
            })

        # 将业务类型关联回原数据
        segment_df = pd.DataFrame(segments)
        df = df.merge(segment_df[['段ID', '业务类型']], on='段ID', how='left')

        return df

    def _match_business(self, direction: str, duration: float, total_flow: float, rules: List[Dict]) -> str:
        """匹配业务类型（参考原工具V6逻辑）"""
        # 方向映射：数据处理中的"下行"/"上行"对应业务规则中的"下载"/"上传"
        direction_map = {'下行': '下载', '上行': '上传'}
        mapped_direction = direction_map.get(direction, direction)

        # 筛选同方向的业务
        same_direction_rules = [r for r in rules if r.get('direction') == mapped_direction and r.get('enabled', True)]

        # 优先级1：有流量区间的业务（应用宝、微信等）
        for rule in same_direction_rules:
            flow_min = rule.get('flow_min')
            flow_max = rule.get('flow_max')

            if flow_min is not None and flow_max is not None:
                flow_mb = total_flow / 8  # Mbps·s → MB
                # 检查流量区间（秒数可以为null）
                calc_seconds = rule.get('calc_seconds')
                if calc_seconds is None or duration >= calc_seconds:
                    if flow_min <= flow_mb <= flow_max:
                        return rule['name']

        # 优先级2：FTP业务（有明确的秒数和最小速率要求）
        for rule in same_direction_rules:
            if 'FTP' in rule['name']:
                calc_seconds = rule.get('calc_seconds')
                min_rate = rule.get('min_rate')

                # 检查秒数
                if calc_seconds is not None and duration < calc_seconds:
                    continue

                # 检查平均速率
                avg_rate = total_flow / duration if duration > 0 else 0
                if min_rate is not None and avg_rate >= min_rate:
                    return rule['name']

        return '未知业务'

    def apply_peak_limiting(self, df: pd.DataFrame, params: Dict) -> pd.DataFrame:
        """削峰处理"""
        df = df.copy()

        dl_limit = params.get('dl_peak_limit', 900)
        ul_limit = params.get('ul_peak_limit', 160)

        # 确定使用哪个速率列（优先MAC，因为覆盖率更高）
        # 支持带和不带后缀的列名
        dl_rate_col = None
        ul_rate_col = None

        for col in df.columns:
            if '下行MAC吞吐率' in col or '下行RLC吞吐率' in col:
                if dl_rate_col is None or ('MAC' in col and 'MAC' not in dl_rate_col):
                    dl_rate_col = col

        for col in df.columns:
            if '上行MAC吞吐率' in col or '上行RLC吞吐率' in col:
                if ul_rate_col is None or ('MAC' in col and 'MAC' not in ul_rate_col):
                    ul_rate_col = col

        print(f"DEBUG: Using columns for peak limiting - DL: {dl_rate_col}, UL: {ul_rate_col}")

        if dl_rate_col:
            df['削峰后下行RLC'] = df[dl_rate_col].apply(lambda x: min(x, dl_limit) if pd.notna(x) else x)
        if ul_rate_col:
            df['削峰后上行RLC'] = df[ul_rate_col].apply(lambda x: min(x, ul_limit) if pd.notna(x) else x)

        return df

    def calculate_statistics(self, df: pd.DataFrame, params: Dict) -> Dict:
        """
        统计计算
        1. 按业务类型聚合
        2. 计算采样点数、均值、峰值
        3. 计算达标占比
        4. 计算TOP10%峰值
        """
        if df.empty:
            print("DEBUG: DataFrame is empty")
            return {}

        if '业务类型' not in df.columns:
            print("DEBUG: '业务类型' column not found in DataFrame")
            print(f"DEBUG: Available columns: {df.columns.tolist()[:10]}...")
            return {}

        statistics = {}
        dl_pass_threshold = params.get('dl_pass_threshold', 100)
        ul_pass_threshold = params.get('ul_pass_threshold', 20)

        # 确定使用哪个速率列（优先MAC，因为覆盖率更高）
        # 支持带和不带后缀的列名
        dl_rate_col = None
        ul_rate_col = None

        for col in df.columns:
            if '下行MAC吞吐率' in col or '下行RLC吞吐率' in col:
                if dl_rate_col is None or ('MAC' in col and 'MAC' not in dl_rate_col):
                    dl_rate_col = col

        for col in df.columns:
            if '上行MAC吞吐率' in col or '上行RLC吞吐率' in col:
                if ul_rate_col is None or ('MAC' in col and 'MAC' not in ul_rate_col):
                    ul_rate_col = col

        print(f"DEBUG: Using columns for statistics - DL: {dl_rate_col}, UL: {ul_rate_col}")
        print(f"DEBUG: Processing {len(df)} rows with {len(df['业务类型'].unique())} business types")

        # 按业务类型分组统计
        for business_type, group_df in df.groupby('业务类型'):
            if business_type == '未知业务' or business_type != business_type:  # 跳过NaN
                continue

            direction = group_df['方向'].iloc[0] if '方向' in group_df.columns else ''
            stats = {
                '业务类型': business_type,
                '方向': direction,
                '测试次数': group_df['段ID'].nunique() if '段ID' in group_df.columns else 0,
                '采样点数': len(group_df)
            }

            # 下行统计
            if direction == '下行' and dl_rate_col and dl_rate_col in group_df.columns:
                dl_rates = group_df[dl_rate_col].dropna()
                if not dl_rates.empty:
                    stats['均值速率'] = dl_rates.mean()
                    stats['峰值速率'] = dl_rates.max()

                    # 削峰后统计
                    if '削峰后下行RLC' in group_df.columns:
                        clipped_dl = group_df['削峰后下行RLC'].dropna()
                        stats['削峰后均值速率'] = clipped_dl.mean()
                        stats['达标占比'] = (clipped_dl >= dl_pass_threshold).sum() / len(clipped_dl) * 100 if len(clipped_dl) > 0 else 0

                        # TOP10%峰值
                        sorted_rates = clipped_dl.sort_values(ascending=False)
                        top10_count = max(1, int(len(sorted_rates) * 0.1))
                        stats['TOP10峰值'] = sorted_rates.head(top10_count).mean()

            # 上行统计
            if direction == '上传' and ul_rate_col and ul_rate_col in group_df.columns:
                ul_rates = group_df[ul_rate_col].dropna()
                if not ul_rates.empty:
                    stats['均值速率'] = ul_rates.mean()
                    stats['峰值速率'] = ul_rates.max()

                    # 削峰后统计
                    if '削峰后上行RLC' in group_df.columns:
                        clipped_ul = group_df['削峰后上行RLC'].dropna()
                        stats['削峰后均值速率'] = clipped_ul.mean()
                        stats['达标占比'] = (clipped_ul >= ul_pass_threshold).sum() / len(clipped_ul) * 100 if len(clipped_ul) > 0 else 0

                        # TOP10%峰值
                        sorted_rates = clipped_ul.sort_values(ascending=False)
                        top10_count = max(1, int(len(sorted_rates) * 0.1))
                        stats['TOP10峰值'] = sorted_rates.head(top10_count).mean()

            statistics[business_type] = stats

        return statistics

    def process(self, file_path: str, params: Dict, rules: List[Dict]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict]:
        """
        完整处理流程
        返回：(原表整理, 速率详表, 统计结果)
        """
        # 1. 解析输入文件
        self.raw_data = self.parse_input_file(file_path)

        # 2. 数据清洗
        self.cleaned_data = self.clean_data(self.raw_data)

        # 3. 过滤筛选
        self.filtered_data = self.filter_data(self.cleaned_data, params)

        # 4. 方向识别
        rate_df = self.identify_direction(self.filtered_data, rules)

        # 5. 业务分段
        rate_df = self.segment_business(rate_df, rules)

        # 6. 削峰处理
        rate_df = self.apply_peak_limiting(rate_df, params)

        # 7. 统计计算
        self.statistics = self.calculate_statistics(rate_df, params)

        self.rate_details = rate_df

        return self.cleaned_data, rate_df, self.statistics


# ==================== Excel导出 ====================

class ExcelExporter:
    """Excel导出器"""

    def __init__(self):
        self.wb = None

    def export(self, output_path: str, cleaned_data: pd.DataFrame, rate_details: pd.DataFrame, statistics: Dict):
        """导出Excel文件，包含3个sheet"""
        self.wb = Workbook()

        # 删除默认sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # 创建三个sheet
        self._create_statistics_sheet(statistics)
        self._create_rate_details_sheet(rate_details)
        self._create_cleaned_data_sheet(cleaned_data)

        # 保存文件
        self.wb.save(output_path)

    def _create_statistics_sheet(self, statistics: Dict):
        """创建统计结果sheet"""
        ws = self.wb.create_sheet('统计结果')

        # 表头
        headers = ['业务类型', '方向', '测试次数', '采样点数', '均值速率(Mbps)',
                   '峰值速率(Mbps)', '削峰后均值(Mbps)', '达标占比(%)', 'TOP10%峰值(Mbps)']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')

        # 数据行
        for row, (business_type, stats) in enumerate(statistics.items(), 2):
            ws.cell(row=row, column=1, value=business_type)
            ws.cell(row=row, column=2, value=stats.get('方向', ''))
            ws.cell(row=row, column=3, value=stats.get('测试次数', 0))
            ws.cell(row=row, column=4, value=stats.get('采样点数', 0))
            ws.cell(row=row, column=5, value=round(stats.get('均值速率', 0), 2))
            ws.cell(row=row, column=6, value=round(stats.get('峰值速率', 0), 2))
            ws.cell(row=row, column=7, value=round(stats.get('削峰后均值速率', 0), 2))
            ws.cell(row=row, column=8, value=round(stats.get('达标占比', 0), 2))
            ws.cell(row=row, column=9, value=round(stats.get('TOP10峰值', 0), 2))

    def _create_rate_details_sheet(self, rate_details: pd.DataFrame):
        """创建速率详表sheet"""
        ws = self.wb.create_sheet('速率详表')

        if rate_details.empty:
            ws.cell(row=1, column=1, value='无数据')
            return

        # 选择关键列
        key_columns = ['序号', '采集时间', '来源', '基站标识', '小区标识', '业务类型',
                       '下行RLC吞吐率', '上行RLC吞吐率', '削峰后下行RLC', '削峰后上行RLC', '方向']

        available_columns = [col for col in key_columns if col in rate_details.columns]

        # 表头
        for col, header in enumerate(available_columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # 数据行
        for row_idx, row_data in rate_details.iterrows():
            for col, col_name in enumerate(available_columns, 1):
                value = row_data.get(col_name, '')
                if pd.isna(value):
                    value = ''
                ws.cell(row=row_idx + 2, column=col, value=value)

    def _create_cleaned_data_sheet(self, cleaned_data: pd.DataFrame):
        """创建原表整理sheet"""
        ws = self.wb.create_sheet('原表整理')

        if cleaned_data.empty:
            ws.cell(row=1, column=1, value='无数据')
            return

        # 表头
        for col, header in enumerate(cleaned_data.columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # 数据行
        for row_idx, row_data in cleaned_data.iterrows():
            for col, value in enumerate(row_data, 1):
                if pd.isna(value):
                    value = ''
                ws.cell(row=row_idx + 2, column=col, value=value)


# ==================== 工作线程 ====================

class ProcessingThread(QThread):
    """数据处理工作线程"""
    progress_signal = Signal(str)
    finished_signal = Signal(bool, str, object, object, object)

    def __init__(self, processor: DataProcessor, file_path: str, params: Dict, rules: List[Dict]):
        super().__init__()
        self.processor = processor
        self.file_path = file_path
        self.params = params
        self.rules = rules

    def run(self):
        try:
            print("DEBUG: ProcessingThread.run() started")
            self.progress_signal.emit("正在解析输入文件...")

            cleaned, rate_details, statistics = self.processor.process(
                self.file_path, self.params, self.rules
            )

            self.progress_signal.emit("正在生成结果...")
            print(f"DEBUG: Processing complete, statistics count: {len(statistics)}")
            self.finished_signal.emit(True, "处理完成", cleaned, rate_details, statistics)
        except Exception as e:
            print(f"DEBUG: ProcessingThread error: {e}")
            import traceback
            traceback.print_exc()
            self.finished_signal.emit(False, f"处理失败: {str(e)}", None, None, None)


# ==================== 主窗口 ====================

class MainWindow(QMainWindow):
    """主窗口"""

    def __init__(self):
        super().__init__()
        self.config_manager = ConfigManager()
        self.config = self.config_manager.config
        self.processor = None
        self.input_files = []
        self.processing_thread = None

        self.init_ui()
        self.load_config_to_ui()

    def init_ui(self):
        """初始化UI"""
        self.setWindowTitle('5G用户级公共监控速率统计工具 V1.00')
        self.setGeometry(100, 100, 1400, 900)

        # 主窗口
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)

        # 左侧面板
        left_panel = self.create_left_panel()
        main_layout.addWidget(left_panel, 1)

        # 右侧面板
        right_panel = self.create_right_panel()
        main_layout.addWidget(right_panel, 2)

    def create_left_panel(self) -> QWidget:
        """创建左侧配置面板"""
        panel = QWidget()
        layout = QVBoxLayout(panel)

        # 选项卡
        tab_widget = QTabWidget()

        # 全局参数选项卡
        global_params_tab = self.create_global_params_tab()
        tab_widget.addTab(global_params_tab, '全局参数')

        # 业务规则选项卡
        business_rules_tab = self.create_business_rules_tab()
        tab_widget.addTab(business_rules_tab, '业务规则')

        layout.addWidget(tab_widget)

        # 文件选择区
        file_group = QGroupBox('文件输入')
        file_layout = QVBoxLayout()

        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText('选择输入文件...')
        file_layout.addWidget(self.file_path_edit)

        file_btn_layout = QHBoxLayout()
        select_file_btn = QPushButton('选择文件')
        select_file_btn.clicked.connect(self.select_input_file)
        file_btn_layout.addWidget(select_file_btn)
        file_layout.addLayout(file_btn_layout)

        file_group.setLayout(file_layout)
        layout.addWidget(file_group)

        # 操作按钮
        btn_layout = QHBoxLayout()
        process_btn = QPushButton('开始统计')
        process_btn.setStyleSheet('QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 10px; }')
        process_btn.clicked.connect(self.start_processing)
        btn_layout.addWidget(process_btn)

        export_btn = QPushButton('导出Excel')
        export_btn.clicked.connect(self.export_excel)
        btn_layout.addWidget(export_btn)

        layout.addLayout(btn_layout)

        # 进度条
        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)

        # 日志区
        log_group = QGroupBox('运行日志')
        log_layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout)
        layout.addWidget(log_group, 1)

        return panel

    def create_global_params_tab(self) -> QWidget:
        """创建全局参数选项卡"""
        tab = QWidget()
        layout = QFormLayout(tab)

        # QCI
        self.qci_spin = QSpinBox()
        self.qci_spin.setRange(0, 99)
        self.qci_spin.setValue(0)  # 0表示不过滤QCI
        self.qci_spin.setSpecialValueText("全部")
        layout.addRow('QCI(0=全部):', self.qci_spin)

        # 下行削峰速率
        self.dl_peak_spin = QDoubleSpinBox()
        self.dl_peak_spin.setRange(0, 5000)
        self.dl_peak_spin.setValue(1000)
        layout.addRow('下行削峰速率(Mbps):', self.dl_peak_spin)

        # 上行削峰速率
        self.ul_peak_spin = QDoubleSpinBox()
        self.ul_peak_spin.setRange(0, 1000)
        self.ul_peak_spin.setValue(200)
        layout.addRow('上行削峰速率(Mbps):', self.ul_peak_spin)

        # 下行达标门限
        self.dl_pass_spin = QDoubleSpinBox()
        self.dl_pass_spin.setRange(0, 1000)
        self.dl_pass_spin.setValue(100)
        layout.addRow('下行达标门限(Mbps):', self.dl_pass_spin)

        # 上行达标门限
        self.ul_pass_spin = QDoubleSpinBox()
        self.ul_pass_spin.setRange(0, 100)
        self.ul_pass_spin.setValue(20)
        layout.addRow('上行达标门限(Mbps):', self.ul_pass_spin)

        # 场景类型
        self.scenario_combo = QComboBox()
        self.scenario_combo.addItems(['场景', '道路'])
        layout.addRow('场景类型:', self.scenario_combo)

        # 道路均值门限
        self.road_threshold_spin = QDoubleSpinBox()
        self.road_threshold_spin.setRange(0, 2000)
        self.road_threshold_spin.setValue(800)
        layout.addRow('道路均值门限(Mbps):', self.road_threshold_spin)

        return tab

    def create_business_rules_tab(self) -> QWidget:
        """创建业务规则选项卡"""
        tab = QWidget()
        layout = QVBoxLayout(tab)

        # 规则表格
        self.rules_table = QTableWidget()
        self.rules_table.setColumnCount(8)
        self.rules_table.setHorizontalHeaderLabels([
            '业务名称', '方向', '计算秒数', '最小速率',
            '流量Min', '流量Max', '中位流量', '启用'
        ])
        self.rules_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.rules_table)

        # 操作按钮
        btn_layout = QHBoxLayout()
        add_rule_btn = QPushButton('添加规则')
        add_rule_btn.clicked.connect(self.add_business_rule)
        btn_layout.addWidget(add_rule_btn)

        delete_rule_btn = QPushButton('删除规则')
        delete_rule_btn.clicked.connect(self.delete_business_rule)
        btn_layout.addWidget(delete_rule_btn)

        layout.addLayout(btn_layout)

        return tab

    def create_right_panel(self) -> QWidget:
        """创建右侧结果展示面板"""
        panel = QWidget()
        layout = QVBoxLayout(panel)

        # 选项卡
        tab_widget = QTabWidget()

        # 统计结果选项卡
        stats_tab = self.create_statistics_tab()
        tab_widget.addTab(stats_tab, '统计结果')

        # 速率详表选项卡
        details_tab = self.create_details_tab()
        tab_widget.addTab(details_tab, '速率详表')

        # 原表整理选项卡
        cleaned_tab = self.create_cleaned_tab()
        tab_widget.addTab(cleaned_tab, '原表整理')

        layout.addWidget(tab_widget)

        return panel

    def create_statistics_tab(self) -> QWidget:
        """创建统计结果选项卡"""
        tab = QWidget()
        layout = QVBoxLayout(tab)

        self.stats_table = QTableWidget()
        self.stats_table.setColumnCount(9)
        self.stats_table.setHorizontalHeaderLabels([
            '业务类型', '方向', '测试次数', '采样点数', '均值速率',
            '峰值速率', '削峰后均值', '达标占比(%)', 'TOP10%峰值'
        ])
        self.stats_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.stats_table)

        return tab

    def create_details_tab(self) -> QWidget:
        """创建速率详表选项卡"""
        tab = QWidget()
        layout = QVBoxLayout(tab)

        self.details_table = QTableWidget()
        layout.addWidget(self.details_table)

        return tab

    def create_cleaned_tab(self) -> QWidget:
        """创建原表整理选项卡"""
        tab = QWidget()
        layout = QVBoxLayout(tab)

        self.cleaned_table = QTableWidget()
        layout.addWidget(self.cleaned_table)

        return tab

    def load_config_to_ui(self):
        """加载配置到UI"""
        # 全局参数
        params = self.config.get('global_params', {})
        self.qci_spin.setValue(params.get('qci', 9))
        self.dl_peak_spin.setValue(params.get('dl_peak_limit', 1000))
        self.ul_peak_spin.setValue(params.get('ul_peak_limit', 200))
        self.dl_pass_spin.setValue(params.get('dl_pass_threshold', 100))
        self.ul_pass_spin.setValue(params.get('ul_pass_threshold', 20))
        self.scenario_combo.setCurrentText(params.get('scenario_type', '场景'))
        self.road_threshold_spin.setValue(params.get('road_mean_threshold', 800))

        # 业务规则
        rules = self.config.get('business_rules', [])
        self.rules_table.setRowCount(len(rules))

        for row, rule in enumerate(rules):
            self.rules_table.setItem(row, 0, QTableWidgetItem(rule.get('name', '')))
            self.rules_table.setItem(row, 1, QTableWidgetItem(rule.get('direction', '')))
            self.rules_table.setItem(row, 2, QTableWidgetItem(str(rule.get('calc_seconds', ''))))
            self.rules_table.setItem(row, 3, QTableWidgetItem(str(rule.get('min_rate', ''))))
            self.rules_table.setItem(row, 4, QTableWidgetItem(str(rule.get('flow_min', ''))))
            self.rules_table.setItem(row, 5, QTableWidgetItem(str(rule.get('flow_max', ''))))
            self.rules_table.setItem(row, 6, QTableWidgetItem(str(rule.get('median_flow', ''))))

            enabled_item = QTableWidgetItem()
            enabled_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            enabled_item.setCheckState(Qt.Checked if rule.get('enabled', True) else Qt.Unchecked)
            self.rules_table.setItem(row, 7, enabled_item)

    def select_input_file(self):
        """选择输入文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, '选择输入文件', '', 'Excel Files (*.xlsx *.xls)'
        )
        if file_path:
            self.input_files = [file_path]
            self.file_path_edit.setText(file_path)
            self.log_message(f"已选择文件: {file_path}")

    def get_params_from_ui(self) -> Dict:
        """从UI获取参数"""
        return {
            'qci': self.qci_spin.value(),
            'dl_peak_limit': self.dl_peak_spin.value(),
            'ul_peak_limit': self.ul_peak_spin.value(),
            'dl_pass_threshold': self.dl_pass_spin.value(),
            'ul_pass_threshold': self.ul_pass_spin.value(),
            'scenario_type': self.scenario_combo.currentText(),
            'road_mean_threshold': self.road_threshold_spin.value()
        }

    def get_rules_from_ui(self) -> List[Dict]:
        """从UI获取业务规则"""
        def safe_float(text: str) -> Optional[float]:
            """安全地将字符串转换为float，空字符串或'None'返回None"""
            if not text or text.strip() == '' or text.strip() == 'None':
                return None
            try:
                return float(text)
            except ValueError:
                return None

        rules = []
        for row in range(self.rules_table.rowCount()):
            enabled = self.rules_table.item(row, 7).checkState() == Qt.Checked
            rule = {
                'id': row + 1,
                'name': self.rules_table.item(row, 0).text(),
                'direction': self.rules_table.item(row, 1).text(),
                'calc_seconds': safe_float(self.rules_table.item(row, 2).text()) or 0,
                'min_rate': safe_float(self.rules_table.item(row, 3).text()) or 0,
                'flow_min': safe_float(self.rules_table.item(row, 4).text()),
                'flow_max': safe_float(self.rules_table.item(row, 5).text()),
                'median_flow': safe_float(self.rules_table.item(row, 6).text()),
                'enabled': enabled
            }
            rules.append(rule)
        return rules

    @Slot()
    def start_processing(self):
        """开始处理"""
        print("DEBUG: start_processing called")
        print(f"DEBUG: input_files = {self.input_files}")

        if not self.input_files:
            QMessageBox.warning(self, '警告', '请先选择输入文件')
            return

        file_path = self.input_files[0]
        params = self.get_params_from_ui()
        rules = self.get_rules_from_ui()

        print(f"DEBUG: file_path = {file_path}")
        print(f"DEBUG: params = {params}")
        print(f"DEBUG: rules count = {len(rules)}")

        self.log_message("=" * 50)
        self.log_message("开始处理...")
        self.log_message(f"文件: {os.path.basename(file_path)}")
        self.log_message(f"QCI: {params['qci']}")
        self.log_message(f"下行削峰: {params['dl_peak_limit']} Mbps")
        self.log_message(f"上行削峰: {params['ul_peak_limit']} Mbps")
        self.progress_bar.setValue(0)

        # 禁用按钮防止重复点击
        self.sender().setEnabled(False)
        self.sender().setText("处理中...")

        # 创建处理器
        self.processor = DataProcessor(self.config)

        # 创建工作线程
        self.processing_thread = ProcessingThread(self.processor, file_path, params, rules)
        self.processing_thread.progress_signal.connect(self.update_progress)
        self.processing_thread.finished_signal.connect(self.on_processing_finished)
        print("DEBUG: Starting thread...")
        self.processing_thread.start()

    @Slot(str)
    def update_progress(self, message: str):
        """更新进度"""
        self.log_message(message)
        # 简单的进度模拟
        current = self.progress_bar.value()
        if current < 90:
            self.progress_bar.setValue(min(current + 10, 90))

    @Slot(bool, str, object, object, object)
    def on_processing_finished(self, success: bool, message: str, cleaned: pd.DataFrame,
                               rate_details: pd.DataFrame, statistics: Dict):
        """处理完成回调"""
        print(f"DEBUG: on_processing_finished called, success={success}")
        self.progress_bar.setValue(100)
        self.log_message(message)
        self.log_message("=" * 50)

        # 重新启用按钮
        for btn in self.findChildren(QPushButton):
            if btn.text() in ["处理中...", "开始统计"]:
                btn.setEnabled(True)
                btn.setText("开始统计")

        if success:
            # 显示统计结果
            self.display_statistics(statistics)
            self.display_rate_details(rate_details)
            self.display_cleaned_data(cleaned)

            # 保存结果供导出使用
            self.last_cleaned = cleaned
            self.last_rate_details = rate_details
            self.last_statistics = statistics

            self.log_message(f"✓ 统计完成！共{len(statistics)}个业务类型")
            QMessageBox.information(self, '完成', '数据处理完成！')
        else:
            self.log_message(f"✗ 处理失败: {message}")
            QMessageBox.critical(self, '错误', f'处理失败: {message}')

    def display_statistics(self, statistics: Dict):
        """显示统计结果"""
        self.stats_table.setRowCount(len(statistics))

        for row, (business_type, stats) in enumerate(statistics.items()):
            self.stats_table.setItem(row, 0, QTableWidgetItem(business_type))
            self.stats_table.setItem(row, 1, QTableWidgetItem(str(stats.get('方向', ''))))
            self.stats_table.setItem(row, 2, QTableWidgetItem(str(stats.get('测试次数', 0))))
            self.stats_table.setItem(row, 3, QTableWidgetItem(str(stats.get('采样点数', 0))))
            self.stats_table.setItem(row, 4, QTableWidgetItem(f"{stats.get('均值速率', 0):.2f}"))
            self.stats_table.setItem(row, 5, QTableWidgetItem(f"{stats.get('峰值速率', 0):.2f}"))
            self.stats_table.setItem(row, 6, QTableWidgetItem(f"{stats.get('削峰后均值速率', 0):.2f}"))
            self.stats_table.setItem(row, 7, QTableWidgetItem(f"{stats.get('达标占比', 0):.2f}"))
            self.stats_table.setItem(row, 8, QTableWidgetItem(f"{stats.get('TOP10峰值', 0):.2f}"))

    def display_rate_details(self, rate_details: pd.DataFrame):
        """显示速率详表"""
        if rate_details is None or rate_details.empty:
            return

        # 选择关键列显示
        key_columns = ['序号', '采集时间', '来源', '业务类型', '下行RLC吞吐率', '上行RLC吞吐率']
        available_columns = [col for col in key_columns if col in rate_details.columns]

        self.details_table.setColumnCount(len(available_columns))
        self.details_table.setHorizontalHeaderLabels(available_columns)
        self.details_table.setRowCount(min(len(rate_details), 1000))  # 限制显示行数

        for row in range(min(len(rate_details), 1000)):
            for col, col_name in enumerate(available_columns):
                value = rate_details.iloc[row][col_name]
                if pd.isna(value):
                    value = ''
                self.details_table.setItem(row, col, QTableWidgetItem(str(value)))

    def display_cleaned_data(self, cleaned_data: pd.DataFrame):
        """显示原表整理"""
        if cleaned_data is None or cleaned_data.empty:
            return

        self.cleaned_table.setColumnCount(len(cleaned_data.columns))
        self.cleaned_table.setHorizontalHeaderLabels(cleaned_data.columns)
        self.cleaned_table.setRowCount(min(len(cleaned_data), 100))

        for row in range(min(len(cleaned_data), 100)):
            for col, col_name in enumerate(cleaned_data.columns):
                value = cleaned_data.iloc[row][col_name]
                if pd.isna(value):
                    value = ''
                self.cleaned_table.setItem(row, col, QTableWidgetItem(str(value)))

    def export_excel(self):
        """导出Excel"""
        if not hasattr(self, 'last_statistics') or self.last_statistics is None:
            QMessageBox.warning(self, '警告', '请先处理数据')
            return

        output_path, _ = QFileDialog.getSaveFileName(
            self, '保存Excel文件', '', 'Excel Files (*.xlsx)'
        )

        if output_path:
            try:
                exporter = ExcelExporter()
                exporter.export(output_path, self.last_cleaned, self.last_rate_details, self.last_statistics)
                self.log_message(f"Excel已导出: {output_path}")
                QMessageBox.information(self, '成功', 'Excel文件导出成功！')
            except Exception as e:
                QMessageBox.critical(self, '错误', f'导出失败: {str(e)}')

    def add_business_rule(self):
        """添加业务规则"""
        row = self.rules_table.rowCount()
        self.rules_table.insertRow(row)

        for col in range(8):
            self.rules_table.setItem(row, col, QTableWidgetItem(''))

        enabled_item = QTableWidgetItem()
        enabled_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
        enabled_item.setCheckState(Qt.Checked)
        self.rules_table.setItem(row, 7, enabled_item)

    def delete_business_rule(self):
        """删除业务规则"""
        current_row = self.rules_table.currentRow()
        if current_row >= 0:
            self.rules_table.removeRow(current_row)

    def log_message(self, message: str):
        """输出日志消息"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.log_text.append(f"[{timestamp}] {message}")


# ==================== 主程序入口 ====================

def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')

    window = MainWindow()
    window.show()

    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
