#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
核心分析：参考文件中26个store_l对应的时间窗口内，每个store_l真正的MAC总flow是多少？
如果总flow经常<1500MB，那分类阈值需要降低。
"""

import sys, os
import numpy as np
import pandas as pd
import openpyxl
from importlib import util as import_util

sys.path.insert(0, '/Users/sun/ClaudeCode/先甲工信部工具')
spec = import_util.spec_from_file_location("tool",
    "/Users/sun/ClaudeCode/先甲工信部工具/5G用户级公共监控速率统计工具-V1.04.py")
tool = import_util.module_from_spec(spec)
spec.loader.exec_module(tool)

uproc = tool.UCMProcessor({'dl_peak_limit': 900, 'ul_peak_limit': 160})
paths = [
    '/Users/sun/ClaudeCode/先甲工信部工具/联通/mmf20260703115328-电信.xlsx',
    '/Users/sun/ClaudeCode/先甲工信部工具/联通/mmf20260703115334-电信.xlsx',
    '/Users/sun/ClaudeCode/先甲工信部工具/联通/mmf20260703115336-联通.xlsx'
]
params = {'dl_peak_limit': 900, 'ul_peak_limit': 160, 'down_min': 50, 'up_min': 10, 'match_mode': 'auto'}
agg = uproc.parse_mmf(paths, params)

# 从参考文件中提取每个store_l的时间窗口
ref = '/Users/sun/ClaudeCode/先甲工信部工具/联通/呼叫详情_按时间排序-20260707.xlsx'
wb = openpyxl.load_workbook(ref, data_only=True)
ws = wb['联通']

store_l_windows = []
for r in range(2, ws.max_row + 1):
    biz = str(ws.cell(r, 2).value or '')
    if '应用商店' not in biz:
        continue
    st = ws.cell(r, 3).value
    et = ws.cell(r, 4).value
    big = ws.cell(r, 11).value  # 大文件下载速率，只有store_l行才有
    store_l_windows.append({'row': r, 'start': st, 'end': et, 'big_rate': big})

# 每组3行：总行(store_s开始-大包结束) + store_s行 + store_l行
# store_l行：开始时间是大文件实际开始，完成时间是大文件结束，big_rate=None
# 判断store_l行的特征：大文件下载速率为空，且名称是大文件
# 分组：每3行为一组
groups = []
for g in range(0, len(store_l_windows), 3):
    rows = store_l_windows[g:g+3]
    if len(rows) < 3:
        continue
    # 第3行是store_l（大文件行）
    sl = rows[2]
    groups.append(sl)

print(f"store_l 时间窗口数: {len(groups)}")
flow_vals = []
for g in groups:
    st = pd.to_datetime(g['start'])
    et = pd.to_datetime(g['end'])
    mask = (agg['t'] >= st) & (agg['t'] <= et)
    sub = agg[mask]
    dl_flow = sub['dl_mac'].sum() / 8
    dur = len(sub)
    avg_rate = sub['dl_rlc'].replace(0, np.nan).mean()
    flow_vals.append(dl_flow)
    print(f"  {st.strftime('%H:%M:%S')}-{et.strftime('%H:%M:%S')}: flow={dl_flow:.0f}MB dur={dur}s avg_rate={avg_rate:.1f}")

flow_vals = np.array(flow_vals)
print(f"\n统计:")
print(f"  min={flow_vals.min():.0f}MB max={flow_vals.max():.0f}MB median={np.median(flow_vals):.0f}MB mean={flow_vals.mean():.0f}MB")
print(f"  >=1500MB: {np.sum(flow_vals >= 1500)}个")
print(f"  >=1200MB: {np.sum(flow_vals >= 1200)}个")
print(f"  >=1000MB: {np.sum(flow_vals >= 1000)}个")
print(f"  >=900MB: {np.sum(flow_vals >= 900)}个")
print(f"  >=800MB: {np.sum(flow_vals >= 800)}个")
print(f"  >=700MB: {np.sum(flow_vals >= 700)}个")
print(f"  >=600MB: {np.sum(flow_vals >= 600)}个")
print(f"  <700MB: {np.sum(flow_vals < 700)}个")
