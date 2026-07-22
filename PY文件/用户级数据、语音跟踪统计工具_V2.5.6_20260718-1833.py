#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
用户级数据、语音跟踪统计工具 V2.5.7
GUI: 双标签(速率统计 / 语音VQI) → 选择文件(MMF+可选基准) → 处理 → 内嵌查看结果
CLI: python3 this.py --cmd
"""
import sys, os, re
import pandas as pd, numpy as np
from datetime import datetime, timedelta, time
from collections import Counter, defaultdict

# ===== 通用文件读取函数（支持xlsx/xls/csv，优先calamine，csv用read_csv）=====
def _read_file(f, nrows=None):
    """读取xlsx/xls/csv/mmf/tmf文件，返回DataFrame
    nrows: 可选，只读取前n行（用于快速判断文件类型）
    """
    low = f.lower()
    if low.endswith('.csv'):
        # 尝试多种编码读取csv
        for enc in ['gbk', 'utf-8', 'latin-1']:
            try:
                return pd.read_csv(f, encoding=enc, nrows=nrows)
            except Exception:
                continue
        return pd.read_csv(f, encoding='latin-1', on_bad_lines='skip', nrows=nrows)
    try:
        return pd.read_excel(f, engine='calamine', nrows=nrows)
    except Exception:
        try:
            return pd.read_excel(f, nrows=nrows)
        except Exception:
            try:
                return pd.read_excel(f, engine='openpyxl', nrows=nrows)
            except Exception:
                return pd.read_excel(f, engine='xlrd')

def _has_message_type_col(df):
    """判断DataFrame是否有MessageType列（支持多种列名）"""
    for col in ['MessageType', 'Message Type', 'message_type', '消息类型']:
        if col in df.columns:
            return True
    # 如果第二列是Source且有Time列，可能是mmf数据文件
    if len(df.columns) >= 2 and 'Time' in df.columns and 'Source' in df.columns:
        return False  # mmf数据文件
    return False

# ===== 预编译正则（性能优化）=====
_PARSE_RE = re.compile(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})[\.\(](\d+)')
_PARSE_MS_RE = re.compile(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\.(\d+)')

def parse_time(t):
    m = _PARSE_RE.match(str(t))
    if m:
        return datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S') + timedelta(milliseconds=int(m.group(2)))
    return None

# ===== 计算批注辅助函数 =====
def _add_calc_comment(cell, formula_desc, calc_process, source_col, numerator_vals=None, denominator_vals=None, numerator_desc='分子', denominator_desc='分母'):
    """为单元格添加计算批注，显示公式、分子分母值"""
    if numerator_vals is None or len(numerator_vals) == 0:
        vs_str = '[]'
    else:
        vs = [str(round(v, 2)) for v in numerator_vals]
        if len(vs) <= 20:
            vs_str = '/'.join(vs)
        else:
            vs_str = '/'.join(vs[:5]) + '/.../' + '/'.join(vs[-5:])
    
    if denominator_vals is None or len(denominator_vals) == 0:
        dv_str = '[]'
    else:
        dv = [str(round(v, 2)) for v in denominator_vals]
        if len(dv) <= 20:
            dv_str = '/'.join(dv)
        else:
            dv_str = '/'.join(dv[:5]) + '/.../' + '/'.join(dv[-5:])
    
    txt = (f"公式: {formula_desc}\n"
           f"计算过程: {calc_process}\n\n"
           f"{numerator_desc}: {vs_str}\n"
           f"{denominator_desc}: {dv_str}\n\n"
           f"数据来源: {source_col} 列")
    from openpyxl.comments import Comment
    c = Comment(txt, "Tool")
    c.width = 500
    c.height = 250
    cell.comment = c

# ===== 阈值 =====
DL_ACTIVE_BIG, UL_ACTIVE_BIG = 50.0, 10.0
DL_ACTIVE_SMALL, UL_ACTIVE_SMALL = 5.0, 5.0
GAP_MERGE = 3.0
# FTP时长阈值（默认20秒，由ftp_duration参数动态调整）
FTP_PAIR_GAP = 15.0
STORE_L_DL_MIN, STORE_L_DL_MAX = 8800, 10100
WX_L_DUR_MIN, WX_L_DUR_MAX, WX_L_UL_MIN = 15, 60, 800
DL_BIZ = {'FTP下载', '应用商店小文件下载', '应用商店大文件下载'}
UL_BIZ = {'FTP上传', '微信小包发送', '微信大包发送'}
REQ_BIZ = {'FTP下载', 'FTP上传', '应用商店小文件下载', '应用商店大文件下载', '微信小包发送', '微信大包发送'}


def detect_segments(ts, dl, ul, dl_th, ul_th, gap, r1=None, r2=None):
    active = (dl > dl_th) | (ul > ul_th); segs = []; i = 0; n = len(ts)
    while i < n:
        if r1 is not None and ts[i] < r1: i += 1; continue
        if r2 is not None and ts[i] > r2: break
        if active[i]:
            si = i; j = i + 1
            while j < n:
                if r2 is not None and ts[j] > r2: break
                if not active[j]:
                    k = j
                    while k < n and not active[k]:
                        if (ts[k] - ts[j - 1]) / np.timedelta64(1, 's') > gap: break
                        k += 1
                    if k < n and active[k] and (r2 is None or ts[k] <= r2): j = k; continue
                    else: break
                j += 1
            st = ts[si]; et = ts[j - 1]; dur = (et - st) / np.timedelta64(1, 's') + 1
            segs.append({'st': st, 'et': et, 'dur': dur, 'dl': float(dl[si:j].sum()), 'ul': float(ul[si:j].sum())})
            i = j
        else: i += 1
    return segs


def process(files, qci_list=None, dl_clip=1000, ul_clip=200, callback=None, cancel_check=None, progress_cb=None, base_file=None, time_filter=None, ftp_duration=20, merge_raw=True):
    if qci_list is None: qci_list = [5, 6, 7]

    # 根据ftp_duration动态调整FTP时长阈值
    # 用户选N秒 → (N-1.5, N+3) 的左右范围
    ftp_dur_min = ftp_duration - 1.5
    ftp_dur_max = ftp_duration + 3.0

    """统一的处理函数：从MMF文件生成5Sheet输出"""
    def log(msg, pct=None):
        if cancel_check and cancel_check():
            raise KeyboardInterrupt('用户取消')
        import time as _time
        _t = _time.time()
        if not hasattr(log, 'last_time'):
            log.last_time = _t
        elapsed = _t - log.last_time if pct is not None else 0
        prefix = f"({elapsed:.1f}秒) " if pct is not None and elapsed > 0 else ""
        print(msg)
        if callback: callback(f"{prefix}{msg}")
        if pct is not None and progress_cb: progress_cb(pct)
        log.last_time = _t

    # 展开压缩文件(zip 自动解压，支持嵌套 zip-in-zip 递归解压)
    import zipfile, tempfile, shutil, glob
    _tmp_dirs = []
    def _expand_zip(zip_path):
        """递归解压 zip(含嵌套 zip)，返回所有 xlsx/xls 路径"""
        _tmp = tempfile.mkdtemp(prefix='mmf_unzip_'); _tmp_dirs.append(_tmp)
        _xs = []
        try:
            with zipfile.ZipFile(zip_path) as _zf: _zf.extractall(_tmp)
            _xs = glob.glob(os.path.join(_tmp, '**', '*.xlsx'), recursive=True) + \
                  glob.glob(os.path.join(_tmp, '**', '*.xls'), recursive=True)
            # 内嵌 zip 递归解压
            _inner_zips = glob.glob(os.path.join(_tmp, '**', '*.zip'), recursive=True)
            for _iz in _inner_zips:
                _xs.extend(_expand_zip(_iz))
        except Exception as _e:
            log(f"  解压失败 {os.path.basename(zip_path)}: {_e}")
        return _xs

    _expanded = []
    for _f in (files or []):
        if str(_f).lower().endswith('.zip'):
            _xs = _expand_zip(_f)
            log(f"  解压 {os.path.basename(_f)}: 找到 {len(_xs)} 个Excel(含嵌套)", 5)
            _expanded.extend(_xs)
        else:
            _expanded.append(_f)
    files = _expanded
    nf = len(files) if files else 1
    log(f"加载MMF文件（共 {nf} 个）...", 5)
    dfs = []
    for fi, f in enumerate(files):
        log(f"  读取文件 {fi+1}/{nf}: {os.path.basename(f)}", 5 + int(15 * (fi+1) / nf))
        # 读全部列（详细过程需展示mmf所有列）；优先用通用函数
        df_i = _read_file(f)
        rlc_dl = [c for c in df_i.columns if 'RLC' in c and 'Downlink' in c]
        rlc_ul = [c for c in df_i.columns if 'RLC' in c and 'Uplink' in c]
        df_i['_t'] = df_i['Time'].apply(parse_time)
        if rlc_dl:
            df_i['_dl'] = pd.to_numeric(df_i[rlc_dl[0]], errors='coerce').fillna(0) / 1e6
        else:
            df_i['_dl'] = 0
        if rlc_ul:
            df_i['_ul'] = pd.to_numeric(df_i[rlc_ul[0]], errors='coerce').fillna(0) / 1e6
        else:
            df_i['_ul'] = 0
        dfs.append(df_i)
    df = pd.concat(dfs, ignore_index=True)
    log("  合并完成，按时间排序...", 22)
    # 按 C列(Time/_t) 升序排列，保证详细过程表按时间顺序
    df = df.sort_values('_t', kind='stable').reset_index(drop=True)

    # ===== 时间段过滤 =====
    if time_filter:
        use_time, t_start, t_end = time_filter
        if use_time and t_start is not None and t_end is not None:
            # 提取时间部分（忽略日期），只保留 HH:MM:SS
            t_start_t = t_start.toPython() if hasattr(t_start, 'toPython') else t_start
            t_end_t = t_end.toPython() if hasattr(t_end, 'toPython') else t_end
            # 将 _t 的时间部分提取出来做比较
            df['_time_only'] = df['_t'].dt.time
            mask = (df['_time_only'] >= t_start_t) & (df['_time_only'] <= t_end_t)
            df = df[mask].copy()
            df = df.drop('_time_only', axis=1)
            log(f"  时间段过滤: {t_start_t} ~ {t_end_t}, 剩余 {len(df)} 行", 22)

    log("QCI过滤+秒聚合...", 28)
    df7 = df[df['QCI'].isin(qci_list)].dropna(subset=['_t']).copy()
    if len(df7) == 0:
        # 检查是否有其他QCI的数据
        available_qci = df['QCI'].dropna().unique()
        avail_str = ', '.join([str(int(q)) for q in available_qci if pd.notna(q) and q > 0])
        if avail_str:
            log(f"警告: 数据中无QCI={qci_list}的记录，可用QCI: {avail_str}", 28)
            log(f"建议: 重新选择QCI后再次运行", 28)
            # 生成空输出文件
            version = 'V2.5.7'
            timestamp = datetime.now().strftime('%Y%m%d-%H%M')
            out = f'联通/电信最终输出_{version}_{timestamp}.xlsx'
            with pd.ExcelWriter(out, engine='openpyxl') as w:
                pd.DataFrame({'提示': [f'数据中无QCI={qci_list}的记录，可用QCI: {avail_str}']}).to_excel(w, sheet_name='提示', index=False)
            log(f"完成(空): {out}", 100)
            return out
        else:
            raise ValueError("数据中无有效的QCI记录")
    df7['_s'] = df7['_t'].dt.floor('s')
    sec = df7.groupby('_s').agg({'_dl': 'max', '_ul': 'max'}).sort_index().reset_index()
    dl_v, ul_v, ts_v = sec['_dl'].fillna(0).values, sec['_ul'].fillna(0).values, sec['_s'].values

    log("业务识别...", 35)
    big = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_BIG, UL_ACTIVE_BIG, GAP_MERGE)

    def cb(s):
        d, dl, ul = s['dur'], s['dl'], s['ul']
        if ftp_dur_min <= d <= ftp_dur_max: return 'FTP候选'
        if dl > ul and STORE_L_DL_MIN <= dl <= STORE_L_DL_MAX and d > 6: return '应用商店大文件下载'
        if WX_L_DUR_MIN <= d <= WX_L_DUR_MAX and ul > dl and ul >= WX_L_UL_MIN: return '微信大包发送'
    for s in big: s['biz'] = cb(s); s['rnd'] = None

    log("FTP配对+小业务识别...", 40)
    fc = [s for s in big if s['biz'] == 'FTP候选']; paired = set(); fp = []
    for i in range(len(fc)):
        if id(fc[i]) in paired: continue
        for j in range(i + 1, len(fc)):
            if id(fc[j]) in paired: continue
            g = (fc[j]['st'] - fc[i]['et']) / np.timedelta64(1, 's')
            if 0 <= g < FTP_PAIR_GAP:
                fp.append((fc[i], fc[j])); paired.add(id(fc[i])); paired.add(id(fc[j])); break
            elif g >= FTP_PAIR_GAP: break

    def fd(s): return 'FTP下载' if s['dl'] > s['ul'] else 'FTP上传'
    for s in fc:
        if id(s) not in paired:
            if STORE_L_DL_MIN <= s['dl'] <= STORE_L_DL_MAX and s['dl'] > s['ul']: s['biz'] = '应用商店大文件下载'
            else: s['biz'] = fd(s)

    al = sorted(big, key=lambda x: x['st'])
    for s in al: s['业务'] = s['biz']
    rn = 0; pi = []
    for f1, f2 in fp:
        rn += 1; f1['rnd'] = rn; f1['业务'] = fd(f1); f2['rnd'] = rn; f2['业务'] = fd(f2)
        pi.append((rn, f2['et'], f1['st']))

    # ===== 无FTP配对时：直接使用大业务作为轮次起点 =====
    # 若没有FTP配对，则按时间顺序给每个大业务编号轮次
    if not fp:
        log("无FTP配对，使用大业务作为轮次起点...", 42)
        large_bizs = [s for s in al if s['业务'] in ('应用商店大文件下载', '微信大包发送', 'FTP下载', 'FTP上传')]
        for i, s in enumerate(large_bizs):
            rn += 1
            s['rnd'] = rn
            pi.append((rn, s['et'], s['st']))
        log(f"  识别到 {rn} 个大业务作为轮次", 43)
    for idx, (rnd, pe, _) in enumerate(pi):
        ns = pi[idx + 1][2] if idx + 1 < len(pi) else (al[-1]['et'] + np.timedelta64(1, 's'))
        for s in al:
            if s['et'] > pe and s['st'] < ns and s['rnd'] is None and s['业务'] in ('应用商店大文件下载', '微信大包发送'):
                s['rnd'] = rnd

    tb = sorted([s for s in al if s['业务'] in ('FTP下载', 'FTP上传', '应用商店大文件下载', '微信大包发送')],
                key=lambda x: x['st'])
    P2 = {'应用商店大文件下载': '应用商店小文件下载', '微信大包发送': '微信小包发送'}
    small = []
    for idx, (rnd, pe, ps) in enumerate(pi):
        we = pi[idx + 1][2] if idx + 1 < len(pi) else (tb[-1]['et'] + np.timedelta64(1, 's'))
        bw = sorted([s for s in tb if s['st'] >= pe and s['st'] < we and s['业务'] in (
        '应用商店大文件下载', '微信大包发送')], key=lambda x: x['st'])
        prev = pe
        for bs in bw:
            st2 = P2.get(bs['业务'])
            if st2 and prev < bs['st'] and (bs['st'] - prev) / np.timedelta64(1, 's') > 1:
                gs = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_SMALL, UL_ACTIVE_SMALL, GAP_MERGE, prev, bs['st'])
                vd = [s for s in gs if s['dl'] + s['ul'] >= 10 and s['dur'] >= 0.5]
                if st2 == '应用商店小文件下载':
                    dls = [s for s in vd if s['dl'] > s['ul']]
                    if dls: bm = max(dls, key=lambda x: x['dl']); bm['业务'] = st2; bm['rnd'] = rnd; small.append(bm)
                else:
                    uls = [s for s in vd if s['ul'] > s['dl']]
                    if uls: bm = max(uls, key=lambda x: x['ul']); bm['业务'] = st2; bm['rnd'] = rnd; small.append(bm)
            prev = bs['et']

    # 检测第一个FTP对之前的不完整轮次
    if pi:
        first_ftp_start = pi[0][2]
        data_start = ts_v[0]
        if first_ftp_start > data_start + np.timedelta64(30, 's'):
            # 找第一个FTP对之前的大业务(应用商店大/微信大)
            pre_big = sorted([s for s in al if s['st'] < first_ftp_start and s['业务'] in (
                '应用商店大文件下载', '微信大包发送')], key=lambda x: x['st'])
            rng_start = data_start
            if pre_big:
                for bs in pre_big:
                    st2 = P2.get(bs['业务'])
                    if st2 and (bs['st'] - rng_start) / np.timedelta64(1, 's') > 1:
                        gs = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_SMALL, UL_ACTIVE_SMALL,
                                             GAP_MERGE, rng_start, bs['st'])
                        vd = [s for s in gs if s['dl'] + s['ul'] >= 10 and s['dur'] >= 0.5]
                        if st2 == '应用商店小文件下载':
                            dls = [s for s in vd if s['dl'] > s['ul']]
                            if dls: bm = max(dls, key=lambda x: x['dl']); bm['业务'] = st2; bm['rnd'] = 0; small.append(bm)
                        else:
                            uls = [s for s in vd if s['ul'] > s['dl']]
                            if uls: bm = max(uls, key=lambda x: x['ul']); bm['业务'] = st2; bm['rnd'] = 0; small.append(bm)
                    rng_start = bs['et']
            # 最后一个大业务(或数据开头) 到 first_ftp_start 之间
            if first_ftp_start > rng_start + np.timedelta64(1, 's'):
                gs = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_SMALL, UL_ACTIVE_SMALL,
                                     GAP_MERGE, rng_start, first_ftp_start)
                vd = [s for s in gs if s['dl'] + s['ul'] >= 10 and s['dur'] >= 0.5]
                dls = [s for s in vd if s['dl'] > s['ul']]
                if dls: bm = max(dls, key=lambda x: x['dl']); bm['业务'] = '应用商店小文件下载'; bm['rnd'] = 0; small.append(bm)
                uls = [s for s in vd if s['ul'] > s['dl']]
                if uls: bm = max(uls, key=lambda x: x['ul']); bm['业务'] = '微信小包发送'; bm['rnd'] = 0; small.append(bm)

    af = sorted(al + small, key=lambda x: x['st'])
    for s in af: s['有效'] = s['业务'] is not None and s.get('rnd') is not None
    rm = defaultdict(list)
    for s in af:
        if s['有效'] and s['业务'] and s.get('rnd') is not None: rm[s['rnd']].append(s)
    for rnd, segs in rm.items():
        seen = set(); cnt = 0
        for s in sorted(segs, key=lambda x: x['st']):
            if s['业务'] in seen or cnt >= 6: s['有效'] = False
            else: seen.add(s['业务']); cnt += 1

    valid = [s for s in af if s['有效'] and s['业务']]
    log(f"识别完成: {len(valid)}个业务段, {rn}轮")

    # ===== 标记 =====
    df = pd.concat([df, pd.DataFrame({
        '下载速率': df['_dl'], '上传速率': df['_ul'],
        '业务识别': None, '持续时长': None, '速率(Mbps)': None,
    }, index=df.index)], axis=1)

    for seg in valid:
        st, et, biz = seg['st'], seg['et'], seg['业务']
        sm = (df['_t'] >= st) & (df['_t'] <= et + timedelta(seconds=1)) & (df['QCI'].isin(qci_list))
        col = '_dl' if biz in DL_BIZ else '_ul'

        # 所有业务：只标记有流量的行(>0)
        cv_all = df.loc[sm, col].fillna(0)
        nz = cv_all > 0
        mm = sm & nz.reindex(df.index, fill_value=False)

        if not mm.any():
            continue  # 无流量则跳过

        # FTP：从第一个有流量的行开始，检查前后小流量补齐到10秒
        if biz in ('FTP下载', 'FTP上传'):
            # 找所有有流量的行
            flow_vals = df.loc[mm, col].fillna(0).tolist()
            flow_times = df.loc[mm, '_t'].tolist()

            if len(flow_vals) == 0:
                continue

            # 起点：从第一个有流量的行开始
            first_idx = 0

            # 终点：最后一个有流量的行
            last_idx = len(flow_vals) - 1

            # 计算时长
            first_t = flow_times[first_idx]
            last_t = flow_times[last_idx]
            md = (last_t - first_t).total_seconds()

            # 如果时长不是10秒左右，检查前后是否有小流量可以补齐
            # 向前找：在segment开始时间之前3秒内，找有流量的行
            if md < 9.5 or md > 10.5:
                # 向前找小流量
                ft0 = st - timedelta(seconds=3)
                pm = (df['_t'] >= ft0) & (df['_t'] < first_t) & (df['QCI'].isin(qci_list))
                if pm.any():
                    pv = df.loc[pm, col].fillna(0)
                    pv_nz = pv > 0
                    if pv_nz.any():
                        # 有前置流量，扩展起点
                        first_t = df.loc[pv[pv_nz].index[0], '_t']
                        md = (last_t - first_t).total_seconds()

                # 向后找小流量
                et0 = et + timedelta(seconds=3)
                pm2 = (df['_t'] > last_t) & (df['_t'] <= et0) & (df['QCI'].isin(qci_list))
                if pm2.any():
                    pv2 = df.loc[pm2, col].fillna(0)
                    pv2_nz = pv2 > 0
                    if pv2_nz.any():
                        # 有后置流量，扩展终点
                        last_t = df.loc[pv2[pv2_nz].index[-1], '_t']
                        md = (last_t - first_t).total_seconds()

            # 重新计算标记范围
            mm_final = (df['_t'] >= first_t) & (df['_t'] <= last_t) & (df['QCI'].isin(qci_list)) & (df[col].fillna(0) > 0)
            md = (last_t - first_t).total_seconds()

            df.loc[mm_final, '业务识别'] = biz
            df.loc[mm_final, '持续时长'] = round(md, 3)
            vs = pd.to_numeric(df.loc[mm_final, '下载速率'] if biz in DL_BIZ else df.loc[mm_final, '上传速率'], errors='coerce').fillna(0).tolist()
            ar = sum(vs) / md if md > 0 else 0
            df.loc[mm_final, '速率(Mbps)'] = round(ar, 2)
            seg['mark_st'], seg['mark_et'], seg['mark_dur'] = first_t, last_t, md
            continue

        # 其他业务：直接标记有流量的行
        first_t = df.loc[mm, '_t'].min()
        last_t = df.loc[mm, '_t'].max()
        md = (last_t - first_t).total_seconds()

        df.loc[mm, '业务识别'] = biz
        df.loc[mm, '持续时长'] = round(md, 3)
        if biz in DL_BIZ: vs = pd.to_numeric(df.loc[mm, '下载速率'], errors='coerce').fillna(0).tolist()
        else: vs = pd.to_numeric(df.loc[mm, '上传速率'], errors='coerce').fillna(0).tolist()
        ar = sum(vs) / md if md > 0 else 0
        df.loc[mm, '速率(Mbps)'] = round(ar, 2)
        seg['mark_st'], seg['mark_et'], seg['mark_dur'] = first_t, last_t, md

    rd = defaultdict(set)
    for s in valid:
        r = s.get('rnd')
        if r is not None: rd[r].add(s['业务'])
    rl = {}; ii = 0
    for r in sorted(rd.keys()):
        if rd[r] >= REQ_BIZ: rl[r] = str(r)
        else: ii += 1; rl[r] = f"不完整{ii}"
    df['轮次'] = None
    for s in valid:
        r = s.get('rnd'); lb = rl.get(r) if r is not None else ''
        if lb:
            ms2, me2 = s.get('mark_st', s['st']), s.get('mark_et', s['et'])
            # 只在有业务标记的行填轮次
            mm2 = (df['_t'] >= ms2) & (df['_t'] <= me2) & (df['QCI'].isin(qci_list)) & df['业务识别'].notna()
            df.loc[mm2, '轮次'] = lb

    # ===== 详细过程 =====
    # 先创建筛选列（A列）
    # E-N列：速率(Mbps)、持续时长、轮次、businessType、基准_业务类型、基准_开始时间、基准_结束时间、基准_速率类指标、基准_数值、基准_时长类指标、基准_数值.1
    OC = ['下载速率', '上传速率', '速率(Mbps)', '持续时长', '轮次']
    log("生成详细过程...", 48)
    df_detail = pd.DataFrame({'筛选': None, 'id': range(1, len(df) + 1), 'Time': df['Time']})
    for c in OC: df_detail[c] = df[c].values if c in df.columns else None
    df_detail['businessType'] = df['业务识别'].values

    # 加载基准数据并标记到详细过程
    log("加载基准数据...", 53)
    has_base = True

    # 优先使用用户传入的基准文件
    if base_file and os.path.exists(base_file):
        try:
            # 尝试读取用户基准文件
            try:
                base = pd.read_excel(base_file, sheet_name='呼叫详情', engine='calamine')
            except:
                base = pd.read_excel(base_file, engine='calamine')

            if base is None or len(base) == 0:
                has_base = False
            else:
                log(f"使用用户基准文件: {os.path.basename(base_file)}", 53)
        except Exception as e:
            log(f"用户基准加载失败: {e}", 53)
            has_base = False
            base = pd.DataFrame()
    else:
        # 尝试加载默认基准文件
        try:
            base = pd.read_excel('工信部业务指标_呼叫详情_整理.xlsx', sheet_name='电信').sort_values('开始时间')
            if base is None or len(base) == 0:
                has_base = False
            else:
                # 检查基准日期与MMF数据日期是否匹配
                base_start = pd.to_datetime(base['开始时间'].iloc[0])
                data_start = df['_t'].dropna().min() if df['_t'].notna().any() else None
                if data_start is not None:
                    if abs((data_start - base_start).days) > 1:
                        log(f"基准日期({base_start.date()})与数据日期({data_start.date()})不匹配，跳过基准", 55)
                        has_base = False
                        base = pd.DataFrame()
        except Exception as e:
            log(f"基准加载失败: {e}", 53)
            has_base = False
            base = pd.DataFrame()
    log("基准对齐...", 55)
    # 初始化基准列
    base_cols = ['基准_业务类型', '基准_开始时间', '基准_结束时间', '基准_速率类指标', '基准_数值', '基准_时长类指标', '基准_数值.1']
    for bc in base_cols:
        df_detail[bc] = None

    # 对齐基准数据到详细过程（向量化：预取整列，避免逐行 df.loc，提速约50倍）
    ts = df['_t']
    ts_valid = ts[ts.notna()]
    for _, brow in base.iterrows():
        b_start = pd.to_datetime(brow['开始时间'])
        b_end = pd.to_datetime(brow['结束时间'])
        b_biz = str(brow['业务类型'])
        if len(ts_valid) == 0:
            break
        # 一次向量化求 b_start 最近行
        diff_s = (ts_valid - b_start).abs()
        best_start_idx = diff_s.idxmin()
        if diff_s.loc[best_start_idx].total_seconds() >= 5:
            continue
        # 从 best_start_idx 起向量化求 b_end 最近行
        sub = ts_valid[ts_valid.index >= best_start_idx]
        best_end_idx = sub.index[(sub - b_end).abs().argmin()] if len(sub) > 0 else best_start_idx
        # 批量切片赋值（替代逐行 loc）
        lo, hi = best_start_idx, min(best_end_idx, len(df_detail) - 1)
        if lo <= hi:
            df_detail.loc[lo:hi, '基准_业务类型'] = b_biz
            df_detail.loc[lo:hi, '基准_开始时间'] = str(brow['开始时间'])
            df_detail.loc[lo:hi, '基准_结束时间'] = str(brow['结束时间'])
            df_detail.loc[lo:hi, '基准_速率类指标'] = str(brow['速率类指标'])
            df_detail.loc[lo:hi, '基准_数值'] = brow['数值']
            df_detail.loc[lo:hi, '基准_时长类指标'] = str(brow['时长类指标'])
            df_detail.loc[lo:hi, '基准_数值.1'] = brow['数值.1']

    # 添加_t列
    df_detail['_t'] = df['_t'].values

    # 将QCI、Downlink RLC、Uplink RLC三列移到_t之后
    special_cols = ['QCI', 'Downlink RLC Throughput(bps)', 'Uplink RLC Throughput(bps)']
    for sc in special_cols:
        if sc in df.columns:
            df_detail[sc] = df[sc].values

    # 添加其他列
    for c in df.columns:
        if c not in df_detail.columns and c not in ('id', '_t', '_dl', '_ul', '_s', 'businessType', '业务识别'):
            df_detail[c] = df[c].values

    # 计算筛选列（向量化）：check_cols 任一非空则为1
    check_cols = ['速率(Mbps)', '持续时长', '轮次', 'businessType',
                  '基准_业务类型', '基准_开始时间', '基准_结束时间', '基准_速率类指标', '基准_数值', '基准_时长类指标', '基准_数值.1']
    has = pd.Series(False, index=df_detail.index)
    for col in check_cols:
        s = df_detail[col]
        ss = s.astype(str).str.strip()
        has = has | (s.notna() & (ss != '') & (ss.str.lower() != 'nan'))
    df_detail['筛选'] = has.astype(int).values

    # ===== 汇总 =====
    sr = []
    sum_raw_data = {}  # 用于批注：{biz: {rates, durs, col_name, r_start, r_end, sub}}
    for biz in sorted(DL_BIZ | UL_BIZ):
        sub = df_detail[df_detail['businessType'] == biz].dropna(subset=['_t']).sort_values('_t')
        if len(sub) == 0: continue
        dl_cn = [c for c in sub.columns if 'RLC' in c and 'Downlink' in c]
        ul_cn = [c for c in sub.columns if 'RLC' in c and 'Uplink' in c]
        if biz in DL_BIZ and dl_cn:
            col_name = dl_cn[0]
            rates_raw = pd.to_numeric(sub[col_name], errors='coerce').fillna(0) / 1e6
            rates = rates_raw[rates_raw > 0]
        elif biz in UL_BIZ and ul_cn:
            col_name = ul_cn[0]
            rates_raw = pd.to_numeric(sub[col_name], errors='coerce').fillna(0) / 1e6
            rates = rates_raw[rates_raw > 0]
        else:
            rates = pd.Series([], dtype=float); col_name = ''
        # 时长
        ts2 = sub['_t']; durs = []
        if len(ts2) > 0:
            s1 = ts2.iloc[0]; p = s1
            for t in ts2.iloc[1:]:
                if (t - p).total_seconds() > 5: durs.append((p - s1).total_seconds() + 1); s1 = t
                p = t
            durs.append((p - s1).total_seconds() + 1)
        # 记录原始数据位置（在详细过程df中的行号）
        r_start = sub.index[0] + 2  # +2 因为Excel第1行表头, 第2行索引0
        r_end = sub.index[-1] + 2
        sum_raw_data[biz] = {'rates': rates, 'durs': durs, 'col_name': col_name,
                             'r_start': r_start, 'r_end': r_end, 'sub': sub}
        row = {'业务类型': biz}
        if biz in DL_BIZ and len(rates) > 0:
            row['应用层FTP下载速率1000M以上占比(%)'] = round((rates > dl_clip).sum() / len(rates) * 100, 2)
            row['应用层FTP下载速率100M以下占比(%)'] = round((rates < 100).sum() / len(rates) * 100, 2)
            row['应用层平均下载速率(Mbps)'] = round(rates.mean(), 2)
            row['削峰应用层平均下载速率(Mbps)'] = round(rates.clip(upper=dl_clip).mean(), 2)
            sv = rates.sort_values(ascending=False); tn = max(1, int(len(sv) * 0.1))
            row['下行削峰TOP10%峰值速率'] = round(sv.head(tn).mean(), 2)
        elif biz in UL_BIZ and len(rates) > 0:
            row['应用层FTP上传速率200M以上占比(%)'] = round((rates > ul_clip).sum() / len(rates) * 100, 2)
            row['应用层FTP上传速率5M以下占比(%)'] = round((rates < 5).sum() / len(rates) * 100, 2)
            row['应用层平均上传速率'] = round(rates.mean(), 2)
            row['削峰应用层平均上传速率(Mbps)'] = round(rates.clip(upper=ul_clip).mean(), 2)
            sv = rates.sort_values(ascending=False); tn = max(1, int(len(sv) * 0.1))
            row['上行削峰TOP10%峰值速率'] = round(sv.head(tn).mean(), 2)
        if durs:
            row['业务时长平均值(s)'] = round(np.mean(durs), 2)
            row['业务时长中位值(s)'] = round(np.median(durs), 0)
        sr.append(row)
    df_summary = pd.DataFrame(sr)

    # ===== 汇总批注 + D/E列颜色 =====
    # 收集每业务的原始数据，用于批注
    raw_data = {}
    for biz in sorted(DL_BIZ | UL_BIZ):
        sub = df_detail[df_detail['businessType'] == biz].dropna(subset=['_t']).sort_values('_t')
        if len(sub) == 0: continue
        dl_cn = [c for c in sub.columns if 'RLC' in c and 'Downlink' in c]
        ul_cn = [c for c in sub.columns if 'RLC' in c and 'Uplink' in c]
        if biz in DL_BIZ and dl_cn:
            rates = pd.to_numeric(sub[dl_cn[0]], errors='coerce').fillna(0) / 1e6
            raw_data[biz] = {'rates': rates, 'durs': [], 'col': dl_cn[0], 'rows': (sub.index.min(), sub.index.max())}
        elif biz in UL_BIZ and ul_cn:
            rates = pd.to_numeric(sub[ul_cn[0]], errors='coerce').fillna(0) / 1e6
            raw_data[biz] = {'rates': rates, 'durs': [], 'col': ul_cn[0], 'rows': (sub.index.min(), sub.index.max())}
        ts2 = sub['_t']; durs = []
        if len(ts2) > 0:
            s1 = ts2.iloc[0]; p = s1
            for t in ts2.iloc[1:]:
                if (t - p).total_seconds() > 5: durs.append((p - s1).total_seconds() + 1); s1 = t
                p = t
            durs.append((p - s1).total_seconds() + 1)
        if biz in raw_data:
            raw_data[biz]['durs'] = durs

    # 等 ExcelWriter 写入完毕后，对汇总 sheet 加批注
    # 标注在 ExcelWriter 的 with 块内完成

    # ===== 对比横表（基准驱动，逐行匹配）=====
    BIZ_ORDER = ['FTP下载', 'FTP上传', '应用商店小文件下载', '应用商店大文件下载', '微信小包发送', '微信大包发送']
    rate_nm = {'FTP下载': '下载平均速率(Mbps)', 'FTP上传': '上传平均速率(Mbps)',
               '应用商店小文件下载': '小文件下载速率(Mbps)', '应用商店大文件下载': '大文件下载速率(Mbps)',
               '微信小包发送': '微信小包发送速率(Mbps)', '微信大包发送': '微信大包发送速率(Mbps)'}

    def fmt_time(t_str):
        m = _PARSE_RE.match(str(t_str))
        if m:
            dt = datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S')
            return f"{m.group(1)}({m.group(2)})", dt
        # 尝试毫秒格式 .xxx
        m2 = _PARSE_MS_RE.match(str(t_str))
        if m2:
            dt = datetime.strptime(m2.group(1), '%Y-%m-%d %H:%M:%S')
            return f"{m2.group(1)}({m2.group(2)[:3]})", dt
        return str(t_str), None

    def fmt_base_time(t):
        """基准时间格式化为 2026-07-02 22:27:36.557 样式(点+3位毫秒)，方便CTRL+F对照基准文件"""
        if t is None: return ''
        try:
            ts = pd.Timestamp(t)
            if pd.isna(ts): return str(t)
            return ts.strftime('%Y-%m-%d %H:%M:%S.') + f"{ts.microsecond // 1000:03d}"
        except Exception:
            return str(t)

    # 给df_detail添加_t列便于时间搜索
    df_detail['_t'] = df['_t'].values

    # 基准已在上方(详细过程模块)读取并按开始时间排序，此处直接复用，避免重复读盘
    base_sorted = base.reset_index(drop=True)

    # 按业务序列切轮：每轮以 FTP下载 开头(标准顺序
    # FTP下载→FTP上传→应用商店小文件→应用商店大文件→微信小包→微信大包)。
    # 第一个 FTP下载 之前的业务归为开头不完整轮次；末尾不足6个为末尾不完整轮次。
    # (替代 V2.0.9 的"每6条硬切一组"——硬切遇到开头/末尾不完整轮次会整体错位)
    base_rounds = []
    cur = []
    for i in range(len(base_sorted)):
        biz = str(base_sorted.iloc[i]['业务类型'])
        if biz == 'FTP下载' and cur:
            base_rounds.append(cur); cur = []
        cur.append(base_sorted.iloc[i])
    if cur:
        base_rounds.append(cur)

    # 逐轮逐业务匹配(基准驱动)。判定"对得上"不用固定时间窗口，而看业务连续段：
    #   对基准轮内业务 Bk，匹配区间由相邻业务边界决定——
    #     下界 = 上一个基准业务结束时间(轮首无下界)
    #     上界 = 下一个基准业务开始时间(轮末用自身结束时间)
    #   代码侧同业务段只要开始时间落在该区间且未被占用即算对得上；
    #   整轮6业务完整且全部对上才编号，否则该轮标"不计轮次"。
    all_items = []
    rnd_counter = 0
    used = set()  # 已占用的 df_detail 行索引，避免一代码段被多轮重复匹配
    for rnd_bizs in base_rounds:
        n = len(rnd_bizs)
        all_matched = True
        grp_items = []

        for k in range(n):
            brow = rnd_bizs[k]
            b_biz = str(brow['业务类型'])
            b_start_str = str(brow['开始时间'])
            b_end_str = str(brow['结束时间'])
            b_start = pd.to_datetime(brow['开始时间'])
            b_end = pd.to_datetime(brow['结束时间'])

            # 就近匹配(B方案)：在所有未占用的同业务行里，找离 b_start 最近的连续段
            # 距离<=300秒(约一轮间隔)才算同轮，避免跨轮串；used 防一段被多轮重复占用
            cand = df_detail[(df_detail['businessType'] == b_biz) &
                             (~df_detail.index.isin(used))]
            code_data = None
            if len(cand) > 0:
                dist = (cand['_t'] - b_start).abs()
                best_idx = dist.idxmin()
                if dist.loc[best_idx] <= pd.Timedelta(seconds=300):
                    # 取 best_idx 所在的连续行块(行号连续的同业务未占用行)
                    cand_idx_sorted = cand.index.sort_values()
                    pos = cand_idx_sorted.get_loc(best_idx)
                    lo = pos
                    while lo - 1 >= 0 and cand_idx_sorted[lo - 1] == cand_idx_sorted[lo] - 1:
                        lo -= 1
                    hi = pos
                    while hi + 1 < len(cand_idx_sorted) and cand_idx_sorted[hi + 1] == cand_idx_sorted[hi] + 1:
                        hi += 1
                    seg_idxs = list(cand_idx_sorted[lo:hi + 1])
                    used.update(seg_idxs)
                    cand_s = df_detail.loc[seg_idxs].sort_values('_t')
                    first = cand_s.iloc[0]
                    last = cand_s.iloc[-1]

                    c_start = str(first['Time'])
                    c_end = str(last['Time'])
                    c_start_fmt, _ = fmt_time(c_start)
                    c_end_fmt, _ = fmt_time(c_end)

                    code_data = {
                        '开始时间': c_start_fmt if c_start_fmt else c_start,
                        '结束时间': c_end_fmt if c_end_fmt else c_end,
                        '速率类指标': rate_nm.get(b_biz, ''),
                        '数值': first.get('速率(Mbps)'),
                        '时长类指标': '业务时长(秒)',
                        '数值.1': first.get('持续时长'),
                    }
                else:
                    all_matched = False
            else:
                all_matched = False

            # 基准数据(开始/结束时间用 .557 点样式，与基准文件一致便于CTRL+F)
            base_data = {
                '开始时间': fmt_base_time(brow['开始时间']),
                '结束时间': fmt_base_time(brow['结束时间']),
                '速率类指标': str(brow.get('速率类指标', '')),
                '数值': brow.get('数值'),
                '时长类指标': str(brow.get('时长类指标', '')),
                '数值.1': brow.get('数值.1'),
            }

            grp_items.append({
                'biz': b_biz,
                '基准': base_data,
                '代码': code_data,
            })

        complete = (n == 6)
        rnd_label = str(rnd_counter + 1) if (complete and all_matched) else '不计轮次'
        if complete and all_matched:
            rnd_counter += 1
        for item in grp_items: item['rnd_label'] = rnd_label
        all_items.extend(grp_items)

    # 偏差计算辅助函数
    def _pct(cv, bv):
        try:
            if pd.isna(cv) or pd.isna(bv): return '-'
            cf = float(cv); bf = float(bv)
            return '-' if bf == 0 else round((cf - bf) / bf * 100, 2)
        except Exception:
            return '-'

    def _time_diff(c_time, b_time):
        try:
            m = _PARSE_RE.match(str(c_time)); n = _PARSE_RE.match(str(b_time))
            if m and n:
                cdt = datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S') + timedelta(milliseconds=int(m.group(2)))
                bdt = datetime.strptime(n.group(1), '%Y-%m-%d %H:%M:%S') + timedelta(milliseconds=int(n.group(2)))
                return round((cdt - bdt).total_seconds(), 2)
            return '-'
        except Exception:
            return '-'

    # 构建df_compare：每业务2行(基准+代码)，偏差在右侧3列横铺
    cr = []
    for item in all_items:
        rnd = item['rnd_label']
        biz = item['biz']
        bd = item['基准']
        cd = item['代码']
        rpct = _pct(cd['数值'], bd['数值']) if cd else '-'
        dpct = _pct(cd['数值.1'], bd['数值.1']) if cd else '-'
        tdiff = _time_diff(cd['开始时间'], bd['开始时间']) if cd else '-'
        # 基准行（偏差列有值）
        cr.append({'轮次': rnd, '业务类型': biz, '来源': '基准',
                   '开始时间': bd['开始时间'], '结束时间': bd['结束时间'],
                   '速率类指标': bd['速率类指标'], '数值': bd['数值'],
                   '时长类指标': bd['时长类指标'], '数值.1': bd['数值.1'],
                   '速率偏差(%)': rpct, '时长偏差(%)': dpct, '时间差(秒)': tdiff})
        # 代码行（偏差列留空，由合并单元格覆盖）
        if cd:
            cr.append({'轮次': rnd, '业务类型': biz, '来源': '代码',
                       '开始时间': cd['开始时间'], '结束时间': cd['结束时间'],
                       '速率类指标': cd['速率类指标'], '数值': cd['数值'],
                       '时长类指标': cd['时长类指标'], '数值.1': cd['数值.1'],
                       '速率偏差(%)': '', '时长偏差(%)': '', '时间差(秒)': ''})
        else:
            cr.append({'轮次': rnd, '业务类型': biz, '来源': '代码',
                       '开始时间': '缺失', '结束时间': '缺失',
                       '速率类指标': '缺失', '数值': '缺失',
                       '时长类指标': '缺失', '数值.1': '缺失',
                       '速率偏差(%)': '', '时长偏差(%)': '', '时间差(秒)': ''})

    df_compare = pd.DataFrame(cr).reset_index(drop=True)

    # ===== 汇总：所有业务 =====
    def _calc_summary(sub_detail, dl_clip, ul_clip):
        """从 df_detail 子集计算汇总行"""
        sr = []
        for biz in sorted(DL_BIZ | UL_BIZ):
            sub = sub_detail[sub_detail['businessType'] == biz].dropna(subset=['_t']).sort_values('_t')
            if len(sub) == 0: continue
            dl_cn = [c for c in sub.columns if 'RLC' in c and 'Downlink' in c]
            ul_cn = [c for c in sub.columns if 'RLC' in c and 'Uplink' in c]
            if biz in DL_BIZ and dl_cn:
                cn = dl_cn[0]; rates = pd.to_numeric(sub[cn], errors='coerce').fillna(0) / 1e6; rates = rates[rates > 0]
            elif biz in UL_BIZ and ul_cn:
                cn = ul_cn[0]; rates = pd.to_numeric(sub[cn], errors='coerce').fillna(0) / 1e6; rates = rates[rates > 0]
            else: rates = pd.Series([], dtype=float); cn = ''
            ts2 = sub['_t']; durs = []
            if len(ts2) > 0:
                s1 = ts2.iloc[0]; p = s1
                for t in ts2.iloc[1:]:
                    if (t - p).total_seconds() > 5: durs.append((p - s1).total_seconds() + 1); s1 = t
                    p = t
                durs.append((p - s1).total_seconds() + 1)
            row = {'业务类型': biz}
            is_dl = biz in DL_BIZ; lim = dl_clip if is_dl else ul_clip; lo = 100 if is_dl else 5
            if len(rates) > 0:
                row['阈值以上占比(%)'] = round((rates > lim).sum() / len(rates) * 100, 2)
                row['5M/100M以下占比(%)'] = round((rates < lo).sum() / len(rates) * 100, 2)
                row['平均速率(Mbps)'] = round(rates.mean(), 2)
                row['削峰平均速率(Mbps)'] = round(rates.clip(upper=lim).mean(), 2)
                sv = rates.sort_values(ascending=False); tn = max(1, int(len(sv) * 0.1))
                row['削峰TOP10%峰值速率'] = round(sv.head(tn).mean(), 2)
            else:
                for k in ['阈值以上占比(%)','5M/100M以下占比(%)','平均速率(Mbps)','削峰平均速率(Mbps)','削峰TOP10%峰值速率']:
                    row[k] = '-'
            if durs:
                row['业务时长平均值(s)'] = round(np.mean(durs), 2)
                row['业务时长中位值(s)'] = round(np.median(durs), 0)
            sr.append(row)
        return pd.DataFrame(sr)

    log("生成汇总...", 70)
    df_summary_all = _calc_summary(df_detail, dl_clip, ul_clip)
    complete_mask = df_detail['轮次'].astype(str).str.isdigit()
    df_detail_complete = df_detail[complete_mask]
    df_summary_complete = _calc_summary(df_detail_complete, dl_clip, ul_clip)

    # 生成输出文件名
    version = 'V2.5.7'
    timestamp = datetime.now().strftime('%Y%m%d-%H%M')
    out = f'联通/电信最终输出_{version}_{timestamp}.xlsx'

    log("写入详细过程...", 78)
    with pd.ExcelWriter(out, engine='openpyxl') as w:
        df_detail.to_excel(w, sheet_name='数据业务-详细过程', index=False)

        from openpyxl.styles import Alignment, PatternFill, Border, Side, Font as OpFont
        from openpyxl.utils import get_column_letter as _gcl
        from openpyxl.comments import Comment
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # ===== 详细过程格式 =====
        ws = w.sheets['数据业务-详细过程']
        # D/E列：无流量填"-"，有值设0.00格式；D="-"的行 A列设0(随A=1筛选隐藏)
        for row_idx in range(2, ws.max_row + 1):
            if cancel_check and row_idx % 5000 == 0 and cancel_check():
                raise KeyboardInterrupt('用户取消')
            d_is_dash = False
            for col_letter in ['D', 'E']:
                cell = ws[f'{col_letter}{row_idx}']
                v = cell.value
                if v is None or (isinstance(v, (int, float)) and v == 0):
                    orig = df_detail.iloc[row_idx - 2, (4 if col_letter == 'D' else 5) - 1] if row_idx - 2 < len(df_detail) else None
                    if orig is None or (isinstance(orig, (int, float)) and orig == 0):
                        cell.value = '-'
                        if col_letter == 'D': d_is_dash = True
                    else:
                        cell.number_format = '0.00'
                else:
                    cell.number_format = '0.00'
            if d_is_dash:
                ws.cell(row=row_idx, column=1).value = 0
        for col_idx in range(1, ws.max_column + 1):
            mx = len(str(ws.cell(row=1, column=col_idx).value or ''))
            for r in range(2, min(102, ws.max_row + 1)):
                cv = ws.cell(row=r, column=col_idx).value
                if cv is not None: mx = max(mx, min(len(str(cv)), 30))
            ws.column_dimensions[_gcl(col_idx)].width = max(mx + 2, 10)
        for cell in ws[1]: cell.alignment = Alignment(wrap_text=True)
        ws.cell(row=1, column=4).value = '下载速率（Mbps）'
        ws.cell(row=1, column=5).value = '上传速率(Mbps)'
        # 冻结到第1行 I列(A~H列+表头冻结)
        ws.freeze_panes = 'I2'
        from openpyxl.worksheet.filters import FilterColumn
        ws.auto_filter.ref = ws.dimensions
        ws.auto_filter.add_filter_column(0, ['1'])
        # 无基准时隐藏 J~Q 列(基准7列 J~P + _t Q)
        if not has_base:
            for col_idx in range(10, 18):  # J=10 ~ Q=17
                ws.column_dimensions[_gcl(col_idx)].hidden = True
        # D/E 列颜色 (F~I条件)
        fill_d = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
        fill_e = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        for col_letter, fill in [('D', fill_d), ('E', fill_e)]:
            for cell in ws[col_letter]:
                if cell.row <= 1: continue
                if cell.value is None or str(cell.value) in ('nan','None','-',''): continue
                r = cell.row
                if any(ws.cell(row=r,column=c).value is not None and str(ws.cell(row=r,column=c).value) not in ('nan','None','')
                       for c in range(6,10)):
                    cell.fill = fill

        # ===== 汇总格式 =====
        log("写入汇总...", 85)
        if cancel_check and cancel_check(): raise KeyboardInterrupt('用户取消')
        wb = w.book
        ws_sum = wb.create_sheet('汇总')
        hdr_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        # 表头(合并格式)
        hdrs = ['业务类型',
                '应用层FTP上传/下载速率\n阈值以上占比(%)',
                '应用层FTP上传/下载速率\n5M/100M以下占比(%)',
                '应用层平均上传/下载\n速率(Mbps)',
                '削峰应用层平均上传/下载\n速率(Mbps)',
                '上行/下行削峰TOP10%\n峰值速率',
                '业务时长平均值(s)',
                '业务时长中位值(s)']
        for ci, h in enumerate(hdrs, 1):
            c = ws_sum.cell(row=1, column=ci, value=h); c.alignment = center; c.fill = hdr_fill
            c.font = OpFont(bold=True); c.border = border
        for ci in range(1, 9):
            ws_sum.column_dimensions[_gcl(ci)].width = 20

        # "所有业务"标题行
        ws_sum.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        tc = ws_sum.cell(row=2, column=1, value='▼ 所有识别业务')
        tc.alignment = center; tc.font = OpFont(bold=True, size=11)
        tc.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        for ci in range(1, 9): ws_sum.cell(row=2, column=ci).border = border

        # 汇总数据行(从row3开始)
        for ri in range(len(df_summary_all)):
            for ci in range(1, 9):
                cv = df_summary_all.iloc[ri, ci - 1] if ci - 1 < len(df_summary_all.columns) else None
                cell = ws_sum.cell(row=ri + 3, column=ci, value=cv if not (isinstance(cv, float) and pd.isna(cv)) else '')
                cell.alignment = center; cell.border = border
        all_data_rows = len(df_summary_all)

        # "仅完整轮次"标题行
        comp_start = all_data_rows + 4
        ws_sum.merge_cells(start_row=comp_start, start_column=1, end_row=comp_start, end_column=8)
        tc2 = ws_sum.cell(row=comp_start, column=1, value='▼ 仅完整轮次（代码侧6业务齐全）')
        tc2.alignment = center; tc2.font = OpFont(bold=True, size=11)
        tc2.fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
        for ci in range(1, 9): ws_sum.cell(row=comp_start, column=ci).border = border

        for ri in range(len(df_summary_complete)):
            for ci in range(1, 9):
                cv = df_summary_complete.iloc[ri, ci - 1] if ci - 1 < len(df_summary_complete.columns) else None
                cell = ws_sum.cell(row=ri + comp_start + 1, column=ci, value=cv if not (isinstance(cv, float) and pd.isna(cv)) else '')
                cell.alignment = center; cell.border = border

        # 自适应行高 + 冻结
        for r in range(1, comp_start + all_data_rows + 2):
            ws_sum.row_dimensions[r].height = 22
        ws_sum.freeze_panes = 'A3'

        # ===== 汇总批注 + 超链接 =====
        def _fmt_vals(vals, prec=2):
            if vals is None or len(vals) == 0: return '[]'
            vs = [round(v, prec) if prec > 0 else int(round(v)) for v in vals]
            if len(vs) <= 20:
                return '/'.join(str(v) for v in vs)
            return '/'.join(str(v) for v in vs[:5]) + '/.../' + '/'.join(str(v) for v in vs[-5:])

        def _add_comments(sub_detail, row_offset):
            for bi, biz in enumerate(sorted(DL_BIZ | UL_BIZ)):
                sub = sub_detail[sub_detail['businessType'] == biz].dropna(subset=['_t']).sort_values('_t')
                if len(sub) == 0: continue
                dl_cn = [c for c in sub.columns if 'RLC' in c and 'Downlink' in c]
                ul_cn = [c for c in sub.columns if 'RLC' in c and 'Uplink' in c]
                if biz in DL_BIZ and dl_cn: cn = dl_cn[0]
                elif biz in UL_BIZ and ul_cn: cn = ul_cn[0]
                else: continue
                rates = pd.to_numeric(sub[cn], errors='coerce').fillna(0) / 1e6; rates = rates[rates > 0]
                ts2 = sub['_t']; durs = []
                if len(ts2) > 0:
                    s1 = ts2.iloc[0]; p = s1
                    for t in ts2.iloc[1:]:
                        if (t - p).total_seconds() > 5: durs.append((p - s1).total_seconds() + 1); s1 = t
                        p = t
                    durs.append((p - s1).total_seconds() + 1)
                is_dl = biz in DL_BIZ; lim = dl_clip if is_dl else ul_clip; lo = 100 if is_dl else 5
                n = len(rates)
                sr = sub.index; r_start = sr[0] + 2; r_end = sr[-1] + 2
                row_idx = bi + row_offset

                for cell_col in range(2, 8):
                    cell = ws_sum.cell(row=row_idx, column=cell_col)
                    cv = cell.value
                    if cv is None or str(cv) in ('nan','None','-',''): continue
                    cp = cell_col - 2
                    if cp in (0, 1):  # above/below pct → 计数
                        if cp == 0:
                            mol = [r for r in rates if r > lim]
                            cnt_label = f"计数(rate>{lim})={len(mol)}"
                            op_cn = f"COUNT(rate > {lim}) / COUNT(all) x 100\n= 统计速率大于{lim}的采样数 / 总采样数 x 100"
                        else:
                            mol = [r for r in rates if r < lo]
                            cnt_label = f"计数(rate<{lo})={len(mol)}"
                            op_cn = f"COUNT(rate < {lo}) / COUNT(all) x 100\n= 统计速率小于{lo}的采样数 / 总采样数 x 100"
                        calc = f"= {len(mol)} / {n} x 100"
                    elif cp in (2, 3, 4):  # avg/clip/top10 → 求和
                        if cp == 2:
                            mol = list(rates); sum_v = sum(rates)
                            sum_label = f"求和={sum_v:.1f}"; cnt_label = f"计数={n}"
                            op_cn = "SUM(rate) / COUNT(rate)\n= 各速率求和 / 总采样数"
                            calc = f"= {sum_v:.1f} / {n}"
                        elif cp == 3:
                            clipped = [min(r, lim) for r in rates]; sum_v = sum(clipped); mol = clipped
                            sum_label = f"求和(MIN≤{lim})={sum_v:.1f}"; cnt_label = f"计数={n}"
                            op_cn = f"SUM(MIN(rate,{lim})) / COUNT(rate)\n= 各速率取{lim}上限后求和 / 总采样数"
                            calc = f"= {sum_v:.1f} / {n}"
                        elif cp == 4:
                            sr_rates = sorted(rates, reverse=True)
                            nt = max(1, int(len(sr_rates) * 0.1))
                            top_vals = sr_rates[:nt]; sum_v = sum(top_vals); mol = top_vals
                            sum_label = f"求和(TOP10%)={sum_v:.1f}"; cnt_label = f"计数(TOP10%)={nt}"
                            op_cn = "AVERAGE(top 10%)\n= 前10%最大速率的平均值"
                            calc = f"= {sum_v:.1f} / {nt}"
                    elif cp == 5:
                        continue
                    else:
                        continue
                    col_letter = _gcl(sub.columns.get_loc(cn) + 1)
                    txt = (f"{op_cn}\n{calc}\n= {cv}\n\n"
                           f"分子: {sum_label if cp in (2,3,4) else cnt_label}\n"
                           f"分子值: {_fmt_vals(mol)}\n"
                           f"分母: {cnt_label if cp in (2,3,4) else f'计数={n}'}\n"
                           f"分母值: {_fmt_vals(list(rates))}\n\n"
                           f"分子来源: {cn} 列, 第 {r_start}~{r_end} 行\n"
                           f"分母来源: {cn} 列, 第 {r_start}~{r_end} 行")
                    c = Comment(txt, "Tool"); cell.comment = c; c.width = 500; c.height = 220
                    cell.hyperlink = f"#'数据业务-详细过程'!{col_letter}{r_start}"

                # G/H 列：业务时长批注
                dur_vals = [d for d in durs if d > 0]
                if not dur_vals: continue
                for dur_col, dur_label in [(7, 'AVERAGE'), (8, 'MEDIAN')]:
                    cell = ws_sum.cell(row=row_idx, column=dur_col)
                    cv = cell.value
                    if cv is None or str(cv) in ('nan','None','-'): continue
                    if dur_col == 7:
                        dur_v = round(np.mean(dur_vals), 2)
                        txt = (f"AVERAGE(duration)\n= 各业务段时间求和 / 总段数\n"
                               f"求和={sum(dur_vals):.1f} / 计数={len(dur_vals)}\n= {cv}\n\n"
                               f"值: {_fmt_vals(dur_vals)}\n\n"
                               f"来源: 业务时长(秒) 列, 第 {r_start}~{r_end} 行")
                    else:
                        dur_v = round(np.median(dur_vals), 0)
                        txt = (f"MEDIAN(duration)\n= 业务段时间中位数\n"
                               f"计数={len(dur_vals)}\n= {cv}\n\n"
                               f"值: {_fmt_vals(dur_vals)}\n\n"
                               f"来源: 业务时长(秒) 列, 第 {r_start}~{r_end} 行")
                    c = Comment(txt, "Tool"); cell.comment = c; c.width = 420; c.height = 180

        # 上表批注
        _add_comments(df_detail, 3)
        # 下表批注
        _add_comments(df_detail_complete, comp_start + 1)

        # ===== 对比横表 =====
        log("写入对比横表...", 93)
        if cancel_check and cancel_check(): raise KeyboardInterrupt('用户取消')
        wb = w.book
        ws_h = wb.create_sheet('数据业务-对比')
        hdr = ['轮次', '业务类型', '来源', '开始时间', '结束时间', '速率类指标', '数值', '时长类指标', '数值.1',
               '速率偏差(%)', '时长偏差(%)', '时间差(秒)']
        for ci, h in enumerate(hdr, 1):
            c = ws_h.cell(row=1, column=ci, value=h)
            c.alignment = center; c.fill = hdr_fill; c.font = OpFont(bold=True); c.border = border

        color1 = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        color2 = PatternFill(start_color='E5FFCC', end_color='E5FFCC', fill_type='solid')
        color_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        biz_color_map = {bn: (color1 if i % 2 == 0 else color2) for i, bn in enumerate(BIZ_ORDER)}

        hr = 2; rnd_start_row = 2; prev_rnd = None
        for ci in range(0, len(df_compare)):
            row = df_compare.iloc[ci]
            rnd = str(row['轮次']); biz = row['业务类型']; src = str(row['来源'])

            # 轮次切换：合并上一轮的轮次列
            if rnd != prev_rnd and prev_rnd is not None:
                ws_h.merge_cells(start_row=rnd_start_row, start_column=1, end_row=hr - 1, end_column=1)
                ws_h.cell(row=rnd_start_row, column=1).alignment = center
                rnd_start_row = hr
            prev_rnd = rnd

            if '缺失' in str(row.get('开始时间', '')): fill = color_yellow
            else: fill = biz_color_map.get(biz, color1)

            vals = [rnd if hr == rnd_start_row else '', biz, src,
                    str(row['开始时间']) if pd.notna(row.get('开始时间')) else '',
                    str(row['结束时间']) if pd.notna(row.get('结束时间')) else '',
                    str(row['速率类指标']) if pd.notna(row.get('速率类指标')) else '',
                    row['数值'] if pd.notna(row.get('数值')) else '',
                    str(row['时长类指标']) if pd.notna(row.get('时长类指标')) else '',
                    row['数值.1'] if pd.notna(row.get('数值.1')) else '',
                    row['速率偏差(%)'] if pd.notna(row.get('速率偏差(%)')) else '',
                    row['时长偏差(%)'] if pd.notna(row.get('时长偏差(%)')) else '',
                    row['时间差(秒)'] if pd.notna(row.get('时间差(秒)')) else '']
            for ri, v in enumerate(vals):
                c = ws_h.cell(row=hr, column=ri + 1, value=v if v != '' else '')
                c.alignment = center; c.fill = fill; c.border = border
            hr += 1
        if prev_rnd is not None and hr > rnd_start_row:
            ws_h.merge_cells(start_row=rnd_start_row, start_column=1, end_row=hr - 1, end_column=1)
            ws_h.cell(row=rnd_start_row, column=1).alignment = center

        # 每2行合并：业务类型(2) + 偏差列(10,11,12)
        for r in range(2, hr, 2):
            if r + 1 > hr - 1: break
            for mc in [2, 10, 11, 12]:
                ws_h.merge_cells(start_row=r, start_column=mc, end_row=r + 1, end_column=mc)
                ws_h.cell(row=r, column=mc).alignment = center

        for cidx in range(1, len(hdr) + 1):
            mx = 0
            for ridx in range(1, min(hr, 120)):
                cv = ws_h.cell(row=ridx, column=cidx).value
                if cv is not None: mx = max(mx, min(len(str(cv)), 30))
            ws_h.column_dimensions[_gcl(cidx)].width = max(mx + 2, 10)
        ws_h.freeze_panes = 'C2'

    log(f"完成! {out}", 100)
    # 清理解压临时目录
    for _d in _tmp_dirs:
        try: shutil.rmtree(_d)
        except Exception: pass
    return out


def _detect_operator(filepath):
    """从文件路径判断运营商：联通/电信"""
    low = filepath.lower()
    if '电信' in low or 'dx' in low:
        return '电信'
    if '联通' in low or 'lt' in low or 'cu' in low:
        return '联通'
    return '未知'

def _detect_file_type(filename):
    """判断文件类型：voice/voice_vqi 或 data/monitoring

    通过文件名和列名综合判断：
    - 如果文件名包含nr_virtual_ue_trace或.tmf，或有MessageType列 → 返回"语音"
    - 否则 → 返回"数据"
    """
    name = filename.lower()
    if 'nr_virtual_ue_trace' in name or name.endswith('.tmf'):
        return '语音'
    if 'user_common_monitoring' in name or name.endswith('.mmf'):
        return '数据'
    # 对于其他文件，需要读取后通过列名判断，这里默认返回数据
    # 实际使用时会在process_vqi中通过_has_message_type_col再次判断
    return '数据'

_VQI_RE = re.compile(r'DlE2eVqi=(\d+)')

def _extract_vqi_vale(detail_info):
    """从Detail Info提取DlE2eVqi值，排除65535"""
    if pd.isna(detail_info):
        return None
    m = _VQI_RE.search(str(detail_info))
    if m:
        val = int(m.group(1))
        return val if val != 65535 else None
    return None

def process_vqi(files, callback=None, cancel_check=None, progress_cb=None, time_filter=None, time_range=None, merge_raw=True):
    """VQI处理函数：自动分类语音/数据文件，生成语音统计输出

    参数:
        files: 输入文件列表
        callback: 日志回调函数
        cancel_check: 取消检查函数
        progress_cb: 进度回调函数
        time_filter: 时间段过滤 (use_time, start_time, end_time)
        time_range: 时间起止范围，用于汇总表。若为None则从数据自动计算
    """
    def log(msg, pct=None):
        if cancel_check and cancel_check():
            raise KeyboardInterrupt('用户取消')
        import time as _time
        _t = _time.time()
        if not hasattr(log, 'last_time'):
            log.last_time = _t
        elapsed = _t - log.last_time if pct is not None else 0
        prefix = f"({elapsed:.1f}秒) " if pct is not None and elapsed > 0 else ""
        print(msg)
        if callback: callback(f"{prefix}{msg}")
        if pct is not None and progress_cb: progress_cb(pct)
        log.last_time = _t

    import zipfile, tempfile, shutil, glob

    # 展开压缩文件
    _tmp_dirs = []
    def _expand_zip(zip_path):
        _tmp = tempfile.mkdtemp(prefix='vqi_unzip_'); _tmp_dirs.append(_tmp)
        _xs = []
        try:
            with zipfile.ZipFile(zip_path) as _zf: _zf.extractall(_tmp)
            _xs = glob.glob(os.path.join(_tmp, '**', '*.xlsx'), recursive=True) + \
                  glob.glob(os.path.join(_tmp, '**', '*.xls'), recursive=True) + \
                  glob.glob(os.path.join(_tmp, '**', '*.mmf'), recursive=True) + \
                  glob.glob(os.path.join(_tmp, '**', '*.tmf'), recursive=True)
            _inner_zips = glob.glob(os.path.join(_tmp, '**', '*.zip'), recursive=True)
            for _iz in _inner_zips:
                _xs.extend(_expand_zip(_iz))
        except Exception as _e:
            log(f"  解压失败 {os.path.basename(zip_path)}: {_e}")
        return _xs

    # 展开所有文件
    all_files = []
    for _f in (files or []):
        if str(_f).lower().endswith('.zip'):
            _xs = _expand_zip(_f)
            log(f"  解压 {os.path.basename(_f)}: 找到 {len(_xs)} 个文件(含嵌套)", 5)
            all_files.extend(_xs)
        else:
            all_files.append(_f)

    if not all_files:
        log("无有效文件", 100)
        return None

    log(f"共 {len(all_files)} 个文件", 5)

    # 分类文件：读取每个文件检查是否有MessageType列
    # 有MessageType → 语音文件；没有 → 数据文件（跳过处理）
    voice_files = []
    data_files = []
    data_files_skipped = 0
    for f in all_files:
        try:
            df_test = _read_file(f, nrows=3)
            has_msg = _has_message_type_col(df_test)
            log(f"  文件 {os.path.basename(f)}: {'✓ MessageType列' if has_msg else 'x 无MessageType列'}", 6)
            if has_msg:
                voice_files.append(f)
            else:
                data_files.append(f)
        except Exception as e:
            log(f"  文件 {os.path.basename(f)}: 读取失败 - {e}", 6)
            data_files_skipped += 1

    log(f"  语音文件: {len(voice_files)} 个, 数据文件: {len(data_files)} 个, 跳过的数据文件: {data_files_skipped} 个", 8)

    # ===== 读取语音文件 =====
    voice_dfs = []
    voice_file_info = {}  # 记录每个文件的基本信息（行数、时间范围等）
    for fi, f in enumerate(voice_files):
        log(f"  读取语音文件 {fi+1}/{len(voice_files)}: {os.path.basename(f)}", 10 + int(10 * (fi+1) / max(len(voice_files), 1)))
        operator = _detect_operator(f)
        try:
            df_i = _read_file(f)
        except Exception:
            log(f"    读取失败: {f}", 10)
            continue
        # 统一MessageType列名（支持Message Type、message_type等变体）
        if 'Message Type' in df_i.columns and 'MessageType' not in df_i.columns:
            df_i = df_i.rename(columns={'Message Type': 'MessageType'})
        elif 'message_type' in df_i.columns and 'MessageType' not in df_i.columns:
            df_i = df_i.rename(columns={'message_type': 'MessageType'})
        df_i['_t'] = df_i['Time'].apply(parse_time)
        df_i['运营商'] = operator
        df_i['业务类型'] = '语音业务'
        # 记录来源文件名，用于后续按文件统计
        filename_base = os.path.basename(f)
        df_i['_source_file'] = filename_base
        voice_dfs.append(df_i)

        # 记录该文件的时间范围和行数
        times = df_i['_t'].dropna()
        time_range_str = ''
        if len(times) > 0:
            time_range_str = f"{times.min().strftime('%Y-%m-%d %H:%M:%S')} ~ {times.max().strftime('%Y-%m-%d %H:%M:%S')}"
        voice_file_info[filename_base] = {
            'operator': operator,
            'row_count': len(df_i),
            'time_range': time_range_str,
            'time_start': times.min() if len(times) > 0 else None,
            'time_end': times.max() if len(times) > 0 else None
        }

    # 合并所有语音数据
    if voice_dfs:
        df_voice_raw = pd.concat(voice_dfs, ignore_index=True)
        df_voice_raw = df_voice_raw.sort_values('_t', kind='stable').reset_index(drop=True)
    else:
        df_voice_raw = pd.DataFrame()

    log(f"语音数据合并完成: {len(df_voice_raw)} 行", 22)

    # ===== 数据文件处理 =====
    log("处理数据文件...", 23)
    data_dfs = []
    data_file_info = {}
    for fi, f in enumerate(data_files):
        if cancel_check and cancel_check(): raise KeyboardInterrupt('用户取消')
        log(f"  读取数据文件 {fi+1}/{len(data_files)}: {os.path.basename(f)}", 24 + int(5 * (fi+1) / max(len(data_files), 1)))
        operator = _detect_operator(f)
        try:
            df_i = _read_file(f)
        except Exception:
            log(f"    读取失败: {f}", 24)
            continue
        has_time = 'Time' in df_i.columns
        has_qci = 'QCI' in df_i.columns
        has_rlc = any('RLC' in c and ('Downlink' in c or 'Uplink' in c) for c in df_i.columns)
        if not (has_time and (has_qci or has_rlc)):
            log(f"    跳过非数据文件(无Time/QCI/RLC列)", 24)
            data_files_skipped += 1
            continue
        
        df_i['_t'] = df_i['Time'].apply(parse_time)
        df_i['运营商'] = operator
        df_i['业务类型'] = '数据业务'
        
        rlc_dl = [c for c in df_i.columns if 'RLC' in c and 'Downlink' in c]
        rlc_ul = [c for c in df_i.columns if 'RLC' in c and 'Uplink' in c]
        if rlc_dl:
            df_i['_dl_rlc'] = pd.to_numeric(df_i[rlc_dl[0]], errors='coerce').fillna(0) / 1e6
        else:
            df_i['_dl_rlc'] = 0
        if rlc_ul:
            df_i['_ul_rlc'] = pd.to_numeric(df_i[rlc_ul[0]], errors='coerce').fillna(0) / 1e6
        else:
            df_i['_ul_rlc'] = 0
        
        data_dfs.append(df_i)
        
        dl_vals = df_i['_dl_rlc'].dropna().values
        ul_vals = df_i['_ul_rlc'].dropna().values
        dl_nonzero = dl_vals[dl_vals > 0]
        ul_nonzero = ul_vals[ul_vals > 0]
        
        dl_peak = 0
        if len(dl_nonzero) > 0:
            tn = max(1, int(len(dl_nonzero) * 0.1))
            dl_sorted = np.sort(dl_nonzero)[::-1]
            dl_peak = round(min(float(np.mean(dl_sorted[:tn])), 1000), 2)
        
        dl_avg = 0
        if len(dl_nonzero) > 0:
            dl_avg = round(min(float(np.mean(dl_nonzero)), 1000), 2)
        
        ul_peak = 0
        if len(ul_nonzero) > 0:
            tn = max(1, int(len(ul_nonzero) * 0.1))
            ul_sorted = np.sort(ul_nonzero)[::-1]
            ul_peak = round(min(float(np.mean(ul_sorted[:tn])), 200), 2)
        
        ul_avg = 0
        if len(ul_nonzero) > 0:
            ul_avg = round(min(float(np.mean(ul_nonzero)), 200), 2)
        
        times = df_i['_t'].dropna()
        time_range_str = ''
        if len(times) > 0:
            time_range_str = f"{times.min().strftime('%Y-%m-%d %H:%M:%S')} ~ {times.max().strftime('%Y-%m-%d %H:%M:%S')}"
        
        data_file_info[os.path.basename(f)] = {
            'operator': operator,
            'row_count': len(df_i),
            'time_range': time_range_str,
            'dl_peak': dl_peak,
            'dl_avg': dl_avg,
            'ul_peak': ul_peak,
            'ul_avg': ul_avg
        }
    
    if data_dfs:
        df_data_raw = pd.concat(data_dfs, ignore_index=True)
        df_data_raw = df_data_raw.sort_values('_t', kind='stable').reset_index(drop=True)
    else:
        df_data_raw = pd.DataFrame()
    
    log(f"数据文件处理完成: {len(data_dfs)} 个文件, {len(df_data_raw)} 行", 28)
    
    data_operator_indicators = {}
    for op in ['电信', '联通']:
        op_data = [info for fname, info in data_file_info.items() if info['operator'] == op]
        if op_data:
            dl_peaks = [d['dl_peak'] for d in op_data if d['dl_peak'] > 0]
            dl_avgs = [d['dl_avg'] for d in op_data if d['dl_avg'] > 0]
            ul_peaks = [d['ul_peak'] for d in op_data if d['ul_peak'] > 0]
            ul_avgs = [d['ul_avg'] for d in op_data if d['ul_avg'] > 0]
            data_operator_indicators[op] = {
                '下行前10%峰值速率': round(np.mean(dl_peaks), 2) if dl_peaks else '',
                '下行均值速率': round(np.mean(dl_avgs), 2) if dl_avgs else '',
                '上行前10%峰值速率': round(np.mean(ul_peaks), 2) if ul_peaks else '',
                '上行均值速率': round(np.mean(ul_avgs), 2) if ul_avgs else ''
            }
        else:
            data_operator_indicators[op] = {
                '下行前10%峰值速率': '',
                '下行均值速率': '',
                '上行前10%峰值速率': '',
                '上行均值速率': ''
            }


    # ===== 时间段过滤 =====
    if time_filter and len(df_voice_raw) > 0:
        use_time, t_start, t_end = time_filter
        if use_time and t_start is not None and t_end is not None:
            t_start_t = t_start.toPython() if hasattr(t_start, 'toPython') else t_start
            t_end_t = t_end.toPython() if hasattr(t_end, 'toPython') else t_end
            df_voice_raw['_time_only'] = df_voice_raw['_t'].dt.time
            mask = (df_voice_raw['_time_only'] >= t_start_t) & (df_voice_raw['_time_only'] <= t_end_t)
            df_voice_raw = df_voice_raw[mask].copy()
            df_voice_raw = df_voice_raw.drop('_time_only', axis=1)
            log(f"  时间段过滤: {t_start_t} ~ {t_end_t}, 剩余 {len(df_voice_raw)} 行", 22)

    # ===== 筛选MessageType非空行 =====
    if 'MessageType' in df_voice_raw.columns and len(df_voice_raw) > 0:
        df_msg = df_voice_raw[df_voice_raw['MessageType'].notna() &
                              (df_voice_raw['MessageType'].astype(str).str.strip() != '')].copy()
    else:
        df_msg = pd.DataFrame()
    log(f"  MessageType非空行数: {len(df_msg)}", 28)

    # ===== 语音统计（INVITE→180→BYE→200） =====
    log("识别语音通话...", 35)
    voice_records = []
    rnd = 0

    if len(df_msg) > 0 and 'MessageType' in df_msg.columns:
        # 向量化：找出所有INVITE的位置
        invite_mask = df_msg['MessageType'].astype(str) == 'SIP_REQ_INVITE'
        invite_positions = df_msg.index[invite_mask].tolist()
        n = len(df_msg)

        # 过滤掉连续的INVITE（时间差<1秒算连续，跳过中间的）
        valid_invite_positions = []
        for idx_val in invite_positions:
            if valid_invite_positions:
                prev_val = valid_invite_positions[-1]
                if abs((df_msg.loc[idx_val, '_t'] - df_msg.loc[prev_val, '_t']).total_seconds()) < 1.0:
                    continue  # 跳过连续INVITE
            valid_invite_positions.append(idx_val)

        invite_positions = valid_invite_positions
        total_invites = len(invite_positions)
        log(f"  找到 {total_invites} 个有效INVITE", 36)

        for ii, pos in enumerate(invite_positions):
            t1 = df_msg.loc[pos, '_t']
            t1_time = df_msg.loc[pos, 'Time']
            operator = df_msg.loc[pos, '运营商']

            # 搜索范围：从当前INVITE到下一个INVITE（或到数据末尾）
            next_pos = invite_positions[ii + 1] if ii + 1 < len(invite_positions) else n + pos
            end_j = next_pos - 1  # 下一个INVITE的前一行

            if pos + 1 > end_j:
                # 两个INVITE之间没有其他行
                rnd += 1
                voice_records.append({
                    '轮次': rnd,
                    '运营商': operator,
                    '开始时间': t1_time,
                    '结束时间': t1_time,
                    '呼叫建立时长(s)': '',
                    '是否完整通话': '否',
                    '_source_file': df_msg.loc[pos].get('_source_file', '')  # 记录来源文件名
                })
                continue

            # 获取搜索范围内的数据子集（向量化操作）
            search_range = df_msg.loc[pos + 1:end_j]

            # 找SIP_RSP_180（第一个）
            r180 = search_range[search_range['MessageType'].astype(str) == 'SIP_RSP_180']
            t2 = r180.iloc[0]['_t'] if len(r180) > 0 else None

            # 找SIP_REQ_BYE
            bye = search_range[search_range['MessageType'].astype(str) == 'SIP_REQ_BYE']
            found_bye_200 = False
            bye_time_raw = None
            last_time_raw = search_range.iloc[-1]['Time'] if len(search_range) > 0 else t1_time

            if len(bye) > 0:
                bye_time = bye.iloc[0]['_t']
                bye_time_raw = bye.iloc[0]['Time']
                bye_pos = bye.index[0]
                # 在BYE附近1秒内找200
                k = bye_pos + 1
                while k <= bye_pos + 20 and k < n:
                    check_msg = str(df_msg.loc[k, 'MessageType'])
                    if check_msg == 'SIP_REQ_INVITE':
                        break
                    if check_msg == 'SIP_RSP_200' and (df_msg.loc[k, '_t'] - bye_time).total_seconds() <= 1:
                        found_bye_200 = True
                        last_time_raw = df_msg.loc[k, 'Time']
                        break
                    k += 1
                if found_bye_200:
                    last_time_raw = bye_time_raw
            else:
                # 无BYE，结束时间=INVITE之后最后一个消息的时间
                last_time_raw = search_range.iloc[-1]['Time'] if len(search_range) > 0 else t1_time

            # 呼叫建立时长
            call_setup_duration = ''
            if t2 is not None:
                call_setup_duration = round((t2 - t1).total_seconds(), 2)

            # 完整通话
            is_complete = '是' if found_bye_200 else '否'

            rnd += 1
            voice_records.append({
                '轮次': rnd,
                '运营商': operator,
                '开始时间': t1_time,
                '结束时间': last_time_raw,
                '呼叫建立时长(s)': call_setup_duration,
                '是否完整通话': is_complete,
                '_source_file': df_msg.loc[pos].get('_source_file', '')  # 记录来源文件名
            })

    log(f"识别完成: {len(voice_records)} 轮语音通话", 50)

    # ===== DlE2eVqi提取（从Detail Info列）=====
    log("提取DlE2eVqi值...", 55)
    vqi_records = []

    # 找Detail Info列（支持中英文名）
    detail_col = None
    if len(df_voice_raw) > 0:
        for col_name in ['Detail Info', '详细信息', 'DetailInfo', 'detail_info']:
            if col_name in df_voice_raw.columns:
                detail_col = col_name
                break

    if detail_col and len(df_voice_raw) > 0:
        for idx, row in df_voice_raw.iterrows():
            vqi = _extract_vqi_vale(row[detail_col])
            if vqi is not None:
                mos = vqi / 100.0
                is_quality = '是' if mos > 3.5 else '否'
                vqi_records.append({
                    '时间': row.get('Time', ''),
                    '运营商': row.get('运营商', _detect_operator(str(row.get('文件名', '')))),
                    'DlE2eVqi原始值': vqi,
                    'MOS值': round(mos, 2),
                    '是否优质(MOS>3.5)': is_quality,
                    '_source_file': row.get('_source_file', '')
                })

    df_vqi = pd.DataFrame(vqi_records)

    # 计算优质占比
    if len(df_vqi) > 0:
        total = len(df_vqi)
        quality = (df_vqi['是否优质(MOS>3.5)'] == '是').sum()
        quality_pct = round(quality / total * 100, 2)
    else:
        total, quality, quality_pct = 0, 0, 0

    # ===== 生成输出 =====
    # 原始数据sheet：所有语音文件的全部行（加运营商和业务类型列）
    df_raw_out = df_voice_raw.copy()
    if '_t' in df_raw_out.columns:
        df_raw_out = df_raw_out.drop('_t', axis=1)
    if '_time_only' in df_raw_out.columns:
        df_raw_out = df_raw_out.drop('_time_only', axis=1)

    # 语音统计sheet
    df_voice_out = pd.DataFrame(voice_records)
    if len(df_voice_out) == 0:
        df_voice_out = pd.DataFrame(columns=['轮次', '运营商', '开始时间', '结束时间', '呼叫建立时长(s)', '是否完整通话'])

    # MOS3.5计算明细sheet（底部加汇总）
    if len(df_vqi) > 0:
        # 底部汇总行
        summary_row = pd.DataFrame([{
            '时间': '汇总',
            '运营商': '',
            'DlE2eVqi原始值': '',
            'MOS值': '',
            '是否优质(MOS>3.5)': f'分子={quality} / 分母={total} = {quality_pct}%'
        }])
        df_vqi_out = pd.concat([df_vqi, summary_row], ignore_index=True)
    else:
        df_vqi_out = pd.DataFrame(columns=['时间', '运营商', 'DlE2eVqi原始值', 'MOS值', '是否优质(MOS>3.5)'])

    # ===== 按文件名统计 =====
    file_stats = []

    # 语音文件统计（使用voice_file_info和_source_file字段）
    for fname_base, f_info in voice_file_info.items():
        operator = f_info['operator']
        file_row_count = f_info['row_count']
        file_time_range_str = f_info['time_range']

        # 从voice_records中筛选属于该文件的轮次（通过_source_file匹配）
        file_voice_records = [vr for vr in voice_records if vr.get('_source_file') == fname_base]
        file_rounds = len(file_voice_records)

        # 平均呼叫建立时长（只统计有值的轮次）
        setup_vals = []
        for vr in file_voice_records:
            if vr.get('呼叫建立时长(s)') != '' and pd.notna(vr.get('呼叫建立时长(s)')):
                try:
                    setup_vals.append(float(vr['呼叫建立时长(s)']))
                except:
                    pass
        file_avg_setup = round(np.mean(setup_vals), 2) if setup_vals else ''

        # 优质通话占比(MOS3.5)：从vqi_records中筛选属于该文件的记录
        file_vqi_records = [r for r in vqi_records if r.get('_source_file') == fname_base]
        file_quality_pct = ''
        if file_vqi_records:
            total_file = len(file_vqi_records)
            quality_file = sum(1 for r in file_vqi_records if r.get('是否优质(MOS>3.5)') == '是')
            if total_file > 0:
                file_quality_pct = f"{round(quality_file / total_file * 100, 2)}%"

        file_stats.append({
            '文件名': fname_base,
            '运营商': operator,
            '文件类型': '语音',
            '行数': file_row_count,
            '时间范围': file_time_range_str,
            '通话轮次': file_rounds,
            '平均呼叫建立时长(s)': file_avg_setup,
            '优质通话占比(MOS3.5)': file_quality_pct,
            '下行前10%峰值速率': '',
            '下行均值速率': '',
            '上行前10%峰值速率': '',
            '上行均值速率': ''
        })
            '文件名': filename_base,
            '运营商': operator,
            '文件类型': file_type,
            '行数': file_row_count,
            '时间范围': file_time_range_str,
            '通话轮次': file_rounds,
            '平均呼叫建立时长(s)': file_avg_setup,
            '优质通话占比(MOS3.5)': file_quality_pct,
            '下行前10%峰值速率': '',
            '下行均值速率': '',
            '上行前10%峰值速率': '',
            '上行均值速率': ''
        })

    # 添加数据文件行到按文件名统计
    for fname, info in data_file_info.items():
        file_stats.append({
            '文件名': fname,
            '运营商': info['operator'],
            '文件类型': '数据',
            '行数': info['row_count'],
            '时间范围': info['time_range'],
            '通话轮次': 0,
            '平均呼叫建立时长(s)': '',
            '优质通话占比(MOS3.5)': '',
            '下行前10%峰值速率': info['dl_peak'],
            '下行均值速率': info['dl_avg'],
            '上行前10%峰值速率': info['ul_peak'],
            '上行均值速率': info['ul_avg']
        })
    
    df_file_stats = pd.DataFrame(file_stats)

    # ===== 按运营商统计 =====
    operator_stats = []
    for op in ['电信', '联通']:
        op_voice = df_voice_out[df_voice_out['运营商'] == op]
        op_vqi = df_vqi[df_vqi['运营商'] == op] if len(df_vqi) > 0 else pd.DataFrame()

        voice_count = len(op_voice)
        call_rounds = len(op_voice)

        # 平均呼叫建立时长（只统计有值的轮次）
        avg_setup = ''
        if len(op_voice) > 0 and '呼叫建立时长(s)' in op_voice.columns:
            setup_vals = pd.to_numeric(op_voice['呼叫建立时长(s)'], errors='coerce').dropna()
            if len(setup_vals) > 0:
                avg_setup = round(setup_vals.mean(), 2)

        # 优质通话占比
        quality_pct_op = ''
        if len(op_vqi) > 0:
            total_op = len(op_vqi)
            quality_op = (op_vqi['是否优质(MOS>3.5)'] == '是').sum()
            if total_op > 0:
                quality_pct_op = f"{round(quality_op / total_op * 100, 2)}%"

        # 语音文件数
        voice_file_count = len(df_file_stats[df_file_stats['运营商'] == op]) if len(df_file_stats) > 0 else 0

        operator_stats.append({
            '运营商': op,
            '语音文件数': voice_file_count,
            '通话轮次': call_rounds,
            '平均呼叫建立时长(s)': avg_setup,
            '优质通话占比(MOS3.5)': quality_pct_op,
            '下行前10%峰值速率': '',
            '下行均值速率': '',
            '上行前10%峰值速率': '',
            '上行均值速率': '',
            '微信大文件上传中位数时长(s)': '',
            '微信大文件发送平均速率(Mbps)': '',
            '微信小文件发送平均时延(s)': '',
            '微信小文件发送平均速率(Mbps)': '',
            '商店大app平均时延(s)': '',
            '商店大app平均速率(Mbps)': '',
            '商店小app平均时延(s)': '',
            '商店小app平均速率(Mbps)': ''
        })

    df_operator_stats = pd.DataFrame(operator_stats)

    # ===== 时间起止 =====
    if time_range:
        time_range_str = time_range
    elif time_filter and time_filter[0] and time_filter[1] is not None and time_filter[2] is not None:
        time_range_str = f"{time_filter[1].toString('HH:mm:ss')} ~ {time_filter[2].toString('HH:mm:ss')}"
    else:
        # 从数据中计算
        if len(df_voice_raw) > 0 and '_t' in df_voice_raw.columns and df_voice_raw['_t'].notna().any():
            t_min = df_voice_raw['_t'].min()
            t_max = df_voice_raw['_t'].max()
            time_range_str = f"{t_min.strftime('%Y-%m-%d %H:%M:%S')} ~ {t_max.strftime('%Y-%m-%d %H:%M:%S')}"
        else:
            time_range_str = ''

    # ===== 汇总-数据和语音 =====
    summary_data = []

    # 标准行
    summary_data.append({
        '运营商': '标准',
        '文件名': '',
        '起止时间': '',
        '呼叫建立时长(s)': 1,
        '优质通话占比(MOS3.5)': '95%',
        '下行前10%峰值速率': 1000,
        '下行均值速率': 800,
        '上行前10%峰值速率': 200,
        '上行均值速率': 160,
        '微信大文件上传中位数时长(s)': 30,
        '微信大文件发送平均速率(Mbps)': 50,
        '微信小文件发送平均时延(s)': 2,
        '微信小文件发送平均速率(Mbps)': 20,
        '商店大app平均时延(s)': 20,
        '商店大app平均速率(Mbps)': 400,
        '商店小app平均时延(s)': 4,
        '商店小app平均速率(Mbps)': 200
    })

    # 电信行、联通行
    for op in ['电信', '联通']:
        op_voice = df_voice_out[df_voice_out['运营商'] == op]
        op_vqi = df_vqi[df_vqi['运营商'] == op] if len(df_vqi) > 0 else pd.DataFrame()

        # 呼叫建立时长
        avg_setup = ''
        if len(op_voice) > 0 and '呼叫建立时长(s)' in op_voice.columns:
            setup_vals = pd.to_numeric(op_voice['呼叫建立时长(s)'], errors='coerce').dropna()
            if len(setup_vals) > 0:
                avg_setup = round(setup_vals.mean(), 2)

        # 优质通话占比
        quality_pct_op = ''
        if len(op_vqi) > 0:
            total_op = len(op_vqi)
            quality_op = (op_vqi['是否优质(MOS>3.5)'] == '是').sum()
            if total_op > 0:
                quality_pct_op = f"{round(quality_op / total_op * 100, 2)}%"

        op_data_indicators = data_operator_indicators.get(op, {})
        summary_data.append({
            '运营商': op,
            '文件名': f"{len(df_file_stats[df_file_stats['运营商'] == op]) if len(df_file_stats) > 0 else 0}个语音文件" if len(df_file_stats) > 0 else '',
            '起止时间': time_range_str,
            '呼叫建立时长(s)': avg_setup,
            '优质通话占比(MOS3.5)': quality_pct_op,
            '下行前10%峰值速率': op_data_indicators.get('下行前10%峰值速率', ''),
            '下行均值速率': op_data_indicators.get('下行均值速率', ''),
            '上行前10%峰值速率': op_data_indicators.get('上行前10%峰值速率', ''),
            '上行均值速率': op_data_indicators.get('上行均值速率', ''),
            '微信大文件上传中位数时长(s)': '',
            '微信大文件发送平均速率(Mbps)': '',
            '微信小文件发送平均时延(s)': '',
            '微信小文件发送平均速率(Mbps)': '',
            '商店大app平均时延(s)': '',
            '商店大app平均速率(Mbps)': '',
            '商店小app平均时延(s)': '',
            '商店小app平均速率(Mbps)': ''
        })

    df_summary_data_voice = pd.DataFrame(summary_data)

    # 输出文件名
    version = 'V2.5.7'
    timestamp = datetime.now().strftime('%Y%m%d-%H%M')
    out = f'联通/语音VQI输出_{version}_{timestamp}.xlsx'

    log("写入输出文件...", 70)
    with pd.ExcelWriter(out, engine='openpyxl') as w:
        from openpyxl.styles import Alignment, PatternFill, Border, Side, Font as OpFont
        from openpyxl.utils import get_column_letter as _gcl
        from openpyxl.comments import Comment as _Comment
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        hdr_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

        # 写入统计sheet（小数据量）
        df_voice_out.to_excel(w, sheet_name='语音指标统计', index=False)
        df_vqi_out.to_excel(w, sheet_name='MOS3.5计算明细', index=False)
        df_summary_data_voice.to_excel(w, sheet_name='汇总-数据和语音', index=False)
        df_file_stats.to_excel(w, sheet_name='按文件名', index=False)
        df_operator_stats.to_excel(w, sheet_name='按运营商', index=False)

        # ===== 原始数据sheet（大数据量，只输出前20000行）=====
        if merge_raw:
            log("  写入原始数据...", 75)
            df_raw_write = df_raw_out.copy()
            if len(df_raw_write) > 20000:
                df_raw_write = df_raw_write.head(20000)
                log(f"  语音原始数据超过20000行，只输出前20000行", 76)
            df_raw_write.to_excel(w, sheet_name='语音业务-原始文件', index=False)
            ws_raw = w.sheets['语音业务-原始文件']
            for cell in ws_raw[1]:
                cell.alignment = center; cell.fill = hdr_fill; cell.font = OpFont(bold=True)
            ws_raw.freeze_panes = 'A2'
            if ws_raw.max_column >= 1:
                ws_raw.column_dimensions['A'].width = 22
            if ws_raw.max_column >= 2:
                ws_raw.column_dimensions['B'].width = 22

            if len(df_data_raw) > 0:
                df_data_write = df_data_raw.copy()
                for col in ['_t', '_dl_rlc', '_ul_rlc']:
                    if col in df_data_write.columns:
                        df_data_write = df_data_write.drop(col, axis=1)
                if len(df_data_write) > 20000:
                    df_data_write = df_data_write.head(20000)
                    log(f"  数据原始数据超过20000行，只输出前20000行", 77)
                df_data_write.to_excel(w, sheet_name='数据业务-原始文件', index=False)
                ws_data = w.sheets['数据业务-原始文件']
                for cell in ws_data[1]:
                    cell.alignment = center; cell.fill = hdr_fill; cell.font = OpFont(bold=True)
                ws_data.freeze_panes = 'A2'

        # ===== 统计sheet格式化（数据量小，全量格式化）=====
        log("  格式化统计sheet...", 82)

        ws_voice = w.sheets['语音指标统计']
        for cell in ws_voice[1]:
            cell.alignment = center; cell.fill = hdr_fill; cell.font = OpFont(bold=True); cell.border = border
        for ri in range(2, ws_voice.max_row + 1):
            for ci in range(1, ws_voice.max_column + 1):
                cell = ws_voice.cell(row=ri, column=ci)
                cell.alignment = center; cell.border = border
                if ci == 5:
                    cv = cell.value
                    if cv is None or str(cv) == '' or str(cv) == 'nan':
                        cell.comment = _Comment("无RSP_180", "Tool")
        for col_idx in range(1, ws_voice.max_column + 1):
            ws_voice.column_dimensions[_gcl(col_idx)].width = 20
        ws_voice.freeze_panes = 'A2'

        ws_vqi = w.sheets['MOS3.5计算明细']
        for cell in ws_vqi[1]:
            cell.alignment = center; cell.fill = hdr_fill; cell.font = OpFont(bold=True); cell.border = border
        for ri in range(2, ws_vqi.max_row + 1):
            for ci in range(1, ws_vqi.max_column + 1):
                cell = ws_vqi.cell(row=ri, column=ci)
                cell.alignment = center; cell.border = border
                if ri == ws_vqi.max_row:
                    cell.fill = green_fill
        for col_idx in range(1, ws_vqi.max_column + 1):
            ws_vqi.column_dimensions[_gcl(col_idx)].width = 22
        ws_vqi.freeze_panes = 'A2'

        ws_summary = w.sheets['汇总-数据和语音']
        for cell in ws_summary[1]:
            cell.alignment = center; cell.fill = hdr_fill; cell.font = OpFont(bold=True); cell.border = border
        for ri in range(2, ws_summary.max_row + 1):
            for ci in range(1, ws_summary.max_column + 1):
                cell = ws_summary.cell(row=ri, column=ci)
                cell.alignment = center; cell.border = border
                if ri == 2:
                    cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        for col_idx in range(1, ws_summary.max_column + 1):
            ws_summary.column_dimensions[_gcl(col_idx)].width = 18
        ws_summary.freeze_panes = 'A2'
        for ri in range(3, ws_summary.max_row + 1):
            for ci in range(2, ws_summary.max_column + 1):
                cell = ws_summary.cell(row=ri, column=ci)
                cv = cell.value
                if cv is not None and isinstance(cv, (int, float)):
                    col_name = ws_summary.cell(row=1, column=ci).value or f'列{ci}'
                    txt = f"公式: {col_name}\n计算过程: 从原始数据统计\n\n数据来源: {col_name} 列"
                    c = _Comment(txt, "Tool"); c.width = 400; c.height = 150
                    cell.comment = c

        ws_file = w.sheets['按文件名']
        for cell in ws_file[1]:
            cell.alignment = center; cell.fill = hdr_fill; cell.font = OpFont(bold=True); cell.border = border
        for ri in range(2, ws_file.max_row + 1):
            for ci in range(1, ws_file.max_column + 1):
                cell = ws_file.cell(row=ri, column=ci)
                cell.alignment = center; cell.border = border
        for col_idx in range(1, ws_file.max_column + 1):
            ws_file.column_dimensions[_gcl(col_idx)].width = 20
        ws_file.freeze_panes = 'A2'
        for ri in range(2, ws_file.max_row + 1):
            for ci in range(2, ws_file.max_column + 1):
                cell = ws_file.cell(row=ri, column=ci)
                cv = cell.value
                if cv is not None and isinstance(cv, (int, float)):
                    col_name = ws_file.cell(row=1, column=ci).value or f'列{ci}'
                    txt = f"公式: {col_name}\n计算过程: 从原始数据统计\n\n数据来源: {col_name} 列"
                    c = _Comment(txt, "Tool"); c.width = 400; c.height = 150
                    cell.comment = c

        ws_op = w.sheets['按运营商']
        for cell in ws_op[1]:
            cell.alignment = center; cell.fill = hdr_fill; cell.font = OpFont(bold=True); cell.border = border
        for ri in range(2, ws_op.max_row + 1):
            for ci in range(1, ws_op.max_column + 1):
                cell = ws_op.cell(row=ri, column=ci)
                cell.alignment = center; cell.border = border
        for col_idx in range(1, ws_op.max_column + 1):
            ws_op.column_dimensions[_gcl(col_idx)].width = 20
        ws_op.freeze_panes = 'A2'
        for ri in range(2, ws_op.max_row + 1):
            for ci in range(2, ws_op.max_column + 1):
                cell = ws_op.cell(row=ri, column=ci)
                cv = cell.value
                if cv is not None and isinstance(cv, (int, float)):
                    col_name = ws_op.cell(row=1, column=ci).value or f'列{ci}'
                    txt = f"公式: {col_name}\n计算过程: 从原始数据统计\n\n数据来源: {col_name} 列"
                    c = _Comment(txt, "Tool"); c.width = 400; c.height = 150
                    cell.comment = c

    log(f"完成! {out}", 100)
    # 清理解压临时目录
    for _d in _tmp_dirs:
        try: shutil.rmtree(_d)
        except Exception: pass
    return out


# ===== GUI 包装 =====
try:
    from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QPushButton, QLabel, QFileDialog, QMessageBox, QProgressBar, QTextEdit,
        QGroupBox, QCheckBox, QSpinBox, QDialog, QSplitter, QGridLayout,
        QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView, QTimeEdit, QComboBox)
    from PySide6.QtCore import Qt, QThread, Signal, Slot, QTime, QEvent
    from PySide6.QtGui import QFont
    GUI_OK = True
except:
    GUI_OK = False


if GUI_OK:
    class WorkerThread(QThread):
        progress = Signal(str)
        progress_pct = Signal(int)
        done = Signal(str)

        def __init__(self, files, qci_list, dl_clip, ul_clip, base_file=None, time_filter=None, ftp_duration=20, merge_raw=True):
            super().__init__()
            self.files = files; self.qci_list = qci_list; self.dl_clip = dl_clip; self.ul_clip = ul_clip
            self.base_file = base_file; self.time_filter = time_filter; self.ftp_duration = ftp_duration
            self.merge_raw = merge_raw
            self._cancel = False; self._pct = 0

        def cancel(self): self._cancel = True

        def run(self):
            try:
                out = process(self.files, self.qci_list, self.dl_clip, self.ul_clip,
                              callback=self.l, cancel_check=lambda: self._cancel,
                              progress_cb=lambda p: self.progress_pct.emit(int(p)),
                              base_file=self.base_file, time_filter=self.time_filter,
                              ftp_duration=self.ftp_duration, merge_raw=self.merge_raw)
                self.done.emit('' if self._cancel else out)
            except KeyboardInterrupt:
                self.done.emit('')

        def l(self, msg):
            self.progress.emit(msg)

    class VqiWorkerThread(QThread):
        progress = Signal(str)
        progress_pct = Signal(int)
        done = Signal(str)

        def __init__(self, files, time_filter=None, merge_raw=True):
            super().__init__()
            self.files = files
            self.time_filter = time_filter
            self._cancel = False
            self.merge_raw = merge_raw

        def cancel(self): self._cancel = True

        def run(self):
            try:
                out = process_vqi(self.files,
                                  callback=self.l, cancel_check=lambda: self._cancel,
                                  progress_cb=lambda p: self.progress_pct.emit(int(p)),
                                  time_filter=self.time_filter,
                                  merge_raw=self.merge_raw)
                self.done.emit('' if self._cancel else out)
            except KeyboardInterrupt:
                self.done.emit('')

        def l(self, msg):
            self.progress.emit(msg)

    class MainWindow(QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("用户级数据、语音跟踪统计工具 V2.5.7")
            self.files = []; self.last_out = None
            self.base_file = None  # 用户可选的基准文件
            # VQI相关实例变量
            self.vqi_files = []
            self.vqi_base_file = None
            self.vqi_last_out = None
            scr = QApplication.primaryScreen()
            self.scr_w = scr.geometry().width() if scr else 1440
            sh = scr.geometry().height() if scr else 1080
            self.fs = max(11, int(round(11 * sh / 1080)))
            self.setMinimumSize(1200, 750)
            self._build()
            self.showMaximized()

        def _font(self, big=0):
            return QFont("PingFang SC", self.fs + big, QFont.Bold if big >= 2 else QFont.Normal)

        def eventFilter(self, obj, event):
            # 让QTabWidget和QSplitter上的拖放事件也传递到MainWindow
            if event.type() == QEvent.DragEnter:
                self.dragEnterEvent(event)
                return True
            elif event.type() == QEvent.Drop:
                self.dropEvent(event)
                return True
            return super().eventFilter(obj, event)

        def dragEnterEvent(self, event):
            # 始终接受拖入事件，让MainWindow统一处理
            if event.mimeData().hasUrls():
                event.acceptProposedAction()
            else:
                event.ignore()

        def dropEvent(self, event):
            new = []
            for u in event.mimeData().urls():
                f = u.toLocalFile()
                if os.path.isdir(f):
                    for root, dirs, files in os.walk(f):
                        for fn in files:
                            low = fn.lower()
                            if low.endswith(('.xlsx', '.xls', '.zip', '.csv', '.mmf', '.tmf')):
                                new.append(os.path.join(root, fn))
                else:
                    low = f.lower()
                    if low.endswith(('.xlsx', '.xls', '.zip', '.csv', '.mmf', '.tmf')):
                        new.append(f)
            if new:
                # 检查当前激活的是哪个标签
                current_tab = self.centralWidget().currentIndex()
                if current_tab == 1:  # VQI标签
                    self.vqi_files = list(dict.fromkeys((self.vqi_files or []) + new))
                    display_lines = [os.path.basename(f) for f in self.vqi_files]
                    self.vqi_te.setText("\n".join(display_lines))
                    self.vqi_br.setEnabled(True)
                    if hasattr(self, 'vqi_lt'):
                        self.vqi_lt.append(f"拖入 {len(new)} 个文件到VQI，当前共 {len(self.vqi_files)} 个")
                else:  # 速率统计标签
                    self.files = list(dict.fromkeys((self.files or []) + new))
                    display_lines = [os.path.basename(f) for f in self.files]
                    self.te.setText("\n".join(display_lines))
                    self.br.setEnabled(True)
                    self.lt.append(f"拖入 {len(new)} 个文件，当前共 {len(self.files)} 个")
            event.acceptProposedAction()

        # 解决拖放不起作用的问题：用事件过滤器让中央控件转发拖放事件
        def disable_child_drops(self, widget):
            for child in widget.findChildren(QWidget):
                child.setAcceptDrops(False)
                child.installEventFilter(self)
                self.disable_child_drops(child)

        def _build(self):
            self.setAcceptDrops(True)
            tabs = QTabWidget()
            tabs.addTab(self._build_rate_tab(), "5G用户级公共监控速率统计工具")
            tabs.addTab(self._build_vqi_tab(), "5G虚拟用户跟踪语音VQI工具")
            self.setCentralWidget(tabs)
            # 禁用所有子控件的拖放
            self.disable_child_drops(self)

        def _build_rate_tab(self):
            sp = QSplitter(Qt.Horizontal)
            # ===== 左：上设置 + 下运行 =====
            left = QWidget(); ll = QVBoxLayout(left)
            ll.setContentsMargins(4, 4, 4, 4)
            ll.setSpacing(4)
            title = QLabel("5G用户级公共监控速率统计工具")
            title.setFont(self._font(2)); title.setAlignment(Qt.AlignCenter); ll.addWidget(title)
            # 输入文件（选择/清空紧挨着，拖入框明显、高度翻倍）
            gf = QGroupBox("输入文件"); gfl = QVBoxLayout(gf)
            gfl.setContentsMargins(4, 4, 4, 4); gfl.setSpacing(2)
            # 选择/清空按钮紧挨着（不加stretch）
            bf = QHBoxLayout(); bf.setSpacing(4)
            self.bs = QPushButton("选择MMF文件"); self.bs.clicked.connect(self.sf)
            self.bc = QPushButton("清空"); self.bc.clicked.connect(self.cf)
            bf.addWidget(self.bs); bf.addWidget(self.bc); gfl.addLayout(bf)
            # 拖入框：只显示文件名，高度翻倍(300)，边框明显
            self.te = QTextEdit(); self.te.setReadOnly(True); self.te.setMaximumHeight(300)
            self.te.setPlaceholderText("可以将文件或是文件夹拖入，支持各类文件类型")
            self.te.setStyleSheet("QTextEdit{font-size:12px;background:#fff;border:2px solid #666;border-radius:4px;padding:4px;}")
            gfl.addWidget(self.te); ll.addWidget(gf)
            # 参数设置（紧凑，紧接输入文件）
            gp = QGroupBox("参数设置"); gpl = QGridLayout(gp)
            gpl.setContentsMargins(4, 4, 4, 4); gpl.setSpacing(2)
            gpl.addWidget(QLabel("QCI:"), 0, 0)
            self.cb6 = QCheckBox("6"); self.cb7 = QCheckBox("7")
            self.cb7.setChecked(True)  # 默认7
            self.cb6.toggled.connect(lambda c: self.cb7.setChecked(not c) if c else None)
            self.cb7.toggled.connect(lambda c: self.cb6.setChecked(not c) if c else None)
            gpl.addWidget(self.cb6, 0, 1); gpl.addWidget(self.cb7, 0, 2)
            gpl.addWidget(QLabel("下行削峰(Mbps):"), 1, 0)
            self.sd = QSpinBox(); self.sd.setRange(100, 5000); self.sd.setValue(1000); gpl.addWidget(self.sd, 1, 1, 1, 3)
            gpl.addWidget(QLabel("上行削峰(Mbps):"), 2, 0)
            self.su = QSpinBox(); self.su.setRange(50, 1000); self.su.setValue(200); gpl.addWidget(self.su, 2, 1, 1, 3)
            # 时间段限制（第3行）
            self.cb_time = QCheckBox(); gpl.addWidget(self.cb_time, 3, 0)
            gpl.addWidget(QLabel("开始时间:"), 3, 1)
            self.time_start = QTimeEdit(); self.time_start.setDisplayFormat("HH:mm:ss"); self.time_start.setTime(QTime(0, 0, 0))
            gpl.addWidget(self.time_start, 3, 2)
            gpl.addWidget(QLabel("结束时间:"), 4, 1)
            self.time_end = QTimeEdit(); self.time_end.setDisplayFormat("HH:mm:ss"); self.time_end.setTime(QTime(23, 59, 59))
            gpl.addWidget(self.time_end, 4, 2)
            # FTP时长（第5行）
            gpl.addWidget(QLabel("FTP时长:"), 5, 0)
            self.ftp_duration_combo = QComboBox()
            self.ftp_duration_combo.addItems(["10秒", "20秒", "30秒"])
            self.ftp_duration_combo.setCurrentIndex(1)  # 默认20秒
            self.cb_merge_raw = QCheckBox("合并原始数据到结果文件")
            self.cb_merge_raw.setChecked(True)
            gpl.addWidget(self.cb_merge_raw, 6, 0, 1, 3)
            gpl.addWidget(self.ftp_duration_combo, 5, 1, 1, 2)
            ll.addWidget(gp)
            # 运行（与参数设置紧接）
            gh = QGroupBox("运行"); ghl = QHBoxLayout(gh)
            ghl.setContentsMargins(4, 4, 4, 4); ghl.setSpacing(4)
            self.br = QPushButton("开始处理"); self.br.clicked.connect(self.run); self.br.setEnabled(False)
            self.bcancel = QPushButton("取消"); self.bcancel.clicked.connect(self.cancel); self.bcancel.setEnabled(False)
            self.babout = QPushButton("关于"); self.babout.clicked.connect(self.about)
            ghl.addWidget(self.br); ghl.addWidget(self.bcancel); ghl.addWidget(self.babout); ll.addWidget(gh)
            # 运行结果（打开按钮右下角，日志左下方剩余空间）
            gr = QGroupBox("运行结果"); grl = QVBoxLayout(gr)
            grl.setContentsMargins(4, 4, 4, 4); grl.setSpacing(2)
            # 进度条 + 打开按钮同行（按钮放右下）
            pb_open = QHBoxLayout()
            self.pb = QProgressBar(); pb_open.addWidget(self.pb)
            self.bopen = QPushButton("打开输出文件"); self.bopen.clicked.connect(self.open_out); self.bopen.setEnabled(False)
            pb_open.addWidget(self.bopen); grl.addLayout(pb_open)
            # 日志占剩余空间
            self.lt = QTextEdit(); self.lt.setReadOnly(True); self.lt.setStyleSheet("QTextEdit{font-size:11px;}")
            grl.addWidget(self.lt, 1); ll.addWidget(gr)
            left.setFont(self._font())
            # ===== 右：结果内嵌查看(子标签+表格) =====
            right = QWidget(); rl = QVBoxLayout(right)
            rh = QLabel("结果数据（内嵌查看，默认筛选A=1）"); rh.setFont(self._font(1)); rl.addWidget(rh)
            self.result_tabs = QTabWidget()
            self.tbl_detail = QTableWidget(); self.tbl_detail.setSelectionBehavior(QTableWidget.SelectItems)
            self.tbl_summary = QTableWidget(); self.tbl_compare = QTableWidget()
            self.result_tabs.addTab(self.tbl_detail, "数据业务-详细过程")
            self.result_tabs.addTab(self.tbl_summary, "汇总")
            self.result_tabs.addTab(self.tbl_compare, "数据业务-对比")
            rl.addWidget(self.result_tabs)
            right.setFont(self._font())
            sp.addWidget(left); sp.addWidget(right)
            sp.setStretchFactor(0, 1); sp.setStretchFactor(1, 5)  # 左1/6，右5/6
            sp.setSizes([int(self.scr_w * 0.167), int(self.scr_w * 0.833)])
            return sp

        def _build_vqi_tab(self):
            sp = QSplitter(Qt.Horizontal)
            # ===== 左：上设置 + 下运行 =====
            left = QWidget(); ll = QVBoxLayout(left)
            ll.setContentsMargins(4, 4, 4, 4)
            ll.setSpacing(4)
            title = QLabel("5G虚拟用户跟踪语音VQI工具")
            title.setFont(self._font(2)); title.setAlignment(Qt.AlignCenter); ll.addWidget(title)
            # 输入文件（选择/清空紧挨着，拖入框明显、高度翻倍）
            gf = QGroupBox("输入文件"); gfl = QVBoxLayout(gf)
            gfl.setContentsMargins(4, 4, 4, 4); gfl.setSpacing(2)
            # 选择/清空按钮紧挨着（不加stretch）
            bf = QHBoxLayout(); bf.setSpacing(4)
            self.vqi_bs = QPushButton("选择MMF文件"); self.vqi_bs.clicked.connect(self.vqi_sf)
            self.vqi_bc = QPushButton("清空"); self.vqi_bc.clicked.connect(self.vqi_cf)
            bf.addWidget(self.vqi_bs); bf.addWidget(self.vqi_bc); gfl.addLayout(bf)
            # 拖入框：只显示文件名，高度翻倍(300)，边框明显
            self.vqi_te = QTextEdit(); self.vqi_te.setReadOnly(True); self.vqi_te.setMaximumHeight(300)
            self.vqi_te.setPlaceholderText("可以将文件或是文件夹拖入，支持各类文件类型")
            self.vqi_te.setStyleSheet("QTextEdit{font-size:12px;background:#fff;border:2px solid #666;border-radius:4px;padding:4px;}")
            gfl.addWidget(self.vqi_te); ll.addWidget(gf)
            # 参数设置（紧凑，紧接输入文件）
            gp = QGroupBox("参数设置"); gpl = QGridLayout(gp)
            gpl.setContentsMargins(4, 4, 4, 4); gpl.setSpacing(2)
            # 时间段限制
            self.vqi_cb_time = QCheckBox(); gpl.addWidget(self.vqi_cb_time, 0, 0)
            gpl.addWidget(QLabel("开始时间:"), 0, 1)
            self.vqi_time_start = QTimeEdit(); self.vqi_time_start.setDisplayFormat("HH:mm:ss"); self.vqi_time_start.setTime(QTime(0, 0, 0))
            gpl.addWidget(self.vqi_time_start, 0, 2)
            gpl.addWidget(QLabel("结束时间:"), 1, 1)
            self.vqi_time_end = QTimeEdit(); self.vqi_time_end.setDisplayFormat("HH:mm:ss"); self.vqi_time_end.setTime(QTime(23, 59, 59))
            gpl.addWidget(self.vqi_time_end, 1, 2)

            self.vqi_cb_merge_raw = QCheckBox("合并原始数据到结果文件")
            self.vqi_cb_merge_raw.setChecked(True)
            gpl.addWidget(self.vqi_cb_merge_raw, 2, 0, 1, 3)
            ll.addWidget(gp)
            # 运行（与参数设置紧接）
            gh = QGroupBox("运行"); ghl = QHBoxLayout(gh)
            ghl.setContentsMargins(4, 4, 4, 4); ghl.setSpacing(4)
            self.vqi_br = QPushButton("开始处理"); self.vqi_br.clicked.connect(self.vqi_run); self.vqi_br.setEnabled(False)
            self.vqi_bcancel = QPushButton("取消"); self.vqi_bcancel.clicked.connect(self.vqi_cancel); self.vqi_bcancel.setEnabled(False)
            self.vqi_babout = QPushButton("关于"); self.vqi_babout.clicked.connect(self.about)
            ghl.addWidget(self.vqi_br); ghl.addWidget(self.vqi_bcancel); ghl.addWidget(self.vqi_babout); ll.addWidget(gh)
            # 运行结果（打开按钮右下角，日志左下方剩余空间）
            gr = QGroupBox("运行结果"); grl = QVBoxLayout(gr)
            grl.setContentsMargins(4, 4, 4, 4); grl.setSpacing(2)
            # 进度条 + 打开按钮同行（按钮放右下）
            pb_open = QHBoxLayout()
            self.vqi_pb = QProgressBar(); pb_open.addWidget(self.vqi_pb)
            self.vqi_bopen = QPushButton("打开输出文件"); self.vqi_bopen.clicked.connect(self.vqi_open_out); self.vqi_bopen.setEnabled(False)
            pb_open.addWidget(self.vqi_bopen); grl.addLayout(pb_open)
            # 日志占剩余空间
            self.vqi_lt = QTextEdit(); self.vqi_lt.setReadOnly(True); self.vqi_lt.setStyleSheet("QTextEdit{font-size:11px;}")
            grl.addWidget(self.vqi_lt, 1); ll.addWidget(gr)
            left.setFont(self._font())
            # ===== 右：结果内嵌查看(子标签+表格) =====
            right = QWidget(); rl = QVBoxLayout(right)
            rh = QLabel("结果数据（内嵌查看）"); rh.setFont(self._font(1)); rl.addWidget(rh)
            self.vqi_result_tabs = QTabWidget()
            self.vqi_tbl_raw = QTableWidget(); self.vqi_tbl_raw.setSelectionBehavior(QTableWidget.SelectItems)
            self.vqi_tbl_voice = QTableWidget(); self.vqi_tbl_voice.setSelectionBehavior(QTableWidget.SelectItems)
            self.vqi_tbl_mos = QTableWidget(); self.vqi_tbl_mos.setSelectionBehavior(QTableWidget.SelectItems)
            self.vqi_result_tabs.addTab(self.vqi_tbl_raw, "原始数据")
            self.vqi_result_tabs.addTab(self.vqi_tbl_voice, "语音指标统计")
            self.vqi_result_tabs.addTab(self.vqi_tbl_mos, "MOS3.5计算明细")
            self.vqi_tbl_data_raw = QTableWidget()
            self.vqi_tbl_data_raw.setSelectionBehavior(QTableWidget.SelectItems)
            self.vqi_result_tabs.addTab(self.vqi_tbl_data_raw, "数据业务-原始文件")
            rl.addWidget(self.vqi_result_tabs)
            right.setFont(self._font())
            sp.addWidget(left); sp.addWidget(right)
            sp.setStretchFactor(0, 1); sp.setStretchFactor(1, 5)  # 左1/6，右5/6
            sp.setSizes([int(self.scr_w * 0.167), int(self.scr_w * 0.833)])
            return sp

        # ===== VQI相关方法 =====
        def vqi_sf(self):
            fs, _ = QFileDialog.getOpenFileNames(self, "选择文件", "", "Excel/CSV/压缩 (*.xlsx *.xls *.csv *.zip);;All (*)")
            if fs:
                self.vqi_files = list(dict.fromkeys((self.vqi_files or []) + fs))
                self.vqi_te.setText("\n".join(os.path.basename(f) for f in self.vqi_files))
                self.vqi_br.setEnabled(True)

        def vqi_cf(self):
            self.vqi_files = []
            self.vqi_te.clear()
            self.vqi_br.setEnabled(False)

        def vqi_run(self):
            if not self.vqi_files:
                QMessageBox.warning(self, "提示", "请先选择MMF文件")
                return

            # 时间段过滤
            time_filter = None
            if self.vqi_cb_time.isChecked():
                t_start = self.vqi_time_start.time()
                t_end = self.vqi_time_end.time()
                time_filter = (True, t_start, t_end)

            self.vqi_br.setEnabled(False); self.vqi_bcancel.setEnabled(True)
            self.vqi_pb.setValue(0)
            self.vqi_lt.append(f"开始处理...")
            if time_filter:
                self.vqi_lt.append(f"时间段过滤: {t_start.toString('HH:mm:ss')} ~ {t_end.toString('HH:mm:ss')}")
            self.vqi_thread = VqiWorkerThread(self.vqi_files, time_filter=time_filter, merge_raw=self.vqi_cb_merge_raw.isChecked())
            self.vqi_thread.progress.connect(self.vqi_upd)
            self.vqi_thread.progress_pct.connect(lambda p: self.vqi_pb.setValue(p))
            self.vqi_thread.done.connect(self.vqi_dn)
            self.vqi_thread.start()

        def vqi_cancel(self):
            if hasattr(self, 'vqi_thread') and self.vqi_thread.isRunning():
                self.vqi_thread.cancel()
                self.vqi_lt.append("已请求取消，将在当前阶段结束后停止...")

        @Slot(str)
        def vqi_upd(self, msg):
            self.vqi_lt.append(msg)

        @Slot(str)
        def vqi_dn(self, path):
            self.vqi_bcancel.setEnabled(False)
            if path:
                self.vqi_pb.setValue(100)
                self.vqi_last_out = path
                self.vqi_bopen.setEnabled(True)
                self.vqi_lt.append(f"完成! {path}")
                self.show_vqi_result(path)
            else:
                self.vqi_pb.setValue(0)
                self.vqi_lt.append("已取消。")
            self.vqi_br.setEnabled(True)

        def vqi_open_out(self):
            if self.vqi_last_out:
                os.system(f'open "{self.vqi_last_out}"')

        def show_vqi_result(self, path):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, data_only=True)
                sheet_list = wb.sheetnames
                # 显示语音原始数据sheet（如果存在）
                if '语音业务-原始文件' in sheet_list:
                    self._fill_vqi_table(self.vqi_tbl_raw, wb['语音业务-原始文件'], path=path, sheet='语音业务-原始文件')
                else:
                    self.vqi_tbl_raw.setRowCount(0); self.vqi_tbl_raw.setColumnCount(0)
                # 显示语音指标统计
                if '语音指标统计' in sheet_list:
                    self._fill_vqi_table(self.vqi_tbl_voice, wb['语音指标统计'], path=path, sheet='语音指标统计')
                else:
                    self.vqi_tbl_voice.setRowCount(0); self.vqi_tbl_voice.setColumnCount(0)
                # 显示MOS3.5计算明细
                if 'MOS3.5计算明细' in sheet_list:
                    self._fill_vqi_table(self.vqi_tbl_mos, wb['MOS3.5计算明细'], path=path, sheet='MOS3.5计算明细')
                elif '汇总-数据和语音' in sheet_list:
                    self._fill_vqi_table(self.vqi_tbl_mos, wb['汇总-数据和语音'], path=path, sheet='汇总-数据和语音')
                else:
                    self.vqi_tbl_mos.setRowCount(0); self.vqi_tbl_mos.setColumnCount(0)
                cnt = len(sheet_list)
                self.vqi_lt.append(f"结果已内嵌显示（{cnt}个Sheet）")
            except Exception as e:
                self.vqi_lt.append(f"内嵌显示失败: {e}")

        def _fill_vqi_table(self, tbl, ws, path, sheet):
            from PySide6.QtGui import QColor
            tbl.clearContents(); tbl.setRowCount(0)
            rows = list(ws.iter_rows())
            if not rows: return
            headers = [str(c.value) if c.value is not None else '' for c in rows[0]]
            tbl.setColumnCount(len(headers)); tbl.setHorizontalHeaderLabels(headers)
            r_idx = 0
            for row in rows[1:]:
                tbl.insertRow(r_idx)
                for ci, cell in enumerate(row):
                    v = cell.value
                    item = QTableWidgetItem('' if v is None else str(v))
                    # 颜色(复制xlsx单元格fill)
                    try:
                        fill = cell.fill
                        if fill and fill.start_color and fill.start_color.rgb:
                            rgb = str(fill.start_color.rgb)
                            if rgb not in ('00000000', '0'):
                                item.setBackground(QColor('#' + rgb[-6:]))
                    except Exception: pass
                    tbl.setItem(r_idx, ci, item)
                r_idx += 1
            tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            tbl.setSortingEnabled(False)

        # ===== 速率统计相关方法 =====
        def sf(self):
            fs, _ = QFileDialog.getOpenFileNames(self, "选择文件", "", "所有支持文件 (*.xlsx *.xls *.csv *.zip *.mmf *.tmf);;Excel/CSV (*.xlsx *.xls *.csv);;All (*)")
            if fs:
                for f in fs:
                    fn = os.path.basename(f).lower()
                    # 判断是基准文件还是MMF文件
                    if '基准' in fn or '呼叫详情' in fn or '整理' in fn:
                        self.base_file = f
                        self.lt.append(f"已选择基准文件: {os.path.basename(f)}")
                    else:
                        self.files = list(dict.fromkeys((self.files or []) + [f]))
                if self.files:
                    self.te.setText("\n".join(os.path.basename(f) for f in self.files))
                    self.br.setEnabled(True)
                if self.base_file:
                    self.te.append(f"\n[基准] {os.path.basename(self.base_file)}")

        def cf(self):
            self.files = []; self.base_file = None
            self.te.clear(); self.br.setEnabled(False)

        def run(self):
            if not self.files: QMessageBox.warning(self, "提示", "请先选择MMF文件"); return
            qci_list = []
            if self.cb6.isChecked(): qci_list.append(6)
            if self.cb7.isChecked(): qci_list.append(7)
            if not qci_list: qci_list = [7]

            # 时间段过滤
            time_filter = None
            if self.cb_time.isChecked():
                t_start = self.time_start.time()
                t_end = self.time_end.time()
                time_filter = (True, t_start, t_end)

            # FTP时长
            ftp_duration_text = self.ftp_duration_combo.currentText()
            ftp_duration = int(ftp_duration_text.replace('秒', ''))

            self.br.setEnabled(False); self.bcancel.setEnabled(True)
            self.pb.setValue(0); self.lt.append(f"开始处理... QCI={qci_list}, FTP时长={ftp_duration}秒")
            if time_filter:
                self.lt.append(f"时间段过滤: {t_start.toString('HH:mm:ss')} ~ {t_end.toString('HH:mm:ss')}")
            self.thread = WorkerThread(self.files, qci_list, self.sd.value(), self.su.value(),
                                       base_file=self.base_file, time_filter=time_filter,
                                       ftp_duration=ftp_duration,
                                       merge_raw=self.cb_merge_raw.isChecked())
            self.thread.progress.connect(self.upd)
            self.thread.progress_pct.connect(lambda p: self.pb.setValue(p))
            self.thread.done.connect(self.dn); self.thread.start()

        def cancel(self):
            if hasattr(self, 'thread') and self.thread.isRunning():
                self.thread.cancel(); self.lt.append("已请求取消，将在当前阶段结束后停止...")

        @Slot(str)
        def upd(self, msg): self.lt.append(msg)

        @Slot(str)
        def dn(self, path):
            self.bcancel.setEnabled(False)
            if path:
                self.pb.setValue(100)
                self.last_out = path; self.bopen.setEnabled(True)
                self.lt.append(f"完成! {path}")
                self.show_result(path)   # 内嵌显示结果
            else:
                self.pb.setValue(0); self.lt.append("已取消。")
            self.br.setEnabled(True)

        def open_out(self):
            if self.last_out: os.system(f'open "{self.last_out}"')

        # ===== 内嵌结果查看(基础版：显示+默认A=1筛选+颜色) =====
        def show_result(self, path):
            try:
                import openpyxl
                from PySide6.QtGui import QColor
                wb = openpyxl.load_workbook(path, data_only=True)
                self._fill_table(self.tbl_detail, wb['数据业务-详细过程'], filter_a1=True, path=path, sheet='数据业务-详细过程')
                self._fill_table(self.tbl_summary, wb['汇总'], filter_a1=False, path=path, sheet='汇总')
                self._fill_table(self.tbl_compare, wb['数据业务-对比'], filter_a1=False, path=path, sheet='数据业务-对比')
                self.lt.append("结果已内嵌显示（详细过程默认筛选A=1）")
            except Exception as e:
                self.lt.append(f"内嵌显示失败: {e}")

        def _fill_table(self, tbl, ws, filter_a1, path, sheet):
            from PySide6.QtGui import QColor
            tbl.clearContents(); tbl.setRowCount(0)
            rows = list(ws.iter_rows())
            if not rows: return
            headers = [str(c.value) if c.value is not None else '' for c in rows[0]]
            tbl.setColumnCount(len(headers)); tbl.setHorizontalHeaderLabels(headers)
            r_idx = 0
            for ri, row in enumerate(rows[1:], start=2):
                if filter_a1:
                    a = row[0].value if row else None
                    if a is None or str(a) != '1': continue
                tbl.insertRow(r_idx)
                for ci, cell in enumerate(row):
                    v = cell.value
                    item = QTableWidgetItem('' if v is None else str(v))
                    # 颜色(复制xlsx单元格fill)
                    try:
                        fill = cell.fill
                        if fill and fill.start_color and fill.start_color.rgb:
                            rgb = str(fill.start_color.rgb)
                            if rgb not in ('00000000', '0'):
                                item.setBackground(QColor('#' + rgb[-6:]))
                    except Exception: pass
                    # 批注(setToolTip, hover显示; 点击显示下一轮完善)
                    try:
                        if cell.comment:
                            item.setToolTip(cell.comment.text)
                    except Exception: pass
                    tbl.setItem(r_idx, ci, item)
                r_idx += 1
            tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            tbl.setSortingEnabled(False)

        def about(self):
            d = QDialog(self); d.setWindowTitle("关于")
            d.resize(560, 600)
            v = QVBoxLayout(d)
            te = QTextEdit(); te.setReadOnly(True)
            te.setHtml(
                "<h3 style='text-align:center'>用户级数据、语音跟踪统计工具</h3>"
                "<p><b>版本:</b> V2.5.7 &nbsp; <b>开发者:</b> 孙晓军 &nbsp; "
                "<b>联系方式:</b> 317827@qq.com</p><hr/>"
                "<h4>更新记录</h4>"
                "<p><b>2026-07-18</b></p>"
                "<ul>"
"<li><b>V2.5.7</b> 修复按文件名统计表显示0的问题：1)为每条voice_record/vqi_record添加_source_file字段；2)使用voice_file_info字典记录每个文件的基本信息；3)按_source_file字段精确匹配统计每个文件的轮次/MOS指标</li>\n"
"<li><b>V2.5.6</b> 新增\"合并原始数据到结果文件\"复选框(速率统计/VQI标签各一),默认勾选</li>\n"
"<li>新增批量计算批注:汇总-数据和语音、按文件名、按运营商Sheet的数值单元格自动添加计算批注</li>\n"
"<li>数据文件处理:无MessageType列的数据文件(mmf)不再跳过,读取并计算下行/上行速率指标</li>\n"
"<li>VQI输出Sheet优化:新增\"数据业务-原始文件\"Sheet,按文件名/按运营商表含完整数据业务指标</li>\n"
"<li>性能优化:文件类型检测只读前3行(nrows=3);缓存文件列表避免重复读取</li>\n"
                "<li><b>V2.5.5</b> 修复文件类型显示为unknown问题：_detect_file_type返回语音/数据而非unknown</li>"
                "<li>按文件名表新增指标列：通话轮次、平均呼叫建立时长(s)、优质通话占比(MOS3.5)</li>"
                "<li>按运营商表新增完整指标列：下行前10%峰值速率等12列（当前留空）</li>"
                "<li>语音指标统计sheet中对空呼叫建立时长单元格加批注无RSP_180</li>"
                "<li>平均呼叫建立时长只统计有值的轮次</li>"
                "<li><b>V2.5.3</b> 呼叫建立时长单位从毫秒改为秒，精确到小数点后2位</li>"
                "<li>新增汇总-数据和语音Sheet：标准行+电信+联通，包含语音指标和数据业务指标列</li>"
                "<li>新增按文件名Sheet：每个文件一行，包含文件名、运营商、文件类型、行数、时间范围</li>"
                "<li>新增按运营商Sheet：电信/联通各一行，包含语音文件数、通话轮次、平均呼叫建立时长、优质通话占比</li>"
                "<li>process_vqi函数新增time_range参数，支持外部传入时间起止</li>"
                "<li><b>V2.5.2</b> 全面重构语音VQI处理逻辑：自动分类语音/数据文件、识别INVITE(取最后一个连续)/180/BYE/200完整通话</li>"
                "<li>新增MOS3.5优质通话占比计算：从Detail Info提取DlE2eVqi值，排除65535，大于3.5为优质</li>"
                "<li>VQI输出3个Sheet：原始数据(加运营商/业务类型列)、语音指标统计、MOS3.5计算明细</li>"
                "<li>支持.tmf文件(语音)和.mmf文件(数据)的自动分类</li>"
                "<li><b>V2.5.1</b> 修复拖入文件不起作用的问题；支持csv文件类型</li>"
                "<li><b>V2.5.0</b> 新增语音VQI标签页：实现完整的语音VQI处理功能</li>"
                "<li>VQI功能：读取MMF文件，筛选MessageType非空行，识别SIP信令（INVITE/180/BYE/200），计算呼叫建立时长，判定完整通话</li>"
                "<li>VQI输出：原始数据sheet（MessageType非空的所有行）+ 语音统计sheet（轮次/开始时间/结束时间/呼叫建立时长/是否完整通话）</li>"
                "<li>FTP时长选项：参数设置新增FTP时长下拉选择（10秒/20秒/30秒），默认20秒，动态调整FTP候选检测阈值</li>"
                "<li>Sheet重命名：详细过程→数据业务-详细过程，对比横表→数据业务-对比</li>"
                "</ul>"
                "<p><b>2026-07-17</b></p>"
                "<ul>"
                "<li><b>V2.4.5</b> 参数设置增加时间段限制：复选框启用后，可设置开始/结束时间(HH:mm:ss)</li>"
                "<li>启用时只处理该时间段内的数据，未启用则全量处理</li>"
                "<li><b>V2.4.4</b> 支持用户自定义基准文件</li>"
                "</ul>"
                "<p><b>V2.4.2</b></p>"
                "<ul>"
                "<li>工具改名：5G用户级公共监控速率统计工具 → 用户级数据、语音跟踪统计工具</li>"
                "<li>主界面改为双标签(QTabWidget)：标签1=5G用户级公共监控速率统计工具，标签2=5G虚拟用户跟踪用户语音VQI工具(空壳待接入算法)</li>"
                "<li>速率统计标签布局调整：左侧上设置+下运行(进度/日志)，右侧腾出内嵌显示结果3个sheet(详细过程/汇总/对比横表)</li>"
                "<li>右侧内嵌QTableWidget：默认筛选A=1、复制xlsx颜色、批注(hover显示)</li>"
                "<li>输入文件区：选择/清空按钮紧挨着，拖入框高度翻倍(300px)、边框加粗明显</li>"
                "<li>关于对话框：更新记录按天汇总，同一天多次更新合并列出</li>"
                "</ul>"
                "<p><b>V2.4.1</b> <i>(2026-07-17)</i></p>"
                "<ul>"
                "<li>商店小文件识别修复：GAP_MERGE 2.0→3.0，防止商店大段被切碎误判FTP；8783段等真小文件恢复识别</li>"
                "<li>去掉商店小文件上行占比约束（RLC/MAC占比均不可靠），回到max(dl)选取</li>"
                "<li>对比横表改就近匹配：找离基准最近的同业务段，距离≤300秒，完整轮0缺失</li>"
                "<li>支持嵌套zip：递归解压(zip里套zip)</li>"
                "<li>取消按钮修复：写Excel各阶段加检查点，长循环每5000行检查</li>"
                "<li>清理重复sort_values；fmt_time改预编译正则</li>"
                "</ul>"
                "<p><b>V2.4.0</b> <i>(2026-07-16)</i></p>"
                "<ul>"
                "<li>性能优化：预编译正则 + usecols 只读必要列，MMF加载提速</li>"
                "<li>汇总批注：每格显示公式+实际数值（hover查看）</li>"
                "<li>D/E列颜色标记：详细过程 D列浅蓝/E列浅绿</li>"
                "</ul>"
                "<p><b>V2.3.0</b> <i>(2026-07-16)</i></p>"
                "<ul>"
                "<li>性能优化：基准对齐/筛选列向量化，处理耗时 107秒→14秒(约7.7倍)</li>"
                "<li>新增：取消按钮(阶段点中断)、拖拽文件入窗口、真实进度百分比</li>"
                "<li>对比竖表美化：业务/轮次配色、合并同轮次、冻结首行、行高</li>"
                "</ul>"
                "<p><b>V2.2.0</b> <i>(2026-07-16)</i></p>"
                "<ul>"
                "<li>对比增加「偏差」行/列：(代码−基准)/基准 百分比，速率与时长各算</li>"
                "<li>对比横表每轮由2列扩为3列(代码+基准+偏差)</li>"
                "<li>新增「对比竖表」sheet(横表转置)</li>"
                "<li>主界面改左右分栏(左35%设置/右65%结果)，全屏+自适应分辨率+高分屏字号放大+可拖动+打开输出按钮</li>"
                "</ul>"
                "<p><b>V2.1.0</b> <i>(2026-07-16)</i></p>"
                "<ul>"
                "<li>对比横表匹配改为「业务连续段」判定 + 按 FTP下载 切轮</li>"
                "<li>不再用固定时间窗口(-5~+60s)和硬切6条分组</li>"
                "<li>整轮6业务完整且全对上才编号，否则标「不计轮次」</li>"
                "</ul>"
                "<p><b>V2.0.9</b></p>"
                "<ul>"
                "<li>修复对比横表代码业务结束时间缺失</li>"
                "<li>新增第一个 FTP对 之前不完整轮次的识别</li>"
                "<li>版本号统一更新(文件名/关于/需求文档)</li>"
                "</ul>"
                "<p><b>V2.0.1</b></p>"
                "<ul><li>QCI 支持 5/6/7 多选；主界面增加「关于」</li></ul>"
                "<p><b>V2.0.0</b></p>"
                "<ul>"
                "<li>从零重写，三页向导改单页</li>"
                "<li>自动识别 FTP/商店/微信 6 类业务</li>"
                "<li>从原始 MMF 文件直接生成详细过程/对比/汇总 3 个 Sheet</li>"
                "<li>削峰阈值可调，轮次完整性检查</li>"
                "</ul>"
            )
            v.addWidget(te)
            bb = QPushButton("关闭"); bb.clicked.connect(d.accept)
            v.addWidget(bb)
            d.exec_()


if __name__ == '__main__':
    if GUI_OK and len(sys.argv) == 1:
        app = QApplication(sys.argv); app.setStyle('Fusion')
        w = MainWindow(); w.show(); sys.exit(app.exec())
    else:
        out = process(
            ['联通/mmf20260703115328-电信.xlsx', '联通/mmf20260703115334-电信.xlsx'],
            qci_list=[5,6,7], dl_clip=1000, ul_clip=200
        )
        print(f"输出: {out}")
