#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
关键诊断：为什么22个被matched的store_l时间都不对（前面5个的参考时间 vs 检测时间）
以及为什么26个里面有这么多MISS
"""

import sys, os
import numpy as np
import pandas as pd
import openpyxl
from importlib import util as import_util

spec = import_util.spec_from_file_location('tool', '5G用户级公共监控速率统计工具-V1.04.py')
tool = import_util.module_from_spec(spec)
spec.loader.exec_module(tool)

uproc = tool.UCMProcessor({'dl_peak_limit': 900, 'ul_peak_limit': 160})
paths = ['联通/mmf20260703115328-电信.xlsx','联通/mmf20260703115334-电信.xlsx','联通/mmf20260703115336-联通.xlsx']
params = {'dl_peak_limit': 900, 'ul_peak_limit': 160, 'down_min': 50, 'up_min': 10, 'match_mode': 'auto'}
agg = uproc.parse_mmf(paths, params)
uproc._fn = 'mmf_test'

round_results = uproc._match_auto(agg, params, 50, 10)

# 参考文件store_l
wb = openpyxl.load_workbook('联通/呼叫详情_按时间排序-20260707.xlsx', data_only=True)
ws = wb['联通']
ref_entries = []
for r in range(2, ws.max_row+1):
    biz = str(ws.cell(r,2).value or '')
    if '应用商店' not in biz: continue
    ref_entries.append({'st':ws.cell(r,3).value, 'et':ws.cell(r,4).value, 'big':ws.cell(r,11).value})

# 选store_l行（每组第3行）
ref_sl = [ref_entries[i] for i in range(2, len(ref_entries), 3)]

# 工具检测到的store_l
tool_sl = []
for rd in round_results:
    if 'store_l' in rd:
        s = rd['store_l']
        tool_sl.append(s)

print(f"参考store_l: {len(ref_sl)}, 工具store_l: {len(tool_sl)}")
print()

# 检查工具store_l中有多少能匹配参考
for ts in tool_sl:
    matched = False
    for rs in ref_sl:
        rst = pd.to_datetime(rs['st'])
        ret = pd.to_datetime(rs['et'])
        st_gap = abs((ts['start_t'] - rst).total_seconds())
        et_gap = abs((ts['end_t'] - ret).total_seconds())
        if st_gap < 5 and et_gap < 5:
            matched = True
            break
    if matched:
        print(f"  MATCH: tool {ts['start_t'].strftime('%H:%M:%S')}-{ts['end_t'].strftime('%H:%M:%S')} flow={ts['flow_mb']:.0f}MB")
    else:
        print(f"  EXTRA: tool {ts['start_t'].strftime('%H:%M:%S')}-{ts['end_t'].strftime('%H:%M:%S')} flow={ts['flow_mb']:.0f}MB (不在参考中)")

print()
# 检查参考store_l中有多少被工具匹配
for rs in ref_sl:
    rst = pd.to_datetime(rs['st'])
    ret = pd.to_datetime(rs['et'])
    matched = False
    for ts in tool_sl:
        st_gap = abs((ts['start_t'] - rst).total_seconds())
        et_gap = abs((ts['end_t'] - ret).total_seconds())
        if st_gap < 5 and et_gap < 5:
            matched = True
            break
    if not matched:
        # 看看这个时间段对应的工具分类段
        mask = (agg['t'] >= rst) & (agg['t'] <= ret)
        dl_flow = agg.loc[mask, 'dl_mac'].sum() / 8

        # 查找这个时间附近tool检测到的所有段（所有key）
        nearby = []
        for rd in round_results:
            for k, s in rd.items():
                if s['start_t'] <= ret and s['end_t'] >= rst:
                    nearby.append((k, s['start_t'].strftime('%H:%M:%S'), s['end_t'].strftime('%H:%M:%S'), s['flow_mb']))

        print(f"  MISS ref {rst.strftime('%H:%M:%S')}-{ret.strftime('%H:%M:%S')} dl_flow={dl_flow:.0f}MB | nearby={nearby}")

wb.close()