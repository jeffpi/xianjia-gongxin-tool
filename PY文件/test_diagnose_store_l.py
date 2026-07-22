#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
诊断：检查参考文件中应用商店大包的分布，以及哪些小时段是store_l但flow不够
"""

import sys, os, json, re
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
import numpy as np
import pandas as pd
import openpyxl

sys.path.insert(0, '/Users/sun/ClaudeCode/先甲工信部工具')
from importlib import util as import_util
spec = import_util.spec_from_file_location("tool",
    "/Users/sun/ClaudeCode/先甲工信部工具/5G用户级公共监控速率统计工具-V1.04.py")
tool = import_util.module_from_spec(spec)
spec.loader.exec_module(tool)

uproc = tool.UCMProcessor({'dl_peak_limit': 900, 'ul_peak_limit': 160})
paths = [
    '/Users/sun/ClaudeCode/先甲工信部工具/联通/mmf20260703115328-电信.xlsx',
    '/Users/sun/ClaudeCode/先甲工信部工具/联通/mmf20260703115334-电信.xlsx',
    '/Users/sun/ClaudeCode/先甲工信部工具/联通/mmf20260703115336-联通.xlsx'
]
params = {'dl_peak_limit': 900, 'ul_peak_limit': 160, 'down_min': 50, 'up_min': 10, 'match_mode': 'auto'}
agg = uproc.parse_mmf(paths, params)

# 读参考文件 -> 获取每个store_l的开始时间和结束时间
ref = '/Users/sun/ClaudeCode/先甲工信部工具/联通/呼叫详情_按时间排序-20260707.xlsx'
wb = openpyxl.load_workbook(ref, data_only=True)
ws = wb['联通']
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print(f"参考文件行数={ws.max_row} 列数={ws.max_column}")
print(f"表头: {headers}")

# 找各列索引
biz_col = None; st_col = None; et_col = None; rate_col = None; dur_col = None; dl_col = None
for c, h in enumerate(headers, 1):
    hs = str(h) if h else ''
    if '业务类型' in hs: biz_col = c
    if '开始时间' in hs: st_col = c
    if '结束时间' in hs: et_col = c
    if '下载平均速率' in hs: rate_col = c
    if '业务时长' in hs: dur_col = c
    if '大文件下载速率' in hs: dl_col = c

# 读参考文件中所有行（应用商店、微信、FTP的统一行）
ref_times = []
for r in range(2, ws.max_row + 1):
    biz = str(ws.cell(r, biz_col).value or '')
    st = ws.cell(r, st_col).value
    et = ws.cell(r, et_col).value
    dl = ws.cell(r, dl_col).value if dl_col else None

    if '应用商店' in biz:
        name = '应用商店'
        # 判断是store_l还是store_s：查看有没有大文件下载速率
        ref_times.append({'type': name, 'row': r, 'start': st, 'end': et, 'dl_rate': dl, 'big_file_rate': dl})

wb.close()
print(f"\n参考文件中 应用商店 共 {len(ref_times)} 行")
for t in ref_times:
    print(f"  行{t['row']}: {t['start']} -> {t['end']} 大文件速率={t['big_file_rate']}")

# 现在从UCM中，查看这些时间对应的MAC flow
print("\n============ 参考文件中store_l时间段 对应的UCM秒级聚合flow ============")
for t in ref_times:
    st = pd.to_datetime(t['start']) if t['start'] else None
    et = pd.to_datetime(t['end']) if t['end'] else None
    if st and et:
        mask = (agg['t'] >= st) & (agg['t'] <= et)
        sub = agg[mask]
        if len(sub):
            dl_flow = sub['dl_mac'].sum() / 8
            ul_flow = sub['ul_mac'].sum() / 8
            dur = len(sub)
            avg_rate = sub['dl_rlc'].replace(0, np.nan).mean()
            print(f"  {st.strftime('%H:%M:%S')}-{et.strftime('%H:%M:%S')}: "
                  f"dur={dur}s dl_flow={dl_flow:.0f}MB avg_rate={avg_rate:.1f}Mbps")
        else:
            print(f"  {st.strftime('%H:%M:%S')}-{et.strftime('%H:%M:%S')}: 无数据")