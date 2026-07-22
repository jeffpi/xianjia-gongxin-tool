#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
终极诊断：看store_l缺失的轮次和FTP被误判的原因
"""

import numpy as np
import pandas as pd
from importlib import util as import_util

spec = import_util.spec_from_file_location('tool', '5G用户级公共监控速率统计工具-V1.04.py')
tool = import_util.module_from_spec(spec)
spec.loader.exec_module(tool)

uproc = tool.UCMProcessor({'dl_peak_limit': 900, 'ul_peak_limit': 160})
paths = ['联通/mmf20260703115328-电信.xlsx','联通/mmf20260703115334-电信.xlsx','联通/mmf20260703115336-联通.xlsx']
params = {'dl_peak_limit': 900, 'ul_peak_limit': 160, 'down_min': 50, 'up_min': 10, 'match_mode': 'auto'}
agg = uproc.parse_mmf(paths, params)
uproc._fn = 'test'

round_results = uproc._match_auto(agg, params, 50, 10)

# 看前几个轮次中，哪些有ftp_dl但没有store_l
for i, rd in enumerate(round_results[:5]):
    has_sl = 'store_l' in rd
    has_ftp = 'ftp_dl' in rd
    dl_keys = [k for k in rd if k in ('ftp_dl', 'store_s', 'store_l')]
    print(f"轮{i}: has_store_l={has_sl}, has_ftp_dl={has_ftp}, keys={dl_keys}")
    for k in dl_keys:
        s = rd[k]
        print(f"  {k}: flow={s['flow_mb']:.0f}MB dur={s['duration']:.1f}s rate={s['rate']:.1f} {s['start_t'].strftime('%H:%M:%S')}-{s['end_t'].strftime('%H:%M:%S')}")
    print()

# 看store_s时间附近的ftp_dl
print("\n===== 参考store_l时间附近的段 =====")
import openpyxl
wb = openpyxl.load_workbook('联通/呼叫详情_按时间排序-20260707.xlsx', data_only=True)
ws = wb['联通']
ref_entries = []
for r in range(2, ws.max_row+1):
    biz = str(ws.cell(r,2).value or '')
    if '应用商店' not in biz: continue
    ref_entries.append({'st':ws.cell(r,3).value, 'et':ws.cell(r,4).value})
ref_sl = [ref_entries[i] for i in range(2, len(ref_entries), 3)]

# 对前5个参考store_l时间段，看附近所有检测到的段
for idx in range(5):
    sl = ref_sl[idx]
    rst = pd.to_datetime(sl['st'])
    ret = pd.to_datetime(sl['et'])
    # 扩大窗口前后各30秒
    win_start = rst - pd.Timedelta(seconds=30)
    win_end = ret + pd.Timedelta(seconds=30)

    nearby = []
    for rd in round_results:
        for k, s in rd.items():
            if s['start_t'] <= win_end and s['end_t'] >= win_start:
                nearby.append((k, s))
    nearby.sort(key=lambda x: x[1]['start_t'])

    print(f"\nRef SL#{idx}: {rst.strftime('%H:%M:%S')}-{ret.strftime('%H:%M:%S')} [前后30s]")
    for k, s in nearby:
        label = f"  {k:8s}: flow={s['flow_mb']:6.0f}MB dur={s['duration']:5.1f}s rate={s['rate']:6.1f} {s['start_t'].strftime('%H:%M:%S')}-{s['end_t'].strftime('%H:%M:%S')}"
        # 如果这个段在参考窗口内，标记*
        if s['start_t'] <= ret and s['end_t'] >= rst:
            label += " *"
        print(label)
    print()

wb.close()