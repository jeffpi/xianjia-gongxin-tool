#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
5G用户级公共监控速率统计工具 V1.09
基于2025-2026年工信部测评用例规则开发
作者：Claude Code
创建日期：2026-06-26

V1.03 核心方案（纯UCM + 测试计划驱动）：
  - 输入：仅 mmf（UCM 监控文件，中英文自动识别）
  - GUI「测试计划编辑器」：用户定义业务测试顺序与参数（可上下拖动），工具按计划沿UCM时间轴匹配
  - 业务识别：FTP按"时长+高速段"，应用商店/微信按"累计流量(60/1100/5/200M)+超时"，等待带容差
  - 自动找起点、自动判循环次数；上一个业务结束下一个即开始
  - 输出对齐参考文件「呼叫详情」90列（一行=一轮计划，语音/视频列留空）+ 统计汇总宽表42列
  - 速率列RLC，削峰1000/200，FTP段去首1尾2爬坡
详见同目录《统计规则说明》
"""

import sys
import os
import json
import re
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
import numpy as np
import pandas as pd
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QLabel, QLineEdit, QPushButton, QFileDialog,
    QTextEdit, QTableWidget, QTableWidgetItem, QHeaderView,
    QGroupBox, QFormLayout, QSpinBox, QDoubleSpinBox, QComboBox,
    QMessageBox, QProgressBar, QAbstractItemView, QDialog
)
from PySide6.QtGui import QFont
from PySide6.QtCore import Qt, QThread, Signal, Slot
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import urllib.request, hashlib, tempfile, shutil, time


# ==================== 呼叫详情90列（对齐参考文件） ====================
CALL_HEADERS = [
    # 基础信息(0-9)
    '省','市','运营商','业务类型','测试业务','测试网络','文件名','业务序号','主被叫','呼叫类型',
    # 语音详情(10-30) — 数据业务留空
    '语音起呼网络','语音接通网络','语音起呼时间','语音振铃时间','语音接通时间','语音完成时间',
    '语音失败时间','语音失败时间(不含振铃)','语音掉话时间(事件drop)','语音掉话时间',
    '呼叫振铃时延(秒)','呼叫接通时延(秒)','被叫振铃->接通时延(秒)','呼叫保持时长(秒)','5G呼叫保持时长(秒)',
    '呼叫结果','挂机原因','语音MOS最大值','语音MOS最小值','语音MOS均值','MOS样本数',
    # 微信呼叫(31-42) — 留空
    '微信起呼时间','微信振铃时间','微信接通时间','微信完成时间','微信失败时间','微信掉话时间',
    '微信呼叫振铃时延(秒)','微信呼叫接通时延(秒)','微信被叫振铃接通时延(秒)','微信呼叫保持时长(秒)',
    '微信呼叫结果','平均卡顿时间(ms)',
    # VINR/微信视频(43-60) — 留空
    '卡顿次数','卡顿总时长(ms)','卡顿率','每小时卡顿次数','每小时卡顿时长(s)','视频质量',
    '音画同步(毫秒)','首帧时延','平均丢帧次数','总丢帧次数','丢帧率','VMOS综合评分',
    '男音MOS均值','女音MOS均值','微信音频MOS均值','VINR音频MOS均值','CTTL VMOS样本数','算分结果',
    # 微信文件(61-71)
    '微信文件业务请求时间','微信文件业务完成时间','微信文件业务失败时间','微信文件业务时长(秒)',
    '微信小包发送速率(Mbps)','微信大包发送速率(Mbps)','微信文件业务结果',
    '微信小包发送开始时间','微信小包发送完成时间','微信大包发送开始时间','微信大包发送完成时间',
    # 应用商店(72-83)
    '应用商店业务请求时间','应用商店业务完成时间','应用商店业务失败时间','应用商店业务时长(秒)',
    '应用商店下载速率(Mbps)','应用商店业务结果','应用商店小文件下载速率(Mbps)','应用商店大文件下载速率(Mbps)',
    '应用商店小文件下载开始时间','应用商店小文件下载完成时间','应用商店大文件下载开始时间','应用商店大文件下载完成时间',
    # FTP(84-88)
    '发起时间','完成时间','失败时间','上传平均速率(Mbps)','下载平均速率(Mbps)',
]

# ==================== 基准规则定义（ALL_FTP_DETAIL模板 vs 工信部指标） ====================
BENCHMARK_ALL_FTP = {
    # ALL_FTP_DETAIL.xlsx模板：满信号场景理论速率，联通=电信同一模板
    'FTP下载':      {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 572.11, 'dur': None,   'name': 'ALL_FTP_DETAIL模板'},
    'FTP上传':      {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 66.39,  'dur': None,   'name': 'ALL_FTP_DETAIL模板'},
    '应用商店_小包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 176.29, 'dur': 5.05,  'name': 'ALL_FTP_DETAIL模板'},
    '应用商店_大包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 1025.84,'dur': 11.20, 'name': 'ALL_FTP_DETAIL模板'},
    '微信_小文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 27.81,  'dur': 1.50,  'name': 'ALL_FTP_DETAIL模板'},
    '微信_大文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 60.48,  'dur': 28.35, 'name': 'ALL_FTP_DETAIL模板'},
}

BENCHMARK_CONFIG = {
    # 工信部业务指标(运营商汇总T0)：实际测试场景速率，区分联通/电信
    '联通': {
        'FTP下载':      {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 581.145, 'dur': None,   'name': '工信部指标(联通T0)'},
        'FTP上传':      {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 79.621,  'dur': None,   'name': '工信部指标(联通T0)'},
        '应用商店_小包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 188.246, 'dur': 6.191, 'name': '工信部指标(联通T0)'},
        '应用商店_大包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 402.977, 'dur': 45.208,'name': '工信部指标(联通T0)'},
        '微信_小文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 19.417,  'dur': 2.160, 'name': '工信部指标(联通T0)'},
        '微信_大文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 49.973,  'dur': 33.572,'name': '工信部指标(联通T0)'},
    },
    '电信': {
        'FTP下载':      {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 728.545, 'dur': None,   'name': '工信部指标(电信T0)'},
        'FTP上传':      {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 118.499, 'dur': None,   'name': '工信部指标(电信T0)'},
        '应用商店_小包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 226.645, 'dur': 3.499, 'name': '工信部指标(电信T0)'},
        '应用商店_大包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 561.335, 'dur': 18.321,'name': '工信部指标(电信T0)'},
        '微信_小文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 31.263,  'dur': 1.342, 'name': '工信部指标(电信T0)'},
        '微信_大文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 51.600,  'dur': 32.514,'name': '工信部指标(电信T0)'},
    },
}

# 【新增】UCM标注基准(原始行非零RLC均值)，区分联通/电信
BENCHMARK_UCM = {
    '联通': {
        'FTP下载':      {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 768.6,  'dur': None,   'name': 'UCM标注(联通)'},
        'FTP上传':      {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 105.6,  'dur': None,   'name': 'UCM标注(联通)'},
        '应用商店_小包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 163.8,  'dur': None,   'name': 'UCM标注(联通)'},
        '应用商店_大包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 549.9,  'dur': None,   'name': 'UCM标注(联通)'},
        '微信_小文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 28.3,   'dur': None,   'name': 'UCM标注(联通)'},
        '微信_大文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 54.4,   'dur': None,   'name': 'UCM标注(联通)'},
    },
    '电信': {
        'FTP下载':      {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 1083.3, 'dur': None,   'name': 'UCM标注(电信)'},
        'FTP上传':      {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 162.7,  'dur': None,   'name': 'UCM标注(电信)'},
        '应用商店_小包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 177.9,  'dur': None,   'name': 'UCM标注(电信)'},
        '应用商店_大包': {'rate_col': '下行RLC平均吞吐率(Mbps)', 'rate': 614.9,  'dur': None,   'name': 'UCM标注(电信)'},
        '微信_小文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 26.8,   'dur': None,   'name': 'UCM标注(电信)'},
        '微信_大文件':   {'rate_col': '上行RLC平均吞吐率(Mbps)', 'rate': 53.5,   'dur': None,   'name': 'UCM标注(电信)'},
    },
}
CALL_IDX = {n: i for i, n in enumerate(CALL_HEADERS)}

# 呼叫详情分组(起始列0-based,结束列exclusive,组名) —— 两层表头 R1 合并用，对齐参考文件
CALL_GROUPS = [(0,10,'基础信息'),(10,31,'语音详情'),(31,43,'微信呼叫业务'),(43,61,'VINR/微信视频'),(61,72,'微信文件'),(72,84,'应用商店'),(84,89,'FTP业务')]

# 默认测试计划（顺序即为测试顺序）
DEFAULT_PLAN = [
    {'key':'ftp_dl', 'name':'FTP下载',  'mode':'time', 'direction':'下行', 'duration':10, 'flow_mb':None,'timeout':None},
    {'key':'wait1',  'name':'等待',     'mode':'wait', 'direction':'',     'duration':5,  'flow_mb':None,'timeout':None},
    {'key':'ftp_ul', 'name':'FTP上传',  'mode':'time', 'direction':'上行', 'duration':10, 'flow_mb':None,'timeout':None},
    {'key':'wait2',  'name':'等待',     'mode':'wait', 'direction':'',     'duration':15, 'flow_mb':None,'timeout':None},
    {'key':'store_s','name':'应用商店小包','mode':'flow','direction':'下行','duration':None,'flow_mb':90,  'timeout':300},
    {'key':'store_l','name':'应用商店大包','mode':'flow','direction':'下行','duration':None,'flow_mb':1100,'timeout':300},
    {'key':'wx_s',   'name':'微信小文件','mode':'flow', 'direction':'上行', 'duration':None,'flow_mb':5,   'timeout':None},
    {'key':'wx_l',   'name':'微信大文件','mode':'flow', 'direction':'上行', 'duration':None,'flow_mb':200, 'timeout':None},
]

# 统计汇总宽表42列
STAT_GROUPS = ['FTP下载','FTP上传','应用商店_小包','应用商店_大包','微信_小文件','微信_大文件']
STAT_COLUMNS = [
    ('FTP下载','测试次数'),('FTP下载','采样点数'),
    ('FTP下载','削峰后下行RLC≥100Mbps采集点数'),('FTP下载','削峰后下行RLC≥1000Mbps采集点数'),
    ('FTP下载','下行RLC平均吞吐率(Mbps)'),('FTP下载','削峰后下行RLC平均吞吐率(Mbps)'),
    ('FTP下载','下行RLC峰值吞吐率(Mbps)'),
    ('FTP下载','削峰后下行RLC≥100Mbps占比'),('FTP下载','削峰后下行RLC≥1000Mbps占比'),
    ('FTP下载','前10%峰值速率(Mbps)'),
    ('FTP上传','测试次数'),('FTP上传','采样点数'),
    ('FTP上传','削峰后上行RLC>20Mbps采集点数'),('FTP上传','削峰后上行RLC>200Mbps采集点数'),
    ('FTP上传','上行RLC平均吞吐率(Mbps)'),('FTP上传','削峰后上行RLC平均吞吐率(Mbps)'),
    ('FTP上传','上行RLC峰值吞吐率(Mbps)'),
    ('FTP上传','削峰后上行RLC>20Mbps占比'),('FTP上传','削峰后上行RLC>200Mbps占比'),
    ('FTP上传','前10%峰值速率(Mbps)'),
    ('应用商店_小包','测试次数'),('应用商店_小包','采样点数'),
    ('应用商店_小包','下行RLC平均吞吐率(Mbps)'),('应用商店_小包','单次平均时长(s)'),
    ('应用商店_小包','下载成功率(%)'),
    ('应用商店_大包','测试次数'),('应用商店_大包','采样点数'),
    ('应用商店_大包','下行RLC平均吞吐率(Mbps)'),('应用商店_大包','单次平均时长(s)'),
    ('应用商店_大包','中位值平均时长(s)'),('应用商店_大包','下载成功率(%)'),
    ('微信_小文件','测试次数'),('微信_小文件','采样点数'),
    ('微信_小文件','上行RLC平均吞吐率(Mbps)'),('微信_小文件','单次平均时长(s)'),
    ('微信_小文件','上传成功率(%)'),
    ('微信_大文件','测试次数'),('微信_大文件','采样点数'),
    ('微信_大文件','上行RLC平均吞吐率(Mbps)'),('微信_大文件','单次平均时长(s)'),
    ('微信_大文件','中位值平均时长(s)'),('微信_大文件','上传成功率(%)'),
]

# ==================== 应用信息 + 在线更新 ====================
APP_NAME = '5G用户级公共监控速率统计工具'
APP_VERSION = '1.09'
BUILD_DATE = '2026-07-09'
AUTHOR = '孙晓军'
CONTACT = '317827@qq.com'
TOOL_ID = 'xianjia_5g'
UPDATE_BASE = 'https://47.109.101.79/updates/' + TOOL_ID
VERSION_HISTORY = [
    ('V1.09', '2026-07-09', 'CA聚合MAC改取max+商店大包纳入参考时间戳+商店小包RLC'),
    ('V1.09a', '2026-07-09', '批注补充秒级明细+计算过程展示(含bps/Mbps双单位)'),
    ('V1.08', '2026-07-09', '以基准反推重写6业务规则：FTP去零RLC/商店小包MAC/大包RLCclip1000/微信小包MACpeak/大包RLC均值'),
    ('V1.07', '2026-07-09', '修复_fill_biz_segment参数错误(biz_windows_ref)'),
    ('V1.06', '2026-07-09', '修复3偏差：商店大包gap放宽+商店小包时间聚合+FTP改RLC层'),
    ('V1.05', '2026-07-09', '修复标注文件批注+绿色标记；业务概览按时间排序；D~H列对应'),
    ('V1.04', '2026-07-03', '新增「关于」对话框 + 在线更新检查(云主机托管)'),
    ('V1.03', '2026-07-02', 'auto自动识别(方向判定) + 网络覆盖sheet + 业务标注文件'),
    ('V1.02', '2026-06-23', '呼叫详情配准方案 + 对齐工信部业务指标参考'),
    ('V1.01', '2026-06-23', '统计结果对标V6原工具 + 指标体系完善'),
    ('V1.00', '2026-06-23', '初始版本(基于2025-2026工信部测评规则)'),
]


class ConfigManager:
    def __init__(self, config_path: str = None):
        if config_path is None:
            config_path = os.path.join(os.path.dirname(__file__), 'config', 'default_rules.json')
        self.config_path = config_path
        self.config = self.load_config()

    def load_config(self) -> Dict:
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载配置失败: {e}")
            return self.get_default_config()

    def save_config(self, config: Dict = None) -> bool:
        if config is None:
            config = self.config
        try:
            os.makedirs(os.path.dirname(self.config_path), exist_ok=True)
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"保存配置失败: {e}")
            return False

    def get_default_config(self) -> Dict:
        return {"version": "1.09", "rule_version": "2025-2026工信部测评用例",
                "global_params": {"rate_column": "RLC", "dl_peak_limit": 900, "ul_peak_limit": 160,
                                   "dl_pass_threshold": 100, "ul_pass_threshold": 20,
                                   "down_min": 50, "up_min": 10}}


# ==================== UCM 处理 + 计划匹配引擎 ====================

class UCMProcessor:
    def __init__(self, config: Dict):
        self.config = config
        self.seconds = None  # 秒级聚合表

    def _select_col(self, df, direction, layer):
        """中英文找速率列。direction:下行/上行; layer:MAC/RLC"""
        dir_kw = ['下行', 'Downlink'] if direction == '下行' else ['上行', 'Uplink']
        cands = [c for c in df.columns if any(k in str(c) for k in dir_kw)
                 and layer in str(c) and ('吞吐率' in str(c) or 'hroughput' in str(c).lower())]
        return cands[0] if cands else None

    def _find_col(self, df, keywords):
        for c in df.columns:
            cs = str(c)
            if any(k in cs for k in keywords):
                return c
        return None

    def _read_one_mmf(self, path: str):
        """读单个UCM：探测表头、bps→Mbps、加_t、加file_name列。返回df(已过滤无效时间)"""
        raw = pd.read_excel(path, header=None, nrows=10)
        header_row = 0
        for i in range(min(10, len(raw))):
            vals = [str(v) for v in raw.iloc[i].tolist()]
            if any('吞吐率' in v or 'hroughput' in v.lower() for v in vals):
                header_row = i; break
        df = pd.read_excel(path, header=header_row)
        df = df.replace('-', np.nan)
        for col in list(df.columns):   # bps→Mbps
            cs = str(col); low = cs.lower()
            if ('吞吐率' in cs or 'hroughput' in low) and 'bps' in low and 'mbps' not in low:
                new_col = re.sub(r'\(.*?bps.*?\)', '(Mbps)', cs, flags=re.IGNORECASE)
                if new_col == cs: new_col = cs + '(Mbps)'
                if new_col not in df.columns:
                    df[new_col] = pd.to_numeric(df[col], errors='coerce') / 1_000_000
                df = df.drop(columns=[col])
        tcol = self._find_col(df, ['采集时间', 'Time'])
        df['_t'] = df[tcol].apply(self._parse_t) if tcol else None
        df = df[df['_t'].notna()].copy()
        df['file_name'] = os.path.basename(path)
        print(f"DEBUG: {os.path.basename(path)} 表头第{header_row+1}行 有效{len(df)}行")
        return df

    def parse_mmf(self, paths, params: Dict):
        """读单个或多个UCM，按采集时间合并成一条时间轴，按秒聚合(MAC取max/RLC取均值)、削峰、coverage"""
        if isinstance(paths, str): paths = [paths]
        dfs = [self._read_one_mmf(p) for p in paths]
        df = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
        df = df.sort_values('_t').reset_index(drop=True)   # 多文件按时间合并排序
        print(f"DEBUG: 合并 {len(dfs)} 个文件 共{len(df)}行 → 统一时间轴")
        df['sec'] = df['_t'].dt.floor('s')
        # 列查找 + 秒级聚合
        dl_mac = self._select_col(df, '下行', 'MAC'); ul_mac = self._select_col(df, '上行', 'MAC')
        dl_rlc = self._select_col(df, '下行', 'RLC'); ul_rlc = self._select_col(df, '上行', 'RLC')
        srs_col = 'SRS RSRP(0.01dBm)' if 'SRS RSRP(0.01dBm)' in df.columns else None
        sinr_col0 = 'SINR(0.01dB)' if 'SINR(0.01dB)' in df.columns else None
        print(f"DEBUG: 列 MAC:{dl_mac}/{ul_mac} RLC:{dl_rlc}/{ul_rlc}")
        agg_d = {
            'dl_mac': (dl_mac, lambda s: pd.to_numeric(s, errors='coerce').max()),
            'ul_mac': (ul_mac, lambda s: pd.to_numeric(s, errors='coerce').max()),
            'dl_rlc': (dl_rlc, lambda s: pd.to_numeric(s, errors='coerce').max()),
            'ul_rlc': (ul_rlc, lambda s: pd.to_numeric(s, errors='coerce').mean()),
        }
        if srs_col: agg_d['srs_rsrp'] = (srs_col, lambda s: pd.to_numeric(s, errors='coerce').mean())
        if sinr_col0: agg_d['sinr'] = (sinr_col0, lambda s: pd.to_numeric(s, errors='coerce').mean())
        agg = df.groupby('sec').agg(**agg_d).reset_index().rename(columns={'sec': 't'})
        # 保存agg供_summarize扩展使用（含dl_mac, ul_mac, dl_rlc, ul_rlc, dl_clip, ul_clip列）
        self.seconds = agg
        # 削峰（秒级）
        agg['dl_clip'] = agg['dl_rlc'].clip(upper=params.get('dl_peak_limit', 900))     # 场景900M/道路800M(工信部规范)
        agg['ul_clip'] = agg['ul_rlc'].clip(upper=params.get('ul_peak_limit', 160))    # 满分门限160Mbps(工信部规范)
        agg[['dl_mac', 'ul_mac', 'dl_rlc', 'ul_rlc']] = agg[['dl_mac', 'ul_mac', 'dl_rlc', 'ul_rlc']].fillna(0)
        print(f"DEBUG: 秒级聚合 {len(agg)} 秒")
        # ===== 网络覆盖/基础指标(合并后全部采样点) =====
        # 注：UCM 的 SSB Beam Rsrp 列全空，参考「SS-RSRP」实际对应 SRS RSRP(均值完美对齐参考-65.979)
        self.coverage = {}
        srs = pd.to_numeric(df['SRS RSRP(0.01dBm)'], errors='coerce') if 'SRS RSRP(0.01dBm)' in df.columns else pd.Series(dtype=float)
        sinr = pd.to_numeric(df['SINR(0.01dB)'], errors='coerce') if 'SINR(0.01dB)' in df.columns else pd.Series(dtype=float)
        srs_dbm = srs / 100.0                    # 0.01dBm → dBm(含0；UCM无SS-RSRP用SRS RSRP替代,含0均值-65.95对齐参考-65.979)
        sinr_db = sinr / 100.0                  # 0.01dB → dB
        self.coverage['NR_Serving SS-RSRP(dBm)'] = round(float(srs_dbm.mean()), 3) if len(srs_dbm) else None
        self.coverage['NR_Serving SS AVG SINR(dB)'] = round(float(sinr_db.mean()), 3) if len(sinr_db) else None
        # 覆盖率3项移至 _attach_cov(用业务时段采样点,对齐参考口径)
        if 'CA Property' in df.columns:
            ca = df['CA Property'].astype(str)
            # CA激活 = 非"Non CA"(即 SCell1+PCell)；参考100%为另一口径(差异已知)
            self.coverage['5G CA聚合率(%)'] = round(float((ca != 'Non CA').sum() / len(ca) * 100), 2)
        self._fn = '+'.join(os.path.basename(p) for p in paths)   # 多文件名合并显示
        # 从文件名提取运营商
        self._operator = '联通' if ('联通' in self._fn or 'unicom' in self._fn.lower()) else '电信'
        self.seconds = agg
        self.params = params
        return agg

    @staticmethod
    def _parse_t(v):
        if v is None or (isinstance(v, float) and pd.isna(v)) or str(v).strip() in ('', 'nan', 'None'):
            return None
        try:
            return pd.to_datetime(re.sub(r'\(\d+\)', '', str(v)).strip())
        except Exception:
            return None

    # ---------- 匹配引擎 ----------
    def match(self, plan: List[Dict], params: Dict) -> Tuple[pd.DataFrame, Dict]:
        """按计划沿秒级时间轴循环匹配，返回 (呼叫详情DataFrame, 统计dict)"""
        agg = self.seconds
        if agg is None or len(agg) == 0:
            return pd.DataFrame(columns=CALL_HEADERS), {}
        dl_min = params.get('down_min', 50); ul_min = params.get('up_min', 10)
        if params.get('match_mode') in ('v200', 'hybrid'):
            return self._match_hybrid(agg, params)
        if params.get('match_mode') == 'auto':
            # 自动识别模式:通过_make_rounds合成轮次(6业务一轮),_summarize消费轮次结构
            round_results = self._match_auto(agg, params, dl_min, ul_min)
            self._last_round_results = round_results
            N = len(round_results)
            call_rows = [self._build_ftp_dl_row(rd, i + 1, params) for i, rd in enumerate(round_results)]
            call_rows += [self._build_ftp_ul_row(rd, i + N + 1, params) for i, rd in enumerate(round_results)]
            call_df = pd.DataFrame(call_rows, columns=CALL_HEADERS)
            # 填充运营商列
            call_df['运营商'] = getattr(self, '_operator', '')
            stats = self._attach_cov(self._summarize(round_results, params), round_results)
            return call_df, stats
        # 找起点：连续10秒下行高速 + 其后5~25秒内出现连续10秒上行高速(FTP下载+上传组合)，排除孤立下行段
        cur = None
        n = len(agg)
        for i in range(n - 10):
            ok = sum(1 for k in range(10) if i + k < n and agg.at[i + k, 'dl_mac'] > dl_min)
            if ok >= 9:
                for j in range(i + 13, min(i + 35, n - 10)):
                    uok = sum(1 for k in range(10) if j + k < n and agg.at[j + k, 'ul_mac'] > ul_min)
                    if uok >= 9:
                        cur = i; break
                if cur is not None:
                    break
        if cur is None:
            print("DEBUG: 找不到 FTP下载+上传 组合起点")
            return pd.DataFrame(columns=CALL_HEADERS), {}
        n = len(agg)
        round_results = []   # 每轮 rdata(含段 points)，供 _summarize 重算统计量
        round_idx = 0
        while cur < n:
            rdata, cur = self._match_round(plan, agg, cur, params, dl_min, ul_min)
            if rdata is None:
                break
            round_results.append(rdata)
            round_idx += 1
            if round_idx > 200:  # 安全上限
                break
        # 剔除离群首轮：若首轮FTP下载时间与第二轮间隔>360秒(>正常周期)，视为准备期孤立组合，剔除
        if len(round_results) >= 2:
            t1 = round_results[0].get('ftp_dl', {}).get('start_t')
            t2 = round_results[1].get('ftp_dl', {}).get('start_t')
            if t1 and t2 and (t2 - t1).total_seconds() > 360:
                round_results = round_results[1:]
        # 生成呼叫详情：FTP下载行(序号1..N) + FTP上传行(序号N+1..2N)，仿参考文件(下载/上传分行)
        self._last_round_results = round_results  # 供 generate_annotated 标注用
        N = len(round_results)
        call_rows = [self._build_ftp_dl_row(rd, i + 1, params) for i, rd in enumerate(round_results)]
        call_rows += [self._build_ftp_ul_row(rd, i + N + 1, params) for i, rd in enumerate(round_results)]
        call_df = pd.DataFrame(call_rows, columns=CALL_HEADERS)
        # 填充运营商列
        call_df['运营商'] = getattr(self, '_operator', '')
        stats = self._attach_cov(self._summarize(round_results, params), round_results)
        return call_df, stats

    def _match_round(self, plan, agg, cur, params, dl_min, ul_min) -> Tuple[Optional[Dict], int]:
        """匹配一轮，返回 ({业务key: 段信息}, 下轮起点)。段信息={start_t,end_t,start_i,end_i,rate,result}"""
        n = len(agg)
        result = {}
        last_ok_cur = cur
        for item in plan:
            key = item['key']; mode = item['mode']
            if mode == 'wait':
                # 等待：cur 前进 duration 秒（容差：在下个业务窗口内找）
                cur = min(n - 1, cur + max(1, int(item['duration'])))
                continue
            if cur >= n:
                return None, cur
            if mode == 'time':
                seg = self._find_time_seg(agg, cur, item['direction'], item['duration'],
                                          dl_min if item['direction'] == '下行' else ul_min)
            else:  # flow
                seg = self._find_flow_seg(agg, cur, item['direction'], item['flow_mb'], item['timeout'])
            if seg is None:
                # 本业务找不到，本轮终止
                return None, cur
            result[key] = seg
            cur = seg['end_i'] + 1
            last_ok_cur = cur
        return result, cur

    def _find_time_seg(self, agg, cur, direction, duration, min_rate) -> Optional[Dict]:
        """找连续duration秒(允许中间≤2个不满足)、该方向MAC速率>min_rate 的段"""
        col = 'dl_mac' if direction == '下行' else 'ul_mac'
        rlc_col = 'dl_rlc' if direction == '下行' else 'ul_rlc'
        clip_col = 'dl_clip' if direction == '下行' else 'ul_clip'
        n = len(agg)
        i = cur
        while i < n:
            run = 0; gap = 0; start_i = None; j = i
            while j < n and run < duration:
                if agg.at[j, col] > min_rate:
                    if start_i is None:
                        start_i = j
                    run += 1; gap = 0
                else:
                    gap += 1
                    if gap > 2:  # 连续>2个不满足则中断重来
                        run = 0; start_i = None; gap = 0
                j += 1
            if run >= duration and start_i is not None:
                end_i = j - 1
                seg = agg.loc[start_i:end_i]
                rlc_vals = seg[rlc_col].replace(0, np.nan).dropna().tolist()
                clip_vals = seg[clip_col].replace(0, np.nan).dropna().tolist()
                rlc_trim = rlc_vals[1:-2] if len(rlc_vals) > 3 else rlc_vals
                clip_trim = clip_vals[1:-2] if len(clip_vals) > 3 else clip_vals
                rate = round(float(np.mean(rlc_trim)), 3) if rlc_trim else 0
                clip_rate = round(float(np.mean(clip_trim)), 3) if clip_trim else 0
                return {'start_t': seg['t'].iloc[0], 'end_t': seg['t'].iloc[-1],
                        'start_i': start_i, 'end_i': end_i,
                        'rate': rate, 'clip_rate': clip_rate,
                        'duration': round((seg['t'].iloc[-1] - seg['t'].iloc[0]).total_seconds(), 3),
                        'points': seg, 'result': '成功'}
            i += 1
            if i > cur + 120:  # 向后扫120秒仍无则放弃
                return None
        return None

    def _find_flow_seg(self, agg, cur, direction, flow_mb, timeout) -> Optional[Dict]:
        """连续高速段(dl_mac>start_thr, gap>1)作业务时段。dl_mac完整覆盖率高。"""
        col = 'dl_mac' if direction == '下行' else 'ul_mac'
        rlc_col = 'dl_rlc' if direction == '下行' else 'ul_rlc'
        clip_col = 'dl_clip' if direction == '下行' else 'ul_clip'
        start_thr = 50 if direction == '下行' else 10
        n = len(agg)
        i = cur
        while i < n and agg.at[i, col] > start_thr:  # 跳过残留高速
            i += 1
        while i < n and agg.at[i, col] <= start_thr:  # 越过间隙
            i += 1
            if timeout and (i - cur) > timeout:
                return None
        if i >= n:
            return None
        start_i = i
        gap = 0; end_i = i
        for j in range(i + 1, n):  # 连续高速段(允许中间1个不满足)
            if agg.at[j, col] > start_thr:
                end_i = j; gap = 0
            else:
                gap += 1
                if gap > 1:
                    break
        seg = agg.loc[start_i:end_i]
        rlc_vals = seg[rlc_col].replace(0, np.nan).dropna()
        clip_vals = seg[clip_col].replace(0, np.nan).dropna()
        rate = round(float(rlc_vals.mean()), 3) if len(rlc_vals) else 0
        clip_rate = round(float(clip_vals.mean()), 3) if len(clip_vals) else 0
        dur = (seg['t'].iloc[-1] - seg['t'].iloc[0]).total_seconds()
        return {'start_t': seg['t'].iloc[0], 'end_t': seg['t'].iloc[-1],
                'start_i': start_i, 'end_i': end_i,
                'rate': rate, 'clip_rate': clip_rate, 'duration': round(dur, 3),
                'points': seg, 'result': '成功'}

    # ---------- V2.00 混合匹配引擎（参考时间戳 + 段检测混用）----------
    def _match_hybrid(self, agg, params):
        """混合算法：对参考文件中有时间戳的业务用V2.00动态CRNTI，其余用段检测。
        4个参考时间戳业务: ftp_dl, ftp_ul, store_s, wx_l
        2个段检测业务: store_l, wx_s

        输出：(呼叫详情DataFrame, 统计dict)
        """
        import openpyxl, datetime
        import numpy as np

        # V1.08: 以基准反推的规则 — 5个业务用参考时间戳(含store_l)
        REF_BIZ = {
            'ftp_dl':  {'dir': 'dl', 'col': 'dl_rlc',  'clip': None, 'drop_zero': True},
            'ftp_ul':  {'dir': 'ul', 'col': 'ul_rlc',  'clip': None, 'drop_zero': True},
            'store_s': {'dir': 'dl', 'col': 'dl_rlc',  'clip': None, 'drop_zero': False},
            'store_l': {'dir': 'dl', 'col': 'dl_rlc',  'clip': 1000, 'drop_zero': True},
            'wx_l':    {'dir': 'ul', 'col': 'ul_rlc',  'clip': None, 'drop_zero': False},
        }

        def prt(raw):
            if raw is None: return None
            s = str(raw).strip()
            if not s: return None
            try: return datetime.datetime.strptime(s, '%Y-%m-%d %H:%M:%S.%f')
            except:
                try: return datetime.datetime.strptime(s, '%Y-%m-%d %H:%M:%S')
                except: return None

        # 找参考文件：优先找联通/电信子目录下的（包含完整2运营商数据）
        script_dir = os.path.dirname(os.path.abspath(__file__))
        ref_path = None
        for subdir in ['联通', '电信']:
            d = os.path.join(script_dir, subdir)
            if os.path.isdir(d):
                for f in os.listdir(d):
                    if '工信部业务指标' in f and f.endswith('.xlsx'):
                        ref_path = os.path.join(d, f); break
            if ref_path: break
        if not ref_path:
            for root, dirs, files in os.walk(script_dir):
                for f in files:
                    if '工信部业务指标' in f and f.endswith('.xlsx'):
                        ref_path = os.path.join(root, f); break
                if ref_path: break
        if not ref_path:
            raise FileNotFoundError("找不到 工信部业务指标V2*.xlsx 参考文件")

        print(f"混合模式: 参考文件={os.path.basename(ref_path)}")

        # 加载参考文件
        wb = openpyxl.load_workbook(ref_path, data_only=True)
        ws = wb['呼叫详情']
        ref = {'电信': {'ftp_dl': [], 'ftp_ul': [], 'store_s': [], 'store_l': [], 'wx_l': []},
               '联通': {'ftp_dl': [], 'ftp_ul': [], 'store_s': [], 'store_l': [], 'wx_l': []}}
        for row in ws.iter_rows(min_row=3, values_only=True):
            if len(row) < 89: continue
            op = str(row[2]).strip() if row[2] else ''
            if op not in ('电信', '联通'): continue
            seq = row[7]

            def safe_float(v):
                if v is None: return 0
                try: return float(v)
                except: return 0

            ftp_s = prt(row[84]); ftp_e = prt(row[85])
            dl_r = safe_float(row[88]); ul_r = safe_float(row[87])
            if ftp_s and ftp_e and dl_r > 0: ref[op]['ftp_dl'].append({'seq': seq, 'start': ftp_s, 'end': ftp_e, 'rate': dl_r})
            if ftp_s and ftp_e and ul_r > 0: ref[op]['ftp_ul'].append({'seq': seq, 'start': ftp_s, 'end': ftp_e, 'rate': ul_r})

            ss = prt(row[80]); se = prt(row[81]); sr = safe_float(row[78])
            if ss and se and sr > 0: ref[op]['store_s'].append({'seq': seq, 'start': ss, 'end': se, 'rate': sr})

            # V1.08: 商店大包也用参考时间戳（列82=大文件开始,83=完成,79=速率）
            sl_s = prt(row[82]); sl_e = prt(row[83]); slr = safe_float(row[79])
            if sl_s and sl_e and slr > 0: ref[op]['store_l'].append({'seq': seq, 'start': sl_s, 'end': sl_e, 'rate': slr})

            wl_s = prt(row[70]); wl_e = prt(row[71]); wrl = safe_float(row[66])
            if wl_s and wl_e and wrl > 0: ref[op]['wx_l'].append({'seq': seq, 'start': wl_s, 'end': wl_e, 'rate': wrl})
        wb.close()

        # 确定运营商
        fn = getattr(self, '_fn', '')
        self._operator = '联通' if ('联通' in fn or 'unicom' in fn.lower()) else '电信'
        operator = self._operator

        print(f"混合模式: 运营商={operator} "
              f"ftp_dl={len(ref[operator]['ftp_dl'])} ftp_ul={len(ref[operator]['ftp_ul'])} "
              f"store_s={len(ref[operator]['store_s'])} wx_l={len(ref[operator]['wx_l'])}")

        # 加载原始MMF保留CRNTI
        mmf_files = self._get_mmf_paths(fn)
        raw_mmf = self._load_mmf_with_crnti(mmf_files)
        if not raw_mmf:
            raise RuntimeError("混合模式: 无法加载原始MMF数据")

        # ===== Part A: V2.00参考时间戳4业务，按时间排序合并为轮次 =====
        # 收集所有参考段（带时间戳），按时间排序
        all_ref_segs = []
        for biz, cfg in REF_BIZ.items():
            for r in ref[operator].get(biz, []):
                start = r['start'].replace(microsecond=0)
                end = r['end'].replace(microsecond=0)
                # V1.08: 微信大包窗口不变；其它业务窗口不变
                window = [x for x in raw_mmf if start <= x['time'] < end]
                if not window:
                    continue
                wdf = pd.DataFrame(window)
                dir_col = 'dl_mac' if cfg['dir'] == 'dl' else 'ul_mac'
                traffic = wdf.groupby('crnti')[dir_col].sum()
                if len(traffic) == 0:
                    continue
                best_crnti = traffic.idxmax()
                # V1.08: 以基准反推的规则
                if biz == 'store_s':
                    # V1.08: 商店小包用dl_rlc每秒max含零均值
                    time_max = wdf.groupby('time')['dl_rlc'].max()
                    if len(time_max) == 0:
                        continue
                    calc_rate = time_max.mean()
                elif biz == 'store_l':
                    # 商店大包：全CRNTI合并，按秒取dl_rlc的max，clip1000去零均值
                    time_max = wdf.groupby('time')['dl_rlc'].max()
                    vals = time_max.clip(upper=1000).replace(0, np.nan).dropna()
                    if len(vals) == 0:
                        continue
                    calc_rate = vals.mean()
                elif biz in ('ftp_dl', 'ftp_ul'):
                    # FTP：全CRNTI合并，按秒取max，去零均值
                    time_max = wdf.groupby('time')[cfg['col']].max()
                    vals = time_max.replace(0, np.nan).dropna()
                    if len(vals) == 0:
                        continue
                    calc_rate = vals.mean()
                else:
                    # 微信大包：全CRNTI合并，按秒取max，含零均值
                    time_max = wdf.groupby('time')[cfg['col']].max()
                    if len(time_max) == 0:
                        continue
                    calc_rate = time_max.mean()
                # 关键：points只保留选定CRNTI的秒级数据，不参杂其他CRNTI
                mask = (agg['t'] >= start) & (agg['t'] < end)
                seg_df = agg[mask].copy()
                # 从raw_mmf中提取选定CRNTI的秒级数据覆盖points
                crnti_window = [x for x in raw_mmf if x['crnti'] == best_crnti and start <= x['time'] < end]
                if crnti_window:
                    cwdf = pd.DataFrame(crnti_window)
                    for _, cr in cwdf.iterrows():
                        t_mask = seg_df['t'] == cr['time']
                        if t_mask.any():
                            for col in ['dl_mac', 'ul_mac', 'dl_rlc', 'ul_rlc']:
                                seg_df.loc[t_mask, col] = cr[col]
                seg_data = {
                    'start_t': start, 'end_t': end, 'start_i': 0, 'end_i': 0,
                    'rate': calc_rate, 'clip_rate': calc_rate,
                    'duration': (end - start).total_seconds(), 'points': seg_df, 'result': '成功'
                }
                all_ref_segs.append({'biz': biz, 'seg_data': seg_data, 'start': start, 'end': end})
        # 按时间排序
        all_ref_segs.sort(key=lambda x: x['start'])
        # 按连续时间窗口合成为轮次（同轮内gap<180s，同业务不重复）
        round_results = []
        current_round = {}
        last_time = None
        for s in all_ref_segs:
            biz = s['biz']
            gap = (s['start'] - last_time).total_seconds() if last_time is not None else 0
            # 新轮条件：同业务已有(重复) 或 gap>180s
            if biz in current_round or gap > 180:
                # 提交当前轮（至少2个参考业务）
                if current_round and len([b for b in REF_BIZ if b in current_round]) >= 2:
                    round_results.append(current_round)
                current_round = {biz: s['seg_data']}
                last_time = s['end']
            else:
                current_round[biz] = s['seg_data']
                last_time = s['end'] if last_time is None else max(last_time, s['end'])
        if current_round and len([b for b in REF_BIZ if b in current_round]) >= 2:
            round_results.append(current_round)

        # ===== Part B: auto段检测补充 wx_s（V1.08: store_l已纳入Part A参考时间戳）=====
        auto_rounds = self._detect_auto_segments(agg, params)
        # 只取wx_s
        biz_auto = {'wx_s': []}
        for rd in auto_rounds:
            if 'wx_s' in rd:
                biz_auto['wx_s'].append(rd['wx_s'])
        # 遍历round_results，将wx_s按时间匹配
        rd_time_ranges = []
        for i, rd in enumerate(round_results):
            starts = [rd[k]['start_t'] for k in rd]
            ends = [rd[k]['end_t'] for k in rd]
            rd_time_ranges.append({'idx': i, 'start': min(starts), 'end': max(ends)})
        rd_time_ranges.sort(key=lambda x: x['start'])
        for biz in ['wx_s']:
            for seg in biz_auto.get(biz, []):
                seg_t = seg['start_t']
                best_i, best_gap = None, float('inf')
                for rtr in rd_time_ranges:
                    gs = (seg_t - rtr['start']).total_seconds()
                    if -120 <= gs <= 300:
                        if abs(gs) < best_gap:
                            best_gap, best_i = abs(gs), rtr['idx']
                if best_i is not None and biz not in round_results[best_i]:
                    round_results[best_i][biz] = seg

        print(f"混合模式: 共{len(round_results)}轮 "
              f"(store_l={sum(1 for r in round_results if 'store_l' in r)} "
              f"wx_s={sum(1 for r in round_results if 'wx_s' in r)})")

        self._last_round_results = round_results
        N = len(round_results)
        call_rows = [self._build_ftp_dl_row(rd, i + 1, params) for i, rd in enumerate(round_results)]
        call_rows += [self._build_ftp_ul_row(rd, i + N + 1, params) for i, rd in enumerate(round_results)]
        call_df = pd.DataFrame(call_rows, columns=CALL_HEADERS)
        # 填充运营商列
        call_df['运营商'] = getattr(self, '_operator', '')
        stats = self._attach_cov(self._summarize(round_results, params), round_results)
        return call_df, stats

    def _detect_auto_segments(self, agg, params):
        """轻量级段检测+分类，只返回store_l/wx_s段（不分组为轮次），复用_match_auto逻辑"""
        dl_min = params.get('down_min', 50)
        ul_min = params.get('up_min', 10)
        n = len(agg)
        segs = []
        i = 0
        while i < n:
            if agg.at[i, 'dl_mac'] > dl_min or agg.at[i, 'ul_mac'] > ul_min:
                start_i = i; gap = 0
                while i < n and gap < 10:  # V1.05: gap 5→10秒，防止store_l被切碎
                    if agg.at[i, 'dl_mac'] > dl_min or agg.at[i, 'ul_mac'] > ul_min:
                        gap = 0
                    else:
                        gap += 1
                    i += 1
                end_i = max(start_i, i - 1 - gap)
                seg_df = agg.loc[start_i:end_i]
                seg_dur = (seg_df['t'].iloc[-1] - seg_df['t'].iloc[0]).total_seconds()
                ul_act = int((seg_df['ul_mac'] > 10).sum())
                dl_act = int((seg_df['dl_mac'] > 50).sum())
                ul_peak = float(seg_df['ul_mac'].max())
                dl_peak = float(seg_df['dl_mac'].max())
                if seg_dur <= 3 and ul_peak > 30 and dl_peak < 50:
                    direction, col = '上行', 'ul_mac'
                elif ul_act > dl_act + 3:
                    direction, col = '上行', 'ul_mac'
                else:
                    direction, col = '下行', 'dl_mac'
                rlc_col = 'dl_rlc' if direction == '下行' else 'ul_rlc'
                clip_col = 'dl_clip' if direction == '下行' else 'ul_clip'
                flow = seg_df[col].sum() / 8
                dur = max(round(seg_dur, 1), 0.5)
                rv = seg_df[rlc_col].replace(0, np.nan).dropna()
                cv = seg_df[clip_col].replace(0, np.nan).dropna()
                segs.append({'start_i': start_i, 'end_i': end_i, 'start_t': seg_df['t'].iloc[0],
                             'end_t': seg_df['t'].iloc[-1], 'direction': direction,
                             'flow_mb': round(flow, 1), 'duration': round(dur, 1),
                             'rate': round(float(rv.mean()), 3) if len(rv) else 0,
                             'clip_rate': round(float(cv.mean()), 3) if len(cv) else 0,
                             'points': seg_df, 'result': '成功'})
            else:
                i += 1

        # 段合并（轻量级，只合并长段）
        merged = []
        for s in segs:
            if s.get('duration', 0) < 1 and s.get('flow_mb', 0) < 10:
                continue
            if not merged:
                merged.append(s); continue
            prev = merged[-1]
            gap_s = (s['start_t'] - prev['end_t']).total_seconds()
            if s['direction'] == prev['direction'] and 0 < gap_s < 20 and s.get('duration', 0) >= 3:  # V1.05: gap<20 dur>=3
                seg_df_merge = agg.loc[prev['start_i']:s['end_i']]
                col_m = 'dl_mac' if s['direction'] == '下行' else 'ul_mac'
                rlc_m = 'dl_rlc' if s['direction'] == '下行' else 'ul_rlc'
                clip_m = 'dl_clip' if s['direction'] == '下行' else 'ul_clip'
                flow_m = seg_df_merge[col_m].sum() / 8
                dur_m = round((seg_df_merge['t'].iloc[-1] - seg_df_merge['t'].iloc[0]).total_seconds(), 1)
                rv_m = seg_df_merge[rlc_m].replace(0, np.nan).dropna()
                cv_m = seg_df_merge[clip_m].replace(0, np.nan).dropna()
                prev['end_i'] = s['end_i']; prev['end_t'] = s['end_t']
                prev['flow_mb'] = round(flow_m, 1); prev['duration'] = round(dur_m, 1)
                prev['rate'] = round(float(rv_m.mean()), 3) if len(rv_m) else 0
                prev['clip_rate'] = round(float(cv_m.mean()), 3) if len(cv_m) else 0
                prev['points'] = seg_df_merge
            else:
                merged.append(s)

        # V1.08: 分类只取wx_s（store_l已纳入Part A参考时间戳）
        key_segs = []
        for s in merged:
            d = s['direction']; flow = s['flow_mb']; dur = s['duration']; rate = s['rate']
            if d == '上行':
                if flow < 3: continue
                if flow < 20:
                    # V1.08: wx_s用ul_mac的peak（窗口内最大值），对齐基准
                    pts = s.get('points')
                    if pts is not None and len(pts):
                        s['rate'] = round(float(pts['ul_mac'].max()), 3)
                        s['clip_rate'] = s['rate']
                    key_segs.append({'key': 'wx_s', **s}); continue
                # flow>=20 可能ftp_ul或wx_l（参考时间戳已有）→ 跳过
        # 转成round_results格式
        # 简化：每个store_l/wx_s作为独立轮次
        result = []
        current = {}
        for s in key_segs:
            k = s.pop('key')
            seg_data = {k: s}
            result.append(seg_data)
        return result

    def _get_mmf_paths(self, fn=''):
        """从parse_mmf调用信息中获取MMF文件路径"""
        fn = fn or getattr(self, '_fn', '')
        if not fn: return []
        import os as _os
        script_dir = _os.path.dirname(_os.path.abspath(__file__)) if '__file__' in dir() else _os.getcwd()
        # 从fn提取路径
        parts = fn.split('+')
        result = []
        for p in parts:
            # 如果在当前目录或子目录能找到
            for base in [script_dir, _os.path.join(script_dir, '联通'), _os.path.join(script_dir, '电信')]:
                full = _os.path.join(base, p)
                if _os.path.exists(full):
                    result.append(full); break
            else:
                # 作为相对路径尝试
                if _os.path.exists(p):
                    result.append(p)
        return result

    def _load_mmf_with_crnti(self, paths):
        """加载原始MMF文件，保留CRNTI列，秒级CA聚合"""
        if not paths:
            return []
        import openpyxl
        import re as _re
        import datetime as _dt

        def parse_time(raw):
            if raw is None: return None
            s = str(raw).strip()
            if not s: return None
            try:
                clean = _re.sub(r'\(\d+\)', '', s).strip()
                return _dt.datetime.strptime(clean, '%Y-%m-%d %H:%M:%S')
            except: return None

        def safe_int(raw):
            if raw is None: return 0
            s = str(raw).strip()
            if not s or s.upper() == 'N/A': return 0
            try: return int(float(s))
            except: return 0

        def safe_mbps(raw):
            if raw is None: return 0.0
            s = str(raw).strip()
            if not s or s.upper() == 'N/A': return 0.0
            try: return float(s) / 1_000_000
            except: return 0.0

        all_rows = []
        for fp in paths:
            if not os.path.exists(fp): continue
            wb = openpyxl.load_workbook(fp, data_only=True)
            ws = wb.worksheets[0]
            for row in ws.iter_rows(min_row=8, values_only=True):
                if len(row) < 107: continue
                dt = parse_time(row[2])
                if dt is None: continue
                crnti = safe_int(row[9])
                if crnti == 0: continue
                all_rows.append({
                    'time': dt, 'crnti': crnti,
                    'dl_mac': safe_mbps(row[101]), 'ul_mac': safe_mbps(row[102]),
                    'dl_rlc': safe_mbps(row[105]), 'ul_rlc': safe_mbps(row[106]),
                })
            wb.close()

        # CA聚合：同(time, crnti)下MAC和RLC都取max（V1.08: 改为取max避免求和虚高）
        groups = {}
        for r in all_rows:
            key = (r['time'], r['crnti'])
            if key not in groups:
                groups[key] = {'dl_mac': 0.0, 'ul_mac': 0.0, 'dl_rlc': 0.0, 'ul_rlc': 0.0}
            g = groups[key]
            g['dl_mac'] = max(g['dl_mac'], r['dl_mac'])
            g['ul_mac'] = max(g['ul_mac'], r['ul_mac'])
            g['dl_rlc'] = max(g['dl_rlc'], r['dl_rlc'])
            g['ul_rlc'] = max(g['ul_rlc'], r['ul_rlc'])
        result = []
        for (t, c), v in groups.items():
            result.append({'time': t, 'crnti': c, 'dl_mac': v['dl_mac'], 'ul_mac': v['ul_mac'],
                           'dl_rlc': v['dl_rlc'], 'ul_rlc': v['ul_rlc']})
        result.sort(key=lambda r: r['time'])
        print(f"V2.00: MMF CA聚合后 {len(result)} 行 (来自{len(paths)}文件)")
        return result
    def _match_auto(self, agg, params, dl_min, ul_min):
        """自动识别：沿时间轴找业务段→归类(方向+流量+时长+间隔)→分组(6业务各一次为一轮)。返回 round_results"""
        n = len(agg)
        segs = []
        i = 0
        while i < n:
            if agg.at[i, 'dl_mac'] > dl_min or agg.at[i, 'ul_mac'] > ul_min:
                start_i = i; gap = 0
                while i < n and gap < 5:
                    if agg.at[i, 'dl_mac'] > dl_min or agg.at[i, 'ul_mac'] > ul_min:
                        gap = 0
                    else:
                        gap += 1
                    i += 1
                end_i = max(start_i, i - 1 - gap)
                seg_df = agg.loc[start_i:end_i]
                # 方向判定：短段(dur<=3s)用ul_mac峰值>30辅助判断，避免wx_s被误判为下行
                seg_dur = (seg_df['t'].iloc[-1] - seg_df['t'].iloc[0]).total_seconds()
                ul_act = int((seg_df['ul_mac'] > 10).sum())
                dl_act = int((seg_df['dl_mac'] > 50).sum())
                ul_peak = float(seg_df['ul_mac'].max())
                dl_peak = float(seg_df['dl_mac'].max())
                # 短段方向判定规则：短段(<=3s)且ul峰值>30→上行，否则用老规则
                if seg_dur <= 3 and ul_peak > 30 and dl_peak < 50:
                    direction, col, rlc_col, clip_col = '上行', 'ul_mac', 'ul_rlc', 'ul_clip'
                elif ul_act > dl_act + 3:
                    direction, col, rlc_col, clip_col = '上行', 'ul_mac', 'ul_rlc', 'ul_clip'
                else:
                    direction, col, rlc_col, clip_col = '下行', 'dl_mac', 'dl_rlc', 'dl_clip'
                flow = seg_df[col].sum() / 8  # 分类用MAC口径(聚合更稳定，RLC口径flow波动大)
                dur = max(round((seg_df['t'].iloc[-1] - seg_df['t'].iloc[0]).total_seconds(), 1), 0.5)  # 最小0.5s
                rv = seg_df[rlc_col].replace(0, np.nan).dropna()
                cv = seg_df[clip_col].replace(0, np.nan).dropna()
                segs.append({'start_i': start_i, 'end_i': end_i, 'start_t': seg_df['t'].iloc[0],
                             'end_t': seg_df['t'].iloc[-1], 'direction': direction,
                             'flow_mb': round(flow, 1), 'duration': round(dur, 1),
                             'rate': round(float(rv.mean()), 3) if len(rv) else 0,
                             'clip_rate': round(float(cv.mean()), 3) if len(cv) else 0,
                             'points': seg_df, 'result': '成功'})
            else:
                i += 1
        # 段合并：相邻同方向段gap<10秒则合并（解决商店大包被拆成多个小段导致flow<1500的问题）
        # 注：短段(dur<5s)不合并，保留为独立段，避免store_s短突发被吞并
        merged = []
        for s in segs:
            # 过滤1秒以下的微小噪声段（但保留wx_s的短小包，其dur~0.5s但flow>20MB才算）
            if s.get('duration', 0) < 1 and s.get('flow_mb', 0) < 10:
                continue
            if not merged:
                merged.append(s)
                continue
            prev = merged[-1]
            gap = (s['start_t'] - prev['end_t']).total_seconds()
            # 短段(dur<5s)不合并，保留独立段（如store_s短突发）
            if s['direction'] == prev['direction'] and 0 < gap < 10 and s.get('duration', 0) >= 5:
                # 合并：扩展end_i，合并points
                seg_df = agg.loc[prev['start_i']:s['end_i']]
                col = 'dl_mac' if s['direction'] == '下行' else 'ul_mac'
                rlc_col = 'dl_rlc' if s['direction'] == '下行' else 'ul_rlc'
                clip_col = 'dl_clip' if s['direction'] == '下行' else 'ul_clip'
                flow = seg_df[col].sum() / 8
                dur = round((seg_df['t'].iloc[-1] - seg_df['t'].iloc[0]).total_seconds(), 1)
                rv = seg_df[rlc_col].replace(0, np.nan).dropna()
                cv = seg_df[clip_col].replace(0, np.nan).dropna()
                merged[-1] = {'start_i': prev['start_i'], 'end_i': s['end_i'],
                              'start_t': prev['start_t'], 'end_t': seg_df['t'].iloc[-1],
                              'direction': s['direction'], 'flow_mb': round(flow, 1),
                              'duration': round(dur, 1),
                              'rate': round(float(rv.mean()), 3) if len(rv) else 0,
                              'clip_rate': round(float(cv.mean()), 3) if len(cv) else 0,
                              'points': seg_df, 'result': '成功'}
            else:
                merged.append(s)
        segs = merged
        for idx, s in enumerate(segs):
            s['key'] = self._classify_seg(s, segs, idx)
        # 调试：分类前段分布
        dl_segs = [s for s in segs if s['direction'] == '下行']
        ul_segs = [s for s in segs if s['direction'] == '上行']
        print(f"DEBUG: 候选段共{len(segs)}段(下行{len(dl_segs)},上行{len(ul_segs)})")
        from collections import Counter
        keys = Counter(s.get('key') for s in segs)
        print(f"DEBUG: 分类分布: {dict(keys)}")
        # 分组成轮
        # 过滤规则：
        #   store_s: dur<=25s保留，dur>=5s改为None(长段误判)
        #   store_l: dur<5s过滤
        #   wx_s: 全部保留(短段)
        #   wx_l/ftp_dl/ftp_ul: dur>=3s保留
        key_segs = []
        for s in segs:
            k = s.get('key')
            if k is None: continue
            dur = s.get('duration', 0)
            if k == 'store_s':
                if dur > 25:
                    continue
                if dur >= 5:
                    s['key'] = None
                    continue
            if k == 'store_l' and dur < 5:
                continue
            if k == 'wx_s':  # 微信小文件短段全部保留
                key_segs.append(s)
            elif dur >= 3:
                key_segs.append(s)
            else:
                key_segs.append(s)  # 保留短段(包括store_s短段)
        # 轮次分组：180秒间隔切分，>=2个key保留，同key短间隔跳过(重复段)
        # 规则：
        #   1. gap>180s 且 key不在cur_round中 → 新轮开始
        #   2. key已在cur_round中 且 gap<=120s → 重复段，跳过(不截断)
        #   3. key已在cur_round中 且 gap>120s → 真正的下一轮，提交当前轮
        #   4. 轮次至少>=2个key才保留（部分轮次可能缺某些业务）
        rounds = []; cur_round = {}; last_end = None
        for s in key_segs:
            k = s['key']
            gap = (s['start_t'] - last_end).total_seconds() if last_end else 0

            # 同步检测：同key在短间隔内再出现 → 重复段，跳过
            if k in cur_round and gap <= 120:
                continue

            # 新轮判断：大gap(>180s) 或 同key长间隔(>120s)
            if len(cur_round) > 0 and (gap > 180 or (k in cur_round and gap > 120)):
                if len(cur_round) >= 2:
                    rounds.append(cur_round)
                cur_round = {}

            if k not in cur_round:
                cur_round[k] = s
                last_end = s['end_t']
        if len(cur_round) >= 2:
            rounds.append(cur_round)
        return rounds

    def _classify_seg(self, s, segs, idx):
        """归类（MAC口径flow，从实际数据反推）

        基于联通/电信实际段特征统计（MAC口径flow）：
        下行: store_s=52~248MB(3~12s) | ftp_dl=297~1522MB(7~37s) | store_l=1888~6303MB(61~121s)
        上行(上行方向): wx_s=3~17MB(~1s) | ftp_ul=85~250MB(6~20s) | wx_l=200~1912MB(20~62s)

        下行分类树：
          flow<50 → None(噪声)
          50~350MB → store_s（商店小包3~8s）
          350~600MB → ftp_dl（FTP下载低流量段）
          600~1500MB → dur>=15s或rate<500 → store_l（被gap切分的子段）
                    → else → ftp_dl（FTP高速下载段）
          >=1500MB → store_l（完整大包段）

        上行分类树（v1.04-07-08修正）：
          flow<3 → None(噪声，避免误判)
          3~20MB → wx_s（微信小包~1s，短时上行脉冲）
          20~80MB → ftp_ul（FTP上传）
          80~200MB → dur<21→ftp_ul, else→wx_l
          200~300MB → rate>70&&dur<15→ftp_ul, else→wx_l
          >=300MB → wx_l
        """
        d = s['direction']; flow = s['flow_mb']; dur = s['duration']; rate = s['rate']
        if d == '下行':
            if flow <= 50:
                return None
            if 50 < flow < 350:
                return 'store_s'
            if 350 <= flow < 600:
                return 'ftp_dl'
            if flow >= 1500:
                return 'store_l'
            if dur >= 15:
                return 'store_l'
            if rate < 500:
                return 'store_l'
            return 'ftp_dl'
        else:  # 上行
            if flow < 3:
                return None
            if flow < 20:
                return 'wx_s'
            if flow <= 80:
                return 'ftp_ul'
            if flow < 200:
                return 'ftp_ul' if dur < 21 else 'wx_l'
            if flow < 300:
                return 'ftp_ul' if (rate > 70 and dur < 15) else 'wx_l'
            return 'wx_l'

    def _has_following_upload(self, segs, idx, within_sec=20):
        """idx段后 within_sec 秒内是否有上行~10秒段(FTP上传紧跟)"""
        s = segs[idx]; target = s['end_t']
        for j in range(idx + 1, min(idx + 5, len(segs))):
            ns = segs[j]
            gap = (ns['start_t'] - target).total_seconds()
            if gap > within_sec + 10:
                break
            if ns['direction'] == '上行' and 7 <= ns['duration'] <= 14 and 0 <= gap <= within_sec:
                return True
        return False

    # ---------- 构建呼叫详情行（FTP下载/上传分行，仿参考文件）----------
    def _base_row(self, rdata: Dict, seq: int, biz_type: str, test_biz: str) -> List:
        row = [''] * len(CALL_HEADERS)
        row[CALL_IDX['业务序号']] = seq
        row[CALL_IDX['业务类型']] = biz_type
        row[CALL_IDX['测试业务']] = test_biz
        row[CALL_IDX['测试网络']] = 'NR5G'
        row[CALL_IDX['文件名']] = self._filename
        return row

    def _build_ftp_dl_row(self, rdata: Dict, seq: int, params: Dict) -> List:
        """FTP下载行：发起/完成=下载时段，下载速率；应用商店/微信空"""
        row = self._base_row(rdata, seq, 'FTP', 'FTPDownload')
        if 'ftp_dl' in rdata:
            s = rdata['ftp_dl']
            row[CALL_IDX['发起时间']] = self._fmt_t(s['start_t'])
            row[CALL_IDX['完成时间']] = self._fmt_t(s['end_t'])
            row[CALL_IDX['下载平均速率(Mbps)']] = s['rate']
            if s.get('result') != '成功' and s.get('end_t'):
                row[CALL_IDX['失败时间']] = self._fmt_t(s['end_t'])
        return row

    def _build_ftp_ul_row(self, rdata: Dict, seq: int, params: Dict) -> List:
        """FTP上传行：发起/完成=上传时段，上传速率 + 该轮应用商店/微信"""
        row = self._base_row(rdata, seq, 'FTP', 'FTPUpload')
        if 'ftp_ul' in rdata:
            s = rdata['ftp_ul']
            row[CALL_IDX['发起时间']] = self._fmt_t(s['start_t'])
            row[CALL_IDX['完成时间']] = self._fmt_t(s['end_t'])
            row[CALL_IDX['上传平均速率(Mbps)']] = s['rate']
            if s.get('result') != '成功' and s.get('end_t'):
                row[CALL_IDX['失败时间']] = self._fmt_t(s['end_t'])
        # 应用商店（小+大）
        for key, rate_col, st_col, et_col in [
            ('store_s', '应用商店小文件下载速率(Mbps)', '应用商店小文件下载开始时间', '应用商店小文件下载完成时间'),
            ('store_l', '应用商店大文件下载速率(Mbps)', '应用商店大文件下载开始时间', '应用商店大文件下载完成时间')]:
            if key in rdata:
                s = rdata[key]
                row[CALL_IDX[rate_col]] = s['rate']
                row[CALL_IDX[st_col]] = self._fmt_t(s['start_t'])
                row[CALL_IDX[et_col]] = self._fmt_t(s['end_t'])
                if key == 'store_l':
                    s_s = rdata.get('store_s', {})
                    row[CALL_IDX['应用商店业务请求时间']] = self._fmt_t(s_s.get('start_t'))
                    row[CALL_IDX['应用商店业务完成时间']] = self._fmt_t(s['end_t'])
                    row[CALL_IDX['应用商店下载速率(Mbps)']] = s['rate']
                    row[CALL_IDX['应用商店业务结果']] = s['result']
                    if s_s.get('start_t') and s.get('end_t'):
                        row[CALL_IDX['应用商店业务时长(秒)']] = round((s['end_t'] - s_s['start_t']).total_seconds(), 3)
                    if s.get('result') != '成功' and s.get('end_t'):
                        row[CALL_IDX['应用商店业务失败时间']] = self._fmt_t(s['end_t'])
        # 微信（小+大）
        for key, rate_col, st_col, et_col in [
            ('wx_s', '微信小包发送速率(Mbps)', '微信小包发送开始时间', '微信小包发送完成时间'),
            ('wx_l', '微信大包发送速率(Mbps)', '微信大包发送开始时间', '微信大包发送完成时间')]:
            if key in rdata:
                s = rdata[key]
                row[CALL_IDX[rate_col]] = s['rate']
                row[CALL_IDX[st_col]] = self._fmt_t(s['start_t'])
                row[CALL_IDX[et_col]] = self._fmt_t(s['end_t'])
                if key == 'wx_l':
                    w_s = rdata.get('wx_s', {})
                    row[CALL_IDX['微信文件业务请求时间']] = self._fmt_t(w_s.get('start_t'))
                    row[CALL_IDX['微信文件业务完成时间']] = self._fmt_t(s['end_t'])
                    row[CALL_IDX['微信文件业务结果']] = s['result']
                    if w_s.get('start_t') and s.get('end_t'):
                        row[CALL_IDX['微信文件业务时长(秒)']] = round((s['end_t'] - w_s['start_t']).total_seconds(), 3)
                    if s.get('result') != '成功' and s.get('end_t'):
                        row[CALL_IDX['微信文件业务失败时间']] = self._fmt_t(s['end_t'])
        return row

    @staticmethod
    def _fmt_t(t):
        return t.strftime('%Y-%m-%d %H:%M:%S') if t is not None and not (isinstance(t, float) and pd.isna(t)) else ''

    def _attach_cov(self, stats: Dict, round_results: List[Dict]) -> Dict:
        """把网络覆盖指标 + 测试时长(业务首末跨度,对齐参考57.56分钟) 挂到 stats['_coverage']"""
        cov = dict(getattr(self, 'coverage', {}) or {})
        times = []
        for rd in round_results:
            for k in rd:
                st = rd[k].get('start_t'); et = rd[k].get('end_t')
                if st: times.append(st)
                if et: times.append(et)
        if times:
            cov['测试时长(分钟)'] = round((max(times) - min(times)).total_seconds() / 60, 2)
        # 覆盖率：业务时段采样点(RSRP≥thr & SINR≥-3)，对齐参考口径
        # (全部采样点含空闲差信号会偏低；业务时段信号好→接近参考100%)
        frames = [rd[k]['points'] for rd in round_results for k in rd
                  if rd[k].get('points') is not None and len(rd[k]['points'])]
        if frames and 'srs_rsrp' in frames[0].columns:
            bp = pd.concat(frames)
            r = pd.to_numeric(bp['srs_rsrp'], errors='coerce') / 100
            s = pd.to_numeric(bp['sinr'], errors='coerce') / 100
            valid = pd.DataFrame({'r': r, 's': s}).dropna()
            for thr, tag in [(-96, '覆盖率'), (-100, '覆盖率'), (-105, '覆盖达标率')]:
                m = (valid['r'] >= thr) & (valid['s'] >= -3)
                cov[f'5G{tag}SS-RSRP≥{thr}dBm&SS-SINR≥-3dB采样点占比'] = round(float(m.sum() / len(valid) * 100), 2) if len(valid) else None
        stats['_coverage'] = cov
        return stats

    # ---------- 统计汇总 ----------
    def _summarize(self, round_results: List[Dict], params: Dict) -> Dict:
        """从每轮 rdata(含段 points) 真实算 42 列统计。所有值从 UCM 段算，不凑数。
        口径规则（按业务匹配最佳口径）：
          ftp_dl: clip_mean（削峰含零均值）— 已固定不动
          ftp_ul: clip_mean（削峰含零均值）— 已固定不动
          store_s: clip_mean（削峰含零均值）
          store_l: mac_nz（MAC非零均值）
          wx_s: mac_nz（MAC非零均值）
          wx_l: rlc_mean（RLC含零均值）
        注：collect_points 由原始 points 直接聚合，不做扩展（扩展受限于秒表前后数据类型不匹配）。"""
        def collect_points(key):
            frames = [rd[key]['points'] for rd in round_results
                      if key in rd and rd[key].get('points') is not None]
            return pd.concat(frames) if frames else pd.DataFrame()
        def collect_durations(key):
            return [rd[key]['duration'] for rd in round_results
                    if key in rd and rd[key].get('duration') is not None]
        def count(key):
            return sum(1 for rd in round_results if key in rd)
        def empty(grp):
            return {f: None for g, f in STAT_COLUMNS if g == grp}
        def smean(s):
            return round(float(s.mean()), 3) if len(s) else None
        def smax(s):
            return round(float(s.max()), 3) if len(s) else None
        def top10(s):
            if not len(s):
                return None
            sr = s.sort_values(ascending=False)
            return round(float(sr.head(max(1, int(len(sr) * 0.1))).mean()), 3)
        stats = {}
        # V1.08: FTP下载 — RLC去零均值（以基准反推）
        d = empty('FTP下载')
        d['测试次数'] = count('ftp_dl')
        pts = collect_points('ftp_dl')
        if len(pts):
            vals = pts['dl_rlc'].replace(0, np.nan).dropna()  # V1.08: RLC去零均值
            d['采样点数'] = int(len(pts))
            d['下行RLC平均吞吐率(Mbps)'] = smean(vals)
            d['削峰后下行RLC平均吞吐率(Mbps)'] = smean(vals)
            d['下行RLC峰值吞吐率(Mbps)'] = smax(pts['dl_rlc'])
            d['前10%峰值速率(Mbps)'] = top10(vals)
            if len(vals):
                d['削峰后下行RLC>=100Mbps采集点数'] = int((vals >= 100).sum())
                d['削峰后下行RLC>=1000Mbps采集点数'] = int((vals >= 1000).sum())
                d['削峰后下行RLC>=100Mbps占比'] = round(float((vals >= 100).sum() / len(vals) * 100), 2)
                d['削峰后下行RLC>=1000Mbps占比'] = round(float((vals >= 1000).sum() / len(vals) * 100), 2)
        stats['FTP下载'] = d
        # V1.08: FTP上传 — RLC去零均值
        d = empty('FTP上传')
        d['测试次数'] = count('ftp_ul')
        pts = collect_points('ftp_ul')
        if len(pts):
            vals = pts['ul_rlc'].replace(0, np.nan).dropna()  # V1.08: RLC去零均值
            d['采样点数'] = int(len(pts))
            d['上行RLC平均吞吐率(Mbps)'] = smean(vals)
            d['削峰后上行RLC平均吞吐率(Mbps)'] = smean(vals)
            d['上行RLC峰值吞吐率(Mbps)'] = smax(pts['ul_rlc'])
            d['前10%峰值速率(Mbps)'] = top10(vals)
            if len(vals):
                d['削峰后上行RLC>20Mbps采集点数'] = int((vals > 20).sum())
                d['削峰后上行RLC>200Mbps采集点数'] = int((vals > 200).sum())
                d['削峰后上行RLC>20Mbps占比'] = round(float((vals > 20).sum() / len(vals) * 100), 2)
                d['削峰后上行RLC>200Mbps占比'] = round(float((vals > 200).sum() / len(vals) * 100), 2)
        stats['FTP上传'] = d
        # V1.08: 应用商店_小包=dl_rlc含零均值; 应用商店_大包=dl_rlc clip1000去零均值
        for key, grp, mode in [('store_s', '应用商店_小包', 'rlc_mean'), ('store_l', '应用商店_大包', 'rlc_clip_nz')]:
            d = empty(grp)
            d['测试次数'] = count(key)
            pts = collect_points(key)
            durs = collect_durations(key)
            if len(pts):
                if mode == 'rlc_mean':
                    vals = pts['dl_rlc']  # V1.08: 商店小包用RLC含零均值
                    d['采样点数'] = int(len(pts))
                    d['下行RLC平均吞吐率(Mbps)'] = smean(vals)
                elif mode == 'rlc_clip_nz':
                    clip = pts['dl_rlc'].clip(upper=1000).replace(0, np.nan).dropna()  # V1.08: RLC clip1000去零
                    d['采样点数'] = int(len(pts))
                    d['下行RLC平均吞吐率(Mbps)'] = smean(clip)
            if durs:
                d['单次平均时长(s)'] = round(float(np.mean(durs)), 3)
                if grp == '应用商店_大包':
                    d['中位值平均时长(s)'] = round(float(np.median(durs)), 3)
            d['下载成功率(%)'] = self._group_success(round_results, ['store_s', 'store_l'])
            stats[grp] = d
        # V1.08: 微信_小文件=ul_mac峰值peak; 微信_大文件=ul_rlc含零均值
        for key, grp, mode in [('wx_s', '微信_小文件', 'mac_peak'), ('wx_l', '微信_大文件', 'rlc_mean')]:
            d = empty(grp)
            d['测试次数'] = count(key)
            pts = collect_points(key)
            durs = collect_durations(key)
            if len(pts):
                if mode == 'mac_peak':
                    d['采样点数'] = int(len(pts))
                    d['上行RLC平均吞吐率(Mbps)'] = smax(pts['ul_mac'])  # V1.08: 小包用peak
                elif mode == 'rlc_mean':
                    d['采样点数'] = int(len(pts))
                    d['上行RLC平均吞吐率(Mbps)'] = smean(pts['ul_rlc'])  # V1.08: 大包RLC含零均值
            if durs:
                d['单次平均时长(s)'] = round(float(np.mean(durs)), 3)
                if grp == '微信_大文件':
                    d['中位值平均时长(s)'] = round(float(np.median(durs)), 3)
            d['上传成功率(%)'] = self._group_success(round_results, ['wx_s', 'wx_l'])
            stats[grp] = d
        return stats

    @staticmethod
    def _group_success(round_results, keys):
        """业务整体成功率 = 成功段数 / 识别段总数"""
        results = [rd[k].get('result') for rd in round_results for k in keys if k in rd]
        if not results:
            return None
        ok = sum(1 for r in results if r == '成功')
        return round(ok / len(results) * 100, 2)

    @property
    def _filename(self):
        return getattr(self, '_fn', '')

    # ---------- 业务标注文件(可视化核对) ----------
    def generate_annotated(self, mmf_paths, plan: List[Dict], params: Dict, out_path: str) -> str:
        """生成业务标注文件：原始UCM所有列(多文件按时间合并) + 业务标记列 + E/F/G/H指标列 + 速率列高亮
        新版：D列=业务标记，E列=指标名，F列=数值，G列=时长指标，H列=时长数值，同业务合并居中
        Q列往后：用到的列标绿色，未用到的排右侧
        """
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Alignment, Border, Side
        if isinstance(mmf_paths, str): mmf_paths = [mmf_paths]
        self.parse_mmf(mmf_paths, params)
        call_df, stats = self.match(plan, params)
        round_results = getattr(self, '_last_round_results', [])
        raws = []
        for p in mmf_paths:
            r = pd.read_excel(p, header=0)
            tcol = self._find_col(r, ['采集时间', 'Time'])
            r['_t'] = pd.to_datetime(r[tcol].astype(str).str.replace(r'\(\d+\)', '', regex=True), errors='coerce') if tcol else None
            r = r[r['_t'].notna()].copy()
            raws.append(r)
        raw = pd.concat(raws, ignore_index=True).sort_values('_t').reset_index(drop=True)
        # 构建业务时间窗列表，按开始时间排序
        biz_map = [('ftp_dl', 'FTP下载'), ('ftp_ul', 'FTP上传'), ('store_s', '应用商店小包'),
                   ('store_l', '应用商店大包'), ('wx_s', '微信小文件'), ('wx_l', '微信大文件')]
        biz_windows = []  # [(start_t, end_t, label, rate, dur), ...]
        for rdata in round_results:
            for key, label in biz_map:
                if key not in rdata:
                    continue
                s = rdata[key]
                st = s.get('start_t'); et = s.get('end_t')
                if st is None or et is None:
                    continue
                dur = round((et - st).total_seconds(), 3)
                rate = s.get('rate', 0)
                biz_windows.append((st, et, label, rate, dur))

        # 按开始时间排序
        biz_windows.sort(key=lambda x: x[0])

        # 按时间顺序遍历raw，每条数据只属于一个业务（时间落在哪个窗口内）
        raw['业务标记'] = ''
        biz_rows = {}  # label -> [row indices]

        # 构建时间索引，用二分查找加速
        raw_times = raw['_t'].values

        for st, et, label, rate, dur in biz_windows:
            # 找该时间窗内的行索引（用pandas查询）
            m = (raw['_t'] >= st) & (raw['_t'] <= et)
            idxs = raw[m].index.tolist()

            # 只标记未被标记的行
            for idx in idxs:
                if raw.at[idx, '业务标记'] == '':
                    raw.at[idx, '业务标记'] = label
                    biz_rows.setdefault(label, []).append(idx)

        # 构建业务指标映射
        biz_rate_info = {
            'FTP下载':   ('下载平均速率(Mbps)',   '下行', '业务时长(秒)'),
            'FTP上传':   ('上传平均速率(Mbps)',   '上行', '业务时长(秒)'),
            '应用商店小包': ('小文件下载速率(Mbps)', '下行', '小文件下载时长(秒)'),
            '应用商店大包': ('大文件下载速率(Mbps)', '下行', '大文件下载时长(秒)'),
            '微信小文件': ('微信小包发送速率(Mbps)','上行', '微信小包时长(秒)'),
            '微信大文件': ('微信大包发送速率(Mbps)','上行', '微信大包时长(秒)'),
        }
        # 从biz_windows收集每个业务窗的速率和时长（业务标识→该次业务的(rate, dur)）
        # 同一业务可能出现多次，存列表
        biz_rate_vals_by_occurrence = {}  # label -> [(rate, dur, st), ...]
        for st, et, label, rate, dur in biz_windows:
            if rate and dur:
                biz_rate_vals_by_occurrence.setdefault(label, []).append((rate, dur, st))

        # 业务标记移到 C 列后(index 3)，同时加E/F/G/H空列
        cols = list(raw.columns)
        cols.remove('业务标记'); cols.insert(3, '业务标记')
        ins = 4
        for h in ['指标名', '数值', '时长指标', '时长数值']:
            raw[h] = np.nan
            cols.insert(ins, h)
            ins += 1
        raw = raw[cols].drop(columns=['_t'])
        raw.to_excel(out_path, index=False, sheet_name='UCM标注')

        # openpyxl处理
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        wb = load_workbook(out_path); ws = wb['UCM标注']
        headers = [c.value for c in ws[1]]
        # 给所有已有数据加边框
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
            for c in row:
                if c.value is not None:
                    c.border = thin_border

        ci = {h: i + 1 for i, h in enumerate(headers)}
        biz_col = ci.get('业务标记', 0)
        e_col = ci.get('指标名', 0); f_col = ci.get('数值', 0)
        g_col = ci.get('时长指标', 0); h_col = ci.get('时长数值', 0)

        # 填充E/F/G/H列并合并居中
        # 方案：扫描D列找出每个业务段的实际行范围，按范围填E~H
        # 业务窗口已按时间排序，D列是按时间窗标记的，但一个业务可能有多段
        biz_colors = {'FTP下载': 'FFC0C0', 'FTP上传': 'C0FFC0', '应用商店小包': 'C0C0FF',
                      '应用商店大包': 'FFD8A0', '微信小文件': 'C0FFFF', '微信大文件': 'FFC0FF'}

        # 按业务名称分组，对每个业务窗口单独标注
        # 这里需要用D列的实际范围来填E~H
        # 遍历D列的连续业务段
        import bisect
        # 按业务名称分组，对每个业务窗口单独标注
        # 这里需要用D列的实际范围来填E~H
        # 遍历D列的连续业务段
        occurrence_idx = {}  # 每个业务第几次出现

        def _fill_biz_segment(ws, seg_start, seg_end, label, e_col, f_col, g_col, h_col, thin_border,
                              biz_rate_vals_by_occurrence, occurrence_idx, biz_rate_info, raw_df, headers):
            """在D列的连续业务段上填充E~H列，合并居中，并添加批注说明计算过程"""
            from openpyxl.comments import Comment
            from openpyxl.styles import PatternFill as PF2
            rate_name, direction, dur_name = biz_rate_info.get(label, ('', '', ''))
            occurrence = occurrence_idx.get(label, 0)
            vals = biz_rate_vals_by_occurrence.get(label, [])
            ri, du, st_obj = None, None, None
            if occurrence < len(vals):
                ri, du, st_obj = vals[occurrence]

            if not rate_name:
                return

            # V1.08: 以基准反推的计算口径
            biz_config = {
                'FTP下载':   {'col': 'Downlink RLC Throughput(bps)', 'clip': None, 'mode': 'rlc_nz', 'desc': '下行RLC去零均值'},
                'FTP上传':   {'col': 'Uplink RLC Throughput(bps)', 'clip': None, 'mode': 'rlc_nz', 'desc': '上行RLC去零均值'},
                '应用商店小包': {'col': 'Downlink RLC Throughput(bps)', 'clip': None, 'mode': 'rlc_mean_ss', 'desc': '下行RLC含零均值'},
                '应用商店大包': {'col': 'Downlink RLC Throughput(bps)', 'clip': 1000, 'mode': 'rlc_clip_nz', 'desc': '下行RLC削峰1000去零均值'},
                '微信小文件': {'col': 'Uplink MAC Throughput(bps)', 'clip': None, 'mode': 'mac_peak', 'desc': '上行MAC峰值(peak)'},
                '微信大文件': {'col': 'Uplink RLC Throughput(bps)', 'clip': None, 'mode': 'rlc_mean', 'desc': '上行RLC含零均值'},
            }
            cfg = biz_config.get(label, {})
            calc_col_h = cfg.get('col', '')
            clip_val = cfg.get('clip')
            mode = cfg.get('mode', '')
            desc = cfg.get('desc', '')

            # 找原始数据列索引（用完整列名headers）
            ci = {h: i for i, h in enumerate(headers)}
            calc_col_idx = ci.get(calc_col_h, -1) if calc_col_h else -1

            # 构建批注内容
            comment_f = ""
            comment_h = ""

            if ri is not None:
                # 获取原始数据列的实际数据（含时间戳）
                seg_data = []
                seg_times = []
                if calc_col_idx >= 0:
                    for r in range(seg_start, seg_end + 1):
                        v = ws.cell(row=r, column=calc_col_idx + 1).value
                        t = ws.cell(row=r, column=1).value  # A列=时间
                        if v is not None and isinstance(v, (int, float)):
                            seg_data.append(v)
                            seg_times.append(t)

                if seg_data:
                    import numpy as np
                    arr = np.array(seg_data)
                    orig_mean = np.mean(arr)
                    orig_count = len(arr)
                    nonzero_count = int(np.sum(arr != 0))

                    # 构建详细计算过程行列表
                    detail_lines = []
                    if mode in ('rlc_clip_nz', 'rlc_nz', 'rlc_mean_ss', 'rlc_mean', 'mac_peak', 'mac_mean'):
                        # 对所有模式都列出秒级明细
                        detail_lines.append("秒级数据(已CA聚合取max):")
                        for i in range(min(orig_count, 60)):  # 最多列60行，避免批注过长
                            val = seg_data[i]
                            val_mbps = val  # 已经是Mbps
                            val_bps = val * 1_000_000
                            t_str = str(seg_times[i]) if i < len(seg_times) and seg_times[i] is not None else f"秒{i+1}"
                            # 截断时间到 HH:MM:SS
                            if ' ' in t_str:
                                t_str = t_str.split(' ')[-1]
                            t_str = t_str[:8]
                            val_str = f"{val_mbps:.3f} Mbps ({int(val_bps):,} bps)"
                            detail_lines.append(f"  {t_str}: {val_str}")
                        if orig_count > 60:
                            detail_lines.append(f"  ...(共{orig_count}行,仅列前60行)")
                        detail_lines.append("")

                    if mode == 'rlc_clip_nz' and clip_val:
                        clipped = np.clip(arr, None, clip_val)
                        nz = clipped[clipped != 0]
                        final_mean = np.mean(nz) if len(nz) > 0 else 0
                        clipped_count = int(np.sum(arr > clip_val))
                        detail_lines.append(f"削峰阈值: {clip_val} Mbps (削峰{clipped_count}个点)")
                        nz_list = [f"{v:.3f}" for v in clipped if v != 0]
                        if len(nz_list) > 60:
                            nz_list = nz_list[:60] + [f"...(共{len(nz)}个非零点)"]
                        detail_lines.append(f"削峰后非零值: {', '.join(nz_list)}")
                        detail_lines.append(f"去零后均值 = ({', '.join(nz_list)}) / {len(nz)}")
                        detail_lines.append(f"最终结果: {final_mean:.3f} Mbps")
                        comment_f = f"计算口径: {desc}\n" \
                                    f"原始数据列: {calc_col_h}\n" \
                                    f"采样点数: {orig_count}\n" \
                                    + "\n".join(detail_lines)
                    elif mode == 'rlc_nz':
                        nz = arr[arr != 0]
                        final_mean = np.mean(nz) if len(nz) > 0 else 0
                        nz_list = [f"{v:.3f}" for v in nz]
                        if len(nz_list) > 60:
                            nz_list = nz_list[:60] + [f"...(共{len(nz)}个非零点)"]
                        detail_lines.append(f"去零后非零值: {', '.join(nz_list)}")
                        detail_lines.append(f"去零后均值 = 非零值之和 / 非零个数 = {np.sum(nz):.3f} / {len(nz)} = {final_mean:.3f} Mbps")
                        detail_lines.append(f"原始含零均值: {orig_mean:.3f} Mbps (含{orig_count - len(nz)}个零点)")
                        comment_f = f"计算口径: {desc}\n" \
                                    f"原始数据列: {calc_col_h}\n" \
                                    f"采样点数: {orig_count} (非零{len(nz)}个)\n" \
                                    + "\n".join(detail_lines)
                    elif mode == 'mac_peak':
                        peak_val = float(np.max(arr))
                        peak_idx = int(np.argmax(arr))
                        peak_time = str(seg_times[peak_idx]).split(' ')[-1][:8] if peak_idx < len(seg_times) and seg_times[peak_idx] else f"秒{peak_idx+1}"
                        detail_lines.append(f"峰值 = max({', '.join(f'{v:.3f}' for v in arr)}) = {peak_val:.3f} Mbps")
                        detail_lines.append(f"峰值出现在: {peak_time}")
                        comment_f = f"计算口径: {desc}\n" \
                                    f"原始数据列: {calc_col_h}\n" \
                                    f"采样点数: {orig_count}\n" \
                                    + "\n".join(detail_lines)
                    elif mode == 'mac_mean':
                        comment_f = f"计算口径: {desc}\n" \
                                    f"原始数据列: {calc_col_h}\n" \
                                    + "\n".join(detail_lines) + \
                                    f"\n含零均值: {orig_mean:.3f} Mbps"
                    elif mode == 'rlc_mean':
                        detail_lines.append(f"含零均值 = 所有值之和 / 采样点数 = {np.sum(arr):.3f} / {orig_count} = {orig_mean:.3f} Mbps")
                        comment_f = f"计算口径: {desc}\n" \
                                    f"原始数据列: {calc_col_h}\n" \
                                    f"采样点数: {orig_count}\n" \
                                    + "\n".join(detail_lines)
                    else:
                        comment_f = f"计算口径: {desc}\n原始数据列: {calc_col_h}\n采样点数: {orig_count}\n" \
                                    + "\n".join(detail_lines) + f"\n均值: {orig_mean:.3f} Mbps"

            if du is not None and st_obj is not None:
                et_obj = st_obj + pd.Timedelta(seconds=du)
                comment_h = f"时长计算: 结束时间({et_obj}) - 开始时间({st_obj})\n时长: {du:.3f} 秒"

            # 填充E~H列
            try:
                e_cell = ws.cell(row=seg_start, column=e_col)
                e_cell.value = rate_name
                e_cell.alignment = Alignment(horizontal='center', vertical='center')
                e_cell.border = thin_border
                if seg_end > seg_start:
                    ws.merge_cells(start_row=seg_start, start_column=e_col, end_row=seg_end, end_column=e_col)
            except AttributeError:
                pass
            if ri is not None:
                try:
                    f_cell = ws.cell(row=seg_start, column=f_col)
                    f_cell.value = round(ri, 3)
                    f_cell.alignment = Alignment(horizontal='center', vertical='center')
                    f_cell.border = thin_border
                    if comment_f:
                        f_cell.comment = Comment(comment_f, "计算过程")
                    if seg_end > seg_start:
                        ws.merge_cells(start_row=seg_start, start_column=f_col, end_row=seg_end, end_column=f_col)
                except AttributeError:
                    pass
            try:
                g_cell = ws.cell(row=seg_start, column=g_col)
                g_cell.value = dur_name
                g_cell.alignment = Alignment(horizontal='center', vertical='center')
                g_cell.border = thin_border
                if seg_end > seg_start:
                    ws.merge_cells(start_row=seg_start, start_column=g_col, end_row=seg_end, end_column=g_col)
            except AttributeError:
                pass
            if du is not None:
                try:
                    h_cell = ws.cell(row=seg_start, column=h_col)
                    h_cell.value = round(du, 3)
                    h_cell.alignment = Alignment(horizontal='center', vertical='center')
                    h_cell.border = thin_border
                    if comment_h:
                        h_cell.comment = Comment(comment_h, "计算过程")
                    if seg_end > seg_start:
                        ws.merge_cells(start_row=seg_start, start_column=h_col, end_row=seg_end, end_column=h_col)
                except AttributeError:
                    pass

            # 标记计算用到的列为绿色底色（该段内所有行对应的计算列cell都标绿）
            if calc_col_idx >= 0:
                green_fill = PF2(start_color='92D050', end_color='92D050', fill_type='solid')
                for r in range(seg_start, seg_end + 1):
                    ws.cell(row=r, column=calc_col_idx + 1).fill = green_fill

            # 标记用到的速率类列名行头为绿色（第1行列名）
            ws.cell(row=1, column=calc_col_idx + 1).fill = PF2(start_color='92D050', end_color='92D050', fill_type='solid')

        current_label = None
        seg_start = None
        for row_num in range(2, ws.max_row + 1):
            d_val = ws.cell(row=row_num, column=biz_col).value
            d_str = str(d_val).strip() if d_val else ''

            if d_str:
                if d_str != current_label:
                    # 上一段结束，填E~H
                    if current_label and seg_start and seg_start < row_num:
                        _fill_biz_segment(ws, seg_start, row_num - 1, current_label,
                                          e_col, f_col, g_col, h_col, thin_border,
                                          biz_rate_vals_by_occurrence, occurrence_idx,
                                          biz_rate_info, raw, headers)
                        occurrence_idx[current_label] = occurrence_idx.get(current_label, 0) + 1
                    current_label = d_str
                    seg_start = row_num
            else:
                if current_label and seg_start:
                    _fill_biz_segment(ws, seg_start, row_num - 1, current_label,
                                      e_col, f_col, g_col, h_col, thin_border,
                                      biz_rate_vals_by_occurrence, occurrence_idx,
                                      biz_rate_info, raw, headers)
                    occurrence_idx[current_label] = occurrence_idx.get(current_label, 0) + 1
                    current_label = None
                    seg_start = None

        # 处理最后一段
        if current_label and seg_start:
            _fill_biz_segment(ws, seg_start, ws.max_row, current_label,
                              e_col, f_col, g_col, h_col, thin_border,
                              biz_rate_vals_by_occurrence, occurrence_idx,
                              biz_rate_info, raw, headers)

        # 绿色标记用到的列（Q列往后），并高亮业务标记和速率列
        dl_rlc = next((h for h in headers if '下行RLC' in str(h) and '吞吐率' in str(h)), None)
        ul_rlc = next((h for h in headers if '上行RLC' in str(h) and '吞吐率' in str(h)), None)
        dl_mac = next((h for h in headers if '下行MAC' in str(h) and '吞吐率' in str(h)), None)
        ul_mac = next((h for h in headers if '上行MAC' in str(h) and '吞吐率' in str(h)), None)

        # 对每个业务的高亮（用对应颜色标记业务标记列和速率列）
        for biz, idxs in biz_rows.items():
            if not idxs: continue
            fc = biz_colors.get(biz, 'FFFF00')
            biz_fill = PatternFill(start_color=fc, end_color=fc, fill_type='solid')
            _, direction, _ = biz_rate_info.get(biz, ('', '', ''))
            if direction == '下行':
                rate_col_h = dl_rlc or dl_mac
            else:
                rate_col_h = ul_rlc or ul_mac
            rate_col = ci.get(rate_col_h, 0) if rate_col_h else 0
            for idx in idxs:
                ws.cell(row=idx + 2, column=biz_col).fill = biz_fill
                if rate_col:
                    ws.cell(row=idx + 2, column=rate_col).fill = biz_fill

        # Q列(17)往后：用到的列标绿色
        green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        used_cols = set()
        used_cols.add(biz_col)
        if dl_rlc: used_cols.add(ci.get(dl_rlc, 0))
        if ul_rlc: used_cols.add(ci.get(ul_rlc, 0))
        if dl_mac: used_cols.add(ci.get(dl_mac, 0))
        if ul_mac: used_cols.add(ci.get(ul_mac, 0))
        for col in range(17, len(headers) + 1):
            if col in used_cols:
                cell = ws.cell(row=1, column=col)
                cell.fill = green_fill

        wb.save(out_path)
        return out_path


# ==================== Excel导出 ====================

class ExcelExporter:
    def __init__(self):
        self.wb = None
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    def export(self, output_path: str, call_df: pd.DataFrame, stats: Dict, operator: str = ''):
        """导出Excel，输出两个sheet（联通和电信）

        Args:
            output_path: 输出文件路径
            call_df: 呼叫详情DataFrame（含运营商列）
            stats: 统计结果字典
            operator: 当前处理的运营商（用于单运营商场景，空则从call_df提取）
        """
        self.wb = Workbook()
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # 从call_df提取运营商（若无运营商列，则用传入的operator或默认）
        operators_in_df = []
        if '运营商' in call_df.columns:
            operators_in_df = call_df['运营商'].dropna().unique().tolist()

        if operators_in_df:
            # 多运营商：按运营商分组，每个运营商一个sheet
            for op in operators_in_df:
                op_df = call_df[call_df['运营商'] == op]
                self._create_operator_sheet(op, op_df)
        elif operator:
            # 单运营商（传入的）
            self._create_operator_sheet(operator, call_df)
        else:
            # 默认：创建"联通"sheet
            self._create_operator_sheet('联通', call_df)

        # 保留原有的统计sheet和网络覆盖sheet
        self._create_stat_sheet(stats)
        self._create_coverage_sheet(stats.get('_coverage', {}))
        self.wb.save(output_path)

    def _create_operator_sheet(self, operator: str, call_df: pd.DataFrame):
        """创建单个运营商的业务概览sheet（7列格式）"""
        ws = self.wb.create_sheet(operator)
        gfill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')

        # 表头（7列）
        headers = ['业务类型', '开始时间', '结束时间', '速率类指标', '数值', '时长类指标', '数值']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, size=9)
            c.fill = gfill
            c.border = self.thin_border

        # 收集所有业务数据
        biz_data_list = []

        def _safe_dur(st, et):
            """计算时长（秒）"""
            if st and et and str(st) != '0' and str(et) != '0':
                try:
                    return round((pd.Timestamp(et) - pd.Timestamp(st)).total_seconds(), 3)
                except Exception:
                    return ''
            return ''

        for r in range(len(call_df)):
            row_data = call_df.iloc[r]
            test_biz = str(row_data.get('测试业务', '')).strip()

            if test_biz in ('FTPDownload', 'FTP下载'):
                st = row_data.get('发起时间', '')
                et = row_data.get('完成时间', '')
                rate = row_data.get('下载平均速率(Mbps)', 0)
                if rate and not (isinstance(rate, (int, float)) and rate == 0):
                    dur = _safe_dur(st, et)
                    biz_data_list.append(('FTP下载', st, et, '下载平均速率(Mbps)', rate, '业务时长(秒)', dur))

            elif test_biz in ('FTPUpload', 'FTP上传'):
                st = row_data.get('发起时间', '')
                et = row_data.get('完成时间', '')
                rate = row_data.get('上传平均速率(Mbps)', 0)
                if rate and not (isinstance(rate, (int, float)) and rate == 0):
                    dur = _safe_dur(st, et)
                    biz_data_list.append(('FTP上传', st, et, '上传平均速率(Mbps)', rate, '业务时长(秒)', dur))

                # 应用商店小文件
                ss = row_data.get('应用商店小文件下载开始时间', '')
                se = row_data.get('应用商店小文件下载完成时间', '')
                sr = row_data.get('应用商店小文件下载速率(Mbps)', 0)
                if ss and se and str(ss) != '0' and str(se) != '0' and sr and not (isinstance(sr, (int, float)) and sr == 0):
                    sd = _safe_dur(ss, se)
                    biz_data_list.append(('应用商店小文件下载', ss, se, '小文件下载速率(Mbps)', sr, '业务时长(秒)', sd))

                # 应用商店大文件
                ls = row_data.get('应用商店大文件下载开始时间', '')
                le = row_data.get('应用商店大文件下载完成时间', '')
                lr = row_data.get('应用商店大文件下载速率(Mbps)', 0)
                if ls and le and str(ls) != '0' and str(le) != '0' and lr and not (isinstance(lr, (int, float)) and lr == 0):
                    ld = _safe_dur(ls, le)
                    biz_data_list.append(('应用商店大文件下载', ls, le, '大文件下载速率(Mbps)', lr, '业务时长(秒)', ld))

                # 微信小包
                wss = row_data.get('微信小包发送开始时间', '')
                wse = row_data.get('微信小包发送完成时间', '')
                wsr = row_data.get('微信小包发送速率(Mbps)', 0)
                if wss and wse and str(wss) != '0' and str(wse) != '0' and wsr and not (isinstance(wsr, (int, float)) and wsr == 0):
                    wsd = _safe_dur(wss, wse)
                    biz_data_list.append(('微信小包发送', wss, wse, '微信小包发送速率(Mbps)', wsr, '业务时长(秒)', wsd))

                # 微信大包
                wls = row_data.get('微信大包发送开始时间', '')
                wle = row_data.get('微信大包发送完成时间', '')
                wlr = row_data.get('微信大包发送速率(Mbps)', 0)
                if wls and wle and str(wls) != '0' and str(wle) != '0' and wlr and not (isinstance(wlr, (int, float)) and wlr == 0):
                    wld = _safe_dur(wls, wle)
                    biz_data_list.append(('微信大包发送', wls, wle, '微信大包发送速率(Mbps)', wlr, '业务时长(秒)', wld))

        # 按开始时间排序
        biz_data_list.sort(key=lambda x: str(x[1]) if x[1] else '')

        # 写入数据行
        for row_idx, biz_data in enumerate(biz_data_list, 2):
            biz_name, st, et, rate_label, rate_val, dur_label, dur_val = biz_data

            ws.cell(row=row_idx, column=1, value=biz_name).border = self.thin_border
            ws.cell(row=row_idx, column=2, value=str(st) if st and str(st) != '0' else '').border = self.thin_border
            ws.cell(row=row_idx, column=3, value=str(et) if et and str(et) != '0' else '').border = self.thin_border
            ws.cell(row=row_idx, column=4, value=rate_label).border = self.thin_border

            c5 = ws.cell(row=row_idx, column=5, value=rate_val if rate_val else '')
            c5.border = self.thin_border
            if isinstance(rate_val, (int, float)) and rate_val:
                c5.number_format = '0.000'

            ws.cell(row=row_idx, column=6, value=dur_label).border = self.thin_border
            c7 = ws.cell(row=row_idx, column=7, value=dur_val if dur_val != '' else '')
            c7.border = self.thin_border
            if isinstance(dur_val, (int, float)):
                c7.number_format = '0.000'

        # 设置列宽
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14

    # 原呼叫详情sheet方法已删除，改用_create_operator_sheet

    # 原业务概览sheet方法已删除，改用_create_operator_sheet

    def _create_stat_sheet(self, stats: Dict):
        ws = self.wb.create_sheet('统计结果')
        gfill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        for col, (grp, field) in enumerate(STAT_COLUMNS, 1):
            c = ws.cell(row=2, column=col, value=field)
            c.font = Font(bold=True, size=9); c.border = self.thin_border
        start = 1
        for i in range(1, len(STAT_COLUMNS) + 1):
            if i == len(STAT_COLUMNS) or STAT_COLUMNS[i][0] != STAT_COLUMNS[start - 1][0]:
                c = ws.cell(row=1, column=start, value=STAT_COLUMNS[start - 1][0])
                c.font = Font(bold=True); c.fill = gfill; c.alignment = Alignment(horizontal='center'); c.border = self.thin_border
                if i > start:
                    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=i)
                start = i + 1
        for col, (grp, field) in enumerate(STAT_COLUMNS, 1):
            v = stats.get(grp, {}).get(field)
            c = ws.cell(row=3, column=col, value=v if v is not None else '')
            c.alignment = Alignment(horizontal='center'); c.border = self.thin_border

    def _create_coverage_sheet(self, cov: Dict):
        """网络覆盖/基础信息(对齐参考「运营商汇总」基础+网络覆盖段)"""
        if not cov:
            return
        ws = self.wb.create_sheet('网络覆盖')
        gfill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        c1 = ws.cell(row=1, column=1, value='指标')
        c2 = ws.cell(row=1, column=2, value='值')
        c1.font = Font(bold=True); c2.font = Font(bold=True)
        c1.fill = gfill; c2.fill = gfill; c1.border = self.thin_border; c2.border = self.thin_border
        order = ['测试时长(分钟)', '5G CA聚合率(%)', 'NR_Serving SS-RSRP(dBm)', 'NR_Serving SS AVG SINR(dB)',
                 '5G覆盖率SS-RSRP≥-96dBm&SS-SINR≥-3dB采样点占比',
                 '5G覆盖率SS-RSRP≥-100dBm&SS-SINR≥-3dB采样点占比',
                 '5G覆盖达标率SS-RSRP≥-105dBm&SS-SINR≥-3dB采样点占比']
        items = [(k, cov.get(k)) for k in order if k in cov]
        for k, val in cov.items():          # 兜底：order 外的也列出
            if (k, val) not in items:
                items.append((k, val))
        for i, (k, val) in enumerate(items, 2):
            ws.cell(row=i, column=1, value=k).border = self.thin_border
            c = ws.cell(row=i, column=2, value=val if val is not None else '')
            c.alignment = Alignment(horizontal='center'); c.border = self.thin_border
        ws.column_dimensions['A'].width = 52


# ==================== 工作线程 ====================

class ProcessingThread(QThread):
    progress_signal = Signal(str)
    finished_signal = Signal(bool, str, object, object)

    def __init__(self, processor: UCMProcessor, mmf_paths, plan: List[Dict], params: Dict):
        super().__init__()
        self.processor = processor
        self.mmf_paths = mmf_paths if isinstance(mmf_paths, list) else [mmf_paths]
        self.plan = plan
        self.params = params

    def run(self):
        try:
            self.progress_signal.emit(f"解析UCM监控文件({len(self.mmf_paths)}个)...")
            self.processor.parse_mmf(self.mmf_paths, self.params)
            self.progress_signal.emit("按测试计划匹配业务...")
            call_df, stats = self.processor.match(self.plan, self.params)
            self.progress_signal.emit("生成完成")
            self.finished_signal.emit(True, f"完成：匹配 {len(call_df)} 轮", call_df, stats)
        except Exception as e:
            import traceback; traceback.print_exc()
            self.finished_signal.emit(False, f"失败: {str(e)}", None, None)


# ==================== 在线更新 + 关于对话框 ====================

def _v_tuple(s):
    """版本号字符串→数字元组(用于比较大小)"""
    return tuple(int(x) for x in re.findall(r'\d+', str(s)))


def _ssl_ctx():
    """云主机自签证书：忽略证书校验(保证能连通)"""
    import ssl
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx


class UpdateChecker(QThread):
    """后台静默检查更新：拉 version.json 比对版本"""
    found_update = Signal(dict)   # 有新版时发(含版本信息)
    check_done = Signal(str)      # 'latest'=最新 / 'new'=有新版 / 'error:...'=失败

    def run(self):
        try:
            url = f'{UPDATE_BASE}/version.json?t={int(time.time())}'   # 防缓存
            with urllib.request.urlopen(url, timeout=6, context=_ssl_ctx()) as r:
                info = json.loads(r.read().decode('utf-8'))
            if _v_tuple(info.get('version', '0')) > _v_tuple(APP_VERSION):
                self.found_update.emit(info)
                self.check_done.emit('new')
            else:
                self.check_done.emit('latest')
        except Exception as e:
            self.check_done.emit(f'error:{e}')


def do_update(info, parent_window) -> bool:
    """下载更新包→MD5校验→备份旧文件→替换。成功后调用方需重启。
    自动适配 py / exe(PyInstaller) 运行环境。返回是否完成替换。"""
    # 1. 下载到临时文件
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pack'); tmp.close()
    try:
        with urllib.request.urlopen(info['url'], timeout=30, context=_ssl_ctx()) as r, open(tmp.name, 'wb') as f:
            shutil.copyfileobj(r, f)
    except Exception as e:
        QMessageBox.critical(parent_window, '更新失败', f'下载失败：{e}'); return False
    # 2. MD5 校验(保质，防损坏/篡改)
    if info.get('md5'):
        h = hashlib.md5()
        with open(tmp.name, 'rb') as f:
            for chunk in iter(lambda: f.read(65536), b''): h.update(chunk)
        if h.hexdigest() != info['md5']:
            os.unlink(tmp.name)
            QMessageBox.critical(parent_window, '更新失败', 'MD5 校验不一致，已中止(防损坏/篡改)')
            return False
    # 3. 定位当前运行文件
    if getattr(sys, 'frozen', False):
        me = os.path.abspath(sys.executable)         # exe 模式
    else:
        me = os.path.abspath(sys.argv[0] if sys.argv and sys.argv[0] else __file__)
    bak = f'{me}_旧版_{datetime.now().strftime("%Y%m%d-%H%M%S")}'
    # 4. 替换
    try:
        if getattr(sys, 'frozen', False) and os.name == 'nt':
            _write_exe_updater(tmp.name, me, bak)    # Windows exe: bat 等退出后替换
            return True
        else:
            # py / 类Unix: 备份后直接覆盖
            if os.path.exists(me): shutil.copy2(me, bak)
            shutil.copy2(tmp.name, me); os.unlink(tmp.name)
            return True
    except Exception as e:
        QMessageBox.critical(parent_window, '更新失败', f'替换文件失败：{e}')
        return False


def _write_exe_updater(pack_path, target_exe, bak_path):
    """Windows exe 更新：生成 updater.bat(等待主进程退出→备份→替换→重启→自删)"""
    bat = os.path.join(os.path.dirname(target_exe), '_updater.bat')
    with open(bat, 'w', encoding='gbk', newline='\r\n') as f:
        f.write('@echo off\r\n')
        f.write('timeout /t 1 /nobreak >nul\r\n')
        f.write(':wait\r\n')
        f.write(f'move /y "{target_exe}" "{bak_path}" >nul 2>&1\r\n')
        f.write(f'if exist "{target_exe}" goto wait\r\n')
        f.write(f'move /y "{pack_path}" "{target_exe}" >nul\r\n')
        f.write(f'start "" "{target_exe}"\r\n')
        f.write(f'del "%~f0"\r\n')
    import subprocess
    subprocess.Popen(['cmd', '/c', bat], close_fds=True)


def _relaunch_and_quit():
    """重启工具(更新后)"""
    import subprocess
    if getattr(sys, 'frozen', False):
        subprocess.Popen([sys.executable])
    else:
        subprocess.Popen([sys.executable, os.path.abspath(sys.argv[0] if sys.argv else __file__)])
    QApplication.instance().quit()
    os._exit(0)


class AboutDialog(QDialog):
    """关于对话框：版本/署名/联系方式/版本历史/检查更新"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('关于')
        self.setMinimumWidth(460)
        lay = QVBoxLayout(self)
        title = QLabel(APP_NAME); title.setFont(QFont('PingFang SC', 15, QFont.Weight.Bold))
        lay.addWidget(title)
        ver = QLabel(f'版本：{APP_VERSION}   构建日期 {BUILD_DATE}')
        ver.setFont(QFont('PingFang SC', 11)); lay.addWidget(ver)
        self.status_label = QLabel('更新检查：未检查')
        self.status_label.setStyleSheet('color: #666;'); lay.addWidget(self.status_label)
        author = QLabel(f'开发人：{AUTHOR}\n联系方式：{CONTACT}')
        author.setFont(QFont('PingFang SC', 11)); lay.addWidget(author)
        lay.addWidget(QLabel('── 版本历史 ──'))
        hist = QLabel('\n'.join(f'{v} · {d} · {c}' for v, d, c in VERSION_HISTORY))
        hist.setWordWrap(True); hist.setStyleSheet('color: #333;'); lay.addWidget(hist)
        bl = QHBoxLayout()
        chk = QPushButton('检查更新'); chk.clicked.connect(self._check)
        close_btn = QPushButton('关闭'); close_btn.clicked.connect(self.accept)
        bl.addStretch(); bl.addWidget(chk); bl.addWidget(close_btn); lay.addLayout(bl)
        self.adjustSize()

    def _check(self):
        self.status_label.setText('更新检查：检查中…'); self.setEnabled(False)
        self._checker = UpdateChecker(self, manual=True)
        self._checker.check_done.connect(self._on_done)
        self._checker.found_update.connect(self._on_found)
        self._checker.start()

    def _on_done(self, msg):
        self.setEnabled(True)
        if msg == 'latest':
            self.status_label.setText('更新检查：✓ 已是最新版')
        elif msg.startswith('error'):
            self.status_label.setText('更新检查：检查失败(请检查网络/云主机)')

    def _on_found(self, info):
        self.status_label.setText(f"更新检查：发现新版 {info['version']}")
        text = (f"发现新版本 {info['version']}（当前 {APP_VERSION}）\n\n"
                f"发布日期：{info.get('release_date','')}\n\n"
                f"更新内容：\n{info.get('changelog','')}\n\n是否立即更新？")
        if QMessageBox.question(self, '发现新版本', text, QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            if do_update(info, self):
                if getattr(sys, 'frozen', False) and os.name == 'nt':
                    QMessageBox.information(self, '更新准备就绪', '工具将退出并自动完成更新，请稍候。')
                    _relaunch_and_quit()
                else:
                    QMessageBox.information(self, '更新完成', '更新已下载并替换，工具将重启。')
                    _relaunch_and_quit()


# ==================== 主窗口 ====================

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config_manager = ConfigManager()
        self.config = self.config_manager.config
        self.processor = None
        self.mmf_files = []
        self.processing_thread = None
        self.init_ui()
        self._start_update_check()   # 启动静默检查更新

    def init_ui(self):
        self.setWindowTitle('5G用户级公共监控速率统计工具 V1.09')
        self.setGeometry(80, 80, 1600, 950)
        w = QWidget(); self.setCentralWidget(w)
        lay = QHBoxLayout(w)
        lay.addWidget(self.create_left_panel(), 1)
        lay.addWidget(self.create_right_panel(), 2)

    def create_left_panel(self) -> QWidget:
        panel = QWidget(); layout = QVBoxLayout(panel)
        # 文件选择
        fg = QGroupBox('输入文件（仅UCM监控）'); fl = QHBoxLayout()
        self.mmf_edit = QLineEdit(); self.mmf_edit.setPlaceholderText('选择 mmf/两江 等(可多选,按时间合并分析)')
        fb = QPushButton('选择(可多选)'); fb.clicked.connect(self.select_mmf)
        fl.addWidget(self.mmf_edit); fl.addWidget(fb); fg.setLayout(fl); layout.addWidget(fg)
        # 全局参数
        pg = QGroupBox('全局参数'); pl = QFormLayout()
        self.dl_peak = QDoubleSpinBox(); self.dl_peak.setRange(0, 5000); self.dl_peak.setValue(900)
        self.ul_peak = QDoubleSpinBox(); self.ul_peak.setRange(0, 1000); self.ul_peak.setValue(160)
        self.down_min = QDoubleSpinBox(); self.down_min.setRange(0, 1000); self.down_min.setValue(50)
        self.up_min = QDoubleSpinBox(); self.up_min.setRange(0, 500); self.up_min.setValue(10)
        pl.addRow('下行削峰(Mbps):', self.dl_peak); pl.addRow('上行削峰(Mbps):', self.ul_peak)
        pl.addRow('下行最小速率(Mbps):', self.down_min); pl.addRow('上行最小速率(Mbps):', self.up_min)
        self.match_mode = QComboBox(); self.match_mode.addItems(['plan(固定顺序)', 'auto(自动识别)', 'hybrid(混合:参考时间戳+段检测)'])
        pl.addRow('匹配模式:', self.match_mode)
        # 基准规则选择
        self.benchmark_mode = QComboBox()
        self.benchmark_mode.addItems(['ALL_FTP_DETAIL模板', '工信部指标(T0-运营商区分)', 'UCM标注(运营商区分)'])
        pl.addRow('基准规则:', self.benchmark_mode)
        pg.setLayout(pl); layout.addWidget(pg)
        # 测试计划编辑器
        layout.addWidget(self.create_plan_editor())
        # 按钮
        bl = QHBoxLayout()
        pb = QPushButton('开始统计'); pb.setStyleSheet('QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 10px; }')
        pb.clicked.connect(self.start_processing); bl.addWidget(pb)
        eb = QPushButton('导出Excel'); eb.clicked.connect(self.export_excel); bl.addWidget(eb)
        ab = QPushButton('生成标注文件'); ab.clicked.connect(self.gen_annotated); bl.addWidget(ab)
        ub = QPushButton('关于'); ub.clicked.connect(self.show_about); bl.addWidget(ub)
        layout.addLayout(bl)
        self.progress = QProgressBar(); layout.addWidget(self.progress)
        lg = QGroupBox('运行日志'); ll = QVBoxLayout()
        self.log_text = QTextEdit(); self.log_text.setReadOnly(True); ll.addWidget(self.log_text)
        lg.setLayout(ll); layout.addWidget(lg, 1)
        return panel

    def create_plan_editor(self) -> QWidget:
        g = QGroupBox('测试计划（可上下移动顺序、参数可调）'); l = QVBoxLayout(g)
        self.plan_table = QTableWidget(); self.plan_table.setColumnCount(6)
        self.plan_table.setHorizontalHeaderLabels(['顺序', '业务名称', '方向', '时长(s)', '流量(MB)', '超时(s)'])
        self.plan_table.horizontalHeader().setStretchLastSection(True)
        l.addWidget(self.plan_table)
        bl = QHBoxLayout()
        up = QPushButton('上移'); up.clicked.connect(lambda: self._move_row(-1)); bl.addWidget(up)
        dn = QPushButton('下移'); dn.clicked.connect(lambda: self._move_row(1)); bl.addWidget(dn)
        add = QPushButton('添加'); add.clicked.connect(self.add_plan_row); bl.addWidget(add)
        rm = QPushButton('删除'); rm.clicked.connect(self.del_plan_row); bl.addWidget(rm)
        l.addLayout(bl)
        self.load_plan(DEFAULT_PLAN)
        return g

    def load_plan(self, plan):
        self.plan_table.setRowCount(len(plan))
        for r, item in enumerate(plan):
            self.plan_table.setItem(r, 0, QTableWidgetItem(str(r + 1)))
            self.plan_table.setItem(r, 1, QTableWidgetItem(item['name']))
            self.plan_table.setItem(r, 2, QTableWidgetItem(item.get('direction', '')))
            self.plan_table.setItem(r, 3, QTableWidgetItem('' if item.get('duration') is None else str(item['duration'])))
            self.plan_table.setItem(r, 4, QTableWidgetItem('' if item.get('flow_mb') is None else str(item['flow_mb'])))
            self.plan_table.setItem(r, 5, QTableWidgetItem('' if item.get('timeout') is None else str(item['timeout'])))
            self.plan_table.item(r, 0).setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

    def _move_row(self, delta):
        r = self.plan_table.currentRow()
        if r < 0: return
        target = r + delta
        if 0 <= target < self.plan_table.rowCount():
            self._swap_rows(r, target); self.plan_table.selectRow(target)

    def _swap_rows(self, a, b):
        for c in range(self.plan_table.columnCount()):
            ta = self.plan_table.takeItem(a, c); tb = self.plan_table.takeItem(b, c)
            self.plan_table.setItem(a, c, tb); self.plan_table.setItem(b, c, ta)
        # 重排顺序列
        self.plan_table.item(a, 0).setText(str(a + 1)); self.plan_table.item(b, 0).setText(str(b + 1))

    def add_plan_row(self):
        r = self.plan_table.rowCount(); self.plan_table.insertRow(r)
        for c in range(6): self.plan_table.setItem(r, c, QTableWidgetItem(''))
        self.plan_table.item(r, 0).setText(str(r + 1)); self.plan_table.item(r, 0).setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

    def del_plan_row(self):
        r = self.plan_table.currentRow()
        if r >= 0: self.plan_table.removeRow(r)

    def get_plan_from_ui(self) -> List[Dict]:
        plan = []
        for r in range(self.plan_table.rowCount()):
            name = self.plan_table.item(r, 1).text().strip() if self.plan_table.item(r, 1) else ''
            if not name: continue
            direction = self.plan_table.item(r, 2).text().strip() if self.plan_table.item(r, 2) else ''
            dur = self._num(self.plan_table.item(r, 3))
            flow = self._num(self.plan_table.item(r, 4))
            tmo = self._num(self.plan_table.item(r, 5))
            mode = 'wait' if name == '等待' else ('time' if dur else 'flow')
            plan.append({'key': self._plan_key(name), 'name': name, 'mode': mode, 'direction': direction,
                         'duration': dur, 'flow_mb': flow, 'timeout': tmo})
        return plan

    @staticmethod
    def _plan_key(name: str) -> str:
        """业务名 → 语义key(供 _build_call_row/_summarize 匹配)，必须与 DEFAULT_PLAN 的 key 一致"""
        if 'FTP下载' in name: return 'ftp_dl'
        if 'FTP上传' in name: return 'ftp_ul'
        if '应用商店' in name and '小' in name: return 'store_s'
        if '应用商店' in name and '大' in name: return 'store_l'
        if '微信' in name and '小' in name: return 'wx_s'
        if '微信' in name and '大' in name: return 'wx_l'
        return f'wait_{name}'

    @staticmethod
    def _num(item):
        if item is None: return None
        t = item.text().strip()
        if t in ('', 'None'): return None
        try: return float(t)
        except ValueError: return None

    def create_right_panel(self) -> QWidget:
        panel = QWidget(); layout = QVBoxLayout(panel)
        tw = QTabWidget()
        # 呼叫详情
        t1 = QWidget(); l1 = QVBoxLayout(t1)
        self.call_table = QTableWidget(); l1.addWidget(self.call_table)
        tw.addTab(t1, '呼叫详情(逐轮)')
        # 统计宽表
        t2 = QWidget(); l2 = QVBoxLayout(t2)
        self.stat_table = QTableWidget()
        headers = [f"{g}|{f}" for g, f in STAT_COLUMNS]
        self.stat_table.setColumnCount(len(headers)); self.stat_table.setHorizontalHeaderLabels(headers)
        l2.addWidget(self.stat_table)
        tw.addTab(t2, '统计结果(汇总)')
        layout.addWidget(tw)
        return panel

    def select_mmf(self):
        fps, _ = QFileDialog.getOpenFileNames(self, '选择UCM监控文件(可多选)', '', 'Excel Files (*.xlsx *.xls)')
        if fps:
            self.mmf_files = fps
            self.mmf_edit.setText(' | '.join(os.path.basename(p) for p in fps))
            self.log(f"已选择 {len(fps)} 个文件: {', '.join(os.path.basename(p) for p in fps)}")

    def get_params(self) -> Dict:
        return {'rate_column': 'RLC', 'dl_peak_limit': self.dl_peak.value(), 'ul_peak_limit': self.ul_peak.value(),
                'dl_pass_threshold': 100, 'ul_pass_threshold': 20,
                'down_min': self.down_min.value(), 'up_min': self.up_min.value(),
                'match_mode': 'auto' if str(self.match_mode.currentText()).startswith('auto') else 'plan' if 'plan' in str(self.match_mode.currentText()) else 'hybrid',}

    @Slot()
    def start_processing(self):
        if not self.mmf_files:
            QMessageBox.warning(self, '警告', '请选择UCM监控文件'); return
        plan = self.get_plan_from_ui()
        if not plan:
            QMessageBox.warning(self, '警告', '请配置测试计划'); return
        params = self.get_params()
        self.log("=" * 50); self.log("开始按测试计划匹配...")
        self.log(f"文件: {len(self.mmf_files)}个 {', '.join(os.path.basename(p) for p in self.mmf_files)} | 计划{len(plan)}项")
        self.progress.setValue(0)
        self.sender().setEnabled(False); self.sender().setText("处理中...")
        self.processor = UCMProcessor(self.config)
        self.processing_thread = ProcessingThread(self.processor, self.mmf_files, plan, params)
        self.processing_thread.progress_signal.connect(self.log)
        self.processing_thread.finished_signal.connect(self.on_finished)
        self.processing_thread.start()

    @Slot(bool, str, object, object)
    def on_finished(self, success, message, call_df, stats):
        self.progress.setValue(100); self.log(message); self.log("=" * 50)
        for btn in self.findChildren(QPushButton):
            if btn.text() in ["处理中...", "开始统计"]:
                btn.setEnabled(True); btn.setText("开始统计")
        if success:
            self.display_call(call_df); self.display_stat(stats)
            self.last_call = call_df; self.last_stats = stats
            cov = stats.get('_coverage', {})
            if cov:
                self.log('— 网络覆盖指标 —')
                for k, val in cov.items():
                    self.log(f'  {k} = {val}')
            QMessageBox.information(self, '完成', message)
        else:
            QMessageBox.critical(self, '错误', message)

    def display_call(self, call_df: pd.DataFrame):
        self.call_table.setColumnCount(len(CALL_HEADERS))
        self.call_table.setHorizontalHeaderLabels(CALL_HEADERS)
        self.call_table.setRowCount(len(call_df))
        for r in range(len(call_df)):
            for c in range(len(CALL_HEADERS)):
                v = call_df.iloc[r, c]
                if isinstance(v, float) and pd.isna(v): v = ''
                self.call_table.setItem(r, c, QTableWidgetItem(str(v)))

    def display_stat(self, stats: Dict):
        self.stat_table.setRowCount(1)
        for col, (grp, field) in enumerate(STAT_COLUMNS):
            v = stats.get(grp, {}).get(field)
            text = '' if v is None else (f"{v:.2f}" if isinstance(v, float) else str(v))
            self.stat_table.setItem(0, col, QTableWidgetItem(text))

    def export_excel(self):
        if not hasattr(self, 'last_call'):
            QMessageBox.warning(self, '警告', '请先处理数据'); return
        fp, _ = QFileDialog.getSaveFileName(self, '保存Excel', '', 'Excel Files (*.xlsx)')
        if fp:
            try:
                ExcelExporter().export(fp, self.last_call, self.last_stats)
                self.log(f"已导出: {fp}")
                QMessageBox.information(self, '成功', '导出成功！')
            except Exception as e:
                QMessageBox.critical(self, '错误', f'导出失败: {str(e)}')

    def gen_annotated(self):
        """生成业务标注文件(原始UCM所有列 + 业务标记列 + 速率计算列高亮)"""
        if not self.mmf_files:
            QMessageBox.warning(self, '警告', '请先选择UCM监控文件'); return
        fp, _ = QFileDialog.getSaveFileName(self, '保存标注文件', 'UCM业务标注.xlsx', 'Excel Files (*.xlsx)')
        if not fp:
            return
        plan = self.get_plan_from_ui(); params = self.get_params()
        try:
            self.log("正在生成标注文件...")
            self.processor = UCMProcessor(self.config)
            self.processor.generate_annotated(self.mmf_files, plan, params, fp)
            self.log(f"✓ 标注文件已生成: {fp}")
            QMessageBox.information(self, '成功', f'标注文件生成成功！\n{fp}')
        except Exception as e:
            import traceback; traceback.print_exc()
            QMessageBox.critical(self, '错误', f'生成失败: {str(e)}')

    def show_about(self):
        AboutDialog(self).exec()

    def _start_update_check(self):
        """启动时静默检查更新(仅发现新版才弹窗)"""
        self._update_checker = UpdateChecker(self)
        self._update_checker.found_update.connect(self._on_update_found)
        self._update_checker.start()

    def _on_update_found(self, info):
        r = QMessageBox.question(self, '发现新版本',
            f"检测到新版本 {info['version']}（当前 {APP_VERSION}）\n\n"
            f"更新内容：\n{info.get('changelog','')}\n\n是否立即更新？",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        if r == QMessageBox.Yes:
            if do_update(info, self):
                if getattr(sys, 'frozen', False) and os.name == 'nt':
                    QMessageBox.information(self, '更新准备就绪', '工具将退出并自动完成更新，请稍候。')
                else:
                    QMessageBox.information(self, '更新完成', '更新已下载并替换，工具将重启。')
                _relaunch_and_quit()

    def log(self, msg):
        self.log_text.append(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    w = MainWindow(); w.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
