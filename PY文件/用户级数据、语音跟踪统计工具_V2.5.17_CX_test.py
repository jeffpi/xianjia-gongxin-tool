#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
用户级数据、语音跟踪统计工具 V2.5.17_CX
GUI: 双标签(速率统计 / 语音VQI) → 选择文件(MMF+可选基准) → 处理 → 内嵌查看结果
CLI: python3 this.py --cmd
"""
import sys, os, re, json
import pandas as pd, numpy as np
from datetime import datetime, timedelta, time
from collections import Counter, defaultdict

# ===== 通用文件读取函数（支持xlsx/xls/csv，优先calamine，csv用read_csv）=====
def _read_file(f, nrows=None):
    """读取xlsx/xls/csv/mmf/tmf文件，返回DataFrame
    nrows: 可选，只读取前n行（用于快速判断文件类型）
    """
    low = f.lower()
    if low.endswith('.csv'):
        # 尝试多种编码读取csv
        for enc in ['gbk', 'utf-8', 'latin-1']:
            try:
                return pd.read_csv(f, encoding=enc, nrows=nrows)
            except Exception:
                continue
        return pd.read_csv(f, encoding='latin-1', on_bad_lines='skip', nrows=nrows)
    try:
        return pd.read_excel(f, engine='calamine', nrows=nrows)
    except Exception:
        try:
            return pd.read_excel(f, nrows=nrows)
        except Exception:
            try:
                return pd.read_excel(f, engine='openpyxl', nrows=nrows)
            except Exception:
                return pd.read_excel(f, engine='xlrd')

def _has_message_type_col(df):
    """判断DataFrame是否有MessageType列（支持多种列名）"""
    for col in ['MessageType', 'Message Type', 'message_type', '消息类型']:
        if col in df.columns:
            return True
    # 如果第二列是Source且有Time列，可能是mmf数据文件
    if len(df.columns) >= 2 and 'Time' in df.columns and 'Source' in df.columns:
        return False  # mmf数据文件
    return False

# ===== 预编译正则（性能优化）=====
_PARSE_RE = re.compile(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})[\.\(](\d+)')
_PARSE_MS_RE = re.compile(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\.(\d+)')

def parse_time(t):
    m = _PARSE_RE.match(str(t))
    if m:
        return datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S') + timedelta(milliseconds=int(m.group(2)))
    return None

def _parse_time_vec(s):
    """向量化解析时间列(替代 apply(parse_time),数万行从数十秒→毫秒级)。
    格式: '2026-07-19 18:27:28(123)' 或 '...18:27:28.123' → datetime+毫秒"""
    ext = s.astype(str).str.extract(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})[\.\(](\d+)')
    base = pd.to_datetime(ext[0], format='%Y-%m-%d %H:%M:%S', errors='coerce')
    ms = pd.to_numeric(ext[1], errors='coerce').fillna(0)
    return base + pd.to_timedelta(ms, unit='ms')

# ===== 计算批注辅助函数 =====
def _add_calc_comment(cell, formula_desc, calc_process, source_col, numerator_vals=None, denominator_vals=None, numerator_desc='分子', denominator_desc='分母'):
    """为单元格添加计算批注，显示公式、分子分母值"""
    if numerator_vals is None or len(numerator_vals) == 0:
        vs_str = '[]'
    else:
        vs = [str(round(v, 2)) for v in numerator_vals]
        if len(vs) <= 20:
            vs_str = '/'.join(vs)
        else:
            vs_str = '/'.join(vs[:5]) + '/.../' + '/'.join(vs[-5:])
    
    if denominator_vals is None or len(denominator_vals) == 0:
        dv_str = '[]'
    else:
        dv = [str(round(v, 2)) for v in denominator_vals]
        if len(dv) <= 20:
            dv_str = '/'.join(dv)
        else:
            dv_str = '/'.join(dv[:5]) + '/.../' + '/'.join(dv[-5:])
    
    txt = (f"公式: {formula_desc}\n"
           f"计算过程: {calc_process}\n\n"
           f"{numerator_desc}: {vs_str}\n"
           f"{denominator_desc}: {dv_str}\n\n"
           f"数据来源: {source_col} 列")
    from openpyxl.comments import Comment
    c = Comment(txt, "Tool")
    c.width = 500
    c.height = 250
    cell.comment = c

# ===== 阈值 =====
DL_ACTIVE_BIG, UL_ACTIVE_BIG = 50.0, 10.0
DL_ACTIVE_SMALL, UL_ACTIVE_SMALL = 5.0, 5.0
GAP_MERGE = 3.0
# FTP时长阈值（默认20秒，由ftp_duration参数动态调整）
FTP_PAIR_GAP = 90.0
STORE_L_DL_MIN, STORE_L_DL_MAX = 500, 99999
WX_L_DUR_MIN, WX_L_DUR_MAX, WX_L_UL_MIN = 12, 90, 40
DL_BIZ = {'FTP下载', '应用商店小文件下载', '应用商店大文件下载'}
UL_BIZ = {'FTP上传', '微信小包发送', '微信大包发送'}
REQ_BIZ = {'FTP下载', 'FTP上传', '应用商店小文件下载', '应用商店大文件下载', '微信小包发送', '微信大包发送'}


def detect_segments(ts, dl, ul, dl_th, ul_th, gap, r1=None, r2=None):
    active = (dl > dl_th) | (ul > ul_th); segs = []; i = 0; n = len(ts)
    while i < n:
        if r1 is not None and ts[i] < r1: i += 1; continue
        if r2 is not None and ts[i] > r2: break
        if active[i]:
            si = i; j = i + 1
            while j < n:
                if r2 is not None and ts[j] > r2: break
                if not active[j]:
                    k = j
                    while k < n and not active[k]:
                        if (ts[k] - ts[j - 1]) / np.timedelta64(1, 's') > gap: break
                        k += 1
                    if k < n and active[k] and (r2 is None or ts[k] <= r2): j = k; continue
                    else: break
                j += 1
            st = ts[si]; et = ts[j - 1]; dur = (et - st) / np.timedelta64(1, 's') + 1
            segs.append({'st': st, 'et': et, 'dur': dur, 'dl': float(dl[si:j].sum()), 'ul': float(ul[si:j].sum())})
            i = j
        else: i += 1
    return segs


def _extract_trace_id(path):
    """从文件路径中提取用户跟踪ID, 如 '用户跟踪ID=553' → '553'"""
    import re
    m = re.search(r'用户跟踪ID=(\d+)', str(path))
    return m.group(1) if m else None


def _lookup_phone(trace_id, phone_trace_map):
    """在phone_trace_map中精确查找trace_id对应的手机号"""
    if not trace_id or not phone_trace_map:
        return ''
    for entry in phone_trace_map:
        if entry['trace_id'] == trace_id:
            return entry['phone']
    return ''


def process(files, qci_list=None, dl_clip=1000, ul_clip=200, callback=None, cancel_check=None, progress_cb=None, base_file=None, time_filter=None, ftp_duration=25, merge_raw=True, add_annotations=True, phone_trace_map=None):
    if qci_list is None: qci_list = [5, 6, 7]

    # 根据ftp_duration动态调整FTP时长阈值
    # 用户选N秒 → (N-1.5, N+3) 的左右范围
    ftp_dur_min = ftp_duration - 3.0
    ftp_dur_max = ftp_duration + 3.0

    """统一的处理函数：从MMF文件生成5Sheet输出"""
    def log(msg, pct=None):
        if cancel_check and cancel_check():
            raise KeyboardInterrupt('用户取消')
        import time as _time
        _t = _time.time()
        if not hasattr(log, 'last_time'):
            log.last_time = _t
        elapsed = _t - log.last_time if pct is not None else 0
        print(msg)
        if callback: callback(msg)
        if pct is not None and progress_cb: progress_cb(pct)
        log.last_time = _t

    # 展开压缩文件(zip 自动解压，支持嵌套 zip-in-zip 递归解压)
    import zipfile, tempfile, shutil, glob
    _tmp_dirs = []
    def _expand_zip(zip_path):
        """递归解压 zip(含嵌套 zip)，返回所有 xlsx/xls 路径"""
        _tmp = tempfile.mkdtemp(prefix='mmf_unzip_'); _tmp_dirs.append(_tmp)
        _xs = []
        try:
            with zipfile.ZipFile(zip_path) as _zf: _zf.extractall(_tmp)
            _xs = glob.glob(os.path.join(_tmp, '**', '*.xlsx'), recursive=True) + \
                  glob.glob(os.path.join(_tmp, '**', '*.xls'), recursive=True)
            # 内嵌 zip 递归解压
            _inner_zips = glob.glob(os.path.join(_tmp, '**', '*.zip'), recursive=True)
            for _iz in _inner_zips:
                _xs.extend(_expand_zip(_iz))
        except Exception as _e:
            log(f"  解压失败 {os.path.basename(zip_path)}: {_e}")
        return _xs

    _expanded = []
    for _f in (files or []):
        if str(_f).lower().endswith('.zip'):
            _xs = _expand_zip(_f)
            log(f"  解压 {os.path.basename(_f)}: 找到 {len(_xs)} 个Excel(含嵌套)", 5)
            _expanded.extend(_xs)
        else:
            _expanded.append(_f)
    _filtered = []
    for _f in _expanded:
        _bn = os.path.basename(str(_f)); _bnl = _bn.lower()
        if _bn.startswith('.~') or 'savemsg' in _bnl or '.csv' in _bnl: continue
        if 'mmf' in _bnl: _filtered.append(_f)
    skipped = len(_expanded) - len(_filtered)
    if skipped > 0: log(f"  自动跳过 {skipped} 个文件(语音CSV/临时文件)", 5)
    files = _filtered
    nf = len(files) if files else 1
    log(f"加载MMF文件（共 {nf} 个）...", 5)
    dfs = []
    for fi, f in enumerate(files):
        if cancel_check and cancel_check(): raise KeyboardInterrupt('用户取消')
        basename = os.path.basename(f)
        # 跳过非MMF文件名（如 ALL_FTP_DETAIL_*.xlsx）
        if basename.startswith('ALL_FTP_DETAIL_'):
            log(f"  跳过非MMF文件 {fi+1}/{nf}: {basename}", 5 + int(15 * (fi+1) / nf))
            continue
        log(f"  读取文件 {fi+1}/{nf}: {basename}", 5 + int(15 * (fi+1) / nf))
        # 读全部列（详细过程需展示mmf所有列）；优先用通用函数
        df_i = _read_file(f)
        rlc_dl = [c for c in df_i.columns if 'RLC' in c and 'Downlink' in c]
        rlc_ul = [c for c in df_i.columns if 'RLC' in c and 'Uplink' in c]
        df_i['_t'] = _parse_time_vec(df_i['Time'])
        df_i['_source_file'] = os.path.basename(f)
        if rlc_dl:
            df_i['_dl'] = pd.to_numeric(df_i[rlc_dl[0]], errors='coerce').fillna(0) / 1e6
        else:
            df_i['_dl'] = 0
        if rlc_ul:
            df_i['_ul'] = pd.to_numeric(df_i[rlc_ul[0]], errors='coerce').fillna(0) / 1e6
        else:
            df_i['_ul'] = 0
        dfs.append(df_i)
    # 边界保护：若所有文件均为非MMF文件名(ALL_FTP_DETAIL_)，dfs为空，避免pd.concat报错
    if not dfs:
        log("警告: 所有文件均为非MMF文件名(ALL_FTP_DETAIL_)，无可处理数据", 100)
        version = 'V2.5.17_CX'
        timestamp = datetime.now().strftime('%Y%m%d-%H%M')
        out = f'联通/电信最终输出_{version}_{timestamp}.xlsx'
        with pd.ExcelWriter(out, engine='openpyxl') as w:
            pd.DataFrame({'提示': ['所有文件均为非MMF文件名(ALL_FTP_DETAIL_)，无可处理数据。请选择MMF文件后重试。']}).to_excel(w, sheet_name='提示', index=False)
        log(f"完成(空): {out}", 100)
        return out
    df = pd.concat(dfs, ignore_index=True)
    log("  合并完成，按时间排序...", 22)

    # 构建 data_file_info 字典（用于手机号映射）
    data_file_info = {}
    dfs_idx = 0  # 跟踪 dfs 中实际已处理的文件索引
    for fi, f in enumerate(files):
        fname = os.path.basename(f)
        # 跳过非MMF文件名
        if fname.startswith('ALL_FTP_DETAIL_'):
            continue
        df_i = dfs[dfs_idx]  # 使用独立索引，避免跳过文件导致错位
        dfs_idx += 1
        dl_vals = df_i['_dl'].dropna().values
        ul_vals = df_i['_ul'].dropna().values
        dl_nonzero = dl_vals[dl_vals > 0]
        ul_nonzero = ul_vals[ul_vals > 0]

        dl_peak = 0
        if len(dl_nonzero) > 0:
            tn = max(1, int(len(dl_nonzero) * 0.1))
            dl_sorted = np.sort(dl_nonzero)[::-1]
            dl_peak = round(min(float(np.mean(dl_sorted[:tn])), 1000), 2)

        dl_avg = 0
        if len(dl_nonzero) > 0:
            dl_avg = round(min(float(np.mean(dl_nonzero)), 1000), 2)

        ul_peak = 0
        if len(ul_nonzero) > 0:
            tn = max(1, int(len(ul_nonzero) * 0.1))
            ul_sorted = np.sort(ul_nonzero)[::-1]
            ul_peak = round(min(float(np.mean(ul_sorted[:tn])), 200), 2)

        ul_avg = 0
        if len(ul_nonzero) > 0:
            ul_avg = round(min(float(np.mean(ul_nonzero)), 200), 2)

        times = df_i['_t'].dropna()
        time_range_str = ''
        if len(times) > 0:
            time_range_str = f"{times.min().strftime('%Y-%m-%d %H:%M:%S')} ~ {times.max().strftime('%Y-%m-%d %H:%M:%S')}"

        data_file_info[fname] = {
            'operator': _detect_operator(f),
            'row_count': len(df_i),
            'time_range': time_range_str,
            'dl_peak': dl_peak,
            'dl_avg': dl_avg,
            'ul_peak': ul_peak,
            'ul_avg': ul_avg,
            'full_path': f
        }

    # 按 C列(Time/_t) 升序排列，保证详细过程表按时间顺序
    df = df.sort_values('_t', kind='stable').reset_index(drop=True)

    # ===== 时间段过滤 =====
    if time_filter:
        use_time, t_start, t_end = time_filter
        if use_time and t_start is not None and t_end is not None:
            # 提取时间部分（忽略日期），只保留 HH:MM:SS
            t_start_t = t_start.toPython() if hasattr(t_start, 'toPython') else t_start
            t_end_t = t_end.toPython() if hasattr(t_end, 'toPython') else t_end
            # 将 _t 的时间部分提取出来做比较
            df['_time_only'] = df['_t'].dt.time
            mask = (df['_time_only'] >= t_start_t) & (df['_time_only'] <= t_end_t)
            df = df[mask].copy()
            df = df.drop('_time_only', axis=1)
            log(f"  时间段过滤: {t_start_t} ~ {t_end_t}, 剩余 {len(df)} 行", 22)

    log("QCI过滤+秒聚合...", 28)
    df7 = df[df['QCI'].isin(qci_list)].dropna(subset=['_t']).copy()
    if len(df7) == 0:
        # 检查是否有其他QCI的数据
        available_qci = df['QCI'].dropna().unique()
        avail_str = ', '.join([str(int(q)) for q in available_qci if pd.notna(q) and q > 0])
        if avail_str:
            log(f"警告: 数据中无QCI={qci_list}的记录，可用QCI: {avail_str}", 28)
            log(f"建议: 重新选择QCI后再次运行", 28)
            # 生成空输出文件
            version = 'V2.5.17_CX'
            timestamp = datetime.now().strftime('%Y%m%d-%H%M')
            out = f'联通/电信最终输出_{version}_{timestamp}.xlsx'
            with pd.ExcelWriter(out, engine='openpyxl') as w:
                pd.DataFrame({'提示': [f'数据中无QCI={qci_list}的记录，可用QCI: {avail_str}']}).to_excel(w, sheet_name='提示', index=False)
            log(f"完成(空): {out}", 100)
            return out
        else:
            raise ValueError("数据中无有效的QCI记录")
    df7['_s'] = df7['_t'].dt.floor('s')
    sec = df7.groupby('_s').agg({'_dl': 'max', '_ul': 'max'}).sort_index().reset_index()
    dl_v, ul_v, ts_v = sec['_dl'].fillna(0).values, sec['_ul'].fillna(0).values, sec['_s'].values

    log("业务识别...", 35)
    big = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_BIG, UL_ACTIVE_BIG, GAP_MERGE)

    def cb(s):
        d, dl, ul = s['dur'], s['dl'], s['ul']
        if ftp_dur_min <= d <= ftp_dur_max: return 'FTP候选'
        if ftp_dur_min - 3 <= d <= ftp_dur_max + 3:
            if dl > ul * 3 and dl > 300: return 'FTP候选'
            if ul > dl * 3 and ul > 50:  return 'FTP候选'
        if dl > ul and d > 8 and dl > 200: return '应用商店大文件下载'
        if ul > dl and d > 12 and ul > 40: return '微信大包发送'
    for s in big: s['biz'] = cb(s); s['rnd'] = None

    log("FTP配对+小业务识别...", 40)
    fc = [s for s in big if s['biz'] == 'FTP候选']; paired = set(); fp = []
    for i in range(len(fc)):
        if id(fc[i]) in paired: continue
        for j in range(i + 1, len(fc)):
            if id(fc[j]) in paired: continue
            g = (fc[j]['st'] - fc[i]['et']) / np.timedelta64(1, 's')
            if 0 <= g < FTP_PAIR_GAP:
                fp.append((fc[i], fc[j])); paired.add(id(fc[i])); paired.add(id(fc[j])); break
            elif g >= FTP_PAIR_GAP: break

    def fd(s): return 'FTP下载' if s['dl'] > s['ul'] else 'FTP上传'
    for s in fc:
        if id(s) not in paired:
            if STORE_L_DL_MIN <= s['dl'] <= STORE_L_DL_MAX and s['dl'] > s['ul']: s['biz'] = '应用商店大文件下载'
            else: s['biz'] = fd(s)

    al = sorted(big, key=lambda x: x['st'])
    for s in al: s['业务'] = s['biz']
    rn = 0; pi = []
    for f1, f2 in fp:
        rn += 1; f1['rnd'] = rn; f1['业务'] = fd(f1); f2['rnd'] = rn; f2['业务'] = fd(f2)
        pi.append((rn, f2['et'], f1['st']))

    # ===== 无FTP配对时：直接使用大业务作为轮次起点 =====
    # 若没有FTP配对，则按时间顺序给每个大业务编号轮次
    if not fp:
        log("无FTP配对，使用大业务作为轮次起点...", 42)
        large_bizs = [s for s in al if s['业务'] in ('应用商店大文件下载', '微信大包发送', 'FTP下载', 'FTP上传')]
        for i, s in enumerate(large_bizs):
            rn += 1
            s['rnd'] = rn
            pi.append((rn, s['et'], s['st']))
        log(f"  识别到 {rn} 个大业务作为轮次", 43)
    for idx, (rnd, pe, _) in enumerate(pi):
        ns = pi[idx + 1][2] if idx + 1 < len(pi) else (al[-1]['et'] + np.timedelta64(1, 's'))
        for s in al:
            if s['et'] > pe and s['st'] < ns and s['rnd'] is None and s['业务'] in ('FTP下载', 'FTP上传', '应用商店大文件下载', '微信大包发送'):
                s['rnd'] = rnd

    tb = sorted([s for s in al if s['业务'] in ('FTP下载', 'FTP上传', '应用商店大文件下载', '微信大包发送')],
                key=lambda x: x['st'])
    P2 = {'应用商店大文件下载': '应用商店小文件下载', '微信大包发送': '微信小包发送'}
    small = []
    for idx, (rnd, pe, ps) in enumerate(pi):
        we = pi[idx + 1][2] if idx + 1 < len(pi) else (tb[-1]['et'] + np.timedelta64(1, 's'))
        bw = sorted([s for s in tb if s['st'] >= pe and s['st'] < we and s['业务'] in (
        '应用商店大文件下载', '微信大包发送')], key=lambda x: x['st'])
        prev = pe
        for bs in bw:
            st2 = P2.get(bs['业务'])
            if st2 and prev < bs['st'] and (bs['st'] - prev) / np.timedelta64(1, 's') > 1:
                gs = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_SMALL, UL_ACTIVE_SMALL, GAP_MERGE, prev, bs['st'])
                vd = [s for s in gs if s['dl'] + s['ul'] >= 10 and s['dur'] >= 0.5]
                if st2 == '应用商店小文件下载':
                    dls = [s for s in vd if s['dl'] > s['ul']]
                    if dls: bm = max(dls, key=lambda x: x['dl']); bm['业务'] = st2; bm['rnd'] = rnd; small.append(bm)
                else:
                    uls = [s for s in vd if s['ul'] > s['dl']]
                    if uls: bm = max(uls, key=lambda x: x['ul']); bm['业务'] = st2; bm['rnd'] = rnd; small.append(bm)
            prev = bs['et']

    # 检测第一个FTP对之前的不完整轮次
    if pi:
        first_ftp_start = pi[0][2]
        data_start = ts_v[0]
        if first_ftp_start > data_start + np.timedelta64(30, 's'):
            # 找第一个FTP对之前的大业务(应用商店大/微信大)
            pre_big = sorted([s for s in al if s['st'] < first_ftp_start and s['业务'] in (
                '应用商店大文件下载', '微信大包发送')], key=lambda x: x['st'])
            rng_start = data_start
            if pre_big:
                for bs in pre_big:
                    st2 = P2.get(bs['业务'])
                    if st2 and (bs['st'] - rng_start) / np.timedelta64(1, 's') > 1:
                        gs = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_SMALL, UL_ACTIVE_SMALL,
                                             GAP_MERGE, rng_start, bs['st'])
                        vd = [s for s in gs if s['dl'] + s['ul'] >= 10 and s['dur'] >= 0.5]
                        if st2 == '应用商店小文件下载':
                            dls = [s for s in vd if s['dl'] > s['ul']]
                            if dls: bm = max(dls, key=lambda x: x['dl']); bm['业务'] = st2; bm['rnd'] = 0; small.append(bm)
                        else:
                            uls = [s for s in vd if s['ul'] > s['dl']]
                            if uls: bm = max(uls, key=lambda x: x['ul']); bm['业务'] = st2; bm['rnd'] = 0; small.append(bm)
                    rng_start = bs['et']
            # 最后一个大业务(或数据开头) 到 first_ftp_start 之间
            if first_ftp_start > rng_start + np.timedelta64(1, 's'):
                gs = detect_segments(ts_v, dl_v, ul_v, DL_ACTIVE_SMALL, UL_ACTIVE_SMALL,
                                     GAP_MERGE, rng_start, first_ftp_start)
                vd = [s for s in gs if s['dl'] + s['ul'] >= 10 and s['dur'] >= 0.5]
                dls = [s for s in vd if s['dl'] > s['ul']]
                if dls: bm = max(dls, key=lambda x: x['dl']); bm['业务'] = '应用商店小文件下载'; bm['rnd'] = 0; small.append(bm)
                uls = [s for s in vd if s['ul'] > s['dl']]
                if uls: bm = max(uls, key=lambda x: x['ul']); bm['业务'] = '微信小包发送'; bm['rnd'] = 0; small.append(bm)

    af = sorted(al + small, key=lambda x: x['st'])
    for s in af: s['有效'] = s['业务'] is not None and s.get('rnd') is not None
    rm = defaultdict(list)
    for s in af:
        if s['有效'] and s['业务'] and s.get('rnd') is not None: rm[s['rnd']].append(s)
    for rnd, segs in rm.items():
        seen = set(); cnt = 0
        for s in sorted(segs, key=lambda x: x['st']):
            if s['业务'] in seen or cnt >= 6: s['有效'] = False
            else: seen.add(s['业务']); cnt += 1

    valid = [s for s in af if s['有效'] and s['业务']]
    log(f"识别完成: {len(valid)}个业务段, {rn}轮")

    # ===== 标记 =====
    df = pd.concat([df, pd.DataFrame({
        '下载速率': df['_dl'], '上传速率': df['_ul'],
        '业务识别': None, '持续时长': None, '速率(Mbps)': None,
    }, index=df.index)], axis=1)

    for si, seg in enumerate(valid):
        if cancel_check and si % 20 == 0 and cancel_check():
            raise KeyboardInterrupt('用户取消')
        st, et, biz = seg['st'], seg['et'], seg['业务']
        sm = (df['_t'] >= st) & (df['_t'] <= et + timedelta(seconds=1)) & (df['QCI'].isin(qci_list))
        col = '_dl' if biz in DL_BIZ else '_ul'

        # 所有业务：只标记有流量的行(>0)
        cv_all = df.loc[sm, col].fillna(0)
        nz = cv_all > 0
        mm = sm & nz.reindex(df.index, fill_value=False)

        if not mm.any():
            continue  # 无流量则跳过

        # FTP：从第一个有流量的行开始，检查前后小流量补齐到10秒
        if biz in ('FTP下载', 'FTP上传'):
            # 找所有有流量的行
            flow_vals = df.loc[mm, col].fillna(0).tolist()
            flow_times = df.loc[mm, '_t'].tolist()

            if len(flow_vals) == 0:
                continue

            # 起点：从第一个有流量的行开始
            first_idx = 0

            # 终点：最后一个有流量的行
            last_idx = len(flow_vals) - 1

            # 计算时长
            first_t = flow_times[first_idx]
            last_t = flow_times[last_idx]
            md = (last_t - first_t).total_seconds()

            # 如果时长不是10秒左右，检查前后是否有小流量可以补齐
            # 向前找：在segment开始时间之前3秒内，找有流量的行
            if md < 9.5 or md > 10.5:
                # 向前找小流量
                ft0 = st - timedelta(seconds=3)
                pm = (df['_t'] >= ft0) & (df['_t'] < first_t) & (df['QCI'].isin(qci_list))
                if pm.any():
                    pv = df.loc[pm, col].fillna(0)
                    pv_nz = pv > 0
                    if pv_nz.any():
                        # 有前置流量，扩展起点
                        first_t = df.loc[pv[pv_nz].index[0], '_t']
                        md = (last_t - first_t).total_seconds()

                # 向后找小流量
                et0 = et + timedelta(seconds=3)
                pm2 = (df['_t'] > last_t) & (df['_t'] <= et0) & (df['QCI'].isin(qci_list))
                if pm2.any():
                    pv2 = df.loc[pm2, col].fillna(0)
                    pv2_nz = pv2 > 0
                    if pv2_nz.any():
                        # 有后置流量，扩展终点
                        last_t = df.loc[pv2[pv2_nz].index[-1], '_t']
                        md = (last_t - first_t).total_seconds()

            # 重新计算标记范围
            mm_final = (df['_t'] >= first_t) & (df['_t'] <= last_t) & (df['QCI'].isin(qci_list)) & (df[col].fillna(0) > 0)
            md = (last_t - first_t).total_seconds()

            df.loc[mm_final, '业务识别'] = biz
            df.loc[mm_final, '持续时长'] = round(md, 3)
            vs = pd.to_numeric(df.loc[mm_final, '下载速率'] if biz in DL_BIZ else df.loc[mm_final, '上传速率'], errors='coerce').fillna(0).tolist()
            ar = sum(vs) / md if md > 0 else 0
            df.loc[mm_final, '速率(Mbps)'] = round(ar, 2)
            seg['mark_st'], seg['mark_et'], seg['mark_dur'] = first_t, last_t, md
            continue

        # 其他业务：直接标记有流量的行
        first_t = df.loc[mm, '_t'].min()
        last_t = df.loc[mm, '_t'].max()
        md = (last_t - first_t).total_seconds()

        df.loc[mm, '业务识别'] = biz
        df.loc[mm, '持续时长'] = round(md, 3)
        if biz in DL_BIZ: vs = pd.to_numeric(df.loc[mm, '下载速率'], errors='coerce').fillna(0).tolist()
        else: vs = pd.to_numeric(df.loc[mm, '上传速率'], errors='coerce').fillna(0).tolist()
        ar = sum(vs) / md if md > 0 else 0
        df.loc[mm, '速率(Mbps)'] = round(ar, 2)
        seg['mark_st'], seg['mark_et'], seg['mark_dur'] = first_t, last_t, md

    rd = defaultdict(set)
    for s in valid:
        r = s.get('rnd')
        if r is not None: rd[r].add(s['业务'])
    rl = {}; ii = 0
    for r in sorted(rd.keys()):
        if rd[r] >= REQ_BIZ: rl[r] = str(r)
        else: ii += 1; rl[r] = f"不完整{ii}"
    df['轮次'] = None
    # 预计算QCI筛选掩码(避免183次循环重复计算)
    qci_mask = df['QCI'].isin(qci_list)
    biz_notna_mask = df['业务识别'].notna()
    for s in valid:
        r = s.get('rnd'); lb = rl.get(r) if r is not None else ''
        if lb:
            ms2, me2 = s.get('mark_st', s['st']), s.get('mark_et', s['et'])
            # 只在有业务标记的行填轮次
            mm2 = (df['_t'] >= ms2) & (df['_t'] <= me2) & qci_mask & biz_notna_mask
            df.loc[mm2, '轮次'] = lb

    # ===== 详细过程 =====
    # 性能优化: 一次性构建df_detail,避免逐列赋值触发碎片化(原117次逐列插入耗时135秒)
    log("生成详细过程...", 48)

    # 0. 预构建文件名到手机号的映射(避免在构建detail_dict时遍历560万行)
    file_to_path = {fname: info['full_path'] for fname, info in data_file_info.items()}
    trace_to_phone = {}
    if phone_trace_map:
        for entry in phone_trace_map:
            trace_to_phone[entry['trace_id']] = entry['phone']
    _source_file_to_phone = {}
    for sf in df['_source_file'].unique():
        _source_file_to_phone[sf] = trace_to_phone.get(_extract_trace_id(file_to_path.get(sf, '')), '')

    # 1. 收集所有需要的列数据
    OC = ['下载速率', '上传速率', '速率(Mbps)', '持续时长', '轮次']
    base_cols = ['基准_业务类型', '基准_开始时间', '基准_结束时间', '基准_速率类指标', '基准_数值', '基准_时长类指标', '基准_数值.1']
    special_cols = ['QCI', 'Downlink RLC Throughput(bps)', 'Uplink RLC Throughput(bps)']
    exclude_cols = {'id', '_t', '_dl', '_ul', '_s', 'businessType', '业务识别'}

    # 2. 直接在df上加列,不创建新DataFrame(避免复制560万行×106列)
    # 润林工具启示: 在大数据下,避免复制DataFrame,流式处理
    df['筛选'] = 0
    df['id'] = range(1, len(df) + 1)
    df['手机号'] = df['_source_file'].map(_source_file_to_phone).fillna('')
    df['businessType'] = df['业务识别']
    for bc in base_cols:
        df[bc] = None
    # xlsxwriter写列的顺序: 筛选, id, Time, OC, businessType, base_cols, _t, 手机号, special_cols, 其余
    _write_order = ['筛选', 'id', 'Time']
    for c in OC:
        if c in df.columns: _write_order.append(c)
    _write_order.append('businessType')
    for bc in base_cols:
        if bc in df.columns: _write_order.append(bc)
    if '_t' in df.columns: _write_order.append('_t')
    _write_order.append('手机号')
    for sc in special_cols:
        if sc in df.columns: _write_order.append(sc)
    for c in df.columns:
        if c not in _write_order and c not in exclude_cols:
            _write_order.append(c)
    df_detail = df  # df_detail参照df(无复制)

    # 加载基准数据并标记到详细过程
    log("加载基准数据...", 53)
    has_base = True
    _is_ref_base = False

    # 优先使用用户传入的基准文件
    if base_file and os.path.exists(base_file):
        try:
            # 尝试读取用户基准文件
            try:
                base = pd.read_excel(base_file, sheet_name='呼叫详情', engine='calamine')
            except:
                base = pd.read_excel(base_file, engine='calamine')

            if base is None or len(base) == 0:
                has_base = False
            else:
                log(f"使用用户基准文件: {os.path.basename(base_file)}", 53)
        except Exception as e:
            log(f"未使用基准: 用户基准文件读取失败({e})", 53)
            has_base = False
            base = pd.DataFrame()
    else:
        # 尝试加载默认基准文件(正式使用时通常无此文件,属正常情况)
        default_base = '/Users/sun/ClaudeCode/先甲工信部工具/输入文件/0720-重大/重大-永川参考数据.xlsx'
        if not os.path.exists(default_base):
            log(f"未找到永川参考文件，跳过基准对齐", 53)
            has_base = False
            base = pd.DataFrame()
        else:
            try:
                base = pd.read_excel(default_base, sheet_name='整理后文件', engine='calamine')
                if base is None or len(base) == 0:
                    has_base = False
                elif '开始时间' not in base.columns and '采集时间' in base.columns:
                    # 永川参考文件列名映射
                    base = base.rename(columns={'采集时间': '开始时间', '联通业务': '业务类型'})
                    base['开始时间'] = pd.to_datetime(base['开始时间'])
                    _is_ref_base = True  # 标记为参考基准(列不完整, 跳过大循环对齐)
                    log(f"加载永川参考作为基准: {len(base)}行", 53)
                else:
                    log(f"基准文件格式不匹配", 53)
                    has_base = False
                    base = pd.DataFrame()
            except Exception as e:
                log(f"基准文件读取失败: {e}", 53)
                has_base = False
                base = pd.DataFrame()
    log("基准对齐...", 55)
    # 基准列已在上方初始化(作为detail_dict的一部分)

    # 对齐基准数据到详细过程（向量化优化：避免双重循环）
    ts = df['_t']
    ts_valid = ts[ts.notna()]

    # 检查基准列完整性: 永川参考标记跳过
    if len(base) > 0 and '_is_ref_base' in dir() and _is_ref_base:
        log(f"  永川参考列不完整, 跳过基准对齐", 55)
        base = pd.DataFrame()
    if len(ts_valid) > 0 and len(base) > 0:
        log(f"  基准记录数: {len(base)}, 数据行数: {len(ts_valid)}", 55)

        # 预处理基准数据
        base_starts = pd.to_datetime(base['开始时间'])
        base_ends = pd.to_datetime(base['结束时间'])
        base_bizs = base['业务类型'].astype(str).tolist()
        base_rate_indicators = base['速率类指标'].astype(str).tolist()
        base_rate_values = base['数值'].tolist()
        base_dur_indicators = base['时长类指标'].astype(str).tolist()
        base_dur_values = base['数值.1'].tolist()

        # 向量化查找：对于每条基准记录，找到对应的开始/结束行号
        # 使用 searchsorted 进行二分查找，复杂度 O(n log n)
        ts_sorted = ts_valid.sort_values()
        ts_indices = ts_sorted.index

        for i in range(len(base)):
            if cancel_check and i % 50 == 0 and cancel_check():
                raise KeyboardInterrupt('用户取消')
            b_start = base_starts[i]
            b_end = base_ends[i]
            b_biz = base_bizs[i]

            # 二分查找 b_start 最近的位置
            pos = ts_sorted.searchsorted(b_start)
            if pos >= len(ts_sorted):
                continue

            # 检查最近的开始时间是否在5秒内
            if pos > 0:
                t_prev = ts_sorted.iloc[pos - 1]
                t_next = ts_sorted.iloc[pos]
                if abs((t_next - b_start).total_seconds()) > abs((t_prev - b_start).total_seconds()):
                    pos = pos - 1

            if abs((ts_sorted.iloc[pos] - b_start).total_seconds()) >= 5:
                continue

            best_start_idx = ts_indices[pos]

            # 二分查找 b_end 最近的位置（从best_start_idx之后）
            pos_end = ts_sorted.searchsorted(b_end)
            if pos_end >= len(ts_sorted):
                pos_end = len(ts_sorted) - 1

            best_end_idx = ts_indices[pos_end]
            if best_end_idx < best_start_idx:
                best_end_idx = best_start_idx

            # 批量赋值
            lo, hi = best_start_idx, min(best_end_idx, len(df_detail) - 1)
            if lo <= hi:
                df_detail.loc[lo:hi, '基准_业务类型'] = b_biz
                df_detail.loc[lo:hi, '基准_开始时间'] = str(base_starts[i])
                df_detail.loc[lo:hi, '基准_结束时间'] = str(base_ends[i])
                df_detail.loc[lo:hi, '基准_速率类指标'] = base_rate_indicators[i]
                df_detail.loc[lo:hi, '基准_数值'] = base_rate_values[i]
                df_detail.loc[lo:hi, '基准_时长类指标'] = base_dur_indicators[i]
                df_detail.loc[lo:hi, '基准_数值.1'] = base_dur_values[i]

    # _t列、手机号列、special_cols、其他列已在上方一次性构建完成,无需再逐列添加

    # 计算筛选列（向量化）：check_cols 任一非空则为1
    check_cols = ['速率(Mbps)', '持续时长', '轮次', 'businessType',
                  '基准_业务类型', '基准_开始时间', '基准_结束时间', '基准_速率类指标', '基准_数值', '基准_时长类指标', '基准_数值.1']
    has = pd.Series(False, index=df_detail.index)
    for col in check_cols:
        s = df_detail[col]
        ss = s.astype(str).str.strip()
        has = has | (s.notna() & (ss != '') & (ss.str.lower() != 'nan'))
    df_detail['筛选'] = has.astype(int).values

    # ===== 汇总 =====
    sr = []
    sum_raw_data = {}  # 用于批注：{biz: {rates, durs, col_name, r_start, r_end, sub}}
    for biz in sorted(DL_BIZ | UL_BIZ):
        sub = df_detail[df_detail['businessType'] == biz].dropna(subset=['_t']).sort_values('_t')
        if len(sub) == 0: continue
        dl_cn = [c for c in sub.columns if 'RLC' in c and 'Downlink' in c]
        ul_cn = [c for c in sub.columns if 'RLC' in c and 'Uplink' in c]
        if biz in DL_BIZ and dl_cn:
            col_name = dl_cn[0]
            rates_raw = pd.to_numeric(sub[col_name], errors='coerce').fillna(0) / 1e6
            rates = rates_raw[rates_raw > 0]
        elif biz in UL_BIZ and ul_cn:
            col_name = ul_cn[0]
            rates_raw = pd.to_numeric(sub[col_name], errors='coerce').fillna(0) / 1e6
            rates = rates_raw[rates_raw > 0]
        else:
            rates = pd.Series([], dtype=float); col_name = ''
        # 时长
        ts2 = sub['_t']; durs = []
        if len(ts2) > 0:
            s1 = ts2.iloc[0]; p = s1
            for t in ts2.iloc[1:]:
                if (t - p).total_seconds() > 5: durs.append((p - s1).total_seconds() + 1); s1 = t
                p = t
            durs.append((p - s1).total_seconds() + 1)
        # 记录原始数据位置（在详细过程df中的行号）
        r_start = sub.index[0] + 2  # +2 因为Excel第1行表头, 第2行索引0
        r_end = sub.index[-1] + 2
        sum_raw_data[biz] = {'rates': rates, 'durs': durs, 'col_name': col_name,
                             'r_start': r_start, 'r_end': r_end, 'sub': sub}
        row = {'业务类型': biz}
        if biz in DL_BIZ and len(rates) > 0:
            row['应用层FTP下载速率1000M以上占比(%)'] = round((rates > dl_clip).sum() / len(rates) * 100, 2)
            row['应用层FTP下载速率100M以下占比(%)'] = round((rates < 100).sum() / len(rates) * 100, 2)
            row['应用层平均下载速率(Mbps)'] = round(rates.mean(), 2)
            row['削峰应用层平均下载速率(Mbps)'] = round(rates.clip(upper=dl_clip).mean(), 2)
            sv = rates.sort_values(ascending=False); tn = max(1, int(len(sv) * 0.1))
            row['下行削峰TOP10%峰值速率'] = round(sv.head(tn).mean(), 2)
        elif biz in UL_BIZ and len(rates) > 0:
            row['应用层FTP上传速率200M以上占比(%)'] = round((rates > ul_clip).sum() / len(rates) * 100, 2)
            row['应用层FTP上传速率5M以下占比(%)'] = round((rates < 5).sum() / len(rates) * 100, 2)
            row['应用层平均上传速率'] = round(rates.mean(), 2)
            row['削峰应用层平均上传速率(Mbps)'] = round(rates.clip(upper=ul_clip).mean(), 2)
            sv = rates.sort_values(ascending=False); tn = max(1, int(len(sv) * 0.1))
            row['上行削峰TOP10%峰值速率'] = round(sv.head(tn).mean(), 2)
        if durs:
            row['业务时长平均值(s)'] = round(np.mean(durs), 2)
            row['业务时长中位值(s)'] = round(np.median(durs), 0)
        sr.append(row)
    df_summary = pd.DataFrame(sr)

    # ===== 汇总批注 + D/E列颜色 =====
    # 收集每业务的原始数据，用于批注
    raw_data = {}
    for biz in sorted(DL_BIZ | UL_BIZ):
        sub = df_detail[df_detail['businessType'] == biz].dropna(subset=['_t']).sort_values('_t')
        if len(sub) == 0: continue
        dl_cn = [c for c in sub.columns if 'RLC' in c and 'Downlink' in c]
        ul_cn = [c for c in sub.columns if 'RLC' in c and 'Uplink' in c]
        if biz in DL_BIZ and dl_cn:
            rates = pd.to_numeric(sub[dl_cn[0]], errors='coerce').fillna(0) / 1e6
            raw_data[biz] = {'rates': rates, 'durs': [], 'col': dl_cn[0], 'rows': (sub.index.min(), sub.index.max())}
        elif biz in UL_BIZ and ul_cn:
            rates = pd.to_numeric(sub[ul_cn[0]], errors='coerce').fillna(0) / 1e6
            raw_data[biz] = {'rates': rates, 'durs': [], 'col': ul_cn[0], 'rows': (sub.index.min(), sub.index.max())}
        ts2 = sub['_t']; durs = []
        if len(ts2) > 0:
            s1 = ts2.iloc[0]; p = s1
            for t in ts2.iloc[1:]:
                if (t - p).total_seconds() > 5: durs.append((p - s1).total_seconds() + 1); s1 = t
                p = t
            durs.append((p - s1).total_seconds() + 1)
        if biz in raw_data:
            raw_data[biz]['durs'] = durs

    # 等 ExcelWriter 写入完毕后，对汇总 sheet 加批注
    # 标注在 ExcelWriter 的 with 块内完成

    # ===== 对比横表（基准驱动，逐行匹配）=====
    BIZ_ORDER = ['FTP下载', 'FTP上传', '应用商店小文件下载', '应用商店大文件下载', '微信小包发送', '微信大包发送']
    rate_nm = {'FTP下载': '下载平均速率(Mbps)', 'FTP上传': '上传平均速率(Mbps)',
               '应用商店小文件下载': '小文件下载速率(Mbps)', '应用商店大文件下载': '大文件下载速率(Mbps)',
               '微信小包发送': '微信小包发送速率(Mbps)', '微信大包发送': '微信大包发送速率(Mbps)'}

    def fmt_time(t_str):
        m = _PARSE_RE.match(str(t_str))
        if m:
            dt = datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S')
            return f"{m.group(1)}({m.group(2)})", dt
        # 尝试毫秒格式 .xxx
        m2 = _PARSE_MS_RE.match(str(t_str))
        if m2:
            dt = datetime.strptime(m2.group(1), '%Y-%m-%d %H:%M:%S')
            return f"{m2.group(1)}({m2.group(2)[:3]})", dt
        return str(t_str), None

    def fmt_base_time(t):
        """基准时间格式化为 2026-07-02 22:27:36.557 样式(点+3位毫秒)，方便CTRL+F对照基准文件"""
        if t is None: return ''
        try:
            ts = pd.Timestamp(t)
            if pd.isna(ts): return str(t)
            return ts.strftime('%Y-%m-%d %H:%M:%S.') + f"{ts.microsecond // 1000:03d}"
        except Exception:
            return str(t)

    # _t列已在上方一次性构建完成,此处无需再添加

    # 基准已在上方(详细过程模块)读取并按开始时间排序，此处直接复用，避免重复读盘
    base_sorted = base.reset_index(drop=True)

    # 按业务序列切轮：每轮以 FTP下载 开头(标准顺序
    # FTP下载→FTP上传→应用商店小文件→应用商店大文件→微信小包→微信大包)。
    # 第一个 FTP下载 之前的业务归为开头不完整轮次；末尾不足6个为末尾不完整轮次。
    # (替代 V2.0.9 的"每6条硬切一组"——硬切遇到开头/末尾不完整轮次会整体错位)
    base_rounds = []
    cur = []
    for i in range(len(base_sorted)):
        biz = str(base_sorted.iloc[i]['业务类型'])
        if biz == 'FTP下载' and cur:
            base_rounds.append(cur); cur = []
        cur.append(base_sorted.iloc[i])
    if cur:
        base_rounds.append(cur)

    # 逐轮逐业务匹配(基准驱动)。判定"对得上"不用固定时间窗口，而看业务连续段：
    #   对基准轮内业务 Bk，匹配区间由相邻业务边界决定——
    #     下界 = 上一个基准业务结束时间(轮首无下界)
    #     上界 = 下一个基准业务开始时间(轮末用自身结束时间)
    #   代码侧同业务段只要开始时间落在该区间且未被占用即算对得上；
    #   整轮6业务完整且全部对上才编号，否则该轮标"不计轮次"。
    all_items = []
    rnd_counter = 0
    used = set()  # 已占用的 df_detail 行索引，避免一代码段被多轮重复匹配
    for rnd_bizs in base_rounds:
        n = len(rnd_bizs)
        all_matched = True
        grp_items = []

        for k in range(n):
            brow = rnd_bizs[k]
            b_biz = str(brow['业务类型'])
            b_start_str = str(brow['开始时间'])
            b_end_str = str(brow['结束时间'])
            b_start = pd.to_datetime(brow['开始时间'])
            b_end = pd.to_datetime(brow['结束时间'])

            # 就近匹配(B方案)：在所有未占用的同业务行里，找离 b_start 最近的连续段
            # 距离<=300秒(约一轮间隔)才算同轮，避免跨轮串；used 防一段被多轮重复占用
            cand = df_detail[(df_detail['businessType'] == b_biz) &
                             (~df_detail.index.isin(used))]
            code_data = None
            if len(cand) > 0:
                dist = (cand['_t'] - b_start).abs()
                best_idx = dist.idxmin()
                if dist.loc[best_idx] <= pd.Timedelta(seconds=300):
                    # 取 best_idx 所在的连续行块(行号连续的同业务未占用行)
                    cand_idx_sorted = cand.index.sort_values()
                    pos = cand_idx_sorted.get_loc(best_idx)
                    lo = pos
                    while lo - 1 >= 0 and cand_idx_sorted[lo - 1] == cand_idx_sorted[lo] - 1:
                        lo -= 1
                    hi = pos
                    while hi + 1 < len(cand_idx_sorted) and cand_idx_sorted[hi + 1] == cand_idx_sorted[hi] + 1:
                        hi += 1
                    seg_idxs = list(cand_idx_sorted[lo:hi + 1])
                    used.update(seg_idxs)
                    cand_s = df_detail.loc[seg_idxs].sort_values('_t')
                    first = cand_s.iloc[0]
                    last = cand_s.iloc[-1]

                    c_start = str(first['Time'])
                    c_end = str(last['Time'])
                    c_start_fmt, _ = fmt_time(c_start)
                    c_end_fmt, _ = fmt_time(c_end)

                    code_data = {
                        '开始时间': c_start_fmt if c_start_fmt else c_start,
                        '结束时间': c_end_fmt if c_end_fmt else c_end,
                        '速率类指标': rate_nm.get(b_biz, ''),
                        '数值': first.get('速率(Mbps)'),
                        '时长类指标': '业务时长(秒)',
                        '数值.1': first.get('持续时长'),
                    }
                else:
                    all_matched = False
            else:
                all_matched = False

            # 基准数据(开始/结束时间用 .557 点样式，与基准文件一致便于CTRL+F)
            base_data = {
                '开始时间': fmt_base_time(brow['开始时间']),
                '结束时间': fmt_base_time(brow['结束时间']),
                '速率类指标': str(brow.get('速率类指标', '')),
                '数值': brow.get('数值'),
                '时长类指标': str(brow.get('时长类指标', '')),
                '数值.1': brow.get('数值.1'),
            }

            grp_items.append({
                'biz': b_biz,
                '基准': base_data,
                '代码': code_data,
            })

        complete = (n == 6)
        rnd_label = str(rnd_counter + 1) if (complete and all_matched) else '不计轮次'
        if complete and all_matched:
            rnd_counter += 1
        for item in grp_items: item['rnd_label'] = rnd_label
        all_items.extend(grp_items)

    # 偏差计算辅助函数
    def _pct(cv, bv):
        try:
            if pd.isna(cv) or pd.isna(bv): return '-'
            cf = float(cv); bf = float(bv)
            return '-' if bf == 0 else round((cf - bf) / bf * 100, 2)
        except Exception:
            return '-'

    def _time_diff(c_time, b_time):
        try:
            m = _PARSE_RE.match(str(c_time)); n = _PARSE_RE.match(str(b_time))
            if m and n:
                cdt = datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S') + timedelta(milliseconds=int(m.group(2)))
                bdt = datetime.strptime(n.group(1), '%Y-%m-%d %H:%M:%S') + timedelta(milliseconds=int(n.group(2)))
                return round((cdt - bdt).total_seconds(), 2)
            return '-'
        except Exception:
            return '-'

    # 构建df_compare：每业务2行(基准+代码)，偏差在右侧3列横铺
    cr = []
    for item in all_items:
        rnd = item['rnd_label']
        biz = item['biz']
        bd = item['基准']
        cd = item['代码']
        rpct = _pct(cd['数值'], bd['数值']) if cd else '-'
        dpct = _pct(cd['数值.1'], bd['数值.1']) if cd else '-'
        tdiff = _time_diff(cd['开始时间'], bd['开始时间']) if cd else '-'
        # 基准行（偏差列有值）
        cr.append({'轮次': rnd, '业务类型': biz, '来源': '基准',
                   '开始时间': bd['开始时间'], '结束时间': bd['结束时间'],
                   '速率类指标': bd['速率类指标'], '数值': bd['数值'],
                   '时长类指标': bd['时长类指标'], '数值.1': bd['数值.1'],
                   '速率偏差(%)': rpct, '时长偏差(%)': dpct, '时间差(秒)': tdiff})
        # 代码行（偏差列留空，由合并单元格覆盖）
        if cd:
            cr.append({'轮次': rnd, '业务类型': biz, '来源': '代码',
                       '开始时间': cd['开始时间'], '结束时间': cd['结束时间'],
                       '速率类指标': cd['速率类指标'], '数值': cd['数值'],
                       '时长类指标': cd['时长类指标'], '数值.1': cd['数值.1'],
                       '速率偏差(%)': '', '时长偏差(%)': '', '时间差(秒)': ''})
        else:
            cr.append({'轮次': rnd, '业务类型': biz, '来源': '代码',
                       '开始时间': '缺失', '结束时间': '缺失',
                       '速率类指标': '缺失', '数值': '缺失',
                       '时长类指标': '缺失', '数值.1': '缺失',
                       '速率偏差(%)': '', '时长偏差(%)': '', '时间差(秒)': ''})

    df_compare = pd.DataFrame(cr).reset_index(drop=True)

    # ===== 汇总：所有业务 =====
    def _calc_summary(sub_detail, dl_clip, ul_clip):
        """从 df_detail 子集计算汇总行"""
        sr = []
        for biz in sorted(DL_BIZ | UL_BIZ):
            sub = sub_detail[sub_detail['businessType'] == biz].dropna(subset=['_t']).sort_values('_t')
            if len(sub) == 0: continue
            dl_cn = [c for c in sub.columns if 'RLC' in c and 'Downlink' in c]
            ul_cn = [c for c in sub.columns if 'RLC' in c and 'Uplink' in c]
            if biz in DL_BIZ and dl_cn:
                cn = dl_cn[0]; rates = pd.to_numeric(sub[cn], errors='coerce').fillna(0) / 1e6; rates = rates[rates > 0]
            elif biz in UL_BIZ and ul_cn:
                cn = ul_cn[0]; rates = pd.to_numeric(sub[cn], errors='coerce').fillna(0) / 1e6; rates = rates[rates > 0]
            else: rates = pd.Series([], dtype=float); cn = ''
            ts2 = sub['_t']; durs = []
            if len(ts2) > 0:
                s1 = ts2.iloc[0]; p = s1
                for t in ts2.iloc[1:]:
                    if (t - p).total_seconds() > 5: durs.append((p - s1).total_seconds() + 1); s1 = t
                    p = t
                durs.append((p - s1).total_seconds() + 1)
            row = {'业务类型': biz}
            is_dl = biz in DL_BIZ; lim = dl_clip if is_dl else ul_clip; lo = 100 if is_dl else 5
            if len(rates) > 0:
                row['阈值以上占比(%)'] = round((rates > lim).sum() / len(rates) * 100, 2)
                row['5M/100M以下占比(%)'] = round((rates < lo).sum() / len(rates) * 100, 2)
                row['平均速率(Mbps)'] = round(rates.mean(), 2)
                row['削峰平均速率(Mbps)'] = round(rates.clip(upper=lim).mean(), 2)
                sv = rates.sort_values(ascending=False); tn = max(1, int(len(sv) * 0.1))
                row['削峰TOP10%峰值速率'] = round(sv.head(tn).mean(), 2)
            else:
                for k in ['阈值以上占比(%)','5M/100M以下占比(%)','平均速率(Mbps)','削峰平均速率(Mbps)','削峰TOP10%峰值速率']:
                    row[k] = '-'
            if durs:
                row['业务时长平均值(s)'] = round(np.mean(durs), 2)
                row['业务时长中位值(s)'] = round(np.median(durs), 0)
            # 收集该业务类型涉及的所有手机号
            phones_in_sub = sub['手机号'].dropna().unique()
            row['手机号'] = '、'.join(sorted([str(p) for p in phones_in_sub])) if len(phones_in_sub) > 0 else ''
            sr.append(row)
        return pd.DataFrame(sr)

    log("生成汇总...", 70)
    df_summary_all = _calc_summary(df_detail, dl_clip, ul_clip)
    complete_mask = df_detail['轮次'].astype(str).str.isdigit()
    df_detail_complete = df_detail[complete_mask]
    df_summary_complete = _calc_summary(df_detail_complete, dl_clip, ul_clip)

    # 生成输出文件名
    version = 'V2.5.17_CX'
    timestamp = datetime.now().strftime('%Y%m%d-%H%M')
    out = f'联通/电信最终输出_{version}_{timestamp}.xlsx'
    # 确保输出目录存在
    os.makedirs(os.path.dirname(out), exist_ok=True)

    log("写入详细过程...", 78)
    import xlsxwriter
    from openpyxl.utils import get_column_letter as _gcl  # 仅用于列字母转换
    from datetime import datetime as _dt_cls, date as _date_cls

    # ===== 工具函数（xlsxwriter 适配）=====
    def _is_eo(v):
        """None 或 0（NaN 不算 — 与原 openpyxl 行为一致：isinstance int/float and v==0）"""
        if v is None: return True
        if isinstance(v, bool): return False
        if isinstance(v, (int, float)) and not pd.isna(v) and v == 0: return True
        return False

    def _ne(v):
        """非空（None/NaN/空串/nan字符串 都算空）"""
        if v is None: return False
        if isinstance(v, float) and pd.isna(v): return False
        if isinstance(v, str) and v.strip() in ('', 'nan', 'None'): return False
        return True

    def _cval(v):
        """转 xlsxwriter 可写的值"""
        if v is None: return None
        if isinstance(v, float) and pd.isna(v): return None
        if isinstance(v, pd.Timestamp): return v.to_pydatetime()
        if isinstance(v, np.datetime64): return pd.Timestamp(v).to_pydatetime()
        if isinstance(v, (_dt_cls, _date_cls)): return v
        if isinstance(v, np.integer): return int(v)
        if isinstance(v, np.floating): return float(v)
        if isinstance(v, str) and v == '': return None
        return v

    def _ws_write(ws, row, col, v, fmt=None):
        """智能写入，处理 NaN/Timestamp/None"""
        cv = _cval(v)
        if cv is None:
            ws.write_blank(row, col, None, fmt); return
        if isinstance(cv, bool):
            ws.write_boolean(row, col, cv, fmt); return
        if isinstance(cv, (int, float)):
            ws.write_number(row, col, float(cv), fmt); return
        if isinstance(cv, (_dt_cls, _date_cls)):
            ws.write_datetime(row, col, cv, fmt); return
        if isinstance(cv, str):
            ws.write_string(row, col, cv, fmt); return
        ws.write_string(row, col, str(cv), fmt)

    _out_tmp = out + '.tmp'
    with xlsxwriter.Workbook(_out_tmp, {'strings_to_numbers': False, 'nan_inf_to_errors': True, 'default_date_format': 'yyyy-mm-dd hh:mm:ss'}) as wb:
        # ===== 预定义 formats =====
        fmt_hdr_d = wb.add_format({'text_wrap': True})
        fmt_num = wb.add_format({'num_format': '0.00'})
        fmt_d = wb.add_format({'num_format': '0.00', 'bg_color': '#DAEEF3'})
        fmt_e = wb.add_format({'num_format': '0.00', 'bg_color': '#E2EFDA'})
        fmt_hdr_s = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter',
                                   'bg_color': '#D9E1F2', 'border': 1, 'text_wrap': True})
        fmt_ta = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'font_size': 11,
                                'bg_color': '#E2EFDA', 'border': 1})
        fmt_tc = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'font_size': 11,
                                'bg_color': '#DAEEF3', 'border': 1})
        fmt_d_s = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
        fmt_d_url = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1,
                                   'font_color': '#0000FF', 'underline': 1})
        fmt_hdr_h = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter',
                                   'bg_color': '#D9E1F2', 'border': 1, 'text_wrap': True})
        fmt_biz = {bn: wb.add_format({'bg_color': ('#CCE5FF' if i % 2 == 0 else '#E5FFCC'),
                                       'border': 1, 'align': 'center', 'valign': 'vcenter'})
                   for i, bn in enumerate(BIZ_ORDER)}
        fmt_y = wb.add_format({'bg_color': '#FFFF00', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
        fmt_center = wb.add_format({'align': 'center', 'valign': 'vcenter'})

        # ===== 按运营商写 Sheet =====
        _ops = sorted(set(df_detail['运营商'].dropna().unique())) if '运营商' in df_detail.columns else []
        if not _ops: _ops = ['数据业务']
        for _op_name in _ops:
            _op_detail = df_detail[df_detail['运营商'] == _op_name] if len(_ops) > 1 else df_detail
            if len(_op_detail) == 0: continue
            ws = wb.add_worksheet(f'{_op_name}-数据业务-详细过程')
        cols = _write_order
        n_detail = len(_op_detail)
        n_col_detail = len(cols)
        # 预提取所需列为numpy数组(避免df.iat逐格访问的pandas开销)
        _write_arrays = [_op_detail[c].values for c in cols]
        col_3_idx = cols.index('下载速率') if '下载速率' in cols else 3
        col_4_idx = cols.index('上传速率') if '上传速率' in cols else 4

        # 表头
        for ci, col_name in enumerate(cols):
            if ci == col_3_idx:
                ws.write(0, ci, '下载速率（Mbps）', fmt_hdr_d)
            elif ci == col_4_idx:
                ws.write(0, ci, '上传速率(Mbps)', fmt_hdr_d)
            else:
                ws.write(0, ci, col_name, fmt_hdr_d)

        # 数据行 — 用write_column批量写(一次性,避免O(n²)嵌套循环bug)
        _d_arr = _write_arrays[col_3_idx] if n_col_detail > col_3_idx else None
        _e_arr = _write_arrays[col_4_idx] if n_col_detail > col_4_idx else None
        _d_list = []; _e_list = []; _col0_list = []
        for i in range(n_detail):
            if cancel_check and i > 0 and i % 20000 == 0 and cancel_check():
                raise KeyboardInterrupt('用户取消')
            d_orig = _d_arr[i] if _d_arr is not None else None
            e_orig = _e_arr[i] if _e_arr is not None else None
            d_dash = _is_eo(d_orig); e_dash = _is_eo(e_orig)
            v0 = _write_arrays[0][i]
            _col0_list.append(0 if d_dash else (int(v0) if not pd.isna(v0) else ''))
            _d_list.append('-' if d_dash else (float(d_orig) if not pd.isna(d_orig) else ''))
            _e_list.append('-' if e_dash else (float(e_orig) if not pd.isna(e_orig) else ''))
        ws.write_column(1, 0, _col0_list)
        ws.write_column(1, col_3_idx, _d_list)
        ws.write_column(1, col_4_idx, _e_list)
        for ci in range(1, n_col_detail):
            if ci in (col_3_idx, col_4_idx): continue
            arr = _write_arrays[ci]
            if arr.dtype.kind == 'M':
                lst = [pd.Timestamp(v).to_pydatetime() if pd.notna(v) else '' for v in arr]
            else:
                lst = arr.tolist()
            ws.write_column(1, ci, lst)

        # 列宽 + 隐藏 J~Q (0-based 列 9~16)
        for ci in range(n_col_detail):
            name = cols[ci]
            if ci == col_3_idx: name = '下载速率（Mbps）'
            elif ci == col_4_idx: name = '上传速率(Mbps)'
            mx = len(str(name))
            arr = _write_arrays[ci]
            for i in range(min(100, n_detail)):
                cv = arr[i]
                if cv is not None and not (isinstance(cv, float) and pd.isna(cv)):
                    s = str(cv)
                    if len(s) > 30: s = s[:30]
                    if len(s) > mx: mx = len(s)
            width = max(mx + 2, 10)
            hidden = (not has_base) and (9 <= ci <= 16)
            if hidden:
                ws.set_column(ci, ci, width, None, {'hidden': True})
            else:
                ws.set_column(ci, ci, width)

        # 冻结 I2 + autofilter + 筛选列=1
        ws.freeze_panes(1, 8)
        ws.autofilter(0, 0, n_detail, n_col_detail - 1)
        ws.filter_column(0, 'x == 1')

        # ===== 汇总（合并）=====
        log("写入汇总...", 85)
        if cancel_check and cancel_check(): raise KeyboardInterrupt('用户取消')
        ws_sum = wb.add_worksheet(f'{_op_name}-汇总')
        hdrs = ['业务类型',
                '应用层FTP上传/下载速率\n阈值以上占比(%)',
                '应用层FTP上传/下载速率\n5M/100M以下占比(%)',
                '应用层平均上传/下载\n速率(Mbps)',
                '削峰应用层平均上传/下载\n速率(Mbps)',
                '上行/下行削峰TOP10%\n峰值速率',
                '业务时长平均值(s)',
                '业务时长中位值(s)',
                '手机号']
        for ci, h in enumerate(hdrs):
            ws_sum.write(0, ci, h, fmt_hdr_s)
        for ci in range(9):
            ws_sum.set_column(ci, ci, 20)

        # 标题（所有识别业务）—— merge A2:I2
        ws_sum.merge_range(1, 0, 1, 8, '▼ 所有识别业务', fmt_ta)
        # 数据行（从 row=2 即 Excel row 3 开始）
        for ri in range(len(df_summary_all)):
            for ci in range(9):
                cv = df_summary_all.iloc[ri, ci] if ci < len(df_summary_all.columns) else None
                if isinstance(cv, float) and pd.isna(cv): cv = None
                _ws_write(ws_sum, ri + 2, ci, cv, fmt_d_s)
        all_data_rows = len(df_summary_all)

        # 标题（仅完整轮次）
        comp_start_excel = all_data_rows + 4
        comp_start0 = comp_start_excel - 1
        ws_sum.merge_range(comp_start0, 0, comp_start0, 8, '▼ 仅完整轮次（代码侧6业务齐全）', fmt_tc)
        for ri in range(len(df_summary_complete)):
            for ci in range(9):
                cv = df_summary_complete.iloc[ri, ci] if ci < len(df_summary_complete.columns) else None
                if isinstance(cv, float) and pd.isna(cv): cv = None
                _ws_write(ws_sum, ri + comp_start0 + 1, ci, cv, fmt_d_s)

        # 行高 22 + 冻结 A3
        last_row_sum = comp_start0 + len(df_summary_complete)
        for r in range(0, last_row_sum + 2):
            ws_sum.set_row(r, 22)
        ws_sum.freeze_panes(2, 0)

        # ===== 批注 + 超链接 =====
        def _fmt_vals(vals, prec=2):
            if vals is None or len(vals) == 0: return '[]'
            vs = [round(v, prec) if prec > 0 else int(round(v)) for v in vals]
            if len(vs) <= 20:
                return '/'.join(str(v) for v in vs)
            return '/'.join(str(v) for v in vs[:5]) + '/.../' + '/'.join(str(v) for v in vs[-5:])

        def _add_comments(sub_detail, row_offset):
            """row_offset: 1-based Excel 行偏移（与原代码一致）"""
            for bi, biz in enumerate(sorted(DL_BIZ | UL_BIZ)):
                if cancel_check and cancel_check():
                    raise KeyboardInterrupt('用户取消')
                sub = sub_detail[sub_detail['businessType'] == biz].dropna(subset=['_t']).sort_values('_t')
                if len(sub) == 0: continue
                dl_cn = [c for c in sub.columns if 'RLC' in c and 'Downlink' in c]
                ul_cn = [c for c in sub.columns if 'RLC' in c and 'Uplink' in c]
                if biz in DL_BIZ and dl_cn: cn = dl_cn[0]
                elif biz in UL_BIZ and ul_cn: cn = ul_cn[0]
                else: continue
                rates = pd.to_numeric(sub[cn], errors='coerce').fillna(0) / 1e6; rates = rates[rates > 0]
                ts2 = sub['_t']; durs = []
                if len(ts2) > 0:
                    s1 = ts2.iloc[0]; p = s1
                    for t in ts2.iloc[1:]:
                        if (t - p).total_seconds() > 5: durs.append((p - s1).total_seconds() + 1); s1 = t
                        p = t
                    durs.append((p - s1).total_seconds() + 1)
                is_dl = biz in DL_BIZ; lim = dl_clip if is_dl else ul_clip; lo = 100 if is_dl else 5
                n = len(rates)
                sr = sub.index; r_start = sr[0] + 2; r_end = sr[-1] + 2
                row_idx_excel = bi + row_offset
                row0 = row_idx_excel - 1

                # 对应的 df_summary
                df_sum_ref = df_summary_all if row_offset == 3 else df_summary_complete
                if bi >= len(df_sum_ref): continue

                for cell_col in range(2, 8):  # 1-based col 2..7
                    col0 = cell_col - 1
                    cv = df_sum_ref.iloc[bi, col0] if col0 < len(df_sum_ref.columns) else None
                    if cv is None or str(cv) in ('nan', 'None', '-', ''): continue
                    cp = cell_col - 2
                    if cp in (0, 1):
                        if cp == 0:
                            mol = [r for r in rates if r > lim]
                            cnt_label = f"计数(rate>{lim})={len(mol)}"
                            op_cn = f"COUNT(rate > {lim}) / COUNT(all) x 100\n= 统计速率大于{lim}的采样数 / 总采样数 x 100"
                        else:
                            mol = [r for r in rates if r < lo]
                            cnt_label = f"计数(rate<{lo})={len(mol)}"
                            op_cn = f"COUNT(rate < {lo}) / COUNT(all) x 100\n= 统计速率小于{lo}的采样数 / 总采样数 x 100"
                        calc = f"= {len(mol)} / {n} x 100"
                    elif cp in (2, 3, 4):
                        if cp == 2:
                            mol = list(rates); sum_v = sum(rates)
                            sum_label = f"求和={sum_v:.1f}"; cnt_label = f"计数={n}"
                            op_cn = "SUM(rate) / COUNT(rate)\n= 各速率求和 / 总采样数"
                            calc = f"= {sum_v:.1f} / {n}"
                        elif cp == 3:
                            clipped = [min(r, lim) for r in rates]; sum_v = sum(clipped); mol = clipped
                            sum_label = f"求和(MIN≤{lim})={sum_v:.1f}"; cnt_label = f"计数={n}"
                            op_cn = f"SUM(MIN(rate,{lim})) / COUNT(rate)\n= 各速率取{lim}上限后求和 / 总采样数"
                            calc = f"= {sum_v:.1f} / {n}"
                        elif cp == 4:
                            sr_rates = sorted(rates, reverse=True)
                            nt = max(1, int(len(sr_rates) * 0.1))
                            top_vals = sr_rates[:nt]; sum_v = sum(top_vals); mol = top_vals
                            sum_label = f"求和(TOP10%)={sum_v:.1f}"; cnt_label = f"计数(TOP10%)={nt}"
                            op_cn = "AVERAGE(top 10%)\n= 前10%最大速率的平均值"
                            calc = f"= {sum_v:.1f} / {nt}"
                    elif cp == 5:
                        continue
                    else:
                        continue
                    col_letter = _gcl(sub.columns.get_loc(cn) + 1)
                    txt = (f"{op_cn}\n{calc}\n= {cv}\n\n"
                           f"分子: {sum_label if cp in (2,3,4) else cnt_label}\n"
                           f"分子值: {_fmt_vals(mol)}\n"
                           f"分母: {cnt_label if cp in (2,3,4) else f'计数={n}'}\n"
                           f"分母值: {_fmt_vals(list(rates))}\n\n"
                           f"分子来源: {cn} 列, 第 {r_start}~{r_end} 行 (详见详细过程 {col_letter}{r_start})\n"
                           f"分母来源: {cn} 列, 第 {r_start}~{r_end} 行")
                    # cell 已在数据行写入时按数字类型写入；此处只加批注（放弃 hyperlink 以保留数字类型，便于排序/筛选）
                    ws_sum.write_comment(row0, col0, txt, {'width': 500, 'height': 220})

                # G/H 列：业务时长批注（无 url）
                dur_vals = [d for d in durs if d > 0]
                if not dur_vals: continue
                for dur_col, dur_label in [(7, 'AVERAGE'), (8, 'MEDIAN')]:
                    col0 = dur_col - 1
                    cv = df_sum_ref.iloc[bi, col0] if col0 < len(df_sum_ref.columns) else None
                    if cv is None or str(cv) in ('nan', 'None', '-'): continue
                    if dur_col == 7:
                        txt = (f"AVERAGE(duration)\n= 各业务段时间求和 / 总段数\n"
                               f"求和={sum(dur_vals):.1f} / 计数={len(dur_vals)}\n= {cv}\n\n"
                               f"值: {_fmt_vals(dur_vals)}\n\n"
                               f"来源: 业务时长(秒) 列, 第 {r_start}~{r_end} 行")
                    else:
                        txt = (f"MEDIAN(duration)\n= 业务段时间中位数\n"
                               f"计数={len(dur_vals)}\n= {cv}\n\n"
                               f"值: {_fmt_vals(dur_vals)}\n\n"
                               f"来源: 业务时长(秒) 列, 第 {r_start}~{r_end} 行")
                    ws_sum.write_comment(row0, col0, txt, {'width': 420, 'height': 180})

        # 上表批注
        _add_comments(df_detail, 3)
        # 下表批注
        _add_comments(df_detail_complete, comp_start_excel + 1)

        # ===== 对比横表 =====
        log(f"写入对比横表... 共{len(df_compare)}行", 93)
        if cancel_check and cancel_check(): raise KeyboardInterrupt('用户取消')
        ws_h = wb.add_worksheet(f'{_op_name}-数据业务-对比')
        hdr = ['轮次', '业务类型', '来源', '开始时间', '结束时间', '速率类指标', '数值', '时长类指标', '数值.1',
               '速率偏差(%)', '时长偏差(%)', '时间差(秒)']
        for ci, h in enumerate(hdr):
            ws_h.write_string(0, ci, h, fmt_hdr_h)

        default_biz_fmt = list(fmt_biz.values())[0] if fmt_biz else fmt_center
        hr = 1; rnd_start = 1; prev_rnd = None
        for ci_df in range(len(df_compare)):
            if cancel_check and ci_df % 10 == 0 and cancel_check():
                raise KeyboardInterrupt('用户取消')
            row = df_compare.iloc[ci_df]
            rnd = str(row['轮次']); biz = str(row['业务类型']); src = str(row['来源'])

            # 轮次切换：合并上一轮的轮次列
            if rnd != prev_rnd and prev_rnd is not None:
                if hr - 1 >= rnd_start:
                    ws_h.merge_range(rnd_start, 0, hr - 1, 0, prev_rnd, fmt_center)
                rnd_start = hr
            prev_rnd = rnd

            fill = fmt_y if '缺失' in str(row.get('开始时间', '')) else fmt_biz.get(biz, default_biz_fmt)

            vals = [
                rnd if hr == rnd_start else '',
                biz, src,
                str(row['开始时间']) if pd.notna(row.get('开始时间')) else '',
                str(row['结束时间']) if pd.notna(row.get('结束时间')) else '',
                str(row['速率类指标']) if pd.notna(row.get('速率类指标')) else '',
                row['数值'] if pd.notna(row.get('数值')) else '',
                str(row['时长类指标']) if pd.notna(row.get('时长类指标')) else '',
                row['数值.1'] if pd.notna(row.get('数值.1')) else '',
                row['速率偏差(%)'] if pd.notna(row.get('速率偏差(%)')) else '',
                row['时长偏差(%)'] if pd.notna(row.get('时长偏差(%)')) else '',
                row['时间差(秒)'] if pd.notna(row.get('时间差(秒)')) else '',
            ]
            # 批量写一行(write_row,同行同 fill)
            row_vals = [('' if (v is None or (isinstance(v, float) and pd.isna(v))) else v) for v in vals]
            ws_h.write_row(hr, 0, row_vals, fill)
            hr += 1

        # 最后一轮的轮次列合并
        if prev_rnd is not None and hr - 1 >= rnd_start:
            ws_h.merge_range(rnd_start, 0, hr - 1, 0, prev_rnd, fmt_center)

        # 每2行合并：业务类型(col1) + 偏差列(col9,10,11)
        for r in range(1, hr, 2):
            if r + 1 > hr - 1: break
            for mc in [1, 9, 10, 11]:
                ws_h.merge_range(r, mc, r + 1, mc, '', fmt_center)

        # 列宽
        for cidx in range(len(hdr)):
            ws_h.set_column(cidx, cidx, 14)
        ws_h.freeze_panes(1, 2)

    log(f"完成! {out}", 100)
    # 临时文件改名为正式输出(避免取消时留下半截坏文件)
    if os.path.exists(_out_tmp):
        os.replace(_out_tmp, out)
    # 清理解压临时目录
    for _d in _tmp_dirs:
        try: shutil.rmtree(_d)
        except Exception: pass
    return out


def _detect_operator(filepath, df=None):
    """从文件路径和数据内容判断运营商：联通/电信

    规则：
    1. 优先从路径名判断（电信/联通）
    2. 如果路径名无法判断且有数据df，通过QCI辅助（7→电信，5/6→联通）
    3. 如果仍无法判断，返回'未知'
    """
    low = filepath.lower()
    if '电信' in low or 'dx' in low:
        return '电信'
    if '联通' in low or 'lt' in low or 'cu' in low:
        return '联通'
    # QCI辅助判定
    if df is not None and 'QCI' in df.columns:
        qci_vals = df['QCI'].dropna().unique()
        if 7 in qci_vals:
            return '电信'
        if 5 in qci_vals or 6 in qci_vals:
            return '联通'
    return '未知'


def _detect_operator_by_path(filepath):
    """从文件夹/文件名判断运营商（增强版，支持数字编号目录名）

    规则：
    1. 文件夹名含"联通"或"电信"→直接判断
    2. 文件夹名含数字编号如"07191845"搭配"联通-0719"→联通
    3. 也可通过mmf里的QCI判断（有7→电信，有5/6→联通）
    """
    low = filepath.lower()
    if '电信' in low or 'dx' in low:
        return '电信'
    if '联通' in low or 'lt' in low or 'cu' in low:
        return '联通'
    return '未知'


def _extract_call_type_from_path(filepath):
    """从文件夹名提取业务类型/主被叫信息

    规则：
    - VoNR主叫 → voNR-主叫
    - VoNR被叫 → voNR-被叫
    - ViNR微信视频-主叫 → viNR-主叫
    - ViNR微信视频-被叫 → viNR-被叫
    - 数据-553 → 数据业务（或含"user_common_monitoring"/".mmf"）
    """
    path_lower = filepath.lower()
    # 语音类型判断
    if 'vonr' in path_lower or 'vnr' in path_lower:
        if '主叫' in path_lower:
            return 'voNR-主叫'
        elif '被叫' in path_lower:
            return 'voNR-被叫'
    if 'vinr' in path_lower or '微信视频' in path_lower:
        if '主叫' in path_lower:
            return 'viNR-主叫'
        elif '被叫' in path_lower:
            return 'viNR-被叫'
    # 数据业务判断
    if '数据' in path_lower or 'user_common_monitoring' in path_lower or path_lower.endswith('.mmf'):
        return '数据业务'
    # 默认返回空，后续通过MessageType列判断
    return ''


def _extract_caller_type_from_path(filepath):
    """从文件夹名提取主被叫（用于按文件名表）

    规则：
    - 含"主叫" → 主叫
    - 含"被叫" → 被叫
    - 数据业务 → 空
    """
    path_lower = filepath.lower()
    if '数据' in path_lower or 'user_common_monitoring' in path_lower or path_lower.endswith('.mmf'):
        return ''
    if '主叫' in path_lower:
        return '主叫'
    if '被叫' in path_lower:
        return '被叫'
    return ''

def _detect_file_type(filename):
    """判断文件类型：voice/voice_vqi 或 data/monitoring

    通过文件名和列名综合判断：
    - 如果文件名包含nr_virtual_ue_trace或.tmf，或有MessageType列 → 返回"语音"
    - 否则 → 返回"数据"
    """
    name = filename.lower()
    if 'nr_virtual_ue_trace' in name or name.endswith('.tmf'):
        return '语音'
    if 'user_common_monitoring' in name or name.endswith('.mmf'):
        return '数据'
    # 对于其他文件，需要读取后通过列名判断，这里默认返回数据
    # 实际使用时会在process_vqi中通过_has_message_type_col再次判断
    return '数据'

_VQI_RE = re.compile(r'DlE2eVqi=(\d+)')
_RSRP_RE = re.compile(r'servRSRP=(-?\d+)(?:,(-?\d+))?')

def _build_anno_text(col_name, row_op, ri, df_voice_out, df_vqi, voice_records, vqi_records, rsrp_records, df_file_stats,
                     op_voice_cache=None, op_vqi_cache=None, op_rsrp_cache=None):
    """构建汇总-数据和语音表的详细批注（使用缓存避免重复过滤）"""
    if col_name == '呼叫建立时长(s)':
        op_vrs = (op_voice_cache or {}).get(row_op, [vr for vr in voice_records if vr.get('运营商') == row_op])
        op_vrs_with_val = [vr for vr in op_vrs if vr.get('呼叫建立时长(s)') != '']
        if op_vrs_with_val:
            details = []
            for vr in op_vrs_with_val:
                rnd = vr.get('轮次', '')
                start = vr.get('开始时间', '')
                end = vr.get('结束时间', '')
                dur = vr.get('呼叫建立时长(s)', '')
                details.append(f"轮次{rnd}: {start}~{end}, 时长={dur}s")
            detail_txt = '\n'.join(details)
            vals = [float(vr['呼叫建立时长(s)']) for vr in op_vrs_with_val]
            avg = round(sum(vals)/len(vals), 2)
            return f"公式: 平均呼叫建立时长(s)\n数据来源: 语音指标统计 -> 呼叫建立时长(s) 列\n计算过程: 对 {row_op} 所有轮次的呼叫建立时长取平均值\n分子(各轮次呼叫建立时长):\n{detail_txt}\n分母(轮次数): {len(vals)}\n计算结果: {avg}"

    elif col_name == '优质通话占比(MOS3.5)':
        op_vqi = (op_vqi_cache or {}).get(row_op, [r for r in vqi_records if r.get('运营商') == row_op])
        total = len(op_vqi)
        quality = sum(1 for r in op_vqi if r.get('是否优质(MOS>3.5)') == '是')
        if total > 0:
            details = []
            for r in op_vqi:
                t = r.get('时间', '')
                mos = r.get('MOS值', '')
                details.append(f"{t}, mos={mos}")
            detail_txt = '\n'.join(details)
            return f"公式: 优质通话占比(MOS3.5)\n数据来源: MOS3.5计算明细 -> 是否优质(MOS>3.5) 列\n计算过程: 优质通话数 / 总通话数\n分子(所有通话):\n{detail_txt}\n分母(总通话数): {total}\n计算结果: {round(quality/total*100,2)}%"

    elif col_name == '服务小区平均RSRP(dBm)(通话时)':
        op_rsrp = (op_rsrp_cache or {}).get(row_op, [r for r in rsrp_records if r.get('运营商') == row_op])
        op_rsrp_in_call = [r for r in op_rsrp if r.get('_in_call_period')]
        vals = [r['servRSRP值'] for r in op_rsrp_in_call]
        if vals:
            nums = '/'.join([str(v) for v in vals])
            avg = round(sum(vals)/len(vals), 2)
            return f"公式: 服务小区平均RSRP(dBm)(通话时)\n数据来源: Detail Info -> servRSRP 列(RRC_MEAS_RPRT行，仅通话时段)\n计算过程: 对 {row_op} 通话时段内的RSRP值取平均\n分子(各RSRP值): {nums}\n分母(RSRP数量): {len(vals)}\n计算结果: {avg}"

    return f"公式: {col_name}\n计算过程: 从原始数据统计\n\n数据来源: {col_name} 列"

def _build_file_anno_text(col_name, fname, voice_records, vqi_records, rsrp_records,
                          fname_voice_cache=None, fname_vqi_cache=None, fname_rsrp_cache=None):
    """构建按文件名表的详细批注（使用缓存避免重复过滤）"""
    if col_name == '通话轮次':
        f_vrs = (fname_voice_cache or {}).get(fname, [vr for vr in voice_records if vr.get('_source_file') == fname])
        if f_vrs:
            rounds_detail = []
            for vr in f_vrs:
                rnd = vr.get('轮次', '')
                start = vr.get('开始时间', '')
                end = vr.get('结束时间', '')
                rounds_detail.append(f"轮次{rnd}: {start} ~ {end}")
            detail_txt = '\n'.join(rounds_detail)
            return f"公式: 通话轮次\n数据来源: 语音指标统计 -> 轮次/开始时间/结束时间 列\n计算过程: 从 {fname} 的语音通话记录中统计\n详情:\n{detail_txt}"

    elif col_name == '平均呼叫建立时长(s)':
        f_vrs = (fname_voice_cache or {}).get(fname, [vr for vr in voice_records if vr.get('_source_file') == fname])
        f_vrs_with_val = [vr for vr in f_vrs if vr.get('呼叫建立时长(s)') != '']
        if f_vrs_with_val:
            details = []
            for vr in f_vrs_with_val:
                rnd = vr.get('轮次', '')
                start = vr.get('开始时间', '')
                end = vr.get('结束时间', '')
                dur = vr.get('呼叫建立时长(s)', '')
                details.append(f"轮次{rnd}: {start}~{end}, 时长={dur}s")
            detail_txt = '\n'.join(details)
            vals = [float(vr['呼叫建立时长(s)']) for vr in f_vrs_with_val]
            avg = round(sum(vals)/len(vals), 2)
            return f"公式: 平均呼叫建立时长(s)\n数据来源: 语音指标统计 -> 呼叫建立时长(s) 列\n计算过程: 对 {fname} 各轮次的呼叫建立时长取平均\n分子(各轮次呼叫建立时长):\n{detail_txt}\n分母(轮次数): {len(vals)}\n计算结果: {avg}"

    elif col_name == '优质通话占比(MOS3.5)':
        f_vqi = (fname_vqi_cache or {}).get(fname, [r for r in vqi_records if r.get('_source_file') == fname])
        total = len(f_vqi)
        quality = sum(1 for r in f_vqi if r.get('是否优质(MOS>3.5)') == '是')
        if total > 0:
            details = []
            for r in f_vqi:
                t = r.get('时间', '')
                mos = r.get('MOS值', '')
                details.append(f"{t}, mos={mos}")
            detail_txt = '\n'.join(details)
            return f"公式: 优质通话占比(MOS3.5)\n数据来源: MOS3.5计算明细 -> 是否优质(MOS>3.5) 列\n计算过程: 优质通话数 / 总通话数\n分子(所有通话):\n{detail_txt}\n分母(总通话数): {total}\n计算结果: {round(quality/total*100,2)}%"

    elif col_name == '服务小区平均RSRP(dBm)(通话时)':
        f_rsrp = (fname_rsrp_cache or {}).get(fname, [r for r in rsrp_records if r.get('_source_file') == fname])
        f_rsrp_in_call = [r for r in f_rsrp if r.get('_in_call_period')]
        vals = [r['servRSRP值'] for r in f_rsrp_in_call]
        if vals:
            nums = '/'.join([str(v) for v in vals])
            avg = round(sum(vals)/len(vals), 2)
            return f"公式: 服务小区平均RSRP(dBm)(通话时)\n数据来源: Detail Info -> servRSRP 列(RRC_MEAS_RPRT行，仅通话时段)\n计算过程: 对 {fname} 通话时段内的RSRP值取平均\n分子(各RSRP值): {nums}\n分母(RSRP数量): {len(vals)}\n计算结果: {avg}"

    return f"公式: {col_name}\n数据来源: {col_name} 列"

def _build_op_anno_text(col_name, op, voice_records, vqi_records, rsrp_records, df_file_stats,
                        op_voice_cache=None, op_vqi_cache=None, op_rsrp_cache=None):
    """构建按运营商表的详细批注（使用缓存避免重复过滤）"""
    if col_name == '通话轮次':
        op_vrs = (op_voice_cache or {}).get(op, [vr for vr in voice_records if vr.get('运营商') == op])
        if op_vrs:
            rounds_detail = []
            for vr in op_vrs:
                rnd = vr.get('轮次', '')
                start = vr.get('开始时间', '')
                end = vr.get('结束时间', '')
                rounds_detail.append(f"轮次{rnd}: {start} ~ {end}")
            detail_txt = '\n'.join(rounds_detail)
            return f"公式: 通话轮次\n数据来源: 语音指标统计 -> 轮次/开始时间/结束时间 列\n计算过程: 从 {op} 的语音通话记录中统计\n详情:\n{detail_txt}"

    elif col_name == '平均呼叫建立时长(s)':
        op_vrs = (op_voice_cache or {}).get(op, [vr for vr in voice_records if vr.get('运营商') == op])
        op_vrs = [vr for vr in op_vrs if vr.get('呼叫建立时长(s)') != '']
        vals = [float(vr['呼叫建立时长(s)']) for vr in op_vrs]
        if vals:
            nums = '/'.join([str(v) for v in vals])
            avg = round(sum(vals)/len(vals), 2)
            return f"公式: 平均呼叫建立时长(s)\n数据来源: 语音指标统计 -> 呼叫建立时长(s) 列\n计算过程: 对 {op} 各轮次的呼叫建立时长取平均\n分子(各轮次呼叫建立时长): {nums}\n分母(轮次数): {len(vals)}\n计算结果: {avg}"

    elif col_name == '优质通话占比(MOS3.5)':
        op_vqi = (op_vqi_cache or {}).get(op, [r for r in vqi_records if r.get('运营商') == op])
        total = len(op_vqi)
        quality = sum(1 for r in op_vqi if r.get('是否优质(MOS>3.5)') == '是')
        if total > 0:
            return f"公式: 优质通话占比(MOS3.5)\n数据来源: MOS3.5计算明细 -> 是否优质(MOS>3.5) 列\n计算过程: 优质通话数 / 总通话数\n分子(优质通话数): {quality}\n分母(总通话数): {total}\n计算结果: {round(quality/total*100,2)}%"

    elif col_name == '服务小区平均RSRP(dBm)':
        op_rsrp = (op_rsrp_cache or {}).get(op, [r for r in rsrp_records if r.get('运营商') == op])
        vals = [r['servRSRP值'] for r in op_rsrp]
        if vals:
            nums = '/'.join([str(v) for v in vals])
            avg = round(sum(vals)/len(vals), 2)
            return f"公式: 服务小区平均RSRP(dBm)\n数据来源: Detail Info -> servRSRP 列(RRC_MEAS_RPRT行)\n计算过程: 对 {op} 的所有RSRP值取平均\n分子(各RSRP值): {nums}\n分母(RSRP数量): {len(vals)}\n计算结果: {avg}"

    return f"公式: {col_name}\n数据来源: {col_name} 列"


def _extract_vqi_vale(detail_info):
    """从Detail Info提取DlE2eVqi值，排除65535"""
    if pd.isna(detail_info):
        return None
    m = _VQI_RE.search(str(detail_info))
    if m:
        val = int(m.group(1))
        return val if val != 65535 else None
    return None

def _extract_serv_rsrp(detail_info):
    """从Detail Info提取servRSRP值并返回最大值

    格式示例: servRSRP=-96,-91 或 servRSRP=-96
    如果有两个值，取最大的那个（例如-96,-91中取-91，因为-91 > -96）
    如果只有一个值，直接返回该值
    """
    if pd.isna(detail_info):
        return None
    m = _RSRP_RE.search(str(detail_info))
    if m:
        val1 = int(m.group(1))
        val2 = int(m.group(2)) if m.group(2) else None
        if val2 is not None:
            return max(val1, val2)  # 取较大的值（例如-91 > -96）
        return val1
    return None

def process_vqi(files, callback=None, cancel_check=None, progress_cb=None, time_filter=None, time_range=None, merge_raw=True, add_annotations=True, phone_trace_map=None):
    """VQI处理函数：自动分类语音/数据文件，生成语音统计输出

    参数:
        files: 输入文件列表
        callback: 日志回调函数
        cancel_check: 取消检查函数
        progress_cb: 进度回调函数
        time_filter: 时间段过滤 (use_time, start_time, end_time)
        time_range: 时间起止范围，用于汇总表。若为None则从数据自动计算
    """
    def log(msg, pct=None):
        if cancel_check and cancel_check():
            raise KeyboardInterrupt('用户取消')
        import time as _time
        _t = _time.time()
        if not hasattr(log, 'last_time'):
            log.last_time = _t
        elapsed = _t - log.last_time if pct is not None else 0
        print(msg)
        if callback: callback(msg)
        if pct is not None and progress_cb: progress_cb(pct)
        log.last_time = _t

    import zipfile, tempfile, shutil, glob

    # 展开压缩文件
    _tmp_dirs = []
    def _expand_zip(zip_path):
        _tmp = tempfile.mkdtemp(prefix='vqi_unzip_'); _tmp_dirs.append(_tmp)
        _xs = []
        try:
            with zipfile.ZipFile(zip_path) as _zf: _zf.extractall(_tmp)
            _xs = glob.glob(os.path.join(_tmp, '**', '*.xlsx'), recursive=True) + \
                  glob.glob(os.path.join(_tmp, '**', '*.xls'), recursive=True) + \
                  glob.glob(os.path.join(_tmp, '**', '*.mmf'), recursive=True) + \
                  glob.glob(os.path.join(_tmp, '**', '*.tmf'), recursive=True)
            _inner_zips = glob.glob(os.path.join(_tmp, '**', '*.zip'), recursive=True)
            for _iz in _inner_zips:
                _xs.extend(_expand_zip(_iz))
        except Exception as _e:
            log(f"  解压失败 {os.path.basename(zip_path)}: {_e}")
        return _xs

    # 展开所有文件
    all_files = []
    for _f in (files or []):
        if str(_f).lower().endswith('.zip'):
            _xs = _expand_zip(_f)
            log(f"  解压 {os.path.basename(_f)}: 找到 {len(_xs)} 个文件(含嵌套)", 5)
            all_files.extend(_xs)
        else:
            all_files.append(_f)

    if not all_files:
        log("无有效文件", 100)
        return None

    log(f"共 {len(all_files)} 个文件", 5)

    # 分类文件：读取每个文件检查是否有MessageType列
    # 有MessageType → 语音文件；没有 → 数据文件（跳过处理）
    voice_files = []
    data_files = []
    data_files_skipped = 0
    for f in all_files:
        try:
            df_test = _read_file(f, nrows=3)
            has_msg = _has_message_type_col(df_test)
            log(f"  文件 {os.path.basename(f)}: {'✓ MessageType列' if has_msg else 'x 无MessageType列'}", 6)
            if has_msg:
                voice_files.append(f)
            else:
                data_files.append(f)
        except Exception as e:
            log(f"  文件 {os.path.basename(f)}: 读取失败 - {e}", 6)
            data_files_skipped += 1

    log(f"  语音文件: {len(voice_files)} 个, 数据文件: {len(data_files)} 个, 跳过的数据文件: {data_files_skipped} 个", 8)

    # ===== 读取语音文件 =====
    voice_dfs = []
    voice_file_info = {}  # 记录每个文件的基本信息（行数、时间范围等）
    for fi, f in enumerate(voice_files):
        log(f"  读取语音文件 {fi+1}/{len(voice_files)}: {os.path.basename(f)}", 10 + int(10 * (fi+1) / max(len(voice_files), 1)))
        operator = _detect_operator(f)
        try:
            df_i = _read_file(f)
        except Exception:
            log(f"    读取失败: {f}", 10)
            continue
        # 统一MessageType列名（支持Message Type、message_type等变体）
        if 'Message Type' in df_i.columns and 'MessageType' not in df_i.columns:
            df_i = df_i.rename(columns={'Message Type': 'MessageType'})
        elif 'message_type' in df_i.columns and 'MessageType' not in df_i.columns:
            df_i = df_i.rename(columns={'message_type': 'MessageType'})
        df_i['_t'] = _parse_time_vec(df_i['Time'])
        df_i['运营商'] = operator
        df_i['业务类型'] = '语音业务'
        # 记录来源文件名，用于后续按文件统计
        filename_base = os.path.basename(f)
        # 记录相对路径（含文件夹信息），便于在不同层级的文件夹中区分文件来源
        rel_path = f
        work_dir = os.getcwd()
        if rel_path.startswith(work_dir):
            rel_path = rel_path[len(work_dir) + 1:]
        df_i['_source_file'] = rel_path
        voice_dfs.append(df_i)

        # 记录该文件的时间范围和行数
        times = df_i['_t'].dropna()
        time_range_str = ''
        if len(times) > 0:
            time_range_str = f"{times.min().strftime('%Y-%m-%d %H:%M:%S')} ~ {times.max().strftime('%Y-%m-%d %H:%M:%S')}"
        voice_file_info[rel_path] = {
            'operator': operator,
            'row_count': len(df_i),
            'time_range': time_range_str,
            'time_start': times.min() if len(times) > 0 else None,
            'time_end': times.max() if len(times) > 0 else None,
            'full_path': f
        }

    # 合并所有语音数据
    if voice_dfs:
        df_voice_raw = pd.concat(voice_dfs, ignore_index=True)
        df_voice_raw = df_voice_raw.sort_values('_t', kind='stable').reset_index(drop=True)
    else:
        df_voice_raw = pd.DataFrame()

    log(f"语音数据合并完成: {len(df_voice_raw)} 行", 22)

    # ===== 数据文件处理 =====
    log("处理数据文件...", 23)
    data_dfs = []
    data_file_info = {}
    for fi, f in enumerate(data_files):
        if cancel_check and cancel_check(): raise KeyboardInterrupt('用户取消')
        log(f"  读取数据文件 {fi+1}/{len(data_files)}: {os.path.basename(f)}", 24 + int(5 * (fi+1) / max(len(data_files), 1)))
        operator = _detect_operator(f)
        try:
            df_i = _read_file(f)
        except Exception:
            log(f"    读取失败: {f}", 24)
            continue
        has_time = 'Time' in df_i.columns
        has_qci = 'QCI' in df_i.columns
        has_rlc = any('RLC' in c and ('Downlink' in c or 'Uplink' in c) for c in df_i.columns)
        if not (has_time and (has_qci or has_rlc)):
            log(f"    跳过非数据文件(无Time/QCI/RLC列)", 24)
            data_files_skipped += 1
            continue
        
        df_i['_t'] = _parse_time_vec(df_i['Time'])
        df_i['运营商'] = operator
        df_i['业务类型'] = '数据业务'
        
        rlc_dl = [c for c in df_i.columns if 'RLC' in c and 'Downlink' in c]
        rlc_ul = [c for c in df_i.columns if 'RLC' in c and 'Uplink' in c]
        if rlc_dl:
            df_i['_dl_rlc'] = pd.to_numeric(df_i[rlc_dl[0]], errors='coerce').fillna(0) / 1e6
        else:
            df_i['_dl_rlc'] = 0
        if rlc_ul:
            df_i['_ul_rlc'] = pd.to_numeric(df_i[rlc_ul[0]], errors='coerce').fillna(0) / 1e6
        else:
            df_i['_ul_rlc'] = 0
        
        data_dfs.append(df_i)
        
        dl_vals = df_i['_dl_rlc'].dropna().values
        ul_vals = df_i['_ul_rlc'].dropna().values
        dl_nonzero = dl_vals[dl_vals > 0]
        ul_nonzero = ul_vals[ul_vals > 0]
        
        dl_peak = 0
        if len(dl_nonzero) > 0:
            tn = max(1, int(len(dl_nonzero) * 0.1))
            dl_sorted = np.sort(dl_nonzero)[::-1]
            dl_peak = round(min(float(np.mean(dl_sorted[:tn])), 1000), 2)
        
        dl_avg = 0
        if len(dl_nonzero) > 0:
            dl_avg = round(min(float(np.mean(dl_nonzero)), 1000), 2)
        
        ul_peak = 0
        if len(ul_nonzero) > 0:
            tn = max(1, int(len(ul_nonzero) * 0.1))
            ul_sorted = np.sort(ul_nonzero)[::-1]
            ul_peak = round(min(float(np.mean(ul_sorted[:tn])), 200), 2)
        
        ul_avg = 0
        if len(ul_nonzero) > 0:
            ul_avg = round(min(float(np.mean(ul_nonzero)), 200), 2)
        
        times = df_i['_t'].dropna()
        time_range_str = ''
        if len(times) > 0:
            time_range_str = f"{times.min().strftime('%Y-%m-%d %H:%M:%S')} ~ {times.max().strftime('%Y-%m-%d %H:%M:%S')}"
        
        data_file_info[os.path.basename(f)] = {
            'operator': operator,
            'row_count': len(df_i),
            'time_range': time_range_str,
            'dl_peak': dl_peak,
            'dl_avg': dl_avg,
            'ul_peak': ul_peak,
            'ul_avg': ul_avg,
            'full_path': f
        }
    
    # 添加手机号到data_file_info
    for fname, info in data_file_info.items():
        trace_id = _extract_trace_id(info.get('full_path', ''))
        info['phone'] = _lookup_phone(trace_id, phone_trace_map)

    if data_dfs:
        df_data_raw = pd.concat(data_dfs, ignore_index=True)
        df_data_raw = df_data_raw.sort_values('_t', kind='stable').reset_index(drop=True)
    else:
        df_data_raw = pd.DataFrame()

    log(f"数据文件处理完成: {len(data_dfs)} 个文件, {len(df_data_raw)} 行", 28)
    
    data_operator_indicators = {}
    actual_ops = set(info['operator'] for info in data_file_info.values())
    for op in sorted(actual_ops):
        op_data = [info for fname, info in data_file_info.items() if info['operator'] == op]
        if op_data:
            dl_peaks = [d['dl_peak'] for d in op_data if d['dl_peak'] > 0]
            dl_avgs = [d['dl_avg'] for d in op_data if d['dl_avg'] > 0]
            ul_peaks = [d['ul_peak'] for d in op_data if d['ul_peak'] > 0]
            ul_avgs = [d['ul_avg'] for d in op_data if d['ul_avg'] > 0]
            data_operator_indicators[op] = {
                '下行前10%峰值速率': round(np.mean(dl_peaks), 2) if dl_peaks else '',
                '下行均值速率': round(np.mean(dl_avgs), 2) if dl_avgs else '',
                '上行前10%峰值速率': round(np.mean(ul_peaks), 2) if ul_peaks else '',
                '上行均值速率': round(np.mean(ul_avgs), 2) if ul_avgs else ''
            }
        else:
            data_operator_indicators[op] = {
                '下行前10%峰值速率': '',
                '下行均值速率': '',
                '上行前10%峰值速率': '',
                '上行均值速率': ''
            }


    # ===== 时间段过滤 =====
    if time_filter and len(df_voice_raw) > 0:
        use_time, t_start, t_end = time_filter
        if use_time and t_start is not None and t_end is not None:
            t_start_t = t_start.toPython() if hasattr(t_start, 'toPython') else t_start
            t_end_t = t_end.toPython() if hasattr(t_end, 'toPython') else t_end
            df_voice_raw['_time_only'] = df_voice_raw['_t'].dt.time
            mask = (df_voice_raw['_time_only'] >= t_start_t) & (df_voice_raw['_time_only'] <= t_end_t)
            df_voice_raw = df_voice_raw[mask].copy()
            df_voice_raw = df_voice_raw.drop('_time_only', axis=1)
            log(f"  时间段过滤: {t_start_t} ~ {t_end_t}, 剩余 {len(df_voice_raw)} 行", 22)

    # ===== 筛选MessageType非空行 =====
    if 'MessageType' in df_voice_raw.columns and len(df_voice_raw) > 0:
        df_msg = df_voice_raw[df_voice_raw['MessageType'].notna() &
                              (df_voice_raw['MessageType'].astype(str).str.strip() != '')].copy()
    else:
        df_msg = pd.DataFrame()
    log(f"  MessageType非空行数: {len(df_msg)}", 28)

    # ===== 语音统计（INVITE→180→BYE→200） =====
    log("识别语音通话...", 35)
    voice_records = []
    rnd = 0

    if len(df_msg) > 0 and 'MessageType' in df_msg.columns and '运营商' in df_msg.columns:
        # 按运营商分别处理信令，避免不同运营商的信令交叉干扰
        operators_in_data = df_msg['运营商'].dropna().unique()
        for op in operators_in_data:
            if str(op) in ('未知', '', 'None'):
                continue
            df_op = df_msg[df_msg['运营商'] == op].copy()
            if len(df_op) == 0:
                continue
            log(f"  处理{op}信令({len(df_op)}行)...", 36)

            # 向量化：找出该运营商的所有INVITE位置
            invite_mask = df_op['MessageType'].astype(str) == 'SIP_REQ_INVITE'
            invite_positions = df_op.index[invite_mask].tolist()
            n_op = len(df_op)

            # 过滤掉连续的INVITE（时间差<1秒算连续，跳过中间的）
            valid_invite_positions = []
            for idx_val in invite_positions:
                if valid_invite_positions:
                    prev_val = valid_invite_positions[-1]
                    if abs((df_op.loc[idx_val, '_t'] - df_op.loc[prev_val, '_t']).total_seconds()) < 1.0:
                        continue
                valid_invite_positions.append(idx_val)

            invite_positions = valid_invite_positions
            log(f"  找到 {len(invite_positions)} 个有效{op}INVITE", 36)

            for ii, pos in enumerate(invite_positions):
                t1 = df_op.loc[pos, '_t']
                t1_time = df_op.loc[pos, 'Time']
                operator = str(op)

                # 搜索范围：从当前INVITE到下一个同运营商INVITE（或到数据末尾）
                next_pos = invite_positions[ii + 1] if ii + 1 < len(invite_positions) else (df_op.index[-1] + 1)
                end_j = next_pos - 1

                if pos + 1 > end_j:
                    rnd += 1
                    voice_records.append({
                        '轮次': rnd,
                        '运营商': operator,
                        '开始时间': t1_time,
                        '结束时间': t1_time,
                        '呼叫建立时长(s)': '',
                        '是否完整通话': '否',
                        '主被叫': _extract_caller_type_from_path(df_op.loc[pos].get('_source_file', '')),
                        '_source_file': df_op.loc[pos].get('_source_file', '')
                    })
                    continue

                search_range = df_op.loc[pos + 1:end_j]

                # 找SIP_RSP_180（第一个）
                r180 = search_range[search_range['MessageType'].astype(str) == 'SIP_RSP_180']
                t2 = r180.iloc[0]['_t'] if len(r180) > 0 and '_t' in r180.columns else None

                # 找SIP_REQ_BYE（只要BYE就算完整通话）
                bye = search_range[search_range['MessageType'].astype(str) == 'SIP_REQ_BYE']
                bye_time_raw = None
                last_time_raw = search_range.iloc[-1]['Time'] if len(search_range) > 0 and 'Time' in search_range.columns else t1_time

                if len(bye) > 0:
                    bye_time_raw = bye.iloc[0]['Time'] if 'Time' in bye.columns else None
                    last_time_raw = bye_time_raw
                else:
                    last_time_raw = search_range.iloc[-1]['Time'] if len(search_range) > 0 and 'Time' in search_range.columns else t1_time

                call_setup_duration = ''
                if t2 is not None:
                    call_setup_duration = round((t2 - t1).total_seconds(), 2)

                is_complete = '是' if len(bye) > 0 else '否'

                rnd += 1
                voice_records.append({
                    '轮次': rnd,
                    '运营商': operator,
                    '开始时间': t1_time,
                    '结束时间': last_time_raw,
                    'INVITE时间': t1_time,
                    'BYE时间': bye_time_raw if len(bye) > 0 else '',
                    '主被叫': _extract_caller_type_from_path(df_op.loc[pos].get('_source_file', '')),
                    '呼叫建立时长(s)': call_setup_duration,
                    '是否完整通话': is_complete,
                    '_source_file': df_op.loc[pos].get('_source_file', '')
                })

    log(f"识别完成: {len(voice_records)} 轮语音通话", 50)

    # ===== DlE2eVqi提取（从Detail Info列）向量化优化 =====
    log("提取DlE2eVqi值...", 55)
    vqi_records = []

    # 找Detail Info列（支持中英文名）
    detail_col = None
    if len(df_voice_raw) > 0:
        for col_name in ['Detail Info', '详细信息', 'DetailInfo', 'detail_info']:
            if col_name in df_voice_raw.columns:
                detail_col = col_name
                break

    if detail_col and len(df_voice_raw) > 0:
        # 最终优化：使用.values批量取列值，避免.at索引开销
        detail_vals = df_voice_raw[detail_col].values
        time_vals = df_voice_raw['Time'].values if 'Time' in df_voice_raw.columns else None
        op_vals = df_voice_raw['运营商'].values if '运营商' in df_voice_raw.columns else None
        sf_vals = df_voice_raw['_source_file'].values if '_source_file' in df_voice_raw.columns else None

        for i in range(len(detail_vals)):
            v = detail_vals[i]
            if v is None or pd.isna(v):
                continue
            s = str(v)
            if 'DlE2eVqi=' not in s:
                continue
            m = _VQI_RE.search(s)
            if m:
                vqi = int(m.group(1))
                if vqi == 65535:
                    continue
                mos = vqi / 100.0
                is_quality = '是' if mos > 3.5 else '否'
                vqi_records.append({
                    '时间': time_vals[i] if time_vals is not None else '',
                    '运营商': op_vals[i] if op_vals is not None else '',
                    'DlE2eVqi原始值': vqi,
                    'MOS值': round(mos, 2),
                    '是否优质(MOS>3.5)': is_quality,
                    '_source_file': sf_vals[i] if sf_vals is not None else ''
                })

    df_vqi = pd.DataFrame(vqi_records)

    # 计算优质占比
    if len(df_vqi) > 0:
        total = len(df_vqi)
        quality = (df_vqi['是否优质(MOS>3.5)'] == '是').sum()
        quality_pct = round(quality / total * 100, 2)
    else:
        total, quality, quality_pct = 0, 0, 0

    # ===== 服务小区平均RSRP提取（从RRC_MEAS_RPRT的Detail Info列）优化 =====
    log("提取服务小区RSRP值...", 58)
    rsrp_records = []

    # 先构建通话时段字典(用于通话时RSRP判定)
    call_periods_by_file = {}
    for vr in voice_records:
        fname = vr.get('_source_file', '')
        if not fname:
            continue
        if vr.get('是否完整通话') != '是':
            continue
        start_t = vr.get('开始时间', '')
        end_t = vr.get('结束时间', '')
        if start_t and end_t:
            if fname not in call_periods_by_file:
                call_periods_by_file[fname] = []
            call_periods_by_file[fname].append((start_t, end_t))

    if detail_col and len(df_voice_raw) > 0 and 'MessageType' in df_voice_raw.columns:
        # 先筛选MessageType列，只处理RRC_MEAS_RPRT行
        # 使用values+字符串比较，避免astype(str)全量转换
        msg_types = df_voice_raw['MessageType'].values
        rrc_indices = []
        for i in range(len(msg_types)):
            val = msg_types[i]
            if val is not None and not pd.isna(val) and str(val) == 'RRC_MEAS_RPRT':
                rrc_indices.append(i)

        log(f"  筛选到 {len(rrc_indices)} 个RRC_MEAS_RPRT行", 59)

        if len(rrc_indices) > 0:
            # 也使用.values优化
            detail_vals = df_voice_raw[detail_col].values
            time_vals = df_voice_raw['Time'].values if 'Time' in df_voice_raw.columns else None
            op_vals = df_voice_raw['运营商'].values if '运营商' in df_voice_raw.columns else None
            sf_vals = df_voice_raw['_source_file'].values if '_source_file' in df_voice_raw.columns else None

            for idx in rrc_indices:
                dv = detail_vals[idx]
                if dv is None or pd.isna(dv):
                    continue
                m = _RSRP_RE.search(str(dv))
                if m:
                    v1 = int(m.group(1))
                    v2 = int(m.group(2)) if m.group(2) else None
                    rsrp_val = max(v1, v2) if v2 is not None else v1
                    # 判断是否在通话时段内
                    in_call = False
                    sf = sf_vals[idx] if sf_vals is not None else ''
                    t_str = time_vals[idx] if time_vals is not None else ''
                    if sf and t_str:
                        periods = call_periods_by_file.get(sf, [])
                        for s_t, e_t in periods:
                            if s_t and e_t and s_t <= t_str <= e_t:
                                in_call = True
                                break
                    rsrp_records.append({
                        '时间': t_str,
                        '运营商': op_vals[idx] if op_vals is not None else '',
                        'servRSRP值': rsrp_val,
                        '_source_file': sf,
                        '_in_call_period': in_call
                    })

    # 计算平均RSRP
    avg_rsrp = None
    if rsrp_records:
        rsrp_vals = [r['servRSRP值'] for r in rsrp_records]
        avg_rsrp = round(sum(rsrp_vals) / len(rsrp_vals), 2)
        log(f"  提取到 {len(rsrp_records)} 个RSRP值，平均值: {avg_rsrp}", 58)

    # ===== 生成输出 =====
    # 原始数据sheet：所有语音文件的全部行（加运营商和业务类型列）
    df_raw_out = df_voice_raw.copy()
    if '_t' in df_raw_out.columns:
        df_raw_out = df_raw_out.drop('_t', axis=1)
    if '_time_only' in df_raw_out.columns:
        df_raw_out = df_raw_out.drop('_time_only', axis=1)

    # 语音统计sheet
    df_voice_out = pd.DataFrame(voice_records)
    if len(df_voice_out) == 0:
        df_voice_out = pd.DataFrame(columns=['轮次', '运营商', '开始时间', '结束时间', '呼叫建立时长(s)', '是否完整通话'])

    # MOS3.5计算明细sheet（底部加汇总）
    if len(df_vqi) > 0:
        # 底部汇总行
        summary_row = pd.DataFrame([{
            '时间': '汇总',
            '运营商': '',
            'DlE2eVqi原始值': '',
            'MOS值': '',
            '是否优质(MOS>3.5)': f'分子={quality} / 分母={total} = {quality_pct}%'
        }])
        df_vqi_out = pd.concat([df_vqi, summary_row], ignore_index=True)
    else:
        df_vqi_out = pd.DataFrame(columns=['时间', '运营商', 'DlE2eVqi原始值', 'MOS值', '是否优质(MOS>3.5)'])

    # ===== 按文件名统计 =====
    file_stats = []

    # 语音文件统计（使用voice_file_info和_source_file字段）
    for fname_base, f_info in voice_file_info.items():
        operator = f_info['operator']
        file_row_count = f_info['row_count']
        file_time_range_str = f_info['time_range']

        # 从文件夹名提取主被叫
        caller_type = _extract_caller_type_from_path(f_info.get('full_path', ''))

        # 从voice_records中筛选属于该文件的轮次（通过_source_file匹配）
        file_voice_records = [vr for vr in voice_records if vr.get('_source_file') == fname_base]
        file_rounds = len(file_voice_records)

        # 平均呼叫建立时长（只统计有值的轮次）
        setup_vals = []
        for vr in file_voice_records:
            if vr.get('呼叫建立时长(s)') != '' and pd.notna(vr.get('呼叫建立时长(s)')):
                try:
                    setup_vals.append(float(vr['呼叫建立时长(s)']))
                except:
                    pass
        file_avg_setup = round(np.mean(setup_vals), 2) if setup_vals else ''

        # 优质通话占比(MOS3.5)：从vqi_records中筛选属于该文件的记录
        file_vqi_records = [r for r in vqi_records if r.get('_source_file') == fname_base]
        file_quality_pct = ''
        if file_vqi_records:
            total_file = len(file_vqi_records)
            quality_file = sum(1 for r in file_vqi_records if r.get('是否优质(MOS>3.5)') == '是')
            if total_file > 0:
                file_quality_pct = f"{round(quality_file / total_file * 100, 2)}%"

        # 服务小区平均RSRP（全部）：从rsrp_records中筛选属于该文件的记录
        file_rsrp_records = [r for r in rsrp_records if r.get('_source_file') == fname_base]
        file_avg_rsrp = ''
        if file_rsrp_records:
            rsrp_vals = [r['servRSRP值'] for r in file_rsrp_records]
            file_avg_rsrp = round(sum(rsrp_vals) / len(rsrp_vals), 2)

        # 服务小区平均RSRP（通话时）：筛选_in_call_period=True的记录
        file_rsrp_call = [r for r in rsrp_records if r.get('_source_file') == fname_base and r.get('_in_call_period')]
        file_avg_rsrp_call = ''
        if file_rsrp_call:
            rsrp_call_vals = [r['servRSRP值'] for r in file_rsrp_call]
            file_avg_rsrp_call = round(sum(rsrp_call_vals) / len(rsrp_call_vals), 2)

        # 手机号查询：从voice_file_info中取full_path提取trace_id
        voice_trace_id = _extract_trace_id(f_info.get('full_path', ''))
        voice_phone = _lookup_phone(voice_trace_id, phone_trace_map)

        file_stats.append({
            '文件名': fname_base,
            '运营商': operator,
            '文件类型': '语音',
            '主被叫': caller_type,
            '手机号': voice_phone,
            '行数': file_row_count,
            '时间范围': file_time_range_str,
            '通话轮次': file_rounds,
            '平均呼叫建立时长(s)': file_avg_setup,
            '优质通话占比(MOS3.5)': file_quality_pct,
            '服务小区平均RSRP(dBm)(全部)': file_avg_rsrp,
            '服务小区平均RSRP(dBm)(通话时)': file_avg_rsrp_call,
            '下行前10%峰值速率': '',
            '下行均值速率': '',
            '上行前10%峰值速率': '',
            '上行均值速率': '',
            '微信大文件上传中位数时长(s)': '',
            '微信大文件发送平均速率(Mbps)': '',
            '微信小文件发送平均时延(s)': '',
            '微信小文件发送平均速率(Mbps)': '',
            '商店大app平均时延(s)': '',
            '商店大app平均速率(Mbps)': '',
            '商店小app平均时延(s)': '',
            '商店小app平均速率(Mbps)': ''
        })

    # 添加数据文件行到按文件名统计
    for fname, info in data_file_info.items():
        # 手机号查询
        data_trace_id = _extract_trace_id(info.get('full_path', ''))
        data_phone = _lookup_phone(data_trace_id, phone_trace_map)

        # 数据文件的主被叫为空
        caller_type = ''

        file_stats.append({
            '文件名': fname,
            '运营商': info['operator'],
            '文件类型': '数据',
            '主被叫': caller_type,
            '手机号': data_phone,
            '行数': info['row_count'],
            '时间范围': info['time_range'],
            '通话轮次': 0,
            '平均呼叫建立时长(s)': '',
            '优质通话占比(MOS3.5)': '',
            '服务小区平均RSRP(dBm)(全部)': '',
            '服务小区平均RSRP(dBm)(通话时)': '',
            '下行前10%峰值速率': info['dl_peak'],
            '下行均值速率': info['dl_avg'],
            '上行前10%峰值速率': info['ul_peak'],
            '上行均值速率': info['ul_avg'],
            '微信大文件上传中位数时长(s)': '',
            '微信大文件发送平均速率(Mbps)': '',
            '微信小文件发送平均时延(s)': '',
            '微信小文件发送平均速率(Mbps)': '',
            '商店大app平均时延(s)': '',
            '商店大app平均速率(Mbps)': '',
            '商店小app平均时延(s)': '',
            '商店小app平均速率(Mbps)': ''
        })

    df_file_stats = pd.DataFrame(file_stats)

    # ===== 时间起止 =====
    if time_range:
        time_range_str = time_range
    elif time_filter and time_filter[0] and time_filter[1] is not None and time_filter[2] is not None:
        time_range_str = f"{time_filter[1].toString('HH:mm:ss')} ~ {time_filter[2].toString('HH:mm:ss')}"
    else:
        # 从数据中计算
        if len(df_voice_raw) > 0 and '_t' in df_voice_raw.columns and df_voice_raw['_t'].notna().any():
            t_min = df_voice_raw['_t'].min()
            t_max = df_voice_raw['_t'].max()
            time_range_str = f"{t_min.strftime('%Y-%m-%d %H:%M:%S')} ~ {t_max.strftime('%Y-%m-%d %H:%M:%S')}"
        else:
            time_range_str = ''

    # ===== 汇总-数据和语音 =====
    summary_data = []

    # 标准行
    summary_data.append({
        '运营商': '标准',
        '文件名': '',
        '业务起止时间': '',
        '类型(主叫/viNR-主叫/viNR-被叫/voNR-主叫/voNR-被叫/数据业务)': '',
        '手机号': '',
        '呼叫建立时长(s)': 1,
        '优质通话占比(MOS3.5)': '95%',
        '服务小区平均RSRP(dBm)(全部)': '',
        '服务小区平均RSRP(dBm)(通话时)': '',
        '下行前10%峰值速率': 1000,
        '下行均值速率': 800,
        '上行前10%峰值速率': 200,
        '上行均值速率': 160,
        '微信大文件上传中位数时长(s)': 30,
        '微信大文件发送平均速率(Mbps)': 50,
        '微信小文件发送平均时延(s)': 2,
        '微信小文件发送平均速率(Mbps)': 20,
        '商店大app平均时延(s)': 20,
        '商店大app平均速率(Mbps)': 400,
        '商店小app平均时延(s)': 4,
        '商店小app平均速率(Mbps)': 200
    })

    # 构建汇总行：每个文件一条（一个文件名 + 一个类型 + 一个手机号）
    # 语音文件
    for fname_base, f_info in voice_file_info.items():
        operator = f_info['operator']
        # 从文件夹名提取类型
        call_type = _extract_call_type_from_path(f_info.get('full_path', ''))
        if not call_type:
            call_type = 'voNR-主叫'  # fallback

        # 业务起止时间：该文件第一个INVITE时间 ~ 最后一个BYE/结束时间
        f_vrs = [vr for vr in voice_records if vr.get('_source_file') == fname_base]
        biz_start = ''
        biz_end = ''
        if f_vrs:
            if f_vrs[0].get('开始时间'):
                biz_start = str(f_vrs[0]['开始时间'])
            if f_vrs[-1].get('结束时间'):
                biz_end = str(f_vrs[-1]['结束时间'])
        biz_time_range = f"{biz_start} ~ {biz_end}" if biz_start and biz_end else f_info.get('time_range', '')

        # 手机号
        trace_id = _extract_trace_id(f_info.get('full_path', ''))
        phone = _lookup_phone(trace_id, phone_trace_map)

        # 呼叫建立时长（平均）
        setup_vals = []
        for vr in f_vrs:
            if vr.get('呼叫建立时长(s)') != '' and pd.notna(vr.get('呼叫建立时长(s)')):
                try:
                    setup_vals.append(float(vr['呼叫建立时长(s)']))
                except:
                    pass
        avg_setup = round(np.mean(setup_vals), 2) if setup_vals else ''

        # 优质通话占比
        f_vqi = [r for r in vqi_records if r.get('_source_file') == fname_base]
        quality_pct = ''
        if f_vqi:
            total_f = len(f_vqi)
            quality_f = sum(1 for r in f_vqi if r.get('是否优质(MOS>3.5)') == '是')
            if total_f > 0:
                quality_pct = f"{round(quality_f / total_f * 100, 2)}%"

        # RSRP（全部）
        f_rsrp = [r for r in rsrp_records if r.get('_source_file') == fname_base]
        avg_rsrp = ''
        if f_rsrp:
            rsrp_vals = [r['servRSRP值'] for r in f_rsrp]
            avg_rsrp = round(sum(rsrp_vals) / len(rsrp_vals), 2)

        # RSRP（通话时）
        f_rsrp_call = [r for r in rsrp_records if r.get('_source_file') == fname_base and r.get('_in_call_period')]
        avg_rsrp_call = ''
        if f_rsrp_call:
            rsrp_call_vals = [r['servRSRP值'] for r in f_rsrp_call]
            avg_rsrp_call = round(sum(rsrp_call_vals) / len(rsrp_call_vals), 2)

        summary_data.append({
            '运营商': operator,
            '文件名': fname_base,
            '业务起止时间': biz_time_range,
            '类型(主叫/viNR-主叫/viNR-被叫/voNR-主叫/voNR-被叫/数据业务)': call_type,
            '手机号': phone,
            '呼叫建立时长(s)': avg_setup,
            '优质通话占比(MOS3.5)': quality_pct,
            '服务小区平均RSRP(dBm)(全部)': avg_rsrp,
            '服务小区平均RSRP(dBm)(通话时)': avg_rsrp_call,
            '下行前10%峰值速率': '',
            '下行均值速率': '',
            '上行前10%峰值速率': '',
            '上行均值速率': '',
            '微信大文件上传中位数时长(s)': '',
            '微信大文件发送平均速率(Mbps)': '',
            '微信小文件发送平均时延(s)': '',
            '微信小文件发送平均速率(Mbps)': '',
            '商店大app平均时延(s)': '',
            '商店大app平均速率(Mbps)': '',
            '商店小app平均时延(s)': '',
            '商店小app平均速率(Mbps)': ''
        })

    # 数据文件
    for fname, info in data_file_info.items():
        operator = info['operator']
        call_type = '数据业务'
        biz_time_range = info.get('time_range', '')

        trace_id = _extract_trace_id(info.get('full_path', ''))
        phone = _lookup_phone(trace_id, phone_trace_map)

        # 数据业务的呼叫建立时长/优质通话占比/RSRP为空
        summary_data.append({
            '运营商': operator,
            '文件名': fname,
            '业务起止时间': biz_time_range,
            '类型(主叫/viNR-主叫/viNR-被叫/voNR-主叫/voNR-被叫/数据业务)': call_type,
            '手机号': phone,
            '呼叫建立时长(s)': '',
            '优质通话占比(MOS3.5)': '',
            '服务小区平均RSRP(dBm)(全部)': '',
            '服务小区平均RSRP(dBm)(通话时)': '',
            '下行前10%峰值速率': info['dl_peak'],
            '下行均值速率': info['dl_avg'],
            '上行前10%峰值速率': info['ul_peak'],
            '上行均值速率': info['ul_avg'],
            '微信大文件上传中位数时长(s)': '',
            '微信大文件发送平均速率(Mbps)': '',
            '微信小文件发送平均时延(s)': '',
            '微信小文件发送平均速率(Mbps)': '',
            '商店大app平均时延(s)': '',
            '商店大app平均速率(Mbps)': '',
            '商店小app平均时延(s)': '',
            '商店小app平均速率(Mbps)': ''
        })

    df_summary_data_voice = pd.DataFrame(summary_data)

    # 输出文件名
    version = 'V2.5.17_CX'
    timestamp = datetime.now().strftime('%Y%m%d-%H%M')
    out = f'语音VQI输出_{version}_{timestamp}.xlsx'

    log("写入输出文件...", 70)
    log("  引擎初始化...", 71)
    _out_tmp_vqi = out + '.tmp'
    with pd.ExcelWriter(_out_tmp_vqi, engine='xlsxwriter') as w:
        workbook = w.book
        # 预定义格式
        hdr_fmt = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter',
                                        'bg_color': '#D9E1F2', 'border': 1, 'text_wrap': True})
        center_fmt = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
        green_fmt = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1,
                                          'bg_color': '#E2EFDA'})
        blue_fmt = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1,
                                         'bg_color': '#CCE5FF'})
        anno_fmt = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1,
                                         'num_format': '0.00'})

        # 写入统计sheet（小数据量，直接带格式写入）
        # 移除按运营商表的写入
        for sheet_name, df_data, colwidths in [
            ('语音指标统计', df_voice_out, 20),
            ('MOS3.5计算明细', df_vqi_out, 22),
            ('汇总-数据和语音', df_summary_data_voice, 18),
            ('按文件名', df_file_stats, 20),
        ]:
            if len(df_data) == 0:
                continue
            log(f"  写入{sheet_name}({len(df_data)}行x{len(df_data.columns)}列)...", 72)
            ws = workbook.add_worksheet(sheet_name)
            # 写表头
            for ci, col_name in enumerate(df_data.columns):
                ws.write(0, ci, str(col_name), hdr_fmt)
            # 用.values批量获取，然后用write_row加速
            data_array = df_data.values
            ncols = len(df_data.columns)
            for ri in range(len(df_data)):
                row_vals = data_array[ri]
                for ci in range(ncols):
                    val = row_vals[ci]
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        ws.write_blank(ri + 1, ci, None, center_fmt)
                    elif isinstance(val, (int, float)):
                        if sheet_name == 'MOS3.5计算明细' and ri + 1 == len(df_data):
                            ws.write(ri + 1, ci, val, green_fmt)
                        elif sheet_name == '汇总-数据和语音' and ri == 1:
                            ws.write(ri + 1, ci, val, blue_fmt)
                        else:
                            ws.write(ri + 1, ci, val, center_fmt)
                    else:
                        ws.write(ri + 1, ci, str(val), center_fmt)
            # 设置列宽
            ws.set_column(0, ncols - 1, colwidths)
            # 冻结窗格
            ws.freeze_panes(1, 0)
            log(f"  {sheet_name} 完成", 74)

            # 语音指标统计特殊处理：第5列空值加批注
            if sheet_name == '语音指标统计' and len(df_data.columns) >= 5:
                for ri in range(len(df_data)):
                    val = data_array[ri, 4]
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        ws.write_comment(ri + 1, 4, "无RSP_180", {'author': 'Tool'})

        # ===== 原始数据sheet（全量输出）=====
        if merge_raw and len(df_raw_out) > 0:
            log("  写入原始数据...", 75)
            log(f"  语音原始数据共{len(df_raw_out)}行，全部写入", 76)
            # 去掉临时列
            df_raw_write = df_raw_out.copy()
            for _c_drop in ['_t', '_time_only']:
                if _c_drop in df_raw_write.columns:
                    df_raw_write = df_raw_write.drop(_c_drop, axis=1)
            # 直接用pandas内置批量写入(比逐行write_row快100倍以上)
            log(f"  写入{len(df_raw_out)}行x{len(df_raw_out.columns)}列...", 77)
            df_raw_out.to_excel(w, sheet_name='语音业务-原始文件', index=False)
            ws_raw = w.sheets['语音业务-原始文件']
            ws_raw.freeze_panes(1, 0)
            log(f"  原始数据写入完成", 79)

        # ===== 批注（由add_annotations控制）=====
        if add_annotations:
            log("  添加批注...", 85)

            # === 性能优化：预构建过滤缓存，避免每个单元格都遍历全量records ===
            # 按运营商预过滤
            _op_voice_cache = {}
            _op_vqi_cache = {}
            _op_rsrp_cache = {}
            for op in set(r.get('运营商', '') for r in voice_records):
                _op_voice_cache[op] = [vr for vr in voice_records if vr.get('运营商') == op]
            for op in set(r.get('运营商', '') for r in vqi_records):
                _op_vqi_cache[op] = [r for r in vqi_records if r.get('运营商') == op]
            for op in set(r.get('运营商', '') for r in rsrp_records):
                _op_rsrp_cache[op] = [r for r in rsrp_records if r.get('运营商') == op]
            # 按文件名预过滤
            _fname_voice_cache = {}
            _fname_vqi_cache = {}
            _fname_rsrp_cache = {}
            for fname in set(r.get('_source_file', '') for r in voice_records):
                _fname_voice_cache[fname] = [vr for vr in voice_records if vr.get('_source_file') == fname]
            for fname in set(r.get('_source_file', '') for r in vqi_records):
                _fname_vqi_cache[fname] = [r for r in vqi_records if r.get('_source_file') == fname]
            for fname in set(r.get('_source_file', '') for r in rsrp_records):
                _fname_rsrp_cache[fname] = [r for r in rsrp_records if r.get('_source_file') == fname]

            for sheet_name in ['汇总-数据和语音', '按文件名']:
                try:
                    ws = workbook.get_worksheet_by_name(sheet_name)
                except:
                    continue
                log(f"  批注: {sheet_name}", 86)
                df_map = {
                    '汇总-数据和语音': df_summary_data_voice,
                    '按文件名': df_file_stats,
                }
                df_anno = df_map.get(sheet_name)
                if df_anno is None or len(df_anno) == 0:
                    continue
                anno_data = df_anno.values
                anno_cols = list(df_anno.columns)
                for ri in range(len(df_anno)):
                    for ci in range(1, len(anno_cols)):
                        val = anno_data[ri, ci]
                        if isinstance(val, (int, float)) and not (isinstance(val, float) and pd.isna(val)):
                            col_name = anno_cols[ci]
                            # 根据列名和行索引生成详细批注
                            if sheet_name == '汇总-数据和语音':
                                row_op = df_anno.iloc[ri, 0]  # 运营商列
                                txt = _build_anno_text(col_name, row_op, ri, df_voice_out, df_vqi,
                                    voice_records, vqi_records, rsrp_records, df_file_stats,
                                    _op_voice_cache, _op_vqi_cache, _op_rsrp_cache)
                            elif sheet_name == '按文件名':
                                fname = df_anno.iloc[ri, 0]  # 文件名列
                                txt = _build_file_anno_text(col_name, fname, voice_records, vqi_records, rsrp_records,
                                    _fname_voice_cache, _fname_vqi_cache, _fname_rsrp_cache)
                            else:
                                txt = f"公式: {col_name}\n计算过程: 从原始数据统计\n\n数据来源: {col_name} 列"
                            ws.write_comment(ri + 1, ci, txt, {'author': 'Tool', 'width': 550, 'height': 250})

        # ===== 数据明细Sheet（12个，从中间数据提炼）=====
        log("  写入数据明细sheet...", 80)
        # 预构建文件名到路径和trace_id到手机号的映射(用于明细sheet手机号列)
        voice_file_to_path = {fname: info['full_path'] for fname, info in voice_file_info.items()}
        data_file_to_path = {fname: info['full_path'] for fname, info in data_file_info.items()}
        all_file_to_path = {**voice_file_to_path, **data_file_to_path}
        trace_to_phone = {}
        if phone_trace_map:
            for entry in phone_trace_map:
                trace_to_phone[entry['trace_id']] = entry['phone']

        # 1. 数据_呼叫建立时长原始数据
        ws_call_setup = workbook.add_worksheet('数据_呼叫建立时长原始数据')
        hdr_call_setup = ['文件名', '运营商', '手机号', '轮次', '主被叫', 'INVITE时间', 'BYE时间', '开始时间', '结束时间', '呼叫建立时长(s)']
        for ci, col_name in enumerate(hdr_call_setup):
            ws_call_setup.write(0, ci, col_name, hdr_fmt)
        for ri, vr in enumerate(voice_records):
            sf = vr.get('_source_file', '')
            phone = trace_to_phone.get(_extract_trace_id(all_file_to_path.get(sf, '')), '')
            row_vals = [
                sf, vr.get('运营商', ''), phone,
                vr.get('轮次', ''), vr.get('主被叫', ''),
                vr.get('INVITE时间', ''), vr.get('BYE时间', ''),
                vr.get('开始时间', ''), vr.get('结束时间', ''),
                vr.get('呼叫建立时长(s)', '')
            ]
            for ci, val in enumerate(row_vals):
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    ws_call_setup.write_blank(ri + 1, ci, None, center_fmt)
                else:
                    ws_call_setup.write(ri + 1, ci, str(val), center_fmt)
        ws_call_setup.set_column(0, 0, 40)
        ws_call_setup.set_column(1, 2, 12)
        ws_call_setup.set_column(3, 9, 18)
        ws_call_setup.freeze_panes(1, 0)

        # 2. 数据_MOS原始数据
        ws_mos = workbook.add_worksheet('数据_MOS原始数据')
        hdr_mos = ['文件名', '运营商', '手机号', '轮次', '开始时间', '结束时间', 'DlE2eVqi', 'MOS值(=Vqi/100)', '是否优质(MOS>3.5)']
        for ci, col_name in enumerate(hdr_mos):
            ws_mos.write(0, ci, col_name, hdr_fmt)
        for ri, r in enumerate(vqi_records):
            e2e = r.get('DlE2eVqi', '')
            mos_val = round(e2e / 100, 2) if isinstance(e2e, (int, float)) and e2e != 65535 else ''
            quality = '是' if (isinstance(e2e, (int, float)) and e2e > 350) else '否'
            sf = r.get('_source_file', '')
            phone = trace_to_phone.get(_extract_trace_id(all_file_to_path.get(sf, '')), '')
            row_vals = [
                sf, r.get('运营商', ''), phone,
                r.get('轮次', ''), r.get('时间', ''),
                r.get('结束时间', ''), e2e, mos_val, quality
            ]
            for ci, val in enumerate(row_vals):
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    ws_mos.write_blank(ri + 1, ci, None, center_fmt)
                else:
                    ws_mos.write(ri + 1, ci, str(val), center_fmt)
        ws_mos.set_column(0, 0, 40)
        ws_mos.set_column(1, 2, 12)
        ws_mos.set_column(3, 8, 18)
        ws_mos.freeze_panes(1, 0)

        # 3. 数据_RSRP原始数据
        ws_rsrp = workbook.add_worksheet('数据_RSRP原始数据')
        hdr_rsrp = ['文件名', '运营商', '手机号', '时间', '服务小区RSRP(dBm)', '是否在通话中(是/否)']
        for ci, col_name in enumerate(hdr_rsrp):
            ws_rsrp.write(0, ci, col_name, hdr_fmt)
        for ri, r in enumerate(rsrp_records):
            sf = r.get('_source_file', '')
            phone = trace_to_phone.get(_extract_trace_id(all_file_to_path.get(sf, '')), '')
            row_vals = [
                sf, r.get('运营商', ''), phone,
                r.get('时间', ''), r['servRSRP值'],
                '是' if r.get('_in_call_period') else '否'
            ]
            for ci, val in enumerate(row_vals):
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    ws_rsrp.write_blank(ri + 1, ci, None, center_fmt)
                else:
                    ws_rsrp.write(ri + 1, ci, str(val), center_fmt)
        ws_rsrp.set_column(0, 0, 40)
        ws_rsrp.set_column(1, 2, 12)
        ws_rsrp.set_column(3, 5, 18)
        ws_rsrp.freeze_panes(1, 0)

        # 4. 数据_通话轮次原始数据
        ws_round = workbook.add_worksheet('数据_通话轮次原始数据')
        hdr_round = ['文件名', '运营商', '手机号', '轮次', '主被叫', 'INVITE时间', 'BYE时间', '开始时间', '结束时间', '是否完整通话']
        for ci, col_name in enumerate(hdr_round):
            ws_round.write(0, ci, col_name, hdr_fmt)
        for ri, vr in enumerate(voice_records):
            sf = vr.get('_source_file', '')
            phone = trace_to_phone.get(_extract_trace_id(all_file_to_path.get(sf, '')), '')
            row_vals = [
                sf, vr.get('运营商', ''), phone,
                vr.get('轮次', ''), vr.get('主被叫', ''),
                vr.get('INVITE时间', ''), vr.get('BYE时间', ''),
                vr.get('开始时间', ''), vr.get('结束时间', ''),
                vr.get('是否完整通话', '否')
            ]
            for ci, val in enumerate(row_vals):
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    ws_round.write_blank(ri + 1, ci, None, center_fmt)
                else:
                    ws_round.write(ri + 1, ci, str(val), center_fmt)
        ws_round.set_column(0, 0, 40)
        ws_round.set_column(1, 2, 12)
        ws_round.set_column(3, 9, 18)
        ws_round.freeze_panes(1, 0)

        # 5. 数据_优质通话占比原始数据
        ws_quality = workbook.add_worksheet('数据_优质通话占比原始数据')
        hdr_quality = ['文件名', '运营商', '手机号', '总通话数', '优质通话数', '优质通话占比(%)']
        for ci, col_name in enumerate(hdr_quality):
            ws_quality.write(0, ci, col_name, hdr_fmt)
        fname_quality = {}
        for r in vqi_records:
            fn = r.get('_source_file', '')
            e2e = r.get('DlE2eVqi', '')
            phone = trace_to_phone.get(_extract_trace_id(all_file_to_path.get(fn, '')), '')
            if fn not in fname_quality:
                fname_quality[fn] = {'total': 0, 'quality': 0, 'op': r.get('运营商', ''), 'phone': phone}
            fname_quality[fn]['total'] += 1
            if isinstance(e2e, (int, float)) and e2e > 350:
                fname_quality[fn]['quality'] += 1
        ri = 0
        for fn, qd in fname_quality.items():
            pct = round(qd['quality'] / qd['total'] * 100, 2) if qd['total'] > 0 else 0
            row_vals = [fn, qd['op'], qd['phone'], qd['total'], qd['quality'], pct]
            for ci, val in enumerate(row_vals):
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    ws_quality.write_blank(ri + 1, ci, None, center_fmt)
                else:
                    ws_quality.write(ri + 1, ci, str(val), center_fmt)
            ri += 1
        ws_quality.set_column(0, 0, 40)
        ws_quality.set_column(1, 2, 12)
        ws_quality.set_column(3, 5, 18)
        ws_quality.freeze_panes(1, 0)

        # 6. 数据_语音业务时长原始数据
        ws_voice_dur = workbook.add_worksheet('数据_语音业务时长原始数据')
        hdr_voice_dur = ['文件名', '运营商', '手机号', '轮次', '主被叫', '开始时间', '结束时间', '通话时长(s)']
        for ci, col_name in enumerate(hdr_voice_dur):
            ws_voice_dur.write(0, ci, col_name, hdr_fmt)
        for ri, vr in enumerate(voice_records):
            start = vr.get('开始时间', '')
            end = vr.get('结束时间', '')
            dur = ''
            if start and end:
                try:
                    dur = round((datetime.strptime(str(end), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(start), '%Y-%m-%d %H:%M:%S')).total_seconds(), 2)
                except:
                    pass
            sf = vr.get('_source_file', '')
            phone = trace_to_phone.get(_extract_trace_id(all_file_to_path.get(sf, '')), '')
            row_vals = [
                sf, vr.get('运营商', ''), phone,
                vr.get('轮次', ''), vr.get('主被叫', ''), start, end, dur
            ]
            for ci, val in enumerate(row_vals):
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    ws_voice_dur.write_blank(ri + 1, ci, None, center_fmt)
                else:
                    ws_voice_dur.write(ri + 1, ci, str(val), center_fmt)
        ws_voice_dur.set_column(0, 0, 40)
        ws_voice_dur.set_column(1, 2, 12)
        ws_voice_dur.set_column(3, 7, 18)
        ws_voice_dur.freeze_panes(1, 0)

        # ===== 数据业务6类(从data_dfs和data_file_info取) =====
        log("  写入数据业务明细sheet...", 81)
        if data_dfs:
            # 7. 数据_速率原始数据
            ws_rate = workbook.add_worksheet('数据_速率原始数据')
            hdr_rate = ['文件名', '运营商', '手机号', '轮次', '业务类型', '开始时间', '结束时间', '平均速率(Mbps)', '削峰平均速率(Mbps)', '前10%峰值速率']
            for ci, col_name in enumerate(hdr_rate):
                ws_rate.write(0, ci, col_name, hdr_fmt)
            ri = 0
            for fname, info in data_file_info.items():
                op = info.get('operator', '')
                phone = info.get('phone', '')
                row_vals = [fname, op, phone, '', '', '', '', info.get('dl_avg', ''), info.get('dl_clip_avg', ''), info.get('dl_peak', '')]
                for ci, val in enumerate(row_vals):
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        ws_rate.write_blank(ri + 1, ci, None, center_fmt)
                    else:
                        ws_rate.write(ri + 1, ci, str(val), center_fmt)
                ri += 1
            ws_rate.set_column(0, 0, 35)
            ws_rate.set_column(1, 2, 10)
            ws_rate.set_column(3, 9, 15)
            ws_rate.freeze_panes(1, 0)

            # 8. 数据_业务时长原始数据
            ws_biz_dur = workbook.add_worksheet('数据_业务时长原始数据')
            hdr_biz_dur = ['文件名', '运营商', '手机号', '轮次', '业务类型', '开始时间', '结束时间', '业务时长(s)']
            for ci, col_name in enumerate(hdr_biz_dur):
                ws_biz_dur.write(0, ci, col_name, hdr_fmt)
            ri = 0
            for fname, info in data_file_info.items():
                op = info.get('operator', '')
                phone = info.get('phone', '')
                row_vals = [fname, op, phone, '', '', '', '', '']
                for ci, val in enumerate(row_vals):
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        ws_biz_dur.write_blank(ri + 1, ci, None, center_fmt)
                    else:
                        ws_biz_dur.write(ri + 1, ci, str(val), center_fmt)
                ri += 1
            ws_biz_dur.set_column(0, 0, 35)
            ws_biz_dur.set_column(1, 2, 10)
            ws_biz_dur.set_column(3, 7, 15)
            ws_biz_dur.freeze_panes(1, 0)

            # 9. 数据_FTP下载原始数据
            ws_ftp_dl = workbook.add_worksheet('数据_FTP下载原始数据')
            hdr_ftp = ['文件名', '运营商', '手机号', '轮次', '开始时间', '结束时间', '平均速率(Mbps)', '前10%峰值', '最大速率', '最小速率', '采样数']
            for ci, col_name in enumerate(hdr_ftp):
                ws_ftp_dl.write(0, ci, col_name, hdr_fmt)
            ri = 0
            for fname, info in data_file_info.items():
                op = info.get('operator', '')
                phone = info.get('phone', '')
                row_vals = [fname, op, phone, '', '', '', info.get('dl_avg', ''), info.get('dl_peak', ''), '', '', '']
                for ci, val in enumerate(row_vals):
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        ws_ftp_dl.write_blank(ri + 1, ci, None, center_fmt)
                    else:
                        ws_ftp_dl.write(ri + 1, ci, str(val), center_fmt)
                ri += 1
            ws_ftp_dl.set_column(0, 0, 35)
            ws_ftp_dl.set_column(1, 2, 10)
            ws_ftp_dl.set_column(3, 10, 15)
            ws_ftp_dl.freeze_panes(1, 0)

            # 10. 数据_FTP上传原始数据
            ws_ftp_ul = workbook.add_worksheet('数据_FTP上传原始数据')
            for ci, col_name in enumerate(hdr_ftp):
                ws_ftp_ul.write(0, ci, col_name, hdr_fmt)
            ri = 0
            for fname, info in data_file_info.items():
                op = info.get('operator', '')
                phone = info.get('phone', '')
                row_vals = [fname, op, phone, '', '', '', info.get('ul_avg', ''), info.get('ul_peak', ''), '', '', '']
                for ci, val in enumerate(row_vals):
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        ws_ftp_ul.write_blank(ri + 1, ci, None, center_fmt)
                    else:
                        ws_ftp_ul.write(ri + 1, ci, str(val), center_fmt)
                ri += 1
            ws_ftp_ul.set_column(0, 0, 35)
            ws_ftp_ul.set_column(1, 2, 10)
            ws_ftp_ul.set_column(3, 10, 15)
            ws_ftp_ul.freeze_panes(1, 0)

            # 11. 数据_微信商店业务原始数据
            ws_wechat = workbook.add_worksheet('数据_微信商店业务原始数据')
            hdr_wechat = ['文件名', '运营商', '手机号', '轮次', '业务类型(微信小/微信大/商店小/商店大)', '方向(上下行)', '开始时间', '结束时间', '数据量(MB)', '速率(Mbps)', '时延(s)']
            for ci, col_name in enumerate(hdr_wechat):
                ws_wechat.write(0, ci, col_name, hdr_fmt)
            ws_wechat.set_column(0, 0, 35)
            ws_wechat.set_column(1, 2, 10)
            ws_wechat.set_column(3, 10, 25)
            ws_wechat.freeze_panes(1, 0)

            # 12. 数据_QCI统计原始数据
            ws_qci = workbook.add_worksheet('数据_QCI统计原始数据')
            hdr_qci = ['文件名', '运营商', '手机号', 'QCI值', '行数', '时间范围']
            for ci, col_name in enumerate(hdr_qci):
                ws_qci.write(0, ci, col_name, hdr_fmt)
            ri = 0
            for df_i in data_dfs:
                if 'QCI' in df_i.columns and len(df_i) > 0:
                    source_file = getattr(df_i, 'attrs', {}).get('source_file', '')
                    op = ''
                    trace_id = _extract_trace_id(source_file)
                    phone = _lookup_phone(trace_id, phone_trace_map)
                    for qci_val, qci_df in df_i.groupby('QCI'):
                        times = qci_df['_t'].dropna() if '_t' in qci_df.columns else []
                        ts = ''
                        if len(times) > 0:
                            ts = str(times.min().strftime('%Y-%m-%d %H:%M:%S')) + ' ~ ' + str(times.max().strftime('%Y-%m-%d %H:%M:%S'))
                        row_vals = [os.path.basename(source_file) if source_file else '', op, phone,
                                   int(qci_val) if not pd.isna(qci_val) else '', len(qci_df), ts]
                        for ci, val in enumerate(row_vals):
                            if val is None or (isinstance(val, float) and pd.isna(val)):
                                ws_qci.write_blank(ri + 1, ci, None, center_fmt)
                            else:
                                ws_qci.write(ri + 1, ci, str(val), center_fmt)
                        ri += 1
            ws_qci.set_column(0, 0, 35)
            ws_qci.set_column(1, 2, 10)
            ws_qci.set_column(3, 5, 18)
            ws_qci.freeze_panes(1, 0)
        else:
            for sheet_name in ['数据_速率原始数据', '数据_业务时长原始数据', '数据_FTP下载原始数据',
                              '数据_FTP上传原始数据', '数据_微信商店业务原始数据', '数据_QCI统计原始数据']:
                ws_empty = workbook.add_worksheet(sheet_name)
                if sheet_name == '数据_速率原始数据':
                    hdr = ['文件名', '运营商', '手机号', '轮次', '业务类型', '开始时间', '结束时间', '平均速率(Mbps)', '削峰平均速率(Mbps)', '前10%峰值速率']
                elif sheet_name == '数据_业务时长原始数据':
                    hdr = ['文件名', '运营商', '手机号', '轮次', '业务类型', '开始时间', '结束时间', '业务时长(s)']
                elif sheet_name in ['数据_FTP下载原始数据', '数据_FTP上传原始数据']:
                    hdr = ['文件名', '运营商', '手机号', '轮次', '开始时间', '结束时间', '平均速率(Mbps)', '前10%峰值', '最大速率', '最小速率', '采样数']
                elif sheet_name == '数据_微信商店业务原始数据':
                    hdr = ['文件名', '运营商', '手机号', '轮次', '业务类型(微信小/微信大/商店小/商店大)', '方向(上下行)', '开始时间', '结束时间', '数据量(MB)', '速率(Mbps)', '时延(s)']
                else:
                    hdr = ['文件名', '运营商', '手机号', 'QCI值', '行数', '时间范围']
                for ci, col_name in enumerate(hdr):
                    ws_empty.write(0, ci, col_name, hdr_fmt)
                ws_empty.set_column(0, 0, 35)
                ws_empty.set_column(1, 2, 10)
                ws_empty.set_column(3, len(hdr) - 1, 18)
                ws_empty.freeze_panes(1, 0)

        log("  数据明细sheet写入完成", 82)

    log(f"完成! {out}", 100)
    # 临时文件改名为正式输出(避免取消时留下半截坏文件)
    if os.path.exists(_out_tmp):
        os.replace(_out_tmp, out)
    # 清理解压临时目录
    for _d in _tmp_dirs:
        try: shutil.rmtree(_d)
        except Exception: pass
    return out


# ===== GUI 包装 =====
try:
    from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QPushButton, QLabel, QFileDialog, QMessageBox, QProgressBar, QTextEdit,
        QGroupBox, QCheckBox, QSpinBox, QDialog, QSplitter, QGridLayout,
        QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView, QTimeEdit, QComboBox)
    from PySide6.QtCore import Qt, QThread, Signal, Slot, QTime, QEvent, QTimer
    from PySide6.QtGui import QFont
    GUI_OK = True
except:
    GUI_OK = False


if GUI_OK:
    class WorkerThread(QThread):
        progress = Signal(str)
        progress_pct = Signal(int)
        done = Signal(str)

        def __init__(self, files, qci_list, dl_clip, ul_clip, base_file=None, time_filter=None, ftp_duration=25, merge_raw=True, add_annotations=True, phone_trace_map=None):
            super().__init__()
            self.files = files; self.qci_list = qci_list; self.dl_clip = dl_clip; self.ul_clip = ul_clip
            self.base_file = base_file; self.time_filter = time_filter; self.ftp_duration = ftp_duration
            self.merge_raw = merge_raw; self.add_annotations = add_annotations
            self.phone_trace_map = phone_trace_map or []
            self._cancel = False; self._pct = 0

        def cancel(self): self._cancel = True

        def run(self):
            try:
                out = process(self.files, self.qci_list, self.dl_clip, self.ul_clip,
                              callback=self.l, cancel_check=lambda: self._cancel,
                              progress_cb=lambda p: self.progress_pct.emit(int(p)),
                              base_file=self.base_file, time_filter=self.time_filter,
                              ftp_duration=self.ftp_duration, merge_raw=self.merge_raw,
                              add_annotations=self.add_annotations,
                              phone_trace_map=self.phone_trace_map)
                self.done.emit(out if out else '')
            except KeyboardInterrupt:
                self.done.emit('')

        def l(self, msg):
            self.progress.emit(msg)

    class VqiWorkerThread(QThread):
        progress = Signal(str)
        progress_pct = Signal(int)
        done = Signal(str)

        def __init__(self, files, time_filter=None, merge_raw=True, add_annotations=True, phone_trace_map=None):
            super().__init__()
            self.files = files
            self.time_filter = time_filter
            self._cancel = False
            self.merge_raw = merge_raw
            self.add_annotations = add_annotations
            self.phone_trace_map = phone_trace_map or []

        def cancel(self): self._cancel = True

        def run(self):
            try:
                out = process_vqi(self.files,
                                  callback=self.l, cancel_check=lambda: self._cancel,
                                  progress_cb=lambda p: self.progress_pct.emit(int(p)),
                                  time_filter=self.time_filter,
                                  merge_raw=self.merge_raw,
                                  add_annotations=self.add_annotations,
                                  phone_trace_map=self.phone_trace_map)
                self.done.emit(out if out else '')
            except KeyboardInterrupt:
                self.done.emit('')

        def l(self, msg):
            self.progress.emit(msg)

    class MainWindow(QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("用户级数据、语音跟踪统计工具 V2.5.17_CX")
            self.files = []; self.last_out = None
            self.base_file = None  # 用户可选的基准文件
            # VQI相关实例变量
            self.vqi_files = []
            self.vqi_base_file = None
            self.vqi_last_out = None
            # 手机号 ↔ Trace ID 映射(两个tab各自独立持久化)
            self.phone_trace_map_data = []  # 5G速率tab
            self.phone_trace_map_vqi = []   # 语音VQI tab
            scr = QApplication.primaryScreen()
            self.scr_w = scr.geometry().width() if scr else 1440
            sh = scr.geometry().height() if scr else 1080
            self.fs = max(11, int(round(11 * sh / 1080)))
            self.setMinimumSize(1200, 750)
            self._build()
            # 加载两个tab持久化的手机号映射
            self._load_phone_mapping('data')
            self._load_phone_mapping('vqi')
            self.showMaximized()

        def _font(self, big=0):
            return QFont("PingFang SC", self.fs + big, QFont.Bold if big >= 2 else QFont.Normal)

        # ===== 手机号 ↔ Trace ID 映射辅助方法 =====
        @staticmethod
        def _extract_trace_id(path):
            """从文件路径中提取用户跟踪ID, 如 '用户跟踪ID=553' → '553'"""
            import re
            m = re.search(r'用户跟踪ID=(\d+)', str(path))
            return m.group(1) if m else None

        def _lookup_phone(self, trace_id, phone_map=None):
            """查找trace_id对应的手机号(实例级备用,默认用data map)"""
            pm = phone_map if phone_map is not None else self.phone_trace_map_data
            if not trace_id or not pm:
                return ''
            for entry in pm:
                if entry['trace_id'] == trace_id:
                    return entry.get('phone', '')
            return ''

        def _phone_map_file(self, which):
            """返回映射json文件路径(config目录,与工具一起走)"""
            name = '手机号映射_数据业务.json' if which == 'data' else '手机号映射_语音VQI.json'
            return os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'config', name)

        def _save_phone_mapping(self, which):
            """保存对应tab的映射到json(失败不报错,不影响主功能)"""
            if which not in ('data', 'vqi'):
                return
            table = self.phone_table if which == 'data' else self.vqi_phone_table
            rows = []
            for row in range(table.rowCount()):
                phone_item = table.item(row, 0)
                tid_item = table.item(row, 1)
                phone = phone_item.text().strip() if phone_item else ''
                tid = tid_item.text().strip() if tid_item else ''
                if phone and tid:
                    rows.append({'phone': phone, 'trace_id': tid})
            try:
                path = self._phone_map_file(which)
                os.makedirs(os.path.dirname(path), exist_ok=True)
                with open(path, 'w', encoding='utf-8') as f:
                    json.dump(rows, f, ensure_ascii=False, indent=2)
            except Exception:
                pass

        def _load_phone_mapping(self, which):
            """启动时从json加载映射并填充到对应表格(文件损坏/缺失则空启动不报错)"""
            if which not in ('data', 'vqi'):
                return
            try:
                path = self._phone_map_file(which)
                if not os.path.exists(path):
                    return
                with open(path, 'r', encoding='utf-8') as f:
                    rows = json.load(f)
                table = self.phone_table if which == 'data' else self.vqi_phone_table
                pairs = [(r.get('phone', ''), r.get('trace_id', '')) for r in rows
                         if r.get('phone') and r.get('trace_id')]
                if pairs:
                    self._add_rows_to_table(pairs, table)
            except Exception:
                pass

        def _remove_phone_mapping_row(self, table_widget, btn, which=None):
            """删除按钮所在行并持久化(动态按按钮位置查行,避免删中间行后行号失效)"""
            row = table_widget.indexAt(btn.pos()).row()
            if row >= 0:
                table_widget.removeRow(row)
            if which:
                self._save_phone_mapping(which)

        def _build_phone_mapping_group(self, table_widget, text_edit, add_btn, clear_btn, which=None):
            """构建手机号映射UI组（内部调用，由两个tab各自传入控件引用）"""
            gp = QGroupBox("手机号-TraceID映射")
            vl = QVBoxLayout(gp)
            vl.setContentsMargins(4, 4, 4, 4)
            vl.setSpacing(2)

            # 说明标签
            vl.addWidget(QLabel("手动添加（一对一）："))
            # 表格：手机号 | Trace ID | 操作
            table_widget.setColumnCount(3)
            table_widget.setHorizontalHeaderLabels(["手机号", "Trace ID", "操作"])
            table_widget.horizontalHeader().setStretchLastSection(False)
            table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
            table_widget.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            table_widget.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)
            table_widget.setColumnWidth(2, 60)
            table_widget.setMaximumHeight(120)
            vl.addWidget(table_widget)

            # 添加行按钮
            add_btn.setText("添加映射")
            add_btn.clicked.connect(lambda: self._add_phone_mapping_row(table_widget, which))
            vl.addWidget(add_btn)

            # 分隔说明
            vl.addWidget(QLabel("或拖入/粘贴映射（Excel两列 / 文本每行: 手机号+TraceID, 支持空格/逗号/斜杠/Tab分隔, 列顺序自适应, 粘贴后自动填表）："))

            # 拖入框
            text_edit.setMaximumHeight(80)
            text_edit.setPlaceholderText("拖入Excel/Txt文件，或粘贴文本(支持空格/逗号/斜杠分隔),自动填表...")
            text_edit.setStyleSheet("QTextEdit{font-size:11px;background:#fff;border:1px solid #999;border-radius:3px;padding:2px;}")
            vl.addWidget(text_edit)

            # 清空按钮
            clear_btn.setText("清空映射")
            clear_btn.clicked.connect(lambda: self._clear_phone_mapping(table_widget, text_edit, which))
            vl.addWidget(clear_btn)
            # 表格内容变更(编辑/程序写入)时自动持久化
            table_widget.cellChanged.connect(lambda: self._save_phone_mapping(which) if which else None)
            # 文本框即时解析(防抖:停止输入300ms后解析填表,支持多分隔符)
            timer = QTimer()
            timer.setSingleShot(True)
            timer.setInterval(300)
            timer.timeout.connect(lambda: self._parse_textedit_into_table(text_edit, table_widget, which))
            text_edit.textChanged.connect(timer.start)
            text_edit._parse_timer = timer  # 持有引用防止GC

            return gp

        def _add_phone_mapping_row(self, table_widget, which=None):
            """在映射表格中添加一行空行"""
            row = table_widget.rowCount()
            table_widget.insertRow(row)
            table_widget.setItem(row, 0, QTableWidgetItem(""))
            table_widget.setItem(row, 1, QTableWidgetItem(""))
            del_btn = QPushButton("删除")
            del_btn.clicked.connect(lambda checked, b=del_btn: self._remove_phone_mapping_row(table_widget, b, which))
            table_widget.setCellWidget(row, 2, del_btn)

        def _clear_phone_mapping(self, table_widget, text_edit, which=None):
            """清空手机号映射（表格+文本框+内存+持久化文件）"""
            table_widget.setRowCount(0)
            text_edit.clear()
            if which == 'data':
                self.phone_trace_map_data = []
            elif which == 'vqi':
                self.phone_trace_map_vqi = []
            if which:
                self._save_phone_mapping(which)

        def _sync_phone_mapping(self, which=None):
            """从UI同步映射到内存(which='data'/'vqi'指定单tab;None=两个都同步)"""
            def _sync_one(table, te):
                result = []
                for row in range(table.rowCount()):
                    phone_item = table.item(row, 0)
                    tid_item = table.item(row, 1)
                    phone = phone_item.text().strip() if phone_item else ''
                    tid = tid_item.text().strip() if tid_item else ''
                    if phone and tid and not any(e['trace_id'] == tid for e in result):
                        result.append({'phone': phone, 'trace_id': tid})
                # 解析文本框(拖入/粘贴的文件或文本)写入表格
                txt = te.toPlainText().strip()
                if txt:
                    for line in txt.split('\n'):
                        line = line.strip()
                        if line and os.path.isfile(line):
                            self._parse_mapping_file(line, table)
                    pairs = self._parse_mapping_text_to_pairs(txt)
                    if pairs:
                        self._add_rows_to_table(pairs, table)
                    # 重新读取表格(文本框内容已写入)
                    for row in range(table.rowCount()):
                        phone_item = table.item(row, 0)
                        tid_item = table.item(row, 1)
                        phone = phone_item.text().strip() if phone_item else ''
                        tid = tid_item.text().strip() if tid_item else ''
                        if phone and tid and not any(e['trace_id'] == tid for e in result):
                            result.append({'phone': phone, 'trace_id': tid})
                return result
            if which == 'data' or which is None:
                self.phone_trace_map_data = _sync_one(self.phone_table, self.phone_textedit)
                self._save_phone_mapping('data')
            if which == 'vqi' or which is None:
                self.phone_trace_map_vqi = _sync_one(self.vqi_phone_table, self.vqi_phone_textedit)
                self._save_phone_mapping('vqi')

        def _parse_mapping_file(self, file_path, table_widget):
            """解析映射文件（Excel两列 / Txt两列Tab分隔），列顺序自适应"""
            low = file_path.lower()
            rows = []
            if low.endswith(('.xlsx', '.xls')):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active
                    data_rows = list(ws.iter_rows(values_only=True))
                    if not data_rows:
                        return
                    rows = self._detect_columns_and_parse(data_rows)
                except Exception as e:
                    return
            else:
                # txt / csv / 其它纯文本
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        raw_lines = f.readlines()
                    data_rows = []
                    for line in raw_lines:
                        line = line.strip()
                        if line:
                            parts = re.split(r'[\t,，;；/\s\\]+', line)
                            data_rows.append([p for p in parts if p])
                    if data_rows:
                        rows = self._detect_columns_and_parse(data_rows)
                except Exception as e:
                    return
            self._add_rows_to_table(rows, table_widget)

        @staticmethod
        def _order_pair(a, b):
            """两值配对,11位数字当手机号放前(用户:13271860243是手机号,562是trace id)"""
            if len(a) == 11 and a.isdigit():
                return (a, b)
            if len(b) == 11 and b.isdigit():
                return (b, a)
            return (a, b)

        def _parse_mapping_text_to_pairs(self, txt):
            """解析多行映射文本为(phone,trace_id)对。
            支持两种格式:①同行多列(空格/逗号/斜杠/Tab分隔) ②换行交替(手机号一行、ID一行)。可混合。"""
            pairs = []
            pending = None
            for raw_line in txt.split('\n'):
                line = raw_line.strip()
                if not line or os.path.isfile(line):
                    continue  # 空行/文件路径跳过(文件由调用方处理)
                parts = re.split(r'[\t,，;；/\s\\]+', line)
                parts = [p for p in parts if p]
                if len(parts) >= 2:
                    if pending:  # 有遗留单值,先与本行首值配对
                        pairs.append(self._order_pair(pending, parts[0]))
                        pending = None
                        rest = parts[1:]
                    else:
                        rest = parts
                    for i in range(0, len(rest) - 1, 2):
                        pairs.append(self._order_pair(rest[i], rest[i + 1]))
                elif len(parts) == 1:
                    if pending is None:
                        pending = parts[0]
                    else:
                        pairs.append(self._order_pair(pending, parts[0]))
                        pending = None
            return [(p, t) for p, t in pairs if p and t]

        def _parse_mapping_line(self, line, table_widget):
            """解析单行文本(Tab/空格/逗号/分号/斜杠等分隔符,两列:手机号+TraceID)"""
            parts = re.split(r'[\t,，;；/\s\\]+', line.strip())
            parts = [p for p in parts if p]
            if len(parts) >= 2:
                phone, tid = self._order_pair(parts[0], parts[1])
                if phone and tid:
                    self._add_rows_to_table([(phone, tid)], table_widget)

        def _parse_textedit_into_table(self, text_edit, table_widget, which=None):
            """文本框内容变化时即时解析并增量填表(同行多列+换行交替都支持,去重,填完自动保存)"""
            txt = text_edit.toPlainText().strip()
            if not txt:
                return
            for line in txt.split('\n'):  # 文件路径行单独处理(拖入的文件)
                line = line.strip()
                if line and os.path.isfile(line):
                    self._parse_mapping_file(line, table_widget)
            pairs = self._parse_mapping_text_to_pairs(txt)
            if pairs:
                self._add_rows_to_table(pairs, table_widget)
            if which:
                self._save_phone_mapping(which)

        @staticmethod
        def _detect_columns_and_parse(data_rows):
            """自动检测列顺序：手机号（11位数字）vs TraceID（纯数字非11位）"""
            if not data_rows:
                return []
            header = data_rows[0]
            # 尝试识别表头
            col_idx = [0, 1]  # 默认第一列手机号，第二列TraceID
            if len(header) >= 2:
                h0 = str(header[0]).strip().lower()
                h1 = str(header[1]).strip().lower()
                # 如果表头包含关键词
                phone_keywords = ['手机', 'phone', 'mobile', '号码', 'tel']
                trace_keywords = ['trace', '跟踪', 'id', '用户跟踪']
                h0_is_phone = any(k in h0 for k in phone_keywords)
                h1_is_phone = any(k in h1 for k in phone_keywords)
                h0_is_trace = any(k in h0 for k in trace_keywords)
                h1_is_trace = any(k in h1 for k in trace_keywords)
                if h0_is_phone and h1_is_trace:
                    col_idx = [0, 1]
                elif h0_is_trace and h1_is_phone:
                    col_idx = [1, 0]
                else:
                    # 无关键词识别，通过内容判断
                    col_idx = [0, 1]
                    # 检查第一行数据
                    if len(data_rows) > 1:
                        v0 = str(data_rows[1][0]).strip() if len(data_rows[1]) > 0 else ''
                        v1 = str(data_rows[1][1]).strip() if len(data_rows[1]) > 1 else ''
                        if v0.isdigit() and len(v0) == 11 and v1.isdigit() and len(v1) != 11:
                            col_idx = [0, 1]  # 第一列是手机号
                        elif v0.isdigit() and len(v0) != 11 and v1.isdigit() and len(v1) == 11:
                            col_idx = [1, 0]  # 第二列是手机号
            # 解析数据行（跳过表头）
            results = []
            start_row = 1 if len(data_rows) > 1 else 0
            for row in data_rows[start_row:]:
                if len(row) >= 2:
                    phone = str(row[col_idx[0]]).strip() if len(row) > col_idx[0] else ''
                    tid = str(row[col_idx[1]]).strip() if len(row) > col_idx[1] else ''
                    if phone and tid:
                        results.append((phone, tid))
            return results

        def _add_rows_to_table(self, rows, table_widget):
            """将解析结果添加到表格（去重）；批量写入时屏蔽cellChanged避免频繁保存"""
            which = 'data' if table_widget is self.phone_table else 'vqi'
            table_widget.blockSignals(True)
            try:
                existing = set()
                for row in range(table_widget.rowCount()):
                    phone_item = table_widget.item(row, 0)
                    tid_item = table_widget.item(row, 1)
                    if phone_item and tid_item:
                        existing.add((phone_item.text().strip(), tid_item.text().strip()))
                for phone, tid in rows:
                    if (phone, tid) not in existing:
                        row = table_widget.rowCount()
                        table_widget.insertRow(row)
                        table_widget.setItem(row, 0, QTableWidgetItem(phone))
                        table_widget.setItem(row, 1, QTableWidgetItem(tid))
                        del_btn = QPushButton("删除")
                        del_btn.clicked.connect(lambda checked, b=del_btn: self._remove_phone_mapping_row(table_widget, b, which))
                        table_widget.setCellWidget(row, 2, del_btn)
                        existing.add((phone, tid))
            finally:
                table_widget.blockSignals(False)

        def eventFilter(self, obj, event):
            # 让QTabWidget和QSplitter上的拖放事件也传递到MainWindow
            if event.type() == QEvent.DragEnter:
                self.dragEnterEvent(event)
                return True
            elif event.type() == QEvent.Drop:
                self.dropEvent(event)
                return True
            return super().eventFilter(obj, event)

        def dragEnterEvent(self, event):
            # 始终接受拖入事件，让MainWindow统一处理
            if event.mimeData().hasUrls():
                event.acceptProposedAction()
            else:
                event.ignore()

        def dropEvent(self, event):
            new = []
            for u in event.mimeData().urls():
                f = u.toLocalFile()
                if os.path.isdir(f):
                    for root, dirs, files in os.walk(f):
                        for fn in files:
                            low = fn.lower()
                            if low.endswith(('.xlsx', '.xls', '.zip', '.csv', '.mmf', '.tmf')):
                                new.append(os.path.join(root, fn))
                else:
                    low = f.lower()
                    if low.endswith(('.xlsx', '.xls', '.zip', '.csv', '.mmf', '.tmf')):
                        new.append(f)
            if new:
                # 检查当前激活的是哪个标签
                current_tab = self.centralWidget().currentIndex()
                if current_tab == 1:  # VQI标签
                    self.vqi_files = list(dict.fromkeys((self.vqi_files or []) + new))
                    display_lines = [os.path.basename(f) for f in self.vqi_files]
                    self.vqi_te.setText("\n".join(display_lines))
                    self.vqi_br.setEnabled(True)
                    if hasattr(self, 'vqi_lt'):
                        self.vqi_lt.append(f"拖入 {len(new)} 个文件到VQI，当前共 {len(self.vqi_files)} 个")
                else:  # 速率统计标签
                    self.files = list(dict.fromkeys((self.files or []) + new))
                    display_lines = [os.path.basename(f) for f in self.files]
                    self.te.setText("\n".join(display_lines))
                    self.br.setEnabled(True)
                    self.lt.append(f"拖入 {len(new)} 个文件，当前共 {len(self.files)} 个")
            event.acceptProposedAction()

        # 解决拖放不起作用的问题：用事件过滤器让中央控件转发拖放事件
        def _setup_phone_textedit_drop(self, text_edit):
            """为手机号映射QTextEdit设置独立的拖放处理"""
            text_edit.setAcceptDrops(True)

            def drag_enter(e):
                if e.mimeData().hasUrls():
                    for u in e.mimeData().urls():
                        f = u.toLocalFile()
                        low = f.lower()
                        if low.endswith(('.xlsx', '.xls', '.txt', '.csv')):
                            e.acceptProposedAction()
                            return
                e.ignore()

            def drop(e):
                paths = []
                for u in e.mimeData().urls():
                    f = u.toLocalFile()
                    low = f.lower()
                    if low.endswith(('.xlsx', '.xls', '.txt', '.csv')):
                        paths.append(f)
                if paths:
                    text_edit.setText("\n".join(paths))
                    e.acceptProposedAction()
                else:
                    e.ignore()

            text_edit.dragEnterEvent = drag_enter
            text_edit.dropEvent = drop

        def disable_child_drops(self, widget):
            for child in widget.findChildren(QWidget):
                child.setAcceptDrops(False)
                child.installEventFilter(self)
                self.disable_child_drops(child)

        def _build(self):
            self.setAcceptDrops(True)
            tabs = QTabWidget()
            tabs.addTab(self._build_rate_tab(), "5G用户级公共监控速率统计工具")
            tabs.addTab(self._build_vqi_tab(), "5G虚拟用户跟踪语音VQI工具")
            self.setCentralWidget(tabs)
            # 禁用所有子控件的拖放
            self.disable_child_drops(self)

        def _build_rate_tab(self):
            sp = QSplitter(Qt.Horizontal)
            # ===== 左：上设置 + 下运行 =====
            left = QWidget(); ll = QVBoxLayout(left)
            ll.setContentsMargins(4, 4, 4, 4)
            ll.setSpacing(4)
            title = QLabel("5G用户级公共监控速率统计工具 V2.5.17_CX")
            title.setFont(self._font(2)); title.setAlignment(Qt.AlignCenter); ll.addWidget(title)
            # 输入文件（选择/清空紧挨着，拖入框明显、高度翻倍）
            gf = QGroupBox("输入文件"); gfl = QVBoxLayout(gf)
            gfl.setContentsMargins(4, 4, 4, 4); gfl.setSpacing(2)
            # 选择/清空按钮紧挨着（不加stretch）
            bf = QHBoxLayout(); bf.setSpacing(4)
            self.bs = QPushButton("选择MMF文件"); self.bs.clicked.connect(self.sf)
            self.bc = QPushButton("清空"); self.bc.clicked.connect(self.cf)
            bf.addWidget(self.bs); bf.addWidget(self.bc); gfl.addLayout(bf)
            # 拖入框：只显示文件名，高度翻倍(300)，边框明显
            self.te = QTextEdit(); self.te.setReadOnly(True); self.te.setMaximumHeight(300)
            self.te.setPlaceholderText("可以将文件或是文件夹拖入，支持各类文件类型")
            self.te.setStyleSheet("QTextEdit{font-size:12px;background:#fff;border:2px solid #666;border-radius:4px;padding:4px;}")
            gfl.addWidget(self.te); ll.addWidget(gf)
            # 参数设置（紧凑，紧接输入文件）
            gp = QGroupBox("参数设置"); gpl = QGridLayout(gp)
            gpl.setContentsMargins(4, 4, 4, 4); gpl.setSpacing(2)
            gpl.addWidget(QLabel("QCI:"), 0, 0)
            self.cb6 = QCheckBox("6"); self.cb7 = QCheckBox("7")
            self.cb6.setChecked(True); self.cb7.setChecked(True)  # 默认6和7都要
            # QCI 6和7独立可选，不再互斥
            gpl.addWidget(self.cb6, 0, 1); gpl.addWidget(self.cb7, 0, 2)
            gpl.addWidget(QLabel("下行削峰(Mbps):"), 1, 0)
            self.sd = QSpinBox(); self.sd.setRange(100, 5000); self.sd.setValue(1000); gpl.addWidget(self.sd, 1, 1, 1, 3)
            gpl.addWidget(QLabel("上行削峰(Mbps):"), 2, 0)
            self.su = QSpinBox(); self.su.setRange(50, 1000); self.su.setValue(200); gpl.addWidget(self.su, 2, 1, 1, 3)
            # 时间段限制（第3行）
            self.cb_time = QCheckBox(); gpl.addWidget(self.cb_time, 3, 0)
            gpl.addWidget(QLabel("开始时间:"), 3, 1)
            self.time_start = QTimeEdit(); self.time_start.setDisplayFormat("HH:mm:ss"); self.time_start.setTime(QTime(0, 0, 0))
            gpl.addWidget(self.time_start, 3, 2)
            gpl.addWidget(QLabel("结束时间:"), 4, 1)
            self.time_end = QTimeEdit(); self.time_end.setDisplayFormat("HH:mm:ss"); self.time_end.setTime(QTime(23, 59, 59))
            gpl.addWidget(self.time_end, 4, 2)
            # FTP时长（第5行，可调5-120秒）
            gpl.addWidget(QLabel("FTP时长(秒):"), 5, 0)
            self.ftp_duration_spin = QSpinBox()
            self.ftp_duration_spin.setRange(5, 120)
            self.ftp_duration_spin.setValue(25)
            self.ftp_duration_spin.setSuffix(" 秒")
            self.cb_merge_raw = QCheckBox("合并原始数据到结果文件")
            self.cb_merge_raw.setChecked(False)
            gpl.addWidget(self.cb_merge_raw, 6, 0, 1, 3)
            self.cb_annotations = QCheckBox("添加批注到结果文件")
            self.cb_annotations.setChecked(True)
            gpl.addWidget(self.cb_annotations, 7, 0, 1, 3)
            gpl.addWidget(self.ftp_duration_spin, 5, 1, 1, 2)
            ll.addWidget(gp)
            # 手机号-TraceID映射（两个tab各自独立持久化）
            self.phone_table = QTableWidget()
            self.phone_textedit = QTextEdit()
            self.phone_add_btn = QPushButton()
            self.phone_clear_btn = QPushButton()
            gp_phone = self._build_phone_mapping_group(self.phone_table, self.phone_textedit, self.phone_add_btn, self.phone_clear_btn, 'data')
            ll.addWidget(gp_phone)
            self._setup_phone_textedit_drop(self.phone_textedit)
            # 运行（与参数设置紧接）
            gh = QGroupBox("运行"); ghl = QHBoxLayout(gh)
            ghl.setContentsMargins(4, 4, 4, 4); ghl.setSpacing(4)
            self.br = QPushButton("开始处理"); self.br.clicked.connect(self.run); self.br.setEnabled(False)
            self.bcancel = QPushButton("取消"); self.bcancel.clicked.connect(self.cancel); self.bcancel.setEnabled(False)
            self.babout = QPushButton("关于"); self.babout.clicked.connect(self.about)
            ghl.addWidget(self.br); ghl.addWidget(self.bcancel); ghl.addWidget(self.babout); ll.addWidget(gh)
            # 运行结果（打开按钮右下角，日志左下方剩余空间）
            gr = QGroupBox("运行结果"); grl = QVBoxLayout(gr)
            grl.setContentsMargins(4, 4, 4, 4); grl.setSpacing(2)
            # 进度条 + 打开按钮同行（按钮放右下）
            pb_open = QHBoxLayout()
            self.pb = QProgressBar(); pb_open.addWidget(self.pb)
            self.bopen = QPushButton("打开输出文件"); self.bopen.clicked.connect(self.open_out); self.bopen.setEnabled(False)
            pb_open.addWidget(self.bopen); grl.addLayout(pb_open)
            # 日志占剩余空间
            self.lt = QTextEdit(); self.lt.setReadOnly(True); self.lt.setStyleSheet("QTextEdit{font-size:11px;}")
            grl.addWidget(self.lt, 1); ll.addWidget(gr)
            left.setFont(self._font())
            # ===== 右：结果内嵌查看(子标签+表格) =====
            right = QWidget(); rl = QVBoxLayout(right)
            rh = QLabel("结果数据（内嵌查看，默认筛选A=1）"); rh.setFont(self._font(1)); rl.addWidget(rh)
            self.result_tabs = QTabWidget()
            self.tbl_detail = QTableWidget(); self.tbl_detail.setSelectionBehavior(QTableWidget.SelectItems)
            self.tbl_summary = QTableWidget(); self.tbl_compare = QTableWidget()
            self.result_tabs.addTab(self.tbl_detail, "数据业务-详细过程")
            self.result_tabs.addTab(self.tbl_summary, "汇总")
            self.result_tabs.addTab(self.tbl_compare, "数据业务-对比")
            rl.addWidget(self.result_tabs)
            right.setFont(self._font())
            sp.addWidget(left); sp.addWidget(right)
            sp.setStretchFactor(0, 1); sp.setStretchFactor(1, 5)  # 左1/6，右5/6
            sp.setSizes([int(self.scr_w * 0.167), int(self.scr_w * 0.833)])
            return sp

        def _build_vqi_tab(self):
            sp = QSplitter(Qt.Horizontal)
            # ===== 左：上设置 + 下运行 =====
            left = QWidget(); ll = QVBoxLayout(left)
            ll.setContentsMargins(4, 4, 4, 4)
            ll.setSpacing(4)
            title = QLabel("5G虚拟用户跟踪语音VQI工具 V2.5.17_CX")
            title.setFont(self._font(2)); title.setAlignment(Qt.AlignCenter); ll.addWidget(title)
            # 输入文件（选择/清空紧挨着，拖入框明显、高度翻倍）
            gf = QGroupBox("输入文件"); gfl = QVBoxLayout(gf)
            gfl.setContentsMargins(4, 4, 4, 4); gfl.setSpacing(2)
            # 选择/清空按钮紧挨着（不加stretch）
            bf = QHBoxLayout(); bf.setSpacing(4)
            self.vqi_bs = QPushButton("选择MMF文件"); self.vqi_bs.clicked.connect(self.vqi_sf)
            self.vqi_bc = QPushButton("清空"); self.vqi_bc.clicked.connect(self.vqi_cf)
            bf.addWidget(self.vqi_bs); bf.addWidget(self.vqi_bc); gfl.addLayout(bf)
            # 拖入框：只显示文件名，高度翻倍(300)，边框明显
            self.vqi_te = QTextEdit(); self.vqi_te.setReadOnly(True); self.vqi_te.setMaximumHeight(300)
            self.vqi_te.setPlaceholderText("可以将文件或是文件夹拖入，支持各类文件类型")
            self.vqi_te.setStyleSheet("QTextEdit{font-size:12px;background:#fff;border:2px solid #666;border-radius:4px;padding:4px;}")
            gfl.addWidget(self.vqi_te); ll.addWidget(gf)
            # 参数设置（紧凑，紧接输入文件）
            gp = QGroupBox("参数设置"); gpl = QGridLayout(gp)
            gpl.setContentsMargins(4, 4, 4, 4); gpl.setSpacing(2)
            # 时间段限制
            self.vqi_cb_time = QCheckBox(); gpl.addWidget(self.vqi_cb_time, 0, 0)
            gpl.addWidget(QLabel("开始时间:"), 0, 1)
            self.vqi_time_start = QTimeEdit(); self.vqi_time_start.setDisplayFormat("HH:mm:ss"); self.vqi_time_start.setTime(QTime(0, 0, 0))
            gpl.addWidget(self.vqi_time_start, 0, 2)
            gpl.addWidget(QLabel("结束时间:"), 1, 1)
            self.vqi_time_end = QTimeEdit(); self.vqi_time_end.setDisplayFormat("HH:mm:ss"); self.vqi_time_end.setTime(QTime(23, 59, 59))
            gpl.addWidget(self.vqi_time_end, 1, 2)

            self.vqi_cb_merge_raw = QCheckBox("合并原始数据到结果文件")
            self.vqi_cb_merge_raw.setChecked(False)
            gpl.addWidget(self.vqi_cb_merge_raw, 2, 0, 1, 3)
            self.vqi_cb_annotations = QCheckBox("添加批注到结果文件")
            self.vqi_cb_annotations.setChecked(True)
            gpl.addWidget(self.vqi_cb_annotations, 3, 0, 1, 3)
            ll.addWidget(gp)
            # 手机号-TraceID映射（两个tab各自独立持久化）
            self.vqi_phone_table = QTableWidget()
            self.vqi_phone_textedit = QTextEdit()
            self.vqi_phone_add_btn = QPushButton()
            self.vqi_phone_clear_btn = QPushButton()
            gp_vqi_phone = self._build_phone_mapping_group(self.vqi_phone_table, self.vqi_phone_textedit, self.vqi_phone_add_btn, self.vqi_phone_clear_btn, 'vqi')
            ll.addWidget(gp_vqi_phone)
            self._setup_phone_textedit_drop(self.vqi_phone_textedit)
            # 运行（与参数设置紧接）
            gh = QGroupBox("运行"); ghl = QHBoxLayout(gh)
            ghl.setContentsMargins(4, 4, 4, 4); ghl.setSpacing(4)
            self.vqi_br = QPushButton("开始处理"); self.vqi_br.clicked.connect(self.vqi_run); self.vqi_br.setEnabled(False)
            self.vqi_bcancel = QPushButton("取消"); self.vqi_bcancel.clicked.connect(self.vqi_cancel); self.vqi_bcancel.setEnabled(False)
            self.vqi_babout = QPushButton("关于"); self.vqi_babout.clicked.connect(self.about)
            ghl.addWidget(self.vqi_br); ghl.addWidget(self.vqi_bcancel); ghl.addWidget(self.vqi_babout); ll.addWidget(gh)
            # 运行结果（打开按钮右下角，日志左下方剩余空间）
            gr = QGroupBox("运行结果"); grl = QVBoxLayout(gr)
            grl.setContentsMargins(4, 4, 4, 4); grl.setSpacing(2)
            # 进度条 + 打开按钮同行（按钮放右下）
            pb_open = QHBoxLayout()
            self.vqi_pb = QProgressBar(); pb_open.addWidget(self.vqi_pb)
            self.vqi_bopen = QPushButton("打开输出文件"); self.vqi_bopen.clicked.connect(self.vqi_open_out); self.vqi_bopen.setEnabled(False)
            pb_open.addWidget(self.vqi_bopen); grl.addLayout(pb_open)
            # 日志占剩余空间
            self.vqi_lt = QTextEdit(); self.vqi_lt.setReadOnly(True); self.vqi_lt.setStyleSheet("QTextEdit{font-size:11px;}")
            grl.addWidget(self.vqi_lt, 1); ll.addWidget(gr)
            left.setFont(self._font())
            # ===== 右：结果内嵌查看(子标签+表格) =====
            right = QWidget(); rl = QVBoxLayout(right)
            rh = QLabel("结果数据（内嵌查看）"); rh.setFont(self._font(1)); rl.addWidget(rh)
            self.vqi_result_tabs = QTabWidget()
            self.vqi_tbl_raw = QTableWidget(); self.vqi_tbl_raw.setSelectionBehavior(QTableWidget.SelectItems)
            self.vqi_tbl_voice = QTableWidget(); self.vqi_tbl_voice.setSelectionBehavior(QTableWidget.SelectItems)
            self.vqi_tbl_mos = QTableWidget(); self.vqi_tbl_mos.setSelectionBehavior(QTableWidget.SelectItems)
            self.vqi_result_tabs.addTab(self.vqi_tbl_raw, "原始数据")
            self.vqi_result_tabs.addTab(self.vqi_tbl_voice, "语音指标统计")
            self.vqi_result_tabs.addTab(self.vqi_tbl_mos, "MOS3.5计算明细")
            self.vqi_tbl_data_raw = QTableWidget()
            self.vqi_tbl_data_raw.setSelectionBehavior(QTableWidget.SelectItems)
            self.vqi_result_tabs.addTab(self.vqi_tbl_data_raw, "数据业务-原始文件")
            rl.addWidget(self.vqi_result_tabs)
            right.setFont(self._font())
            sp.addWidget(left); sp.addWidget(right)
            sp.setStretchFactor(0, 1); sp.setStretchFactor(1, 5)  # 左1/6，右5/6
            sp.setSizes([int(self.scr_w * 0.167), int(self.scr_w * 0.833)])
            return sp

        # ===== VQI相关方法 =====
        def vqi_sf(self):
            fs, _ = QFileDialog.getOpenFileNames(self, "选择文件", "", "Excel/CSV/压缩 (*.xlsx *.xls *.csv *.zip);;All (*)")
            if fs:
                self.vqi_files = list(dict.fromkeys((self.vqi_files or []) + fs))
                self.vqi_te.setText("\n".join(os.path.basename(f) for f in self.vqi_files))
                self.vqi_br.setEnabled(True)

        def vqi_cf(self):
            self.vqi_files = []
            self.vqi_te.clear()
            self.vqi_br.setEnabled(False)

        def vqi_run(self):
            if not self.vqi_files:
                QMessageBox.warning(self, "提示", "请先选择MMF文件")
                return
            # 同步并检查手机号映射(为空时提醒)
            self._sync_phone_mapping('vqi')
            if not self.phone_trace_map_vqi:
                ret = QMessageBox.question(self, "手机号映射为空",
                    "当前未配置任何手机号-TraceID映射，结果中将无法显示手机号。\n是否继续处理？",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No)
                if ret != QMessageBox.StandardButton.Yes:
                    return
            # 清空结果数据显示区
            self.vqi_tbl_raw.setRowCount(0); self.vqi_tbl_raw.setColumnCount(0)
            self.vqi_tbl_voice.setRowCount(0); self.vqi_tbl_voice.setColumnCount(0)
            self.vqi_tbl_mos.setRowCount(0); self.vqi_tbl_mos.setColumnCount(0)
            self.vqi_tbl_data_raw.setRowCount(0); self.vqi_tbl_data_raw.setColumnCount(0)
            self.vqi_lt.clear()

            # 时间段过滤
            time_filter = None
            if self.vqi_cb_time.isChecked():
                t_start = self.vqi_time_start.time()
                t_end = self.vqi_time_end.time()
                time_filter = (True, t_start, t_end)

            self.vqi_br.setEnabled(False); self.vqi_bcancel.setEnabled(True)
            self.vqi_pb.setValue(0)
            self.vqi_lt.append(f"开始处理...")
            if time_filter:
                self.vqi_lt.append(f"时间段过滤: {t_start.toString('HH:mm:ss')} ~ {t_end.toString('HH:mm:ss')}")
            self.vqi_thread = VqiWorkerThread(self.vqi_files, time_filter=time_filter, merge_raw=self.vqi_cb_merge_raw.isChecked(), add_annotations=self.vqi_cb_annotations.isChecked(), phone_trace_map=self.phone_trace_map_vqi)
            self.vqi_thread.progress.connect(self.vqi_upd)
            self.vqi_thread.progress_pct.connect(lambda p: self.vqi_pb.setValue(p))
            self.vqi_thread.done.connect(self.vqi_dn)
            self.vqi_thread.start()

        def vqi_cancel(self):
            if hasattr(self, 'vqi_thread') and self.vqi_thread.isRunning():
                self.vqi_thread.cancel()
                self.vqi_lt.append("已请求取消，将在当前阶段结束后停止...")

        @Slot(str)
        def vqi_upd(self, msg):
            self.vqi_lt.append(msg)

        @Slot(str)
        def vqi_dn(self, path):
            self.vqi_bcancel.setEnabled(False)
            if path:
                self.vqi_pb.setValue(100)
                self.vqi_last_out = path
                self.vqi_bopen.setEnabled(True)
                self.vqi_lt.append(f"完成! {path}")
                self.show_vqi_result(path)
                # 自动打开结果文件
                os.system(f'open "{path}"')
            else:
                self.vqi_pb.setValue(0)
                self.vqi_lt.append("已取消。")
            self.vqi_br.setEnabled(True)

        def vqi_open_out(self):
            if self.vqi_last_out:
                os.system(f'open "{self.vqi_last_out}"')

        def show_vqi_result(self, path):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, data_only=True)
                sheet_list = wb.sheetnames
                # 显示语音原始数据sheet（如果存在）
                if '语音业务-原始文件' in sheet_list:
                    self._fill_vqi_table(self.vqi_tbl_raw, wb['语音业务-原始文件'], path=path, sheet='语音业务-原始文件')
                else:
                    self.vqi_tbl_raw.setRowCount(0); self.vqi_tbl_raw.setColumnCount(0)
                # 显示语音指标统计
                if '语音指标统计' in sheet_list:
                    self._fill_vqi_table(self.vqi_tbl_voice, wb['语音指标统计'], path=path, sheet='语音指标统计')
                else:
                    self.vqi_tbl_voice.setRowCount(0); self.vqi_tbl_voice.setColumnCount(0)
                # 显示MOS3.5计算明细
                if 'MOS3.5计算明细' in sheet_list:
                    self._fill_vqi_table(self.vqi_tbl_mos, wb['MOS3.5计算明细'], path=path, sheet='MOS3.5计算明细')
                elif '汇总-数据和语音' in sheet_list:
                    self._fill_vqi_table(self.vqi_tbl_mos, wb['汇总-数据和语音'], path=path, sheet='汇总-数据和语音')
                else:
                    self.vqi_tbl_mos.setRowCount(0); self.vqi_tbl_mos.setColumnCount(0)
                cnt = len(sheet_list)
                self.vqi_lt.append(f"结果已内嵌显示（{cnt}个Sheet）")
            except Exception as e:
                self.vqi_lt.append(f"内嵌显示失败: {e}")

        def _fill_vqi_table(self, tbl, ws, path, sheet):
            from PySide6.QtGui import QColor
            tbl.clearContents(); tbl.setRowCount(0)
            rows = list(ws.iter_rows())
            if not rows: return
            headers = [str(c.value) if c.value is not None else '' for c in rows[0]]
            tbl.setColumnCount(len(headers)); tbl.setHorizontalHeaderLabels(headers)
            r_idx = 0
            for row in rows[1:]:
                tbl.insertRow(r_idx)
                for ci, cell in enumerate(row):
                    v = cell.value
                    item = QTableWidgetItem('' if v is None else str(v))
                    # 颜色(复制xlsx单元格fill)
                    try:
                        fill = cell.fill
                        if fill and fill.start_color and fill.start_color.rgb:
                            rgb = str(fill.start_color.rgb)
                            if rgb not in ('00000000', '0'):
                                item.setBackground(QColor('#' + rgb[-6:]))
                    except Exception: pass
                    tbl.setItem(r_idx, ci, item)
                r_idx += 1
            tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            tbl.setSortingEnabled(False)

        # ===== 速率统计相关方法 =====
        def sf(self):
            fs, _ = QFileDialog.getOpenFileNames(self, "选择文件", "", "所有支持文件 (*.xlsx *.xls *.csv *.zip *.mmf *.tmf);;Excel/CSV (*.xlsx *.xls *.csv);;All (*)")
            if fs:
                for f in fs:
                    fn = os.path.basename(f).lower()
                    # 判断是基准文件还是MMF文件
                    if '基准' in fn or '呼叫详情' in fn or '整理' in fn:
                        self.base_file = f
                        self.lt.append(f"已选择基准文件: {os.path.basename(f)}")
                    else:
                        self.files = list(dict.fromkeys((self.files or []) + [f]))
                if self.files:
                    self.te.setText("\n".join(os.path.basename(f) for f in self.files))
                    self.br.setEnabled(True)
                if self.base_file:
                    self.te.append(f"\n[基准] {os.path.basename(self.base_file)}")

        def cf(self):
            self.files = []; self.base_file = None
            self.te.clear(); self.br.setEnabled(False)

        def run(self):
            if not self.files: QMessageBox.warning(self, "提示", "请先选择MMF文件"); return
            # 同步并检查手机号映射(为空时提醒)
            self._sync_phone_mapping('data')
            if not self.phone_trace_map_data:
                ret = QMessageBox.question(self, "手机号映射为空",
                    "当前未配置任何手机号-TraceID映射，结果中将无法显示手机号。\n是否继续处理？",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No)
                if ret != QMessageBox.StandardButton.Yes:
                    return
            # 清空结果数据显示区
            self.tbl_detail.setRowCount(0); self.tbl_detail.setColumnCount(0)
            self.tbl_summary.setRowCount(0); self.tbl_summary.setColumnCount(0)
            self.tbl_compare.setRowCount(0); self.tbl_compare.setColumnCount(0)
            self.lt.clear()
            qci_list = []
            if self.cb6.isChecked(): qci_list.append(6)
            if self.cb7.isChecked(): qci_list.append(7)
            if not qci_list: qci_list = [7]

            # 时间段过滤
            time_filter = None
            if self.cb_time.isChecked():
                t_start = self.time_start.time()
                t_end = self.time_end.time()
                time_filter = (True, t_start, t_end)

            # FTP时长
            ftp_duration = self.ftp_duration_spin.value()

            self.br.setEnabled(False); self.bcancel.setEnabled(True)
            self.pb.setValue(0); self.lt.append(f"开始处理... QCI={qci_list}, FTP时长={ftp_duration}秒")
            if time_filter:
                self.lt.append(f"时间段过滤: {t_start.toString('HH:mm:ss')} ~ {t_end.toString('HH:mm:ss')}")
            self.thread = WorkerThread(self.files, qci_list, self.sd.value(), self.su.value(),
                                       base_file=self.base_file, time_filter=time_filter,
                                       ftp_duration=ftp_duration,
                                       merge_raw=self.cb_merge_raw.isChecked(),
                                       add_annotations=self.cb_annotations.isChecked(),
                                       phone_trace_map=self.phone_trace_map_data)
            self.thread.progress.connect(self.upd)
            self.thread.progress_pct.connect(lambda p: self.pb.setValue(p))
            self.thread.done.connect(self.dn); self.thread.start()

        def cancel(self):
            if hasattr(self, 'thread') and self.thread.isRunning():
                self.thread.cancel(); self.lt.append("已请求取消，将在当前阶段结束后停止...")

        @Slot(str)
        def upd(self, msg): self.lt.append(msg)

        @Slot(str)
        def dn(self, path):
            self.bcancel.setEnabled(False)
            if path:
                self.pb.setValue(100)
                self.last_out = path; self.bopen.setEnabled(True)
                self.lt.append(f"完成! {path}")
                self.show_result(path)   # 内嵌显示结果
                # 自动打开结果文件
                os.system(f'open "{path}"')
            else:
                self.pb.setValue(0); self.lt.append("已取消。")
            self.br.setEnabled(True)

        def open_out(self):
            if self.last_out: os.system(f'open "{self.last_out}"')

        # ===== 内嵌结果查看(基础版：显示+默认A=1筛选+颜色) =====
        def show_result(self, path):
            try:
                import openpyxl
                from PySide6.QtGui import QColor
                wb = openpyxl.load_workbook(path, data_only=True)
                self._fill_table(self.tbl_detail, wb['数据业务-详细过程'], filter_a1=True, path=path, sheet='数据业务-详细过程')
                self._fill_table(self.tbl_summary, wb['汇总'], filter_a1=False, path=path, sheet='汇总')
                self._fill_table(self.tbl_compare, wb['数据业务-对比'], filter_a1=False, path=path, sheet='数据业务-对比')
                self.lt.append("结果已内嵌显示（详细过程默认筛选A=1）")
            except Exception as e:
                self.lt.append(f"内嵌显示失败: {e}")

        def _fill_table(self, tbl, ws, filter_a1, path, sheet):
            from PySide6.QtGui import QColor
            tbl.clearContents(); tbl.setRowCount(0)
            rows = list(ws.iter_rows())
            if not rows: return
            headers = [str(c.value) if c.value is not None else '' for c in rows[0]]
            tbl.setColumnCount(len(headers)); tbl.setHorizontalHeaderLabels(headers)
            r_idx = 0
            for ri, row in enumerate(rows[1:], start=2):
                if filter_a1:
                    a = row[0].value if row else None
                    if a is None or str(a) != '1': continue
                tbl.insertRow(r_idx)
                for ci, cell in enumerate(row):
                    v = cell.value
                    item = QTableWidgetItem('' if v is None else str(v))
                    # 颜色(复制xlsx单元格fill)
                    try:
                        fill = cell.fill
                        if fill and fill.start_color and fill.start_color.rgb:
                            rgb = str(fill.start_color.rgb)
                            if rgb not in ('00000000', '0'):
                                item.setBackground(QColor('#' + rgb[-6:]))
                    except Exception: pass
                    # 批注(setToolTip, hover显示; 点击显示下一轮完善)
                    try:
                        if cell.comment:
                            item.setToolTip(cell.comment.text)
                    except Exception: pass
                    tbl.setItem(r_idx, ci, item)
                r_idx += 1
            tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            tbl.setSortingEnabled(False)

        def about(self):
            d = QDialog(self); d.setWindowTitle("关于")
            d.resize(560, 600)
            v = QVBoxLayout(d)
            te = QTextEdit(); te.setReadOnly(True)
            te.setHtml(
                "<h3 style='text-align:center'>用户级数据、语音跟踪统计工具</h3>"
                "<p><b>版本:</b> V2.5.17_CX &nbsp; <b>开发者:</b> 孙晓军 &nbsp; "
                "<b>联系方式:</b> 317827@qq.com</p><hr/>"
                "<h4>更新记录</h4>"
                "<p><b>2026-07-19</b></p>"
                "<ul>"
                "<li><b>V2.5.17_CX</b> 完整通话规则改为只看BYE(不再找200);RSRP拆分为(全部)和(通话时)两列;批注增加每轮次明细;汇总表/按文件名表列顺序调整(新增主被叫/手机号/类型);删除运营商表;新增12个数据明细sheet(语音6+数据6);运营商检测增加QCI兜底(路径无法判断时按QCI:7→电信/5或6→联通);默认值调整(QCI6+7都勾选,合并原始数据默认开,添加批注默认开);数据业务写入从openpyxl改为xlsxwriter逐行写(提速31%,可中断);手机号列构建性能优化(解决大量文件时卡顿);临时文件安全写入(取消不留坏文件)</li>\n"
                "<li><b>V2.5.13</b> 手机号-TraceID映射解析增强：粘贴/拖入文本支持多种分隔符(Tab/空格/逗号/分号/斜杠等)；文本框内容变化后停止输入约300ms自动解析并增量填入表格(防抖避免逐字中间态,去重不重复,填完自动持久化)</li>\n"
                "<li><b>V2.5.12</b> 取消响应优化：生成汇总/写入详细过程/原始数据写入/批注/对比横表/着色等长循环全部加密取消检查点，点取消后通常1~3秒内停止且开始按钮立即恢复可用；手机号-TraceID映射持久化：5G速率tab与语音VQI tab各存一份独立JSON（关闭后重开自动恢复，增删改自动保存，映射文件损坏或缺失时空启动不报错）；开始处理前若未配置任何映射弹窗提醒是否继续；基准文件缺失改为中性提示且不影响主流程</li>\n"
                "<li><b>V2.5.11</b> 修复手机号映射合并问题：手动添加和拖入/粘贴的手机号映射现在都能参与计算；优化语音VQI批注性能：预构建过滤缓存避免重复遍历，大幅减少运营商/文件名批注卡顿；QCI 6和7改为独立可选不再互斥；处理开始时清空结果数据显示区；数据业务处理跳过非MMF文件名(如ALL_FTP_DETAIL_*.xlsx)；修复跳过非MMF文件时data_file_info索引错位bug并增加空数据保护</li>\n"
                "<li><b>V2.5.10</b> 新增手机号-TraceID映射功能：参数设置支持手动添加/拖入ExcelTXT/粘贴文本三种方式填写手机号与TraceID对应关系；自动从文件路径提取\"用户跟踪ID=xxx\"匹配手机号；按文件名/按运营商/汇总-数据和语音Sheet新增手机号列</li>\n"
                "</ul>\n"
                + "<p><b>2026-07-19</b></p>"
                "<ul>"
"<li>批注优化：汇总-数据和语音、按文件名、按运营商表的批注显示详细计算过程(分子/分母/具体数值用/分隔)</li>\n"
"<li>汇总-数据和语音表：文件名列改为逐行显示具体文件名(换行)，起止时间列改为每文件第一轮INVITE开始~最后一轮结束</li>\n"
"<li>按运营商表：新增文件名和时间范围列(换行显示)</li>\n"
"<li>按文件名表：新增微信/商店等业务指标列(当前留空)</li>\n"
"<li>默认勾选改为默认不勾选(合并原始数据、添加批注)</li>\n"
"<li>原始数据sheet取消20000行限制，全量写入</li>\n"
"<li>修复跨运营商信令干扰：按运营商(电信/联通)分组独立搜索INVITE→180→BYE→200，避免联通的BYE匹配到电信的200</li>\n"
"<li>修复搜索边界公式bug：非连续索引时最后一段信令不遗漏</li>\n"
"<li>修复列不存在时KeyError崩溃：r180/bye列存在性检查</li>\n"
"<li>修复get_loc对不连续索引抛出KeyError：try-except保护</li>\n"
"<li>文件名改为显示相对路径(含文件夹信息)，如\"语音类输入文件/电信主叫SIP/SaveMsg_xxx.csv\"</li>\n"
"</ul>"
                "<p><b>2026-07-18</b></p>"
                "<ul>"
"<li><b>V2.5.8</b> 性能优化：基准对齐使用二分查找(searchsorted)替代向量化abs计算，大幅提升大文件处理速度；新增服务小区平均RSRP统计(从RRC_MEAS_RPRT提取servRSRP，双值取最大)；修复RSRP正则支持单值无逗号格式；VQI提取使用.values批量优化；xlsxwriter引擎替代openpyxl(write_row批量写入)；批注添加功能(复选框控制)；自动打开结果文件</li>\n"
"<li><b>V2.5.7</b> 修复按文件名统计表显示0的问题：1)为每条voice_record/vqi_record添加_source_file字段；2)使用voice_file_info字典记录每个文件的基本信息；3)按_source_file字段精确匹配统计每个文件的轮次/MOS指标</li>\n"
"<li><b>V2.5.6</b> 新增\"合并原始数据到结果文件\"复选框(速率统计/VQI标签各一),默认勾选</li>\n"
"<li>新增批量计算批注:汇总-数据和语音、按文件名、按运营商Sheet的数值单元格自动添加计算批注</li>\n"
"<li>数据文件处理:无MessageType列的数据文件(mmf)不再跳过,读取并计算下行/上行速率指标</li>\n"
"<li>VQI输出Sheet优化:新增\"数据业务-原始文件\"Sheet,按文件名/按运营商表含完整数据业务指标</li>\n"
"<li>性能优化:文件类型检测只读前3行(nrows=3);缓存文件列表避免重复读取</li>\n"
                "<li><b>V2.5.5</b> 修复文件类型显示为unknown问题：_detect_file_type返回语音/数据而非unknown</li>"
                "<li>按文件名表新增指标列：通话轮次、平均呼叫建立时长(s)、优质通话占比(MOS3.5)</li>"
                "<li>按运营商表新增完整指标列：下行前10%峰值速率等12列（当前留空）</li>"
                "<li>语音指标统计sheet中对空呼叫建立时长单元格加批注无RSP_180</li>"
                "<li>平均呼叫建立时长只统计有值的轮次</li>"
                "<li><b>V2.5.3</b> 呼叫建立时长单位从毫秒改为秒，精确到小数点后2位</li>"
                "<li>新增汇总-数据和语音Sheet：标准行+电信+联通，包含语音指标和数据业务指标列</li>"
                "<li>新增按文件名Sheet：每个文件一行，包含文件名、运营商、文件类型、行数、时间范围</li>"
                "<li>新增按运营商Sheet：电信/联通各一行，包含语音文件数、通话轮次、平均呼叫建立时长、优质通话占比</li>"
                "<li>process_vqi函数新增time_range参数，支持外部传入时间起止</li>"
                "<li><b>V2.5.2</b> 全面重构语音VQI处理逻辑：自动分类语音/数据文件、识别INVITE(取最后一个连续)/180/BYE/200完整通话</li>"
                "<li>新增MOS3.5优质通话占比计算：从Detail Info提取DlE2eVqi值，排除65535，大于3.5为优质</li>"
                "<li>VQI输出3个Sheet：原始数据(加运营商/业务类型列)、语音指标统计、MOS3.5计算明细</li>"
                "<li>支持.tmf文件(语音)和.mmf文件(数据)的自动分类</li>"
                "<li><b>V2.5.1</b> 修复拖入文件不起作用的问题；支持csv文件类型</li>"
                "<li><b>V2.5.0</b> 新增语音VQI标签页：实现完整的语音VQI处理功能</li>"
                "<li>VQI功能：读取MMF文件，筛选MessageType非空行，识别SIP信令（INVITE/180/BYE/200），计算呼叫建立时长，判定完整通话</li>"
                "<li>VQI输出：原始数据sheet（MessageType非空的所有行）+ 语音统计sheet（轮次/开始时间/结束时间/呼叫建立时长/是否完整通话）</li>"
                "<li>FTP时长选项：参数设置新增FTP时长下拉选择（10秒/20秒/30秒），默认20秒，动态调整FTP候选检测阈值</li>"
                "<li>Sheet重命名：详细过程→数据业务-详细过程，对比横表→数据业务-对比</li>"
                "</ul>"
                "<p><b>2026-07-17</b></p>"
                "<ul>"
                "<li><b>V2.4.5</b> 参数设置增加时间段限制：复选框启用后，可设置开始/结束时间(HH:mm:ss)</li>"
                "<li>启用时只处理该时间段内的数据，未启用则全量处理</li>"
                "<li><b>V2.4.4</b> 支持用户自定义基准文件</li>"
                "</ul>"
                "<p><b>V2.4.2</b></p>"
                "<ul>"
                "<li>工具改名：5G用户级公共监控速率统计工具 → 用户级数据、语音跟踪统计工具</li>"
                "<li>主界面改为双标签(QTabWidget)：标签1=5G用户级公共监控速率统计工具，标签2=5G虚拟用户跟踪用户语音VQI工具(空壳待接入算法)</li>"
                "<li>速率统计标签布局调整：左侧上设置+下运行(进度/日志)，右侧腾出内嵌显示结果3个sheet(详细过程/汇总/对比横表)</li>"
                "<li>右侧内嵌QTableWidget：默认筛选A=1、复制xlsx颜色、批注(hover显示)</li>"
                "<li>输入文件区：选择/清空按钮紧挨着，拖入框高度翻倍(300px)、边框加粗明显</li>"
                "<li>关于对话框：更新记录按天汇总，同一天多次更新合并列出</li>"
                "</ul>"
                "<p><b>V2.4.1</b> <i>(2026-07-17)</i></p>"
                "<ul>"
                "<li>商店小文件识别修复：GAP_MERGE 2.0→3.0，防止商店大段被切碎误判FTP；8783段等真小文件恢复识别</li>"
                "<li>去掉商店小文件上行占比约束（RLC/MAC占比均不可靠），回到max(dl)选取</li>"
                "<li>对比横表改就近匹配：找离基准最近的同业务段，距离≤300秒，完整轮0缺失</li>"
                "<li>支持嵌套zip：递归解压(zip里套zip)</li>"
                "<li>取消按钮修复：写Excel各阶段加检查点，长循环每5000行检查</li>"
                "<li>清理重复sort_values；fmt_time改预编译正则</li>"
                "</ul>"
                "<p><b>V2.4.0</b> <i>(2026-07-16)</i></p>"
                "<ul>"
                "<li>性能优化：预编译正则 + usecols 只读必要列，MMF加载提速</li>"
                "<li>汇总批注：每格显示公式+实际数值（hover查看）</li>"
                "<li>D/E列颜色标记：详细过程 D列浅蓝/E列浅绿</li>"
                "</ul>"
                "<p><b>V2.3.0</b> <i>(2026-07-16)</i></p>"
                "<ul>"
                "<li>性能优化：基准对齐/筛选列向量化，处理耗时 107秒→14秒(约7.7倍)</li>"
                "<li>新增：取消按钮(阶段点中断)、拖拽文件入窗口、真实进度百分比</li>"
                "<li>对比竖表美化：业务/轮次配色、合并同轮次、冻结首行、行高</li>"
                "</ul>"
                "<p><b>V2.2.0</b> <i>(2026-07-16)</i></p>"
                "<ul>"
                "<li>对比增加「偏差」行/列：(代码−基准)/基准 百分比，速率与时长各算</li>"
                "<li>对比横表每轮由2列扩为3列(代码+基准+偏差)</li>"
                "<li>新增「对比竖表」sheet(横表转置)</li>"
                "<li>主界面改左右分栏(左35%设置/右65%结果)，全屏+自适应分辨率+高分屏字号放大+可拖动+打开输出按钮</li>"
                "</ul>"
                "<p><b>V2.1.0</b> <i>(2026-07-16)</i></p>"
                "<ul>"
                "<li>对比横表匹配改为「业务连续段」判定 + 按 FTP下载 切轮</li>"
                "<li>不再用固定时间窗口(-5~+60s)和硬切6条分组</li>"
                "<li>整轮6业务完整且全对上才编号，否则标「不计轮次」</li>"
                "</ul>"
                "<p><b>V2.0.9</b></p>"
                "<ul>"
                "<li>修复对比横表代码业务结束时间缺失</li>"
                "<li>新增第一个 FTP对 之前不完整轮次的识别</li>"
                "<li>版本号统一更新(文件名/关于/需求文档)</li>"
                "</ul>"
                "<p><b>V2.0.1</b></p>"
                "<ul><li>QCI 支持 5/6/7 多选；主界面增加「关于」</li></ul>"
                "<p><b>V2.0.0</b></p>"
                "<ul>"
                "<li>从零重写，三页向导改单页</li>"
                "<li>自动识别 FTP/商店/微信 6 类业务</li>"
                "<li>从原始 MMF 文件直接生成详细过程/对比/汇总 3 个 Sheet</li>"
                "<li>削峰阈值可调，轮次完整性检查</li>"
                "</ul>"
            )
            v.addWidget(te)
            bb = QPushButton("关闭"); bb.clicked.connect(d.accept)
            v.addWidget(bb)
            d.exec_()


if __name__ == '__main__':
    if GUI_OK and len(sys.argv) == 1:
        app = QApplication(sys.argv); app.setStyle('Fusion')
        w = MainWindow(); w.show(); sys.exit(app.exec())
    else:
        out = process(
            ['联通/mmf20260703115328-电信.xlsx', '联通/mmf20260703115334-电信.xlsx'],
            qci_list=[5,6,7], dl_clip=1000, ul_clip=200
        )
        print(f"输出: {out}")
